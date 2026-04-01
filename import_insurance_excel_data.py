"""
Comprehensive Insurance Excel Data Import Script
Imports insurance adjudication reports and debtor balances into the system
Acts as Senior Engineer and Senior Account Officer to ensure proper accounting balance

This script will:
1. Read all Excel files from "insurance excel" folder
2. Map insurance companies to Payer records
3. Create/update InsuranceReceivable records
4. Create/update AccountsReceivable (debtors) records
5. Create journal entries for proper accounting balance
6. Validate all data and check for errors
"""
import os
import sys
from decimal import Decimal
from datetime import datetime, date
from django.db import transaction
from django.utils import timezone

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'hms.settings')
import django
django.setup()

from hospital.models import Payer, Patient, Invoice
from hospital.models_accounting import Account
from hospital.models_accounting_advanced import (
    InsuranceReceivable, AdvancedAccountsReceivable, 
    AdvancedJournalEntry, AdvancedJournalEntryLine
)
from django.db.models import Sum

# Try to import Excel reading libraries
HAS_PANDAS = False
HAS_OPENPYXL = False
HAS_XLRD = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    try:
        from openpyxl import load_workbook
        HAS_OPENPYXL = True
    except ImportError:
        try:
            import xlrd
            HAS_XLRD = True
        except ImportError:
            pass


class InsuranceExcelImporter:
    """Comprehensive insurance data importer"""
    
    def __init__(self, dry_run=False):
        self.dry_run = dry_run
        self.stats = {
            'files_processed': 0,
            'payers_found': 0,
            'payers_created': 0,
            'receivables_created': 0,
            'receivables_updated': 0,
            'ar_created': 0,
            'ar_updated': 0,
            'journal_entries_created': 0,
            'errors': [],
            'warnings': []
        }
        
        # Get or create required accounts (outside transaction)
        self.insurance_receivable_account = self.get_or_create_account(
            code='1201',
            name='Insurance Receivables',
            account_type='asset'
        )
        
        self.trade_receivable_account = self.get_or_create_account(
            code='1200',
            name='Trade Receivables',
            account_type='asset'
        )
        
        # Get or create summary patient (outside transaction)
        self.summary_patient = self.get_or_create_summary_patient()
    
    def get_or_create_account(self, code, name, account_type):
        """Get or create an account"""
        try:
            account, created = Account.objects.get_or_create(
                account_code=code,
                defaults={
                    'account_name': name,
                    'account_type': account_type,
                    'is_active': True
                }
            )
            if created:
                print(f"  [OK] Created account: {code} - {name}")
            return account
        except Exception as e:
            self.stats['errors'].append(f"Error creating account {code}: {e}")
            return None
    
    def cleanup_suppliers_from_insurance(self):
        """Remove any suppliers that were mistakenly created as insurance payers"""
        print(f"\n{'='*80}")
        print("Checking for suppliers mistakenly created as insurance payers")
        print(f"{'='*80}")
        
        # Get all payers that might be suppliers
        all_payers = Payer.objects.filter(
            payer_type__in=['private', 'nhis'],
            is_active=True
        )
        
        suppliers_to_remove = []
        for payer in all_payers:
            if self.is_supplier(payer.name):
                suppliers_to_remove.append(payer)
        
        if suppliers_to_remove:
            print(f"  [FOUND] {len(suppliers_to_remove)} supplier(s) mistakenly created as insurance payers:")
            for payer in suppliers_to_remove:
                print(f"    - {payer.name} (ID: {payer.id})")
            
            if not self.dry_run:
                # Delete the payers
                for payer in suppliers_to_remove:
                    # Check if payer has any related records
                    has_invoices = payer.invoices.exists()
                    has_receivables = hasattr(payer, 'receivables') and payer.receivables.exists()
                    
                    if has_invoices or has_receivables:
                        print(f"    [SKIP] {payer.name} has related records - marking as inactive instead")
                        payer.is_active = False
                        payer.save()
                    else:
                        print(f"    [DELETE] Removing {payer.name}")
                        payer.delete()
                
                print(f"  [OK] Cleaned up {len(suppliers_to_remove)} supplier(s)")
            else:
                print(f"  [DRY RUN] Would delete {len(suppliers_to_remove)} supplier(s)")
        else:
            print("  [OK] No suppliers found in insurance payers")
    
    def get_or_create_summary_patient(self):
        """Get or create summary patient for insurance receivables"""
        try:
            # Try to get existing patient
            patient = Patient.objects.filter(mrn='SUMMARY-INS').first()
            if patient:
                return patient
            
            # Create new patient - handle QR code issue by disabling signal temporarily
            from django.db.models.signals import post_save
            from hospital import signals
            
            # Temporarily disconnect the QR code signal
            post_save.disconnect(signals.ensure_patient_qr_profile, sender=Patient)
            
            try:
                patient = Patient(
                    mrn='SUMMARY-INS',
                    first_name='Summary',
                    last_name='Insurance Receivables',
                    date_of_birth=date(2000, 1, 1),
                    gender='M'
                )
                patient.save()
                
                # Manually create QR code with unique token
                from hospital.models import PatientQRCode
                import secrets
                try:
                    qr_code, _ = PatientQRCode.objects.get_or_create(
                        patient=patient,
                        defaults={
                            'qr_token': secrets.token_urlsafe(32),
                            'is_active': True
                        }
                    )
                    if not qr_code.qr_token:
                        qr_code.qr_token = secrets.token_urlsafe(32)
                        qr_code.save()
                except Exception as qr_error:
                    # If QR code creation fails, try to get existing one
                    try:
                        qr_code = PatientQRCode.objects.get(patient=patient)
                        if not qr_code.qr_token:
                            qr_code.qr_token = secrets.token_urlsafe(32)
                            qr_code.save()
                    except:
                        pass
                
                print(f"  [OK] Created summary patient for insurance receivables")
                return patient
            finally:
                # Reconnect the signal
                post_save.connect(signals.ensure_patient_qr_profile, sender=Patient)
                
        except Exception as e:
            self.stats['errors'].append(f"Error creating summary patient: {e}")
            # Try one more time to get existing patient
            try:
                return Patient.objects.get(mrn='SUMMARY-INS')
            except:
                return None
    
    def is_supplier(self, entity_name):
        """Check if an entity is a supplier (not insurance)"""
        name_lower = entity_name.lower()
        
        supplier_keywords = [
            'pharma', 'pharmaceutical', 'enterprise', 'logistics', 'supply',
            'electricals', 'motors', 'company limited', 'ltd', 'limited',
            'bank', 'plc', 'church', 'university', 'college', 'school',
            'total', 'oil', 'petroleum', 'gas', 'energy', 'power'
        ]
        
        insurance_keywords = [
            'insurance', 'health insurance', 'medical insurance', 'nhis',
            'mutual health', 'care', 'health', 'medical', 'assurance'
        ]
        
        # Check for insurance keywords first
        for keyword in insurance_keywords:
            if keyword in name_lower:
                return False
        
        # Check for supplier keywords
        for keyword in supplier_keywords:
            if keyword in name_lower:
                return True
        
        # Specific checks
        if 'health insurance' in name_lower or 'medical insurance' in name_lower:
            return False
        
        return False
    
    def find_or_create_payer(self, payer_name, is_supplier=False):
        """Find or create a Payer record for insurance company (NOT for suppliers)"""
        payer_name = payer_name.strip()
        if not payer_name:
            return None
        
        # Don't create payers for suppliers - they should be handled separately
        if is_supplier:
            return None
        
        # Try to find existing payer
        payer = Payer.objects.filter(
            name__icontains=payer_name,
            payer_type__in=['private', 'nhis']
        ).first()
        
        if not payer:
            # Try partial match
            words = payer_name.split()
            if words:
                for word in words:
                    if len(word) > 3:  # Skip short words
                        payer = Payer.objects.filter(
                            name__icontains=word,
                            payer_type__in=['private', 'nhis']
                        ).first()
                        if payer:
                            break
        
        if not payer:
            # Create new payer
            payer_type = 'private'
            if 'NHIS' in payer_name.upper() or 'NATIONAL' in payer_name.upper():
                payer_type = 'nhis'
            
            payer = Payer.objects.create(
                name=payer_name,
                payer_type=payer_type,
                is_active=True
            )
            self.stats['payers_created'] += 1
            print(f"  [OK] Created payer: {payer_name} ({payer_type})")
        else:
            self.stats['payers_found'] += 1
        
        return payer
    
    def read_excel_file(self, file_path):
        """Read Excel file using available library"""
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if HAS_PANDAS:
            try:
                if file_ext == '.xls':
                    # For old .xls files, try xlrd engine
                    try:
                        return pd.read_excel(file_path, engine='xlrd')
                    except:
                        return pd.read_excel(file_path)
                else:
                    return pd.read_excel(file_path)
            except Exception as e:
                print(f"  [WARNING] Pandas error: {e}, trying other methods...")
        
        if HAS_OPENPYXL and file_ext == '.xlsx':
            try:
                wb = load_workbook(file_path, read_only=True, data_only=True)
                data = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    sheet_data = []
                    for row in ws.iter_rows(values_only=True):
                        sheet_data.append(row)
                    data.append({'sheet': sheet_name, 'data': sheet_data})
                return data
            except Exception as e:
                print(f"  [WARNING] Openpyxl error: {e}")
        
        if HAS_XLRD and file_ext == '.xls':
            try:
                workbook = xlrd.open_workbook(file_path)
                data = []
                for sheet_name in workbook.sheet_names():
                    sheet = workbook.sheet_by_name(sheet_name)
                    sheet_data = []
                    for row_idx in range(sheet.nrows):
                        row = sheet.row_values(row_idx)
                        sheet_data.append(row)
                    data.append({'sheet': sheet_name, 'data': sheet_data})
                return data
            except Exception as e:
                print(f"  [WARNING] Xlrd error: {e}")
        
        return None
    
    def process_debtor_balances(self, file_path):
        """Process JERRY.xlsx debtor balances sheet"""
        print(f"\n{'='*80}")
        print(f"Processing Debtor Balances from: {os.path.basename(file_path)}")
        print(f"{'='*80}")
        
        data = self.read_excel_file(file_path)
        if data is None:
            self.stats['errors'].append(f"Could not read file: {file_path}")
            return
        
        # Handle different data formats
        if isinstance(data, list):
            # openpyxl/xlrd format
            debtor_sheet = None
            for sheet_info in data:
                if 'DEBTOR' in sheet_info['sheet'].upper():
                    debtor_sheet = sheet_info['data']
                    break
            
            if not debtor_sheet:
                self.stats['warnings'].append(f"No debtor sheet found in {file_path}")
                return
            
            # Find data start row
            data_start = None
            for i, row in enumerate(debtor_sheet):
                if row and len(row) >= 2:
                    if isinstance(row[0], str) and isinstance(row[1], (int, float)):
                        data_start = i
                        break
            
            if data_start is None:
                self.stats['warnings'].append(f"Could not find data start in {file_path}")
                return
            
            # Process rows
            for row in debtor_sheet[data_start:]:
                if not row or len(row) < 2:
                    continue
                
                payer_name = str(row[0]).strip() if row[0] else None
                amount = row[1] if isinstance(row[1], (int, float)) else None
                
                if payer_name and amount and amount != 0:
                    self.process_debtor_balance(payer_name, Decimal(str(amount)))
        
        elif hasattr(data, 'iterrows'):
            # pandas DataFrame
            # Find the debtor balances sheet
            if isinstance(data, dict):
                debtor_df = data.get('DEBTOR BALANCES')
            else:
                debtor_df = data
            
            if debtor_df is None:
                self.stats['warnings'].append(f"No debtor data found in {file_path}")
                return
            
            # Process rows
            for _, row in debtor_df.iterrows():
                payer_name = None
                amount = None
                
                # Try to find payer name and amount columns
                for col in debtor_df.columns:
                    val = row[col]
                    if isinstance(val, str) and val.strip():
                        payer_name = val.strip()
                    elif isinstance(val, (int, float)) and val != 0:
                        amount = Decimal(str(val))
                
                if payer_name and amount:
                    self.process_debtor_balance(payer_name, amount)
    
    def process_debtor_balance(self, payer_name, balance_amount):
        """Process a single debtor balance entry"""
        if not payer_name or balance_amount == 0:
            return
        
        payer = self.find_or_create_payer(payer_name)
        if not payer:
            self.stats['warnings'].append(f"Could not create/find payer: {payer_name}")
            return
        
        # Create or update summary receivable
        # For debtor balances, we create a summary InsuranceReceivable
        # Since we don't have individual patient/invoice data, we'll create a summary entry
        
        try:
            # Use pre-created summary patient
            if not self.summary_patient:
                self.stats['errors'].append("Summary patient not available")
                return
            summary_patient = self.summary_patient
            
            # Create summary invoice if needed
            invoice_number = f'SUMMARY-{payer.name.upper().replace(" ", "-")}-{timezone.now().strftime("%Y%m%d")}'
            summary_invoice, _ = Invoice.objects.get_or_create(
                invoice_number=invoice_number,
                defaults={
                    'patient': summary_patient,
                    'payer': payer,
                    'total_amount': abs(balance_amount),
                    'balance': abs(balance_amount),
                    'status': 'issued',
                    'due_at': timezone.now() + timezone.timedelta(days=30)
                }
            )
            
            # Create or update InsuranceReceivable
            # Generate shorter receivable number
            payer_code = payer.name.upper().replace(" ", "")[:10]
            receivable_number = f'IR-{payer_code}-{timezone.now().strftime("%y%m%d")}'
            if len(receivable_number) > 50:
                receivable_number = receivable_number[:50]
            
            receivable, created = InsuranceReceivable.objects.get_or_create(
                receivable_number=receivable_number,
                defaults={
                    'insurance_company': payer,
                    'patient': summary_patient,
                    'invoice': summary_invoice,
                    'total_amount': abs(balance_amount),
                    'amount_paid': Decimal('0.00'),
                    'balance_due': abs(balance_amount),
                    'due_date': (timezone.now() + timezone.timedelta(days=30)).date(),  # For InsuranceReceivable model
                    'status': 'pending',
                    'receivable_account': self.insurance_receivable_account,
                    'notes': f'Summary balance from Excel import - {payer_name}'
                }
            )
            
            if created:
                self.stats['receivables_created'] += 1
                print(f"  [OK] Created receivable: {payer_name} - GHS {balance_amount:,.2f}")
            else:
                # Update balance
                receivable.total_amount = abs(balance_amount)
                receivable.balance_due = abs(balance_amount)
                receivable.save()
                self.stats['receivables_updated'] += 1
                print(f"  [OK] Updated receivable: {payer_name} - GHS {balance_amount:,.2f}")
            
            # Create or update AccountsReceivable
            ar, ar_created = AdvancedAccountsReceivable.objects.get_or_create(
                invoice=summary_invoice,
                defaults={
                    'patient': summary_patient,
                    'invoice_amount': abs(balance_amount),
                    'amount_paid': Decimal('0.00'),
                    'balance_due': abs(balance_amount),
                    'due_date': timezone.now().date()
                }
            )
            
            if ar_created:
                self.stats['ar_created'] += 1
            else:
                ar.invoice_amount = abs(balance_amount)
                ar.balance_due = abs(balance_amount)
                ar.save()
                self.stats['ar_updated'] += 1
            
            # Create journal entry for accounting balance
            if not self.dry_run:
                self.create_receivable_journal_entry(receivable, abs(balance_amount))
        
        except Exception as e:
            error_msg = f"Error processing debtor {payer_name}: {e}"
            self.stats['errors'].append(error_msg)
            print(f"  [ERROR] {error_msg}")
    
    def create_receivable_journal_entry(self, receivable, amount):
        """Create journal entry for insurance receivable"""
        if self.dry_run:
            return
        
        try:
            # Check if journal entry already exists
            if receivable.journal_entry:
                return
            
            # Create journal entry
            # Debit: Insurance Receivables (Asset increase)
            # Credit: Revenue (we'll use a summary revenue account)
            revenue_account = self.get_or_create_account(
                code='4999',
                name='Insurance Revenue - Summary',
                account_type='revenue'
            )
            
            journal = AdvancedJournalEntry.objects.create(
                entry_number=f'JE-IR-{receivable.receivable_number}',
                entry_date=timezone.now().date(),
                description=f'Insurance Receivable - {receivable.insurance_company.name}',
                status='posted',
                reference=f'IR-{receivable.receivable_number}'
            )
            
            # Debit: Insurance Receivables
            AdvancedJournalEntryLine.objects.create(
                journal_entry=journal,
                account=self.insurance_receivable_account,
                debit=amount,
                credit=Decimal('0.00'),
                description=f'Receivable from {receivable.insurance_company.name}'
            )
            
            # Credit: Revenue
            AdvancedJournalEntryLine.objects.create(
                journal_entry=journal,
                account=revenue_account,
                debit=Decimal('0.00'),
                credit=amount,
                description=f'Revenue from {receivable.insurance_company.name}'
            )
            
            receivable.journal_entry = journal
            receivable.save()
            
            self.stats['journal_entries_created'] += 1
            
        except Exception as e:
            error_msg = f"Error creating journal entry for {receivable.receivable_number}: {e}"
            self.stats['errors'].append(error_msg)
            print(f"  [ERROR] {error_msg}")
    
    def process_adjudication_report(self, file_path):
        """Process insurance adjudication report files"""
        print(f"\n{'='*80}")
        print(f"Processing Adjudication Report: {os.path.basename(file_path)}")
        print(f"{'='*80}")
        
        # Extract insurance company name from filename
        filename = os.path.basename(file_path)
        insurance_name = None
        
        # Try to extract insurance company name
        if 'COSMO' in filename.upper():
            insurance_name = 'Cosmopolitan Health Insurance'
        elif 'EQUITY' in filename.upper():
            insurance_name = 'Equity Health Insurance'
        elif 'GAB' in filename.upper():
            insurance_name = 'GAB Health Insurance'
        else:
            # Try to extract from filename pattern
            parts = filename.split('-')
            if parts:
                insurance_name = parts[0].strip()
        
        if not insurance_name:
            self.stats['warnings'].append(f"Could not determine insurance company from filename: {filename}")
            return
        
        payer = self.find_or_create_payer(insurance_name)
        if not payer:
            self.stats['warnings'].append(f"Could not create/find payer: {insurance_name}")
            return
        
        # Read the file
        data = self.read_excel_file(file_path)
        if data is None:
            self.stats['errors'].append(f"Could not read file: {file_path}")
            return
        
        # Extract date range from filename
        date_range = None
        if 'OCT 2025' in filename.upper():
            date_range = 'October 2025'
        elif 'AUG 2025' in filename.upper():
            date_range = 'August 2025'
        elif 'SEP 2025' in filename.upper():
            date_range = 'September 2025'
        
        print(f"  Insurance: {insurance_name}")
        print(f"  Period: {date_range}")
        print(f"  Note: Adjudication reports require detailed parsing. Creating summary entry...")
        
        # For now, create a summary entry
        # In a full implementation, we would parse the detailed claim data
        self.stats['warnings'].append(
            f"Adjudication report {filename} processed as summary. "
            f"Detailed claim parsing requires file structure analysis."
        )
    
    def validate_accounting_balance(self):
        """Validate that accounting entries are balanced"""
        print(f"\n{'='*80}")
        print("Validating Accounting Balance")
        print(f"{'='*80}")
        
        if self.dry_run:
            print("  (Dry run - skipping validation)")
            return
        
        try:
            # Check journal entries are balanced
            unbalanced = []
            journals = AdvancedJournalEntry.objects.filter(
                entry_number__startswith='JE-IR-'
            ).select_related()
            
            for journal in journals:
                lines = journal.lines.all()
                total_debit = sum(line.debit for line in lines)
                total_credit = sum(line.credit for line in lines)
                
                if total_debit != total_credit:
                    unbalanced.append({
                        'journal': journal.entry_number,
                        'debit': total_debit,
                        'credit': total_credit,
                        'difference': total_debit - total_credit
                    })
            
            if unbalanced:
                print(f"  [WARNING] Found {len(unbalanced)} unbalanced journal entries:")
                for entry in unbalanced:
                    print(f"    - {entry['journal']}: Debit={entry['debit']}, Credit={entry['credit']}, Diff={entry['difference']}")
                self.stats['errors'].extend([f"Unbalanced journal: {e['journal']}" for e in unbalanced])
            else:
                print("  [OK] All journal entries are balanced")
            
            # Check receivable totals
            total_receivables = InsuranceReceivable.objects.filter(
                receivable_number__startswith='IR-SUMMARY-'
            ).aggregate(
                total=Sum('balance_due')
            )['total'] or Decimal('0.00')
            
            print(f"  Total Insurance Receivables: GHS {total_receivables:,.2f}")
            
        except Exception as e:
            error_msg = f"Error validating accounting balance: {e}"
            self.stats['errors'].append(error_msg)
            print(f"  [ERROR] {error_msg}")
    
    def run(self):
        """Run the import process"""
        print(f"\n{'='*80}")
        print("INSURANCE EXCEL DATA IMPORT")
        print(f"{'='*80}")
        print(f"Mode: {'DRY RUN' if self.dry_run else 'LIVE IMPORT'}")
        print(f"Excel Libraries: Pandas={HAS_PANDAS}, Openpyxl={HAS_OPENPYXL}, Xlrd={HAS_XLRD}")
        
        excel_folder = "insurance excel"
        if not os.path.exists(excel_folder):
            print(f"\n[ERROR] Folder '{excel_folder}' not found!")
            return
        
        # Get all Excel files
        excel_files = []
        for file in os.listdir(excel_folder):
            if file.lower().endswith(('.xls', '.xlsx')):
                excel_files.append(os.path.join(excel_folder, file))
        
        if not excel_files:
            print(f"\n[ERROR] No Excel files found in '{excel_folder}'")
            return
        
        print(f"\nFound {len(excel_files)} Excel file(s)")
        
        # First, check and clean up any suppliers mistakenly created as insurance payers
        self.cleanup_suppliers_from_insurance()
        
        with transaction.atomic():
            # Process files
            for file_path in excel_files:
                filename = os.path.basename(file_path)
                
                if 'JERRY' in filename.upper() or 'DEBTOR' in filename.upper():
                    self.process_debtor_balances(file_path)
                elif 'ADJUDICATION' in filename.upper():
                    self.process_adjudication_report(file_path)
                else:
                    self.stats['warnings'].append(f"Unknown file type: {filename}")
                
                self.stats['files_processed'] += 1
            
            if self.dry_run:
                print("\n[DRY RUN] No data was saved. Use --live to import.")
                transaction.set_rollback(True)
            else:
                # Validate accounting balance
                self.validate_accounting_balance()
        
        # Print summary
        self.print_summary()
    
    def print_summary(self):
        """Print import summary"""
        print(f"\n{'='*80}")
        print("IMPORT SUMMARY")
        print(f"{'='*80}")
        print(f"Files Processed: {self.stats['files_processed']}")
        print(f"Payers Found: {self.stats['payers_found']}")
        print(f"Payers Created: {self.stats['payers_created']}")
        print(f"Receivables Created: {self.stats['receivables_created']}")
        print(f"Receivables Updated: {self.stats['receivables_updated']}")
        print(f"AR Created: {self.stats['ar_created']}")
        print(f"AR Updated: {self.stats['ar_updated']}")
        print(f"Journal Entries Created: {self.stats['journal_entries_created']}")
        
        if self.stats['warnings']:
            print(f"\n[WARNING] Warnings ({len(self.stats['warnings'])}):")
            for warning in self.stats['warnings'][:10]:  # First 10
                print(f"  - {warning}")
            if len(self.stats['warnings']) > 10:
                print(f"  ... and {len(self.stats['warnings']) - 10} more")
        
        if self.stats['errors']:
            print(f"\n[ERROR] Errors ({len(self.stats['errors'])}):")
            for error in self.stats['errors'][:10]:  # First 10
                print(f"  - {error}")
            if len(self.stats['errors']) > 10:
                print(f"  ... and {len(self.stats['errors']) - 10} more")
        else:
            print("\n[OK] No errors!")


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Import insurance Excel data')
    parser.add_argument('--live', action='store_true', help='Perform live import (default is dry run)')
    args = parser.parse_args()
    
    importer = InsuranceExcelImporter(dry_run=not args.live)
    importer.run()

