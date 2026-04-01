import openpyxl
import sys

try:
    wb = openpyxl.load_workbook('insurance excel/JERRY.xlsx')
    print('Sheet names:', wb.sheetnames)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f'\n{"="*80}')
        print(f'Sheet: {sheet_name}')
        print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
        print(f'\nFirst 20 rows:')
        
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i > 20:
                break
            # Filter out None values for cleaner output
            row_data = [str(cell) if cell is not None else '' for cell in row]
            print(f'Row {i}: {row_data}')
            
except Exception as e:
    print(f'Error: {e}')
    import traceback
    traceback.print_exc()


