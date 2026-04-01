"""
Patient Data Export Views
Export patients to Excel and PDF formats
"""

from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.utils import timezone
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from datetime import datetime

from .models import Patient


def get_all_patients(query=None, source='all'):
    """Get combined list of all patients"""
    all_patients = []
    
    # Get Django patients
    if source in ['all', 'new']:
        django_patients = Patient.objects.filter(is_deleted=False).select_related('primary_insurance')
        
        if query:
            django_patients = django_patients.filter(
                Q(first_name__icontains=query) |
                Q(last_name__icontains=query) |
                Q(middle_name__icontains=query) |
                Q(mrn__icontains=query) |
                Q(phone_number__icontains=query) |
                Q(national_id__icontains=query)
            )
        
        for p in django_patients:
            all_patients.append({
                'mrn': p.mrn,
                'full_name': p.full_name,
                'first_name': p.first_name,
                'last_name': p.last_name,
                'dob': str(p.date_of_birth),
                'age': p.age,
                'gender': p.get_gender_display(),
                'phone': p.phone_number,
                'email': p.email,
                'address': p.address,
                'blood_type': p.blood_type,
                'insurance': p.insurance_company,
                'source': 'New',
            })
    
    # Get Legacy patients
    if source in ['all', 'legacy']:
        try:
            from .models_legacy_patients import LegacyPatient
            
            legacy_patients = LegacyPatient.objects.all()
            
            if query:
                legacy_patients = legacy_patients.filter(
                    Q(fname__icontains=query) |
                    Q(lname__icontains=query) |
                    Q(mname__icontains=query) |
                    Q(pid__icontains=query) |
                    Q(phone_cell__icontains=query) |
                    Q(email__icontains=query) |
                    Q(pmc_mrn__icontains=query)
                )
            
            for p in legacy_patients:
                # Calculate age
                age = '?'
                try:
                    if p.DOB and len(str(p.DOB)) >= 4:
                        year = int(str(p.DOB)[:4])
                        age = datetime.now().year - year
                except:
                    pass
                
                all_patients.append({
                    'mrn': p.mrn_display,
                    'full_name': p.full_name,
                    'first_name': p.fname,
                    'last_name': p.lname,
                    'dob': p.DOB or '',
                    'age': age,
                    'gender': p.sex or '',
                    'phone': p.display_phone,
                    'email': p.email or '',
                    'address': f"{p.street}, {p.city}, {p.state}".strip(', '),
                    'blood_type': '',
                    'insurance': p.pricelevel or '',
                    'source': 'Legacy',
                })
        except:
            pass
    
    return all_patients


@login_required
def export_patients_excel(request):
    """Export all patients to Excel"""
    query = request.GET.get('q', '').strip()
    source = request.GET.get('source', 'all')
    
    # Get all patients
    patients = get_all_patients(query, source)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Patients'
    
    # Define styles
    header_fill = PatternFill(start_color='7B68EE', end_color='7B68EE', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:M1')
    title_cell = ws['A1']
    title_cell.value = f'PATIENT DATA EXPORT - {source.upper()} PATIENTS'
    title_cell.font = Font(bold=True, size=14, color='7B68EE')
    title_cell.alignment = Alignment(horizontal='center')
    
    # Metadata
    ws['A2'] = f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A3'] = f'Total Records: {len(patients)}'
    ws['A4'] = f'Filter: {query if query else "None"}'
    
    # Headers
    headers = [
        'MRN', 'Full Name', 'First Name', 'Last Name', 'Date of Birth',
        'Age', 'Gender', 'Phone', 'Email', 'Address',
        'Blood Type', 'Insurance/Price Level', 'Source'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    for row_num, patient in enumerate(patients, 7):
        ws.cell(row=row_num, column=1, value=patient['mrn']).border = border
        ws.cell(row=row_num, column=2, value=patient['full_name']).border = border
        ws.cell(row=row_num, column=3, value=patient['first_name']).border = border
        ws.cell(row=row_num, column=4, value=patient['last_name']).border = border
        ws.cell(row=row_num, column=5, value=patient['dob']).border = border
        ws.cell(row=row_num, column=6, value=str(patient['age'])).border = border
        ws.cell(row=row_num, column=7, value=patient['gender']).border = border
        ws.cell(row=row_num, column=8, value=patient['phone']).border = border
        ws.cell(row=row_num, column=9, value=patient['email']).border = border
        ws.cell(row=row_num, column=10, value=patient['address']).border = border
        ws.cell(row=row_num, column=11, value=patient['blood_type']).border = border
        ws.cell(row=row_num, column=12, value=patient['insurance']).border = border
        ws.cell(row=row_num, column=13, value=patient['source']).border = border
    
    # Auto-adjust column widths
    for col_num in range(1, 14):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'patients_{source}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
def export_patients_pdf(request):
    """Export all patients to PDF"""
    query = request.GET.get('q', '').strip()
    source = request.GET.get('source', 'all')
    
    # Get all patients
    patients = get_all_patients(query, source)
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#7B68EE'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    # Title
    title = Paragraph(f'<b>PATIENT DATA EXPORT - {source.upper()} PATIENTS</b>', title_style)
    elements.append(title)
    
    # Metadata
    metadata_style = ParagraphStyle('Metadata', parent=styles['Normal'], fontSize=9, textColor=colors.gray)
    elements.append(Paragraph(f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}', metadata_style))
    elements.append(Paragraph(f'Total Records: {len(patients)}', metadata_style))
    if query:
        elements.append(Paragraph(f'Filter: {query}', metadata_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Table data
    table_data = [[
        'MRN', 'Full Name', 'DOB', 'Age', 'Gender',
        'Phone', 'Email', 'Insurance', 'Source'
    ]]
    
    for patient in patients:
        table_data.append([
            patient['mrn'],
            patient['full_name'][:30],  # Truncate long names
            patient['dob'][:10] if patient['dob'] else '',
            str(patient['age']),
            patient['gender'],
            patient['phone'][:15],
            patient['email'][:25],
            patient['insurance'][:20],
            patient['source']
        ])
    
    # Create table
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        # Header style
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7B68EE')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        
        # Data style
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Borders
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        
        # Alternating rows
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    filename = f'patients_{source}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
def export_patients_csv(request):
    """Export all patients to CSV"""
    import csv
    query = request.GET.get('q', '').strip()
    source = request.GET.get('source', 'all')
    
    # Get all patients
    patients = get_all_patients(query, source)
    
    # Create CSV
    response = HttpResponse(content_type='text/csv')
    filename = f'patients_{source}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Headers
    writer.writerow([
        'MRN', 'Full Name', 'First Name', 'Last Name', 'Date of Birth',
        'Age', 'Gender', 'Phone', 'Email', 'Address',
        'Blood Type', 'Insurance/Price Level', 'Source'
    ])
    
    # Data
    for patient in patients:
        writer.writerow([
            patient['mrn'],
            patient['full_name'],
            patient['first_name'],
            patient['last_name'],
            patient['dob'],
            patient['age'],
            patient['gender'],
            patient['phone'],
            patient['email'],
            patient['address'],
            patient['blood_type'],
            patient['insurance'],
            patient['source'],
        ])
    
    return response




















