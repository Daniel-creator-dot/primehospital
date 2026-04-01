"""
Export lab results to Excel
"""
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Q
from django.utils import timezone
from django.conf import settings

from .models import LabResult

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


@login_required
def export_lab_results_excel(request):
    """Export lab results to Excel"""
    if not EXCEL_AVAILABLE:
        return HttpResponse(
            "Excel export requires openpyxl. Please install it: pip install openpyxl",
            content_type='text/plain',
            status=500
        )
    
    # Get filters from request
    status_filter = request.GET.get('status', '')
    query = request.GET.get('q', '')
    
    # Get lab results with same filters as list view
    results = LabResult.objects.filter(is_deleted=False).select_related(
        'test', 'order__encounter__patient', 'verified_by', 'order__encounter'
    )
    
    if status_filter:
        results = results.filter(status=status_filter)
    
    if query:
        results = results.filter(
            Q(test__name__icontains=query) |
            Q(order__encounter__patient__first_name__icontains=query) |
            Q(order__encounter__patient__last_name__icontains=query) |
            Q(order__encounter__patient__mrn__icontains=query)
        )
    
    # Get all results (not limited like list view)
    results_list = results.order_by('-created')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Lab Results'
    
    # Define styles
    header_fill = PatternFill(start_color='7B68EE', end_color='7B68EE', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = 'LABORATORY RESULTS EXPORT'
    title_cell.font = Font(bold=True, size=14, color='7B68EE')
    title_cell.alignment = Alignment(horizontal='center')
    
    # Metadata
    ws['A2'] = f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A3'] = f'Total Records: {results_list.count()}'
    ws['A4'] = f'Status Filter: {status_filter if status_filter else "All"}'
    ws['A5'] = f'Search Query: {query if query else "None"}'
    
    # Headers
    headers = [
        'Date', 'Time', 'Patient Name', 'MRN', 'Test Name', 'Test Code',
        'Status', 'Result Value', 'Units', 'Abnormal', 'Verified By', 'Notes'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    row_num = 8
    for result in results_list:
        patient = result.order.encounter.patient if result.order and result.order.encounter else None
        patient_name = patient.full_name if patient else 'N/A'
        mrn = patient.mrn if patient else 'N/A'
        
        # Format result value
        result_value = result.value or ''
        if result.details:
            # For panel tests, combine details
            detail_str = ', '.join([f"{k}: {v}" for k, v in result.details.items()])
            if detail_str:
                result_value = detail_str
        if result.qualitative_result:
            result_value = result.qualitative_result
        
        # Format date and time
        result_date = result.created.date() if result.created else ''
        result_time = result.created.time() if result.created else ''
        
        data = [
            result_date,
            result_time,
            patient_name,
            mrn,
            result.test.name if result.test else 'N/A',
            result.test.code if result.test else 'N/A',
            result.get_status_display(),
            result_value,
            result.units or '',
            'Yes' if result.is_abnormal else 'No',
            result.verified_by.user.get_full_name() if result.verified_by and result.verified_by.user else 'N/A',
            result.notes or ''
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Highlight abnormal results
            if result.is_abnormal and col_num == 10:  # Abnormal column
                cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
        
        row_num += 1
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'lab_results_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response










