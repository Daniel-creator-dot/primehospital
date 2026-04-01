"""
Professional Price Import System - ALL SERVICES
Imports consultation, lab test, and treatment prices from Excel file
Handles Cash, Corporate, Insurance, and all insurance company prices
"""
import os
import sys
from decimal import Decimal, InvalidOperation
from django.core.management.base import BaseCommand
from django.utils import timezone
from django.db import transaction
from openpyxl import load_workbook

from hospital.models import ServiceCode
from hospital.models_flexible_pricing import PricingCategory, ServicePrice


class Command(BaseCommand):
    help = 'Import ALL prices (Consultations, Lab Tests, Treatments) from Excel file'
    
    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default='hms/prices/Consult Price List 2025(1).xlsx',
            help='Path to Excel file'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Run without saving to database (validation only)'
        )
        parser.add_argument(
            '--verbose',
            action='store_true',
            help='Show detailed progress'
        )
        parser.add_argument(
            '--section',
            type=str,
            choices=['consultations', 'lab', 'treatments', 'all'],
            default='all',
            help='Which section to import (default: all)'
        )
    
    def handle(self, *args, **options):
        file_path = options['file']
        dry_run = options['dry_run']
        verbose = options['verbose']
        section = options['section']
        
        self.stdout.write(self.style.SUCCESS('=' * 80))
        self.stdout.write(self.style.SUCCESS('PROFESSIONAL PRICE IMPORT SYSTEM - ALL SERVICES'))
        self.stdout.write(self.style.SUCCESS('=' * 80))
        self.stdout.write('')
        
        if dry_run:
            self.stdout.write(self.style.WARNING('DRY RUN MODE - No changes will be saved'))
            self.stdout.write('')
        
        # Validate file exists
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'ERROR: File not found: {file_path}'))
            sys.exit(1)
        
        # Load workbook
        self.stdout.write(f'Loading Excel file: {file_path}')
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            self.stdout.write(self.style.SUCCESS(f'✓ File loaded successfully'))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'ERROR: Could not load Excel file: {e}'))
            sys.exit(1)
        
        # Find sections
        sections = self._find_sections(ws)
        self.stdout.write(f'\nFound {len(sections)} sections in Excel file:')
        for sec in sections:
            self.stdout.write(f'  - {sec["type"]} (starts at row {sec["row"]})')
        
        # Ensure pricing categories exist
        self.stdout.write('')
        self.stdout.write('Setting up pricing categories...')
        categories = self._ensure_categories(ws, dry_run, verbose)
        
        total_stats = {
            'services_processed': 0,
            'services_created': 0,
            'services_updated': 0,
            'prices_created': 0,
            'prices_updated': 0,
            'errors': []
        }
        
        # Process each section
        for sec in sections:
            if section != 'all' and sec['type'].lower() != section:
                continue
            
            self.stdout.write('')
            self.stdout.write(self.style.SUCCESS(f'Processing {sec["type"]}...'))
            self.stdout.write('-' * 80)
            
            # Parse header for this section
            headers = self._parse_headers(ws, sec['row'])
            if not headers:
                self.stdout.write(self.style.WARNING(f'  ⚠️ Could not parse headers for {sec["type"]}'))
                continue
            
            # Identify columns
            column_map = self._identify_columns(headers, sec['type'])
            
            # Import prices for this section
            stats = self._import_section(
                ws, sec, headers, column_map, categories, dry_run, verbose
            )
            
            # Accumulate stats
            for key in total_stats:
                if key == 'errors':
                    total_stats[key].extend(stats.get(key, []))
                else:
                    total_stats[key] += stats.get(key, 0)
        
        # Final summary
        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS('=' * 80))
        self.stdout.write(self.style.SUCCESS('FINAL IMPORT SUMMARY'))
        self.stdout.write(self.style.SUCCESS('=' * 80))
        self.stdout.write(f'  Services processed: {total_stats["services_processed"]}')
        self.stdout.write(f'  Services created: {total_stats["services_created"]}')
        self.stdout.write(f'  Services updated: {total_stats["services_updated"]}')
        self.stdout.write(f'  Prices created: {total_stats["prices_created"]}')
        self.stdout.write(f'  Prices updated: {total_stats["prices_updated"]}')
        self.stdout.write(f'  Errors: {len(total_stats["errors"])}')
        
        if total_stats['errors'] and verbose:
            self.stdout.write('')
            self.stdout.write('Errors:')
            for error in total_stats['errors'][:20]:
                self.stdout.write(f'  {error}')
            if len(total_stats['errors']) > 20:
                self.stdout.write(f'  ... and {len(total_stats["errors"]) - 20} more')
        
        if dry_run:
            self.stdout.write('')
            self.stdout.write(self.style.WARNING('DRY RUN COMPLETE - No data was saved'))
        else:
            self.stdout.write('')
            self.stdout.write(self.style.SUCCESS('✓ Import completed successfully!'))
            self.stdout.write('')
            self.stdout.write('You can now view prices at: /hms/pricing/')
    
    def _find_sections(self, ws):
        """Find all sections in the worksheet"""
        sections = []
        
        for row_idx, row in enumerate(ws.iter_rows(), 1):
            if not row or not row[0] or not row[0].value:
                continue
            
            first_cell = str(row[0].value).strip()
            
            if first_cell == 'No.':
                # Check column 5 for section type
                if len(row) > 4 and row[4].value:
                    col5 = str(row[4].value).strip()
                    if 'SpecialistTypeID' in col5 or 'SpecialistType' in col5:
                        sections.append({
                            'type': 'Consultations',
                            'row': row_idx,
                            'category': 'Consultation'
                        })
                    elif 'LabTestID' in col5 or 'LabTest' in col5:
                        sections.append({
                            'type': 'Lab Tests',
                            'row': row_idx,
                            'category': 'Lab Test'
                        })
                    elif 'TreatmentID' in col5 or 'Treatment' in col5:
                        sections.append({
                            'type': 'Treatments',
                            'row': row_idx,
                            'category': 'Treatment'
                        })
        
        return sections
    
    def _parse_headers(self, ws, header_row):
        """Parse header row"""
        headers = []
        for cell in ws[header_row]:
            headers.append(cell.value if cell.value else "")
        return headers
    
    def _identify_columns(self, headers, section_type):
        """Identify important columns"""
        column_map = {
            'id_col': None,
            'name_col': None,
            'cash': None,
            'corporate': None,
            'insurance': None,
            'insurance_companies': {}
        }
        
        for idx, header in enumerate(headers):
            if not header:
                continue
            
            header_str = str(header).lower()
            
            # ID and Name columns (different for each section)
            if section_type == 'Consultations':
                if 'specialisttypeid' in header_str.replace(' ', ''):
                    column_map['id_col'] = idx
                elif 'specialisttypename' in header_str.replace(' ', ''):
                    column_map['name_col'] = idx
                elif 'visit type' in header_str:
                    column_map['visit_type_col'] = idx
            elif section_type == 'Lab Tests':
                if 'labtestid' in header_str.replace(' ', ''):
                    column_map['id_col'] = idx
                elif 'labtestname' in header_str.replace(' ', ''):
                    column_map['name_col'] = idx
            elif section_type == 'Treatments':
                if 'treatmentid' in header_str.replace(' ', ''):
                    column_map['id_col'] = idx
                elif 'treatmentname' in header_str.replace(' ', ''):
                    column_map['name_col'] = idx
            
            # Price columns (same for all sections)
            if ('cash' in header_str or 'private' in header_str or 
                '100 percent' in header_str or 'mark-up' in header_str):
                if column_map['cash'] is None:
                    column_map['cash'] = idx
            elif ('company' in header_str or 'corporate' in header_str or 
                  'coperate' in header_str):
                if column_map['corporate'] is None:
                    column_map['corporate'] = idx
            elif 'insurance' in header_str and 'other' not in header_str:
                if column_map['insurance'] is None:
                    column_map['insurance'] = idx
            elif '[' in header_str and ']' in header_str:
                # Insurance company column
                column_map['insurance_companies'][idx] = header
        
        return column_map
    
    def _ensure_categories(self, ws, dry_run, verbose):
        """Ensure pricing categories exist"""
        categories = {}
        
        # Cash category
        cash_category, created = self._get_or_create_category(
            name='Cash / Private Patients',
            code='CASH',
            category_type='cash',
            color='#10b981',
            dry_run=dry_run
        )
        categories['cash'] = cash_category
        
        # Corporate category
        corporate_category, created = self._get_or_create_category(
            name='Corporate / Company',
            code='CORPORATE',
            category_type='corporate',
            color='#8b5cf6',
            dry_run=dry_run
        )
        categories['corporate'] = corporate_category
        
        # General Insurance category
        insurance_category, created = self._get_or_create_category(
            name='Insurance (General)',
            code='INSURANCE',
            category_type='insurance',
            color='#3b82f6',
            dry_run=dry_run
        )
        categories['insurance'] = insurance_category
        
        # Get insurance companies from first section header
        headers = self._parse_headers(ws, 1)
        for idx, header in enumerate(headers):
            if header and '[' in str(header) and ']' in str(header):
                company_name = str(header).split(']')[-1].strip()
                if company_name and company_name not in ['COMPANY', 'COMPANY(coperate)']:
                    code = f"INS_{company_name.upper().replace(' ', '_')[:20]}"
                    category, created = self._get_or_create_category(
                        name=f'Insurance - {company_name}',
                        code=code,
                        category_type='insurance',
                        color='#3b82f6',
                        dry_run=dry_run
                    )
                    categories[f'insurance_{idx}'] = category
                    if created and verbose:
                        self.stdout.write(f'  ✓ Created category: Insurance - {company_name}')
        
        return categories
    
    def _get_or_create_category(self, name, code, category_type, color, dry_run):
        """Get or create pricing category"""
        if dry_run:
            class MockCategory:
                def __init__(self, name, code):
                    self.name = name
                    self.code = code
            return MockCategory(name, code), True
        
        category, created = PricingCategory.objects.get_or_create(
            code=code,
            defaults={
                'name': name,
                'category_type': category_type,
                'color_code': color,
                'is_active': True,
                'priority': 100 if category_type == 'cash' else 200
            }
        )
        return category, created
    
    @transaction.atomic
    def _import_section(self, ws, section, headers, column_map, categories, dry_run, verbose):
        """Import prices for a section"""
        stats = {
            'services_processed': 0,
            'services_created': 0,
            'services_updated': 0,
            'prices_created': 0,
            'prices_updated': 0,
            'errors': []
        }
        
        data_start_row = section['row'] + 1
        next_section_row = ws.max_row + 1
        
        # Find next section
        for row_idx in range(data_start_row, ws.max_row + 1):
            row = ws[row_idx]
            if row and row[0] and str(row[0].value).strip() == 'No.':
                next_section_row = row_idx
                break
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, max_row=next_section_row-1), data_start_row):
            if not row or not row[0] or not row[0].value:
                continue
            
            if row[0].value and isinstance(row[0].value, str) and row[0].value.startswith('No.'):
                continue
            
            try:
                row_data = self._extract_row_data(row, column_map, section['type'])
                
                if not row_data.get('name'):
                    continue
                
                stats['services_processed'] += 1
                
                # Create or update service code
                service_code, service_created = self._get_or_create_service(
                    row_data, section['category'], dry_run
                )
                
                if service_created:
                    stats['services_created'] += 1
                else:
                    stats['services_updated'] += 1
                
                # Import prices
                price_stats = self._import_service_prices(
                    service_code, row_data, column_map, categories, dry_run
                )
                
                stats['prices_created'] += price_stats['created']
                stats['prices_updated'] += price_stats['updated']
                
                if verbose and stats['services_processed'] % 50 == 0:
                    self.stdout.write(f'  Processed {stats["services_processed"]} services...')
            
            except Exception as e:
                error_msg = f'Row {row_idx}: {str(e)}'
                stats['errors'].append(error_msg)
                if verbose:
                    self.stdout.write(self.style.ERROR(f'  {error_msg}'))
                # Log to Django logger for debugging
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error importing row {row_idx}: {str(e)}", exc_info=True)
        
        self.stdout.write(f'  ✓ {section["type"]}: {stats["services_processed"]} services, '
                         f'{stats["prices_created"] + stats["prices_updated"]} prices')
        
        return stats
    
    def _extract_row_data(self, row, column_map, section_type):
        """Extract data from row"""
        def get_value(idx):
            if idx is None or idx >= len(row):
                return None
            val = row[idx].value
            return val if val not in (None, '', '0') else None
        
        def get_float(idx):
            val = get_value(idx)
            if val is None:
                return None
            try:
                return float(val)
            except (ValueError, TypeError):
                return None
        
        data = {
            'id': get_value(column_map.get('id_col')),
            'name': get_value(column_map.get('name_col')),
            'cash_price': get_float(column_map.get('cash')),
            'corporate_price': get_float(column_map.get('corporate')),
            'insurance_price': get_float(column_map.get('insurance')),
            'insurance_companies': {}
        }
        
        # Add visit type for consultations
        if section_type == 'Consultations':
            data['visit_type'] = get_value(column_map.get('visit_type_col'))
        
        # Get insurance company prices
        for idx in column_map['insurance_companies'].keys():
            price = get_float(idx)
            if price is not None:
                data['insurance_companies'][idx] = price
        
        return data
    
    def _get_or_create_service(self, row_data, category, dry_run):
        """Get or create service code"""
        name = str(row_data['name']).strip()
        service_id = row_data.get('id', '')
        
        # Generate service code
        if service_id:
            service_code_str = str(service_id)[:20]
        else:
            # Generate from name
            words = name.split()[:3]
            code = ''.join([w[0].upper() for w in words])[:15]
            service_code_str = f"{category[:3].upper()}_{code}"[:20]
        
        # Build description
        if row_data.get('visit_type'):
            description = f"{name} - {row_data['visit_type']}"
        else:
            description = name
        
        if dry_run:
            class MockService:
                def __init__(self, code, description):
                    self.code = code
                    self.description = description
            return MockService(service_code_str, description), True
        
        service, created = ServiceCode.objects.get_or_create(
            code=service_code_str,
            defaults={
                'description': description,
                'category': category,
                'is_active': True
            }
        )
        
        if not created and service.description != description:
            service.description = description
            service.save()
        
        return service, created
    
    def _import_service_prices(self, service_code, row_data, column_map, categories, dry_run):
        """Import prices for a service"""
        stats = {'created': 0, 'updated': 0}
        
        # Cash price
        if row_data['cash_price']:
            stats.update(self._create_or_update_price(
                service_code, categories['cash'], row_data['cash_price'], dry_run
            ))
        
        # Corporate price
        if row_data['corporate_price']:
            stats.update(self._create_or_update_price(
                service_code, categories['corporate'], row_data['corporate_price'], dry_run
            ))
        
        # General insurance price
        if row_data['insurance_price']:
            stats.update(self._create_or_update_price(
                service_code, categories['insurance'], row_data['insurance_price'], dry_run
            ))
        
        # Insurance company prices
        for idx, price in row_data['insurance_companies'].items():
            category_key = f'insurance_{idx}'
            if category_key in categories:
                stats.update(self._create_or_update_price(
                    service_code, categories[category_key], price, dry_run
                ))
        
        return stats
    
    def _create_or_update_price(self, service_code, category, price, dry_run):
        """Create or update a service price"""
        stats = {'created': 0, 'updated': 0}
        
        if dry_run:
            return {'created': 1, 'updated': 0}
        
        try:
            price_decimal = Decimal(str(price))
            
            # Check for existing price with same effective_from date
            effective_date = timezone.now().date()
            existing_price = ServicePrice.objects.filter(
                service_code=service_code,
                pricing_category=category,
                effective_from=effective_date,
                is_deleted=False
            ).first()
            
            if existing_price:
                # Update existing price
                existing_price.price = price_decimal
                existing_price.is_active = True
                existing_price.notes = f'Imported from Excel - {timezone.now().strftime("%Y-%m-%d")}'
                existing_price.save()
                stats['updated'] = 1
            else:
                # Create new price
                ServicePrice.objects.create(
                    service_code=service_code,
                    pricing_category=category,
                    effective_from=effective_date,
                    price=price_decimal,
                    is_active=True,
                    notes=f'Imported from Excel - {timezone.now().strftime("%Y-%m-%d")}'
                )
                stats['created'] = 1
        
        except (InvalidOperation, ValueError) as e:
            # Log invalid price but continue
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Invalid price value '{price}' for service {service_code.code}: {str(e)}")
        except Exception as e:
            # Log other errors but continue
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error creating/updating price for service {service_code.code}: {str(e)}", exc_info=True)
        
        return stats








