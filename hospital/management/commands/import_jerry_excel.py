"""
Import Debtors and Creditors from JERRY.xlsx Excel file
- Creates/updates Payer records for insurance companies (debtors) with payer_type=private/insurance, NOT corporate
- Creates/updates Supplier records for creditors (NOT as insurance)
- Creates InsuranceReceivableEntry records for debtors with balances
- Creates AccountsPayable records for creditors with balances
- Creates accounting journal entries for both
"""

import os
import sys
from decimal import Decimal
from datetime import datetime, timedelta
from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils import timezone
from django.contrib.auth.models import User

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from hospital.models import Payer
from hospital.models_missing_features import Supplier
from hospital.models_primecare_accounting import InsuranceReceivableEntry
from hospital.models_accounting import Account
from hospital.models_accounting_advanced import (
    AccountsPayable, Journal, AdvancedJournalEntry, AdvancedJournalEntryLine
)


class Command(BaseCommand):
    help = 'Import debtors and creditors from JERRY.xlsx Excel file'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default='insurance excel/JERRY.xlsx',
            help='Path to JERRY.xlsx file'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Run without making changes (preview only)'
        )

    def handle(self, *args, **options):
        if not HAS_OPENPYXL:
            self.stdout.write(
                self.style.ERROR('openpyxl is required. Install with: pip install openpyxl')
            )
            return

        file_path = options['file']
        dry_run = options['dry_run']

        if not os.path.exists(file_path):
            self.stdout.write(
                self.style.ERROR(f'File not found: {file_path}')
            )
            return

        if dry_run:
            self.stdout.write(self.style.WARNING('DRY RUN MODE - No changes will be saved'))

        self.stdout.write('='*80)
        self.stdout.write('IMPORTING DEBTORS AND CREDITORS FROM JERRY.XLSX')
        self.stdout.write('='*80)
        self.stdout.write('')

        # Load Excel file
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Error loading Excel file: {e}'))
            return

        # Extract data
        debtors = self.extract_debtors(wb)
        creditors = self.extract_creditors(wb)

        self.stdout.write(f'Found {len(debtors)} debtors (insurance companies)')
        self.stdout.write(f'Found {len(creditors)} creditors (suppliers)')
        self.stdout.write('')

        if dry_run:
            self.stdout.write(self.style.WARNING('DRY RUN - Would create/update:'))
            for debtor in debtors:
                self.stdout.write(f'  Payer: {debtor["name"]} - Balance: GHS {debtor["balance"]:,.2f}')
            for creditor in creditors:
                self.stdout.write(f'  Supplier: {creditor["name"]} - Balance: GHS {creditor["balance"]:,.2f}')
            return

        # Get or create system user
        user = User.objects.filter(is_superuser=True).first()
        if not user:
            user = User.objects.first()
        if not user:
            self.stdout.write(self.style.ERROR('No user found. Please create a superuser first.'))
            return

        # Process debtors (insurance companies)
        self.stdout.write('Processing debtors (insurance companies)...')
        payer_results = self.process_debtors(debtors, user)
        
        # Process creditors (suppliers)
        self.stdout.write('Processing creditors (suppliers)...')
        supplier_results = self.process_creditors(creditors, user)

        # Summary
        self.stdout.write('')
        self.stdout.write('='*80)
        self.stdout.write('IMPORT SUMMARY')
        self.stdout.write('='*80)
        self.stdout.write(f'Payers created: {payer_results["created"]}')
        self.stdout.write(f'Payers updated: {payer_results["updated"]}')
        self.stdout.write(f'Insurance Receivable Entries created: {payer_results["receivable_entries"]}')
        self.stdout.write(f'Suppliers created: {supplier_results["created"]}')
        self.stdout.write(f'Suppliers updated: {supplier_results["updated"]}')
        self.stdout.write(f'Accounts Payable entries created: {supplier_results["ap_entries"]}')
        self.stdout.write(f'Journal entries created: {payer_results["journal_entries"] + supplier_results["journal_entries"]}')
        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS('Import completed successfully!'))

    def extract_debtors(self, wb):
        """Extract debtor balances (insurance companies)"""
        debtors = []
        
        if 'DEBTOR BALANCES' not in wb.sheetnames:
            self.stdout.write(self.style.WARNING('DEBTOR BALANCES sheet not found'))
            return debtors

        ws = wb['DEBTOR BALANCES']
        
        # Find data start row (skip headers)
        data_start = None
        for row_idx in range(1, min(35, ws.max_row + 1)):
            row = [cell.value for cell in ws[row_idx]]
            if row and len(row) >= 2:
                # Check if row has a name and a numeric balance
                name = row[0] if row[0] else None
                balance = row[1] if len(row) > 1 else None
                
                if name and isinstance(name, str) and balance is not None:
                    try:
                        if isinstance(balance, (int, float)):
                            data_start = row_idx
                            break
                    except:
                        pass

        if not data_start:
            # Try default start at row 5
            data_start = 5

        # Extract data
        for row_idx in range(data_start, ws.max_row + 1):
            row = [cell.value for cell in ws[row_idx]]
            if not row or len(row) < 2:
                continue

            company_name = row[0] if row[0] else None
            balance = row[1] if len(row) > 1 else None

            if not company_name or not isinstance(company_name, str):
                continue

            company_name = company_name.strip()
            if not company_name:
                continue

            # Parse balance
            balance_value = None
            if balance is not None:
                try:
                    if isinstance(balance, (int, float)):
                        balance_value = Decimal(str(balance))
                    elif isinstance(balance, str):
                        balance_str = balance.replace('GHC', '').replace('GHS', '').replace(',', '').strip()
                        if balance_str:
                            balance_value = Decimal(balance_str)
                except:
                    pass

            if balance_value is None or balance_value == Decimal('0.00'):
                continue

            debtors.append({
                'name': company_name,
                'balance': balance_value
            })

        return debtors

    def extract_creditors(self, wb):
        """Extract creditor balances (suppliers)"""
        creditors = []
        
        if 'CREDITOR BALANCES' not in wb.sheetnames:
            self.stdout.write(self.style.WARNING('CREDITOR BALANCES sheet not found'))
            return creditors

        ws = wb['CREDITOR BALANCES']
        
        # Find data start row
        data_start = None
        for row_idx in range(1, min(25, ws.max_row + 1)):
            row = [cell.value for cell in ws[row_idx]]
            if row and len(row) >= 3:
                supplier_name = row[1] if row[1] else None
                balance = row[2] if len(row) > 2 else None
                
                if supplier_name and isinstance(supplier_name, str) and balance is not None:
                    try:
                        if isinstance(balance, (int, float)):
                            data_start = row_idx
                            break
                    except:
                        pass

        if not data_start:
            # Try default start at row 3
            data_start = 3

        # Extract data
        for row_idx in range(data_start, ws.max_row + 1):
            row = [cell.value for cell in ws[row_idx]]
            if not row or len(row) < 3:
                continue

            supplier_name = row[1] if row[1] else None
            balance = row[2] if len(row) > 2 else None

            if not supplier_name or not isinstance(supplier_name, str):
                continue

            supplier_name = supplier_name.strip()
            if not supplier_name:
                continue

            # Parse balance
            balance_value = None
            if balance is not None:
                try:
                    if isinstance(balance, (int, float)):
                        balance_value = Decimal(str(balance))
                    elif isinstance(balance, str):
                        balance_str = balance.replace('GHC', '').replace('GHS', '').replace(',', '').strip()
                        if balance_str:
                            balance_value = Decimal(balance_str)
                except:
                    pass

            if balance_value is None or balance_value == Decimal('0.00'):
                continue

            creditors.append({
                'name': supplier_name,
                'balance': balance_value
            })

        return creditors

    @transaction.atomic
    def process_debtors(self, debtors, user):
        """Process debtors - create/update Payers and Insurance Receivable Entries"""
        results = {
            'created': 0,
            'updated': 0,
            'receivable_entries': 0,
            'journal_entries': 0
        }

        # Ensure default accounts exist
        # Use 1201 for Insurance Receivables (as per balance sheet)
        ar_account = self.get_or_create_account('1201', 'Insurance Receivables', 'asset')
        # Use consultation revenue account (4110) as per InsuranceReceivableEntry.match_to_revenue
        revenue_account = self.get_or_create_account('4110', 'Consultation Revenue', 'revenue')

        # Get or create sales journal
        journal, _ = Journal.objects.get_or_create(
            code='SALES',
            defaults={'name': 'Sales Journal', 'journal_type': 'sales'}
        )

        for debtor in debtors:
            name = debtor['name']
            balance = debtor['balance']

            # Create or update Payer - ensure payer_type is 'private' (NOT corporate)
            payer, created = Payer.objects.get_or_create(
                name=name,
                defaults={
                    'payer_type': 'private',  # Private Insurance, NOT corporate
                    'is_active': True
                }
            )

            if created:
                results['created'] += 1
                self.stdout.write(f'  Created Payer: {name} (type: private)')
            else:
                # Update existing payer to ensure it's private/insurance, NOT corporate
                if payer.payer_type == 'corporate':
                    payer.payer_type = 'private'
                    payer.is_active = True
                    payer.save()
                    results['updated'] += 1
                    self.stdout.write(f'  Updated Payer: {name} (changed from corporate to private)')
                else:
                    # Just ensure it's active
                    if not payer.is_active:
                        payer.is_active = True
                        payer.save()
                        results['updated'] += 1
                        self.stdout.write(f'  Updated Payer: {name} (activated)')

            # Create Insurance Receivable Entry if balance > 0
            if balance > 0:
                # Check if entry already exists for this payer with this balance
                existing_entry = InsuranceReceivableEntry.objects.filter(
                    payer=payer,
                    total_amount=balance,
                    status='pending'
                ).first()

                if not existing_entry:
                    # For opening balances, we'll allocate to consultation revenue as default
                    receivable_entry = InsuranceReceivableEntry.objects.create(
                        payer=payer,
                        entry_date=timezone.now().date(),
                        total_amount=balance,
                        consultation_amount=balance,  # Default to consultation revenue
                        outstanding_amount=balance,
                        status='pending',
                        notes=f'Imported from JERRY.xlsx - Opening balance'
                    )
                    results['receivable_entries'] += 1
                    self.stdout.write(f'    Created Insurance Receivable Entry: {receivable_entry.entry_number} - GHS {balance:,.2f}')

                    # Create journal entry for insurance receivable
                    je = self.create_receivable_journal_entry(
                        receivable_entry, payer, balance, ar_account, revenue_account, journal, user
                    )
                    if je:
                        receivable_entry.journal_entry = je
                        receivable_entry.save()
                        results['journal_entries'] += 1
                        self.stdout.write(f'    Created Journal Entry: {je.entry_number}')
                else:
                    self.stdout.write(f'    Skipped - Receivable entry already exists for {name}')

        return results

    @transaction.atomic
    def process_creditors(self, creditors, user):
        """Process creditors - create/update Suppliers and Accounts Payable"""
        results = {
            'created': 0,
            'updated': 0,
            'ap_entries': 0,
            'journal_entries': 0
        }

        # Ensure default accounts exist
        ap_account = self.get_or_create_account('2000', 'Accounts Payable', 'liability')
        expense_account = self.get_or_create_account('5100', 'Operating Expenses', 'expense')

        # Get or create purchase journal
        journal, _ = Journal.objects.get_or_create(
            code='PURCHASE',
            defaults={'name': 'Purchase Journal', 'journal_type': 'purchase'}
        )

        for creditor in creditors:
            name = creditor['name']
            balance = creditor['balance']

            # Create or update Supplier - ensure they are NOT added as insurance
            supplier, created = Supplier.objects.get_or_create(
                name=name,
                defaults={
                    'is_active': True
                }
            )

            if created:
                results['created'] += 1
                self.stdout.write(f'  Created Supplier: {name}')
            else:
                if not supplier.is_active:
                    supplier.is_active = True
                    supplier.save()
                    results['updated'] += 1
                    self.stdout.write(f'  Updated Supplier: {name} (activated)')

            # Create Accounts Payable entry if balance > 0
            if balance > 0:
                # Generate unique bill number
                bill_number = self.generate_bill_number()

                # Check if AP entry already exists
                existing_ap = AccountsPayable.objects.filter(
                    vendor_name=name,
                    amount=balance,
                    balance_due=balance
                ).first()

                if not existing_ap:
                    due_date = timezone.now().date() + timedelta(days=30)  # Default 30 days

                    ap_entry = AccountsPayable.objects.create(
                        bill_number=bill_number,
                        vendor_name=name,
                        vendor_invoice=f'INV-{bill_number}',
                        bill_date=timezone.now().date(),
                        due_date=due_date,
                        amount=balance,
                        amount_paid=Decimal('0.00'),
                        balance_due=balance,
                        description=f'Opening balance from JERRY.xlsx - {name}',
                        supply_type='goods',
                        supplier_is_exempted=False
                    )
                    results['ap_entries'] += 1
                    self.stdout.write(f'    Created Accounts Payable: {ap_entry.bill_number} - GHS {balance:,.2f}')

                    # Create journal entry for accounts payable
                    je = self.create_payable_journal_entry(
                        ap_entry, balance, expense_account, ap_account, journal, user
                    )
                    if je:
                        ap_entry.journal_entry = je
                        ap_entry.save()
                        results['journal_entries'] += 1
                        self.stdout.write(f'    Created Journal Entry: {je.entry_number}')
                else:
                    self.stdout.write(f'    Skipped - AP entry already exists for {name}')

        return results

    def get_or_create_account(self, code, name, account_type):
        """Get or create an account"""
        account, _ = Account.objects.get_or_create(
            account_code=code,
            defaults={
                'account_name': name,
                'account_type': account_type,
                'is_active': True
            }
        )
        return account

    def create_receivable_journal_entry(self, receivable_entry, payer, amount, ar_account, revenue_account, journal, user):
        """Create journal entry for insurance receivable"""
        try:
            je = AdvancedJournalEntry.objects.create(
                journal=journal,
                entry_date=receivable_entry.entry_date,
                description=f'Insurance Receivable - {payer.name} - {receivable_entry.entry_number}',
                reference=receivable_entry.entry_number,
                created_by=user,
                status='draft',
                total_debit=amount,
                total_credit=amount
            )

            # Debit: Accounts Receivable (increase asset)
            AdvancedJournalEntryLine.objects.create(
                journal_entry=je,
                line_number=1,
                account=ar_account,
                description=f'Accounts Receivable - {payer.name}',
                debit_amount=amount,
                credit_amount=Decimal('0.00')
            )

            # Credit: Revenue (increase revenue)
            AdvancedJournalEntryLine.objects.create(
                journal_entry=je,
                line_number=2,
                account=revenue_account,
                description=f'Revenue - {payer.name}',
                debit_amount=Decimal('0.00'),
                credit_amount=amount
            )

            # Post the journal entry
            je.post(user)

            return je
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'    Error creating journal entry: {e}'))
            return None

    def create_payable_journal_entry(self, ap_entry, amount, expense_account, ap_account, journal, user):
        """Create journal entry for accounts payable"""
        try:
            je = AdvancedJournalEntry.objects.create(
                journal=journal,
                entry_date=ap_entry.bill_date,
                description=f'Accounts Payable - {ap_entry.vendor_name} - {ap_entry.bill_number}',
                reference=ap_entry.bill_number,
                created_by=user,
                status='draft',
                total_debit=amount,
                total_credit=amount
            )

            # Debit: Expense (increase expense)
            AdvancedJournalEntryLine.objects.create(
                journal_entry=je,
                line_number=1,
                account=expense_account,
                description=f'Expense - {ap_entry.vendor_name}',
                debit_amount=amount,
                credit_amount=Decimal('0.00')
            )

            # Credit: Accounts Payable (increase liability)
            AdvancedJournalEntryLine.objects.create(
                journal_entry=je,
                line_number=2,
                account=ap_account,
                description=f'Accounts Payable - {ap_entry.vendor_name}',
                debit_amount=Decimal('0.00'),
                credit_amount=amount
            )

            # Post the journal entry
            je.post(user)

            return je
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'    Error creating journal entry: {e}'))
            return None

    def generate_bill_number(self):
        """Generate unique bill number for Accounts Payable"""
        today = timezone.now()
        prefix = f"AP{today.strftime('%Y%m')}"
        
        last_ap = AccountsPayable.objects.filter(
            bill_number__startswith=prefix
        ).order_by('-bill_number').first()
        
        if last_ap:
            try:
                last_num = int(last_ap.bill_number[-6:])
                new_num = last_num + 1
            except:
                new_num = 1
        else:
            new_num = 1
        
        return f"{prefix}{new_num:06d}"

