"""
Professional Consultation Price Import System
Imports consultation prices from Excel file into the database
Handles Cash, Corporate, Insurance, and all insurance company prices
"""
import os
import sys
from decimal import Decimal, InvalidOperation
from django.core.management.base import BaseCommand
from django.utils import timezone
from django.db import transaction
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from hospital.models import ServiceCode
from hospital.models_flexible_pricing import PricingCategory, ServicePrice
from hospital.models_insurance_companies import InsuranceCompany


class Command(BaseCommand):
    help = 'Import consultation prices from Excel file (Professional Grade)'
    
    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default='hms/prices/Consult Price List 2025(1).xlsx',
            help='Path to Excel file'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Run without saving to database (validation only)'
        )
        parser.add_argument(
            '--verbose',
            action='store_true',
            help='Show detailed progress'
        )
    
    def handle(self, *args, **options):
        file_path = options['file']
        dry_run = options['dry_run']
        verbose = options['verbose']
        
        self.stdout.write(self.style.SUCCESS('=' * 80))
        self.stdout.write(self.style.SUCCESS('PROFESSIONAL CONSULTATION PRICE IMPORT SYSTEM'))
        self.stdout.write(self.style.SUCCESS('=' * 80))
        self.stdout.write('')
        
        if dry_run:
            self.stdout.write(self.style.WARNING('DRY RUN MODE - No changes will be saved'))
            self.stdout.write('')
        
        # Validate file exists
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'ERROR: File not found: {file_path}'))
            sys.exit(1)
        
        # Load workbook
        self.stdout.write(f'Loading Excel file: {file_path}')
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            self.stdout.write(self.style.SUCCESS(f'✓ File loaded successfully'))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'ERROR: Could not load Excel file: {e}'))
            sys.exit(1)
        
        # Parse header row
        self.stdout.write('')
        self.stdout.write('Analyzing file structure...')
        headers = self._parse_headers(ws)
        
        if not headers:
            self.stdout.write(self.style.ERROR('ERROR: Could not find header row'))
            sys.exit(1)
        
        self.stdout.write(self.style.SUCCESS(f'✓ Found {len(headers)} columns'))
        
        # Identify key columns
        column_map = self._identify_columns(headers)
        self.stdout.write(f'  Cash column: {column_map["cash"] + 1}')
        self.stdout.write(f'  Corporate column: {column_map["corporate"] + 1}')
        self.stdout.write(f'  Insurance column: {column_map["insurance"] + 1}')
        self.stdout.write(f'  Insurance companies: {len(column_map["insurance_companies"])}')
        
        # Ensure pricing categories exist
        self.stdout.write('')
        self.stdout.write('Setting up pricing categories...')
        categories = self._ensure_categories(column_map, dry_run, verbose)
        
        # Process data rows
        self.stdout.write('')
        self.stdout.write('Processing consultation prices...')
        stats = self._import_prices(ws, headers, column_map, categories, dry_run, verbose)
        
        # Summary
        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS('=' * 80))
        self.stdout.write(self.style.SUCCESS('IMPORT SUMMARY'))
        self.stdout.write(self.style.SUCCESS('=' * 80))
        self.stdout.write(f'  Services processed: {stats["services_processed"]}')
        self.stdout.write(f'  Services created: {stats["services_created"]}')
        self.stdout.write(f'  Services updated: {stats["services_updated"]}')
        self.stdout.write(f'  Prices created: {stats["prices_created"]}')
        self.stdout.write(f'  Prices updated: {stats["prices_updated"]}')
        self.stdout.write(f'  Errors: {stats["errors"]}')
        
        if dry_run:
            self.stdout.write('')
            self.stdout.write(self.style.WARNING('DRY RUN COMPLETE - No data was saved'))
        else:
            self.stdout.write('')
            self.stdout.write(self.style.SUCCESS('✓ Import completed successfully!'))
            self.stdout.write('')
            self.stdout.write('You can now view prices at: /hms/pricing/')
    
    def _parse_headers(self, ws):
        """Parse header row from worksheet"""
        headers = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10), 1):
            row_values = [cell.value for cell in row]
            # Check if this looks like a header row
            if any('cash' in str(val).lower() if val else False for val in row_values):
                headers = row_values
                return headers
        return None
    
    def _identify_columns(self, headers):
        """Identify important columns"""
        column_map = {
            'no': None,
            'specialist_type_id': None,
            'specialist_type_name': None,
            'visit_type': None,
            'cash': None,
            'corporate': None,
            'insurance': None,
            'insurance_companies': {}
        }
        
        for idx, header in enumerate(headers):
            if not header:
                continue
            
            header_str = str(header).lower()
            
            # Basic columns
            if 'no.' in header_str:
                column_map['no'] = idx
            elif 'specialisttypeid' in header_str.replace(' ', ''):
                column_map['specialist_type_id'] = idx
            elif 'specialisttypename' in header_str.replace(' ', ''):
                column_map['specialist_type_name'] = idx
            elif 'visit type' in header_str:
                column_map['visit_type'] = idx
            
            # Price columns
            elif 'cash' in header_str or ('private' in header_str and 'mark-up' in header_str):
                column_map['cash'] = idx
            elif 'company' in header_str or 'corporate' in header_str or 'coperate' in header_str:
                column_map['corporate'] = idx
            elif 'insurance' in header_str and 'other' not in header_str:
                column_map['insurance'] = idx
            elif '[' in header_str and ']' in header_str:
                # Insurance company column (format: [I002] ACE)
                column_map['insurance_companies'][idx] = header
        
        return column_map
    
    def _ensure_categories(self, column_map, dry_run, verbose):
        """Ensure pricing categories exist"""
        categories = {}
        
        # Cash category
        cash_category, created = self._get_or_create_category(
            name='Cash / Private Patients',
            code='CASH',
            category_type='cash',
            color='#10b981',
            dry_run=dry_run
        )
        categories['cash'] = cash_category
        if created and verbose:
            self.stdout.write(f'  ✓ Created category: Cash / Private Patients')
        
        # Corporate category
        corporate_category, created = self._get_or_create_category(
            name='Corporate / Company',
            code='CORPORATE',
            category_type='corporate',
            color='#8b5cf6',
            dry_run=dry_run
        )
        categories['corporate'] = corporate_category
        if created and verbose:
            self.stdout.write(f'  ✓ Created category: Corporate / Company')
        
        # General Insurance category
        insurance_category, created = self._get_or_create_category(
            name='Insurance (General)',
            code='INSURANCE',
            category_type='insurance',
            color='#3b82f6',
            dry_run=dry_run
        )
        categories['insurance'] = insurance_category
        if created and verbose:
            self.stdout.write(f'  ✓ Created category: Insurance (General)')
        
        # Insurance company categories
        for idx, header in column_map['insurance_companies'].items():
            # Extract company name from header (e.g., "[I002] ACE" -> "ACE")
            company_name = str(header).split(']')[-1].strip()
            if not company_name:
                continue
            
            code = f"INS_{company_name.upper().replace(' ', '_')[:20]}"
            category, created = self._get_or_create_category(
                name=f'Insurance - {company_name}',
                code=code,
                category_type='insurance',
                color='#3b82f6',
                dry_run=dry_run
            )
            categories[f'insurance_{idx}'] = category
            if created and verbose:
                self.stdout.write(f'  ✓ Created category: Insurance - {company_name}')
        
        return categories
    
    def _get_or_create_category(self, name, code, category_type, color, dry_run):
        """Get or create pricing category"""
        if dry_run:
            # In dry run, just return a mock object
            class MockCategory:
                def __init__(self, name, code):
                    self.name = name
                    self.code = code
            return MockCategory(name, code), True
        
        category, created = PricingCategory.objects.get_or_create(
            code=code,
            defaults={
                'name': name,
                'category_type': category_type,
                'color_code': color,
                'is_active': True,
                'priority': 100 if category_type == 'cash' else 200
            }
        )
        return category, created
    
    @transaction.atomic
    def _import_prices(self, ws, headers, column_map, categories, dry_run, verbose):
        """Import prices from worksheet"""
        stats = {
            'services_processed': 0,
            'services_created': 0,
            'services_updated': 0,
            'prices_created': 0,
            'prices_updated': 0,
            'errors': []
        }
        
        # Find data start row
        data_start_row = 2  # Assuming header is row 1
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row), data_start_row):
            # Skip empty rows
            if not row or not row[0] or not row[0].value:
                continue
            
            # Skip header-like rows
            if row[0].value and isinstance(row[0].value, str) and row[0].value.startswith('No.'):
                continue
            
            try:
                # Extract row data
                row_data = self._extract_row_data(row, column_map, headers)
                
                if not row_data['specialist_type_name'] or not row_data['visit_type']:
                    continue
                
                stats['services_processed'] += 1
                
                # Create or update service code
                service_code, service_created = self._get_or_create_service(
                    row_data, dry_run
                )
                
                if service_created:
                    stats['services_created'] += 1
                else:
                    stats['services_updated'] += 1
                
                # Import prices
                price_stats = self._import_service_prices(
                    service_code, row_data, column_map, categories, dry_run
                )
                
                stats['prices_created'] += price_stats['created']
                stats['prices_updated'] += price_stats['updated']
                
                if verbose and stats['services_processed'] % 50 == 0:
                    self.stdout.write(f'  Processed {stats["services_processed"]} services...')
            
            except Exception as e:
                error_msg = f'Row {row_idx}: {str(e)}'
                stats['errors'].append(error_msg)
                if verbose:
                    self.stdout.write(self.style.ERROR(f'  {error_msg}'))
                # Log to Django logger for debugging
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error importing row {row_idx}: {str(e)}", exc_info=True)
        
        return stats
    
    def _extract_row_data(self, row, column_map, headers):
        """Extract data from row"""
        def get_value(idx):
            if idx is None or idx >= len(row):
                return None
            val = row[idx].value
            return val if val not in (None, '', '0') else None
        
        def get_float(idx):
            val = get_value(idx)
            if val is None:
                return None
            try:
                return float(val)
            except (ValueError, TypeError):
                return None
        
        return {
            'specialist_type_id': get_value(column_map.get('specialist_type_id')),
            'specialist_type_name': get_value(column_map.get('specialist_type_name')),
            'visit_type': get_value(column_map.get('visit_type')),
            'cash_price': get_float(column_map.get('cash')),
            'corporate_price': get_float(column_map.get('corporate')),
            'insurance_price': get_float(column_map.get('insurance')),
            'insurance_companies': {
                idx: get_float(idx) for idx in column_map['insurance_companies'].keys()
                if get_float(idx) is not None
            }
        }
    
    def _get_or_create_service(self, row_data, dry_run):
        """Get or create service code"""
        specialist_name = str(row_data['specialist_type_name']).strip()
        visit_type = str(row_data['visit_type']).strip()
        
        # Create service code (max 20 chars for ServiceCode.code field)
        service_code_str = row_data.get('specialist_type_id', '')
        if not service_code_str:
            # Generate code from name (truncate to fit 20 char limit)
            # Format: CONS_SPECIALIST_VISIT (abbreviated)
            spec_abbr = ''.join([w[0] for w in specialist_name.split()[:3]])[:8]
            visit_abbr = ''.join([w[0] for w in visit_type.split()[:2]])[:6]
            service_code_str = f"CONS_{spec_abbr}_{visit_abbr}"[:20]
        else:
            # Ensure it fits 20 char limit
            service_code_str = str(service_code_str)[:20]
        
        description = f"{specialist_name} - {visit_type}"
        
        if dry_run:
            class MockService:
                def __init__(self, code, description):
                    self.code = code
                    self.description = description
            return MockService(service_code_str, description), True
        
        service, created = ServiceCode.objects.get_or_create(
            code=service_code_str,
            defaults={
                'description': description,
                'category': 'Consultation',
                'is_active': True
            }
        )
        
        # Update description if service exists
        if not created and service.description != description:
            service.description = description
            service.save()
        
        return service, created
    
    def _import_service_prices(self, service_code, row_data, column_map, categories, dry_run):
        """Import prices for a service"""
        stats = {'created': 0, 'updated': 0}
        
        # Cash price
        if row_data['cash_price']:
            stats.update(self._create_or_update_price(
                service_code, categories['cash'], row_data['cash_price'], dry_run
            ))
        
        # Corporate price
        if row_data['corporate_price']:
            stats.update(self._create_or_update_price(
                service_code, categories['corporate'], row_data['corporate_price'], dry_run
            ))
        
        # General insurance price
        if row_data['insurance_price']:
            stats.update(self._create_or_update_price(
                service_code, categories['insurance'], row_data['insurance_price'], dry_run
            ))
        
        # Insurance company prices
        for idx, price in row_data['insurance_companies'].items():
            category_key = f'insurance_{idx}'
            if category_key in categories:
                stats.update(self._create_or_update_price(
                    service_code, categories[category_key], price, dry_run
                ))
        
        return stats
    
    def _create_or_update_price(self, service_code, category, price, dry_run):
        """Create or update a service price"""
        stats = {'created': 0, 'updated': 0}
        
        if dry_run:
            return {'created': 1, 'updated': 0}
        
        try:
            price_decimal = Decimal(str(price))
            
            # Check for existing price with same effective_from date
            effective_date = timezone.now().date()
            existing_price = ServicePrice.objects.filter(
                service_code=service_code,
                pricing_category=category,
                effective_from=effective_date,
                is_deleted=False
            ).first()
            
            if existing_price:
                # Update existing price
                existing_price.price = price_decimal
                existing_price.is_active = True
                existing_price.notes = f'Imported from Excel - {timezone.now().strftime("%Y-%m-%d")}'
                existing_price.save()
                stats['updated'] = 1
            else:
                # Create new price
                ServicePrice.objects.create(
                    service_code=service_code,
                    pricing_category=category,
                    effective_from=effective_date,
                    price=price_decimal,
                    is_active=True,
                    notes=f'Imported from Excel - {timezone.now().strftime("%Y-%m-%d")}'
                )
                stats['created'] = 1
        
        except (InvalidOperation, ValueError) as e:
            # Log invalid price but continue
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Invalid price value '{price}' for service {service_code.code}: {str(e)}")
        except Exception as e:
            # Log other errors but continue
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error creating/updating price for service {service_code.code}: {str(e)}", exc_info=True)
        
        return stats

