"""
Department-specific dashboard views for Pharmacy, Laboratory, and Imaging
"""
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test, user_passes_test
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.db.models import Q, Count, Sum, Avg, F
from django.db.models.functions import TruncDay
from django.db.utils import OperationalError
from django.core.exceptions import ValidationError
from django.utils import timezone
from datetime import date, timedelta
from decimal import Decimal, ROUND_HALF_UP
from django.urls import reverse
from .models import (
    Order, LabTest, LabResult, Drug, PharmacyStock, Prescription,
    Encounter, Patient, Staff, Department
)
from .models_advanced import ImagingStudy, ImagingImage, ImagingCatalog
from django.http import JsonResponse, Http404, HttpResponse
from django.views.decorators.http import require_http_methods, require_POST
from django.views.decorators.csrf import csrf_exempt
import json
from .services.auto_billing_service import AutoBillingService
from .models_payment_verification import PharmacyDispensing
from .models_accounting import PaymentReceipt
from .views_hod_shift_monitoring import is_hod
from .models_pharmacy_walkin import WalkInPharmacySale
from .models import ServiceCode, Invoice, InvoiceLine, Payer
from .services.pricing_engine_service import pricing_engine
from .utils_billing import get_drug_price_for_prescription, patient_payer_display_labels
from django.db import transaction
import logging

logger = logging.getLogger(__name__)

# Pharmacy “expiring soon” window: dashboard sample list, batch counts, stock list expiring filter
PHARMACY_EXPIRY_FOCUS_DAYS = 70


def _user_can_manage_lab_queue(user):
    if not user.is_authenticated:
        return False
    if getattr(user, 'is_superuser', False) or getattr(user, 'is_staff', False):
        return True
    try:
        from .utils_roles import get_user_role
        return get_user_role(user) in ('lab_technician', 'admin')
    except Exception:
        return False


def _get_patient_payer_info(patient, encounter=None):
    """Thin wrapper around utils_billing.get_patient_payer_info for backward compatibility."""
    from .utils_billing import get_patient_payer_info
    return get_patient_payer_info(patient, encounter)


def _get_encounter_invoice_for_pharmacy(encounter, request=None):
    """Get invoice data for pharmacy verification - encounter's current invoice with pharmacy lines.
    If request is provided and user is pharmacy, detail_url is omitted so UI does not show invoice link."""
    if not encounter:
        return None
    inv = Invoice.objects.filter(
        encounter=encounter, is_deleted=False
    ).select_related('payer').order_by('-created').first()
    if not inv:
        return None
    # Ensure totals are current
    try:
        inv.calculate_totals()
    except Exception:
        pass
    try:
        from .utils_invoice_line import heal_invoice_zero_line_prices
        heal_invoice_zero_line_prices(inv)
    except Exception:
        pass
    # Exclude waived lines
    lines = list(inv.lines.filter(is_deleted=False, waived_at__isnull=True).select_related('service_code', 'prescription').order_by('id'))
    from .utils_roles import is_pharmacy_user
    show_invoice_link = not (request and getattr(request, 'user', None) and is_pharmacy_user(request.user))

    return {
        'id': str(inv.id),
        'invoice_number': inv.invoice_number,
        'status': inv.status,
        'payer_name': getattr(inv.payer, 'name', '') if inv.payer else 'Cash',
        'total_amount': float(inv.total_amount or 0),
        'balance': float(getattr(inv, 'balance', inv.total_amount) or 0),
        'lines': [
            {
                'description': l.description or '',
                'quantity': float(l.quantity or 0),
                'unit_price': float(l.display_unit_price or 0),
                'line_total': float(l.display_line_total or 0),
                'is_pharmacy': l.prescription_id is not None,
                'waived': bool(getattr(l, 'waived_at', None)),
            }
            for l in lines
        ],
        'detail_url': reverse('hospital:invoice_detail', args=[inv.pk]) if show_invoice_link else None,
    }


@login_required
def pharmacy_dashboard(request):
    """World-Class Pharmacy Dashboard - Direct patient service with accounting integration"""
    # Date and name filters from GET params
    filter_name = (request.GET.get('name') or '').strip()
    filter_date_str = (request.GET.get('date') or '').strip()
    filter_date = None
    if filter_date_str:
        try:
            from datetime import datetime
            filter_date = datetime.strptime(filter_date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            filter_date = None
    today = timezone.now().date()
    
    # Get pending medication orders - prioritize by priority level
    # Exclude orders where ALL prescriptions are waived (no drugs to dispense)
    waived_prescription_ids_for_orders = InvoiceLine.objects.filter(
        prescription__isnull=False,
        is_deleted=False,
        waived_at__isnull=False
    ).values_list('prescription_id', flat=True)

    priority_order = {'stat': 0, 'urgent': 1, 'routine': 2}
    pending_orders_qs = Order.objects.filter(
        order_type='medication',
        status='pending',
        is_deleted=False
    ).select_related(
        'encounter__patient',
        'encounter__patient__primary_insurance',
        'requested_by',
    ).defer('encounter__current_activity')

    # Exclude orders where ALL prescriptions are waived or cancelled (nothing left to dispense)
    waived_rx_ids = set(waived_prescription_ids_for_orders)
    order_ids_all_done = []
    for ord in Order.objects.filter(order_type='medication', status='pending', is_deleted=False):
        rxs = list(ord.prescriptions.filter(is_deleted=False).values_list('id', flat=True))
        if not rxs:
            continue
        # Count prescriptions that are "done" (waived invoice line OR cancelled dispensing)
        done_count = 0
        for rx_id in rxs:
            if rx_id in waived_rx_ids:
                done_count += 1
            else:
                try:
                    disp = PharmacyDispensing.objects.get(prescription_id=rx_id)
                    if disp.dispensing_status == 'cancelled' or (disp.quantity_ordered or 0) <= 0:
                        done_count += 1
                except PharmacyDispensing.DoesNotExist:
                    pass  # No dispensing record = still pending
        if done_count >= len(rxs):
            order_ids_all_done.append(ord.id)
    pending_orders_qs = pending_orders_qs.exclude(id__in=order_ids_all_done)
    
    # Apply filters
    if filter_name:
        pending_orders_qs = pending_orders_qs.filter(
            Q(encounter__patient__first_name__icontains=filter_name) |
            Q(encounter__patient__last_name__icontains=filter_name) |
            Q(encounter__patient__mrn__icontains=filter_name)
        )
    if filter_date:
        pending_orders_qs = pending_orders_qs.filter(requested_at__date=filter_date)
    
    pending_medication_orders_total = pending_orders_qs.count()

    # Sort by priority, then by creation time
    pending_orders = sorted(
        pending_orders_qs[:50],
        key=lambda x: (priority_order.get(x.priority, 2), x.requested_at),
        reverse=False
    )[:20]
    
    # Attach payer info so pharmacy knows insurance/corporate vs cash
    for order in pending_orders:
        if order.encounter and order.encounter.patient:
            pat = order.encounter.patient
            enc = order.encounter
            order.payer_info = _get_patient_payer_info(pat, enc)
        else:
            order.payer_info = {'type': 'cash', 'name': 'Cash', 'is_insurance_or_corporate': False}
    
    # Get prescriptions DISPENSED today (not just created) - matches "Dispensed today" label
    target_date = filter_date or today
    today_prescriptions_qs = Prescription.objects.filter(
        is_deleted=False,
        dispensing_record__dispensed_at__date=target_date,
        dispensing_record__dispensing_status__in=('fully_dispensed', 'partially_dispensed'),
        dispensing_record__is_deleted=False
    ).exclude(
        id__in=InvoiceLine.objects.filter(
            prescription__isnull=False,
            is_deleted=False,
            waived_at__isnull=False
        ).values_list('prescription_id', flat=True)
    ).select_related('order__encounter__patient', 'drug', 'prescribed_by', 'dispensing_record').defer('order__encounter__current_activity')
    if filter_name:
        today_prescriptions_qs = today_prescriptions_qs.filter(
            Q(order__encounter__patient__first_name__icontains=filter_name) |
            Q(order__encounter__patient__last_name__icontains=filter_name) |
            Q(order__encounter__patient__mrn__icontains=filter_name)
        )
    today_prescriptions = list(today_prescriptions_qs.order_by('-dispensing_record__dispensed_at')[:20])
    
    # Stock alerts (low stock and expiring soon - exclude already expired)
    expiring_soon = today + timedelta(days=PHARMACY_EXPIRY_FOCUS_DAYS)
    low_stock = list(PharmacyStock.objects.filter(
        quantity_on_hand__lte=F('reorder_level'),
        is_deleted=False
    ).select_related('drug')[:10])
    
    expiring_stock = list(PharmacyStock.objects.filter(
        expiry_date__gte=today,
        expiry_date__lte=expiring_soon,
        quantity_on_hand__gt=0,
        is_deleted=False
    ).select_related('drug').order_by('expiry_date')[:20])
    if not expiring_stock:
        try:
            from .models_procurement import Store
            from .models_inventory_advanced import InventoryBatch
            pharmacy_store = Store.get_pharmacy_store_for_prescriptions()
            if pharmacy_store:
                batches = InventoryBatch.objects.filter(
                    store=pharmacy_store,
                    inventory_item__drug__isnull=False,
                    expiry_date__gte=today,
                    expiry_date__lte=expiring_soon,
                    quantity_remaining__gt=0,
                    is_deleted=False,
                    is_expired=False
                ).select_related('inventory_item__drug').order_by('expiry_date')[:20]
                ExpiringItem = type('ExpiringItem', (), {})
                for b in batches:
                    if b.inventory_item and b.inventory_item.drug:
                        item = ExpiringItem()
                        item.drug = b.inventory_item.drug
                        item.batch_number = b.batch_number or '-'
                        item.expiry_date = b.expiry_date
                        item.quantity_on_hand = b.quantity_remaining
                        expiring_stock.append(item)
        except Exception:
            pass
    
    # Pharmacy statistics
    total_drugs = Drug.objects.filter(is_active=True, is_deleted=False).count()
    total_prescriptions = Prescription.objects.filter(is_deleted=False).count()
    pending_prescriptions = Prescription.objects.filter(
        order__status='pending',
        is_deleted=False
    ).count()
    
    # Total stock value
    stock_value = PharmacyStock.objects.filter(
        is_deleted=False
    ).aggregate(
        total_value=Sum(F('quantity_on_hand') * F('unit_cost'))
    )['total_value'] or Decimal('0')
    
    # Real-time pharmacy revenue + accountability (today = timezone.now().date() at start of view)
    pharmacy_receipt_types = ['pharmacy', 'pharmacy_prescription', 'pharmacy_walkin', 'medication']
    month_start = today.replace(day=1)
    seven_day_window = today - timedelta(days=6)
    recent_window = timezone.now() - timedelta(days=2)
    
    my_receipts_today_qs = PaymentReceipt.objects.filter(
        service_type__in=pharmacy_receipt_types,
        receipt_date__date=today,
        is_deleted=False,
        received_by=request.user
    ).select_related('patient').order_by('-receipt_date')
    
    my_sales_total = my_receipts_today_qs.aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')
    my_sales_count = my_receipts_today_qs.count()
    my_sales_avg_ticket = (my_sales_total / my_sales_count) if my_sales_count else Decimal('0.00')
    
    pharmacy_receipts_today_qs = PaymentReceipt.objects.filter(
        service_type__in=pharmacy_receipt_types,
        receipt_date__date=today,
        is_deleted=False
    )
    pharmacy_revenue_today = pharmacy_receipts_today_qs.aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')
    pharmacy_receipt_count_today = pharmacy_receipts_today_qs.count()
    
    pharmacy_revenue_month = PaymentReceipt.objects.filter(
        service_type__in=pharmacy_receipt_types,
        receipt_date__date__gte=month_start,
        receipt_date__date__lte=today,
        is_deleted=False
    ).aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')
    
    my_sales_share_percentage = 0
    if pharmacy_revenue_today and pharmacy_revenue_today > 0:
        try:
            share = (my_sales_total / pharmacy_revenue_today) * 100
            my_sales_share_percentage = float(round(share, 2))
        except Exception:
            my_sales_share_percentage = 0
    
    pharmacy_daily_trend_qs = PaymentReceipt.objects.filter(
        service_type__in=pharmacy_receipt_types,
        receipt_date__date__gte=seven_day_window,
        receipt_date__date__lte=today,
        is_deleted=False
    ).annotate(day=TruncDay('receipt_date')).values('day').annotate(
        total=Sum('amount_paid'),
        count=Count('id')
    ).order_by('-day')[:7]
    pharmacy_daily_trend = [
        {
            'date': entry['day'].date() if entry['day'] else today,
            'total': entry['total'] or Decimal('0.00'),
            'count': entry['count']
        }
        for entry in reversed(list(pharmacy_daily_trend_qs))
    ]
    
    top_staff_raw = PaymentReceipt.objects.filter(
        service_type__in=pharmacy_receipt_types,
        receipt_date__date=today,
        is_deleted=False,
        received_by__isnull=False
    ).values(
        'received_by__first_name',
        'received_by__last_name',
        'received_by__username'
    ).annotate(
        total=Sum('amount_paid'),
        count=Count('id')
    ).order_by('-total')[:5]
    pharmacy_top_staff = []
    for entry in top_staff_raw:
        first = entry.get('received_by__first_name') or ''
        last = entry.get('received_by__last_name') or ''
        name = f"{first} {last}".strip()
        if not name:
            name = entry.get('received_by__username') or 'Unassigned'
        pharmacy_top_staff.append({
            'name': name,
            'total': entry['total'] or Decimal('0.00'),
            'count': entry['count']
        })
    
    # Live Services Record: show receipts for selected date (or today when no filter)
    _receipts_date = filter_date if filter_date is not None else today
    pharmacy_recent_sales = list(
        PaymentReceipt.objects.filter(
            service_type__in=pharmacy_receipt_types,
            receipt_date__date=_receipts_date,
            is_deleted=False
        ).select_related('patient', 'received_by').order_by('-receipt_date')[:20]
    )
    
    # Walk-in pharmacy sales (OTC) visibility
    try:
        walkin_sales_today_qs = WalkInPharmacySale.objects.filter(
            sale_date__date=today,
            is_deleted=False
        ).select_related('served_by__user').order_by('-sale_date')
        
        my_walkin_sales_qs = walkin_sales_today_qs.filter(
            served_by__user=request.user,
            payment_status='paid'
        )
        
        my_walkin_sales_total = my_walkin_sales_qs.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
        my_walkin_sales_count = my_walkin_sales_qs.count()
        
        walkin_sales_total_today = walkin_sales_today_qs.filter(payment_status='paid').aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
        walkin_sales_count_today = walkin_sales_today_qs.count()
        
        walkin_sales_recent = list(
            walkin_sales_today_qs[:10]
        )
    except OperationalError:
        my_walkin_sales_total = Decimal('0.00')
        my_walkin_sales_count = 0
        walkin_sales_total_today = Decimal('0.00')
        walkin_sales_count_today = 0
        walkin_sales_recent = []
    
    combined_total_revenue = (pharmacy_revenue_today or Decimal('0.00')) + (walkin_sales_total_today or Decimal('0.00'))

    yesterday = today - timedelta(days=1)
    pharmacy_revenue_yesterday = PaymentReceipt.objects.filter(
        service_type__in=pharmacy_receipt_types,
        receipt_date__date=yesterday,
        is_deleted=False,
    ).aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')
    try:
        walkin_yesterday_total = WalkInPharmacySale.objects.filter(
            sale_date__date=yesterday,
            is_deleted=False,
            payment_status='paid',
        ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
    except OperationalError:
        walkin_yesterday_total = Decimal('0.00')
    combined_revenue_yesterday = pharmacy_revenue_yesterday + walkin_yesterday_total
    if combined_revenue_yesterday and combined_revenue_yesterday > 0:
        revenue_dod_change_pct = float(
            ((combined_total_revenue - combined_revenue_yesterday) / combined_revenue_yesterday) * 100
        )
    else:
        revenue_dod_change_pct = None

    my_total_sales = (my_sales_total or Decimal('0.00')) + (my_walkin_sales_total or Decimal('0.00'))
    my_total_transactions = (my_sales_count or 0) + (my_walkin_sales_count or 0)
    my_combined_avg_ticket = (my_total_sales / my_total_transactions) if my_total_transactions else Decimal('0.00')
    my_combined_share = 0
    if combined_total_revenue and combined_total_revenue > 0:
        try:
            my_combined_share = float(round((my_total_sales / combined_total_revenue) * 100, 2))
        except Exception:
            my_combined_share = 0
    
    # Drug Accountability Features
    try:
        from .models_drug_accountability import DrugReturn, DrugAdministrationInventory
        from .models_inventory_advanced import InventoryTransaction
        from .models_procurement import Store, InventoryItem
        
        # Get pharmacy store
        pharmacy_store = Store.objects.filter(store_type='pharmacy', is_deleted=False).first()
        
        # Pending drug returns
        pending_returns = DrugReturn.objects.filter(
            status='pending',
            is_deleted=False
        ).select_related('patient', 'drug', 'requested_by').order_by('-return_date')[:10]
        
        # Recent drug administrations (today)
        recent_administrations = DrugAdministrationInventory.objects.filter(
            administered_at__date=today,
            is_deleted=False
        ).select_related('patient', 'drug', 'administered_by').order_by('-administered_at')[:10]
        
        # Recent inventory transactions (last 7 days)
        recent_transactions = InventoryTransaction.objects.filter(
            store=pharmacy_store,
            transaction_date__gte=seven_day_window,
            is_deleted=False
        ).select_related('inventory_item', 'performed_by').order_by('-transaction_date')[:20] if pharmacy_store else []
        
        # Inventory statistics
        if pharmacy_store:
            inventory_items = InventoryItem.objects.filter(
                store=pharmacy_store,
                is_deleted=False
            )
            total_inventory_items = inventory_items.count()
            low_stock_items = inventory_items.filter(
                quantity_on_hand__lte=F('reorder_level')
            ).count()
            inventory_value = inventory_items.aggregate(
                total=Sum(F('quantity_on_hand') * F('unit_cost'))
            )['total'] or Decimal('0.00')
        else:
            total_inventory_items = 0
            low_stock_items = 0
            inventory_value = Decimal('0.00')
        
        # Today's transactions summary
        today_transactions = InventoryTransaction.objects.filter(
            store=pharmacy_store,
            transaction_date__date=today,
            is_deleted=False
        ) if pharmacy_store else InventoryTransaction.objects.none()
        
        today_receipts = today_transactions.filter(transaction_type='receipt').count()
        today_issues = today_transactions.filter(transaction_type='issue').count()
        today_returns = today_transactions.filter(transaction_type='return_from_dept').count()
        
    except Exception as e:
        # If models don't exist yet, set defaults
        pending_returns = []
        recent_administrations = []
        recent_transactions = []
        total_inventory_items = 0
        low_stock_items = 0
        inventory_value = Decimal('0.00')
        today_receipts = 0
        today_issues = 0
        today_returns = 0
    
    # Get recent medical records for patients with pending orders
    try:
        from .models import MedicalRecord
        patient_ids = [order.encounter.patient.id for order in pending_orders[:10] if order.encounter and order.encounter.patient]
        if patient_ids:
            recent_medical_records = MedicalRecord.objects.filter(
                patient_id__in=patient_ids,
                is_deleted=False
            ).select_related('patient', 'encounter', 'created_by__user').order_by('-created')[:20]
        else:
            recent_medical_records = []
    except Exception:
        recent_medical_records = []
    
    # Check if accountability URLs are available
    accountability_urls_available = False
    try:
        from django.urls import reverse
        reverse('hospital:inventory_accountability_dashboard')
        accountability_urls_available = True
    except Exception:
        accountability_urls_available = False
    
    # Check if user is HOD
    user_is_hod = False
    if request.user.is_authenticated:
        try:
            from .views_hod_scheduling import is_hod
            user_is_hod = is_hod(request.user)
        except:
            pass

    # Dashboard analytics: batch stock + revenue trend bars
    batch_stock_qs = PharmacyStock.objects.filter(is_deleted=False)
    batch_analytics = batch_stock_qs.aggregate(
        batch_lines=Count('id'),
        total_units=Sum('quantity_on_hand'),
    )
    batch_low_stock_count = batch_stock_qs.filter(
        quantity_on_hand__lte=F('reorder_level')
    ).count()
    exp_7 = today + timedelta(days=7)
    batch_expiring_7d_count = batch_stock_qs.filter(
        expiry_date__gte=today,
        expiry_date__lte=exp_7,
        quantity_on_hand__gt=0,
    ).count()
    batch_expiring_focus_count = batch_stock_qs.filter(
        expiry_date__gte=today,
        expiry_date__lte=expiring_soon,
        quantity_on_hand__gt=0,
    ).count()
    distinct_drugs_in_stock = batch_stock_qs.filter(quantity_on_hand__gt=0).values('drug_id').distinct().count()

    if pharmacy_receipt_count_today and pharmacy_receipt_count_today > 0:
        pharmacy_avg_receipt_today = pharmacy_revenue_today / pharmacy_receipt_count_today
    else:
        pharmacy_avg_receipt_today = Decimal('0.00')

    trend_amounts = [d['total'] or Decimal('0') for d in pharmacy_daily_trend]
    pharmacy_trend_max = max(trend_amounts) if trend_amounts else Decimal('0')
    if pharmacy_trend_max <= 0:
        pharmacy_trend_max = Decimal('1')
    pharmacy_daily_trend_bars = []
    for d in pharmacy_daily_trend:
        amt = d['total'] or Decimal('0')
        bar_pct = float((amt / pharmacy_trend_max) * 100) if pharmacy_trend_max else 0.0
        pharmacy_daily_trend_bars.append({
            'date': d['date'],
            'total': amt,
            'count': d['count'],
            'bar_pct': min(100.0, bar_pct),
        })

    stock_alert_total = batch_stock_qs.filter(
        Q(quantity_on_hand__lte=F('reorder_level'))
        | Q(
            expiry_date__gte=today,
            expiry_date__lte=expiring_soon,
            quantity_on_hand__gt=0,
        )
    ).values('id').distinct().count()

    # Cash / corporate / insurance: today's takings + dispensed counts (calendar date)
    _ins_payer_receipt_q = (
        Q(invoice__payer__payer_type='nhis')
        | Q(invoice__payer__payer_type='private')
        | Q(invoice__payer__payer_type='insurance')
    )
    _ins_walkin_q = (
        Q(payer__payer_type='nhis')
        | Q(payer__payer_type='private')
        | Q(payer__payer_type='insurance')
    )
    rev_rx_cash = pharmacy_receipts_today_qs.filter(invoice__payer__payer_type='cash').aggregate(
        t=Sum('amount_paid')
    )['t'] or Decimal('0')
    rev_rx_corp = pharmacy_receipts_today_qs.filter(invoice__payer__payer_type='corporate').aggregate(
        t=Sum('amount_paid')
    )['t'] or Decimal('0')
    rev_rx_ins = pharmacy_receipts_today_qs.filter(_ins_payer_receipt_q).aggregate(t=Sum('amount_paid'))['t'] or Decimal('0')
    rev_rx_tagged = rev_rx_cash + rev_rx_corp + rev_rx_ins
    rev_rx_other = (pharmacy_revenue_today or Decimal('0')) - rev_rx_tagged
    if rev_rx_other < 0:
        rev_rx_other = Decimal('0')

    rev_w_cash = Decimal('0')
    rev_w_corp = Decimal('0')
    rev_w_ins = Decimal('0')
    try:
        wb = WalkInPharmacySale.objects.filter(
            sale_date__date=today, payment_status='paid', is_deleted=False
        )
        rev_w_cash = wb.filter(Q(payer__isnull=True) | Q(payer__payer_type='cash')).aggregate(
            t=Sum('total_amount')
        )['t'] or Decimal('0')
        rev_w_corp = wb.filter(payer__payer_type='corporate').aggregate(t=Sum('total_amount'))['t'] or Decimal('0')
        rev_w_ins = wb.filter(_ins_walkin_q).aggregate(t=Sum('total_amount'))['t'] or Decimal('0')
    except OperationalError:
        pass
    w_tagged = rev_w_cash + rev_w_corp + rev_w_ins
    rev_w_other = (walkin_sales_total_today or Decimal('0')) - w_tagged
    if rev_w_other < 0:
        rev_w_other = Decimal('0')

    waived_rx_disp_ids = InvoiceLine.objects.filter(
        prescription__isnull=False,
        is_deleted=False,
        waived_at__isnull=False,
    ).values_list('prescription_id', flat=True)
    disp_rx_base = PharmacyDispensing.objects.filter(
        dispensed_at__date=today,
        dispensing_status__in=('fully_dispensed', 'partially_dispensed'),
        is_deleted=False,
    ).exclude(prescription_id__in=waived_rx_disp_ids)
    _ins_disp_q = (
        Q(payment_receipt__invoice__payer__payer_type='nhis')
        | Q(payment_receipt__invoice__payer__payer_type='private')
        | Q(payment_receipt__invoice__payer__payer_type='insurance')
    )
    disp_rx_cash = disp_rx_base.filter(payment_receipt__invoice__payer__payer_type='cash').count()
    disp_rx_corp = disp_rx_base.filter(payment_receipt__invoice__payer__payer_type='corporate').count()
    disp_rx_ins = disp_rx_base.filter(_ins_disp_q).count()
    disp_rx_no_receipt = disp_rx_base.filter(payment_receipt__isnull=True).count()

    disp_w_cash = 0
    disp_w_corp = 0
    disp_w_ins = 0
    try:
        wd = WalkInPharmacySale.objects.filter(
            is_dispensed=True,
            dispensed_at__date=today,
            is_deleted=False,
            waived_at__isnull=True,
        )
        disp_w_cash = wd.filter(Q(payer__isnull=True) | Q(payer__payer_type='cash')).count()
        disp_w_corp = wd.filter(payer__payer_type='corporate').count()
        disp_w_ins = wd.filter(_ins_walkin_q).count()
    except OperationalError:
        pass

    pharmacy_payer_channels = {
        'revenue_cash': rev_rx_cash + rev_w_cash,
        'revenue_corporate': rev_rx_corp + rev_w_corp,
        'revenue_insurance': rev_rx_ins + rev_w_ins,
        'revenue_other': rev_rx_other + rev_w_other,
        'dispense_cash': disp_rx_cash + disp_w_cash,
        'dispense_corporate': disp_rx_corp + disp_w_corp,
        'dispense_insurance': disp_rx_ins + disp_w_ins,
        'dispense_rx_no_receipt': disp_rx_no_receipt,
        'as_of_date': today,
    }

    context = {
        'pending_orders': pending_orders,
        'pending_medication_orders_total': pending_medication_orders_total,
        'today_prescriptions': today_prescriptions,
        'filter_date': filter_date_str or '',
        'filter_name': filter_name,
        'low_stock': low_stock,
        'expiring_stock': expiring_stock,
        'total_drugs': total_drugs,
        'total_prescriptions': total_prescriptions,
        'pending_prescriptions': pending_prescriptions,
        'stock_value': stock_value,
        'now': timezone.now(),
        'my_sales_total': my_sales_total,
        'my_sales_count': my_sales_count,
        'my_sales_avg_ticket': my_sales_avg_ticket,
        'my_sales_share_percentage': my_sales_share_percentage,
        'my_recent_pharmacy_receipts': list(my_receipts_today_qs[:8]),
        'my_total_sales': my_total_sales,
        'my_total_transactions': my_total_transactions,
        'my_combined_avg_ticket': my_combined_avg_ticket,
        'my_combined_share_percentage': my_combined_share,
        'pharmacy_revenue_today': pharmacy_revenue_today,
        'pharmacy_receipt_count_today': pharmacy_receipt_count_today,
        'pharmacy_revenue_month': pharmacy_revenue_month,
        'pharmacy_daily_trend': pharmacy_daily_trend,
        'pharmacy_top_staff': pharmacy_top_staff,
        'pharmacy_recent_sales': pharmacy_recent_sales,
        'my_walkin_sales_total': my_walkin_sales_total,
        'my_walkin_sales_count': my_walkin_sales_count,
        'walkin_sales_total_today': walkin_sales_total_today,
        'walkin_sales_count_today': walkin_sales_count_today,
        'walkin_sales_recent': walkin_sales_recent,
        'combined_total_revenue': combined_total_revenue,
        'combined_revenue_yesterday': combined_revenue_yesterday,
        'revenue_dod_change_pct': revenue_dod_change_pct,
        'batch_analytics': batch_analytics,
        'batch_low_stock_count': batch_low_stock_count,
        'batch_expiring_7d_count': batch_expiring_7d_count,
        'batch_expiring_focus_count': batch_expiring_focus_count,
        'pharmacy_expiry_focus_days': PHARMACY_EXPIRY_FOCUS_DAYS,
        'distinct_drugs_in_stock': distinct_drugs_in_stock,
        'pharmacy_avg_receipt_today': pharmacy_avg_receipt_today,
        'pharmacy_daily_trend_bars': pharmacy_daily_trend_bars,
        'stock_alert_total': stock_alert_total,
        # Accountability features
        'pending_returns': pending_returns,
        'recent_administrations': recent_administrations,
        'recent_transactions': recent_transactions,
        'total_inventory_items': total_inventory_items,
        'low_stock_items': low_stock_items,
        'inventory_value': inventory_value,
        'today_receipts': today_receipts,
        'today_issues': today_issues,
        'today_returns': today_returns,
        'pharmacy_store': pharmacy_store,
        'accountability_urls_available': accountability_urls_available,
        'is_hod': user_is_hod,
        'recent_medical_records': recent_medical_records,
        'pharmacy_payer_channels': pharmacy_payer_channels,
    }
    return render(request, 'hospital/pharmacy_dashboard_worldclass.html', context)


@login_required
def pharmacy_my_sales(request):
    """
    Show all sales dispensed/served by the current pharmacist - Cash, Corporate, Insurance.
    Lets pharmacy verify the sales they processed.
    """
    from .models_payment_verification import PharmacyDispensing
    from .models_pharmacy_walkin import WalkInPharmacySale, WalkInPharmacySaleItem
    
    staff = None
    try:
        staff = Staff.objects.get(user=request.user, is_deleted=False)
    except Staff.DoesNotExist:
        pass
    
    filter_date_str = (request.GET.get('date') or '').strip()
    filter_date = None
    if filter_date_str:
        try:
            from datetime import datetime
            filter_date = datetime.strptime(filter_date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            filter_date = None
    today = timezone.now().date()
    date_start = filter_date or today
    date_end = date_start
    
    def payer_type_display(payer):
        if not payer:
            return '—'
        t = getattr(payer, 'payer_type', None) or ''
        return {'cash': 'Cash', 'corporate': 'Corporate', 'nhis': 'NHIS', 'private': 'Insurance'}.get(t, t)
    
    # Prescription dispensings by this pharmacist (dispensed_by)
    prescription_sales = []
    if staff:
        disp_qs = PharmacyDispensing.objects.filter(
            dispensed_by=staff,
            dispensed_at__date__gte=date_start,
            dispensed_at__date__lte=date_end,
            is_deleted=False,
            dispensing_status__in=['partially_dispensed', 'fully_dispensed']
        ).select_related(
            'prescription__drug', 'substitute_drug', 'patient',
            'payment_receipt__invoice__payer'
        ).order_by('-dispensed_at')
        
        for d in disp_qs:
            drug = d.drug_to_dispense
            drug_name = (drug.name if drug else '—') + (f' x{d.quantity_dispensed}' if d.quantity_dispensed else '')
            amount = None
            payer_label = 'Pending'
            if d.payment_receipt:
                amount = d.payment_receipt.amount_paid
                inv = d.payment_receipt.invoice
                if inv and inv.payer:
                    payer_label = payer_type_display(inv.payer)
                else:
                    payer_label = 'Cash'
            else:
                drug = d.drug_to_dispense
                qty = d.quantity_dispensed or d.quantity_ordered or 1
                amount = (drug.unit_price * qty) if drug else Decimal('0.00')
            
            prescription_sales.append({
                'type': 'prescription',
                'date': d.dispensed_at,
                'patient': d.patient.full_name if d.patient else '—',
                'drug': drug_name,
                'amount': amount or Decimal('0.00'),
                'payer': payer_label,
                'receipt': d.payment_receipt.receipt_number if d.payment_receipt else '—',
                'paid': bool(d.payment_receipt),
            })
    
    # Walk-in sales served or dispensed by this pharmacist
    walkin_sales = []
    if staff:
        sale_qs = WalkInPharmacySale.objects.filter(
            Q(served_by=staff) | Q(dispensed_by=staff),
            sale_date__date__gte=date_start,
            sale_date__date__lte=date_end,
            is_deleted=False
        ).select_related('payer', 'patient').prefetch_related('items__drug').order_by('-sale_date')
        
        for s in sale_qs:
            payer_label = payer_type_display(s.payer) if s.payer else 'Cash'
            items_str = ', '.join(
                f"{item.drug.name} x{item.quantity}" for item in s.items.filter(is_deleted=False)[:5]
            )
            if s.items.filter(is_deleted=False).count() > 5:
                items_str += ' …'
            walkin_sales.append({
                'type': 'walkin',
                'date': s.sale_date,
                'patient': s.customer_name or (s.patient.full_name if s.patient else '—'),
                'drug': items_str or '—',
                'amount': s.total_amount or s.amount_paid or Decimal('0.00'),
                'payer': payer_label,
                'receipt': s.sale_number,
                'paid': s.payment_status == 'paid',
                'sale': s,
            })
    
    all_sales = sorted(
        prescription_sales + walkin_sales,
        key=lambda x: x['date'] or timezone.now(),
        reverse=True
    )
    
    def safe_amount(s):
        a = s.get('amount')
        return a if a is not None else Decimal('0.00')
    
    totals = {
        'cash': sum(safe_amount(s) for s in all_sales if s.get('payer') == 'Cash'),
        'corporate': sum(safe_amount(s) for s in all_sales if s.get('payer') == 'Corporate'),
        'insurance': sum(safe_amount(s) for s in all_sales if s.get('payer') in ('NHIS', 'Insurance')),
        'pending': sum(safe_amount(s) for s in all_sales if s.get('payer') == 'Pending' or not s.get('paid', True)),
        'total': sum(safe_amount(s) for s in all_sales),
    }
    
    context = {
        'all_sales': all_sales,
        'totals': totals,
        'filter_date': filter_date_str or today.isoformat(),
        'staff': staff,
    }
    return render(request, 'hospital/pharmacy_my_sales.html', context)


@login_required
def laboratory_dashboard(request):
    """Enhanced Laboratory dashboard with comprehensive analytics and lab-friendly features"""
    from datetime import timedelta as dt_timedelta, datetime as dt_datetime

    today = timezone.now().date()
    week_ago = today - dt_timedelta(days=7)
    month_ago = today - dt_timedelta(days=30)

    # Work list date: default today when opening the page (no ?date= required)
    filter_date = today
    raw_date = (request.GET.get('date') or '').strip()
    if raw_date:
        try:
            filter_date = dt_datetime.strptime(raw_date[:10], '%Y-%m-%d').date()
        except ValueError:
            filter_date = today
    is_today_view = filter_date == today

    priority_order = {'stat': 0, 'urgent': 1, 'routine': 2}

    pending_orders_qs = Order.objects.filter(
        order_type='lab',
        status='pending',
        is_deleted=False,
        created__date=filter_date,
    ).exclude(
        lab_results__status__in=['in_progress', 'completed', 'cancelled']
    ).select_related(
        'encounter__patient',
        'requested_by',
        'requested_by__user'
    ).distinct()

    pending_orders = sorted(
        list(pending_orders_qs[:50]),
        key=lambda x: (priority_order.get(x.priority or 'routine', 2), x.requested_at or x.created),
        reverse=False
    )[:20]

    in_progress_results = list(
        LabResult.objects.filter(
            Q(status='in_progress'),
            Q(created__date=filter_date) | Q(modified__date=filter_date),
            is_deleted=False,
        )
        .select_related('test', 'order', 'order__encounter', 'order__encounter__patient', 'verified_by', 'verified_by__user')
        .select_related('release_record')
        .order_by('-modified')[:50]
    )

    today_results = list(
        LabResult.objects.filter(
            status='completed',
            verified_at__date=filter_date,
            is_deleted=False,
        )
        .select_related('test', 'order', 'order__encounter', 'order__encounter__patient', 'verified_by', 'verified_by__user')
        .select_related('release_record')
        .order_by('-verified_at')[:50]
    )

    pending_results = list(
        LabResult.objects.filter(
            status='pending',
            is_deleted=False,
            created__date=filter_date,
        )
        .select_related('test', 'order', 'order__encounter', 'order__encounter__patient')
        .select_related('release_record')
        .order_by('-created')[:50]
    )

    def _group_results_by_patient_order(results_list):
        groups = {}
        for r in results_list:
            order = r.order
            patient_id = order.encounter.patient_id if order and order.encounter else None
            key = (r.order_id or r.id, patient_id)
            if key not in groups:
                patient = order.encounter.patient if order and order.encounter else None
                groups[key] = {
                    'order': order,
                    'patient': patient,
                    'results': [r],
                    'total_price': float(getattr(r.test, 'price', 0) or 0),
                    'priority': getattr(order, 'priority', 'routine') if order else 'routine',
                    'has_abnormal': getattr(r, 'is_abnormal', False),
                }
            else:
                groups[key]['results'].append(r)
                groups[key]['total_price'] += float(getattr(r.test, 'price', 0) or 0)
                groups[key]['has_abnormal'] = groups[key]['has_abnormal'] or getattr(r, 'is_abnormal', False)
        return sorted(groups.values(), key=lambda g: g['results'][0].created, reverse=True)[:20]

    pending_results_grouped = _group_results_by_patient_order(pending_results)
    in_progress_results_grouped = _group_results_by_patient_order(in_progress_results)
    today_results_grouped = _group_results_by_patient_order(today_results)

    total_tests = LabTest.objects.filter(is_active=True, is_deleted=False).count()
    total_results = LabResult.objects.filter(is_deleted=False).count()

    pending_tests = (
        LabResult.objects.filter(status='pending', is_deleted=False, created__date=filter_date).count()
        + pending_orders_qs.count()
    )

    abnormal_results = LabResult.objects.filter(
        is_abnormal=True,
        is_deleted=False,
        verified_at__gte=week_ago
    ).count()

    completed_results_for_day = LabResult.objects.filter(
        status='completed',
        verified_at__date=filter_date,
        verified_at__isnull=False,
        is_deleted=False,
    ).select_related('order')

    avg_tat_hours = 0
    if completed_results_for_day.exists():
        tat_times = []
        for result in completed_results_for_day[:100]:
            if result.order and result.order.created and result.verified_at:
                tat = (result.verified_at - result.order.created).total_seconds() / 3600
                if tat > 0:
                    tat_times.append(tat)
        if tat_times:
            avg_tat_hours = round(sum(tat_times) / len(tat_times), 1)

    week_completed = LabResult.objects.filter(
        status='completed',
        verified_at__date__gte=week_ago,
        is_deleted=False
    ).count()

    month_completed = LabResult.objects.filter(
        status='completed',
        verified_at__date__gte=month_ago,
        is_deleted=False
    ).count()

    today_revenue = sum([
        float(result.test.price or 0)
        for result in today_results
        if result.test and hasattr(result.test, 'price')
    ])

    top_tests = LabResult.objects.filter(
        is_deleted=False,
        created__gte=week_ago
    ).values('test__name', 'test__code').annotate(
        count=Count('id')
    ).order_by('-count')[:5]

    daily_volume = []
    for i in range(6, -1, -1):
        day = today - dt_timedelta(days=i)
        count = LabResult.objects.filter(
            created__date=day,
            is_deleted=False
        ).count()
        daily_volume.append({
            'date': day.strftime('%b %d'),
            'count': count
        })

    daily_volume_json = json.dumps(daily_volume)

    priority_stats = {
        'stat': Order.objects.filter(
            order_type='lab', status='pending', priority='stat', is_deleted=False, created__date=filter_date
        ).count(),
        'urgent': Order.objects.filter(
            order_type='lab', status='pending', priority='urgent', is_deleted=False, created__date=filter_date
        ).count(),
        'routine': Order.objects.filter(
            order_type='lab', status='pending', priority='routine', is_deleted=False, created__date=filter_date
        ).count(),
    }

    pending_verification = LabResult.objects.filter(
        status='pending',
        is_deleted=False,
        created__date=filter_date,
    ).count()

    overdue_cutoff = timezone.now() - dt_timedelta(hours=24)
    overdue_tests_qs = LabResult.objects.filter(
        status__in=['pending', 'in_progress'],
        created__lt=overdue_cutoff,
        is_deleted=False,
    ).select_related('test', 'order', 'order__encounter', 'order__encounter__patient').select_related(
        'release_record'
    ).order_by('created')
    overdue_tests = overdue_tests_qs.count()
    overdue_lab_results = list(overdue_tests_qs[:100])

    completed_tab_label = 'Completed Today' if is_today_view else f'Completed {filter_date.strftime("%b %d, %Y")}'
    stat_completed_label = 'Completed Today' if is_today_view else f'Completed ({filter_date.strftime("%b %d")})'

    context = {
        'pending_orders': pending_orders,
        'in_progress_results': in_progress_results,
        'today_results': today_results,
        'pending_results': pending_results,
        'pending_results_grouped': pending_results_grouped,
        'in_progress_results_grouped': in_progress_results_grouped,
        'today_results_grouped': today_results_grouped,
        'total_tests': total_tests,
        'total_results': total_results,
        'pending_tests': pending_tests,
        'abnormal_results': abnormal_results,
        'avg_tat_hours': avg_tat_hours,
        'week_completed': week_completed,
        'month_completed': month_completed,
        'today_revenue': today_revenue,
        'top_tests': list(top_tests),
        'daily_volume': daily_volume_json,
        'priority_stats': priority_stats,
        'pending_verification': pending_verification,
        'overdue_tests': overdue_tests,
        'overdue_lab_results': overdue_lab_results,
        'filter_date': filter_date,
        'is_today_view': is_today_view,
        'completed_tab_label': completed_tab_label,
        'stat_completed_label': stat_completed_label,
        'can_remove_overdue': _user_can_manage_lab_queue(request.user),
        'is_lab_hod': is_hod(request.user),
    }
    return render(request, 'hospital/laboratory_dashboard.html', context)


@login_required
@require_POST
def laboratory_remove_overdue_lab(request):
    """Soft-remove a pending/in-progress lab result overdue >24h; void matching invoice lines."""
    if not _user_can_manage_lab_queue(request.user):
        messages.error(request, 'You do not have permission to remove lab queue items.')
        return redirect('hospital:laboratory_dashboard')

    lab_result_id = (request.POST.get('lab_result_id') or '').strip()
    next_url = (request.POST.get('next') or '').strip()
    if next_url and not next_url.startswith('/'):
        next_url = ''

    overdue_cutoff = timezone.now() - timedelta(hours=24)

    try:
        with transaction.atomic():
            lr = LabResult.objects.select_for_update().get(
                pk=lab_result_id,
                is_deleted=False,
            )
            if lr.status not in ('pending', 'in_progress'):
                messages.error(request, 'Only pending or in-progress tests can be removed from this queue.')
                return redirect(next_url or reverse('hospital:laboratory_dashboard'))

            if lr.created >= overdue_cutoff:
                messages.error(request, 'This test is not overdue yet (must be older than 24 hours).')
                return redirect(next_url or reverse('hospital:laboratory_dashboard'))

            try:
                rel = lr.release_record
                if rel and getattr(rel, 'payment_receipt_id', None):
                    messages.error(
                        request,
                        'Cannot remove: a payment receipt is already linked. Use billing/admin if a refund is needed.',
                    )
                    return redirect(next_url or reverse('hospital:laboratory_dashboard'))
            except Exception:
                pass

            lr.status = 'cancelled'
            lr.is_deleted = True
            lr.save(update_fields=['status', 'is_deleted', 'modified'])

            from hospital.views_consultation import _void_lab_invoice_lines_for_lab_result

            _void_lab_invoice_lines_for_lab_result(lr)

            order = lr.order
            if order and order.order_type == 'lab':
                remaining = (
                    order.lab_results.filter(is_deleted=False)
                    .exclude(status='cancelled')
                    .count()
                )
                if remaining == 0:
                    order.status = 'cancelled'
                    order.save(update_fields=['status', 'modified'])

        test_name = getattr(lr.test, 'name', '') or 'Test'
        messages.success(request, f'Removed overdue lab queue item: {test_name}. Billing line voided where applicable.')
    except LabResult.DoesNotExist:
        messages.error(request, 'Lab result not found or already removed.')
    except Exception as exc:
        messages.error(request, f'Could not remove lab item: {exc}')

    return redirect(next_url or reverse('hospital:laboratory_dashboard'))


@login_required
def imaging_dashboard(request):
    """World-Class Imaging/Radiology Dashboard - International Standard PACS-RIS System"""
    today = timezone.now().date()
    now = timezone.now()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)
    
    # Priority ordering
    priority_order = {'stat': 0, 'urgent': 1, 'routine': 2}
    
    # ==================== WORKFLOW QUEUE DATA ====================
    # Ordered (Pending) - DEDUPLICATED
    from django.db import connection
    
    # CRITICAL: Deduplicate orders - keep only most recent per patient per order type per day
    try:
        with connection.cursor() as cursor:
            sql_query = """
                SELECT DISTINCT ON (p.id, o.order_type, DATE(COALESCE(o.requested_at, o.created))) o.id::text
                FROM hospital_order o
                INNER JOIN hospital_encounter e ON e.id = o.encounter_id
                INNER JOIN hospital_patient p ON p.id = e.patient_id
                WHERE o.order_type = 'imaging'
                  AND o.status = 'pending'
                  AND o.is_deleted = false
                  AND e.is_deleted = false
                  AND e.patient_id IS NOT NULL
                ORDER BY p.id, o.order_type, DATE(COALESCE(o.requested_at, o.created)), o.created DESC, o.id DESC
                LIMIT 50
            """
            cursor.execute(sql_query)
            latest_order_ids = [row[0] for row in cursor.fetchall()]
        
        if not latest_order_ids:
            pending_orders = []
        else:
            pending_orders_qs = Order.objects.filter(
                pk__in=latest_order_ids
            ).select_related('encounter__patient', 'requested_by', 'encounter').defer('encounter__current_activity')
            
            pending_orders = sorted(
                list(pending_orders_qs),
                key=lambda x: (priority_order.get(x.priority, 2), x.requested_at or x.created),
                reverse=False
            )[:20]
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"Error deduplicating orders: {str(e)}")
        # Fallback: original query without deduplication
        pending_orders_qs = Order.objects.filter(
            order_type='imaging',
            status='pending',
            is_deleted=False,
            encounter__isnull=False,
            encounter__patient__isnull=False
        ).select_related('encounter__patient', 'requested_by', 'encounter').defer('encounter__current_activity')
        
        pending_orders = sorted(
            list(pending_orders_qs[:50]),
            key=lambda x: (priority_order.get(x.priority, 2), x.requested_at or x.created),
            reverse=False
        )[:20]
    
    # Paid (Orders that are paid but not yet captured)
    # Simplified: Get orders that have imaging studies marked as paid
    paid_orders = Order.objects.filter(
        order_type='imaging',
        status='pending',
        is_deleted=False,
        imaging_studies__is_paid=True,
        imaging_studies__is_deleted=False
    ).select_related('encounter__patient', 'requested_by', 'requested_by__user').prefetch_related('imaging_studies').distinct()[:20]
    
    # Captured (Studies with images but no report)
    captured_studies = ImagingStudy.objects.filter(
        status__in=['completed', 'awaiting_report'],
        is_deleted=False,
        images__isnull=False
    ).exclude(
        report_text__isnull=False
    ).select_related('patient', 'technician__user', 'order', 'order__requested_by', 'order__requested_by__user', 'order__encounter__patient').prefetch_related('images').distinct()[:20]
    
    # Reported (Studies with reports awaiting verification)
    reported_studies = ImagingStudy.objects.filter(
        status='reported',
        report_text__isnull=False,
        report_verified_at__isnull=True,
        is_deleted=False
    ).select_related('patient', 'report_dictated_by__user', 'order', 'order__requested_by', 'order__requested_by__user', 'order__encounter__patient')[:20]
    
    # Released (Verified and released today)
    released_studies = ImagingStudy.objects.filter(
        status='verified',
        report_verified_at__date=today,
        is_deleted=False
    ).select_related('patient', 'report_verified_by__user', 'order', 'order__requested_by', 'order__requested_by__user', 'order__encounter__patient').order_by('-report_verified_at')[:20]
    
    # ==================== STATISTICS ====================
    total_pending = Order.objects.filter(
        order_type='imaging',
        status='pending',
        is_deleted=False
    ).count()
    
    total_in_progress = Order.objects.filter(
        order_type='imaging',
        status='in_progress',
        is_deleted=False
    ).count()
    
    total_completed_today = ImagingStudy.objects.filter(
        status__in=['completed', 'reported', 'verified'],
        performed_at__date=today,
        is_deleted=False
    ).count()
    
    # Revenue today
    today_revenue = sum([
        float(study.paid_amount or 0)
        for study in ImagingStudy.objects.filter(
            paid_at__date=today,
            is_paid=True,
            is_deleted=False
        )
    ])
    
    # Average Turnaround Time (TAT) in hours
    completed_studies = ImagingStudy.objects.filter(
        status='verified',
        performed_at__isnull=False,
        report_verified_at__isnull=False,
        is_deleted=False
    ).exclude(performed_at__isnull=True).exclude(report_verified_at__isnull=True)[:100]
    
    if completed_studies:
        tat_times = []
        for study in completed_studies:
            if study.performed_at and study.report_verified_at:
                delta = study.report_verified_at - study.performed_at
                tat_times.append(delta.total_seconds() / 3600)  # Convert to hours
        avg_tat_hours = sum(tat_times) / len(tat_times) if tat_times else 0
    else:
        avg_tat_hours = 0
    
    # ==================== PRIORITY BREAKDOWN ====================
    stat_count = Order.objects.filter(
        order_type='imaging',
        status='pending',
        priority='stat',
        is_deleted=False
    ).count()
    
    urgent_count = Order.objects.filter(
        order_type='imaging',
        status='pending',
        priority='urgent',
        is_deleted=False
    ).count()
    
    routine_count = Order.objects.filter(
        order_type='imaging',
        status='pending',
        priority='routine',
        is_deleted=False
    ).count()
    
    # ==================== WORKFLOW COUNTS ====================
    ordered_count = total_pending
    paid_count = paid_orders.count()
    captured_count = captured_studies.count()
    reported_count = reported_studies.count()
    released_count = released_studies.count()
    
    # ==================== MODALITY DISTRIBUTION ====================
    modality_counts = {
        'xray': ImagingStudy.objects.filter(modality='xray', created__date=today, is_deleted=False).count(),
        'ct': ImagingStudy.objects.filter(modality='ct', created__date=today, is_deleted=False).count(),
        'mri': ImagingStudy.objects.filter(modality='mri', created__date=today, is_deleted=False).count(),
        'ultrasound': ImagingStudy.objects.filter(modality='ultrasound', created__date=today, is_deleted=False).count(),
        'mammography': ImagingStudy.objects.filter(modality='mammography', created__date=today, is_deleted=False).count(),
    }
    
    # ==================== EQUIPMENT STATUS ====================
    # Placeholder for equipment management - would come from Equipment model if exists
    equipment_list = []
    equipment_operational = 0
    total_equipment = 0
    
    # Try to get equipment data if Equipment model exists
    try:
        from .models_advanced import MedicalEquipment
        equipment_list = list(MedicalEquipment.objects.filter(
            equipment_type__in=['xray', 'ct', 'mri', 'ultrasound'],
            is_deleted=False
        )[:10])
        # MedicalEquipment uses: 'available', 'in_use', 'maintenance', 'out_of_order'
        equipment_operational = MedicalEquipment.objects.filter(
            status__in=['available', 'in_use'],
            is_deleted=False
        ).count()
        total_equipment = MedicalEquipment.objects.filter(is_deleted=False).count()
    except (ImportError, AttributeError):
        # Equipment model not available - use placeholder
        equipment_list = []
        equipment_operational = 0
        total_equipment = 0
    
    # ==================== RECENT REPORTS ====================
    from .models import MedicalRecord
    recent_reports = list(MedicalRecord.objects.filter(
        record_type='imaging',
        is_deleted=False
    ).select_related('patient').order_by('-created')[:10])
    
    context = {
        # Workflow Queue
        'pending_orders': pending_orders,
        'paid_orders': paid_orders,
        'captured_studies': captured_studies,
        'reported_studies': reported_studies,
        'released_studies': released_studies,
        
        # Statistics
        'total_pending': total_pending,
        'total_in_progress': total_in_progress,
        'total_completed_today': total_completed_today,
        'today_revenue': today_revenue,
        'avg_tat_hours': avg_tat_hours,
        
        # Priority Breakdown
        'stat_count': stat_count,
        'urgent_count': urgent_count,
        'routine_count': routine_count,
        
        # Workflow Counts
        'ordered_count': ordered_count,
        'paid_count': paid_count,
        'captured_count': captured_count,
        'reported_count': reported_count,
        'released_count': released_count,
        
        # Modality Distribution
        'modality_counts': modality_counts,
        
        # Equipment
        'equipment_list': equipment_list,
        'equipment_operational': equipment_operational,
        'total_equipment': total_equipment,
        
        # Reports
        'recent_reports': recent_reports,
    }
    
    # Use the comprehensive dashboard template
    return render(request, 'hospital/imaging_dashboard_worldclass.html', context)


@login_required
def imaging_study_detail(request, study_id):
    """View and manage imaging study with images"""
    study = get_object_or_404(
        ImagingStudy.objects.select_related('order', 'order__requested_by', 'order__requested_by__user'),
        pk=study_id,
        is_deleted=False
    )
    
    # Check payment status for cash patients - enforce payment before viewing results
    from hospital.services.auto_billing_service import AutoBillingService
    from hospital.models_payment_verification import ImagingRelease
    from hospital.utils_roles import get_user_role
    
    # Check if user is staff/admin (they can always view)
    user_role = get_user_role(request.user)
    is_staff = (user_role == 'admin' or request.user.is_staff or request.user.is_superuser)
    
    # For cash patients, check payment status
    patient = study.patient
    payer = None
    try:
        payer = patient.primary_insurance
    except:
        pass
    
    is_cash_patient = not payer or (payer and hasattr(payer, 'payer_type') and payer.payer_type == 'cash')
    
    payment_status = None
    can_view_results = True
    payment_required_message = None
    
    if is_cash_patient and not is_staff:
        # Ensure release record exists
        try:
            release_record = study.release_record
        except:
            # Create release record if it doesn't exist
            release_record = ImagingRelease.objects.create(
                imaging_study=study,
                patient=patient,
                release_status='pending_payment'
            )
            # Try to create bill if not exists
            try:
                AutoBillingService.create_imaging_bill(study)
            except:
                pass
        
        # Check payment status
        payment_status = AutoBillingService.check_payment_status('imaging', str(study_id))
        
        # Only allow viewing if payment is verified OR if study is not yet completed/reported
        # (patients can see pending/scheduled studies, but not completed results without payment)
        if study.status in ['completed', 'reported', 'verified'] and not payment_status.get('paid', False):
            can_view_results = False
            payment_required_message = (
                f"❌ PAYMENT REQUIRED! This imaging study requires payment before results can be viewed. "
                f"Please proceed to the cashier to make payment. "
                f"Status: {payment_status.get('message', 'Payment pending')}"
            )
    
    # Get all images for this study
    images = study.images.filter(is_deleted=False).order_by('sequence_number', 'uploaded_at')
    
    # Get current staff
    current_staff = None
    try:
        current_staff = request.user.staff
    except:
        pass
    
    # Handle study assignment
    if request.method == 'POST' and 'assign_to_me' in request.POST:
        try:
            current_staff = request.user.staff
            if current_staff:
                study.assigned_radiologist = current_staff
                study.save(update_fields=['assigned_radiologist'])
                messages.success(request, 'Study assigned to you successfully.')
                return redirect('hospital:imaging_study_detail', study_id=study.id)
        except Exception as e:
            messages.error(request, f'Error assigning study: {str(e)}')
    
    # Handle image upload
    if request.method == 'POST' and 'upload_image' in request.POST:
        try:
            image_file = request.FILES.get('image')
            description = request.POST.get('description', '')
            
            if not image_file:
                messages.error(request, 'Please select an image file to upload.')
            else:
                # Get next sequence number
                last_image = study.images.filter(is_deleted=False).order_by('-sequence_number').first()
                next_sequence = (last_image.sequence_number + 1) if last_image else 1
                
                # Create image
                imaging_image = ImagingImage.objects.create(
                    imaging_study=study,
                    image=image_file,
                    description=description,
                    sequence_number=next_sequence,
                    uploaded_by=current_staff
                )
                
                # Mark study as completed if not already
                if study.status != 'completed':
                    study.status = 'completed'
                    study.performed_at = timezone.now()
                    study.save()
                    
                    # Also mark the order as completed
                    if study.order and study.order.status != 'completed':
                        study.order.status = 'completed'
                        study.order.save()
                
                messages.success(request, f'Image uploaded successfully as image #{next_sequence}. Study marked as complete.')
                return redirect('hospital:imaging_study_detail', study_id=study.id)
        except Exception as e:
            messages.error(request, f'Error uploading image: {str(e)}')
    
    # Handle image deletion
    if request.method == 'POST' and 'delete_image' in request.POST:
        try:
            image_id = request.POST.get('image_id')
            image = get_object_or_404(ImagingImage, pk=image_id, imaging_study=study, is_deleted=False)
            image.is_deleted = True
            image.save()
            messages.success(request, 'Image deleted successfully.')
            return redirect('hospital:imaging_study_detail', study_id=study.id)
        except Exception as e:
            messages.error(request, f'Error deleting image: {str(e)}')
    
    # Get hospital settings for print template
    from .models_settings import HospitalSettings
    settings = HospitalSettings.get_settings()
    
    # Get attached result files for this imaging study
    result_documents = []
    pdf_report_doc = None
    try:
        from .models_medical_records import PatientDocument
        result_documents = list(PatientDocument.objects.filter(
            imaging_study=study,
            document_type='imaging_report',
            is_deleted=False
        ).select_related('uploaded_by__user').order_by('-created'))
        if not result_documents:
            result_documents = list(PatientDocument.objects.filter(
                patient=study.patient,
                encounter=study.encounter,
                document_type='imaging_report',
                is_deleted=False
            ).select_related('uploaded_by__user').order_by('-created')[:5])
        pdf_report_doc = result_documents[0] if result_documents else None
    except:
        pass
    
    context = {
        'study': study,
        'images': images,
        'patient': study.patient,
        'encounter': study.encounter,
        'settings': settings,
        'now': timezone.now(),
        'can_view_results': can_view_results,
        'payment_required_message': payment_required_message,
        'payment_status': payment_status,
        'is_cash_patient': is_cash_patient,
        'pdf_report_doc': pdf_report_doc,
        'result_documents': result_documents,
    }
    
    # If payment required and not staff, show payment required message
    if not can_view_results:
        messages.error(request, payment_required_message)
        # Redirect to cashier payment page if available
        from django.urls import reverse
        try:
            return redirect(reverse('hospital:cashier_process_service_payment', args=['imaging', str(study_id)]))
        except:
            pass
    
    return render(request, 'hospital/imaging_study_detail.html', context)


@login_required
@require_http_methods(["POST"])
def upload_imaging_image(request, study_id):
    """AJAX endpoint for uploading imaging images"""
    study = get_object_or_404(ImagingStudy, pk=study_id, is_deleted=False)
    
    # Get staff profile
    current_staff = None
    try:
        current_staff = Staff.objects.get(user=request.user, is_deleted=False)
    except Staff.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Access denied. Staff profile required.'}, status=403)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Error getting staff profile: {str(e)}')
        return JsonResponse({'success': False, 'error': 'Error accessing staff profile.'}, status=500)
    
    try:
        image_file = request.FILES.get('image')
        description = request.POST.get('description', '').strip()
        
        if not image_file:
            return JsonResponse({'success': False, 'error': 'No image file provided'}, status=400)
        
        # Validate file type - check both content_type and file extension
        allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/dicom', 'application/dicom', 'application/octet-stream', '']
        allowed_extensions = ['.jpg', '.jpeg', '.png', '.dcm', '.dicom']
        
        file_name = image_file.name.lower()
        file_extension = None
        for ext in allowed_extensions:
            if file_name.endswith(ext):
                file_extension = ext
                break
        
        # Check content type or file extension (be lenient with content_type as browsers may not set it correctly)
        content_type_valid = image_file.content_type in allowed_types or not image_file.content_type
        if not content_type_valid and not file_extension:
            return JsonResponse({
                'success': False, 
                'error': f'Invalid file type ({image_file.content_type or "unknown"}). Please upload JPG, PNG, or DICOM images.'
            }, status=400)
        
        # Validate file size (max 50MB for medical images)
        max_size = 50 * 1024 * 1024  # 50MB
        if image_file.size > max_size:
            return JsonResponse({
                'success': False, 
                'error': f'File too large ({image_file.size / 1024 / 1024:.1f}MB). Maximum size is 50MB.'
            }, status=400)
        
        # Get next sequence number
        last_image = study.images.filter(is_deleted=False).order_by('-sequence_number').first()
        next_sequence = (last_image.sequence_number + 1) if last_image else 1
        
        # Create image
        imaging_image = ImagingImage.objects.create(
            imaging_study=study,
            image=image_file,
            description=description,
            sequence_number=next_sequence,
            uploaded_by=current_staff
        )
        
        # Update study status to completed if images are uploaded
        if study.status != 'completed':
            study.status = 'completed'
            if not study.performed_at:
                study.performed_at = timezone.now()
            study.save(update_fields=['status', 'performed_at'])
            
            # Also update the order status
            if study.order and study.order.status != 'completed':
                study.order.status = 'completed'
                study.order.save(update_fields=['status'])
        
        return JsonResponse({
            'success': True,
            'image_id': str(imaging_image.id),
            'image_url': imaging_image.image.url,
            'description': imaging_image.description or '',
            'sequence_number': imaging_image.sequence_number,
            'status': 'completed'
        })
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Error uploading imaging image: {str(e)}', exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def edit_imaging_report(request, study_id):
    """Edit imaging study report"""
    study = get_object_or_404(ImagingStudy, pk=study_id, is_deleted=False)
    
    # Check if user has permission to edit reports (radiologists, doctors, admin)
    current_staff = None
    try:
        current_staff = request.user.staff
    except:
        pass
    
    if not current_staff:
        messages.error(request, 'Access denied. Staff profile required.')
        return redirect('hospital:imaging_study_detail', study_id=study_id)
    
    # Check if user can edit reports
    can_edit = (
        current_staff.profession in ['radiologist', 'doctor', 'admin'] or
        current_staff.user.is_staff
    )
    
    if not can_edit:
        messages.error(request, 'Access denied. Insufficient permissions to edit reports.')
        return redirect('hospital:imaging_study_detail', study_id=study_id)
    
    if request.method == 'POST':
        try:
            # Update report fields
            study.report_text = request.POST.get('report_text', '')
            study.findings = request.POST.get('findings', '')
            study.impression = request.POST.get('impression', '')
            
            # Set dictated by if not already set
            if not study.report_dictated_by:
                study.report_dictated_by = current_staff

            # Save report text first so it's not lost if file upload fails
            study.save()

            # Handle result file upload (PDF, images, Word)
            ALLOWED_IMAGING_FILE_TYPES = {
                'application/pdf',
                'image/jpeg', 'image/jpg', 'image/png', 'image/gif', 'image/webp', 'image/bmp',
                'image/x-png', 'application/msword',
                'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            }
            allowed_extensions = ('.pdf', '.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp', '.doc', '.docx')
            max_file_size = 25 * 1024 * 1024  # 25MB

            if 'result_file' in request.FILES or 'pdf_report' in request.FILES:
                uploaded_file = request.FILES.get('result_file') or request.FILES.get('pdf_report')
                if uploaded_file:
                    fname = getattr(uploaded_file, 'name', '').lower()
                    ext_ok = fname.endswith(allowed_extensions)
                    type_ok = uploaded_file.content_type in ALLOWED_IMAGING_FILE_TYPES or not uploaded_file.content_type
                    if ext_ok and (type_ok or ext_ok):
                        if uploaded_file.size > max_file_size:
                            messages.error(request, f'File too large ({uploaded_file.size // (1024*1024)}MB). Maximum 25MB.')
                        else:
                            try:
                                from .models_medical_records import PatientDocument
                                PatientDocument.objects.create(
                                    patient=study.patient,
                                    encounter=study.encounter,
                                    imaging_study=study,
                                    title=f"Imaging Report - {study.get_modality_display()} - {study.body_part} ({uploaded_file.name[:40]})",
                                    document_type='imaging_report',
                                    uploaded_by=current_staff,
                                    description=f"Report for {study.get_modality_display()} - {study.body_part}",
                                    document_date=timezone.now().date(),
                                    file=uploaded_file,
                                    file_size=uploaded_file.size,
                                    file_type=uploaded_file.content_type or 'application/octet-stream',
                                )
                                messages.success(request, f'Result file "{uploaded_file.name}" saved to patient profile. Doctors can view and download it.')
                            except Exception as doc_error:
                                import logging
                                logger = logging.getLogger(__name__)
                                logger.error(f'Error saving imaging file: {doc_error}', exc_info=True)
                                messages.error(request, f'Could not save file: {str(doc_error)}. Please try again.')
                    else:
                        messages.error(request, 'Invalid file type. Allowed: PDF, JPG, PNG, GIF, WEBP, BMP, DOC, DOCX.')
            
            messages.success(request, 'Report updated successfully.')
            return redirect('hospital:imaging_study_detail', study_id=study_id)
            
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Error updating imaging report: {str(e)}', exc_info=True)
            messages.error(request, f'Error updating report: {str(e)}')
    
    # Get attached result files for this imaging study
    result_documents = []
    pdf_report_doc = None
    try:
        from .models_medical_records import PatientDocument
        result_documents = list(PatientDocument.objects.filter(
            imaging_study=study,
            document_type='imaging_report',
            is_deleted=False
        ).select_related('uploaded_by__user').order_by('-created'))
        if not result_documents:
            result_documents = list(PatientDocument.objects.filter(
                patient=study.patient,
                encounter=study.encounter,
                document_type='imaging_report',
                is_deleted=False
            ).select_related('uploaded_by__user').order_by('-created')[:5])
        pdf_report_doc = result_documents[0] if result_documents else None
    except Exception:
        pass
    
    context = {
        'study': study,
        'patient': study.patient,
        'encounter': study.encounter,
        'result_documents': result_documents,
        'pdf_report_doc': pdf_report_doc,
    }
    return render(request, 'hospital/edit_imaging_report.html', context)


@login_required
def verify_imaging_report(request, study_id):
    """Verify imaging study report"""
    study = get_object_or_404(ImagingStudy, pk=study_id, is_deleted=False)
    
    # Get staff profile
    current_staff = None
    try:
        current_staff = Staff.objects.get(user=request.user, is_deleted=False)
    except Staff.DoesNotExist:
        messages.error(request, 'Access denied. Staff profile required.')
        return redirect('hospital:imaging_study_detail', study_id=study_id)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Error getting staff profile: {str(e)}')
        messages.error(request, 'Error accessing staff profile.')
        return redirect('hospital:imaging_study_detail', study_id=study_id)
    
    # Check if user can verify reports (radiologists, admin, or staff)
    can_verify = (
        current_staff.profession in ['radiologist', 'admin'] or
        current_staff.user.is_staff or
        current_staff.user.is_superuser
    )
    
    if not can_verify:
        messages.error(request, 'Access denied. Insufficient permissions to verify reports.')
        return redirect('hospital:imaging_study_detail', study_id=study_id)
    
    # Check if report exists
    if not study.report_text and not study.findings and not study.impression:
        messages.error(request, 'Cannot verify report. No report content available. Please create a report first.')
        return redirect('hospital:imaging_study_detail', study_id=study_id)
    
    # Check if already verified
    if study.report_verified_at:
        messages.info(request, 'This report has already been verified.')
        return redirect('hospital:imaging_study_detail', study_id=study_id)
    
    # Handle both GET and POST requests (GET for direct link clicks)
    try:
        study.report_verified_by = current_staff
        study.report_verified_at = timezone.now()
        study.status = 'verified'
        study.save(update_fields=['report_verified_by', 'report_verified_at', 'status'])
        
        # Also update order status if it exists
        if study.order and study.order.status != 'completed':
            study.order.status = 'completed'
            study.order.save(update_fields=['status'])
        
        messages.success(request, 'Report verified and released successfully.')
        return redirect('hospital:imaging_study_detail', study_id=study_id)
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Error verifying imaging report: {str(e)}', exc_info=True)
        messages.error(request, f'Error verifying report: {str(e)}')
        return redirect('hospital:imaging_study_detail', study_id=study_id)


def pharmacy_stock_filtered_querysets(request):
    """Same filters as pharmacy stock list: batch stock + pharmacy-category inventory items."""
    from .models_procurement import InventoryCategory, Store, InventoryItem

    query = (request.GET.get('q') or '').strip()
    filter_type = request.GET.get('filter', 'all') or 'all'
    category_filter = (request.GET.get('category') or '').strip()

    pharmacy_category = InventoryCategory.objects.filter(is_for_pharmacy=True, is_active=True).first()

    stock_list = PharmacyStock.objects.filter(is_deleted=False).select_related('drug')
    if category_filter:
        stock_list = stock_list.filter(drug__category=category_filter)

    if pharmacy_category:
        inventory_items = InventoryItem.objects.filter(
            category=pharmacy_category,
            is_deleted=False,
        ).select_related('drug', 'category', 'store').order_by('store__name', 'item_name')
    else:
        inventory_items = InventoryItem.objects.none()

    if query:
        stock_list = stock_list.filter(
            Q(drug__name__icontains=query) |
            Q(drug__generic_name__icontains=query) |
            Q(batch_number__icontains=query)
        )
        if pharmacy_category:
            inventory_items = inventory_items.filter(
                Q(item_name__icontains=query) |
                Q(item_code__icontains=query) |
                Q(description__icontains=query)
            )

    if filter_type == 'low_stock':
        stock_list = stock_list.filter(quantity_on_hand__lte=F('reorder_level'))
        if pharmacy_category:
            inventory_items = inventory_items.filter(quantity_on_hand__lte=F('reorder_level'))
    elif filter_type == 'expiring':
        _t = date.today()
        expiring_soon = _t + timedelta(days=PHARMACY_EXPIRY_FOCUS_DAYS)
        stock_list = stock_list.filter(
            expiry_date__gte=_t,
            expiry_date__lte=expiring_soon,
            quantity_on_hand__gt=0,
        )

    return stock_list, inventory_items, pharmacy_category


@login_required
def pharmacy_stock_list(request):
    """Pharmacy stock/inventory management - shows only pharmacy/pharmaceutical items
    Pharmacy can VIEW; Admin and Procurement/Stores can add and edit."""
    from .models_procurement import Store
    from .views_procurement import can_edit_inventory, can_add_pharmacy_stock, is_pharmacy_staff, is_procurement_staff

    can_edit = can_edit_inventory(request.user) or can_add_pharmacy_stock(request.user)
    can_edit_existing_stock = can_edit_inventory(request.user)
    is_pharmacy = is_pharmacy_staff(request.user)

    query = (request.GET.get('q') or '').strip()
    filter_type = request.GET.get('filter', 'all') or 'all'
    category_filter = (request.GET.get('category') or '').strip()

    stock_list, inventory_items, pharmacy_category = pharmacy_stock_filtered_querysets(request)
    pharmacy_store = Store.objects.filter(store_type='pharmacy').first()

    _td = date.today()
    expiry_threshold = _td + timedelta(days=PHARMACY_EXPIRY_FOCUS_DAYS)
    stock_list_evaluated = list(stock_list.order_by('-created', 'drug__name')[:100])
    inventory_items_evaluated = list(inventory_items)

    from hospital.models import Drug

    bs_all = PharmacyStock.objects.filter(is_deleted=False)
    stock_analytics = {
        'batch_lines': bs_all.count(),
        'total_units': bs_all.aggregate(t=Sum('quantity_on_hand'))['t'] or 0,
        'low_stock_batches': bs_all.filter(quantity_on_hand__lte=F('reorder_level')).count(),
        'expiring_within_window': bs_all.filter(
            expiry_date__gte=_td,
            expiry_date__lte=_td + timedelta(days=PHARMACY_EXPIRY_FOCUS_DAYS),
            quantity_on_hand__gt=0,
        ).count(),
        'stock_value': bs_all.aggregate(v=Sum(F('quantity_on_hand') * F('unit_cost')))['v'] or Decimal('0'),
    }
    filtered_batch_count = stock_list.count()
    inventory_low_batches = 0
    inventory_total_value = Decimal('0')
    if pharmacy_category:
        inventory_low_batches = inventory_items.filter(quantity_on_hand__lte=F('reorder_level')).count()
        inventory_total_value = (
            inventory_items.aggregate(v=Sum(F('quantity_on_hand') * F('unit_cost')))['v'] or Decimal('0')
        )

    context = {
        'stock_list': stock_list_evaluated,
        'inventory_items': inventory_items_evaluated,
        'inventory_count': len(inventory_items_evaluated),
        'stock_analytics': stock_analytics,
        'filtered_batch_count': filtered_batch_count,
        'inventory_low_batches': inventory_low_batches,
        'inventory_total_value': inventory_total_value,
        'query': query,
        'filter_type': filter_type,
        'category_filter': category_filter,
        'drug_categories': Drug.CATEGORIES,
        'expiry_threshold': expiry_threshold.isoformat(),
        'pharmacy_category': pharmacy_category,
        'can_edit': can_edit,
        'can_edit_existing_stock': can_edit_existing_stock,
        'is_pharmacy': is_pharmacy,
        'pharmacy_store': pharmacy_store,
        'pharmacy_expiry_focus_days': PHARMACY_EXPIRY_FOCUS_DAYS,
    }
    return render(request, 'hospital/pharmacy_stock_list.html', context)


@login_required
def pharmacy_stock_print(request):
    """Print / save as PDF: all pharmacy batch stock + inventory rows matching current filters (no 100-row cap)."""
    from .models_settings import HospitalSettings

    stock_list, inventory_items, pharmacy_category = pharmacy_stock_filtered_querysets(request)
    stock_rows = list(stock_list.order_by('-created', 'drug__name'))
    inventory_rows = list(inventory_items)

    settings = HospitalSettings.get_settings()
    category_filter = (request.GET.get('category') or '').strip()
    category_filter_label = (
        dict(Drug.CATEGORIES).get(category_filter, category_filter) if category_filter else ''
    )
    expiry_threshold = date.today() + timedelta(days=PHARMACY_EXPIRY_FOCUS_DAYS)

    context = {
        'stock_list': stock_rows,
        'inventory_items': inventory_rows,
        'query': (request.GET.get('q') or '').strip(),
        'filter_type': request.GET.get('filter', 'all') or 'all',
        'category_filter': category_filter,
        'category_filter_label': category_filter_label,
        'pharmacy_category': pharmacy_category,
        'expiry_threshold': expiry_threshold.isoformat(),
        'settings': settings,
        'generated_at': timezone.now(),
    }
    return render(request, 'hospital/pharmacy_stock_print.html', context)


@login_required
def pharmacy_stock_pdf(request):
    """Download pharmacy inventory as PDF (batch stock + pharmacy inventory items)."""
    try:
        from io import BytesIO
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        return HttpResponse(
            'ReportLab is required for PDF export. Please install it and try again.',
            status=500,
        )

    def _cell(val, max_len=42):
        s = '' if val is None else str(val).replace('\r', ' ').replace('\n', ' ')
        return s[:max_len] if len(s) > max_len else s

    stock_list, inventory_items, _pharmacy_category = pharmacy_stock_filtered_querysets(request)
    stock_rows = list(stock_list.order_by('drug__name', 'expiry_date'))
    inventory_rows = list(inventory_items.order_by('store__name', 'item_name'))

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        topMargin=0.45 * inch,
        bottomMargin=0.45 * inch,
        leftMargin=0.35 * inch,
        rightMargin=0.35 * inch,
    )
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=14,
        textColor=colors.HexColor('#7c3aed'),
        spaceAfter=6,
        alignment=TA_CENTER,
    )
    meta_style = ParagraphStyle('Meta', parent=styles['Normal'], fontSize=8, textColor=colors.grey)

    elements.append(Paragraph('<b>Pharmacy inventory</b>', title_style))
    elements.append(Paragraph(f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M")}', meta_style))
    fq = (request.GET.get('q') or '').strip()
    if fq:
        elements.append(Paragraph(f'Search: {_cell(fq, 90)}', meta_style))
    ft = request.GET.get('filter', 'all') or 'all'
    if ft != 'all':
        elements.append(Paragraph(f'Filter: {ft}', meta_style))
    elements.append(Spacer(1, 0.12 * inch))

    # --- Batch stock (PharmacyStock) ---
    elements.append(Paragraph('<b>Pharmacy stock (by batch)</b>', styles['Heading2']))
    stock_header = [
        'Drug', 'Strength/Form', 'Batch', 'Expiry', 'Qty', 'Reorder', 'Location', 'Unit cost (GHS)',
    ]
    stock_data = [stock_header]
    for s in stock_rows:
        drug = s.drug
        sf = f'{drug.strength} {drug.form}'.strip()
        stock_data.append([
            _cell(drug.name, 36),
            _cell(sf, 22),
            _cell(s.batch_number, 14),
            s.expiry_date.isoformat() if s.expiry_date else '',
            str(s.quantity_on_hand),
            str(s.reorder_level),
            _cell(s.location, 18),
            f'{s.unit_cost:.2f}',
        ])
    if len(stock_data) == 1:
        stock_data.append(['—', '', '', '', '', '', '', 'No rows'])
    st = Table(
        stock_data,
        repeatRows=1,
        colWidths=[1.5 * inch, 1.1 * inch, 0.85 * inch, 0.75 * inch, 0.45 * inch, 0.5 * inch, 1.0 * inch, 0.75 * inch],
    )
    st.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7c3aed')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    elements.append(st)
    elements.append(Spacer(1, 0.2 * inch))

    # --- InventoryItem (pharmacy category) ---
    elements.append(Paragraph('<b>Pharmacy inventory items (procurement)</b>', styles['Heading2']))
    inv_header = [
        'Item', 'Code', 'Store', 'Linked drug', 'Qty', 'UoM', 'Reorder', 'Unit (GHS)', 'Line value (GHS)',
    ]
    inv_data = [inv_header]
    for item in inventory_rows:
        line_val = (item.quantity_on_hand or 0) * (item.unit_cost or Decimal('0'))
        drug_name = item.drug.name if item.drug else '—'
        inv_data.append([
            _cell(item.item_name, 32),
            _cell(item.item_code, 14),
            _cell(item.store.name if item.store else '', 16),
            _cell(drug_name, 28),
            str(item.quantity_on_hand),
            _cell(item.unit_of_measure, 8),
            str(item.reorder_level),
            f'{item.unit_cost:.2f}',
            f'{line_val:.2f}',
        ])
    if len(inv_data) == 1:
        inv_data.append(['—', '', '', '', '', '', '', '', 'No rows'])
    it = Table(
        inv_data,
        repeatRows=1,
        colWidths=[1.35 * inch, 0.8 * inch, 1.0 * inch, 1.2 * inch, 0.4 * inch, 0.45 * inch, 0.45 * inch, 0.6 * inch, 0.75 * inch],
    )
    it.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#5b21b6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    elements.append(it)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    fname = f'pharmacy_inventory_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


@login_required
def pharmacy_stock_excel(request):
    """Polished multi-sheet Excel export: summary + batch stock + procurement pharmacy items."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from .models_settings import HospitalSettings

    HEADER = '5B21B6'
    HEADER_LIGHT = '7C3AED'
    LABEL_BG = 'EDE9FE'
    BORDER_HEX = 'D1D5DB'
    ZEBRA = 'F9FAFB'
    thin = Side(style='thin', color=BORDER_HEX)
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)

    stock_list, inventory_items, pharmacy_category = pharmacy_stock_filtered_querysets(request)
    stock_rows = list(stock_list.order_by('drug__name', 'expiry_date', 'batch_number'))
    inventory_rows = list(inventory_items.order_by('store__name', 'item_name'))

    settings = HospitalSettings.get_settings()
    query = (request.GET.get('q') or '').strip()
    filter_type = request.GET.get('filter', 'all') or 'all'
    category_filter = (request.GET.get('category') or '').strip()
    category_filter_label = (
        dict(Drug.CATEGORIES).get(category_filter, category_filter) if category_filter else 'All categories'
    )
    filter_labels = {
        'all': 'All stock',
        'low_stock': 'Low stock only',
        'expiring': f'Expiring within {PHARMACY_EXPIRY_FOCUS_DAYS} days',
    }
    expiry_cutoff = date.today() + timedelta(days=PHARMACY_EXPIRY_FOCUS_DAYS)
    now = timezone.now()
    gen_str = now.strftime('%Y-%m-%d %H:%M:%S')

    wb = Workbook()

    # ----- Summary -----
    ws_sum = wb.active
    ws_sum.title = 'Summary'
    ws_sum.sheet_properties.tabColor = HEADER_LIGHT
    ws_sum.merge_cells('A1:H1')
    t1 = ws_sum['A1']
    t1.value = 'Pharmacy inventory report'
    t1.font = Font(bold=True, size=22, color=HEADER)
    t1.alignment = Alignment(horizontal='center', vertical='center')
    ws_sum.row_dimensions[1].height = 38
    ws_sum.merge_cells('A2:H2')
    t2 = ws_sum['A2']
    t2.value = settings.hospital_name or 'Hospital'
    t2.font = Font(size=12, color='374151')
    t2.alignment = Alignment(horizontal='center')
    if settings.address:
        ws_sum.merge_cells('A3:H3')
        ws_sum['A3'] = settings.address
        ws_sum['A3'].alignment = Alignment(horizontal='center', wrap_text=True)
        ws_sum['A3'].font = Font(size=10, color='6B7280')
        start_meta = 5
    else:
        start_meta = 4

    meta_rows = [
        ('Generated', gen_str),
        ('Search', query or '—'),
        ('Drug category', category_filter_label),
        ('View filter', filter_labels.get(filter_type, filter_type)),
        ('Batch stock rows', len(stock_rows)),
        ('Inventory item rows', len(inventory_rows)),
    ]
    if pharmacy_category:
        meta_rows.insert(5, ('Procurement pharmacy category', pharmacy_category.name))

    for i, (label, value) in enumerate(meta_rows):
        r = start_meta + i
        c1 = ws_sum.cell(r, 1, label)
        c1.font = Font(bold=True, size=11)
        c1.fill = PatternFill(start_color=LABEL_BG, end_color=LABEL_BG, fill_type='solid')
        c1.border = grid
        c1.alignment = Alignment(vertical='center')
        c2 = ws_sum.cell(r, 2, value)
        c2.border = grid
        c2.alignment = Alignment(vertical='center', wrap_text=True)
        if isinstance(value, str) and len(value) > 60:
            c2.font = Font(size=10)

    ws_sum.column_dimensions['A'].width = 30
    ws_sum.column_dimensions['B'].width = 52
    tip_r = start_meta + len(meta_rows) + 2
    tip = ws_sum.cell(tip_r, 1)
    tip.value = (
        'The Batch stock and Inventory items sheets include filters (Data → Filter), '
        'frozen header rows, and currency formatting (GHS).'
    )
    tip.font = Font(italic=True, size=10, color='6B7280')
    ws_sum.merge_cells(start_row=tip_r, start_column=1, end_row=tip_r, end_column=6)

    def _naive_dt(dt):
        if not dt:
            return None
        if timezone.is_aware(dt):
            return timezone.localtime(dt).replace(tzinfo=None)
        return dt

    def _stock_status(s):
        exp = s.expiry_date
        if exp and exp <= expiry_cutoff and (s.quantity_on_hand or 0) > 0:
            return 'Expiring soon'
        if (s.quantity_on_hand or 0) <= (s.reorder_level or 0):
            return 'Low stock'
        return 'OK'

    # ----- Batch stock sheet -----
    ws_b = wb.create_sheet('Batch stock', 1)
    ws_b.sheet_properties.tabColor = HEADER
    batch_headers = [
        'Drug name',
        'Generic name',
        'Strength',
        'Form',
        'Drug category',
        'Batch number',
        'Expiry date',
        'Qty on hand',
        'Reorder level',
        'Location',
        'Unit cost (GHS)',
        'Extended value (GHS)',
        'Date added',
        'Status',
    ]
    n_batch_cols = len(batch_headers)
    for col, h in enumerate(batch_headers, 1):
        cell = ws_b.cell(1, col, h)
        cell.fill = PatternFill(start_color=HEADER, end_color=HEADER, fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = grid

    for idx, s in enumerate(stock_rows):
        row = 2 + idx
        drug = s.drug
        ext = Decimal(s.quantity_on_hand or 0) * Decimal(s.unit_cost or 0)
        vals = [
            drug.name,
            drug.generic_name or '',
            drug.strength,
            drug.form,
            drug.get_category_display(),
            s.batch_number,
            s.expiry_date,
            int(s.quantity_on_hand or 0),
            int(s.reorder_level or 0),
            s.location or '',
            float(s.unit_cost or 0),
            float(ext),
            _naive_dt(s.created),
            _stock_status(s),
        ]
        zebra = idx % 2 == 1
        for col, val in enumerate(vals, 1):
            c = ws_b.cell(row, col, val)
            c.border = grid
            if zebra:
                c.fill = PatternFill(start_color=ZEBRA, end_color=ZEBRA, fill_type='solid')
            if col == 7 and val:
                c.number_format = 'yyyy-mm-dd'
            if col in (11, 12):
                c.number_format = '#,##0.00'
            if col == 13 and val:
                c.number_format = 'yyyy-mm-dd hh:mm'
            if col in (8, 9):
                c.alignment = Alignment(horizontal='right')

    last_batch = 1 + len(stock_rows)
    if stock_rows:
        tot_r = last_batch + 2
        ws_b.cell(tot_r, 10, 'Total extended value (GHS)').font = Font(bold=True)
        sum_col = get_column_letter(12)
        ws_b.cell(tot_r, 12).value = f'=SUM({sum_col}2:{sum_col}{last_batch})'
        ws_b.cell(tot_r, 12).font = Font(bold=True)
        ws_b.cell(tot_r, 12).number_format = '#,##0.00'
        ws_b.cell(tot_r, 10).border = grid
        ws_b.cell(tot_r, 12).border = grid
    ws_b.auto_filter.ref = f'A1:{get_column_letter(n_batch_cols)}{max(last_batch, 1)}'
    ws_b.freeze_panes = 'A2'
    widths_batch = [28, 22, 12, 12, 36, 14, 12, 10, 10, 18, 14, 16, 18, 14]
    for i, w in enumerate(widths_batch, 1):
        if i <= n_batch_cols:
            ws_b.column_dimensions[get_column_letter(i)].width = w

    # ----- Inventory items sheet -----
    ws_i = wb.create_sheet('Inventory items', 2)
    ws_i.sheet_properties.tabColor = '6D28D9'
    inv_headers = [
        'Item name',
        'Item code',
        'Store',
        'Category',
        'Linked drug',
        'Strength / form',
        'Qty on hand',
        'Unit of measure',
        'Reorder level',
        'Unit cost (GHS)',
        'Line value (GHS)',
        'Active',
        'Status',
    ]
    n_inv_cols = len(inv_headers)
    for col, h in enumerate(inv_headers, 1):
        cell = ws_i.cell(1, col, h)
        cell.fill = PatternFill(start_color=HEADER, end_color=HEADER, fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = grid

    for idx, item in enumerate(inventory_rows):
        row = 2 + idx
        line_val = Decimal(item.quantity_on_hand or 0) * Decimal(item.unit_cost or 0)
        drug = item.drug
        sf = f'{drug.strength} {drug.form}'.strip() if drug else ''
        st = 'Low stock' if (item.quantity_on_hand or 0) <= (item.reorder_level or 0) else (
            'Active' if item.is_active else 'Inactive'
        )
        vals = [
            item.item_name,
            item.item_code or '',
            item.store.name if item.store else '',
            item.category.name if item.category else '',
            drug.name if drug else '',
            sf,
            int(item.quantity_on_hand or 0),
            item.unit_of_measure or '',
            int(item.reorder_level or 0),
            float(item.unit_cost or 0),
            float(line_val),
            'Yes' if item.is_active else 'No',
            st,
        ]
        zebra = idx % 2 == 1
        for col, val in enumerate(vals, 1):
            c = ws_i.cell(row, col, val)
            c.border = grid
            if zebra:
                c.fill = PatternFill(start_color=ZEBRA, end_color=ZEBRA, fill_type='solid')
            if col in (10, 11):
                c.number_format = '#,##0.00'
            if col in (7, 9):
                c.alignment = Alignment(horizontal='right')

    last_inv = 1 + len(inventory_rows)
    if inventory_rows:
        tot_r = last_inv + 2
        ws_i.cell(tot_r, 9, 'Total line value (GHS)').font = Font(bold=True)
        lv_col = get_column_letter(11)
        ws_i.cell(tot_r, 11).value = f'=SUM({lv_col}2:{lv_col}{last_inv})'
        ws_i.cell(tot_r, 11).font = Font(bold=True)
        ws_i.cell(tot_r, 11).number_format = '#,##0.00'
        ws_i.cell(tot_r, 9).border = grid
        ws_i.cell(tot_r, 11).border = grid
    ws_i.auto_filter.ref = f'A1:{get_column_letter(n_inv_cols)}{max(last_inv, 1)}'
    ws_i.freeze_panes = 'A2'
    widths_inv = [30, 16, 20, 18, 26, 18, 12, 12, 12, 14, 16, 10, 14]
    for i, w in enumerate(widths_inv, 1):
        if i <= n_inv_cols:
            ws_i.column_dimensions[get_column_letter(i)].width = w

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    fname = f'pharmacy_inventory_{now.strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


@login_required
def pharmacy_stock_add(request):
    """Add pharmacy stock - HMS frontend form with drug search. Access: Admin, Procurement, Store Managers."""
    from .views_procurement import can_add_pharmacy_stock
    
    if not can_add_pharmacy_stock(request.user):
        messages.error(request, 'You do not have permission to add pharmacy stock.')
        return redirect('hospital:pharmacy_stock_list')
    
    if request.method == 'POST':
        drug_id = request.POST.get('drug_id')
        batch_number = (request.POST.get('batch_number') or '').strip()
        expiry_date_str = request.POST.get('expiry_date')
        location = (request.POST.get('location') or 'Main Pharmacy').strip()
        quantity = request.POST.get('quantity_on_hand', '0')
        reorder_level = request.POST.get('reorder_level', '10')
        unit_cost = request.POST.get('unit_cost', '0')
        total_cost_raw = (request.POST.get('total_cost') or '').strip()
        unit_price = request.POST.get('unit_price', '')
        
        errors = []
        if not drug_id:
            errors.append('Please select a drug from the search.')
        if not expiry_date_str:
            errors.append('Expiry date is required.')
        
        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            try:
                drug = Drug.objects.get(id=drug_id, is_deleted=False, is_active=True)
                expiry_date = date.fromisoformat(expiry_date_str)
                
                if not batch_number:
                    dt = timezone.now()
                    prefix = f"BATCH-{dt.strftime('%Y%m%d')}"
                    same_prefix = PharmacyStock.objects.filter(
                        batch_number__startswith=prefix,
                        is_deleted=False
                    ).count()
                    batch_number = f"{prefix}-{same_prefix + 1:04d}"
                
                quantity_added = int(quantity or 0)
                unit_cost_dec = Decimal(str(unit_cost or 0))
                if total_cost_raw:
                    try:
                        total_dec = Decimal(total_cost_raw)
                        if total_dec > 0 and quantity_added > 0:
                            unit_cost_dec = (total_dec / Decimal(quantity_added)).quantize(
                                Decimal('0.01'), rounding=ROUND_HALF_UP
                            )
                    except (ValueError, TypeError, ArithmeticError):
                        pass
                with transaction.atomic():
                    PharmacyStock.objects.create(
                        drug=drug,
                        batch_number=batch_number,
                        expiry_date=expiry_date,
                        location=location,
                        quantity_on_hand=quantity_added,
                        reorder_level=int(reorder_level or 10),
                        unit_cost=unit_cost_dec,
                        created_by=request.user,
                    )
                    # Update drug selling price if provided (editable on Add Pharmacy Stock)
                    if unit_price not in (None, ''):
                        try:
                            drug.unit_price = Decimal(str(unit_price))
                            drug.save(update_fields=['unit_price'])
                        except (ValueError, TypeError):
                            pass
                added_on = timezone.localtime(timezone.now()).strftime('%d/%m/%Y %I:%M %p')
                messages.success(
                    request,
                    f'Stock added: {drug.name} - Batch {batch_number} | Current stock left: {quantity_added} | Added on: {added_on}'
                )
                return redirect('hospital:pharmacy_stock_list')
            except Drug.DoesNotExist:
                messages.error(request, 'Invalid drug selected.')
            except ValueError as e:
                messages.error(request, str(e))
            except Exception as e:
                messages.error(request, f'Error adding stock: {e}')
    
    context = {'drug_categories': Drug.CATEGORIES}
    return render(request, 'hospital/pharmacy_stock_add.html', context)


@login_required
def pharmacy_stock_edit(request, pk):
    """Edit existing pharmacy stock batch. Admin only."""
    from .views_procurement import can_edit_inventory

    if not can_edit_inventory(request.user):
        messages.error(request, 'You do not have permission to edit pharmacy stock.')
        return redirect('hospital:pharmacy_stock_list')

    stock = get_object_or_404(PharmacyStock, pk=pk, is_deleted=False)

    if request.method == 'POST':
        batch_number = (request.POST.get('batch_number') or '').strip()
        expiry_date_str = request.POST.get('expiry_date')
        location = (request.POST.get('location') or 'Main Pharmacy').strip()
        quantity = request.POST.get('quantity_on_hand', '0')
        reorder_level = request.POST.get('reorder_level', '10')
        unit_cost = request.POST.get('unit_cost', '0')
        total_cost_raw = (request.POST.get('total_cost') or '').strip()
        errors = []
        if not expiry_date_str:
            errors.append('Expiry date is required.')
        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            try:
                qty_int = int(quantity or 0)
                unit_cost_dec = Decimal(str(unit_cost or 0))
                if total_cost_raw:
                    try:
                        total_dec = Decimal(total_cost_raw)
                        if total_dec > 0 and qty_int > 0:
                            unit_cost_dec = (total_dec / Decimal(qty_int)).quantize(
                                Decimal('0.01'), rounding=ROUND_HALF_UP
                            )
                    except (ValueError, TypeError, ArithmeticError):
                        pass
                stock.expiry_date = date.fromisoformat(expiry_date_str)
                stock.location = location
                stock.quantity_on_hand = qty_int
                stock.reorder_level = int(reorder_level or 10)
                stock.unit_cost = unit_cost_dec
                if batch_number:
                    stock.batch_number = batch_number
                stock.save()
                messages.success(request, f'Stock updated: {stock.drug.name} - Batch {stock.batch_number}')
                return redirect('hospital:pharmacy_stock_list')
            except ValueError as e:
                messages.error(request, str(e))
            except Exception as e:
                messages.error(request, f'Error updating stock: {e}')

    context = {
        'stock': stock,
        'drug_categories': Drug.CATEGORIES,
    }
    return render(request, 'hospital/pharmacy_stock_edit.html', context)


@login_required
def lab_results_list(request):
    """List all lab results"""
    from .models_settings import HospitalSettings
    
    status_filter = request.GET.get('status', '')
    query = request.GET.get('q', '')
    print_mode = request.GET.get('print', '')

    results = LabResult.objects.filter(is_deleted=False).select_related(
        'test', 'order__encounter__patient', 'verified_by', 'verified_by__user'
    )

    if status_filter:
        results = results.filter(status=status_filter)

    if query:
        results = results.filter(
            Q(test__name__icontains=query) |
            Q(order__encounter__patient__first_name__icontains=query) |
            Q(order__encounter__patient__last_name__icontains=query) |
            Q(order__encounter__patient__mrn__icontains=query)
        )

    # Get results list (limited to 100)
    results_list = list(results.order_by('-created')[:100])

    # Calculate statistics
    completed_count = sum(1 for r in results_list if r.status == 'completed')
    abnormal_count = sum(1 for r in results_list if r.is_abnormal)

    # If print mode, use print template
    if print_mode:
        settings = HospitalSettings.get_settings()
        context = {
            'results': results_list,
            'status_filter': status_filter,
            'query': query,
            'completed_count': completed_count,
            'abnormal_count': abnormal_count,
            'settings': settings,
            'now': timezone.now(),
        }
        return render(request, 'hospital/lab_results_archive_print.html', context)

    context = {
        'results': results_list,
        'status_filter': status_filter,
        'query': query,
        'completed_count': completed_count,
        'abnormal_count': abnormal_count,
    }
    return render(request, 'hospital/lab_results_list.html', context)


@login_required
def edit_lab_result(request, result_id):
    """Structured edit form for lab results (supports FBC table and qualitative terms)"""
    result = get_object_or_404(LabResult, pk=result_id, is_deleted=False)

    # Use central mapping so the correct template is shown for this test
    from .utils_lab_templates import is_single_value_template

    # All structured templates (FBC, LFT, urine, malaria, etc.) use tabular entry
    if not is_single_value_template(result.test):
        return redirect('hospital:tabular_lab_report', result_id=result_id)

    is_fbc = False  # Single-value only reaches here
    is_single_value = is_single_value_template(result.test)

    if request.method == 'POST':
        # Common fields
        notes = request.POST.get('notes', '').strip()
        qualitative = request.POST.get('qualitative_result', '').strip()
        
        # Handle expected completion datetime
        expected_completion_str = request.POST.get('expected_completion_datetime', '').strip()
        expected_completion = None
        if expected_completion_str:
            try:
                from django.utils.dateparse import parse_datetime
                expected_completion = parse_datetime(expected_completion_str)
                if not expected_completion:
                    # Try parsing as date-time string with timezone
                    from datetime import datetime
                    expected_completion = datetime.strptime(expected_completion_str, '%Y-%m-%dT%H:%M')
                    from django.utils import timezone
                    # Assume local timezone if not provided
                    if timezone.is_naive(expected_completion):
                        expected_completion = timezone.make_aware(expected_completion)
            except (ValueError, TypeError) as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f"Failed to parse expected_completion_datetime: {e}")
        
        # If no explicit datetime set, auto-calculate from tat_minutes for long tests
        if not expected_completion and result.test and result.test.tat_minutes >= 1440:  # >= 1 day
            from django.utils import timezone
            from datetime import timedelta
            start_time = result.created or timezone.now()
            expected_completion = start_time + timedelta(minutes=result.test.tat_minutes)

        # Build details for FBC, single-value, or general components
        details = {}
        if is_fbc:
            # Numeric parameters
            fields = [
                'wbc', 'lymph_count', 'mid_count', 'gran_count',
                'lymph_perc', 'mid_perc', 'gran_perc',
                'plt', 'mpv', 'pdw', 'pct',
                'rbc', 'hgb', 'hct', 'mcv', 'mch', 'mchc', 'rdw_cv', 'rdw_sd',
                'neut_perc', 'mono_perc', 'eos_perc', 'baso_perc'
            ]
            for f in fields:
                val = request.POST.get(f, '').strip()
                if val != '':
                    details[f] = val
        elif is_single_value:
            # Single-value tests (Prolactin, hormones, etc.)
            rv = request.POST.get('result_value', '').strip()
            ru = request.POST.get('result_unit', '').strip()
            if rv:
                details['result_value'] = rv
            if ru:
                details['result_unit'] = ru
        else:
            # Generic component/value pairs (optional future use)
            for i in range(1, 11):
                key = request.POST.get(f'comp_{i}_name', '').strip()
                val = request.POST.get(f'comp_{i}_value', '').strip()
                if key and val:
                    details[key] = val

        # Handle result file upload (PDF, images, Word docs)
        ALLOWED_LAB_FILE_TYPES = {
            'application/pdf',
            'image/jpeg', 'image/png', 'image/gif', 'image/webp', 'image/bmp',
            'application/msword',  # .doc
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document',  # .docx
        }
        if 'result_file' in request.FILES or 'pdf_report' in request.FILES:
            uploaded_file = request.FILES.get('result_file') or request.FILES.get('pdf_report')
            if uploaded_file and (uploaded_file.content_type in ALLOWED_LAB_FILE_TYPES or
                                 getattr(uploaded_file, 'name', '').lower().endswith(('.pdf', '.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp', '.doc', '.docx'))):
                try:
                    from .models_medical_records import PatientDocument
                    current_staff = None
                    try:
                        current_staff = request.user.staff
                    except Exception:
                        pass
                    ext = (uploaded_file.name or '').lower().split('.')[-1] if '.' in (uploaded_file.name or '') else 'pdf'
                    doc = PatientDocument.objects.create(
                        patient=result.order.encounter.patient,
                        encounter=result.order.encounter,
                        lab_result=result,
                        title=f"Lab Report - {result.test.name} ({uploaded_file.name[:50]})",
                        document_type='lab_report',
                        uploaded_by=current_staff,
                        description=f"Lab result file for {result.test.name}",
                        document_date=timezone.now().date(),
                        file=uploaded_file,
                        file_size=uploaded_file.size,
                        file_type=uploaded_file.content_type or f'application/{ext}',
                    )
                    messages.success(request, f'Result file "{uploaded_file.name}" uploaded successfully. Doctors can view and download it.')
                except Exception as doc_error:
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.error(f'Error saving lab result file: {doc_error}', exc_info=True)
                    messages.warning(request, 'Lab result saved, but file upload failed. Please try again.')
            elif uploaded_file:
                messages.error(request, 'Invalid file type. Allowed: PDF, JPG, PNG, GIF, WEBP, BMP, DOC, DOCX.')
        
        # Save onto result
        result.details = details or None
        if is_single_value:
            d = result.details if isinstance(result.details, dict) else {}
            result.value = (d.get('result_value') or '').strip()
            result.units = (d.get('result_unit') or '').strip()
        result.qualitative_result = qualitative
        result.notes = notes
        result.expected_completion_datetime = expected_completion
        # Keep status if already completed else move to in_progress when editing
        if result.status == 'pending':
            result.status = 'in_progress'
        result.save()

        messages.success(request, 'Lab result details saved.')
        return redirect('hospital:laboratory_dashboard')

    # Calculate default expected completion if not set
    default_expected_completion = None
    if not result.expected_completion_datetime and result.test and result.test.tat_minutes:
        from django.utils import timezone
        from datetime import timedelta
        start_time = result.created or timezone.now()
        default_expected_completion = start_time + timedelta(minutes=result.test.tat_minutes)
    
    # Determine if test is long-duration (needs calendar)
    is_long_test = result.test and result.test.tat_minutes and result.test.tat_minutes >= 1440  # >= 1 day
    
    # Get attached result files (PDF, images, etc.) for this lab result
    result_documents = []
    try:
        from .models_medical_records import PatientDocument
        result_documents = list(PatientDocument.objects.filter(
            lab_result=result,
            document_type='lab_report',
            is_deleted=False
        ).select_related('uploaded_by__user').order_by('-created'))
        if not result_documents:
            result_documents = list(PatientDocument.objects.filter(
                patient=result.order.encounter.patient,
                encounter=result.order.encounter,
                document_type='lab_report',
                title__icontains=result.test.name[:30],
                is_deleted=False
            ).select_related('uploaded_by__user').order_by('-created')[:5])
    except Exception:
        pass
    pdf_report_doc = result_documents[0] if result_documents else None
    
    # Pre-fill context
    context = {
        'result': result,
        'is_fbc': is_fbc,
        'is_single_value': is_single_value,
        'details': result.details or {},
        'qualitative_result': result.qualitative_result or '',
        # Standard qualitative options commonly used in labs
        'qualitative_options': ['Seen', 'Not seen', 'Reactive', 'Non-reactive', 'Positive', 'Negative', 'Equivocal', 'Indeterminate'],
        # Expected completion data
        'expected_completion_datetime': result.expected_completion_datetime or default_expected_completion,
        'is_long_test': is_long_test,
        'tat_minutes': result.test.tat_minutes if result.test else None,
        'result_documents': result_documents,
        'pdf_report_doc': pdf_report_doc,
    }
    return render(request, 'hospital/lab_result_edit.html', context)


# ==================== REAL-TIME AJAX ENDPOINTS ====================

@login_required
@csrf_exempt
@require_http_methods(["POST"])
def update_order_status(request, order_id):
    """AJAX endpoint to update order status"""
    order = get_object_or_404(Order, pk=order_id, is_deleted=False)
    action = request.POST.get('action', '')
    current_staff = None
    
    try:
        current_staff = request.user.staff
    except:
        pass
    
    if action == 'start':
        # Mark as in-progress
        order.status = 'in_progress'
        order.save(update_fields=['status', 'modified'])
        return JsonResponse({
            'success': True,
            'message': 'Order marked as in progress',
            'new_status': 'in_progress',
            'status_display': order.get_status_display()
        })
    
    elif action == 'complete':
        # Mark as completed
        order.status = 'completed'
        order.save(update_fields=['status', 'modified'])
        return JsonResponse({
            'success': True,
            'message': 'Order completed',
            'new_status': 'completed',
            'status_display': order.get_status_display()
        })
    
    elif action == 'cancel':
        # Cancel order
        order.status = 'cancelled'
        order.save(update_fields=['status', 'modified'])
        return JsonResponse({
            'success': True,
            'message': 'Order cancelled',
            'new_status': 'cancelled',
            'status_display': order.get_status_display()
        })
    
    else:
        return JsonResponse({
            'success': False,
            'error': 'Invalid action'
        }, status=400)


@login_required
@require_http_methods(["POST"])
def update_lab_result_status(request, result_id):
    """AJAX endpoint to update lab result status"""
    result = get_object_or_404(LabResult, pk=result_id, is_deleted=False)
    action = request.POST.get('action', '')
    value = request.POST.get('value', '')
    notes = request.POST.get('notes', '')
    
    current_staff = None
    try:
        current_staff = request.user.staff
    except:
        pass
    
    order = result.order if hasattr(result, 'order') else None
    
    if action == 'start':
        # Mark lab result as in-progress
        result.status = 'in_progress'
        result.save(update_fields=['status', 'modified'])

        # Move the parent order out of "pending" so it disappears from Pending Lab Orders
        if order and order.status != 'in_progress':
            order.status = 'in_progress'
            order.save(update_fields=['status', 'modified'])

        return JsonResponse({
            'success': True,
            'message': 'Test started',
            'new_status': 'in_progress'
        })
    
    elif action == 'complete':
        # Mark lab result as completed + verified
        result.status = 'completed'
        if value:
            result.value = value
        if notes:
            result.notes = notes
        if current_staff:
            result.verified_by = current_staff
            result.verified_at = timezone.now()
        result.save()

        # Also mark the parent order as completed so it no longer shows with a Start button
        if order and order.status != 'completed':
            order.status = 'completed'
            order.save(update_fields=['status', 'modified'])

        return JsonResponse({
            'success': True,
            'message': 'Test completed',
            'new_status': 'completed'
        })
    
    else:
        return JsonResponse({
            'success': False,
            'error': 'Invalid action'
        }, status=400)


@login_required
def dashboard_stats(request):
    """AJAX endpoint to get real-time dashboard statistics"""
    dashboard_type = request.GET.get('type', 'pharmacy')
    stats = {}
    
    if dashboard_type == 'pharmacy':
        stats = {
            'pending_orders': Order.objects.filter(
                order_type='medication',
                status='pending',
                is_deleted=False
            ).count(),
            'pending_prescriptions': Prescription.objects.filter(
                order__status='pending',
                is_deleted=False
            ).count(),
            'low_stock': PharmacyStock.objects.filter(
                quantity_on_hand__lte=F('reorder_level'),
                is_deleted=False
            ).count(),
        }
    
    elif dashboard_type == 'laboratory':
        stats = {
            'pending_orders': Order.objects.filter(
                order_type='lab',
                status='pending',
                is_deleted=False
            ).count(),
            'in_progress_results': LabResult.objects.filter(
                status='in_progress',
                is_deleted=False
            ).count(),
            'pending_results': LabResult.objects.filter(
                status='pending',
                is_deleted=False
            ).count(),
        }
    
    elif dashboard_type == 'imaging':
        stats = {
            'pending_orders': Order.objects.filter(
                order_type='imaging',
                status='pending',
                is_deleted=False
            ).count(),
            'in_progress_orders': Order.objects.filter(
                order_type='imaging',
                status='in_progress',
                is_deleted=False
            ).count(),
        }
    
    return JsonResponse({
        'success': True,
        'stats': stats,
        'timestamp': timezone.now().isoformat()
    })


@login_required
@require_http_methods(["POST"])
def upload_multiple_imaging_images(request):
    """Upload multiple imaging images via AJAX - World-Class Upload"""
    try:
        order_id = request.POST.get('order_id')
        description = request.POST.get('description', '')
        
        if not order_id:
            return JsonResponse({'success': False, 'error': 'Order ID required'})
        
        order = get_object_or_404(Order, pk=order_id, is_deleted=False)
        
        # Get or create imaging study for this order
        # CRITICAL: Check for existing studies more comprehensively to prevent duplicates
        study = order.imaging_studies.filter(is_deleted=False).first()
        
        if not study:
            # Check for recent duplicate studies for same patient/encounter (within last 10 minutes)
            time_threshold = timezone.now() - timedelta(minutes=10)
            recent_duplicate = ImagingStudy.objects.filter(
                patient=order.encounter.patient,
                encounter=order.encounter,
                is_deleted=False,
                created__gte=time_threshold
            ).exclude(order=order).order_by('-created').first()
            
            if recent_duplicate:
                # Link existing study to this order instead of creating duplicate
                recent_duplicate.order = order
                recent_duplicate.save(update_fields=['order'])
                study = recent_duplicate
            else:
                # Check for duplicate before creating
                existing_study = ImagingStudy.objects.filter(
                    order=order,
                    patient=order.encounter.patient,
                    encounter=order.encounter,
                    is_deleted=False
                ).first()
                
                if not existing_study:
                    # Create new imaging study only if no recent duplicate exists
                    study = ImagingStudy.objects.create(
                        order=order,
                        patient=order.encounter.patient,
                        encounter=order.encounter,
                        modality='xray',  # Default, can be customized
                        body_part='Unknown',  # Can be updated later
                        status='in_progress',
                        scheduled_at=timezone.now(),
                        priority=order.priority
                    )
                else:
                    study = existing_study
        
        # Get current staff
        current_staff = None
        try:
            current_staff = request.user.staff
        except:
            pass
        
        # Upload multiple images
        images = request.FILES.getlist('images')
        if not images:
            return JsonResponse({'success': False, 'error': 'No images uploaded'})
        
        uploaded_count = 0
        last_image = study.images.filter(is_deleted=False).order_by('-sequence_number').first()
        next_sequence = (last_image.sequence_number + 1) if last_image else 1
        
        for image_file in images:
            ImagingImage.objects.create(
                imaging_study=study,
                image=image_file,
                description=description or f'Image {next_sequence}',
                sequence_number=next_sequence,
                uploaded_by=current_staff
            )
            next_sequence += 1
            uploaded_count += 1
        
        # Update order status to completed
        if order.status != 'completed':
            order.status = 'completed'
            order.save()
        
        # Mark imaging study as completed and set performed_at
        if study.status != 'completed':
            study.status = 'completed'
            study.performed_at = timezone.now()
            study.save()
        
        # If study is completed but has no report, set status to awaiting_report
        if study.status == 'completed' and not study.report_text and not study.findings and not study.impression:
            study.status = 'awaiting_report'
            study.save(update_fields=['status'])
        
        # Create medical record for the imaging study
        from .models import MedicalRecord
        record_exists = MedicalRecord.objects.filter(
            patient=study.patient,
            record_type='imaging',
            is_deleted=False,
            content__contains=f'Study ID: {study.id}'
        ).exists()
        
        if not record_exists:
            MedicalRecord.objects.create(
                patient=study.patient,
                encounter=study.encounter,
                record_type='imaging',
                title=f'{study.get_modality_display()} - {study.body_part}',
                content=f'''Imaging Study Completed

Study Type: {study.get_modality_display()}
Body Part: {study.body_part}
Images: {uploaded_count} image(s) uploaded
Study ID: {study.id}
Status: Completed
Performed: {timezone.now().strftime('%Y-%m-%d %H:%M')}

{study.clinical_indication if getattr(study, 'clinical_indication', '') else 'Clinical indication not specified'}

Report pending radiologist review.
''',
                created_by=current_staff
            )
        
        return JsonResponse({
            'success': True,
            'count': uploaded_count,
            'study_id': str(study.id),
            'message': f'Successfully uploaded {uploaded_count} image(s). Study marked as complete.'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def create_imaging_study(request):
    """Create imaging study - handles both GET (form) and POST (create)"""
    from .models_advanced import ImagingStudy
    from datetime import datetime
    
    def parse_datetime(dt_str):
        """Parse datetime string from form input"""
        if not dt_str:
            return timezone.now()
        try:
            # Try ISO format first (datetime-local input)
            if 'T' in dt_str:
                return timezone.make_aware(datetime.strptime(dt_str, '%Y-%m-%dT%H:%M'))
            # Try space format
            return timezone.make_aware(datetime.strptime(dt_str, '%Y-%m-%d %H:%M'))
        except:
            return timezone.now()
    
    # Get pending imaging orders for the form
    pending_orders = Order.objects.filter(
        order_type='imaging',
        status='pending',
        is_deleted=False
    ).select_related('encounter__patient', 'requested_by').order_by('-requested_at')[:50]
    
    if request.method == 'POST':
        # Handle AJAX POST request
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            try:
                order_id = request.POST.get('order_id')
                
                if not order_id:
                    return JsonResponse({'success': False, 'error': 'Order ID required'})
                
                order = get_object_or_404(Order, pk=order_id, is_deleted=False)
                
                # Check if study already exists
                existing_study = order.imaging_studies.filter(is_deleted=False).first()
                if existing_study:
                    return JsonResponse({
                        'success': True,
                        'study_id': str(existing_study.id),
                        'message': 'Study already exists'
                    })
                
                # Get form data
                modality = request.POST.get('modality', 'xray')
                body_part = request.POST.get('body_part', 'Unknown')
                study_type = request.POST.get('study_type', '')
                priority = request.POST.get('priority', order.priority)
                scheduled_at = request.POST.get('scheduled_at')
                
                # Check for duplicate before creating
                existing_study = ImagingStudy.objects.filter(
                    order=order,
                    patient=order.encounter.patient,
                    encounter=order.encounter,
                    modality=modality,
                    body_part=body_part,
                    is_deleted=False
                ).first()
                
                if not existing_study:
                    # Create new imaging study
                    study = ImagingStudy.objects.create(
                        order=order,
                        patient=order.encounter.patient,
                        encounter=order.encounter,
                        modality=modality,
                        body_part=body_part,
                        study_type=study_type,
                        status='scheduled',
                        scheduled_at=parse_datetime(scheduled_at),
                        priority=priority
                    )
                else:
                    study = existing_study
                
                return JsonResponse({
                    'success': True,
                    'study_id': str(study.id),
                    'message': 'Imaging study created successfully',
                    'redirect_url': reverse('hospital:imaging_study_detail', args=[study.id])
                })
                
            except Exception as e:
                return JsonResponse({'success': False, 'error': str(e)})
        
        # Handle regular POST (form submission)
        try:
            order_id = request.POST.get('order_id')
            if not order_id:
                messages.error(request, 'Order ID required')
                return redirect('hospital:imaging_dashboard')
            
            order = get_object_or_404(Order, pk=order_id, is_deleted=False)
            
            # Check if study already exists
            existing_study = order.imaging_studies.filter(is_deleted=False).first()
            if existing_study:
                messages.info(request, 'Study already exists for this order')
                return redirect('hospital:imaging_study_detail', study_id=existing_study.id)
            
            # Get form data
            modality = request.POST.get('modality', 'xray')
            body_part = request.POST.get('body_part', 'Unknown')
            study_type = request.POST.get('study_type', '')
            priority = request.POST.get('priority', order.priority)
            scheduled_at = request.POST.get('scheduled_at')
            
            # Check for duplicate before creating
            existing_study = ImagingStudy.objects.filter(
                order=order,
                patient=order.encounter.patient,
                encounter=order.encounter,
                modality=modality,
                body_part=body_part,
                is_deleted=False
            ).first()
            
            if not existing_study:
                # Create new imaging study
                study = ImagingStudy.objects.create(
                    order=order,
                    patient=order.encounter.patient,
                    encounter=order.encounter,
                    modality=modality,
                    body_part=body_part,
                    study_type=study_type,
                    status='scheduled',
                    scheduled_at=parse_datetime(scheduled_at),
                    priority=priority
                )
            else:
                study = existing_study
                messages.info(request, 'Imaging study already exists for this order.')
                return redirect('hospital:imaging_study_detail', study_id=study.id)
            
            messages.success(request, 'Imaging study created successfully')
            return redirect('hospital:imaging_study_detail', study_id=study.id)
            
        except Exception as e:
            messages.error(request, f'Error creating study: {str(e)}')
            return redirect('hospital:imaging_dashboard')
    
    # GET request - show form
    context = {
        'pending_orders': pending_orders,
        'modality_choices': ImagingStudy.MODALITY_CHOICES,
        'priority_choices': ImagingStudy.PRIORITY_CHOICES,
    }
    
    return render(request, 'hospital/imaging_create_study.html', context)


@login_required
def imaging_add_order(request):
    """Add imaging order from imaging department: select encounter, catalog items (with prices), create order + studies + bills."""
    from .utils_cache import get_cached_imaging_studies

    # Active encounters (last 7 days) for dropdown
    recent_cutoff = timezone.now() - timedelta(days=7)
    encounters = Encounter.objects.filter(
        status='active',
        is_deleted=False,
        started_at__gte=recent_cutoff,
        patient__is_deleted=False,
    ).select_related('patient', 'provider__user').order_by('-started_at')[:80]

    # Imaging catalog with prices
    try:
        catalog_items = get_cached_imaging_studies()
        if hasattr(catalog_items, 'count') and catalog_items.count() == 0:
            catalog_items = ImagingCatalog.objects.filter(
                is_active=True, is_deleted=False
            ).order_by('modality', 'name').only('id', 'name', 'code', 'modality', 'body_part', 'study_type', 'price', 'corporate_price', 'insurance_price')
    except Exception:
        catalog_items = ImagingCatalog.objects.filter(
            is_active=True, is_deleted=False
        ).order_by('modality', 'name').only('id', 'name', 'code', 'modality', 'body_part', 'study_type', 'price', 'corporate_price', 'insurance_price')

    staff = None
    try:
        staff = getattr(request.user, 'staff', None) or Staff.objects.filter(user=request.user, is_deleted=False).first()
    except Exception:
        staff = Staff.objects.filter(user=request.user, is_deleted=False).first()

    if request.method == 'POST':
        if not staff:
            messages.error(request, 'Staff profile required to create imaging orders. Please contact admin.')
            return redirect('hospital:imaging_add_order')
        encounter_id = request.POST.get('encounter_id', '').strip()
        catalog_ids = request.POST.getlist('imaging_catalog_ids')
        priority = request.POST.get('priority', 'routine')
        notes = request.POST.get('notes', '')[:500]

        if not encounter_id or not catalog_ids:
            messages.error(request, 'Please select an encounter and at least one imaging study.')
            return redirect('hospital:imaging_add_order')

        encounter = get_object_or_404(Encounter, pk=encounter_id, is_deleted=False)
        items = ImagingCatalog.objects.filter(
            pk__in=catalog_ids,
            is_active=True,
            is_deleted=False,
        ).order_by('modality', 'name')

        if not items.exists():
            messages.error(request, 'No valid imaging studies selected.')
            return redirect('hospital:imaging_add_order')

        try:
            with transaction.atomic():
                imaging_order = Order.objects.create(
                    encounter=encounter,
                    order_type='imaging',
                    status='pending',
                    priority=priority,
                    notes=notes or None,
                    requested_by=staff,
                    requested_at=timezone.now(),
                )
                created = []
                for cat in items:
                    study = ImagingStudy.objects.create(
                        order=imaging_order,
                        patient=encounter.patient,
                        encounter=encounter,
                        modality=cat.modality,
                        body_part=cat.body_part or '',
                        study_type=cat.code or cat.name,
                        status='scheduled',
                        priority=priority,
                        clinical_indication=notes or '',
                    )
                    created.append(cat.name)
                    try:
                        AutoBillingService.create_imaging_bill(study)
                    except Exception as e:
                        import logging
                        logging.getLogger(__name__).warning('Could not auto-create imaging bill: %s', e)
                names = ', '.join(created)
                messages.success(request, f'Imaging order created: {names}. Patient can proceed to cashier.')
                return redirect('hospital:imaging_dashboard')
        except Exception as e:
            messages.error(request, f'Error creating imaging order: {str(e)}')
            return redirect('hospital:imaging_add_order')

    context = {
        'encounters': encounters,
        'catalog_items': catalog_items,
        'priority_choices': ImagingStudy.PRIORITY_CHOICES,
    }
    return render(request, 'hospital/imaging_add_order.html', context)


@login_required
def imaging_catalog_list(request):
    """List imaging catalog with prices; allow editing prices."""
    from django.core.paginator import Paginator
    from .utils_roles import get_user_role, is_account_or_finance

    # Access control:
    # - Radiology/Scan staff (radiologist role) may rename scan studies
    # - Account/finance and admin may update prices
    try:
        user_role = get_user_role(request.user)
    except Exception:
        user_role = 'staff'

    is_admin = bool(getattr(request.user, 'is_superuser', False)) or user_role == 'admin'
    can_rename = is_admin or user_role == 'radiologist'
    can_edit_prices = is_admin or is_account_or_finance(request.user)
    can_view = can_rename or can_edit_prices

    if not can_view:
        messages.error(request, 'Access denied. Radiology or authorized finance staff only.')
        return redirect('/hms/')

    items = ImagingCatalog.objects.filter(is_deleted=False).order_by('modality', 'name')
    search = request.GET.get('q', '').strip()
    if search:
        items = items.filter(
            Q(name__icontains=search) |
            Q(code__icontains=search) |
            Q(modality__icontains=search) |
            Q(body_part__icontains=search)
        )
    paginator = Paginator(items, 25)
    page = request.GET.get('page', 1)
    items_page = paginator.get_page(page)

    # Handle POST actions: update prices, add item, delete item
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'update_prices':
            if not (can_edit_prices or can_rename):
                messages.error(request, 'Access denied.')
                return redirect('hospital:imaging_catalog_list')
            catalog_id = request.POST.get('catalog_id')
            def _decimal(val, default=None):
                try:
                    s = (val or '').strip()
                    if s == '':
                        return None
                    v = Decimal(s)
                    return v if v >= 0 else (default or Decimal('0.00'))
                except Exception:
                    return default or Decimal('0.00')

            new_name = (request.POST.get('name') or '').strip()
            price_val = _decimal(request.POST.get('price'), Decimal('0.00')) or Decimal('0.00')
            corp_raw = request.POST.get('corporate_price', '').strip()
            ins_raw = request.POST.get('insurance_price', '').strip()
            corp_val = _decimal(corp_raw) if corp_raw != '' else None
            ins_val = _decimal(ins_raw) if ins_raw != '' else None

            if catalog_id:
                obj = ImagingCatalog.objects.filter(pk=catalog_id, is_deleted=False).first()
                if obj:
                    update_fields = []
                    if can_rename and new_name:
                        obj.name = new_name
                        update_fields.append('name')
                    if can_edit_prices:
                        obj.price = price_val
                        obj.corporate_price = corp_val
                        obj.insurance_price = ins_val
                        update_fields.extend(['price', 'corporate_price', 'insurance_price'])
                    if update_fields:
                        obj.save(update_fields=update_fields)
                    try:
                        from django.core.cache import cache
                        cache.delete('hms:active_imaging_studies')
                    except Exception:
                        pass
                    messages.success(request, f'Updated {obj.name}.')
            return redirect('hospital:imaging_catalog_list')
        
        elif action == 'add_imaging_item':
            if not is_admin:
                messages.error(request, 'Access denied. Only admin can add imaging items.')
                return redirect('hospital:imaging_catalog_list')
            # Add new imaging catalog item
            try:
                code = request.POST.get('code', '').strip()
                name = request.POST.get('name', '').strip()
                modality = request.POST.get('modality', 'xray')
                body_part = request.POST.get('body_part', '').strip()
                price = Decimal(request.POST.get('price', '0') or '0')
                corporate_price = request.POST.get('corporate_price', '').strip()
                insurance_price = request.POST.get('insurance_price', '').strip()
                
                if not code or not name:
                    messages.error(request, 'Code and name are required.')
                else:
                    # Check if code already exists
                    if ImagingCatalog.objects.filter(code=code, is_deleted=False).exists():
                        messages.error(request, f'Imaging item with code "{code}" already exists.')
                    else:
                        corp_val = Decimal(corporate_price) if corporate_price else None
                        ins_val = Decimal(insurance_price) if insurance_price else None
                        
                        imaging_item = ImagingCatalog.objects.create(
                            code=code,
                            name=name,
                            modality=modality,
                            body_part=body_part,
                            price=price,
                            corporate_price=corp_val,
                            insurance_price=ins_val,
                            is_active=True
                        )
                        try:
                            from django.core.cache import cache
                            cache.delete('hms:active_imaging_studies')
                        except Exception:
                            pass
                        messages.success(request, f'Imaging item "{imaging_item.name}" added successfully.')
            except Exception as e:
                messages.error(request, f'Error adding imaging item: {str(e)}')
            return redirect('hospital:imaging_catalog_list')
        
        elif action == 'delete_imaging_item':
            if not is_admin:
                messages.error(request, 'Access denied. Only admin can delete imaging items.')
                return redirect('hospital:imaging_catalog_list')
            # Delete imaging catalog item (soft delete)
            try:
                catalog_id = request.POST.get('catalog_id')
                if catalog_id:
                    obj = ImagingCatalog.objects.filter(pk=catalog_id, is_deleted=False).first()
                    if obj:
                        obj.is_deleted = True
                        obj.save()
                        try:
                            from django.core.cache import cache
                            cache.delete('hms:active_imaging_studies')
                        except Exception:
                            pass
                        messages.success(request, f'Imaging item "{obj.name}" deleted successfully.')
                    else:
                        messages.error(request, 'Imaging item not found.')
            except Exception as e:
                messages.error(request, f'Error deleting imaging item: {str(e)}')
            return redirect('hospital:imaging_catalog_list')

    context = {
        'items_page': items_page,
        'search_query': search,
        'can_rename': can_rename,
        'can_edit_prices': can_edit_prices,
        'is_admin': is_admin,
    }
    return render(request, 'hospital/imaging_catalog_list.html', context)


@login_required
def get_pharmacy_order_prescriptions(request, order_id):
    """API endpoint to get prescriptions for a medication order.
    Returns all prescriptions so pharmacy can always see and edit drugs/quantities before sending to company.
    """
    try:
        order = get_object_or_404(Order, pk=order_id, order_type='medication', is_deleted=False)
        
        # Prescriptions that already have an active (non-waived) bill - for UI badge only
        rx_ids_with_active_bill = set(InvoiceLine.objects.filter(
            prescription__order=order,
            is_deleted=False,
            waived_at__isnull=True
        ).values_list('prescription_id', flat=True))
        
        # Return ALL order prescriptions so data-prescription-ids includes every id;
        # then when pharmacy removes one and sends, we receive quantity 0 for that id.
        prescriptions = order.prescriptions.filter(is_deleted=False).select_related('drug')
        
        from .models_payment_verification import (
            PharmacyDispensing,
            PharmacyDispenseHistory,
            PharmacyStockDeductionLog,
        )
        dispensings_by_rx = {
            str(d.prescription_id): d
            for d in PharmacyDispensing.objects.filter(
                prescription__in=prescriptions,
                is_deleted=False
            ).select_related('substitute_drug')
        }
        
        prescriptions_data = []
        patient_for_rx = order.encounter.patient
        payer_rx = getattr(patient_for_rx, 'primary_insurance', None)
        for rx in prescriptions:
            # Use substituted drug from existing dispensing record if any (so re-open shows current bill)
            disp = dispensings_by_rx.get(str(rx.id))
            drug_to_show = (disp.substitute_drug if disp and getattr(disp, 'substitute_drug_id', None) else None) or rx.drug
            try:
                qty_to_show = int(disp.quantity_ordered) if disp and getattr(disp, 'quantity_ordered', None) is not None else int(rx.quantity or 0)
            except (TypeError, ValueError):
                qty_to_show = int(rx.quantity or 0)
            
            drug_price = get_drug_price_for_prescription(drug_to_show, payer=payer_rx)
            
            # Check stock availability for the drug being shown
            stock_available = PharmacyStock.objects.filter(
                drug=drug_to_show,
                is_deleted=False,
                quantity_on_hand__gt=0
            ).aggregate(total=Sum('quantity_on_hand'))['total'] or 0
            
            prescriptions_data.append({
                'id': str(rx.id),
                'drug_id': str(drug_to_show.id),
                'original_drug_id': str(rx.drug.id),
                'drug_name': drug_to_show.name,
                'drug_strength': drug_to_show.strength or '',
                'drug_form': drug_to_show.form or '',
                'dose': rx.dose,
                'frequency': rx.frequency,
                'duration': rx.duration,
                'quantity': qty_to_show,
                'drug_price': float(drug_price),
                'total_price': float(drug_price * qty_to_show),
                'instructions': rx.instructions,
                'stock_available': int(stock_available),
                'has_active_bill': rx.id in rx_ids_with_active_bill,
                'prescribed_on': timezone.localtime(rx.created).strftime('%Y-%m-%d %H:%M') if getattr(rx, 'created', None) else '',
            })
        
        patient = order.encounter.patient
        
        # Payer info for insurance/corporate vs cash workflow (comprehensive check)
        payer_info = _get_patient_payer_info(patient, order.encounter)
        payer_display_labels = patient_payer_display_labels(patient, order.encounter)
        payer = payer_info.get('payer')
        payer_type = payer_info.get('type', 'cash')
        is_insurance_or_corporate = payer_info.get('is_insurance_or_corporate', False)
        payer_name = payer_info.get('name', 'Cash')
        if payer_display_labels and not is_insurance_or_corporate:
            is_insurance_or_corporate = True
        
        # Check insurance exclusion for each drug (insurance/corporate only)
        for rx_data in prescriptions_data:
            rx_data['is_insurance_excluded'] = False
            if is_insurance_or_corporate and payer:
                try:
                    from .models import Drug
                    from .services.insurance_exclusion_service import InsuranceExclusionService
                    drug = Drug.objects.filter(pk=rx_data['drug_id']).first()
                    if drug:
                        result = InsuranceExclusionService(
                            patient=patient, payer=payer, drug=drug
                        ).evaluate()
                        rx_data['is_insurance_excluded'] = result.requires_patient_pay or result.should_block
                except Exception:
                    pass
        
        # Invoice for pharmacy verification (combine view)
        invoice_data = _get_encounter_invoice_for_pharmacy(order.encounter, request=request)

        # When insurance/corporate: bill already sent if all dispensing records are ready_to_dispense (not pending payment)
        sent_to_insurance = False
        if is_insurance_or_corporate and prescriptions_data:
            from .models_payment_verification import PharmacyDispensing
            rx_ids = [p['id'] for p in prescriptions_data]
            pending = PharmacyDispensing.objects.filter(
                prescription_id__in=rx_ids,
                is_deleted=False,
                dispensing_status='pending_payment',
            ).exists()
            sent_to_insurance = not pending

        all_order_rx_ids = [str(rx.id) for rx in order.prescriptions.filter(is_deleted=False)]
        return JsonResponse({
            'success': True,
            'prescriptions': prescriptions_data,
            'all_prescription_ids': all_order_rx_ids,
            'sent_to_insurance': sent_to_insurance,
            'patient': {
                'id': str(patient.id),
                'full_name': patient.full_name,
                'mrn': patient.mrn,
                'age': patient.age,
                'phone_number': patient.phone_number or '',
                'gender': patient.get_gender_display(),
            },
            'payer': {
                'type': payer_type,
                'name': payer_name,
                'is_insurance_or_corporate': is_insurance_or_corporate,
                'corporate_badge_name': payer_info.get('corporate_badge_name') or '',
                'display_labels': payer_display_labels,
            },
            'invoice': invoice_data,
            'order': {
                'id': str(order.id),
                'priority': order.priority,
                'requested_by': order.requested_by.user.get_full_name() if order.requested_by else '',
                'requested_at': order.requested_at.strftime('%Y-%m-%d %H:%M'),
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def check_pharmacy_order_payment_status(request, order_id):
    """Check if payment has been made for pharmacy order (or bill sent to insurance)"""
    try:
        order = get_object_or_404(Order, pk=order_id, order_type='medication', is_deleted=False)
        
        # Get all prescriptions for this order
        prescriptions = order.prescriptions.filter(is_deleted=False)
        
        if not prescriptions.exists():
            return JsonResponse({'success': False, 'error': 'No prescriptions found'})
        
        patient = order.encounter.patient if order.encounter else None
        payment_verified = False
        
        # Insurance/Corporate: Verified when pharmacy dispensing records exist (bill sent to insurer)
        payer_info = _get_patient_payer_info(patient, order.encounter) if patient else {}
        if payer_info.get('is_insurance_or_corporate'):
            from .models_payment_verification import PharmacyDispensing
            if PharmacyDispensing.objects.filter(
                prescription__order=order,
                is_deleted=False,
                dispensing_status='pending_payment'
            ).exists():
                payment_verified = True
        
        # Cash: Check for payment receipt
        if not payment_verified:
            from .models_accounting import PaymentReceipt
            for prescription in prescriptions:
                if not patient:
                    patient = prescription.order.encounter.patient
                receipt_exists = PaymentReceipt.objects.filter(
                    is_deleted=False,
                    patient=patient,
                    service_type='pharmacy_prescription',
                    receipt_date__gte=prescription.created.date()
                ).exists()
                if receipt_exists:
                    payment_verified = True
                    break
        
        return JsonResponse({
            'success': True,
            'payment_verified': payment_verified,
            'prescriptions_count': prescriptions.count()
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@csrf_exempt
@login_required
@require_http_methods(["POST"])
def send_pharmacy_order_to_cashier(request, order_id):
    """Queue pharmacy order to payer/cashier from pharmacy workflow only."""
    try:
        from .utils_roles import is_pharmacy_user

        if not (is_pharmacy_user(request.user) or request.user.is_superuser):
            return JsonResponse({
                'success': False,
                'error': 'Only pharmacy can push medication orders to payer/cashier.',
                'message': 'Doctors should only place prescriptions. Pharmacy must push to payer/cashier.'
            }, status=403)

        order = get_object_or_404(
            Order.objects.select_related('encounter', 'encounter__patient'),
            pk=order_id, order_type='medication', is_deleted=False
        )
        prescriptions = order.prescriptions.filter(is_deleted=False).select_related('drug')
        
        if not prescriptions.exists():
            return JsonResponse({'success': False, 'error': 'No prescriptions found for this order'}, status=400)
        
        patient = order.encounter.patient if order.encounter else None
        if not patient:
            return JsonResponse({'success': False, 'error': 'Patient information missing for this order'}, status=400)

        payer_info = _get_patient_payer_info(patient, order.encounter)
        payer = payer_info.get('payer')

        total_amount = Decimal('0.00')
        created_records = 0
        already_paid = 0
        warnings = []
        
        # Get edited quantities and substitutions from JSON body if available
        quantities_data = {}
        substitutions_data = {}
        if request.content_type == 'application/json':
            try:
                body_data = json.loads(request.body)
                quantities_data = body_data.get('quantities', {})
                substitutions_data = body_data.get('substitutions', {})
            except:
                pass

        # Resolve invoice and create/update bills inside a transaction (_get_or_create_invoice uses select_for_update)
        with transaction.atomic():
            encounter_invoice = None
            if payer and order.encounter:
                encounter_invoice, _ = AutoBillingService._get_or_create_invoice(patient, order.encounter, payer)

            from .models import Drug
            for prescription in prescriptions:
                prescription_id_str = str(prescription.id)
                substitute_drug = None
                if prescription_id_str in substitutions_data:
                    sub = substitutions_data[prescription_id_str]
                    if isinstance(sub, dict) and sub.get('drug_id'):
                        try:
                            substitute_drug = Drug.objects.get(pk=sub['drug_id'], is_deleted=False, is_active=True)
                        except Drug.DoesNotExist:
                            pass

                drug_to_use = substitute_drug if substitute_drug else prescription.drug
                # Use same price as billing (get_drug_price_for_prescription includes insurance markup) so total matches bill
                from .utils_billing import get_drug_price_for_prescription
                unit_price = get_drug_price_for_prescription(drug_to_use, payer=payer)

                # Use edited quantity if provided, otherwise use prescribed quantity
                if prescription_id_str in quantities_data:
                    quantity = Decimal(str(quantities_data[prescription_id_str]))
                else:
                    quantity = Decimal(str(prescription.quantity or 0))

                # Skip prescriptions with 0 quantity (removed from dispensing list by pharmacist)
                if quantity <= 0:
                    dispensing_record = PharmacyDispensing.objects.filter(
                        prescription=prescription,
                        is_deleted=False
                    ).first()
                    if dispensing_record and not dispensing_record.payment_receipt_id:
                        dispensing_record.dispensing_status = 'cancelled'
                        dispensing_record.quantity_ordered = 0
                        dispensing_record.save(update_fields=['dispensing_status', 'quantity_ordered', 'modified'])
                    # Waive existing invoice lines for this prescription so the bill reflects removal
                    waived_invoice_ids = set()
                    waive_user = None
                    _u = getattr(request, 'user', None)
                    if _u is not None and getattr(_u, 'is_authenticated', False):
                        waive_user = _u
                    for line in InvoiceLine.objects.filter(
                        prescription=prescription,
                        is_deleted=False,
                        waived_at__isnull=True
                    ).select_related('invoice'):
                        line.waived_at = timezone.now()
                        line.waiver_reason = 'Removed by pharmacy before sending to payer'
                        if waive_user:
                            line.waived_by = waive_user
                        # Full save: InvoiceLine.save() sets discount_amount + line_total when waived;
                        # update_fields omitting those leaves stale amounts on the invoice.
                        line.save()
                        if getattr(line, 'invoice_id', None):
                            waived_invoice_ids.add(line.invoice_id)
                    for inv_id in waived_invoice_ids:
                        try:
                            inv = Invoice.all_objects.filter(pk=inv_id).first()
                            if inv:
                                inv.update_totals()
                        except Exception:
                            pass
                    continue

                total_amount += (unit_price * quantity)

                dispensing_record = PharmacyDispensing.objects.filter(
                    prescription=prescription,
                    is_deleted=False
                ).first()
                if not dispensing_record:
                    ensure = AutoBillingService.create_pharmacy_dispensing_record_only(prescription)
                    if not ensure.get('success'):
                        warnings.append(
                            ensure.get('message')
                            or ensure.get('error')
                            or 'Could not add prescription to pharmacy queue before billing.'
                        )
                        continue

                # Do not skip when already paid: still update the bill with pharmacy's edits (qty/drug)
                # create_pharmacy_bill updates existing lines when they already exist.

                # Always create/update bill with pharmacy edits when pushing to payer/cashier.
                # Pharmacy's edited drug and quantity become the source of truth.
                billing_result = AutoBillingService.create_pharmacy_bill(
                    prescription,
                    substitute_drug=substitute_drug if substitute_drug else None,
                    quantity_override=int(quantity),
                    payer=payer,
                    invoice=encounter_invoice,
                )
                if billing_result.get('success'):
                    created_records += 1
                    # Stock is reduced when medication is actually dispensed (or at insurer billing
                    # inside create_pharmacy_bill). Do not deduct here for cash — avoids removing
                    # stock before payment and matches "deduct on hand-out" policy.
                else:
                    warnings.append(billing_result.get('message') or 'Unable to create bill.')

            if total_amount <= 0:
                return JsonResponse({
                    'success': False,
                    'error': 'No medications to send. Add at least one medication to dispense.'
                }, status=400)

            if created_records == 0 and already_paid == len(prescriptions):
                return JsonResponse({
                    'success': False,
                    'error': 'All prescriptions for this order are already paid. Proceed to dispensing.'
                }, status=400)

            if created_records == 0 and total_amount > 0:
                return JsonResponse({
                    'success': False,
                    'error': warnings[0] if warnings else 'Unable to create bills. Please try again or contact support.'
                }, status=400)

            # Recalculate encounter invoice totals so the bill reflects all updates/waivers
            if encounter_invoice:
                try:
                    encounter_invoice.refresh_from_db()
                    encounter_invoice.update_totals()
                except Exception:
                    pass

            payer_queue_url = reverse('hospital:pharmacy_pending_dispensing')

            display_amount = str(encounter_invoice.total_amount) if encounter_invoice else str(total_amount)
            balance_amount = str(encounter_invoice.balance) if encounter_invoice else display_amount

            response_data = {
                'success': True,
                'created': created_records,
                'already_paid': already_paid,
                'amount': display_amount,
                'balance': balance_amount,
                'pharmacy_amount': str(total_amount),
                'patient_name': patient.full_name,
                'patient_mrn': patient.mrn,
                'payer_queue_url': payer_queue_url,
            }

            if warnings:
                response_data['warnings'] = warnings

            return JsonResponse(response_data)
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@login_required
@require_http_methods(["POST"])
def dispense_pharmacy_order(request, order_id):
    """Dispense medications for an order - pharmacy workflow only."""
    try:
        from .utils_roles import is_pharmacy_user

        if not (is_pharmacy_user(request.user) or request.user.is_superuser):
            return JsonResponse({
                'success': False,
                'error': 'Only pharmacy can dispense medication orders.',
                'message': 'Doctors should push prescriptions to pharmacy; pharmacy handles payer push and dispensing.'
            }, status=403)

        order = get_object_or_404(Order, pk=order_id, order_type='medication', is_deleted=False)
        
        # Get all prescriptions
        prescriptions = order.prescriptions.filter(is_deleted=False)
        
        if not prescriptions.exists():
            return JsonResponse({'success': False, 'error': 'No prescriptions found'}, status=400)
        
        # PAYMENT VERIFICATION: Check if payment has been made
        from .models_payment_verification import (
            PharmacyDispensing,
            PharmacyDispenseHistory,
            PharmacyStockDeductionLog,
        )
        from .services.unified_receipt_service import UnifiedReceiptService
        
        # Check if patient has paid for this order
        patient = order.encounter.patient if order.encounter else None
        
        if not patient:
            return JsonResponse({
                'success': False, 
                'error': 'Patient not found for this order'
            }, status=400)
        
        # Check if there's a receipt for these prescriptions
        has_payment = False
        payment_receipt = None
        
        # Check each prescription for payment
        for prescription in prescriptions:
            # Check if there's a PharmacyDispensing record with payment
            existing_dispensing = PharmacyDispensing.objects.filter(
                prescription=prescription,
                is_deleted=False
            ).first()
            
            if existing_dispensing and existing_dispensing.payment_receipt:
                has_payment = True
                payment_receipt = existing_dispensing.payment_receipt
                break
        
        # If no payment found, check if there's a receipt for this encounter
        if not has_payment:
            from .models import Receipt
            payment_receipts = Receipt.objects.filter(
                patient=patient,
                is_deleted=False,
                is_cancelled=False
            ).filter(
                Q(encounter=order.encounter) if order.encounter else Q(id__isnull=False)
            ).order_by('-created')
            
            if payment_receipts.exists():
                payment_receipt = payment_receipts.first()
                has_payment = True
        
        # Insurance/Corporate: Allow dispensing when bill has been created (sent to insurer).
        # No cashier payment required - insurer will be billed.
        if not has_payment:
            payer_info = _get_patient_payer_info(patient, order.encounter) if patient else {}
            if payer_info.get('is_insurance_or_corporate'):
                # Check that pharmacy dispensing records exist (bill was created via Send to Insurance)
                dispensing_count = PharmacyDispensing.objects.filter(
                    prescription__order=order,
                    is_deleted=False,
                    dispensing_status__in=['ready_to_dispense', 'partially_dispensed', 'fully_dispensed']
                ).count()
                if dispensing_count > 0:
                    has_payment = True  # Bill sent to insurer = OK to dispense
        
        # ENFORCE PAYMENT: Don't allow dispensing without payment (cash patients) or bill creation (insurance)
        if not has_payment:
            return JsonResponse({
                'success': False,
                'error': 'Payment required before dispensing',
                'message': 'Pharmacy must first push this order to payer/cashier before dispensing.',
                'payment_required': True
            }, status=403)
        
        # Get current staff
        current_staff = None
        try:
            current_staff = request.user.staff
        except:
            pass
        
        # Get quantities from JSON body (for AJAX requests) or POST data
        quantities_data = {}
        if request.content_type == 'application/json':
            import json as json_module
            try:
                body_data = json_module.loads(request.body)
                quantities_data = body_data.get('quantities', {})
            except:
                pass
        
        # Create dispensing records for each prescription
        dispensed_count = 0
        for prescription in prescriptions:
            # Check if already dispensed
            existing = PharmacyDispensing.objects.filter(
                prescription=prescription,
                is_deleted=False
            ).first()
            
            # Get actual quantity to dispense (may be edited by pharmacy)
            # Try JSON first, then POST, then default to prescription quantity
            prescription_id_str = str(prescription.id)
            if prescription_id_str in quantities_data:
                actual_quantity = Decimal(str(quantities_data[prescription_id_str]))
            else:
                actual_quantity = Decimal(str(request.POST.get(f'quantity_{prescription.id}', prescription.quantity)))
            
            # Skip prescriptions with 0 quantity (removed from dispensing list by pharmacist)
            if actual_quantity <= 0:
                continue
            
            drug_to_dispense = (existing.drug_to_dispense if existing else None) or prescription.drug
            qty_int = int(actual_quantity)

            disp_rec = existing
            did_dispense_this_request = False
            newly_dispensed_qty = 0
            note_text = (
                f'Dispensed from pharmacy dashboard (consulting-room origin, '
                f'Edited: {actual_quantity} of {prescription.quantity} prescribed)'
            )

            if not existing:
                # Create dispensing record with payment info
                disp_rec = PharmacyDispensing.objects.create(
                    prescription=prescription,
                    patient=patient,
                    dispensed_by=current_staff,
                    dispensed_at=timezone.now(),
                    quantity_ordered=qty_int,
                    quantity_dispensed=qty_int,
                    payment_receipt=payment_receipt,
                    payment_verified_at=timezone.now(),
                    dispensing_status='fully_dispensed',
                    dispensing_notes=note_text,
                )
                did_dispense_this_request = True
                newly_dispensed_qty = qty_int
            elif existing:
                # Update existing record
                prev_dispensed = int(existing.quantity_dispensed or 0)
                target_dispensed = max(prev_dispensed, qty_int)
                newly_dispensed_qty = max(0, target_dispensed - prev_dispensed)
                existing.dispensed_by = current_staff
                existing.dispensed_at = timezone.now()
                existing.quantity_ordered = max(int(existing.quantity_ordered or 0), qty_int)
                existing.quantity_dispensed = target_dispensed
                existing.payment_receipt = payment_receipt
                existing.payment_verified_at = timezone.now()
                if existing.quantity_dispensed >= int(existing.quantity_ordered or 0):
                    existing.dispensing_status = 'fully_dispensed'
                else:
                    existing.dispensing_status = 'partially_dispensed'
                existing.dispensing_notes = note_text
                existing.save(update_fields=[
                    'dispensed_by',
                    'dispensed_at',
                    'quantity_ordered',
                    'quantity_dispensed',
                    'payment_receipt',
                    'payment_verified_at',
                    'dispensing_status',
                    'dispensing_notes',
                    'modified',
                ])
                disp_rec = existing
                did_dispense_this_request = newly_dispensed_qty > 0

            # If this prescription was already dispensed previously, skip
            if (
                not did_dispense_this_request
                or not disp_rec
                or int(getattr(disp_rec, 'quantity_dispensed', 0) or 0) <= 0
            ):
                continue

            disp_hist = PharmacyDispenseHistory.objects.create(
                dispensing_record=disp_rec,
                prescription=prescription,
                patient=patient,
                patient_name=getattr(patient, 'full_name', str(patient)),
                drug=drug_to_dispense,
                drug_name=getattr(drug_to_dispense, 'name', str(drug_to_dispense)),
                quantity_dispensed=newly_dispensed_qty,
                instructions='',
                notes=note_text,
                counselling_given=False,
                dispensed_by=current_staff,
                dispensed_by_name=current_staff.user.get_full_name() if current_staff and current_staff.user else '',
                payment_receipt=disp_rec.payment_receipt,
                dispensed_at=disp_rec.dispensed_at or timezone.now(),
            )
            dispensed_count += 1

            # Reduce pharmacy stock when drugs are dispensed (skip if already reduced at Send to Payer)
            if (
                drug_to_dispense
                and newly_dispensed_qty > 0
                and not getattr(disp_rec, 'stock_reduced_at', None)
            ):
                from hospital.pharmacy_stock_utils import reduce_pharmacy_stock_once

                reduce_pharmacy_stock_once(
                    drug_to_dispense,
                    newly_dispensed_qty,
                    PharmacyStockDeductionLog.SOURCE_DISPENSE_HISTORY,
                    disp_hist.id,
                )
            elif getattr(disp_rec, 'stock_reduced_at', None):
                logger.info(
                    "Skipping stock deduction for order %s prescription %s; pre-deduct already recorded at %s",
                    getattr(order, 'id', None),
                    getattr(prescription, 'id', None),
                    getattr(disp_rec, 'stock_reduced_at', None),
                )
        
        if dispensed_count == 0:
            return JsonResponse({
                'success': False,
                'error': 'No medications to dispense. Add at least one medication to the dispensing list.'
            }, status=400)
        
        # Update order status
        order.status = 'completed'
        order.save()
        
        # Send customer service review SMS when medicine is dispensed
        if patient:
            try:
                from hospital.services.patient_feedback_service import send_customer_service_review_sms
                send_customer_service_review_sms(
                    patient,
                    message_type='pharmacy_dispensing_feedback',
                    related_object_id=None,
                    related_object_type='PharmacyDispensing',
                )
            except Exception:
                pass
        
        return JsonResponse({
            'success': True,
            'dispensed_count': dispensed_count,
            'message': f'{dispensed_count} prescription(s) dispensed successfully',
            'receipt_number': payment_receipt.receipt_number if payment_receipt else None
        })
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error processing dispensing: {error_details}")
        return JsonResponse({
            'success': False, 
            'error': 'Error processing dispensing',
            'details': str(e)
        }, status=500)


@login_required
def get_imaging_study_images(request, study_id):
    """Get all images for an imaging study via AJAX"""
    try:
        study = get_object_or_404(ImagingStudy, pk=study_id, is_deleted=False)
        
        images = study.images.filter(is_deleted=False).order_by('sequence_number', 'uploaded_at')
        
        images_data = []
        for img in images:
            images_data.append({
                'id': str(img.id),
                'url': img.image.url,
                'description': img.description,
                'sequence_number': img.sequence_number,
                'uploaded_at': img.uploaded_at.strftime('%Y-%m-%d %H:%M'),
                'uploaded_by': img.uploaded_by.user.get_full_name() if img.uploaded_by else 'Unknown'
            })
        
        return JsonResponse({
            'success': True,
            'images': images_data,
            'study': {
                'id': str(study.id),
                'modality': study.modality,
                'body_part': study.body_part,
                'status': study.status,
                'report_text': study.report_text or '',
                'findings': study.findings or '',
                'impression': study.impression or ''
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def imaging_teleradiology(request):
    """Teleradiology & Remote Reporting Dashboard"""
    from .models_advanced import ImagingStudy
    
    today = timezone.now().date()
    
    # Base queryset for remote studies (don't slice yet)
    remote_studies_qs = ImagingStudy.objects.filter(
        status__in=['awaiting_report', 'reporting'],
        is_deleted=False
    ).select_related('patient', 'assigned_radiologist__user', 'order__encounter__patient')
    
    # Calculate statistics BEFORE slicing
    stats = {
        'pending_remote': remote_studies_qs.filter(status='awaiting_report').count(),
        'in_progress_remote': remote_studies_qs.filter(status='reporting').count(),
        'completed_today': ImagingStudy.objects.filter(
            status='verified',
            report_verified_at__date=today,
            is_deleted=False
        ).count(),
    }
    
    # Now slice for display
    remote_studies = remote_studies_qs.order_by('-created')[:50]
    
    context = {
        'remote_studies': remote_studies,
        'stats': stats,
    }
    
    return render(request, 'hospital/imaging_teleradiology.html', context)


@login_required
def imaging_equipment_management(request):
    """Equipment & Quality Control Management Dashboard"""
    try:
        from .models_advanced import MedicalEquipment
        
        equipment_list = MedicalEquipment.objects.filter(
            equipment_type__in=['xray', 'ct', 'mri', 'ultrasound', 'mammography'],
            is_deleted=False
        ).order_by('name')
        
        # Statistics - MedicalEquipment uses: 'available', 'in_use', 'maintenance', 'out_of_order'
        stats = {
            'total': equipment_list.count(),
            'operational': equipment_list.filter(status__in=['available', 'in_use']).count(),
            'maintenance': equipment_list.filter(status='maintenance').count(),
            'down': equipment_list.filter(status='out_of_order').count(),
        }
        
        # Maintenance due soon (next 30 days)
        from datetime import timedelta
        maintenance_due_date = timezone.now().date() + timedelta(days=30)
        maintenance_due = equipment_list.filter(
            next_maintenance_due__lte=maintenance_due_date,
            next_maintenance_due__gte=timezone.now().date()
        )
        
    except (ImportError, AttributeError):
        equipment_list = []
        stats = {'total': 0, 'operational': 0, 'maintenance': 0, 'down': 0}
        maintenance_due = []
    
    context = {
        'equipment_list': equipment_list,
        'stats': stats,
        'maintenance_due': maintenance_due,
    }
    
    return render(request, 'hospital/imaging_equipment.html', context)


# ==================== IMAGING API ENDPOINTS ====================

@login_required
@require_http_methods(["GET"])
def export_imaging_queue(request):
    """Export imaging queue to CSV/Excel"""
    from django.http import HttpResponse
    import csv
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="imaging_queue.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Patient', 'MRN', 'Modality', 'Priority', 'Status', 'Requested By', 'Requested At'])
    
    orders = Order.objects.filter(
        order_type='imaging',
        is_deleted=False
    ).select_related('encounter__patient', 'requested_by__user')[:100]
    
    for order in orders:
        writer.writerow([
            order.encounter.patient.full_name if order.encounter else 'N/A',
            order.encounter.patient.mrn if order.encounter else 'N/A',
            order.get_modality_display() if hasattr(order, 'get_modality_display') else 'N/A',
            order.get_priority_display(),
            order.get_status_display(),
            order.requested_by.user.get_full_name() if order.requested_by and order.requested_by.user else 'N/A',
            order.requested_at.strftime('%Y-%m-%d %H:%M') if order.requested_at else 'N/A'
        ])
    
    return response


@login_required
def start_imaging_order(request, order_id):
    """Start imaging for an order"""
    from django.contrib import messages
    from django.shortcuts import redirect
    from django.http import JsonResponse
    
    # Check if this is an AJAX request (check multiple ways)
    is_ajax = (
        request.headers.get('X-Requested-With') == 'XMLHttpRequest' or
        request.GET.get('ajax') == 'true' or
        request.content_type == 'application/json' or
        '/api/' in request.path
    )
    
    try:
        order = get_object_or_404(Order, pk=order_id, order_type='imaging', is_deleted=False)
        
        # Verify encounter and patient exist
        if not order.encounter:
            error_msg = 'Order has no associated encounter.'
            if is_ajax:
                return JsonResponse({'success': False, 'error': error_msg}, status=400)
            messages.error(request, error_msg)
            return redirect('hospital:radiologist_pending_orders')
        
        if not order.encounter.patient:
            error_msg = 'Order has no associated patient.'
            if is_ajax:
                return JsonResponse({'success': False, 'error': error_msg}, status=400)
            messages.error(request, error_msg)
            return redirect('hospital:radiologist_pending_orders')
        
        # Update order status
        order.status = 'in_progress'
        order.save(update_fields=['status'])
        
        # Get existing imaging studies linked to this order
        existing_studies = ImagingStudy.objects.filter(
            order=order,
            is_deleted=False
        )
        
        if existing_studies.exists():
            # Update all existing studies to in_progress
            now = timezone.now()
            studies_to_update = []
            for existing_study in existing_studies:
                existing_study.status = 'in_progress'
                if not existing_study.started_at:
                    existing_study.started_at = now
                studies_to_update.append(existing_study)
            
            # Bulk update for efficiency
            ImagingStudy.objects.bulk_update(studies_to_update, ['status', 'started_at'], batch_size=50)
            
            # Use the first study for redirect
            study = existing_studies.first()
        else:
            # No existing studies - create a basic study as fallback
            # Try to get study details from ImagingCatalog if available
            from .models_advanced import ImagingCatalog
            
            modality = 'xray'
            body_part = ''
            study_type = 'Imaging Study'
            
            # Try to find any catalog item for this order (check notes or use default)
            catalog_items = ImagingCatalog.objects.filter(is_active=True, is_deleted=False).order_by('name')[:1]
            if catalog_items.exists():
                catalog = catalog_items.first()
                modality = catalog.modality
                body_part = catalog.body_part or ''
                study_type = catalog.code or catalog.name or 'Imaging Study'
            
            # Check for duplicate before creating
            existing_study = ImagingStudy.objects.filter(
                order=order,
                patient=order.encounter.patient,
                encounter=order.encounter,
                modality=modality,
                body_part=body_part,
                is_deleted=False
            ).first()
            
            if not existing_study:
                study = ImagingStudy.objects.create(
                    order=order,
                    patient=order.encounter.patient,
                    encounter=order.encounter,
                    modality=modality,
                    body_part=body_part,
                    study_type=study_type,
                    status='in_progress',
                    priority=order.priority,
                    started_at=timezone.now()
                )
            else:
                study = existing_study
                # Update status if needed
                if study.status != 'in_progress':
                    study.status = 'in_progress'
                    study.started_at = timezone.now()
                    study.save()
        
        # Return JSON for AJAX requests
        if is_ajax:
            return JsonResponse({
                'success': True,
                'message': 'Imaging started successfully',
                'study_id': str(study.id)
            })
        else:
            # Regular request - redirect to study detail page
            messages.success(request, 'Imaging order started successfully.')
            return redirect('hospital:imaging_study_detail', study_id=study.id)
        
    except Order.DoesNotExist:
        error_msg = 'Imaging order not found.'
        if is_ajax:
            return JsonResponse({'success': False, 'error': error_msg}, status=404)
        messages.error(request, error_msg)
        return redirect('hospital:radiologist_pending_orders')
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Error starting imaging order {order_id}: {str(e)}', exc_info=True)
        
        if is_ajax:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
        else:
            messages.error(request, f'Error starting imaging order: {str(e)}')
            return redirect('hospital:radiologist_pending_orders')


@login_required
@require_http_methods(["GET"])
def get_imaging_study_api(request, study_id):
    """Get imaging study details via API"""
    try:
        study = get_object_or_404(ImagingStudy, pk=study_id, is_deleted=False)
        
        return JsonResponse({
            'success': True,
            'study': {
                'id': str(study.id),
                'modality': study.modality,
                'body_part': study.body_part,
                'status': study.status,
                'report_text': study.report_text or '',
                'findings': study.findings or '',
                'impression': study.impression or '',
                'has_critical_findings': study.has_critical_findings,
                'critical_findings': study.critical_findings or '',
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@require_http_methods(["POST"])
def imaging_ai_analysis(request, study_id):
    """AI-assisted imaging analysis (placeholder)"""
    try:
        study = get_object_or_404(ImagingStudy, pk=study_id, is_deleted=False)
        
        # Placeholder for AI analysis
        # In production, this would call an AI service
        findings = [
            'AI analysis feature available',
            'Connect to AI service for automated findings',
            'AI suggestions will appear here when configured'
        ]
        
        return JsonResponse({
            'success': True,
            'findings': findings,
            'note': 'AI analysis service not yet configured. This is a placeholder.'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@require_http_methods(["POST"])
def save_imaging_report_api(request, study_id):
    """Save imaging report via API"""
    try:
        study = get_object_or_404(ImagingStudy, pk=study_id, is_deleted=False)
        
        import json
        data = json.loads(request.body)
        
        study.findings = data.get('findings', '')
        study.impression = data.get('impression', '')
        study.has_critical_findings = data.get('has_critical_findings', False)
        study.critical_findings = data.get('critical_findings', '')
        
        # Get current staff
        try:
            current_staff = request.user.staff
            study.report_dictated_by = current_staff
        except:
            pass
        
        if data.get('status') == 'draft':
            study.status = 'reporting'
        else:
            study.status = 'reported'
            study.report_text = f"FINDINGS:\n{study.findings}\n\nIMPRESSION:\n{study.impression}"
        
        study.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Report saved successfully'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@require_http_methods(["POST"])
def finalize_imaging_report_api(request, study_id):
    """Finalize imaging report via API"""
    try:
        study = get_object_or_404(ImagingStudy, pk=study_id, is_deleted=False)
        
        study.status = 'reported'
        if not study.report_text:
            study.report_text = f"FINDINGS:\n{study.findings}\n\nIMPRESSION:\n{study.impression}"
        study.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Report finalized successfully'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


def is_pharmacy_or_nurse_staff(user):
    """Check if user is pharmacy/nursing/procurement staff for pharmacy operations."""
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    try:
        from .utils_roles import is_procurement_staff as _is_procurement_staff
        if _is_procurement_staff(user):
            return True
    except Exception:
        pass
    try:
        if hasattr(user, 'staff') and user.staff:
            # Check profession
            if user.staff.profession in [
                'pharmacist', 'pharmacy_technician',
                'nurse', 'registered_nurse', 'enrolled_nurse',
                'procurement_officer', 'store_manager', 'inventory_manager'
            ]:
                return True
            # Check department
            if user.staff.department:
                dept_name = user.staff.department.name.lower()
                if (
                    'pharmacy' in dept_name or 'nursing' in dept_name or 'nurse' in dept_name
                    or 'procurement' in dept_name or 'store' in dept_name
                ):
                    return True
    except:
        pass
    # Also allow through groups
    return user.groups.filter(name__in=[
        'Admin', 'Administrator',
        'Pharmacy', 'Pharmacist',
        'Nurse', 'Nursing',
        'Procurement', 'Procurement Officer',
        'Store Manager', 'Inventory Stores Manager',
    ]).exists()


@login_required
def api_pharmacy_consumables_search(request):
    """Search pharmacy items (Drug + Clinical Consumables) for consumables billing. Permission enforced by parent page."""
    try:
        query = (request.GET.get('q') or '').strip()
        if len(query) < 2:
            return JsonResponse({'results': []})
        results = []
        # Search Drugs (pharmacy formulary)
        drugs = Drug.objects.filter(
            Q(name__icontains=query) | Q(generic_name__icontains=query) | Q(atc_code__icontains=query) | Q(strength__icontains=query),
            is_active=True,
            is_deleted=False
        )[:15]
        for d in drugs:
            desc = f"{d.name} {d.strength or ''} {d.form or ''}".strip()
            price = getattr(d, 'unit_price', None) or Decimal('0.00')
            results.append({
                'code': f'DRUG-{d.id}',
                'description': desc,
                'unit_price': str(price),
                'source': 'drug',
            })
        # Search ServiceCode Clinical Consumables
        consumables = ServiceCode.objects.filter(
            category='Clinical Consumables',
            is_active=True,
            is_deleted=False
        ).filter(
            Q(code__icontains=query) | Q(description__icontains=query)
        )[:15]
        seen_codes = {r['code'] for r in results}
        for sc in consumables:
            if sc.code in seen_codes:
                continue
            seen_codes.add(sc.code)
            price = Decimal('30.00')
            try:
                fp = getattr(sc, 'flexible_prices', None)
                if fp:
                    first_price = fp.filter(is_deleted=False).first()
                    if first_price:
                        price = first_price.price
            except Exception:
                pass
            results.append({
                'code': sc.code,
                'description': sc.description or sc.code,
                'unit_price': str(price),
                'source': 'consumable',
            })
        return JsonResponse({'results': results})
    except Exception as e:
        import logging
        logging.getLogger(__name__).exception('Consumables search failed')
        return JsonResponse({'results': [], 'error': str(e)}, status=500)


@login_required
@user_passes_test(is_pharmacy_or_nurse_staff, login_url='/admin/login/')
def pharmacy_add_consumables(request):
    """
    Enhanced pharmacy view for procedure-first consumables billing
    Workflow: Select Patient → Select Procedure → Select Consumables → Bill to Session
    Access: Pharmacy staff, Nurses, Procurement/Stores, and Admins
    """
    from .models_advanced import ProcedureCatalog
    from .services.pricing_engine_service import PricingEngineService
    
    pricing_engine = PricingEngineService()
    
    # Handle AJAX request for creating new procedure
    if request.method == 'POST' and request.POST.get('action') == 'create_procedure':
        from django.http import JsonResponse
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.POST.get('ajax'):
            try:
                code = request.POST.get('code', '').strip()
                name = request.POST.get('name', '').strip()
                category = request.POST.get('category', 'other')
                description = request.POST.get('description', '').strip()
                price = Decimal(request.POST.get('price', '0') or '0')
                
                if not code or not name or price < 0:
                    return JsonResponse({'success': False, 'error': 'Code, name, and price are required.'})
                
                # Check if code already exists
                if ProcedureCatalog.objects.filter(code=code, is_deleted=False).exists():
                    return JsonResponse({'success': False, 'error': f'Procedure code "{code}" already exists.'})
                
                # Create procedure
                procedure = ProcedureCatalog.objects.create(
                    code=code,
                    name=name,
                    category=category,
                    description=description,
                    price=price,
                    is_active=True
                )
                
                return JsonResponse({
                    'success': True,
                    'message': f'Procedure "{name}" created successfully.',
                    'id': str(procedure.id),
                    'code': code,
                    'name': name,
                    'category': category,
                    'price': float(price)
                })
            except Exception as e:
                return JsonResponse({'success': False, 'error': str(e)})
    
    # Handle AJAX request for creating new consumable
    if request.method == 'POST' and request.POST.get('action') == 'create_consumable':
        from django.http import JsonResponse
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.POST.get('ajax'):
            try:
                code = request.POST.get('code', '').strip()
                description = request.POST.get('description', '').strip()
                price = Decimal(request.POST.get('price', '0') or '0')
                
                if not code or not description or price <= 0:
                    return JsonResponse({'success': False, 'error': 'Invalid data provided.'})
                
                # Check if code already exists
                if ServiceCode.objects.filter(code=code, is_deleted=False).exists():
                    return JsonResponse({'success': False, 'error': f'Code "{code}" already exists.'})
                
                # Create service code
                service_code = ServiceCode.objects.create(
                    code=code,
                    description=description,
                    category='Clinical Consumables',
                    is_active=True
                )
                
                # Create default price in flexible pricing
                from .models_flexible_pricing import PricingCategory, ServicePrice
                cash_category = PricingCategory.objects.filter(
                    category_type='cash',
                    is_active=True,
                    is_deleted=False
                ).first()
                
                if cash_category:
                    ServicePrice.objects.create(
                        service_code=service_code,
                        pricing_category=cash_category,
                        price=price,
                        is_active=True
                    )
                
                return JsonResponse({
                    'success': True,
                    'message': f'Consumable "{description}" created successfully.',
                    'code': code,
                    'description': description,
                    'price': float(price)
                })
            except Exception as e:
                return JsonResponse({'success': False, 'error': str(e)})
    
    # Handle AJAX request for creating new procedure
    if request.method == 'POST' and request.POST.get('action') == 'create_procedure':
        from django.http import JsonResponse
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.POST.get('ajax'):
            try:
                code = request.POST.get('code', '').strip()
                name = request.POST.get('name', '').strip()
                category = request.POST.get('category', 'other')
                description = request.POST.get('description', '').strip()
                price = Decimal(request.POST.get('price', '0') or '0')
                
                if not code or not name or price < 0:
                    return JsonResponse({'success': False, 'error': 'Code, name, and price are required.'})
                
                # Check if code already exists
                if ProcedureCatalog.objects.filter(code=code, is_deleted=False).exists():
                    return JsonResponse({'success': False, 'error': f'Procedure code "{code}" already exists.'})
                
                # Create procedure
                procedure = ProcedureCatalog.objects.create(
                    code=code,
                    name=name,
                    category=category,
                    description=description,
                    price=price,
                    is_active=True
                )
                
                return JsonResponse({
                    'success': True,
                    'message': f'Procedure "{name}" created successfully.',
                    'id': str(procedure.id),
                    'code': code,
                    'name': name,
                    'category': category,
                    'price': float(price)
                })
            except Exception as e:
                return JsonResponse({'success': False, 'error': str(e)})
    
    # Handle main form submission
    if request.method == 'POST' and request.POST.get('action') != 'create_consumable':
        patient_id = request.POST.get('patient_id')
        encounter_id = request.POST.get('encounter_id')
        procedure_id = request.POST.get('procedure_id')
        consumables_json = request.POST.get('consumables', '[]')
        nurses_notes = request.POST.get('nurses_notes', '').strip()
        
        if not patient_id:
            messages.error(request, 'Patient is required.')
            return redirect('hospital:pharmacy_add_consumables')
        
        try:
            patient = Patient.objects.get(id=patient_id, is_deleted=False)
        except Patient.DoesNotExist:
            messages.error(request, 'Patient not found.')
            return redirect('hospital:pharmacy_add_consumables')
        
        # Get or create encounter
        encounter = None
        if encounter_id:
            try:
                encounter = Encounter.objects.get(id=encounter_id, patient=patient, is_deleted=False)
            except Encounter.DoesNotExist:
                messages.warning(request, 'Encounter not found. Creating new encounter.')
        
        # Get or create encounter - if no encounter, use most recent active one or create new
        if not encounter:
            encounter = patient.encounters.filter(
                status='active',
                is_deleted=False
            ).order_by('-created').first()
        
        # If still no encounter, create a quick visit encounter for billing
        if not encounter:
            encounter = Encounter.objects.create(
                patient=patient,
                encounter_type='outpatient',
                status='active',
                visit_type='procedure',
                notes='Auto-created for procedure and consumables billing'
            )
        
        # Get or create invoice for encounter
        from .utils_billing import get_or_create_encounter_invoice
        invoice = get_or_create_encounter_invoice(encounter)
        
        if not invoice:
            messages.error(request, 'Failed to create invoice.')
            return redirect('hospital:pharmacy_add_consumables')
        
        # Get patient's payer
        payer = patient.primary_insurance
        if not payer:
            payer = Payer.objects.filter(payer_type='cash', is_active=True, is_deleted=False).first()
            if not payer:
                payer = Payer.objects.create(name='Cash', payer_type='cash', is_active=True)
        
        # Parse consumables from JSON
        try:
            consumables_data = json.loads(consumables_json)
        except json.JSONDecodeError:
            messages.error(request, 'Invalid consumables data.')
            return redirect('hospital:pharmacy_add_consumables')
        
        added_count = 0
        errors = []
        try:
            with transaction.atomic():
                # Add procedure to invoice if selected
                if procedure_id:
                    try:
                        procedure = ProcedureCatalog.objects.get(id=procedure_id, is_deleted=False, is_active=True)
                        
                        # Get procedure price
                        procedure_price = procedure.price or Decimal('0.00')
                        if payer.payer_type == 'corporate' and procedure.corporate_price:
                            procedure_price = procedure.corporate_price
                        elif payer.payer_type in ('nhis', 'insurance', 'private') and procedure.insurance_price:
                            procedure_price = procedure.insurance_price
                        
                        # Get or create service code for procedure
                        procedure_service_code, _ = ServiceCode.objects.get_or_create(
                            code=f'PROC-{procedure.code}',
                            defaults={
                                'description': f'{procedure.name} (Procedure)',
                                'category': 'Procedures',
                                'is_active': True
                            }
                        )
                        
                        # Create invoice line for procedure
                        InvoiceLine.objects.create(
                            invoice=invoice,
                            service_code=procedure_service_code,
                            description=f'{procedure.name} - {procedure.get_category_display()}',
                            quantity=Decimal('1.00'),
                            unit_price=procedure_price,
                            line_total=procedure_price
                        )
                        added_count += 1
                    except ProcedureCatalog.DoesNotExist:
                        errors.append('Selected procedure not found.')
                    except ValidationError as e:
                        err_msg = (e.messages[0] if isinstance(getattr(e, 'messages', None), list) and getattr(e, 'messages', None) else str(e))
                        errors.append(f'Procedure: {err_msg}')
                    except Exception as e:
                        errors.append(f'Error adding procedure: {str(e)}')
            
                # Add consumables to invoice
                for item in consumables_data:
                    service_code_str = item.get('code') or ''
                    quantity = Decimal(str(item.get('quantity', 1)))
                    custom_price = item.get('custom_price')  # Optional override price
                    description_override = item.get('description')
                    
                    if not service_code_str or quantity <= 0:
                        continue
                    
                    try:
                        unit_price = None
                        # Handle DRUG-{id} from pharmacy formulary
                        if service_code_str.startswith('DRUG-'):
                            drug_id = service_code_str.replace('DRUG-', '')
                            try:
                                drug = Drug.objects.get(pk=drug_id, is_deleted=False, is_active=True)
                            except (Drug.DoesNotExist, ValueError):
                                errors.append(f'Drug {service_code_str} not found.')
                                continue
                            unit_price = Decimal(str(custom_price)) if custom_price else (drug.unit_price or Decimal('0.00'))
                            desc = description_override or f"{drug.name} {drug.strength or ''} {drug.form or ''}".strip()
                            service_code, _ = ServiceCode.objects.get_or_create(
                                code=f'DRUG-{drug.id}',
                                defaults={
                                    'description': desc,
                                    'category': 'Pharmacy Services',
                                    'is_active': True
                                }
                            )
                            if not service_code.description or service_code.description != desc:
                                service_code.description = desc
                                service_code.save(update_fields=['description'])
                        else:
                            # Clinical Consumables (existing flow)
                            service_code = ServiceCode.objects.get(
                                code=service_code_str,
                                category='Clinical Consumables',
                                is_active=True,
                                is_deleted=False
                            )
                        
                        # Get price (skip for DRUG - already set above)
                        if unit_price is None:
                            if custom_price:
                                unit_price = Decimal(str(custom_price))
                            else:
                                unit_price = pricing_engine.get_service_price(
                                    service_code=service_code,
                                    patient=patient,
                                    payer=payer
                                )
                        
                        if unit_price <= 0:
                            # Try to get price from flexible pricing (cash category)
                            from .models_flexible_pricing import PricingCategory, ServicePrice
                            cash_category = PricingCategory.objects.filter(
                                category_type='cash',
                                is_active=True,
                                is_deleted=False
                            ).first()
                            if cash_category:
                                cash_price = ServicePrice.objects.filter(
                                    service_code=service_code,
                                    pricing_category=cash_category,
                                    is_active=True,
                                    is_deleted=False
                                ).first()
                                if cash_price:
                                    unit_price = cash_price.price
                            
                            # Fallback to default if still 0
                            if unit_price <= 0:
                                unit_price = Decimal('30.00')
                        
                        # Calculate line total
                        line_total = quantity * unit_price
                        
                        # Check if line already exists (prevent duplicates)
                        existing_line = InvoiceLine.objects.filter(
                            invoice=invoice,
                            service_code=service_code,
                            is_deleted=False
                        ).first()
                        
                        if existing_line:
                            # Update existing line
                            existing_line.quantity += quantity
                            existing_line.line_total = existing_line.quantity * unit_price
                            existing_line.unit_price = unit_price  # Update to current price
                            existing_line.save()
                            added_count += 1
                        else:
                            # Create new line
                            InvoiceLine.objects.create(
                                invoice=invoice,
                                service_code=service_code,
                                description=service_code.description,
                                quantity=quantity,
                                unit_price=unit_price,
                                line_total=line_total
                            )
                            added_count += 1
                            
                    except ServiceCode.DoesNotExist:
                        errors.append(f'Consumable {service_code_str} not found.')
                    except ValidationError as e:
                        err_msg = (e.messages[0] if isinstance(getattr(e, 'messages', None), list) and getattr(e, 'messages', None) else str(e))
                        errors.append(f'{service_code_str}: {err_msg}')
                    except Exception as e:
                        errors.append(f'Error adding {service_code_str}: {str(e)}')
            
                # Update invoice totals
                invoice.update_totals()
                
                # Update invoice status to 'issued' if it was draft
                if invoice.status == 'draft' and invoice.total_amount > 0:
                    invoice.status = 'issued'
                    invoice.save()
                
                # Save nurses notes to encounter
                if nurses_notes:
                    from .models_advanced import ClinicalNote
                    current_staff = None
                    if hasattr(request.user, 'staff'):
                        current_staff = request.user.staff
                    
                    ClinicalNote.objects.create(
                        encounter=encounter,
                        note_type='procedure',
                        notes=nurses_notes,
                        created_by=current_staff
                    )
                    
                    # Also update encounter notes
                    if encounter.notes:
                        encounter.notes += f"\n\n[Nurse's Note - {timezone.now().strftime('%Y-%m-%d %H:%M')}]\n{nurses_notes}"
                    else:
                        encounter.notes = f"[Nurse's Note - {timezone.now().strftime('%Y-%m-%d %H:%M')}]\n{nurses_notes}"
                    encounter.save(update_fields=['notes'])
        
            if added_count > 0:
                messages.success(
                    request,
                    f'Successfully added {added_count} item(s) to patient bill. '
                    f'Total: GHS {invoice.total_amount:.2f}'
                )
            if errors:
                for error in errors[:5]:  # Show first 5 errors
                    messages.warning(request, error)
            
            return redirect('hospital:pharmacy_add_consumables')
        except Exception as e:
            messages.error(request, f'Could not add consumables: {str(e)}')
            return redirect('hospital:pharmacy_add_consumables')
    
    # GET request - show form
    # Get all available procedures
    procedures = ProcedureCatalog.objects.filter(
        is_active=True,
        is_deleted=False
    ).order_by('category', 'name')
    
    # Get procedure categories for filter
    procedure_categories = ProcedureCatalog.PROCEDURE_CATEGORIES
    
    # Get all available consumables with pagination
    from django.core.paginator import Paginator
    consumables_qs = ServiceCode.objects.filter(
        category='Clinical Consumables',
        is_active=True,
        is_deleted=False
    ).prefetch_related('flexible_prices__pricing_category').order_by('code', 'description')
    
    # Paginate consumables (25 per page)
    paginator = Paginator(consumables_qs, 25)
    page = request.GET.get('page', 1)
    consumables_page = paginator.get_page(page)
    
    # Get recent patients with active encounters (for quick selection)
    from datetime import timedelta
    recent_cutoff = timezone.now() - timedelta(days=7)
    recent_patients = Patient.objects.filter(
        encounters__created__gte=recent_cutoff,
        encounters__status='active',
        is_deleted=False
    ).distinct().select_related('primary_insurance').order_by('-encounters__created')[:10]
    
    context = {
        'title': 'Procedure & Consumables Billing',
        'procedures': procedures,
        'procedure_categories': procedure_categories,
        'consumables': consumables_page,
        'consumables_page': consumables_page,  # For pagination
        'recent_patients': recent_patients,
    }
    return render(request, 'hospital/pharmacy_add_consumables.html', context)


@login_required
@user_passes_test(is_pharmacy_or_nurse_staff, login_url='/admin/login/')
def pharmacy_prescribe(request):
    """
    Allow authorized pharmacy operations staff to prescribe drugs for patients.
    Creates Order + Prescription (status pending) -> flows to pharmacy queue for dispensing.
    """
    from .models import Prescription, Order, Drug
    
    # Get recent patients for quick selection
    recent_patients = list(
        Patient.objects.filter(
            encounters__is_deleted=False,
            is_deleted=False
        ).distinct().select_related('primary_insurance').order_by('-encounters__created')[:10]
    )
    
    if request.method == 'POST':
        patient_id = request.POST.get('patient_id')
        drug_id = request.POST.get('drug_id')
        quantity = int(request.POST.get('quantity', 1) or 1)
        dose = (request.POST.get('dosage_instructions') or request.POST.get('dose') or 'As directed')[:100]
        frequency = (request.POST.get('frequency') or 'Once daily')[:50]
        duration_days = request.POST.get('duration_days', '')
        duration_str = f'{duration_days} days' if duration_days and duration_days.isdigit() else (duration_days or 'As directed')[:50]
        route = (request.POST.get('route') or 'oral')[:50]
        
        if not patient_id or not drug_id:
            messages.error(request, 'Patient and drug are required.')
            return redirect('hospital:pharmacy_prescribe')
        
        try:
            patient = Patient.objects.get(id=patient_id, is_deleted=False)
            drug = Drug.objects.get(pk=drug_id, is_deleted=False)
        except (Patient.DoesNotExist, Drug.DoesNotExist):
            messages.error(request, 'Patient or drug not found.')
            return redirect('hospital:pharmacy_prescribe')
        
        try:
            current_staff = Staff.objects.get(user=request.user, is_active=True, is_deleted=False)
        except Staff.DoesNotExist:
            messages.error(request, 'You must be registered as staff to prescribe.')
            return redirect('hospital:pharmacy_prescribe')
        
        # Get or create encounter (reuse active or create outpatient)
        encounter = patient.encounters.filter(
            status='active',
            is_deleted=False
        ).order_by('-created').first()
        
        if not encounter:
            encounter = Encounter.objects.create(
                patient=patient,
                encounter_type='outpatient',
                status='active',
                chief_complaint='Pharmacy prescription',
                notes='Created by pharmacy for prescription',
                provider=current_staff
            )
        
        with transaction.atomic():
            # Use existing pending medication order or create one (avoids MultipleObjectsReturned if duplicates exist)
            med_order = Order.objects.filter(
                encounter=encounter,
                order_type='medication',
                status='pending',
                is_deleted=False,
            ).order_by('-id').first()
            if med_order is None:
                med_order = Order.objects.create(
                    encounter=encounter,
                    order_type='medication',
                    status='pending',
                    is_deleted=False,
                    requested_by=current_staff,
                    priority='routine',
                )
            
            Prescription.objects.create(
                order=med_order,
                drug=drug,
                quantity=quantity,
                dose=dose,
                frequency=frequency,
                duration=duration_str,
                route=route,
                prescribed_by=current_staff,
            )
        
        messages.success(request, f'Prescription for {drug.name} added. It will appear in the pharmacy queue.')
        return redirect('hospital:pharmacy_prescribe')
    
    context = {
        'title': 'Pharmacy Prescribe',
        'recent_patients': recent_patients,
    }
    return render(request, 'hospital/pharmacy_prescribe.html', context)


@login_required
@require_http_methods(["POST"])
def verify_imaging_report_api(request, study_id):
    """Verify and release imaging report via API"""
    try:
        study = get_object_or_404(ImagingStudy, pk=study_id, is_deleted=False)
        
        # Get current staff
        try:
            current_staff = request.user.staff
            study.report_verified_by = current_staff
        except:
            pass
        
        study.report_verified_at = timezone.now()
        study.status = 'verified'
        study.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Report verified and released successfully'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
