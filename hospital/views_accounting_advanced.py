"""
Advanced Accounting Views
Financial reports, dashboards, and accounting operations
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, Q, F, Count
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
from types import SimpleNamespace
import json

from .models_accounting import Account, CostCenter, PaymentReceipt
from .models_accounting_advanced import (
    FiscalYear, AccountingPeriod, Journal, AdvancedJournalEntry, AdvancedJournalEntryLine,
    AdvancedGeneralLedger, PaymentVoucher, ReceiptVoucher,
    Revenue, RevenueCategory, Expense, ExpenseCategory,
    AdvancedAccountsReceivable, AccountsPayable,
    BankAccount, BankTransaction, Budget, BudgetLine
)
from .decorators import role_required


def is_accountant(user):
    """Check if user has accounting permissions"""
    return user.is_superuser or user.is_staff or user.groups.filter(name__in=['Accountant', 'Finance']).exists()


@login_required
@user_passes_test(is_accountant)
def accounting_dashboard(request):
    """Main accounting dashboard with KPIs and quick links"""
    
    # Check if tables exist (database-agnostic)
    from django.db import connection
    from django.core.management.color import no_style
    from django.db import models
    
    tables_exist = True
    try:
        # Use Django's introspection to check if table exists
        db_table = Revenue._meta.db_table
        with connection.cursor() as cursor:
            if connection.vendor == 'postgresql':
                cursor.execute(
                    "SELECT EXISTS (SELECT FROM information_schema.tables WHERE table_schema = 'public' AND table_name = %s)",
                    [db_table]
                )
                result = cursor.fetchone()
                if not result or not result[0]:  # EXISTS returns boolean
                    tables_exist = False
            elif connection.vendor == 'sqlite':
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", [db_table])
                if not cursor.fetchone():
                    tables_exist = False
            elif connection.vendor == 'mysql':
                cursor.execute("SELECT table_name FROM information_schema.tables WHERE table_schema = DATABASE() AND table_name = %s", [db_table])
                if not cursor.fetchone():
                    tables_exist = False
            else:
                # Fallback: try to query the table
                try:
                    cursor.execute(f"SELECT 1 FROM {db_table} LIMIT 1")
                except Exception:
                    tables_exist = False
    except Exception:
        tables_exist = False
    
    # Get current period
    today = timezone.now().date()
    try:
        fiscal_year = FiscalYear.objects.filter(start_date__lte=today, end_date__gte=today, is_closed=False).first()
        accounting_period = AccountingPeriod.objects.filter(start_date__lte=today, end_date__gte=today, is_closed=False).first()
    except:
        fiscal_year = None
        accounting_period = None
        tables_exist = False
    
    # Revenue Statistics (Current Month)
    start_of_month = today.replace(day=1)
    try:
        # Try Revenue model first (advanced accounting)
        total_revenue = Revenue.objects.filter(
            revenue_date__gte=start_of_month,
            revenue_date__lte=today,
            is_deleted=False
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        # Fallback to General Ledger if Revenue model is empty
        if total_revenue == 0:
            from .models_accounting import GeneralLedger, Account
            total_revenue = GeneralLedger.objects.filter(
                account__account_type='revenue',
                transaction_date__gte=start_of_month,
                transaction_date__lte=today,
                is_deleted=False
            ).aggregate(total=Sum('credit_amount'))['credit_amount__sum'] or Decimal('0.00')
            
            # If still 0, try PaymentReceipts
            if total_revenue == 0:
                from .models_accounting import PaymentReceipt
                total_revenue = PaymentReceipt.objects.filter(
                    receipt_date__date__gte=start_of_month,
                    receipt_date__date__lte=today,
                    is_deleted=False
                ).aggregate(total=Sum('amount_paid'))['amount_paid__sum'] or Decimal('0.00')
    except Exception as e:
        # Fallback to General Ledger
        try:
            from .models_accounting import GeneralLedger, Account
            total_revenue = GeneralLedger.objects.filter(
                account__account_type='revenue',
                transaction_date__gte=start_of_month,
                transaction_date__lte=today,
                is_deleted=False
            ).aggregate(total=Sum('credit_amount'))['credit_amount__sum'] or Decimal('0.00')
        except:
            total_revenue = Decimal('0.00')
    
    # Expense Statistics (Current Month)
    try:
        total_expenses = Expense.objects.filter(
            expense_date__gte=start_of_month,
            expense_date__lte=today,
            status='paid'
        ).aggregate(total=Sum('amount'))['total'] or 0
    except:
        total_expenses = 0
    
    # Accounts Receivable
    try:
        total_receivable = AdvancedAccountsReceivable.objects.filter(
            balance_due__gt=0
        ).aggregate(total=Sum('balance_due'))['total'] or 0
        
        overdue_receivable = AdvancedAccountsReceivable.objects.filter(
            is_overdue=True
        ).aggregate(total=Sum('balance_due'))['total'] or 0
    except:
        total_receivable = 0
        overdue_receivable = 0
    
    # Accounts Payable
    # Priority: Use General Ledger (Excel imported balances) if available, otherwise use AccountsPayable model
    total_payable = Decimal('0.00')
    try:
        # First, check General Ledger for AP accounts (Excel imported balances)
        # For Excel imports: debit amounts ARE the balances (independent, different companies)
        ap_accounts = Account.objects.filter(
            account_type='liability',
            account_name__icontains='payable',
            is_deleted=False
        )
        
        for ap_account in ap_accounts:
            # Sum all debit amounts (each is an independent balance from Excel import)
            ap_gl_total = AdvancedGeneralLedger.objects.filter(
                account=ap_account,
                is_voided=False,
                is_deleted=False
            ).aggregate(total=Sum('debit_amount'))['total'] or Decimal('0.00')
            total_payable += ap_gl_total
        
        # If General Ledger has no AP data, fall back to AccountsPayable model
        if total_payable == 0:
            total_payable = AccountsPayable.objects.filter(
                balance_due__gt=0,
                is_deleted=False
            ).aggregate(total=Sum('balance_due'))['total'] or Decimal('0.00')
    except Exception as e:
        # Fallback to AccountsPayable model
        try:
            total_payable = AccountsPayable.objects.filter(
                balance_due__gt=0,
                is_deleted=False
            ).aggregate(total=Sum('balance_due'))['total'] or Decimal('0.00')
        except:
            total_payable = Decimal('0.00')
    
    # Payment Vouchers
    try:
        pending_vouchers = PaymentVoucher.objects.filter(status='pending_approval').count()
        approved_vouchers = PaymentVoucher.objects.filter(status='approved').count()
    except:
        pending_vouchers = 0
        approved_vouchers = 0
    
    # Procurement Requests pending accounts approval
    try:
        from .models_procurement import ProcurementRequest
        pending_procurement_count = ProcurementRequest.objects.filter(
            status='admin_approved',
            is_deleted=False
        ).count()
    except:
        pending_procurement_count = 0
    
    # Journal Entries
    try:
        draft_entries = AdvancedJournalEntry.objects.filter(status='draft').count()
        posted_entries_month = AdvancedJournalEntry.objects.filter(
            entry_date__gte=start_of_month,
            status='posted'
        ).count()
    except:
        draft_entries = 0
        posted_entries_month = 0
    
    # Revenue by Category (Current Month)
    try:
        revenue_by_category = Revenue.objects.filter(
            revenue_date__gte=start_of_month
        ).values('category__name').annotate(
            total=Sum('amount')
        ).order_by('-total')[:5]
    except:
        revenue_by_category = []
    
    # Expenses by Category (Current Month)
    try:
        expenses_by_category = Expense.objects.filter(
            expense_date__gte=start_of_month,
            status='paid'
        ).values('category__name').annotate(
            total=Sum('amount')
        ).order_by('-total')[:5]
    except:
        expenses_by_category = []
    
    # AR Aging Summary
    try:
        ar_aging = {
            'current': AdvancedAccountsReceivable.objects.filter(aging_bucket='current').aggregate(total=Sum('balance_due'))['total'] or 0,
            '0_30': AdvancedAccountsReceivable.objects.filter(aging_bucket='0-30').aggregate(total=Sum('balance_due'))['total'] or 0,
            '31_60': AdvancedAccountsReceivable.objects.filter(aging_bucket='31-60').aggregate(total=Sum('balance_due'))['total'] or 0,
            '61_90': AdvancedAccountsReceivable.objects.filter(aging_bucket='61-90').aggregate(total=Sum('balance_due'))['total'] or 0,
            '90_plus': AdvancedAccountsReceivable.objects.filter(aging_bucket='90+').aggregate(total=Sum('balance_due'))['total'] or 0,
        }
    except:
        ar_aging = {
            'current': 0,
            '0_30': 0,
            '31_60': 0,
            '61_90': 0,
            '90_plus': 0,
        }
    
    # If tables don't exist, show preview template
    if not tables_exist:
        return render(request, 'hospital/accounting_preview.html', {})
    
    context = {
        'tables_exist': tables_exist,
        'fiscal_year': fiscal_year,
        'accounting_period': accounting_period,
        'total_revenue': total_revenue,
        'total_expenses': total_expenses,
        'net_income': total_revenue - total_expenses,
        'total_receivable': total_receivable,
        'overdue_receivable': overdue_receivable,
        'total_payable': total_payable,
        'pending_vouchers': pending_vouchers,
        'approved_vouchers': approved_vouchers,
        'pending_procurement_count': pending_procurement_count,
        'draft_entries': draft_entries,
        'posted_entries_month': posted_entries_month,
        'revenue_by_category': revenue_by_category,
        'expenses_by_category': expenses_by_category,
        'ar_aging': ar_aging,
    }
    
    return render(request, 'hospital/accounting_dashboard.html', context)


@login_required
@user_passes_test(is_accountant)
def accounting_reports_hub(request):
    """Central landing page for all accounting and finance reports"""
    today = timezone.now().date()
    start_of_month = today.replace(day=1)
    start_of_year = today.replace(month=1, day=1)

    def safe_sum(queryset, field_name):
        try:
            return queryset.aggregate(total=Sum(field_name))['total'] or Decimal('0.00')
        except Exception:
            return Decimal('0.00')

    def safe_count(queryset):
        try:
            return queryset.count()
        except Exception:
            return 0

    # Check if advanced accounting tables exist
    from django.db import connection
    tables_exist = True
    try:
        with connection.cursor() as cursor:
            # Database-agnostic table check
            db_table = Revenue._meta.db_table
            if connection.vendor == 'postgresql':
                cursor.execute(
                    "SELECT EXISTS (SELECT FROM information_schema.tables WHERE table_schema = 'public' AND table_name = %s)",
                    [db_table]
                )
                result = cursor.fetchone()
                if not result or not result[0]:  # EXISTS returns boolean
                    tables_exist = False
            elif connection.vendor == 'sqlite':
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", [db_table])
                if not cursor.fetchone():
                    tables_exist = False
            elif connection.vendor == 'mysql':
                cursor.execute("SELECT table_name FROM information_schema.tables WHERE table_schema = DATABASE() AND table_name = %s", [db_table])
                if not cursor.fetchone():
                    tables_exist = False
            else:
                # Fallback: try to query the table
                try:
                    cursor.execute(f"SELECT 1 FROM {db_table} LIMIT 1")
                except Exception:
                    tables_exist = False
    except Exception:
        tables_exist = False

    revenue_mtd = safe_sum(
        Revenue.objects.filter(revenue_date__gte=start_of_month, revenue_date__lte=today),
        'amount'
    )
    revenue_ytd = safe_sum(
        Revenue.objects.filter(revenue_date__gte=start_of_year, revenue_date__lte=today),
        'amount'
    )
    expense_mtd = safe_sum(
        Expense.objects.filter(expense_date__gte=start_of_month, expense_date__lte=today, status='paid'),
        'amount'
    )
    expense_ytd = safe_sum(
        Expense.objects.filter(expense_date__gte=start_of_year, expense_date__lte=today, status='paid'),
        'amount'
    )
    net_income_mtd = revenue_mtd - expense_mtd

    receivable_total = safe_sum(
        AdvancedAccountsReceivable.objects.filter(balance_due__gt=0),
        'balance_due'
    )
    receivable_overdue = safe_sum(
        AdvancedAccountsReceivable.objects.filter(is_overdue=True),
        'balance_due'
    )
    payable_total = safe_sum(
        AccountsPayable.objects.filter(balance_due__gt=0),
        'balance_due'
    )
    pending_vouchers = safe_count(PaymentVoucher.objects.filter(status='pending_approval'))
    approved_vouchers = safe_count(PaymentVoucher.objects.filter(status='approved'))

    report_sections = [
        {
            'title': 'Core Financial Statements',
            'description': 'Run the big three statements plus supporting schedules.',
            'reports': [
                {'name': 'Profit & Loss Statement', 'url_name': 'hospital:profit_loss_statement', 'icon': 'bi bi-file-earmark-bar-graph', 'badge': 'MTD/YTD'},
                {'name': 'Balance Sheet', 'url_name': 'hospital:balance_sheet', 'icon': 'bi bi-bank'},
                {'name': 'Cash Flow Statement', 'url_name': 'hospital:cash_flow_statement', 'icon': 'bi bi-cash-stack'},
                {'name': 'Trial Balance', 'url_name': 'hospital:trial_balance', 'icon': 'bi bi-calculator'},
                {'name': 'General Ledger', 'url_name': 'hospital:general_ledger_report', 'icon': 'bi bi-journal-text'},
            ],
        },
        {
            'title': 'Receivables & Payables',
            'description': 'Track aging buckets and settlement pipelines.',
            'reports': [
                {'name': 'Accounts Receivable Aging', 'url_name': 'hospital:ar_aging_report', 'icon': 'bi bi-receipt'},
                {'name': 'Accounts Payable Report', 'url_name': 'hospital:ap_report', 'icon': 'bi bi-journal-check'},
                {'name': 'Payment Voucher Queue', 'url_name': 'hospital:payment_voucher_list', 'icon': 'bi bi-file-earmark-check'},
                {'name': 'Receipt Vouchers', 'url_name': 'hospital:receipt_voucher_list', 'icon': 'bi bi-receipt-cutoff'},
            ],
        },
        {
            'title': 'Revenue Intelligence',
            'description': 'Drill into department revenue trends and KPIs.',
            'reports': [
                {'name': 'Revenue Streams Dashboard', 'url_name': 'hospital:revenue_streams_dashboard', 'icon': 'bi bi-bar-chart-line'},
                {'name': 'Revenue Report', 'url_name': 'hospital:revenue_report', 'icon': 'bi bi-graph-up'},
                {'name': 'Expense Report', 'url_name': 'hospital:expense_report', 'icon': 'bi bi-graph-down'},
                {'name': 'Budget vs Actual', 'url_name': 'hospital:budget_vs_actual_report', 'icon': 'bi bi-target'},
            ],
        },
        {
            'title': 'Operational Controls',
            'description': 'Close the loop with procurement and budgeting.',
            'reports': [
                {'name': 'Procurement Accounts Approval', 'url_name': 'hospital:accounts_approval_list', 'icon': 'bi bi-clipboard-check'},
                {'name': 'Budget Dashboard', 'url_name': 'hospital:budget_dashboard', 'icon': 'bi bi-pie-chart'},
                {'name': 'Payment Processing Center', 'url_name': 'hospital:centralized_cashier_dashboard', 'icon': 'bi bi-columns-gap'},
            ],
        },
    ]

    context = {
        'tables_exist': tables_exist,
        'today': today,
        'start_of_month': start_of_month,
        'start_of_year': start_of_year,
        'kpis': {
            'revenue_mtd': revenue_mtd,
            'revenue_ytd': revenue_ytd,
            'expense_mtd': expense_mtd,
            'expense_ytd': expense_ytd,
            'net_income_mtd': net_income_mtd,
            'receivable_total': receivable_total,
            'receivable_overdue': receivable_overdue,
            'payable_total': payable_total,
            'pending_vouchers': pending_vouchers,
            'approved_vouchers': approved_vouchers,
        },
        'report_sections': report_sections,
    }

    template = 'hospital/accounting_reports_hub.html'
    return render(request, template, context)


@login_required
@user_passes_test(is_accountant)
def profit_loss_statement(request):
    """Profit & Loss Statement (Income Statement)"""
    
    # Get date range
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if not start_date or not end_date:
        # Default to current month
        today = timezone.now().date()
        start_date = today.replace(day=1)
        end_date = today
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Revenue Accounts
    revenue_accounts = Account.objects.filter(account_type='revenue')
    revenue_data = []
    total_revenue = Decimal('0.00')
    
    for account in revenue_accounts:
        balance = AdvancedGeneralLedger.objects.filter(
            account=account,
            transaction_date__gte=start_date,
            transaction_date__lte=end_date,
            is_voided=False
        ).aggregate(
            total=Sum('credit_amount') - Sum('debit_amount')
        )['total'] or Decimal('0.00')
        
        if balance != 0:
            revenue_data.append({
                'account': account,
                'amount': balance
            })
            total_revenue += balance
    
    # Expense Accounts
    expense_accounts = Account.objects.filter(account_type='expense')
    expense_data = []
    total_expenses = Decimal('0.00')
    
    for account in expense_accounts:
        balance = AdvancedGeneralLedger.objects.filter(
            account=account,
            transaction_date__gte=start_date,
            transaction_date__lte=end_date,
            is_voided=False
        ).aggregate(
            total=Sum('debit_amount') - Sum('credit_amount')
        )['total'] or Decimal('0.00')
        
        if balance != 0:
            expense_data.append({
                'account': account,
                'amount': balance
            })
            total_expenses += balance
    
    # Cost of Services (COGS) breakdown
    cost_of_services = {
        'supplies': Decimal('0.00'),
        'salaries': Decimal('0.00'),
        'lab': Decimal('0.00'),
    }
    
    cost_filters = {
        'supplies': ['supply', 'consumable', 'direct medical'],
        'salaries': ['salary', 'staff', 'payroll'],
        'lab': ['lab', 'laboratory'],
    }
    
    try:
        expense_queryset = Expense.objects.filter(
            expense_date__gte=start_date,
            expense_date__lte=end_date,
            status__in=['approved', 'paid'],
            is_deleted=False
        )
        
        for key, keywords in cost_filters.items():
            keyword_filter = Q()
            for term in keywords:
                keyword_filter |= Q(category__name__icontains=term) | Q(description__icontains=term)
            if keyword_filter:
                cost_of_services[key] = (
                    expense_queryset.filter(keyword_filter).aggregate(total=Sum('amount'))['total']
                    or Decimal('0.00')
                )
    except Exception:
        # If expense tables aren't available yet, keep defaults
        pass
    
    total_cost_of_services = sum(cost_of_services.values())
    
    # Gross Profit & Net Income
    gross_profit = total_revenue - total_cost_of_services
    net_income = total_revenue - total_expenses
    
    revenue_by_category = [
        {
            'category__name': item['account'].account_name if isinstance(item['account'], Account) else str(item['account']),
            'total': item['amount'],
        }
        for item in revenue_data
    ]
    
    expense_by_category = [
        {
            'category__name': item['account'].account_name if isinstance(item['account'], Account) else str(item['account']),
            'total': item['amount'],
        }
        for item in expense_data
    ]
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'date_from': start_date,
        'date_to': end_date,
        'today': timezone.now(),
        'revenue_data': revenue_data,
        'expense_data': expense_data,
        'revenue_by_category': revenue_by_category,
        'expense_by_category': expense_by_category,
        'total_revenue': total_revenue,
        'total_expenses': total_expenses,
        'cost_of_services': cost_of_services,
        'total_cost_of_services': total_cost_of_services,
        'gross_profit': gross_profit,
        'net_income': net_income,
        'report_title': f'Profit & Loss Statement - {start_date} to {end_date}',
    }
    
    return render(request, 'hospital/profit_loss_statement.html', context)


@login_required
@user_passes_test(is_accountant)
def balance_sheet(request):
    """Balance Sheet"""
    
    # Get date
    as_of_date = request.GET.get('as_of_date')
    if not as_of_date:
        as_of_date = timezone.now().date()
    else:
        as_of_date = datetime.strptime(as_of_date, '%Y-%m-%d').date()
    
    def get_account_balance(account_code, account_type='asset', as_of_date=None):
        """Get balance for a specific account"""
        try:
            account = Account.objects.get(account_code=account_code, is_deleted=False)
        except Account.DoesNotExist:
            return Decimal('0.00')
        
        ledger_sum = AdvancedGeneralLedger.objects.filter(
            account=account,
            transaction_date__lte=as_of_date,
            is_voided=False
        ).aggregate(
            debits=Sum('debit_amount'),
            credits=Sum('credit_amount')
        )
        
        # Handle None values from aggregate
        debits = Decimal(str(ledger_sum['debits'])) if ledger_sum['debits'] is not None else Decimal('0.00')
        credits = Decimal(str(ledger_sum['credits'])) if ledger_sum['credits'] is not None else Decimal('0.00')
        
        if account_type in ['asset', 'expense']:
            balance = debits - credits
        else:
            balance = credits - debits
        
        # Ensure we return a Decimal
        return Decimal(str(balance)) if balance is not None else Decimal('0.00')
    
    # ASSETS - Build dictionary structure
    assets = {
        'cash': Decimal('0.00'),
        'bank': Decimal('0.00'),
        'accounts_receivable': Decimal('0.00'),
        'inventory': Decimal('0.00'),
        'prepaid': Decimal('0.00'),
        'equipment': Decimal('0.00'),
        'building': Decimal('0.00'),
        'depreciation': Decimal('0.00'),
    }
    
    # Map account codes to asset categories
    # Cash accounts (1000-1099)
    cash_accounts = Account.objects.filter(
        account_type='asset',
        account_code__startswith='10',
        is_deleted=False
    )
    for account in cash_accounts:
        balance = get_account_balance(account.account_code, 'asset', as_of_date)
        # Ensure balance is Decimal
        try:
            balance = Decimal(str(balance)) if balance is not None else Decimal('0.00')
        except (TypeError, ValueError):
            balance = Decimal('0.00')
        
        account_name_lower = (account.account_name or '').lower()
        account_code = account.account_code or ''
        
        if account_code in ['1000', '1010'] or 'cash' in account_name_lower:
            assets['cash'] += balance
        elif account_code.startswith('102') or 'bank' in account_name_lower:
            assets['bank'] += balance
        else:
            # Default to cash if not specifically bank
            assets['cash'] += balance
    
    # Accounts Receivable (1200-1299)
    ar_accounts = Account.objects.filter(
        account_type='asset',
        account_code__startswith='12',
        is_deleted=False
    )
    for account in ar_accounts:
        balance = get_account_balance(account.account_code, 'asset', as_of_date)
        balance = Decimal(str(balance)) if balance is not None else Decimal('0.00')
        assets['accounts_receivable'] += balance
    
    # Inventory (1300-1399)
    inventory_accounts = Account.objects.filter(
        account_type='asset',
        account_code__startswith='13',
        is_deleted=False
    )
    for account in inventory_accounts:
        balance = get_account_balance(account.account_code, 'asset', as_of_date)
        balance = Decimal(str(balance)) if balance is not None else Decimal('0.00')
        assets['inventory'] += balance
    
    # Prepaid (1400-1499)
    prepaid_accounts = Account.objects.filter(
        account_type='asset',
        account_code__startswith='14',
        is_deleted=False
    )
    for account in prepaid_accounts:
        balance = get_account_balance(account.account_code, 'asset', as_of_date)
        balance = Decimal(str(balance)) if balance is not None else Decimal('0.00')
        assets['prepaid'] += balance
    
    # Fixed Assets (1500-1599) - Enhanced categorization
    fixed_asset_accounts = Account.objects.filter(
        account_type='asset',
        account_code__startswith='15',
        is_deleted=False
    )
    
    # Separate fixed assets by category
    land_buildings = Decimal('0.00')
    equipment_total = Decimal('0.00')
    vehicles_total = Decimal('0.00')
    accumulated_depreciation = Decimal('0.00')
    construction = Decimal('0.00')
    intangible = Decimal('0.00')
    
    for account in fixed_asset_accounts:
        balance = get_account_balance(account.account_code, 'asset', as_of_date)
        balance = Decimal(str(balance)) if balance is not None else Decimal('0.00')
        account_name_lower = account.account_name.lower() if account.account_name else ''
        account_code = int(account.account_code) if account.account_code.isdigit() else 0
        
        # Categorize by account code range
        if 1500 <= account_code <= 1509:
            # Land and Buildings
            if 'depreciation' not in account_name_lower:
                land_buildings += balance
            else:
                accumulated_depreciation += balance
        elif 1510 <= account_code <= 1529:
            # Equipment
            if 'depreciation' not in account_name_lower:
                equipment_total += balance
            else:
                accumulated_depreciation += balance
        elif 1530 <= account_code <= 1539:
            # Vehicles
            if 'depreciation' not in account_name_lower:
                vehicles_total += balance
            else:
                accumulated_depreciation += balance
        elif 1540 <= account_code <= 1559:
            # Accumulated Depreciation (contra-asset, reduces asset value)
            accumulated_depreciation += balance
        elif 1560 <= account_code <= 1569:
            # Construction in Progress
            construction += balance
        elif 1570 <= account_code <= 1579:
            # Intangible Assets
            intangible += balance
        else:
            # Fallback: check by name
            if 'depreciation' not in account_name_lower:
                equipment_total += balance
            else:
                accumulated_depreciation += balance
    
    # Add to equipment and building totals
    assets['equipment'] += equipment_total
    assets['building'] += land_buildings
    assets['depreciation'] += accumulated_depreciation
    
    # Store detailed fixed asset breakdown for template
    fixed_assets_detail = {
        'land_buildings': land_buildings,
        'equipment': equipment_total,
        'vehicles': vehicles_total,
        'accumulated_depreciation': accumulated_depreciation,
        'construction': construction,
        'intangible': intangible,
        'net_fixed_assets': land_buildings + equipment_total + vehicles_total + construction + intangible - accumulated_depreciation
    }
    
    # Building (1600-1699)
    building_accounts = Account.objects.filter(
        account_type='asset',
        account_code__startswith='16',
        is_deleted=False
    )
    for account in building_accounts:
        balance = get_account_balance(account.account_code, 'asset', as_of_date)
        balance = Decimal(str(balance)) if balance is not None else Decimal('0.00')
        account_name_lower = account.account_name.lower() if account.account_name else ''
        if 'depreciation' not in account_name_lower:
            assets['building'] += balance
        else:
            assets['depreciation'] += balance
    
    # Initialize fixed_assets_detail if not already set
    if 'fixed_assets_detail' not in locals():
        fixed_assets_detail = {
            'land_buildings': Decimal('0.00'),
            'equipment': Decimal('0.00'),
            'vehicles': Decimal('0.00'),
            'accumulated_depreciation': Decimal('0.00'),
            'construction': Decimal('0.00'),
            'intangible': Decimal('0.00'),
            'net_fixed_assets': Decimal('0.00')
        }
    
    # Calculate total assets
    total_assets = (
        assets['cash'] + assets['bank'] + assets['accounts_receivable'] +
        assets['inventory'] + assets['prepaid'] + assets['equipment'] +
        assets['building'] - assets['depreciation']
    )
    
    # LIABILITIES - Build dictionary structure
    liabilities = {
        'accounts_payable': Decimal('0.00'),
        'accrued': Decimal('0.00'),
        'short_term_loans': Decimal('0.00'),
        'long_term_loans': Decimal('0.00'),
        'mortgages': Decimal('0.00'),
    }
    
    # Accounts Payable (2000-2099)
    # For Excel imports: debit amounts ARE the balances (independent, different companies)
    # Sum all debit amounts directly, don't calculate running balance
    ap_accounts = Account.objects.filter(
        account_type='liability',
        account_name__icontains='payable',
        is_deleted=False
    )
    for account in ap_accounts:
        # For Excel imports: sum debit amounts (each is an independent balance)
        ap_gl_entries = AdvancedGeneralLedger.objects.filter(
            account=account,
            transaction_date__lte=as_of_date,
            is_voided=False,
            is_deleted=False
        )
        # Sum debit amounts (each debit IS the balance for that entry)
        ap_debit_total = ap_gl_entries.aggregate(total=Sum('debit_amount'))['total'] or Decimal('0.00')
        if ap_debit_total > 0:
            # Use debit amounts as balances (Excel import format)
            liabilities['accounts_payable'] += ap_debit_total
        else:
            # Fallback to normal calculation if no GL entries
            balance = get_account_balance(account.account_code, 'liability', as_of_date)
            balance = Decimal(str(balance)) if balance is not None else Decimal('0.00')
            liabilities['accounts_payable'] += balance
    
    # If no AP from GL, fall back to AccountsPayable model
    if liabilities['accounts_payable'] == 0:
        try:
            ap_model_total = AccountsPayable.objects.filter(
                balance_due__gt=0,
                is_deleted=False
            ).aggregate(total=Sum('balance_due'))['total'] or Decimal('0.00')
            liabilities['accounts_payable'] = ap_model_total
        except:
            pass
    
    # Accrued Expenses (2100-2199)
    accrued_accounts = Account.objects.filter(
        account_type='liability',
        account_code__startswith='21',
        is_deleted=False
    )
    for account in accrued_accounts:
        balance = get_account_balance(account.account_code, 'liability', as_of_date)
        balance = Decimal(str(balance)) if balance is not None else Decimal('0.00')
        liabilities['accrued'] += balance
    
    # Short-term Loans (2200-2299)
    short_term_accounts = Account.objects.filter(
        account_type='liability',
        account_code__startswith='22',
        is_deleted=False
    )
    for account in short_term_accounts:
        balance = get_account_balance(account.account_code, 'liability', as_of_date)
        balance = Decimal(str(balance)) if balance is not None else Decimal('0.00')
        liabilities['short_term_loans'] += balance
    
    # Long-term Loans (2300-2399)
    long_term_accounts = Account.objects.filter(
        account_type='liability',
        account_code__startswith='23',
        is_deleted=False
    )
    for account in long_term_accounts:
        balance = get_account_balance(account.account_code, 'liability', as_of_date)
        balance = Decimal(str(balance)) if balance is not None else Decimal('0.00')
        account_name_lower = account.account_name.lower() if account.account_name else ''
        if 'mortgage' in account_name_lower:
            liabilities['mortgages'] += balance
        else:
            liabilities['long_term_loans'] += balance
    
    # Calculate total liabilities
    total_liabilities = (
        liabilities['accounts_payable'] + liabilities['accrued'] +
        liabilities['short_term_loans'] + liabilities['long_term_loans'] +
        liabilities['mortgages']
    )
    
    # EQUITY
    equity_accounts = Account.objects.filter(account_type='equity', is_deleted=False)
    total_equity = Decimal('0.00')
    for account in equity_accounts:
        balance = get_account_balance(account.account_code, 'equity', as_of_date)
        balance = Decimal(str(balance)) if balance is not None else Decimal('0.00')
        total_equity += balance
    
    # Calculate Net Income (Revenue - Expenses) for current year
    revenue_accounts = Account.objects.filter(account_type='revenue', is_deleted=False)
    total_revenue = Decimal('0.00')
    for account in revenue_accounts:
        balance = get_account_balance(account.account_code, 'revenue', as_of_date)
        balance = Decimal(str(balance)) if balance is not None else Decimal('0.00')
        total_revenue += balance
    
    expense_accounts = Account.objects.filter(account_type='expense', is_deleted=False)
    total_expenses = Decimal('0.00')
    for account in expense_accounts:
        balance = get_account_balance(account.account_code, 'expense', as_of_date)
        balance = Decimal(str(balance)) if balance is not None else Decimal('0.00')
        total_expenses += balance
    
    # Net Income = Revenue - Expenses
    net_income = total_revenue - total_expenses
    
    # Total Equity includes Net Income
    total_equity_with_income = total_equity + net_income
    
    # Ensure all values are properly formatted as Decimal for template
    # Convert all asset values to float for template compatibility
    assets_formatted = {
        'cash': float(assets['cash']),
        'bank': float(assets['bank']),
        'accounts_receivable': float(assets['accounts_receivable']),
        'inventory': float(assets['inventory']),
        'prepaid': float(assets['prepaid']),
        'equipment': float(assets['equipment']),
        'building': float(assets['building']),
        'depreciation': float(assets['depreciation']),
    }
    
    # Convert all liability values to float for template compatibility
    liabilities_formatted = {
        'accounts_payable': float(liabilities['accounts_payable']),
        'accrued': float(liabilities['accrued']),
        'short_term_loans': float(liabilities['short_term_loans']),
        'long_term_loans': float(liabilities['long_term_loans']),
        'mortgages': float(liabilities['mortgages']),
    }
    
    context = {
        'as_of_date': as_of_date,
        'today': timezone.now(),
        'assets': assets_formatted,
        'liabilities': liabilities_formatted,
        'total_assets': float(total_assets),
        'total_liabilities': float(total_liabilities),
        'fixed_assets_detail': {
            'land_buildings': float(fixed_assets_detail.get('land_buildings', 0)),
            'equipment': float(fixed_assets_detail.get('equipment', 0)),
            'vehicles': float(fixed_assets_detail.get('vehicles', 0)),
            'accumulated_depreciation': float(fixed_assets_detail.get('accumulated_depreciation', 0)),
            'construction': float(fixed_assets_detail.get('construction', 0)),
            'intangible': float(fixed_assets_detail.get('intangible', 0)),
            'net_fixed_assets': float(fixed_assets_detail.get('net_fixed_assets', 0))
        },
        'total_equity': float(total_equity),
        'total_revenue': float(total_revenue),
        'total_expenses': float(total_expenses),
        'net_income': float(net_income),
        'total_equity_with_income': float(total_equity_with_income),
        'total_liab_equity': float(total_liabilities + total_equity_with_income),
        'is_balanced': abs(total_assets - (total_liabilities + total_equity_with_income)) < Decimal('0.01'),
        'report_title': f'Balance Sheet as of {as_of_date}',
    }
    
    return render(request, 'hospital/balance_sheet.html', context)


@login_required
@user_passes_test(is_accountant)
def trial_balance(request):
    """Trial Balance Report"""
    
    # Get date
    as_of_date = request.GET.get('as_of_date')
    if not as_of_date:
        as_of_date = timezone.now().date()
    else:
        as_of_date = datetime.strptime(as_of_date, '%Y-%m-%d').date()
    
    # Get all accounts with balances
    accounts = Account.objects.filter(is_active=True).order_by('account_code')
    
    trial_balance_data = []
    accounts_with_balances = []
    total_debits = Decimal('0.00')
    total_credits = Decimal('0.00')
    
    for account in accounts:
        # Calculate balance
        ledger_sum = AdvancedGeneralLedger.objects.filter(
            account=account,
            transaction_date__lte=as_of_date,
            is_voided=False
        ).aggregate(
            debits=Sum('debit_amount'),
            credits=Sum('credit_amount')
        )
        
        debits = ledger_sum['debits'] or Decimal('0.00')
        credits = ledger_sum['credits'] or Decimal('0.00')
        balance = debits - credits
        
        if balance != 0:
            setattr(account, 'balance', balance)
            accounts_with_balances.append(account)
        
        if balance != 0:
            if balance > 0:
                # Debit balance
                trial_balance_data.append({
                    'account': account,
                    'debit': balance,
                    'credit': Decimal('0.00')
                })
                total_debits += balance
            else:
                # Credit balance
                trial_balance_data.append({
                    'account': account,
                    'debit': Decimal('0.00'),
                    'credit': abs(balance)
                })
                total_credits += abs(balance)
    
    # Check if balanced
    is_balanced = abs(total_debits - total_credits) < 0.01
    balance_difference = total_debits - total_credits
    
    context = {
        'as_of_date': as_of_date,
        'accounts': accounts_with_balances,
        'trial_balance_data': trial_balance_data,
        'total_debits': total_debits,
        'total_credits': total_credits,
        'is_balanced': is_balanced,
        'balance_difference': balance_difference,
        'today': timezone.now(),
        'report_title': f'Trial Balance as of {as_of_date}',
    }
    
    return render(request, 'hospital/trial_balance.html', context)


@login_required
@user_passes_test(is_accountant)
def cash_flow_statement(request):
    """Cash Flow Statement"""
    
    # Get date range
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if not start_date or not end_date:
        # Default to current month
        today = timezone.now().date()
        start_date = today.replace(day=1)
        end_date = today
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Operating Activities
    # Use PaymentReceipt for actual cash received (more reliable than Revenue table)
    from .models_accounting import PaymentReceipt
    
    # Convert dates to timezone-aware datetime for comparison
    from django.utils import timezone as tz
    start_datetime = tz.make_aware(datetime.combine(start_date, datetime.min.time()))
    end_datetime = tz.make_aware(datetime.combine(end_date, datetime.max.time()))
    
    cash_from_operations = PaymentReceipt.objects.filter(
        receipt_date__gte=start_datetime,
        receipt_date__lte=end_datetime,
        is_deleted=False
    ).aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')
    
    # Also try Revenue table if PaymentReceipt is empty
    if cash_from_operations == Decimal('0.00'):
        cash_from_operations = Revenue.objects.filter(
            revenue_date__gte=start_date,
            revenue_date__lte=end_date,
            payment_method__in=['cash', 'mobile_money', 'bank_transfer']
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    cash_for_expenses = Expense.objects.filter(
        expense_date__gte=start_date,
        expense_date__lte=end_date,
        status__in=['approved', 'paid']
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    net_operating_cash = cash_from_operations - cash_for_expenses
    
    # Investing Activities (placeholder - can be expanded)
    investing_cash_flow = Decimal('0.00')
    
    # Financing Activities (placeholder - can be expanded)
    financing_cash_flow = Decimal('0.00')
    
    # Net Change
    net_cash_change = net_operating_cash + investing_cash_flow + financing_cash_flow
    
    # Opening and Closing Cash
    cash_accounts = BankAccount.objects.filter(is_active=True, account_type__in=['checking', 'savings'])
    opening_cash = sum(acc.opening_balance for acc in cash_accounts)
    closing_cash = sum(acc.current_balance for acc in cash_accounts)
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        # FIXED: Match template variable names
        'operating_inflows': cash_from_operations,
        'operating_outflows': cash_for_expenses,
        'net_operating': net_operating_cash,
        'investing_outflows': abs(investing_cash_flow),
        'financing_inflows': financing_cash_flow,
        'beginning_cash': opening_cash,
        'net_change': net_cash_change,
        'ending_cash': closing_cash,
        'report_title': f'Cash Flow Statement - {start_date} to {end_date}',
    }
    
    return render(request, 'hospital/cash_flow_statement.html', context)


@login_required
@user_passes_test(is_accountant)
def general_ledger_report(request):
    """General Ledger Report by Account"""
    
    account_id = request.GET.get('account_id')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if not start_date or not end_date:
        today = timezone.now().date()
        start_date = today.replace(day=1)
        end_date = today
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Get ledger entries
    ledger_entries = AdvancedGeneralLedger.objects.filter(
        transaction_date__gte=start_date,
        transaction_date__lte=end_date,
        is_voided=False,
        is_deleted=False
    ).select_related('account', 'journal_entry').order_by('account', 'transaction_date')
    
    if account_id:
        ledger_entries = ledger_entries.filter(account_id=account_id)
        selected_account = Account.objects.filter(id=account_id).first()
    else:
        selected_account = None
    
    # Group by account
    accounts_data = {}
    for entry in ledger_entries:
        account_code = entry.account.account_code
        if account_code not in accounts_data:
            accounts_data[account_code] = {
                'account': entry.account,
                'entries': [],
                'total_debit': Decimal('0.00'),
                'total_credit': Decimal('0.00'),
                'balance': Decimal('0.00')
            }
        
        accounts_data[account_code]['entries'].append(entry)
        accounts_data[account_code]['total_debit'] += entry.debit_amount
        accounts_data[account_code]['total_credit'] += entry.credit_amount
        accounts_data[account_code]['balance'] = accounts_data[account_code]['total_debit'] - accounts_data[account_code]['total_credit']
    
    # For Excel imports: debit/credit amounts ARE the balances (not transaction amounts)
    # Each entry represents a different company's balance - they are INDEPENDENT
    # Always use debit/credit amount directly as balance (don't use stored balance which may be cumulative)
    ledger_with_balance = []
    
    for entry in ledger_entries:
        # For Excel entries: debit/credit amounts ARE the balances
        # Each entry is independent (different companies) - don't calculate running balance
        if entry.debit_amount and entry.debit_amount > 0:
            # Debit amount IS the balance (for asset accounts like AR, or AP balances)
            entry.running_balance = entry.debit_amount
        elif entry.credit_amount and entry.credit_amount > 0:
            # Credit amount IS the balance (for liability accounts)
            entry.running_balance = entry.credit_amount
        elif entry.balance and entry.balance != 0:
            # Fallback to stored balance if no debit/credit
            entry.running_balance = entry.balance
        else:
            entry.running_balance = Decimal('0.00')
        
        ledger_with_balance.append(entry)
    
    # Calculate totals
    total_debits = ledger_entries.aggregate(total=Sum('debit_amount'))['total'] or Decimal('0.00')
    total_credits = ledger_entries.aggregate(total=Sum('credit_amount'))['total'] or Decimal('0.00')
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'gl_entries': ledger_with_balance,  # FIXED: Template expects this variable name
        'accounts_data': accounts_data,
        'selected_account': selected_account,
        'all_accounts': Account.objects.filter(is_active=True).order_by('account_code'),
        'total_debits': total_debits,  # Added for summary
        'total_credits': total_credits,  # Added for summary
        'entry_count': ledger_entries.count(),  # Added for display
        'report_title': 'General Ledger Report',
    }
    
    return render(request, 'hospital/general_ledger_report.html', context)


@login_required
@user_passes_test(is_accountant)
def accounts_receivable_aging(request):
    """Accounts Receivable Aging Report - Includes both AdvancedAccountsReceivable and InsuranceReceivableEntry"""
    from .models_primecare_accounting import InsuranceReceivableEntry
    
    # Get all AR with balances
    receivables = AdvancedAccountsReceivable.objects.filter(
        balance_due__gt=0
    ).select_related('invoice', 'patient').order_by('due_date')
    
    # Also include InsuranceReceivableEntry records - EXCLUDE CORPORATE PAYERS
    insurance_receivables = InsuranceReceivableEntry.objects.filter(
        outstanding_amount__gt=0,
        is_deleted=False
    ).exclude(payer__payer_type='corporate').select_related('payer').order_by('entry_date')  # Only insurance, not corporate
    
    # Group by aging bucket
    aging_summary = {
        'current': {'items': [], 'total': Decimal('0.00')},
        '0-30': {'items': [], 'total': Decimal('0.00')},
        '31-60': {'items': [], 'total': Decimal('0.00')},
        '61-90': {'items': [], 'total': Decimal('0.00')},
        '90+': {'items': [], 'total': Decimal('0.00')},
    }
    
    # Process AdvancedAccountsReceivable
    for ar in receivables:
        bucket = ar.aging_bucket
        if bucket in aging_summary:
            aging_summary[bucket]['items'].append(ar)
            aging_summary[bucket]['total'] += ar.balance_due
    
    # Process InsuranceReceivableEntry - calculate aging
    from datetime import timedelta
    today = timezone.now().date()
    
    for entry in insurance_receivables:
        days_old = (today - entry.entry_date).days
        
        if days_old <= 0:
            bucket = 'current'
        elif days_old <= 30:
            bucket = '0-30'
        elif days_old <= 60:
            bucket = '31-60'
        elif days_old <= 90:
            bucket = '61-90'
        else:
            bucket = '90+'
        
        if bucket in aging_summary:
            # Create a wrapper object for template compatibility
            class InsuranceARWrapper:
                def __init__(self, entry):
                    self.id = entry.id
                    self.invoice = None
                    self.patient = None
                    self.insurance_company = entry.payer
                    self.balance_due = entry.outstanding_amount
                    self.total_amount = entry.total_amount
                    self.due_date = entry.entry_date
                    self.entry_number = entry.entry_number
                    self.is_insurance_entry = True
            
            aging_summary[bucket]['items'].append(InsuranceARWrapper(entry))
            aging_summary[bucket]['total'] += entry.outstanding_amount
    
    grand_total = sum(bucket['total'] for bucket in aging_summary.values())
    
    # Create ar_aging dict for template (expects specific variable names)
    ar_aging_dict = {
        'current': aging_summary['current']['total'],
        '0_30': aging_summary['0-30']['total'],
        '31_60': aging_summary['31-60']['total'],
        '61_90': aging_summary['61-90']['total'],
        '90_plus': aging_summary['90+']['total'],
    }
    
    # DEBUG - Print values being passed
    print("\n" + "="*70)
    print("AR AGING REPORT - VALUES BEING PASSED TO TEMPLATE")
    print("="*70)
    print(f"ar_aging dict: {ar_aging_dict}")
    print(f"total_ar: {grand_total}")
    print(f"Type of current: {type(ar_aging_dict['current'])}")
    print("="*70 + "\n")
    
    context = {
        'aging_summary': aging_summary,
        'ar_aging': ar_aging_dict,  # FIXED: Template expects this variable name
        'ar_list': receivables,  # FIXED: Template expects this for the detailed list
        'grand_total': grand_total,
        'total_ar': grand_total,  # FIXED: Template expects this too
        'report_title': 'Accounts Receivable Aging Report',
    }
    
    return render(request, 'hospital/ar_aging_report.html', context)


@login_required
@user_passes_test(is_accountant)
def corporate_receivables(request):
    """Corporate Receivables Report - Shows only receivables from corporate payers"""
    from .models_primecare_accounting import InsuranceReceivableEntry
    from .models import Payer
    from datetime import timedelta
    
    # Get filter parameters
    payer_id = request.GET.get('payer', '')
    aging_bucket = request.GET.get('aging', '')
    
    # Get corporate AR with balances (filter by invoice payer type)
    receivables = AdvancedAccountsReceivable.objects.filter(
        balance_due__gt=0,
        invoice__payer__payer_type='corporate'
    ).select_related('invoice', 'invoice__payer', 'patient').order_by('due_date')
    
    # Also get InsuranceReceivableEntry records for corporate payers
    # Handle case where table doesn't exist (migration not run)
    try:
        insurance_receivables = InsuranceReceivableEntry.objects.filter(
            outstanding_amount__gt=0,
            is_deleted=False,
            payer__payer_type='corporate'
        ).select_related('payer').order_by('entry_date')
    except Exception as e:
        # Table doesn't exist or other database error - use empty queryset
        from django.db import connection
        from django.core.exceptions import ImproperlyConfigured
        insurance_receivables = InsuranceReceivableEntry.objects.none()
    
    # Filter by specific payer if provided
    if payer_id:
        receivables = receivables.filter(invoice__payer_id=payer_id)
        insurance_receivables = insurance_receivables.filter(payer_id=payer_id)
    
    # Group by aging bucket
    aging_summary = {
        'current': {'items': [], 'total': Decimal('0.00')},
        '0-30': {'items': [], 'total': Decimal('0.00')},
        '31-60': {'items': [], 'total': Decimal('0.00')},
        '61-90': {'items': [], 'total': Decimal('0.00')},
        '90+': {'items': [], 'total': Decimal('0.00')},
    }
    
    # Process AdvancedAccountsReceivable
    for ar in receivables:
        bucket = ar.aging_bucket
        if bucket in aging_summary:
            aging_summary[bucket]['items'].append(ar)
            aging_summary[bucket]['total'] += ar.balance_due
    
    # Process InsuranceReceivableEntry - calculate aging
    today = timezone.now().date()
    
    # Only process if table exists and has data - wrap in try-except to handle missing table
    try:
        if insurance_receivables.exists():
            for entry in insurance_receivables:
                days_old = (today - entry.entry_date).days
                
                if days_old <= 0:
                    bucket = 'current'
                elif days_old <= 30:
                    bucket = '0-30'
                elif days_old <= 60:
                    bucket = '31-60'
                elif days_old <= 90:
                    bucket = '61-90'
                else:
                    bucket = '90+'
                
                # Apply aging bucket filter if provided
                if aging_bucket and bucket != aging_bucket:
                    continue
                
                if bucket in aging_summary:
                    # Create a wrapper object for template compatibility
                    class CorporateARWrapper:
                        def __init__(self, entry, days_old, bucket):
                            self.id = entry.id
                            # Create a fake invoice object for template compatibility
                            fake_invoice = SimpleNamespace()
                            fake_invoice.invoice_number = entry.entry_number
                            fake_invoice.created = entry.entry_date
                            self.invoice = fake_invoice
                            
                            self.patient = None  # No patient for corporate receivables
                            self.payer = entry.payer  # Store payer for company name
                            self.insurance_company = entry.payer  # For compatibility
                            self.balance_due = entry.outstanding_amount
                            self.invoice_amount = entry.total_amount
                            self.amount_paid = entry.amount_received
                            self.due_date = entry.entry_date
                            self.entry_number = entry.entry_number
                            self.is_insurance_entry = True
                            self.aging_bucket = bucket
                            self.days_overdue = days_old if days_old > 0 else 0
                            self.is_overdue = days_old > 0
                    
                    aging_summary[bucket]['items'].append(CorporateARWrapper(entry, days_old, bucket))
                    aging_summary[bucket]['total'] += entry.outstanding_amount
    except Exception:
        # Table doesn't exist or error accessing it - skip processing
        pass
    
    # Apply aging bucket filter to AdvancedAccountsReceivable if provided
    if aging_bucket:
        filtered_receivables = []
        for ar in receivables:
            if ar.aging_bucket == aging_bucket:
                filtered_receivables.append(ar)
        receivables = filtered_receivables
    
    grand_total = sum(bucket['total'] for bucket in aging_summary.values())
    
    # Get all corporate payers for filter dropdown
    corporate_payers = Payer.objects.filter(
        payer_type='corporate',
        is_active=True,
        is_deleted=False
    ).order_by('name')
    
    # Create ar_aging dict for template
    ar_aging_dict = {
        'current': aging_summary['current']['total'],
        '0_30': aging_summary['0-30']['total'],
        '31_60': aging_summary['31-60']['total'],
        '61_90': aging_summary['61-90']['total'],
        '90_plus': aging_summary['90+']['total'],
    }
    
    # Combine all receivables for the list view
    all_receivables_list = list(receivables)
    for bucket_data in aging_summary.values():
        all_receivables_list.extend(bucket_data['items'])
    
    context = {
        'aging_summary': aging_summary,
        'ar_aging': ar_aging_dict,
        'ar_list': all_receivables_list,
        'ar_count': len(all_receivables_list),  # Add count for template
        'grand_total': grand_total,
        'total_ar': grand_total,
        'report_title': 'Corporate Receivables Aging Report',
        'corporate_payers': corporate_payers,
        'selected_payer_id': payer_id,
        'selected_aging': aging_bucket,
        'is_corporate': True,  # Flag to indicate this is corporate receivables
    }
    
    return render(request, 'hospital/ar_aging_report.html', context)


@login_required
def ar_aging_test(request):
    """Test view for AR aging context"""
    # Create ar_aging dict for template (expects specific variable names)
    ar_aging_dict = {
        'current': Decimal('0.00'),
        '0_30': Decimal('0.00'),
        '31_60': Decimal('0.00'),
        '61_90': Decimal('0.00'),
        '90_plus': Decimal('0.00'),
    }
    
    context = {
        'ar_aging': ar_aging_dict,
        'total_ar': Decimal('0.00'),
    }
    
    return render(request, 'hospital/ar_test.html', context)


@login_required
@user_passes_test(is_accountant)
def accounts_payable_report(request):
    """Accounts Payable Report"""
    
    payables = AccountsPayable.objects.filter(
        balance_due__gt=0
    ).order_by('due_date')
    
    # Separate overdue and current
    overdue = payables.filter(is_overdue=True)
    current = payables.filter(is_overdue=False)
    
    overdue_total = overdue.aggregate(total=Sum('balance_due'))['total'] or Decimal('0.00')
    current_total = current.aggregate(total=Sum('balance_due'))['total'] or Decimal('0.00')
    grand_total = overdue_total + current_total
    
    context = {
        'overdue': overdue,
        'current': current,
        'overdue_total': overdue_total,
        'current_total': current_total,
        'grand_total': grand_total,
        'report_title': 'Accounts Payable Report',
    }
    
    return render(request, 'hospital/accounts_payable_report.html', context)


@login_required
@user_passes_test(is_accountant)
def record_ap_payment(request, ap_id):
    """Record payment for a specific Accounts Payable entry"""
    from django.contrib import messages
    from django.db import transaction as db_transaction
    from .models_accounting_advanced import PaymentVoucher, Journal, AdvancedJournalEntry, AdvancedJournalEntryLine
    from .models_accounting import Account
    
    ap = get_object_or_404(AccountsPayable, pk=ap_id, is_deleted=False)
    
    if request.method == 'POST':
        payment_amount = Decimal(request.POST.get('payment_amount', '0'))
        payment_date = request.POST.get('payment_date')
        payment_method = request.POST.get('payment_method', 'bank_transfer')
        bank_account_id = request.POST.get('bank_account')
        payment_reference = request.POST.get('payment_reference', '')
        notes = request.POST.get('notes', '')
        
        if payment_amount <= 0:
            messages.error(request, 'Payment amount must be greater than zero.')
            return redirect('hospital:record_ap_payment', ap_id=ap.id)
        
        if payment_amount > ap.balance_due:
            messages.error(request, f'Payment amount (GHS {payment_amount:,.2f}) cannot exceed balance due (GHS {ap.balance_due:,.2f}).')
            return redirect('hospital:record_ap_payment', ap_id=ap.id)
        
        try:
            with db_transaction.atomic():
                # Update AP balance - refresh from DB first to ensure we have latest values
                ap.refresh_from_db()
                ap.amount_paid += payment_amount
                # balance_due will be auto-calculated in save() method
                # Force recalculation to ensure accuracy
                ap.balance_due = ap.amount - ap.amount_paid
                ap.save()
                
                # Get or create payment account based on payment method
                if bank_account_id:
                    from .models_accounting_advanced import BankAccount
                    try:
                        bank_account = BankAccount.objects.get(pk=bank_account_id, is_deleted=False)
                        payment_account = bank_account.gl_account
                    except BankAccount.DoesNotExist:
                        # Fallback to method-based account
                        payment_account = None
                else:
                    payment_account = None
                
                # If no bank account selected, use payment method to determine account
                if not payment_account:
                    account_map = {
                        'cash': ('1010', 'Cash on Hand', 'asset'),
                        'bank_transfer': ('1020', 'Bank Account', 'asset'),
                        'cheque': ('1020', 'Bank Account', 'asset'),
                        'mobile_money': ('1030', 'Mobile Money', 'asset'),
                    }
                    code, name, acc_type = account_map.get(payment_method, ('1010', 'Cash on Hand', 'asset'))
                    payment_account, _ = Account.objects.get_or_create(
                        account_code=code,
                        defaults={'account_name': name, 'account_type': acc_type}
                    )
                
                # Get AP account
                ap_account, _ = Account.objects.get_or_create(
                    account_code='2100',
                    defaults={'account_name': 'Accounts Payable', 'account_type': 'liability'}
                )
                
                # Create journal entry
                payment_journal, _ = Journal.objects.get_or_create(
                    journal_type='payment',
                    defaults={'name': 'Payment Journal', 'code': 'PJ'}
                )
                
                je = AdvancedJournalEntry.objects.create(
                    journal=payment_journal,
                    entry_date=datetime.strptime(payment_date, '%Y-%m-%d').date() if payment_date else timezone.now().date(),
                    description=f"Payment to {ap.vendor_name} - {ap.bill_number}",
                    reference=payment_reference or ap.bill_number,
                    created_by=request.user,
                    status='posted'
                )
                
                # Debit Accounts Payable (decreases liability)
                AdvancedJournalEntryLine.objects.create(
                    journal_entry=je,
                    line_number=1,
                    account=ap_account,
                    debit_amount=payment_amount,
                    credit_amount=Decimal('0.00'),
                    description=f"Payment to {ap.vendor_name}"
                )
                
                # Credit Payment Account (decreases asset - cash/bank)
                AdvancedJournalEntryLine.objects.create(
                    journal_entry=je,
                    line_number=2,
                    account=payment_account,
                    debit_amount=Decimal('0.00'),
                    credit_amount=payment_amount,
                    description=f"Payment for {ap.bill_number}"
                )
                
                # Update journal entry totals
                je.total_debit = payment_amount
                je.total_credit = payment_amount
                je.save()
                
                # Post journal entry to GL
                je.post(request.user)
                
                messages.success(
                    request,
                    f'Payment of GHS {payment_amount:,.2f} recorded successfully. '
                    f'Balance due: GHS {ap.balance_due:,.2f}'
                )
                return redirect('hospital:ap_report')
        
        except Exception as e:
            messages.error(request, f'Error recording payment: {str(e)}')
            import traceback
            traceback.print_exc()
    
    # GET request - show payment form
    from .models_accounting_advanced import BankAccount
    bank_accounts = BankAccount.objects.filter(is_active=True, is_deleted=False)
    
    context = {
        'ap': ap,
        'bank_accounts': bank_accounts,
    }
    
    return render(request, 'hospital/record_ap_payment.html', context)


@login_required
@user_passes_test(is_accountant)
def budget_variance_report(request):
    """Budget vs Actual Variance Report"""
    
    budget_id = request.GET.get('budget_id')
    
    if budget_id:
        budget = get_object_or_404(Budget, id=budget_id)
    else:
        # Get current budget
        budget = Budget.objects.filter(
            start_date__lte=timezone.now().date(),
            end_date__gte=timezone.now().date()
        ).first()
    
    if not budget:
        context = {
            'error': 'No budget found for current period',
            'all_budgets': Budget.objects.all().order_by('-start_date')
        }
        return render(request, 'hospital/budget_variance_report.html', context)
    
    # Update actual amounts
    for line in budget.lines.all():
        actual = AdvancedGeneralLedger.objects.filter(
            account=line.account,
            transaction_date__gte=budget.start_date,
            transaction_date__lte=budget.end_date,
            is_voided=False
        ).aggregate(
            total=Sum('debit_amount') - Sum('credit_amount')
        )['total'] or Decimal('0.00')
        
        line.actual_amount = abs(actual)
        line.calculate_variance()
    
    # Get lines
    budget_lines = budget.lines.all().select_related('account', 'cost_center').order_by('account__account_code')
    
    # Calculate totals
    total_budgeted = budget_lines.aggregate(total=Sum('budgeted_amount'))['total'] or Decimal('0.00')
    total_actual = budget_lines.aggregate(total=Sum('actual_amount'))['total'] or Decimal('0.00')
    total_variance = total_actual - total_budgeted
    
    context = {
        'budget': budget,
        'budget_lines': budget_lines,
        'total_budgeted': total_budgeted,
        'total_actual': total_actual,
        'total_variance': total_variance,
        'all_budgets': Budget.objects.all().order_by('-start_date'),
        'report_title': f'Budget Variance Report - {budget.name}',
    }
    
    return render(request, 'hospital/budget_variance_report.html', context)


@login_required
@user_passes_test(is_accountant)
def revenue_report(request):
    """Detailed Revenue Report"""
    
    # Accept both legacy (start_date/end_date) and new (date_from/date_to) params
    start_date_param = request.GET.get('date_from') or request.GET.get('start_date')
    end_date_param = request.GET.get('date_to') or request.GET.get('end_date')
    
    category_id = request.GET.get('category') or request.GET.get('category_id')
    service_type = request.GET.get('service_type') or request.GET.get('service')
    
    if not start_date_param or not end_date_param:
        today = timezone.now().date()
        start_date = today.replace(day=1)
        end_date = today
    else:
        try:
            start_date = datetime.strptime(start_date_param, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_param, '%Y-%m-%d').date()
        except ValueError:
            today = timezone.now().date()
            start_date = today.replace(day=1)
            end_date = today
    
    # Get revenues
    revenues = Revenue.objects.filter(
        revenue_date__gte=start_date,
        revenue_date__lte=end_date
    ).select_related('category', 'patient').order_by('-revenue_date')
    
    if category_id:
        revenues = revenues.filter(category_id=category_id)
    
    if service_type:
        service_type_normalized = service_type.lower()
        
        service_filter = Q(service_type__iexact=service_type_normalized)
        service_type_keywords = {
            'pharmacy': ['pharmacy', 'drug', 'dispensary', 'medication', 'medicine', 'rx'],
            'laboratory': ['laboratory', 'lab', 'test'],
            'imaging': ['imaging', 'radiology', 'scan', 'x-ray', 'mri', 'ct'],
            'consultation': ['consultation', 'clinic', 'visit', 'outpatient'],
            'dental': ['dental', 'dentist', 'tooth'],
            'gynecology': ['gyne', 'obstetric', 'maternity', 'antenatal'],
            'surgery': ['surgery', 'procedure', 'theatre', 'operation'],
            'emergency': ['emergency', 'casualty', 'er'],
            'ambulance': ['ambulance', 'ems', 'transport'],
            'admission': ['admission', 'ward', 'inpatient', 'bed'],
            'other': [],
        }
        keywords = service_type_keywords.get(service_type_normalized, [])
        if service_type_normalized not in keywords:
            keywords.append(service_type_normalized)
        if keywords:
            keyword_query = Q()
            for kw in keywords:
                keyword_query |= (
                    Q(category__name__icontains=kw) |
                    Q(category__code__icontains=kw) |
                    Q(description__icontains=kw) |
                    Q(revenue_stream__name__icontains=kw) |
                    Q(revenue_stream__code__icontains=kw)
                )
            service_filter |= keyword_query
        
        # Match payment receipts linked to the revenue's invoice/patient
        receipt_match = Q(invoice__receipts__service_type__iexact=service_type_normalized) | Q(
            invoice__receipts__service_type__in=[
                f"{service_type_normalized}_{suffix}" for suffix in ['prescription', 'walkin', 'sale']
            ]
        )
        service_filter |= receipt_match
        
        revenues = revenues.filter(service_filter).distinct()
        service_type = service_type_normalized
    
    # Base summaries from Revenue entries
    category_summary = {}
    service_summary = {}
    method_summary = {}
    
    revenue_by_category_qs = revenues.values('category__name').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    revenue_by_service_type_qs = revenues.values('service_type').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    revenue_by_method_qs = revenues.values('payment_method').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    total_revenue = revenues.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    entry_count = revenues.count()
    
    for item in revenue_by_category_qs:
        name = item['category__name'] or 'Uncategorized'
        data = category_summary.setdefault(name, {'total': Decimal('0.00'), 'count': 0})
        data['total'] += item['total'] or Decimal('0.00')
        data['count'] += item['count'] or 0
    
    for item in revenue_by_service_type_qs:
        key = item['service_type'] or 'other'
        data = service_summary.setdefault(key, {'total': Decimal('0.00'), 'count': 0})
        data['total'] += item['total'] or Decimal('0.00')
        data['count'] += item['count'] or 0
    
    for item in revenue_by_method_qs:
        method = item['payment_method'] or 'unknown'
        data = method_summary.setdefault(method, {'total': Decimal('0.00'), 'count': 0})
        data['total'] += item['total'] or Decimal('0.00')
        data['count'] += item['count'] or 0
    
    service_type_label_map = dict(Revenue.SERVICE_TYPES)
    
    # Fallback / Augment with PaymentReceipt data (ensures Pharmacy/Lab filters show real revenue)
    receipt_entries = []
    receipts = PaymentReceipt.objects.filter(
        receipt_date__date__gte=start_date,
        receipt_date__date__lte=end_date,
        is_deleted=False
    ).select_related('patient', 'invoice').prefetch_related('invoice__lines__service_code')
    
    if service_type:
        receipt_filter = Q(service_type__iexact=service_type_normalized)
        for kw in keywords:
            receipt_filter |= Q(service_type__icontains=kw)
        receipts = receipts.filter(receipt_filter)
    
    receipt_total = Decimal('0.00')
    for receipt in receipts:
        invoice_lines = list(receipt.invoice.lines.all()) if receipt.invoice else []
        service_code = invoice_lines[0].service_code if invoice_lines else None
        category_name = ''
        if service_code:
            if hasattr(service_code, 'name') and service_code.name:
                category_name = service_code.name
            elif hasattr(service_code, 'description') and service_code.description:
                category_name = service_code.description
            elif hasattr(service_code, 'service_name') and service_code.service_name:
                category_name = service_code.service_name
            elif hasattr(service_code, 'category') and service_code.category:
                category_name = getattr(service_code.category, 'name', '') or getattr(service_code.category, 'description', '')
            else:
                category_name = getattr(service_code, 'code', '') or 'Service'
        
        entry = SimpleNamespace(
            revenue_date=receipt.receipt_date.date(),
            revenue_number=receipt.receipt_number,
            category=None,
            category_name=category_name,
            service_type=receipt.service_type or 'other',
            service_label=service_type_label_map.get(receipt.service_type, receipt.service_type.title() if receipt.service_type else 'Other'),
            description=receipt.notes or f"Receipt {receipt.receipt_number}",
            patient=receipt.patient,
            payment_method=receipt.payment_method,
            amount=receipt.amount_paid,
            notes=receipt.notes,
            is_receipt=True,
        )
        receipt_entries.append(entry)
        receipt_total += receipt.amount_paid or Decimal('0.00')
        
        name = category_name or 'Uncategorized'
        cat_data = category_summary.setdefault(name, {'total': Decimal('0.00'), 'count': 0})
        cat_data['total'] += receipt.amount_paid or Decimal('0.00')
        cat_data['count'] += 1
        
        svc_key = entry.service_type or 'other'
        svc_data = service_summary.setdefault(svc_key, {'total': Decimal('0.00'), 'count': 0})
        svc_data['total'] += receipt.amount_paid or Decimal('0.00')
        svc_data['count'] += 1
        
        method_key = receipt.payment_method or 'unknown'
        meth_data = method_summary.setdefault(method_key, {'total': Decimal('0.00'), 'count': 0})
        meth_data['total'] += receipt.amount_paid or Decimal('0.00')
        meth_data['count'] += 1
    
    report_entries = list(revenues) + receipt_entries
    total_revenue = (total_revenue or Decimal('0.00')) + receipt_total
    entry_count += len(receipt_entries)
    average_revenue = (total_revenue / entry_count) if entry_count else Decimal('0.00')
    
    # Convert summaries to list with percentages
    revenue_by_category = []
    for name, data in category_summary.items():
        percentage = (data['total'] / total_revenue * 100) if total_revenue else Decimal('0.00')
        revenue_by_category.append({
            'category__name': name,
            'total': data['total'],
            'count': data['count'],
            'percentage': percentage
        })
    
    revenue_by_service_type = []
    for key, data in service_summary.items():
        percentage = (data['total'] / total_revenue * 100) if total_revenue else Decimal('0.00')
        revenue_by_service_type.append({
            'service_type': key,
            'label': service_type_label_map.get(key, key.title() if key else 'Other'),
            'total': data['total'],
            'count': data['count'],
            'percentage': percentage
        })
    
    revenue_by_method = []
    for key, data in method_summary.items():
        percentage = (data['total'] / total_revenue * 100) if total_revenue else Decimal('0.00')
        revenue_by_method.append({
            'method': key,
            'total': data['total'],
            'count': data['count'],
            'percentage': percentage
        })
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'date_from': start_date,
        'date_to': end_date,
        'revenues': report_entries,
        'revenue_entries': report_entries,
        'selected_category': str(category_id) if category_id else '',
        'selected_service_type': service_type or '',
        'revenue_by_category': revenue_by_category,
        'revenue_by_service_type': revenue_by_service_type,
        'revenue_by_method': revenue_by_method,
        'total_revenue': total_revenue,
        'average_revenue': average_revenue,
        'entry_count': entry_count,
        'categories': RevenueCategory.objects.filter(is_active=True),
        'service_types': Revenue.SERVICE_TYPES,
        'report_title': f'Revenue Report - {start_date} to {end_date}',
    }
    
    return render(request, 'hospital/revenue_report.html', context)


@login_required
@user_passes_test(is_accountant)
def expense_report(request):
    """Detailed Expense Report"""
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    category_id = request.GET.get('category')
    
    if not start_date or not end_date:
        today = timezone.now().date()
        start_date = today.replace(day=1)
        end_date = today
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Get expenses (include approved AND paid for proper reporting)
    expenses = Expense.objects.filter(
        expense_date__gte=start_date,
        expense_date__lte=end_date,
        status__in=['approved', 'paid']  # FIXED: Include approved expenses from procurement
    ).select_related('category', 'payment_voucher').order_by('-expense_date')
    
    if category_id:
        expenses = expenses.filter(category_id=category_id)
    
    # Summary by category
    expense_by_category = expenses.values('category__name').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    total_expense = expenses.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    count = expenses.count()
    average_expense = (total_expense / count) if count > 0 else Decimal('0.00')
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'expenses': expenses,
        'expense_entries': expenses,  # FIXED: Template expects this variable name
        'expense_by_category': expense_by_category,
        'total_expense': total_expense,
        'total_expenses': total_expense,  # FIXED: Template expects this too
        'average_expense': average_expense,  # FIXED: Added for stats card
        'categories': ExpenseCategory.objects.filter(is_active=True),
        'report_title': f'Expense Report - {start_date} to {end_date}',
    }
    
    return render(request, 'hospital/expense_report.html', context)


@login_required
@role_required('accountant', 'senior_account_officer', 'admin')
def payment_voucher_list(request):
    """World-Class Payment Voucher Management"""
    
    # Get filter parameters
    status = request.GET.get('status', 'all')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search = request.GET.get('search', '')
    
    # Base queryset
    vouchers = PaymentVoucher.objects.filter(is_deleted=False).select_related(
        'requested_by', 'approved_by', 'paid_by', 'expense_account', 'payment_account'
    )
    
    # Apply filters
    if status != 'all':
        vouchers = vouchers.filter(status=status)
    
    if date_from:
        vouchers = vouchers.filter(voucher_date__gte=date_from)
    
    if date_to:
        vouchers = vouchers.filter(voucher_date__lte=date_to)
    
    if search:
        vouchers = vouchers.filter(
            Q(voucher_number__icontains=search) |
            Q(payee_name__icontains=search) |
            Q(description__icontains=search) |
            Q(payment_reference__icontains=search)
        )
    
    vouchers = vouchers.order_by('-voucher_date', '-created')
    
    # Comprehensive Statistics
    stats = {
        'total_pending': PaymentVoucher.objects.filter(status='pending_approval', is_deleted=False).count(),
        'total_approved': PaymentVoucher.objects.filter(status='approved', is_deleted=False).count(),
        'total_paid': PaymentVoucher.objects.filter(status='paid', is_deleted=False).count(),
        'total_all': PaymentVoucher.objects.filter(is_deleted=False).count(),
        'amount_pending': PaymentVoucher.objects.filter(status='pending_approval', is_deleted=False).aggregate(total=Sum('amount'))['total'] or Decimal('0.00'),
        'amount_approved': PaymentVoucher.objects.filter(status='approved', is_deleted=False).aggregate(total=Sum('amount'))['total'] or Decimal('0.00'),
        'amount_paid': PaymentVoucher.objects.filter(status='paid', is_deleted=False).aggregate(total=Sum('amount'))['total'] or Decimal('0.00'),
        'amount_total': PaymentVoucher.objects.filter(is_deleted=False).aggregate(total=Sum('amount'))['total'] or Decimal('0.00'),
    }
    
    context = {
        'vouchers': vouchers[:200],  # Limit for performance
        'stats': stats,
        'status_filter': status,
        'date_from': date_from,
        'date_to': date_to,
        'search_query': search,
    }
    
    return render(request, 'hospital/payment_voucher_list_worldclass.html', context)


@login_required
@user_passes_test(is_accountant)
def mark_voucher_paid(request):
    """Mark payment voucher as paid"""
    if request.method == 'POST':
        voucher_id = request.POST.get('voucher_id')
        payment_date = request.POST.get('payment_date')
        payment_reference = request.POST.get('payment_reference')
        payment_method = request.POST.get('payment_method', 'bank_transfer')
        notes = request.POST.get('notes', '')
        
        try:
            voucher = PaymentVoucher.objects.get(id=voucher_id, is_deleted=False)
            
            if voucher.status != 'approved':
                messages.error(request, 'Only approved vouchers can be marked as paid.')
                return redirect('hospital:payment_voucher_list')
            
            # Update voucher
            voucher.status = 'paid'
            voucher.payment_date = payment_date
            voucher.payment_reference = payment_reference
            voucher.payment_method = payment_method
            voucher.paid_by = request.user
            if notes:
                voucher.notes = (voucher.notes + '\n\n' if voucher.notes else '') + f"Payment Notes: {notes}"
            voucher.save()
            
            # Update linked AP if exists
            try:
                ap_entry = AccountsPayable.objects.filter(
                    payment_voucher=voucher
                ).first()
                if ap_entry:
                    ap_entry.amount_paid = ap_entry.amount
                    ap_entry.balance_due = Decimal('0.00')
                    ap_entry.save()
                    print(f"[ACCOUNTING] ✅ Updated AP: {ap_entry.bill_number} - Marked as paid")
            except Exception as e:
                print(f"[ACCOUNTING] ⚠️ Could not update AP: {e}")
            
            # Post payment to General Ledger
            try:
                from .procurement_accounting_integration import post_payment_to_ledger
                journal_entry = post_payment_to_ledger(voucher)
                if journal_entry:
                    messages.success(
                        request,
                        f'✅ Payment voucher {voucher.voucher_number} marked as paid! '
                        f'Posted to GL: {journal_entry.entry_number}'
                    )
                else:
                    messages.success(request, f'✅ Payment voucher {voucher.voucher_number} marked as paid successfully!')
            except Exception as e:
                print(f"[LEDGER] ⚠️ Could not post to GL: {e}")
                messages.success(request, f'✅ Payment voucher {voucher.voucher_number} marked as paid successfully!')
            
        except PaymentVoucher.DoesNotExist:
            messages.error(request, 'Payment voucher not found.')
        except Exception as e:
            messages.error(request, f'Error marking voucher as paid: {e}')
    
    return redirect('hospital:payment_voucher_list')


@login_required
@user_passes_test(is_accountant)
def export_vouchers_excel(request):
    """Export payment vouchers to Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from datetime import datetime
    
    # Get filtered vouchers (same as list view)
    status = request.GET.get('status', 'all')
    vouchers = PaymentVoucher.objects.filter(is_deleted=False)
    if status != 'all':
        vouchers = vouchers.filter(status=status)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment Vouchers"
    
    # Headers
    headers = ['Voucher #', 'Date', 'Payee', 'Type', 'Description', 'Amount (GHS)', 'Status', 'Payment Date', 'Reference']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for voucher in vouchers:
        ws.append([
            voucher.voucher_number,
            voucher.voucher_date.strftime('%Y-%m-%d'),
            voucher.payee_name,
            voucher.get_payment_type_display(),
            voucher.description[:100],
            float(voucher.amount),
            voucher.get_status_display(),
            voucher.payment_date.strftime('%Y-%m-%d') if voucher.payment_date else '',
            voucher.payment_reference or '',
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="payment_vouchers_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_accountant)
def export_vouchers_pdf(request):
    """Export payment vouchers to PDF"""
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from datetime import datetime
    
    # Get filtered vouchers
    status = request.GET.get('status', 'all')
    vouchers = PaymentVoucher.objects.filter(is_deleted=False)
    if status != 'all':
        vouchers = vouchers.filter(status=status)
    
    # Create PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="payment_vouchers_{datetime.now().strftime("%Y%m%d")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#667eea'),
        spaceAfter=30,
    )
    elements.append(Paragraph('Payment Vouchers Report', title_style))
    elements.append(Paragraph(f'Generated: {datetime.now().strftime("%B %d, %Y %I:%M %p")}', styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Table data
    data = [['Voucher #', 'Date', 'Payee', 'Amount', 'Status']]
    
    for voucher in vouchers[:100]:  # Limit to 100 for PDF
        data.append([
            voucher.voucher_number,
            voucher.voucher_date.strftime('%Y-%m-%d'),
            voucher.payee_name[:30],
            f'GHS {float(voucher.amount):,.2f}',
            voucher.get_status_display(),
        ])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    return response


@login_required
@user_passes_test(is_accountant)
def receipt_voucher_list(request):
    """Receipt Voucher Management"""
    
    search_query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', '').strip()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    receipts = ReceiptVoucher.objects.filter(is_deleted=False).select_related(
        'patient', 'revenue_account', 'received_by'
    ).order_by('-receipt_date')
    
    if status_filter:
        receipts = receipts.filter(status=status_filter)
    
    if start_date:
        try:
            start_date_parsed = datetime.strptime(start_date, '%Y-%m-%d').date()
            receipts = receipts.filter(receipt_date__date__gte=start_date_parsed)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_date_parsed = datetime.strptime(end_date, '%Y-%m-%d').date()
            receipts = receipts.filter(receipt_date__date__lte=end_date_parsed)
        except ValueError:
            pass
    
    if search_query:
        receipts = receipts.filter(
            Q(receipt_number__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(patient__first_name__icontains=search_query) |
            Q(patient__last_name__icontains=search_query) |
            Q(received_from__icontains=search_query)
        )
    
    total_amount = receipts.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    status_counts = ReceiptVoucher.objects.filter(is_deleted=False).values('status').annotate(count=Count('id'))
    
    context = {
        'receipts': receipts,
        'total_amount': total_amount,
        'status_counts': status_counts,
        'selected_status': status_filter,
        'search_query': search_query,
        'start_date': start_date,
        'end_date': end_date,
    }
    
    return render(request, 'hospital/receipt_voucher_list.html', context)


# ==================== API ENDPOINTS ====================

@login_required
@user_passes_test(is_accountant)
def accounting_api_stats(request):
    """API endpoint for accounting statistics"""
    
    today = timezone.now().date()
    start_of_month = today.replace(day=1)
    
    stats = {
        'revenue_today': float(Revenue.objects.filter(revenue_date=today).aggregate(total=Sum('amount'))['total'] or 0),
        'revenue_month': float(Revenue.objects.filter(revenue_date__gte=start_of_month).aggregate(total=Sum('amount'))['total'] or 0),
        'expenses_today': float(Expense.objects.filter(expense_date=today, status='paid').aggregate(total=Sum('amount'))['total'] or 0),
        'expenses_month': float(Expense.objects.filter(expense_date__gte=start_of_month, status='paid').aggregate(total=Sum('amount'))['total'] or 0),
        'total_receivable': float(AdvancedAccountsReceivable.objects.filter(balance_due__gt=0).aggregate(total=Sum('balance_due'))['total'] or 0),
        'overdue_receivable': float(AdvancedAccountsReceivable.objects.filter(is_overdue=True).aggregate(total=Sum('balance_due'))['total'] or 0),
        'total_payable': float(AccountsPayable.objects.filter(balance_due__gt=0).aggregate(total=Sum('balance_due'))['total'] or 0),
        'pending_vouchers': PaymentVoucher.objects.filter(status='pending_approval').count(),
        'approved_vouchers': PaymentVoucher.objects.filter(status='approved').count(),
    }
    
    return JsonResponse(stats)

