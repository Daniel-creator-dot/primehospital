"""
Excel export for accountant (RMC) payroll — runs summary and single-run line detail.
"""
from __future__ import annotations

from io import BytesIO
from typing import Any


def _require_openpyxl():
    try:
        import openpyxl
        from openpyxl.styles import Font
        return openpyxl, Font
    except ImportError as e:
        raise RuntimeError(
            'Excel export requires openpyxl. Install with: pip install openpyxl'
        ) from e


def payroll_runs_summary_to_xlsx_bytes(rows: list[dict[str, Any]]) -> bytes:
    """
    rows: dicts with keys payroll_number, staff_names, emp_codes, period_start, period_end,
    label, pay_date, gross, deductions, net, status, source (dates as date or str).
    """
    openpyxl, Font = _require_openpyxl()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Payroll runs'
    headers = [
        'Payroll #',
        'Staff names',
        'Emp. codes',
        'Period start',
        'Period end',
        'Label',
        'Pay date',
        'Gross (GHS)',
        'Deductions (GHS)',
        'Net (GHS)',
        'Status',
        'Source',
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        ws.append([
            r.get('payroll_number'),
            r.get('staff_names'),
            r.get('emp_codes'),
            r.get('period_start'),
            r.get('period_end'),
            r.get('label'),
            r.get('pay_date'),
            float(r.get('gross') or 0),
            float(r.get('deductions') or 0),
            float(r.get('net') or 0),
            r.get('status'),
            r.get('source'),
        ])
    for col in range(8, 11):
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col).number_format = '#,##0.00'
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def payroll_run_lines_to_xlsx_bytes(
    payroll_number: str,
    period_label: str,
    pay_date,
    status_display: str,
    entries: list,
    has_rmc_columns: bool,
) -> bytes:
    """entries: AccountingPayrollEntry queryset or list with staff, amounts."""
    openpyxl, Font = _require_openpyxl()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Lines'
    ws.append([f'Payroll run: {payroll_number}'])
    ws.append([period_label])
    ws.append([f'Pay date: {pay_date}', f'Status: {status_display}'])
    ws.append([])
    if has_rmc_columns:
        headers = [
            '#',
            'Emp. code',
            'Employee name',
            'Department',
            'Basic salary',
            'Housing',
            'Transport',
            'Other allowance',
            'Medical',
            'Risk / emergency',
            'Gross',
            'SSF 5.5%',
            'PF employee',
            'PAYE',
            'Loan',
            'Other ded.',
            'Deductions',
            'Net pay',
        ]
    else:
        headers = ['Emp. code', 'Employee name', 'Department', 'Gross', 'Deductions', 'Net pay']
    ws.append(headers)
    hdr_row = ws.max_row
    for cell in ws[hdr_row]:
        cell.font = Font(bold=True)
    money_cols: set[int] = set()
    for i, e in enumerate(entries, start=1):
        staff = e.staff
        name = staff.user.get_full_name() or staff.user.username
        emp = (staff.employee_id or '').strip()
        if has_rmc_columns:
            dept = (getattr(e, 'department_snapshot', None) or '').strip() or (
                staff.department.name if getattr(staff, 'department', None) else ''
            )
            row_vals = [
                getattr(e, 'sheet_serial', None) or i,
                emp,
                name,
                dept,
                float(e.basic_salary or 0),
                float(e.housing_allowance or 0),
                float(e.transport_allowance or 0),
                float(e.other_allowances or 0),
                float(e.medical_allowance or 0),
                float(e.risk_emergency_allowance or 0),
                float(e.gross_pay or 0),
                float(e.ssnit_employee or 0),
                float(e.pension_employee or 0),
                float(e.paye_tax or 0),
                float(e.loan_deduction or 0),
                float(e.other_deductions_detail or 0),
                float(e.deductions or 0),
                float(e.net_pay or 0),
            ]
            money_cols.update(range(5, 19))
        else:
            dept = staff.department.name if getattr(staff, 'department', None) else ''
            row_vals = [emp, name, dept, float(e.gross_pay or 0), float(e.deductions or 0), float(e.net_pay or 0)]
            money_cols.update({4, 5, 6})
        ws.append(row_vals)
    first_data = hdr_row + 1
    for row in range(first_data, ws.max_row + 1):
        for c in money_cols:
            ws.cell(row=row, column=c).number_format = '#,##0.00'
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
