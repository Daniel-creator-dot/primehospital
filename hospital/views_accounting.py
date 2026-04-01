"""
Accounting and Financial Management Views
"""
import logging
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q, Sum, Count
from django.http import JsonResponse, HttpResponse
from datetime import date, timedelta
from decimal import Decimal
from .models import Patient, Invoice
from .models_accounting import (
    Account, Transaction, PaymentReceipt, AccountsReceivable,
    GeneralLedger, JournalEntry, JournalEntryLine
)
from .models_workflow import Bill, CashierSession

logger = logging.getLogger(__name__)


def is_accountant(user):
    """Check if user is an accountant"""
    return user.groups.filter(name__in=['Cashier', 'Admin']).exists() or user.is_staff


@login_required
@user_passes_test(is_accountant, login_url='/admin/login/')
def accounting_dashboard(request):
    """Accounting main dashboard with FULL SYNC"""
    today = timezone.now().date()
    
    # Accounts Receivable Summary - Use AdvancedAccountsReceivable (new model)
    from .models_accounting_advanced import AdvancedAccountsReceivable
    ar_total = Decimal('0.00')
    overdue_receivable = Decimal('0.00')
    try:
        ar_total = AdvancedAccountsReceivable.objects.filter(
            balance_due__gt=0,
            is_deleted=False
        ).aggregate(Sum('balance_due'))['balance_due__sum'] or Decimal('0.00')
        
        # Calculate overdue receivables
        overdue_receivable = AdvancedAccountsReceivable.objects.filter(
            balance_due__gt=0,
            is_overdue=True,
            is_deleted=False
        ).aggregate(Sum('balance_due'))['balance_due__sum'] or Decimal('0.00')
        
        ar_by_aging = {
            'current': AdvancedAccountsReceivable.objects.filter(aging_bucket='current', is_deleted=False).aggregate(Sum('balance_due'))['balance_due__sum'] or Decimal('0.00'),
            'aging_31_60': AdvancedAccountsReceivable.objects.filter(aging_bucket='31-60', is_deleted=False).aggregate(Sum('balance_due'))['balance_due__sum'] or Decimal('0.00'),
            'aging_61_90': AdvancedAccountsReceivable.objects.filter(aging_bucket='61-90', is_deleted=False).aggregate(Sum('balance_due'))['balance_due__sum'] or Decimal('0.00'),
            'aging_90_plus': AdvancedAccountsReceivable.objects.filter(aging_bucket='90+', is_deleted=False).aggregate(Sum('balance_due'))['balance_due__sum'] or Decimal('0.00'),
        }
    except Exception as e:
        # Fallback to old model if AdvancedAccountsReceivable doesn't exist
        logger.error(f"Error calculating AR from AdvancedAccountsReceivable: {e}")
        try:
            ar_total = AccountsReceivable.objects.filter(
                outstanding_amount__gt=0,
                is_deleted=False
            ).aggregate(Sum('outstanding_amount'))['outstanding_amount__sum'] or Decimal('0.00')
            
            overdue_receivable = AccountsReceivable.objects.filter(
                outstanding_amount__gt=0,
                is_overdue=True,
                is_deleted=False
            ).aggregate(Sum('outstanding_amount'))['outstanding_amount__sum'] or Decimal('0.00')
            
            ar_by_aging = {
                'current': AccountsReceivable.objects.filter(aging_bucket='current', is_deleted=False).aggregate(Sum('outstanding_amount'))['outstanding_amount__sum'] or Decimal('0.00'),
                'aging_31_60': AccountsReceivable.objects.filter(aging_bucket='31-60', is_deleted=False).aggregate(Sum('outstanding_amount'))['outstanding_amount__sum'] or Decimal('0.00'),
                'aging_61_90': AccountsReceivable.objects.filter(aging_bucket='61-90', is_deleted=False).aggregate(Sum('outstanding_amount'))['outstanding_amount__sum'] or Decimal('0.00'),
                'aging_90_plus': AccountsReceivable.objects.filter(aging_bucket='90+', is_deleted=False).aggregate(Sum('outstanding_amount'))['outstanding_amount__sum'] or Decimal('0.00'),
            }
        except:
            ar_total = Decimal('0.00')
            overdue_receivable = Decimal('0.00')
            ar_by_aging = {
                'current': Decimal('0.00'),
                'aging_31_60': Decimal('0.00'),
                'aging_61_90': Decimal('0.00'),
                'aging_90_plus': Decimal('0.00'),
            }
    
    # Calculate start of month for monthly revenue
    start_of_month = today.replace(day=1)
    
    # Today's revenue from GENERAL LEDGER (source of truth)
    # Note: transaction_date is a DateField, so we don't use __date lookup
    today_revenue_gl = GeneralLedger.objects.filter(
        account__account_type='revenue',
        transaction_date=today,
        is_deleted=False
    ).aggregate(Sum('credit_amount'))['credit_amount__sum'] or Decimal('0.00')
    
    # Today's revenue from PaymentReceipts (for comparison)
    # Note: receipt_date is a DateTimeField, so we use __date lookup
    today_revenue_receipts = PaymentReceipt.objects.filter(
        receipt_date__date=today,
        is_deleted=False
    ).aggregate(Sum('amount_paid'))['amount_paid__sum'] or Decimal('0.00')
    
    # Monthly revenue from GENERAL LEDGER (this month)
    month_revenue_gl = GeneralLedger.objects.filter(
        account__account_type='revenue',
        transaction_date__gte=start_of_month,
        transaction_date__lte=today,
        is_deleted=False
    ).aggregate(Sum('credit_amount'))['credit_amount__sum'] or Decimal('0.00')
    
    # Monthly revenue from PaymentReceipts (this month)
    month_revenue_receipts = PaymentReceipt.objects.filter(
        receipt_date__date__gte=start_of_month,
        receipt_date__date__lte=today,
        is_deleted=False
    ).aggregate(Sum('amount_paid'))['amount_paid__sum'] or Decimal('0.00')
    
    # Today's revenue from Receivables - use invoice issued_at date (accrual basis)
    # Revenue is recognized when invoice is issued, not when AR entry is created
    today_revenue_from_ar = Decimal('0.00')
    try:
        today_ar_entries = AdvancedAccountsReceivable.objects.filter(
            invoice__issued_at__date=today,
            is_deleted=False
        ).select_related('invoice')
        for ar in today_ar_entries:
            # Only include if invoice was issued today
            if ar.invoice and ar.invoice.issued_at and ar.invoice.issued_at.date() == today:
                today_revenue_from_ar += ar.invoice_amount
    except Exception as e:
        logger.warning(f"Error calculating today's revenue from AR: {e}")
        pass
    
    # Also check invoices issued today (even if AR not created yet)
    today_invoices_revenue = Decimal('0.00')
    try:
        today_invoices_revenue = Invoice.objects.filter(
            issued_at__date=today,
            status__in=['issued', 'partially_paid', 'overdue'],
            is_deleted=False
        ).aggregate(Sum('total_amount'))['total_amount__sum'] or Decimal('0.00')
    except:
        pass
    
    # Combine: Use highest value (accrual basis - recognize revenue when service provided)
    # This ensures we show revenue when services are provided, not just when paid
    today_revenue = max(
        today_revenue_gl,
        today_revenue_from_ar,
        today_invoices_revenue,
        today_revenue_receipts
    )
    
    # Monthly revenue
    month_revenue = month_revenue_gl if month_revenue_gl > 0 else month_revenue_receipts
    
    # Calculate sync variance (absolute difference)
    sync_variance = abs(today_revenue_gl - today_revenue_receipts)
    is_synced = today_revenue_gl == today_revenue_receipts
    
    # Recent journal entries (proper accounting view)
    recent_journal_entries = JournalEntry.objects.filter(
        is_deleted=False
    ).select_related('entered_by', 'posted_by').prefetch_related('lines__account').order_by('-entry_date', '-created')[:15]
    
    # Recent transactions (old view, for compatibility)
    recent_transactions = Transaction.objects.filter(
        is_deleted=False
    ).select_related('patient', 'invoice').order_by('-transaction_date')[:20]
    
    # Open cashier sessions
    open_sessions = CashierSession.objects.filter(
        status='open',
        is_deleted=False
    ).select_related('cashier')[:5]
    
    # Get account balances for quick view
    from .models_accounting import Account
    key_accounts = Account.objects.filter(
        account_code__in=['1010', '4010', '4020', '4030', '4040', '4060'],  # Cash, Lab Rev, Pharmacy Rev, Imaging Rev, Consult Rev, Admission Rev
        is_deleted=False
    )
    
    account_balances = {}
    for account in key_accounts:
        # Calculate balance from all GL entries for this account
        gl_entries = GeneralLedger.objects.filter(
            account=account,
            is_deleted=False
        ).aggregate(
            total_debits=Sum('debit_amount'),
            total_credits=Sum('credit_amount')
        )
        
        total_debits = gl_entries['total_debits'] or Decimal('0.00')
        total_credits = gl_entries['total_credits'] or Decimal('0.00')
        
        # Calculate balance based on account type
        if account.account_type in ['asset', 'expense']:
            # Assets and Expenses: Debit increases, Credit decreases
            balance = total_debits - total_credits
        else:
            # Liabilities, Equity, Revenue: Credit increases, Debit decreases
            balance = total_credits - total_debits
        
        account_balances[account.account_code] = {
            'name': account.account_name,
            'balance': balance
        }
    
    # Calculate expenses (if Expense model exists)
    total_expenses = Decimal('0.00')
    try:
        from .models_accounting_advanced import Expense
        total_expenses = Expense.objects.filter(
            expense_date__gte=start_of_month,
            expense_date__lte=today,
            status='paid',
            is_deleted=False
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    except Exception as e:
        logger.warning(f"Could not calculate expenses: {e}")
        total_expenses = Decimal('0.00')
    
    # Calculate net income
    net_income = month_revenue - total_expenses
    
    # Get AR total (already calculated above)
    total_receivable = ar_total
    
    # Calculate Accounts Payable
    # Priority: Use General Ledger (Excel imported balances) if available, otherwise use AccountsPayable model
    total_payable = Decimal('0.00')
    try:
        # First, check General Ledger for AP accounts (Excel imported balances)
        # For Excel imports: debit amounts ARE the balances (independent, different companies)
        from .models_accounting_advanced import AdvancedGeneralLedger
        ap_accounts = Account.objects.filter(
            account_type='liability',
            account_name__icontains='payable',
            is_deleted=False
        )
        
        for ap_account in ap_accounts:
            # Sum all debit amounts (each is an independent balance from Excel import)
            ap_gl_total = AdvancedGeneralLedger.objects.filter(
                account=ap_account,
                is_voided=False,
                is_deleted=False
            ).aggregate(total=Sum('debit_amount'))['total'] or Decimal('0.00')
            total_payable += ap_gl_total
        
        # If General Ledger has no AP data, fall back to AccountsPayable model
        if total_payable == 0:
            from .models_accounting_advanced import AccountsPayable
            total_payable = AccountsPayable.objects.filter(
                balance_due__gt=0,
                is_deleted=False
            ).aggregate(total=Sum('balance_due'))['total'] or Decimal('0.00')
    except Exception as e:
        logger.warning(f"Error calculating Accounts Payable: {e}")
        # Fallback to AccountsPayable model
        try:
            from .models_accounting_advanced import AccountsPayable
            total_payable = AccountsPayable.objects.filter(
                balance_due__gt=0,
                is_deleted=False
            ).aggregate(total=Sum('balance_due'))['total'] or Decimal('0.00')
        except:
            total_payable = Decimal('0.00')
    
    # Payment Vouchers
    try:
        from .models_accounting_advanced import PaymentVoucher
        pending_vouchers = PaymentVoucher.objects.filter(
            status='pending_approval',
            is_deleted=False
        ).count()
    except:
        pending_vouchers = 0
    
    # Draft Journal Entries
    try:
        from .models_accounting_advanced import AdvancedJournalEntry
        draft_entries = AdvancedJournalEntry.objects.filter(
            status='draft',
            is_deleted=False
        ).count()
        posted_entries_month = AdvancedJournalEntry.objects.filter(
            entry_date__gte=start_of_month,
            entry_date__lte=today,
            status='posted',
            is_deleted=False
        ).count()
    except:
        draft_entries = 0
        posted_entries_month = 0
    
    # Revenue by Category (from GL revenue accounts)
    revenue_by_category = []
    try:
        revenue_accounts = Account.objects.filter(
            account_type='revenue',
            account_code__in=['4010', '4020', '4030', '4040', '4060'],
            is_deleted=False
        )
        for account in revenue_accounts:
            revenue = GeneralLedger.objects.filter(
                account=account,
                transaction_date__gte=start_of_month,
                transaction_date__lte=today,
                is_deleted=False
            ).aggregate(Sum('credit_amount'))['credit_amount__sum'] or Decimal('0.00')
            if revenue > 0:
                revenue_by_category.append({
                    'category__name': account.account_name,
                    'total': revenue
                })
    except:
        pass
    
    # Expenses by Category
    expenses_by_category = []
    try:
        from .models_accounting_advanced import ExpenseCategory
        expense_categories = ExpenseCategory.objects.filter(is_active=True, is_deleted=False)
        for category in expense_categories:
            expense_total = Expense.objects.filter(
                category=category,
                expense_date__gte=start_of_month,
                expense_date__lte=today,
                status='paid',
                is_deleted=False
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
            if expense_total > 0:
                expenses_by_category.append({
                    'category__name': category.name,
                    'total': expense_total
                })
    except:
        pass
    
    context = {
        'ar_total': ar_total,
        'ar_by_aging': ar_by_aging,
        'overdue_receivable': overdue_receivable,
        'today_revenue': today_revenue,
        'today_revenue_from_ar': today_revenue_from_ar,
        'today_invoices_revenue': today_invoices_revenue,
        'month_revenue': month_revenue,
        'total_revenue': month_revenue,  # For template compatibility
        'total_expenses': total_expenses,
        'net_income': net_income,
        'total_receivable': total_receivable,
        'total_payable': total_payable,
        'pending_vouchers': pending_vouchers,
        'draft_entries': draft_entries,
        'posted_entries_month': posted_entries_month,
        'revenue_by_category': revenue_by_category,
        'expenses_by_category': expenses_by_category,
        'today_revenue_gl': today_revenue_gl,
        'today_revenue_receipts': today_revenue_receipts,
        'month_revenue_gl': month_revenue_gl,
        'month_revenue_receipts': month_revenue_receipts,
        'sync_variance': sync_variance,
        'is_synced': is_synced,
        'recent_journal_entries': recent_journal_entries,
        'recent_transactions': recent_transactions,
        'open_sessions': open_sessions,
        'account_balances': account_balances,
        'today': today,
        'start_of_month': start_of_month,
    }
    return render(request, 'hospital/accounting_dashboard.html', context)


@login_required
@user_passes_test(is_accountant, login_url='/admin/login/')
def chart_of_accounts(request):
    """Chart of Accounts view - List all accounts with balances"""
    from .models_accounting import GeneralLedger
    
    account_type_filter = request.GET.get('type', '')
    search_query = request.GET.get('search', '')
    
    accounts = Account.objects.filter(is_deleted=False).select_related('parent_account')
    
    if account_type_filter:
        accounts = accounts.filter(account_type=account_type_filter)
    
    if search_query:
        accounts = accounts.filter(
            Q(account_code__icontains=search_query) |
            Q(account_name__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Calculate balances for each account
    account_list = []
    for account in accounts.order_by('account_code'):
        # Get all GL entries for this account up to today
        gl_entries = GeneralLedger.objects.filter(
            account=account,
            is_deleted=False
        ).aggregate(
            total_debits=Sum('debit_amount'),
            total_credits=Sum('credit_amount')
        )
        
        total_debits = gl_entries['total_debits'] or Decimal('0.00')
        total_credits = gl_entries['total_credits'] or Decimal('0.00')
        
        # Calculate balance based on account type
        if account.account_type in ['asset', 'expense']:
            # Assets and Expenses: Debit increases, Credit decreases
            balance = total_debits - total_credits
        else:
            # Liabilities, Equity, Revenue: Credit increases, Debit decreases
            balance = total_credits - total_debits
        
        account_list.append({
            'account': account,
            'balance': balance,
            'total_debits': total_debits,
            'total_credits': total_credits,
        })
    
    # Calculate totals by account type
    type_totals = {}
    for item in account_list:
        acc_type = item['account'].account_type
        if acc_type not in type_totals:
            type_totals[acc_type] = Decimal('0.00')
        type_totals[acc_type] += item['balance']
    
    # Check if PV/Cheque accounts are set up
    from .utils_pv_account_setup import get_pv_expense_accounts, get_pv_payment_accounts
    expense_accounts = get_pv_expense_accounts()
    payment_accounts = get_pv_payment_accounts()
    
    # Check for default accounts
    default_expense_codes = ['5010', '5020', '5030', '5040', '5050']
    default_payment_codes = ['1010', '1020', '1030']
    
    missing_expense = [code for code in default_expense_codes 
                      if not Account.objects.filter(account_code=code, is_active=True).exists()]
    missing_payment = [code for code in default_payment_codes 
                     if not Account.objects.filter(account_code=code, is_active=True).exists()]
    
    pv_accounts_setup = len(missing_expense) == 0 and len(missing_payment) == 0
    
    context = {
        'accounts': account_list,
        'type_totals': type_totals,
        'account_types': Account.ACCOUNT_TYPES,
        'selected_type': account_type_filter,
        'search_query': search_query,
        'pv_accounts_setup': pv_accounts_setup,
        'missing_pv_accounts': len(missing_expense) + len(missing_payment),
    }
    return render(request, 'hospital/chart_of_accounts.html', context)


@login_required
@user_passes_test(is_accountant, login_url='/admin/login/')
def accounts_receivable(request):
    """Accounts Receivable aging report"""
    from .models_accounting_advanced import AdvancedAccountsReceivable
    
    aging_filter = request.GET.get('aging', '')

    # Use AdvancedAccountsReceivable (new model) if available
    use_advanced = False
    try:
        ar_entries = AdvancedAccountsReceivable.objects.filter(
            balance_due__gt=0,
            is_deleted=False
        ).select_related('invoice', 'patient').order_by('due_date')
        use_advanced = True
    except:
        # Fallback to old model
        ar_entries = AccountsReceivable.objects.filter(
            outstanding_amount__gt=0,
            is_deleted=False
        ).select_related('invoice', 'patient').order_by('due_date')
    
    if aging_filter:
        ar_entries = ar_entries.filter(aging_bucket=aging_filter)
    
    # Update aging for all entries
    for entry in ar_entries:
        entry.update_aging()
    
    context = {
        'ar_entries': ar_entries,
        'aging_filter': aging_filter,
        'use_advanced': use_advanced,  # Pass flag to template
    }
    return render(request, 'hospital/accounts_receivable.html', context)


@login_required
@user_passes_test(is_accountant, login_url='/admin/login/')
def general_ledger(request):
    """General Ledger view"""
    account_filter = request.GET.get('account')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    entries = GeneralLedger.objects.filter(is_deleted=False).select_related('account')
    
    if account_filter:
        entries = entries.filter(account_id=account_filter)
    
    if start_date:
        try:
            entries = entries.filter(transaction_date__gte=date.fromisoformat(start_date))
        except ValueError:
            pass
    
    if end_date:
        try:
            entries = entries.filter(transaction_date__lte=date.fromisoformat(end_date))
        except ValueError:
            pass
    
    # Get account balances
    accounts = Account.objects.filter(is_active=True, is_deleted=False)
    
    context = {
        'entries': entries.order_by('-transaction_date')[:500],
        'accounts': accounts,
        'selected_account': account_filter,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'hospital/general_ledger.html', context)


@login_required
@user_passes_test(is_accountant, login_url='/admin/login/')
def trial_balance(request):
    """Trial Balance report - includes both GeneralLedger and AdvancedGeneralLedger"""
    report_date = request.GET.get('date')
    
    if report_date:
        try:
            report_date = date.fromisoformat(report_date)
        except ValueError:
            report_date = timezone.now().date()
    else:
        report_date = timezone.now().date()
    
    # Try to import AdvancedGeneralLedger if available
    try:
        from .models_accounting_advanced import AdvancedGeneralLedger
        HAS_ADVANCED = True
    except ImportError:
        HAS_ADVANCED = False
        AdvancedGeneralLedger = None
    
    accounts_list = Account.objects.filter(is_active=True, is_deleted=False).order_by('account_code')
    
    accounts_with_balance = []
    total_debits = Decimal('0.00')
    total_credits = Decimal('0.00')
    
    for account in accounts_list:
        # Get entries from GeneralLedger
        gl_entries = GeneralLedger.objects.filter(
            account=account,
            transaction_date__lte=report_date,
            is_deleted=False
        ).order_by('transaction_date', 'created')
        
        gl_debits = gl_entries.aggregate(Sum('debit_amount'))['debit_amount__sum'] or Decimal('0.00')
        gl_credits = gl_entries.aggregate(Sum('credit_amount'))['credit_amount__sum'] or Decimal('0.00')
        
        # Get entries from AdvancedGeneralLedger if available
        adv_debits = Decimal('0.00')
        adv_credits = Decimal('0.00')
        adv_entries_list = []
        
        if HAS_ADVANCED and AdvancedGeneralLedger:
            adv_entries = AdvancedGeneralLedger.objects.filter(
                account=account,
                transaction_date__lte=report_date,
                is_voided=False,
                is_deleted=False
            ).order_by('transaction_date', 'created')
            adv_debits = adv_entries.aggregate(Sum('debit_amount'))['debit_amount__sum'] or Decimal('0.00')
            adv_credits = adv_entries.aggregate(Sum('credit_amount'))['credit_amount__sum'] or Decimal('0.00')
            adv_entries_list = list(adv_entries)
        
        # Combine totals from both ledgers
        debits = gl_debits + adv_debits
        credits = gl_credits + adv_credits
        
        # Calculate balance based on account type
        # For assets and expenses: Debit increases, Credit decreases (balance = debits - credits)
        # For liabilities, equity, and revenue: Credit increases, Debit decreases (balance = credits - debits)
        if account.account_type in ['asset', 'expense']:
            balance = debits - credits
        else:
            # liability, equity, revenue
            balance = credits - debits
        
        # Only include accounts with activity
        if debits > 0 or credits > 0:
            # Collect all transaction entries for this account
            all_entries = []
            
            # Add GeneralLedger entries
            try:
                for entry in gl_entries:
                    all_entries.append({
                        'date': entry.transaction_date,
                        'entry_number': entry.entry_number or 'N/A',
                        'description': entry.description or '',
                        'reference_number': entry.reference_number or '',
                        'reference_type': entry.reference_type or '',
                        'debit': entry.debit_amount or Decimal('0.00'),
                        'credit': entry.credit_amount or Decimal('0.00'),
                        'source': 'GeneralLedger'
                    })
            except Exception as e:
                # Log error but continue
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error processing GeneralLedger entries for account {account.account_code}: {e}")
            
            # Add AdvancedGeneralLedger entries
            try:
                for entry in adv_entries_list:
                    # Get reference from journal entry if available
                    ref_number = ''
                    ref_type = ''
                    entry_number = 'N/A'
                    
                    try:
                        if hasattr(entry, 'journal_entry') and entry.journal_entry:
                            if hasattr(entry.journal_entry, 'entry_number'):
                                entry_number = entry.journal_entry.entry_number or 'N/A'
                            if hasattr(entry.journal_entry, 'reference_number'):
                                ref_number = entry.journal_entry.reference_number or ''
                            if hasattr(entry.journal_entry, 'entry_type'):
                                ref_type = entry.journal_entry.entry_type or ''
                    except:
                        pass  # If journal_entry access fails, use defaults
                    
                    all_entries.append({
                        'date': entry.transaction_date,
                        'entry_number': entry_number,
                        'description': getattr(entry, 'description', '') or '',
                        'reference_number': ref_number,
                        'reference_type': ref_type,
                        'debit': entry.debit_amount or Decimal('0.00'),
                        'credit': entry.credit_amount or Decimal('0.00'),
                        'source': 'AdvancedGeneralLedger'
                    })
            except Exception as e:
                # Log error but continue
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error processing AdvancedGeneralLedger entries for account {account.account_code}: {e}")
            
            # Sort all entries by date
            try:
                all_entries.sort(key=lambda x: (x['date'] if x['date'] else date(1900, 1, 1), x['entry_number']))
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error sorting entries for account {account.account_code}: {e}")
            
            # Ensure all values are Decimal, not None - always set defaults
            debits = debits if debits is not None else Decimal('0.00')
            credits = credits if credits is not None else Decimal('0.00')
            balance = balance if balance is not None else Decimal('0.00')
            
            # Ensure entries list exists
            if all_entries is None:
                all_entries = []
            
            # Add balance and entries as attributes to the account object for template access
            # Always set these attributes to prevent "INVALID" in templates
            account.balance = balance
            account.total_debits = debits
            account.total_credits = credits
            account.entries = all_entries
            account.entry_count = len(all_entries) if all_entries else 0
            accounts_with_balance.append(account)
            
            # For trial balance totals, we sum actual debits and credits
            total_debits += debits
            total_credits += credits
    
    balance_difference = total_debits - total_credits
    
    context = {
        'accounts': accounts_with_balance,
        'total_debits': total_debits,
        'total_credits': total_credits,
        'balance_difference': balance_difference,
        'as_of_date': report_date,
        'today': timezone.now(),
    }
    return render(request, 'hospital/trial_balance.html', context)


@login_required
@user_passes_test(is_accountant, login_url='/admin/login/')
def financial_statement_export_excel(request):
    """Export comprehensive financial statement to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    
    statement_type = request.GET.get('type', 'profit_loss')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    today = timezone.now().date()
    if not start_date:
        start_date = today.replace(month=1, day=1)
    else:
        try:
            start_date = date.fromisoformat(start_date)
        except ValueError:
            start_date = today.replace(month=1, day=1)
    
    if not end_date:
        end_date = today
    else:
        try:
            end_date = date.fromisoformat(end_date)
        except ValueError:
            end_date = today
    
    # Calculate revenue and expenses totals first (used in multiple worksheets)
    revenue_accounts_all = Account.objects.filter(account_type='revenue', is_deleted=False).order_by('account_code')
    total_revenue = Decimal('0.00')
    for account in revenue_accounts_all:
        entries = GeneralLedger.objects.filter(
            account=account,
            transaction_date__gte=start_date,
            transaction_date__lte=end_date,
            is_deleted=False
        ).aggregate(
            total_credits=Sum('credit_amount'),
            total_debits=Sum('debit_amount')
        )
        credits = entries['total_credits'] or Decimal('0.00')
        debits = entries['total_debits'] or Decimal('0.00')
        total_revenue += (credits - debits)
    
    expense_accounts_all = Account.objects.filter(account_type='expense', is_deleted=False).order_by('account_code')
    total_expenses = Decimal('0.00')
    for account in expense_accounts_all:
        entries = GeneralLedger.objects.filter(
            account=account,
            transaction_date__gte=start_date,
            transaction_date__lte=end_date,
            is_deleted=False
        ).aggregate(
            total_debits=Sum('debit_amount'),
            total_credits=Sum('credit_amount')
        )
        debits = entries['total_debits'] or Decimal('0.00')
        credits = entries['total_credits'] or Decimal('0.00')
        total_expenses += (debits - credits)
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # Helper function to add header
    def add_header(ws, title, row=1):
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = title
        cell.font = title_font
        cell.alignment = center_align
        ws.row_dimensions[row].height = 25
        return row + 2
    
    # Helper function to format currency
    def format_currency(value):
        return f"GHS {value:,.2f}"
    
    # ===== WORKSHEET 1: Chart of Accounts =====
    ws_coa = wb.create_sheet("Chart of Accounts")
    row = add_header(ws_coa, f"Chart of Accounts - As of {end_date.strftime('%B %d, %Y')}")
    
    # Headers
    headers = ['Account Code', 'Account Name', 'Type', 'Balance', 'Status', 'Description']
    for col, header in enumerate(headers, 1):
        cell = ws_coa.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
        ws_coa.column_dimensions[get_column_letter(col)].width = 20
    
    row += 1
    
    # Get all accounts
    accounts = Account.objects.filter(is_deleted=False).order_by('account_code')
    for account in accounts:
        # Calculate balance
        gl_entries = GeneralLedger.objects.filter(
            account=account,
            transaction_date__lte=end_date,
            is_deleted=False
        ).aggregate(
            total_debits=Sum('debit_amount'),
            total_credits=Sum('credit_amount')
        )
        
        total_debits = gl_entries['total_debits'] or Decimal('0.00')
        total_credits = gl_entries['total_credits'] or Decimal('0.00')
        
        if account.account_type in ['asset', 'expense']:
            balance = total_debits - total_credits
        else:
            balance = total_credits - total_debits
        
        ws_coa.cell(row=row, column=1).value = account.account_code
        ws_coa.cell(row=row, column=2).value = account.account_name
        ws_coa.cell(row=row, column=3).value = account.get_account_type_display()
        ws_coa.cell(row=row, column=4).value = float(balance)
        ws_coa.cell(row=row, column=4).number_format = '#,##0.00'
        ws_coa.cell(row=row, column=5).value = 'Active' if account.is_active else 'Inactive'
        ws_coa.cell(row=row, column=6).value = account.description or ''
        
        for col in range(1, 7):
            ws_coa.cell(row=row, column=col).border = border
        row += 1
    
    # ===== WORKSHEET 2: Income Statement =====
    ws_is = wb.create_sheet("Income Statement")
    row = add_header(ws_is, f"Income Statement - {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}")
    
    # Revenue section
    ws_is.cell(row=row, column=1).value = "REVENUE"
    ws_is.cell(row=row, column=1).font = subheader_font
    row += 1
    
    revenue_accounts = Account.objects.filter(account_type='revenue', is_deleted=False).order_by('account_code')
    total_revenue_ws = Decimal('0.00')  # For worksheet display only
    
    for account in revenue_accounts:
        entries = GeneralLedger.objects.filter(
            account=account,
            transaction_date__gte=start_date,
            transaction_date__lte=end_date,
            is_deleted=False
        ).aggregate(
            total_credits=Sum('credit_amount'),
            total_debits=Sum('debit_amount')
        )
        
        credits = entries['total_credits'] or Decimal('0.00')
        debits = entries['total_debits'] or Decimal('0.00')
        net_revenue = credits - debits
        
        if net_revenue != 0:
            ws_is.cell(row=row, column=1).value = account.account_name
            ws_is.cell(row=row, column=2).value = float(net_revenue)
            ws_is.cell(row=row, column=2).number_format = '#,##0.00'
            ws_is.cell(row=row, column=2).alignment = right_align
            total_revenue_ws += net_revenue
            row += 1
    
    # Total Revenue
    ws_is.cell(row=row, column=1).value = "Total Revenue"
    ws_is.cell(row=row, column=1).font = subheader_font
    ws_is.cell(row=row, column=2).value = float(total_revenue)
    ws_is.cell(row=row, column=2).number_format = '#,##0.00'
    ws_is.cell(row=row, column=2).font = subheader_font
    ws_is.cell(row=row, column=2).alignment = right_align
    row += 2
    
    # Expenses section
    ws_is.cell(row=row, column=1).value = "EXPENSES"
    ws_is.cell(row=row, column=1).font = subheader_font
    row += 1
    
    expense_accounts = Account.objects.filter(account_type='expense', is_deleted=False).order_by('account_code')
    total_expenses_ws = Decimal('0.00')  # For worksheet display only
    
    for account in expense_accounts:
        entries = GeneralLedger.objects.filter(
            account=account,
            transaction_date__gte=start_date,
            transaction_date__lte=end_date,
            is_deleted=False
        ).aggregate(
            total_debits=Sum('debit_amount'),
            total_credits=Sum('credit_amount')
        )
        
        debits = entries['total_debits'] or Decimal('0.00')
        credits = entries['total_credits'] or Decimal('0.00')
        net_expense = debits - credits
        
        if net_expense != 0:
            ws_is.cell(row=row, column=1).value = account.account_name
            ws_is.cell(row=row, column=2).value = float(net_expense)
            ws_is.cell(row=row, column=2).number_format = '#,##0.00'
            ws_is.cell(row=row, column=2).alignment = right_align
            total_expenses_ws += net_expense
            row += 1
    
    # Total Expenses
    ws_is.cell(row=row, column=1).value = "Total Expenses"
    ws_is.cell(row=row, column=1).font = subheader_font
    ws_is.cell(row=row, column=2).value = float(total_expenses)
    ws_is.cell(row=row, column=2).number_format = '#,##0.00'
    ws_is.cell(row=row, column=2).font = subheader_font
    ws_is.cell(row=row, column=2).alignment = right_align
    row += 2
    
    # Net Income
    net_income = total_revenue - total_expenses
    ws_is.cell(row=row, column=1).value = "NET INCOME"
    ws_is.cell(row=row, column=1).font = title_font
    ws_is.cell(row=row, column=2).value = float(net_income)
    ws_is.cell(row=row, column=2).number_format = '#,##0.00'
    ws_is.cell(row=row, column=2).font = title_font
    ws_is.cell(row=row, column=2).alignment = right_align
    
    ws_is.column_dimensions['A'].width = 40
    ws_is.column_dimensions['B'].width = 20
    
    # ===== WORKSHEET 3: Balance Sheet =====
    ws_bs = wb.create_sheet("Balance Sheet")
    row = add_header(ws_bs, f"Balance Sheet - As of {end_date.strftime('%B %d, %Y')}")
    
    # Assets
    ws_bs.cell(row=row, column=1).value = "ASSETS"
    ws_bs.cell(row=row, column=1).font = subheader_font
    row += 1
    
    asset_accounts = Account.objects.filter(account_type='asset', is_deleted=False).order_by('account_code')
    total_assets = Decimal('0.00')
    
    for account in asset_accounts:
        gl_entries = GeneralLedger.objects.filter(
            account=account,
            transaction_date__lte=end_date,
            is_deleted=False
        ).aggregate(
            total_debits=Sum('debit_amount'),
            total_credits=Sum('credit_amount')
        )
        
        total_debits = gl_entries['total_debits'] or Decimal('0.00')
        total_credits = gl_entries['total_credits'] or Decimal('0.00')
        balance = total_debits - total_credits
        
        if balance != 0:
            ws_bs.cell(row=row, column=1).value = account.account_name
            ws_bs.cell(row=row, column=2).value = float(balance)
            ws_bs.cell(row=row, column=2).number_format = '#,##0.00'
            ws_bs.cell(row=row, column=2).alignment = right_align
            total_assets += balance
            row += 1
    
    ws_bs.cell(row=row, column=1).value = "Total Assets"
    ws_bs.cell(row=row, column=1).font = subheader_font
    ws_bs.cell(row=row, column=2).value = float(total_assets)
    ws_bs.cell(row=row, column=2).number_format = '#,##0.00'
    ws_bs.cell(row=row, column=2).font = subheader_font
    row += 2
    
    # Liabilities
    ws_bs.cell(row=row, column=1).value = "LIABILITIES"
    ws_bs.cell(row=row, column=1).font = subheader_font
    row += 1
    
    liability_accounts = Account.objects.filter(account_type='liability', is_deleted=False).order_by('account_code')
    total_liabilities = Decimal('0.00')
    
    for account in liability_accounts:
        gl_entries = GeneralLedger.objects.filter(
            account=account,
            transaction_date__lte=end_date,
            is_deleted=False
        ).aggregate(
            total_debits=Sum('debit_amount'),
            total_credits=Sum('credit_amount')
        )
        
        total_debits = gl_entries['total_debits'] or Decimal('0.00')
        total_credits = gl_entries['total_credits'] or Decimal('0.00')
        balance = total_credits - total_debits
        
        if balance != 0:
            ws_bs.cell(row=row, column=1).value = account.account_name
            ws_bs.cell(row=row, column=2).value = float(balance)
            ws_bs.cell(row=row, column=2).number_format = '#,##0.00'
            ws_bs.cell(row=row, column=2).alignment = right_align
            total_liabilities += balance
            row += 1
    
    ws_bs.cell(row=row, column=1).value = "Total Liabilities"
    ws_bs.cell(row=row, column=1).font = subheader_font
    ws_bs.cell(row=row, column=2).value = float(total_liabilities)
    ws_bs.cell(row=row, column=2).number_format = '#,##0.00'
    ws_bs.cell(row=row, column=2).font = subheader_font
    row += 2
    
    # Equity
    ws_bs.cell(row=row, column=1).value = "EQUITY"
    ws_bs.cell(row=row, column=1).font = subheader_font
    row += 1
    
    equity_accounts = Account.objects.filter(account_type='equity', is_deleted=False).order_by('account_code')
    total_equity = Decimal('0.00')
    
    for account in equity_accounts:
        gl_entries = GeneralLedger.objects.filter(
            account=account,
            transaction_date__lte=end_date,
            is_deleted=False
        ).aggregate(
            total_debits=Sum('debit_amount'),
            total_credits=Sum('credit_amount')
        )
        
        total_debits = gl_entries['total_debits'] or Decimal('0.00')
        total_credits = gl_entries['total_credits'] or Decimal('0.00')
        balance = total_credits - total_debits
        
        if balance != 0:
            ws_bs.cell(row=row, column=1).value = account.account_name
            ws_bs.cell(row=row, column=2).value = float(balance)
            ws_bs.cell(row=row, column=2).number_format = '#,##0.00'
            ws_bs.cell(row=row, column=2).alignment = right_align
            total_equity += balance
            row += 1
    
    # Calculate Net Income YTD (from beginning of year to end_date for Balance Sheet)
    revenue_ytd_data = GeneralLedger.objects.filter(
        account__account_type='revenue',
        transaction_date__lte=end_date,
        is_deleted=False
    ).aggregate(
        total_credits=Sum('credit_amount'),
        total_debits=Sum('debit_amount')
    )
    
    expenses_ytd_data = GeneralLedger.objects.filter(
        account__account_type='expense',
        transaction_date__lte=end_date,
        is_deleted=False
    ).aggregate(
        total_debits=Sum('debit_amount'),
        total_credits=Sum('credit_amount')
    )
    
    total_revenue_ytd = (revenue_ytd_data['total_credits'] or Decimal('0.00')) - (revenue_ytd_data['total_debits'] or Decimal('0.00'))
    total_expenses_ytd = (expenses_ytd_data['total_debits'] or Decimal('0.00')) - (expenses_ytd_data['total_credits'] or Decimal('0.00'))
    net_income_ytd = total_revenue_ytd - total_expenses_ytd
    
    # Add Net Income to Equity
    total_equity_with_income = total_equity + net_income_ytd
    
    ws_bs.cell(row=row, column=1).value = "Net Income (YTD)"
    ws_bs.cell(row=row, column=2).value = float(net_income_ytd)
    ws_bs.cell(row=row, column=2).number_format = '#,##0.00'
    ws_bs.cell(row=row, column=2).alignment = right_align
    row += 1
    
    ws_bs.cell(row=row, column=1).value = "Total Equity"
    ws_bs.cell(row=row, column=1).font = subheader_font
    ws_bs.cell(row=row, column=2).value = float(total_equity_with_income)
    ws_bs.cell(row=row, column=2).number_format = '#,##0.00'
    ws_bs.cell(row=row, column=2).font = subheader_font
    
    ws_bs.column_dimensions['A'].width = 40
    ws_bs.column_dimensions['B'].width = 20
    
    # ===== WORKSHEET 4: General Ledger Detail =====
    ws_gl = wb.create_sheet("General Ledger")
    row = add_header(ws_gl, f"General Ledger - {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}")
    
    # Headers
    headers = ['Date', 'Entry Number', 'Account', 'Description', 'Debit', 'Credit', 'Reference']
    for col, header in enumerate(headers, 1):
        cell = ws_gl.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
        ws_gl.column_dimensions[get_column_letter(col)].width = 18
    
    row += 1
    
    # Get all GL entries
    gl_entries = GeneralLedger.objects.filter(
        transaction_date__gte=start_date,
        transaction_date__lte=end_date,
        is_deleted=False
    ).select_related('account').order_by('transaction_date', 'entry_number')[:5000]  # Limit to 5000 rows
    
    for entry in gl_entries:
        ws_gl.cell(row=row, column=1).value = entry.transaction_date.strftime('%Y-%m-%d')
        ws_gl.cell(row=row, column=2).value = entry.entry_number
        ws_gl.cell(row=row, column=3).value = f"{entry.account.account_code} - {entry.account.account_name}"
        ws_gl.cell(row=row, column=4).value = entry.description[:50]  # Truncate long descriptions
        ws_gl.cell(row=row, column=5).value = float(entry.debit_amount) if entry.debit_amount else 0
        ws_gl.cell(row=row, column=5).number_format = '#,##0.00'
        ws_gl.cell(row=row, column=6).value = float(entry.credit_amount) if entry.credit_amount else 0
        ws_gl.cell(row=row, column=6).number_format = '#,##0.00'
        ws_gl.cell(row=row, column=7).value = entry.reference_number or ''
        
        for col in range(1, 8):
            ws_gl.cell(row=row, column=col).border = border
        row += 1
    
    # ===== WORKSHEET 5: Accounts Receivable =====
    ws_ar = wb.create_sheet("Accounts Receivable")
    row = add_header(ws_ar, f"Accounts Receivable Aging - As of {end_date.strftime('%B %d, %Y')}")
    
    # Headers
    headers = ['Invoice #', 'Patient', 'Due Date', 'Outstanding', 'Days Overdue', 'Aging Bucket']
    for col, header in enumerate(headers, 1):
        cell = ws_ar.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
        ws_ar.column_dimensions[get_column_letter(col)].width = 20
    
    row += 1
    
    # Get AR entries
    ar_entries = AccountsReceivable.objects.filter(
        outstanding_amount__gt=0,
        is_deleted=False
    ).select_related('invoice', 'patient').order_by('due_date')
    
    for ar in ar_entries:
        ar.update_aging()
        ws_ar.cell(row=row, column=1).value = ar.invoice.invoice_number
        ws_ar.cell(row=row, column=2).value = ar.patient.full_name
        ws_ar.cell(row=row, column=3).value = ar.due_date.strftime('%Y-%m-%d')
        ws_ar.cell(row=row, column=4).value = float(ar.outstanding_amount)
        ws_ar.cell(row=row, column=4).number_format = '#,##0.00'
        ws_ar.cell(row=row, column=5).value = ar.days_overdue
        ws_ar.cell(row=row, column=6).value = ar.get_aging_bucket_display()
        
        for col in range(1, 7):
            ws_ar.cell(row=row, column=col).border = border
        row += 1
    
    # ===== WORKSHEET 6: Trial Balance =====
    ws_tb = wb.create_sheet("Trial Balance")
    row = add_header(ws_tb, f"Trial Balance - As of {end_date.strftime('%B %d, %Y')}")
    
    # Headers
    headers = ['Account', 'Debits', 'Credits', 'Balance']
    for col, header in enumerate(headers, 1):
        cell = ws_tb.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
        ws_tb.column_dimensions[get_column_letter(col)].width = 25
    
    row += 1
    
    # Calculate trial balance
    accounts = Account.objects.filter(is_active=True, is_deleted=False).order_by('account_code')
    total_debits = Decimal('0.00')
    total_credits = Decimal('0.00')
    
    for account in accounts:
        entries = GeneralLedger.objects.filter(
            account=account,
            transaction_date__lte=end_date,
            is_deleted=False
        ).aggregate(
            total_debits=Sum('debit_amount'),
            total_credits=Sum('credit_amount')
        )
        
        debits = entries['total_debits'] or Decimal('0.00')
        credits = entries['total_credits'] or Decimal('0.00')
        
        if debits > 0 or credits > 0:
            ws_tb.cell(row=row, column=1).value = f"{account.account_code} - {account.account_name}"
            ws_tb.cell(row=row, column=2).value = float(debits)
            ws_tb.cell(row=row, column=2).number_format = '#,##0.00'
            ws_tb.cell(row=row, column=3).value = float(credits)
            ws_tb.cell(row=row, column=3).number_format = '#,##0.00'
            
            if account.account_type in ['asset', 'expense']:
                balance = debits - credits
            else:
                balance = credits - debits
            
            ws_tb.cell(row=row, column=4).value = float(balance)
            ws_tb.cell(row=row, column=4).number_format = '#,##0.00'
            
            total_debits += debits
            total_credits += credits
            
            for col in range(1, 5):
                ws_tb.cell(row=row, column=col).border = border
            row += 1
    
    # Totals
    ws_tb.cell(row=row, column=1).value = "TOTALS"
    ws_tb.cell(row=row, column=1).font = subheader_font
    ws_tb.cell(row=row, column=2).value = float(total_debits)
    ws_tb.cell(row=row, column=2).number_format = '#,##0.00'
    ws_tb.cell(row=row, column=2).font = subheader_font
    ws_tb.cell(row=row, column=3).value = float(total_credits)
    ws_tb.cell(row=row, column=3).number_format = '#,##0.00'
    ws_tb.cell(row=row, column=3).font = subheader_font
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Financial_Statement_{statement_type}_{end_date.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
@user_passes_test(is_accountant, login_url='/admin/login/')
def financial_statement_export_pdf(request):
    """Export comprehensive financial statement to PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from io import BytesIO
    
    statement_type = request.GET.get('type', 'profit_loss')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    today = timezone.now().date()
    if not start_date:
        start_date = today.replace(month=1, day=1)
    else:
        try:
            start_date = date.fromisoformat(start_date)
        except ValueError:
            start_date = today.replace(month=1, day=1)
    
    if not end_date:
        end_date = today
    else:
        try:
            end_date = date.fromisoformat(end_date)
        except ValueError:
            end_date = today
    
    # Create PDF buffer
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    story = []
    styles = getSampleStyleSheet()
    
    # Title style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Add title
    title = Paragraph(f"Financial Statement Report<br/>{start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}", title_style)
    story.append(title)
    story.append(Spacer(1, 0.3*inch))
    
    # Helper function to create table
    def create_table(data, col_widths=None):
        if col_widths is None:
            col_widths = [3*inch, 2*inch]
        table = Table(data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        return table
    
    # Section: Income Statement
    story.append(Paragraph("INCOME STATEMENT", styles['Heading2']))
    story.append(Spacer(1, 0.2*inch))
    
    # Calculate revenue
    revenue_accounts = Account.objects.filter(account_type='revenue', is_deleted=False).order_by('account_code')
    revenue_data = [['Account', 'Amount']]
    total_revenue = Decimal('0.00')
    
    for account in revenue_accounts:
        entries = GeneralLedger.objects.filter(
            account=account,
            transaction_date__gte=start_date,
            transaction_date__lte=end_date,
            is_deleted=False
        ).aggregate(
            total_credits=Sum('credit_amount'),
            total_debits=Sum('debit_amount')
        )
        
        credits = entries['total_credits'] or Decimal('0.00')
        debits = entries['total_debits'] or Decimal('0.00')
        net_revenue = credits - debits
        
        if net_revenue != 0:
            revenue_data.append([account.account_name, f"GHS {net_revenue:,.2f}"])
            total_revenue += net_revenue
    
    revenue_data.append(['<b>Total Revenue</b>', f"<b>GHS {total_revenue:,.2f}</b>"])
    story.append(create_table(revenue_data))
    story.append(Spacer(1, 0.3*inch))
    
    # Calculate expenses
    expense_accounts = Account.objects.filter(account_type='expense', is_deleted=False).order_by('account_code')
    expense_data = [['Account', 'Amount']]
    total_expenses = Decimal('0.00')
    
    for account in expense_accounts:
        entries = GeneralLedger.objects.filter(
            account=account,
            transaction_date__gte=start_date,
            transaction_date__lte=end_date,
            is_deleted=False
        ).aggregate(
            total_debits=Sum('debit_amount'),
            total_credits=Sum('credit_amount')
        )
        
        debits = entries['total_debits'] or Decimal('0.00')
        credits = entries['total_credits'] or Decimal('0.00')
        net_expense = debits - credits
        
        if net_expense != 0:
            expense_data.append([account.account_name, f"GHS {net_expense:,.2f}"])
            total_expenses += net_expense
    
    expense_data.append(['<b>Total Expenses</b>', f"<b>GHS {total_expenses:,.2f}</b>"])
    story.append(create_table(expense_data))
    story.append(Spacer(1, 0.3*inch))
    
    # Net Income
    net_income = total_revenue - total_expenses
    net_income_data = [['<b>NET INCOME</b>', f"<b>GHS {net_income:,.2f}</b>"]]
    story.append(create_table(net_income_data))
    story.append(PageBreak())
    
    # Section: Balance Sheet
    story.append(Paragraph(f"BALANCE SHEET - As of {end_date.strftime('%B %d, %Y')}", styles['Heading2']))
    story.append(Spacer(1, 0.2*inch))
    
    # Assets
    story.append(Paragraph("ASSETS", styles['Heading3']))
    asset_data = [['Account', 'Balance']]
    total_assets = Decimal('0.00')
    
    asset_accounts = Account.objects.filter(account_type='asset', is_deleted=False).order_by('account_code')
    for account in asset_accounts:
        gl_entries = GeneralLedger.objects.filter(
            account=account,
            transaction_date__lte=end_date,
            is_deleted=False
        ).aggregate(
            total_debits=Sum('debit_amount'),
            total_credits=Sum('credit_amount')
        )
        
        total_debits = gl_entries['total_debits'] or Decimal('0.00')
        total_credits = gl_entries['total_credits'] or Decimal('0.00')
        balance = total_debits - total_credits
        
        if balance != 0:
            asset_data.append([account.account_name, f"GHS {balance:,.2f}"])
            total_assets += balance
    
    asset_data.append(['<b>Total Assets</b>', f"<b>GHS {total_assets:,.2f}</b>"])
    story.append(create_table(asset_data))
    story.append(Spacer(1, 0.3*inch))
    
    # Liabilities
    story.append(Paragraph("LIABILITIES", styles['Heading3']))
    liability_data = [['Account', 'Balance']]
    total_liabilities = Decimal('0.00')
    
    liability_accounts = Account.objects.filter(account_type='liability', is_deleted=False).order_by('account_code')
    for account in liability_accounts:
        gl_entries = GeneralLedger.objects.filter(
            account=account,
            transaction_date__lte=end_date,
            is_deleted=False
        ).aggregate(
            total_debits=Sum('debit_amount'),
            total_credits=Sum('credit_amount')
        )
        
        total_debits = gl_entries['total_debits'] or Decimal('0.00')
        total_credits = gl_entries['total_credits'] or Decimal('0.00')
        balance = total_credits - total_debits
        
        if balance != 0:
            liability_data.append([account.account_name, f"GHS {balance:,.2f}"])
            total_liabilities += balance
    
    liability_data.append(['<b>Total Liabilities</b>', f"<b>GHS {total_liabilities:,.2f}</b>"])
    story.append(create_table(liability_data))
    story.append(Spacer(1, 0.3*inch))
    
    # Equity
    story.append(Paragraph("EQUITY", styles['Heading3']))
    equity_data = [['Account', 'Balance']]
    total_equity = Decimal('0.00')
    
    equity_accounts = Account.objects.filter(account_type='equity', is_deleted=False).order_by('account_code')
    for account in equity_accounts:
        gl_entries = GeneralLedger.objects.filter(
            account=account,
            transaction_date__lte=end_date,
            is_deleted=False
        ).aggregate(
            total_debits=Sum('debit_amount'),
            total_credits=Sum('credit_amount')
        )
        
        total_debits = gl_entries['total_debits'] or Decimal('0.00')
        total_credits = gl_entries['total_credits'] or Decimal('0.00')
        balance = total_credits - total_debits
        
        if balance != 0:
            equity_data.append([account.account_name, f"GHS {balance:,.2f}"])
            total_equity += balance
    
    equity_data.append(['<b>Net Income (YTD)</b>', f"GHS {net_income:,.2f}"])
    total_equity_with_income = total_equity + net_income
    equity_data.append(['<b>Total Equity</b>', f"<b>GHS {total_equity_with_income:,.2f}</b>"])
    story.append(create_table(equity_data))
    story.append(PageBreak())
    
    # Section: Accounts Receivable Aging
    story.append(Paragraph("ACCOUNTS RECEIVABLE AGING", styles['Heading2']))
    story.append(Spacer(1, 0.2*inch))
    
    ar_data = [['Invoice #', 'Patient', 'Due Date', 'Outstanding', 'Days Overdue']]
    ar_entries = AccountsReceivable.objects.filter(
        outstanding_amount__gt=0,
        is_deleted=False
    ).select_related('invoice', 'patient').order_by('due_date')[:100]  # Limit to 100 entries
    
    for ar in ar_entries:
        ar.update_aging()
        ar_data.append([
            ar.invoice.invoice_number,
            ar.patient.full_name[:30],  # Truncate long names
            ar.due_date.strftime('%Y-%m-%d'),
            f"GHS {ar.outstanding_amount:,.2f}",
            str(ar.days_overdue)
        ])
    
    story.append(create_table(ar_data, col_widths=[1.5*inch, 2*inch, 1*inch, 1.2*inch, 1*inch]))
    story.append(PageBreak())
    
    # Section: Chart of Accounts Summary
    story.append(Paragraph("CHART OF ACCOUNTS SUMMARY", styles['Heading2']))
    story.append(Spacer(1, 0.2*inch))
    
    coa_data = [['Account Code', 'Account Name', 'Type', 'Balance']]
    accounts = Account.objects.filter(is_deleted=False).order_by('account_code')[:200]  # Limit to 200 accounts
    
    for account in accounts:
        gl_entries = GeneralLedger.objects.filter(
            account=account,
            transaction_date__lte=end_date,
            is_deleted=False
        ).aggregate(
            total_debits=Sum('debit_amount'),
            total_credits=Sum('credit_amount')
        )
        
        total_debits = gl_entries['total_debits'] or Decimal('0.00')
        total_credits = gl_entries['total_credits'] or Decimal('0.00')
        
        if account.account_type in ['asset', 'expense']:
            balance = total_debits - total_credits
        else:
            balance = total_credits - total_debits
        
        coa_data.append([
            account.account_code,
            account.account_name[:40],  # Truncate long names
            account.get_account_type_display(),
            f"GHS {balance:,.2f}"
        ])
    
    story.append(create_table(coa_data, col_widths=[1.2*inch, 2.5*inch, 1*inch, 1.2*inch]))
    
    # Build PDF
    doc.build(story)
    
    # Get PDF data
    pdf = buffer.getvalue()
    buffer.close()
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    filename = f"Financial_Statement_{statement_type}_{end_date.strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(pdf)
    return response


@login_required
@user_passes_test(is_accountant, login_url='/admin/login/')
def financial_statement(request):
    """Financial statements (P&L, Balance Sheet)"""
    # Check for export requests
    export_format = request.GET.get('export')
    if export_format == 'excel':
        return financial_statement_export_excel(request)
    elif export_format == 'pdf':
        return financial_statement_export_pdf(request)
    
    statement_type = request.GET.get('type', 'profit_loss')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    today = timezone.now().date()
    if not start_date:
        start_date = today.replace(month=1, day=1)
    else:
        try:
            start_date = date.fromisoformat(start_date)
        except ValueError:
            start_date = today.replace(month=1, day=1)
    
    if not end_date:
        end_date = today
    else:
        try:
            end_date = date.fromisoformat(end_date)
        except ValueError:
            end_date = today
    
    context = {
        'statement_type': statement_type,
        'start_date': start_date,
        'end_date': end_date,
    }
    
    if statement_type == 'profit_loss':
        # P&L Statement
        revenue = GeneralLedger.objects.filter(
            account__account_type='revenue',
            transaction_date__gte=start_date,
            transaction_date__lte=end_date,
            is_deleted=False
        ).aggregate(Sum('credit_amount'))['credit_amount__sum'] or Decimal('0.00')
        
        expenses = GeneralLedger.objects.filter(
            account__account_type='expense',
            transaction_date__gte=start_date,
            transaction_date__lte=end_date,
            is_deleted=False
        ).aggregate(Sum('debit_amount'))['debit_amount__sum'] or Decimal('0.00')
        
        context.update({
            'revenue': revenue,
            'expenses': expenses,
            'net_income': revenue - expenses,
        })
    
    elif statement_type == 'balance_sheet':
        # Balance Sheet
        # Assets = Debit balance (debits - credits)
        assets_data = GeneralLedger.objects.filter(
            account__account_type='asset',
            transaction_date__lte=end_date,
            is_deleted=False
        ).aggregate(
            total_debits=Sum('debit_amount'),
            total_credits=Sum('credit_amount')
        )
        
        total_assets = (
            (assets_data['total_debits'] or Decimal('0.00')) - 
            (assets_data['total_credits'] or Decimal('0.00'))
        )
        
        # Liabilities = Credit balance (credits - debits)
        liabilities_data = GeneralLedger.objects.filter(
            account__account_type='liability',
            transaction_date__lte=end_date,
            is_deleted=False
        ).aggregate(
            total_debits=Sum('debit_amount'),
            total_credits=Sum('credit_amount')
        )
        
        total_liabilities = (
            (liabilities_data['total_credits'] or Decimal('0.00')) - 
            (liabilities_data['total_debits'] or Decimal('0.00'))
        )
        
        # Equity = Credit balance (credits - debits)
        equity_data = GeneralLedger.objects.filter(
            account__account_type='equity',
            transaction_date__lte=end_date,
            is_deleted=False
        ).aggregate(
            total_debits=Sum('debit_amount'),
            total_credits=Sum('credit_amount')
        )
        
        total_equity = (
            (equity_data['total_credits'] or Decimal('0.00')) - 
            (equity_data['total_debits'] or Decimal('0.00'))
        )
        
        # Calculate Net Income (Revenue - Expenses) to add to Retained Earnings
        revenue = GeneralLedger.objects.filter(
            account__account_type='revenue',
            transaction_date__lte=end_date,
            is_deleted=False
        ).aggregate(
            total_credits=Sum('credit_amount'),
            total_debits=Sum('debit_amount')
        )
        
        expenses = GeneralLedger.objects.filter(
            account__account_type='expense',
            transaction_date__lte=end_date,
            is_deleted=False
        ).aggregate(
            total_debits=Sum('debit_amount'),
            total_credits=Sum('credit_amount')
        )
        
        total_revenue = (
            (revenue['total_credits'] or Decimal('0.00')) - 
            (revenue['total_debits'] or Decimal('0.00'))
        )
        total_expenses = (
            (expenses['total_debits'] or Decimal('0.00')) - 
            (expenses['total_credits'] or Decimal('0.00'))
        )
        net_income = total_revenue - total_expenses
        
        # Add Net Income to Equity (Retained Earnings)
        total_equity_with_income = total_equity + net_income
        
        # Calculate asset breakdown
        cash_data = GeneralLedger.objects.filter(
            account__account_code__in=['1010', '1020', '1030', '1040'],
            transaction_date__lte=end_date,
            is_deleted=False
        ).aggregate(
            total_debits=Sum('debit_amount'),
            total_credits=Sum('credit_amount')
        )
        
        total_cash = (
            (cash_data['total_debits'] or Decimal('0.00')) - 
            (cash_data['total_credits'] or Decimal('0.00'))
        )
        
        ar_data = GeneralLedger.objects.filter(
            account__account_code='1200',
            transaction_date__lte=end_date,
            is_deleted=False
        ).aggregate(
            total_debits=Sum('debit_amount'),
            total_credits=Sum('credit_amount')
        )
        
        total_ar = (
            (ar_data['total_debits'] or Decimal('0.00')) - 
            (ar_data['total_credits'] or Decimal('0.00'))
        )
        
        context.update({
            'total_assets': total_assets,
            'total_cash': total_cash,
            'total_ar': total_ar,
            'total_liabilities': total_liabilities,
            'total_equity': total_equity,
            'net_income': net_income,
            'total_equity_with_income': total_equity_with_income,
            'total_revenue': total_revenue,
            'total_expenses': total_expenses,
        })
    
    elif statement_type == 'cashflow':
        # Cash Flow Statement
        # Get cash account codes (1010, 1020, 1030, 1040)
        cash_account_codes = ['1010', '1020', '1030', '1040']  # Cash, Card, Mobile, Bank
        
        # Operating Activities - Cash received from revenue
        cash_from_operations = GeneralLedger.objects.filter(
            account__account_code__in=cash_account_codes,
            transaction_date__gte=start_date,
            transaction_date__lte=end_date,
            is_deleted=False
        ).aggregate(
            cash_in=Sum('debit_amount'),
            cash_out=Sum('credit_amount')
        )
        
        cash_receipts = cash_from_operations['cash_in'] or Decimal('0.00')
        cash_payments = cash_from_operations['cash_out'] or Decimal('0.00')
        net_operating_cash = cash_receipts - cash_payments
        
        # Get beginning cash balance (before start_date)
        beginning_cash_entries = GeneralLedger.objects.filter(
            account__account_code__in=cash_account_codes,
            transaction_date__lt=start_date,
            is_deleted=False
        ).aggregate(
            total_debits=Sum('debit_amount'),
            total_credits=Sum('credit_amount')
        )
        
        beginning_cash = (
            (beginning_cash_entries['total_debits'] or Decimal('0.00')) - 
            (beginning_cash_entries['total_credits'] or Decimal('0.00'))
        )
        
        # Ending cash balance
        ending_cash = beginning_cash + net_operating_cash
        
        context.update({
            'cash_receipts': cash_receipts,
            'cash_payments': cash_payments,
            'net_operating_cash': net_operating_cash,
            'beginning_cash': beginning_cash,
            'ending_cash': ending_cash,
            'net_cash_flow': net_operating_cash,
        })
    
    return render(request, 'hospital/financial_statement.html', context)

