"""
Admin Interface for Automatic Attendance
"""

from django.contrib import admin
from django.utils.html import format_html
from .models_auto_attendance import StaffAttendance, AttendanceSummary


@admin.register(StaffAttendance)
class StaffAttendanceAdmin(admin.ModelAdmin):
    list_display = [
        'staff_name',
        'date',
        'check_in_display',
        'check_out_display',
        'status_badge',
        'login_method_badge',
        'shift_info',
        'late_status',
        'working_hours_display',
    ]
    
    list_filter = [
        'date',
        'status',
        'login_method',
        'is_late',
        'staff__department',
    ]
    
    search_fields = [
        'staff__first_name',
        'staff__last_name',
        'staff__employee_id',
    ]
    
    date_hierarchy = 'date'
    
    readonly_fields = [
        'first_login_time',
        'last_login_time',
        'login_count',
        'check_in_ip',
        'working_hours',
    ]
    
    fieldsets = (
        ('Staff & Date', {
            'fields': ('staff', 'date')
        }),
        ('Check In/Out', {
            'fields': (
                'check_in_time',
                'check_out_time',
                'check_in_location',
                'check_in_ip',
            )
        }),
        ('Status', {
            'fields': (
                'status',
                'is_late',
                'late_minutes',
            )
        }),
        ('Login Tracking', {
            'fields': (
                'login_method',
                'first_login_time',
                'last_login_time',
                'login_count',
            )
        }),
        ('Shift', {
            'fields': ('assigned_shift',)
        }),
        ('Notes', {
            'fields': ('notes', 'approved_by')
        }),
    )
    
    actions = ['mark_as_present', 'mark_as_absent', 'export_to_excel']
    
    def staff_name(self, obj):
        staff_name = f"{obj.staff.first_name} {obj.staff.last_name}" if hasattr(obj.staff, 'first_name') else str(obj.staff)
        return staff_name
    staff_name.short_description = 'Staff'
    
    def check_in_display(self, obj):
        if obj.check_in_time:
            return format_html(
                '<strong style="color: green;">{}</strong>',
                obj.check_in_time.strftime('%H:%M')
            )
        return format_html('<span style="color: #999;">Not checked in</span>')
    check_in_display.short_description = 'Check In'
    
    def check_out_display(self, obj):
        if obj.check_out_time:
            return format_html(
                '<strong style="color: blue;">{}</strong>',
                obj.check_out_time.strftime('%H:%M')
            )
        elif obj.check_in_time:
            return format_html('<span style="color: orange;">Still working</span>')
        return '-'
    check_out_display.short_description = 'Check Out'
    
    def status_badge(self, obj):
        colors = {
            'present': '#28a745',
            'absent': '#dc3545',
            'late': '#ffc107',
            'half_day': '#17a2b8',
            'on_leave': '#6c757d',
        }
        color = colors.get(obj.status, '#6c757d')
        return format_html(
            '<span style="background: {}; color: white; padding: 3px 10px; border-radius: 3px; font-size: 11px; font-weight: 600;">{}</span>',
            color,
            obj.get_status_display()
        )
    status_badge.short_description = 'Status'
    
    def login_method_badge(self, obj):
        icons = {
            'password': '🔑',
            'manual': '✍️',
        }
        icon = icons.get(obj.login_method, '📱')
        return format_html(
            '<span title="{}">{} {}</span>',
            obj.get_login_method_display(),
            icon,
            obj.get_login_method_display()
        )
    login_method_badge.short_description = 'Login Method'
    
    def shift_info(self, obj):
        if obj.assigned_shift:
            return format_html(
                '<small>{}</small>',
                obj.assigned_shift.get_shift_type_display()
            )
        return '-'
    shift_info.short_description = 'Shift'
    
    def late_status(self, obj):
        if obj.is_late:
            return format_html(
                '<span style="color: #dc3545;">⚠ {} min late</span>',
                obj.late_minutes
            )
        return format_html('<span style="color: #28a745;">✓ On time</span>')
    late_status.short_description = 'Punctuality'
    
    def working_hours_display(self, obj):
        hours = obj.working_hours
        if hours > 0:
            return f"{hours}h"
        elif obj.check_in_time and not obj.check_out_time:
            return "In progress..."
        return "-"
    working_hours_display.short_description = 'Hours'
    
    def mark_as_present(self, request, queryset):
        updated = queryset.update(status='present')
        self.message_user(request, f'{updated} record(s) marked as present')
    mark_as_present.short_description = 'Mark as Present'
    
    def mark_as_absent(self, request, queryset):
        updated = queryset.update(status='absent')
        self.message_user(request, f'{updated} record(s) marked as absent')
    mark_as_absent.short_description = 'Mark as Absent'
    
    def export_to_excel(self, request, queryset):
        """Export selected attendance records to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from django.http import HttpResponse
            from datetime import datetime
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Staff Attendance"
            
            # Headers
            headers = ['Staff Name', 'Date', 'Check In', 'Check Out', 'Status', 'Login Method', 'Late', 'Late Minutes', 'Working Hours']
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            for attendance in queryset.select_related('staff'):
                staff_name = f"{attendance.staff.first_name} {attendance.staff.last_name}" if hasattr(attendance.staff, 'first_name') else str(attendance.staff)
                ws.append([
                    staff_name,
                    attendance.date.strftime('%Y-%m-%d') if attendance.date else '',
                    attendance.check_in_time.strftime('%H:%M') if attendance.check_in_time else '',
                    attendance.check_out_time.strftime('%H:%M') if attendance.check_out_time else '',
                    attendance.get_status_display(),
                    attendance.get_login_method_display(),
                    'Yes' if attendance.is_late else 'No',
                    attendance.late_minutes,
                    f"{attendance.working_hours:.2f}" if attendance.working_hours else '0.00',
                ])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="attendance_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            wb.save(response)
            return response
        except ImportError:
            from django.http import HttpResponse
            return HttpResponse("Excel export requires openpyxl. Please install it: pip install openpyxl", status=500)
    export_to_excel.short_description = "📊 Export to Excel"


@admin.register(AttendanceSummary)
class AttendanceSummaryAdmin(admin.ModelAdmin):
    list_display = [
        'staff_name',
        'month_year',
        'days_present',
        'days_absent',
        'days_late',
        'total_hours_worked',
        'attendance_percentage_display',
    ]
    
    list_filter = ['year', 'month', 'staff__department']
    search_fields = ['staff__first_name', 'staff__last_name']
    
    readonly_fields = [
        'days_present',
        'days_absent',
        'days_late',
        'days_on_leave',
        'total_hours_worked',
        'attendance_percentage',
    ]
    
    actions = ['recalculate_summaries']
    
    def staff_name(self, obj):
        staff_name = f"{obj.staff.first_name} {obj.staff.last_name}" if hasattr(obj.staff, 'first_name') else str(obj.staff)
        return staff_name
    staff_name.short_description = 'Staff'
    
    def month_year(self, obj):
        months = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                  'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        return f"{months[obj.month]} {obj.year}"
    month_year.short_description = 'Period'
    
    def attendance_percentage_display(self, obj):
        percentage = obj.attendance_percentage
        
        if percentage >= 95:
            color = '#28a745'
        elif percentage >= 80:
            color = '#ffc107'
        else:
            color = '#dc3545'
        
        return format_html(
            '<strong style="color: {};">{:.1f}%</strong>',
            color,
            percentage
        )
    attendance_percentage_display.short_description = 'Attendance %'
    
    def recalculate_summaries(self, request, queryset):
        for summary in queryset:
            summary.calculate_summary()
        self.message_user(request, f'{queryset.count()} summary/ies recalculated')
    recalculate_summaries.short_description = 'Recalculate Attendance'















