"""
Views for Hospital Management System frontend
"""
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import logout as auth_logout
from django.contrib import messages
from django.core.cache import cache
from django.core.paginator import Paginator
from django.core.exceptions import ValidationError
from django.db.models import Q, Sum, F
from django.db import models, transaction, connection, IntegrityError
from django.db.transaction import TransactionManagementError
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.urls import reverse
from datetime import date, timedelta
from hospital.utils_pagination import get_pagination_html
from hospital.decorators import block_pharmacy_from_invoice_and_payment
from uuid import UUID
import csv
from collections import OrderedDict, Counter
from io import BytesIO
import logging

logger = logging.getLogger(__name__)

DEFAULT_QUICK_ACTIONS = [
    {'title': 'New Patient', 'icon': 'person-plus', 'gradient': 'linear-gradient(135deg, #3b82f6 0%, #2563eb 100%)', 'url_name': 'hospital:patient_create'},
    {'title': 'Accounting', 'icon': 'graph-up-arrow', 'gradient': 'linear-gradient(135deg, #10b981 0%, #059669 100%)', 'url_name': 'hospital:accounting_dashboard'},
    {'title': 'Procurement', 'icon': 'clipboard-check', 'gradient': 'linear-gradient(135deg, #f59e0b 0%, #d97706 100%)', 'url_name': 'hospital:admin_approval_list'},
    {'title': 'HOD Scheduling', 'icon': 'calendar-week', 'gradient': 'linear-gradient(135deg, #8b5cf6 0%, #7c3aed 100%)', 'url_name': 'hospital:hod_scheduling_dashboard'},
    {'title': 'My Schedule', 'icon': 'clock-history', 'gradient': 'linear-gradient(135deg, #ec4899 0%, #db2777 100%)', 'url_name': 'hospital:staff_schedule_dashboard'},
    {'title': 'Backups', 'icon': 'shield-check', 'gradient': 'linear-gradient(135deg, #06b6d4 0%, #0891b2 100%)', 'url_name': 'hospital:backup_dashboard'},
    {'title': 'Book Appointment', 'icon': 'calendar-plus', 'gradient': 'linear-gradient(135deg, #10b981 0%, #059669 100%)', 'url_name': 'hospital:frontdesk_appointment_create'},
    {'title': 'Patient Billing', 'icon': 'credit-card', 'gradient': 'linear-gradient(135deg, #f59e0b 0%, #d97706 100%)', 'url_name': 'hospital:cashier_patient_bills'},
    {'title': 'Pharmacy', 'icon': 'capsule', 'gradient': 'linear-gradient(135deg, #8b5cf6 0%, #7c3aed 100%)', 'url_name': 'hospital:pharmacy_dashboard'},
    {'title': 'Laboratory', 'icon': 'flask', 'gradient': 'linear-gradient(135deg, #06b6d4 0%, #0891b2 100%)', 'url_name': 'hospital:laboratory_dashboard'},
    {'title': 'Imaging', 'icon': 'camera', 'gradient': 'linear-gradient(135deg, #ec4899 0%, #db2777 100%)', 'url_name': 'hospital:imaging_dashboard'},
    {'title': 'Pre-employment / Pre-admission', 'icon': 'clipboard2-pulse', 'gradient': 'linear-gradient(135deg, #0d9488 0%, #0f766e 100%)', 'url_name': 'hospital:screening_dashboard'},
    {'title': 'Pricing', 'icon': 'tags', 'gradient': 'linear-gradient(135deg, #f97316 0%, #ea580c 100%)', 'url_name': 'hospital:pricing_dashboard'},
    {'title': 'Insurance', 'icon': 'shield-check', 'gradient': 'linear-gradient(135deg, #667eea 0%, #764ba2 100%)', 'url_name': 'hospital:insurance_management_dashboard'},
    {'title': 'HR Management', 'icon': 'people-fill', 'gradient': 'linear-gradient(135deg, #8b5cf6 0%, #7c3aed 100%)', 'url_name': 'hospital:hr_worldclass_dashboard'},
    {'title': 'Beds', 'icon': 'hospital', 'gradient': 'linear-gradient(135deg, #14b8a6 0%, #0d9488 100%)', 'url_name': 'hospital:bed_availability'},
    {'title': 'KPIs', 'icon': 'graph-up-arrow', 'gradient': 'linear-gradient(135deg, #6366f1 0%, #4f46e5 100%)', 'url_name': 'hospital:kpi_dashboard'},
    {'title': 'Search', 'icon': 'search', 'gradient': 'linear-gradient(135deg, #4b5563 0%, #374151 100%)', 'url_name': 'hospital:global_search'},
]

ROLE_SPECIFIC_QUICK_ACTIONS = {
    'accountant': [
        {'title': 'Accounting Hub', 'icon': 'calculator', 'gradient': 'linear-gradient(135deg, #0ea5e9 0%, #0284c7 100%)', 'url_name': 'hospital:accountant_dashboard'},
        {'title': 'Invoices', 'icon': 'receipt', 'gradient': 'linear-gradient(135deg, #10b981 0%, #059669 100%)', 'url_name': 'hospital:invoice_list'},
        {'title': 'Cashier Hub', 'icon': 'cash-stack', 'gradient': 'linear-gradient(135deg, #f97316 0%, #ea580c 100%)', 'url_name': 'hospital:centralized_cashier_dashboard'},
        {'title': 'Revenue Streams', 'icon': 'graph-up-arrow', 'gradient': 'linear-gradient(135deg, #6366f1 0%, #4f46e5 100%)', 'url_name': 'hospital:revenue_streams_dashboard'},
        {'title': 'Accounts Approval', 'icon': 'clipboard-check', 'gradient': 'linear-gradient(135deg, #f59e0b 0%, #d97706 100%)', 'url_name': 'hospital:accounts_approval_list'},
        {'title': 'Financial Reports', 'icon': 'bar-chart-line', 'gradient': 'linear-gradient(135deg, #14b8a6 0%, #0d9488 100%)', 'url_name': 'hospital:revenue_report'},
    ],
    'senior_account_officer': [
        {'title': 'Senior Dashboard', 'icon': 'shield-check', 'gradient': 'linear-gradient(135deg, #059669 0%, #10b981 100%)', 'url_name': 'hospital:senior_account_officer_dashboard'},
        {'title': 'Account Staff', 'icon': 'people', 'gradient': 'linear-gradient(135deg, #6366f1 0%, #8b5cf6 100%)', 'url_name': 'hospital:account_staff_list'},
        {'title': 'Accounting Dashboard', 'icon': 'calculator', 'gradient': 'linear-gradient(135deg, #0ea5e9 0%, #0284c7 100%)', 'url_name': 'hospital:accountant_comprehensive_dashboard'},
        {'title': 'Cashbook', 'icon': 'journal-text', 'gradient': 'linear-gradient(135deg, #10b981 0%, #059669 100%)', 'url_name': 'hospital:cashbook_list'},
        {'title': 'Bank Reconciliation', 'icon': 'bank', 'gradient': 'linear-gradient(135deg, #6366f1 0%, #4f46e5 100%)', 'url_name': 'hospital:bank_reconciliation_list'},
        {'title': 'Payment Vouchers', 'icon': 'receipt-cutoff', 'gradient': 'linear-gradient(135deg, #f59e0b 0%, #d97706 100%)', 'url_name': 'hospital:pv_list'},
        {'title': 'Journal Entries', 'icon': 'journal', 'gradient': 'linear-gradient(135deg, #14b8a6 0%, #0d9488 100%)', 'url_name': 'hospital:accounting_general_ledger'},
        {'title': 'Profit & Loss', 'icon': 'graph-up', 'gradient': 'linear-gradient(135deg, #8b5cf6 0%, #7c3aed 100%)', 'url_name': 'hospital:accountant_profit_loss'},
    ]
}


def build_quick_actions(role):
    actions = ROLE_SPECIFIC_QUICK_ACTIONS.get(role, DEFAULT_QUICK_ACTIONS)
    resolved = []
    for action in actions:
        url = action.get('url')
        if not url:
            url_name = action.get('url_name')
            if url_name:
                try:
                    url = reverse(url_name)
                except Exception:
                    url = action.get('fallback_url', '#')
        resolved.append({**action, 'url': url})
    return resolved
from .models import (
    Patient, Encounter, Admission, Invoice, InvoiceLine, Bed,
    Appointment, LabResult, PharmacyStock, VitalSign, Order, Prescription, Staff, Drug, LabTest,
    Department, PatientQRCode
)
# Note: In patient_create function, we import Patient as PatientModel to avoid variable shadowing
from .models_settings import HospitalSettings
from .forms import PatientForm, EncounterForm, TabularLabReportForm
from .utils import get_dashboard_stats, get_patient_demographics, get_encounter_statistics, get_dashboard_extra_stats
from .utils_roles import get_user_role, get_user_dashboard_url, user_has_role_access, is_account_or_finance
from .views_procurement import is_admin_user, is_procurement_staff
from .utils_lab_templates import (
    get_lab_result_template_type,
    is_single_value_template,
    get_allowed_detail_keys_for_save,
    get_param_display_name,
    get_param_ref_range,
    get_param_units,
    get_qualitative_flag,
    compute_value_flag,
    infer_units_from_result,
    TEMPLATE_URINE,
    TEMPLATE_STOOL,
    TEMPLATE_FBC,
    TEMPLATE_RFT,
    TEMPLATE_LFT,
    TEMPLATE_MALARIA,
    TEMPLATE_BLOOD_GROUP,
    TEMPLATE_SICKLE,
    TEMPLATE_COAGULATION,
    TEMPLATE_SEROLOGY,
    TEMPLATE_SEMEN,
    TEMPLATE_AFB,
    URINE_ORDERED_KEYS,
    STOOL_ORDERED_KEYS,
    FBC_ORDERED_KEYS,
    RFT_ORDERED_KEYS,
    LFT_ORDERED_KEYS,
    MALARIA_ORDERED_KEYS,
    BLOOD_GROUP_ORDERED_KEYS,
    SICKLE_ORDERED_KEYS,
    COAGULATION_ORDERED_KEYS,
    SEROLOGY_ORDERED_KEYS,
    SEMEN_ORDERED_KEYS,
    AFB_ORDERED_KEYS,
)
from .views_hod_shift_monitoring import is_hod
try:
    from .models_advanced import Queue
except ImportError:
    Queue = None
from .models_workflow import PatientFlowStage
# Reports
from .reports import generate_financial_report


@login_required
def dashboard(request):
    """World-Class Main Dashboard View with Role-Based Routing - OPTIMIZED FOR 300+ USERS"""
    # ===== IMMEDIATE REDIRECT - MUST BE FIRST =====
    # Check and redirect BEFORE any processing to prevent showing wrong content
    if request.user.is_authenticated:
        # IMMEDIATE ACCOUNTANT REDIRECT - Check first before anything else
        try:
            from .utils_roles import get_user_role
            user_role = get_user_role(request.user)
            if user_role == 'accountant':
                return redirect('hospital:accountant_comprehensive_dashboard')
        except Exception:
            pass  # Continue if role detection fails
        
        try:
            # Check if user is a specialist and redirect to specialist dashboard
            try:
                from .models import Staff
                from .models_specialists import SpecialistProfile
                staff = Staff.objects.filter(user=request.user, is_deleted=False).order_by('-created').first()
                if staff:
                    try:
                        specialist_profile = staff.specialist_profile
                        if specialist_profile and specialist_profile.is_active:
                            return redirect('hospital:specialist_personal_dashboard')
                    except AttributeError:
                        pass  # Not a specialist, continue
            except Exception:
                pass  # If check fails, continue to normal flow
            
            # get_user_role is imported at top of file (line 90)
            user_role = get_user_role(request.user)
            
            # FORCE REDIRECT for senior_account_officer - MUST happen immediately
            if user_role == 'senior_account_officer':
                return redirect('hospital:senior_account_officer_dashboard')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Dashboard redirect check failed: {e}")
            # Continue to main dashboard if redirect fails
    
    from decimal import Decimal
    
    # Import optional models with safe fallbacks
    Transaction = None
    PaymentReceipt = None
    try:
        from .models_accounting import Transaction, PaymentReceipt
    except (ImportError, AttributeError, Exception):
        Transaction = None
        PaymentReceipt = None
    
    ImagingStudy = None
    try:
        from .models_advanced import ImagingStudy
    except (ImportError, AttributeError, Exception):
        ImagingStudy = None
    
    # Role-based dashboard routing (with error handling to prevent blocking)
    if request.user.is_authenticated:
        try:
            user_role = get_user_role(request.user)
            
            # FORCE REDIRECT for specific roles - MUST redirect before any rendering
            if user_role == 'senior_account_officer':
                try:
                    return redirect('hospital:senior_account_officer_dashboard')
                except Exception as e:
                    logger.warning(f"Failed to redirect senior_account_officer to dashboard: {e}")
                    # If redirect fails, continue but we'll filter content below
            
            # Redirect to role-specific dashboard (with safe fallback)
            role_redirects = {
                'hr_manager': 'hospital:hr_manager_dashboard',
                'doctor': 'hospital:medical_dashboard',
                'nurse': 'hospital:nurse_dashboard',
                'midwife': 'hospital:midwife_dashboard',  # Midwives have their own dashboard
                'pharmacist': 'hospital:pharmacy_dashboard',
                'lab_technician': 'hospital:laboratory_dashboard',
                'radiologist': 'hospital:radiologist_dashboard',  # Radiologists have their own dashboard
                'receptionist': 'hospital:reception_dashboard',
                'cashier': 'hospital:cashier_dashboard',
                'accountant': 'hospital:accountant_comprehensive_dashboard',  # Fixed: Use comprehensive dashboard
                'admin': 'hospital:admin_dashboard',
                'marketing': 'hospital:marketing_dashboard',
                'inventory_stores_manager': 'hospital:inventory_stores_manager_dashboard',
                'store_manager': 'hospital:inventory_stores_manager_dashboard',
                # Procurement users should NOT land on the generic (legacy) /hms/dashboard/
                'procurement_officer': 'hospital:procurement_dashboard',
                'it': 'hospital:it_operations_dashboard',
                'it_staff': 'hospital:it_operations_dashboard',
            }
            
            redirect_url = role_redirects.get(user_role)
            if redirect_url:
                try:
                    return redirect(redirect_url)
                except Exception as e:
                    # If redirect fails, continue to main dashboard instead of blocking
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.warning(f"Failed to redirect to {redirect_url} for user {request.user.username}: {e}")
        except Exception as e:
            # If role detection fails, continue to main dashboard instead of blocking
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Role detection failed for user {request.user.username}: {e}")
            user_role = 'staff'  # Default fallback
    
    stats = get_dashboard_stats()
    demographics = get_patient_demographics()
    encounter_stats = get_encounter_statistics()
    
    # Get recent encounters for activity feed - OPTIMIZED with only() to fetch minimal fields
    recent_encounters = Encounter.objects.filter(
        is_deleted=False
    ).defer('current_activity', 'notes').select_related(
        'patient', 'provider', 'provider__user', 'location'
    ).only(
        'id', 'started_at', 'status', 'encounter_type',
        'patient__id', 'patient__first_name', 'patient__last_name', 'patient__mrn',
        'provider__id', 'provider__user__first_name', 'provider__user__last_name',
        'location__id', 'location__name'
    ).order_by('-started_at')[:10]
    
    # Convert demographics to JSON for Chart.js
    import json
    demographics_json = json.dumps({
        'gender': demographics['gender'],
        'age_groups': demographics['age_groups'],
    })
    encounter_stats_json = json.dumps(encounter_stats)
    
    # Monthly trends data for charts
    monthly_trends_json = json.dumps({
        'labels': stats.get('month_labels', []),
        'patients': stats.get('monthly_patients', []),
        'encounters': stats.get('monthly_encounters', []),
    })
    
    # Additional dashboard data
    today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
    four_hours_ago = timezone.now() - timedelta(hours=4)
    today_date = timezone.now().date()
    
    # Pending appointments - CACHED for performance
    today_date = timezone.now().date()
    cache_key_appts = f'hms:pending_appts_{today_date}'
    pending_appointments = cache.get(cache_key_appts)
    if pending_appointments is None:
        pending_appointments = Appointment.objects.filter(
            appointment_date__gte=timezone.now(),
            status__in=['scheduled', 'confirmed'],
            is_deleted=False
        ).count()
        cache.set(cache_key_appts, pending_appointments, 300)  # 5 min cache
    
    # Pending lab results - CACHED for performance
    cache_key_lab = f'hms:pending_lab_{today_date}'
    pending_lab_results = cache.get(cache_key_lab)
    if pending_lab_results is None:
        pending_lab_results = LabResult.objects.filter(
            status__in=['pending', 'in_progress'],
            is_deleted=False
        ).count()
        cache.set(cache_key_lab, pending_lab_results, 300)  # 5 min cache
    
    # Low stock alerts - CACHED for performance
    cache_key_stock = f'hms:low_stock_{today_date}'
    low_stock_items = cache.get(cache_key_stock)
    if low_stock_items is None:
        try:
            low_stock_items = PharmacyStock.objects.filter(
                is_deleted=False
            ).extra(where=['quantity_on_hand <= reorder_level']).count()
        except Exception:
            low_stock_items = 0
        cache.set(cache_key_stock, low_stock_items, 600)  # 10 min cache
    
    # Today's queue - CACHED for performance
    cache_key_queue = f'hms:queue_today_{today_date}'
    queue_today = cache.get(cache_key_queue)
    if queue_today is None:
        try:
            queue_today = Queue.objects.filter(
                checked_in_at__date=today_date,
                status='waiting',
                is_deleted=False
            ).count() if Queue else 0
        except Exception:
            queue_today = 0
        cache.set(cache_key_queue, queue_today, 300)  # 5 min cache
    
    # Encounters that need vital signs recorded
    encounters_without_vitals = []
    try:
        encounters_without_vitals = Encounter.objects.filter(
            status='active',
            is_deleted=False,
            started_at__date=timezone.now().date()
        ).exclude(
            pk__in=VitalSign.objects.filter(
                recorded_at__gte=four_hours_ago
            ).values_list('encounter_id', flat=True)
        ).select_related('patient').order_by('-started_at')[:10]
    except Exception:
        pass
    
    # Encounters with critical/abnormal vitals (for nurse dashboard)
    critical_vitals_encounters = []
    try:
        from .services.vital_signs_validator import VitalSignsValidator
        # OPTIMIZED: Prefetch vitals to avoid N+1 queries
        for enc in Encounter.objects.filter(
            status='active',
            is_deleted=False,
            vitals__recorded_at__gte=four_hours_ago
        ).distinct().select_related('patient').prefetch_related('vitals')[:10]:
            latest_vital = enc.get_latest_vitals()
            if latest_vital and latest_vital.systolic_bp:
                vital_data = {
                    'systolic_bp': latest_vital.systolic_bp,
                    'diastolic_bp': latest_vital.diastolic_bp,
                    'pulse': latest_vital.pulse,
                    'temperature': latest_vital.temperature,
                    'spo2': latest_vital.spo2,
                    'respiratory_rate': latest_vital.respiratory_rate,
                }
                try:
                    validation = VitalSignsValidator.validate_all_vitals(
                        vital_data, enc.patient.age, enc.patient.gender
                    )
                    if validation.get('overall', {}).get('status') in ['critical', 'abnormal']:
                        critical_vitals_encounters.append({
                            'encounter': enc,
                            'vital': latest_vital,
                            'validation': validation
                        })
                except Exception:
                    pass
    except Exception:
        pass
    
    # Upcoming appointments for sidebar
    upcoming_appointments = Appointment.objects.filter(
        appointment_date__gte=timezone.now(),
        status__in=['scheduled', 'confirmed'],
        is_deleted=False
    ).select_related('patient', 'provider', 'provider__user', 'department').order_by('appointment_date')[:5]
    
    # Cached dashboard extra stats (financial, department, alert counts) - 5 min TTL
    extra = get_dashboard_extra_stats(today_date)
    prescriptions_today = extra['prescriptions_today']
    active_orders = extra['active_orders']
    discharges_today = extra['discharges_today']
    staff_on_duty = extra['staff_on_duty']
    today_revenue = extra['today_revenue']
    today_payment_count = extra['today_payment_count']
    month_revenue = extra['month_revenue']
    pending_bills_count = extra['pending_bills_count']
    lab_pending = extra['lab_pending']
    lab_completed_today = extra['lab_completed_today']
    pharmacy_pending = extra['pharmacy_pending']
    pharmacy_dispensed_today = extra['pharmacy_dispensed_today']
    imaging_pending = extra['imaging_pending']
    imaging_completed_today = extra['imaging_completed_today']
    
    # ========== ALERTS & NOTIFICATIONS (World-Class) ==========
    alerts = []
    
    # Critical patients (with abnormal vitals)
    if len(critical_vitals_encounters) > 0:
        alerts.append({
            'type': 'danger',
            'icon': 'exclamation-triangle-fill',
            'title': 'Critical Patients',
            'message': f'{len(critical_vitals_encounters)} patient(s) with critical/abnormal vital signs',
            'action_url': '#',
            'action_text': 'View Patients'
        })
    
    # Expiring contracts/certificates alerts (counts from cached extra stats)
    expiring_contracts = extra.get('expiring_contracts', 0)
    expiring_certs = extra.get('expiring_certs', 0)
    if expiring_contracts > 0:
        alerts.append({
            'type': 'warning',
            'icon': 'file-earmark-text-fill',
            'title': 'Contracts Expiring',
            'message': f'{expiring_contracts} contract(s) expiring in the next 30 days',
            'action_url': '/hms/contracts/',
            'action_text': 'View Contracts'
        })
    if expiring_certs > 0:
        alerts.append({
            'type': 'warning',
            'icon': 'award-fill',
            'title': 'Certificates Expiring',
            'message': f'{expiring_certs} certificate(s) expiring in the next 60 days',
            'action_url': '/hms/certificates/list/',
            'action_text': 'View Certificates'
        })
    
    # Low stock items
    if low_stock_items > 0:
        alerts.append({
            'type': 'warning',
            'icon': 'box-seam',
            'title': 'Low Stock Alert',
            'message': f'{low_stock_items} medication(s) running low on stock',
            'action_url': '/hms/pharmacy/',
            'action_text': 'Manage Stock'
        })
    
    # Pending lab results
    if lab_pending > 5:
        alerts.append({
            'type': 'info',
            'icon': 'flask',
            'title': 'Pending Lab Tests',
            'message': f'{lab_pending} lab test(s) awaiting processing',
            'action_url': '/hms/lab/',
            'action_text': 'View Lab Queue'
        })
    
    # Patients without vitals
    if len(encounters_without_vitals) > 0:
        alerts.append({
            'type': 'warning',
            'icon': 'heart-pulse',
            'title': 'Missing Vital Signs',
            'message': f'{len(encounters_without_vitals)} patient(s) need vital signs recorded',
            'action_url': '#',
            'action_text': 'Record Vitals'
        })
    
    # Check if user is HOD
    user_is_hod = False
    if request.user.is_authenticated:
        try:
            user_is_hod = is_hod(request.user)
        except:
            pass
    
    context = {
        'stats': stats,
        'demographics': demographics,
        'demographics_json': demographics_json,
        'encounter_stats_json': encounter_stats_json,
        'monthly_trends_json': monthly_trends_json,
        'recent_encounters': recent_encounters,
        'pending_appointments': pending_appointments,
        'pending_lab_results': pending_lab_results,
        'low_stock_items': low_stock_items,
        'queue_today': queue_today,
        'upcoming_appointments': upcoming_appointments,
        'encounters_without_vitals': encounters_without_vitals,
        'critical_vitals_encounters': critical_vitals_encounters,
        'prescriptions_today': prescriptions_today,
        'active_orders': active_orders,
        'discharges_today': discharges_today,
        'is_hod': user_is_hod,
        # Financial stats
        'today_revenue': today_revenue,
        'today_payment_count': today_payment_count,
        'month_revenue': month_revenue,
        'pending_bills_count': pending_bills_count,
        # Department stats
        'lab_pending': lab_pending,
        'lab_completed_today': lab_completed_today,
        'pharmacy_pending': pharmacy_pending,
        'pharmacy_dispensed_today': pharmacy_dispensed_today,
        'imaging_pending': imaging_pending,
        'imaging_completed_today': imaging_completed_today,
        # Alerts
        'alerts': alerts,
        'staff_on_duty': staff_on_duty,
        'user_role': user_role,
        'quick_actions': build_quick_actions(user_role),
    }
    
    # Procurement Approval Notifications
    try:
        from .models_procurement import ProcurementRequest
        
        # Check if user has admin approval permission
        if request.user.has_perm('hospital.can_approve_procurement_admin') or request.user.is_superuser:
            pending_admin_approvals = ProcurementRequest.objects.filter(
                status='submitted',
                is_deleted=False
            ).count()
            context['pending_admin_approvals'] = pending_admin_approvals
        
        # Check if user has accounts approval permission
        if request.user.has_perm('hospital.can_approve_procurement_accounts') or request.user.is_superuser:
            pending_accounts_approvals = ProcurementRequest.objects.filter(
                status='admin_approved',
                is_deleted=False
            ).count()
            context['pending_accounts_approvals'] = pending_accounts_approvals
    except Exception:
        # If procurement module not available, skip
        pass
    
    return render(request, 'hospital/dashboard.html', context)


@login_required
def end_session(request):
    """
    Explicit 'End Session' action for any user.
    Logs the user out and ends their active UserSession record via signals.
    """
    # Default to main HMS dashboard (which will enforce login as needed)
    next_url = request.GET.get('next') or '/hms/'
    try:
        auth_logout(request)
    except TransactionManagementError:
        logger.warning("Broken transaction detected during logout; rolling back connection.")
        try:
            connection.rollback()
        except Exception as rollback_error:
            logger.error("Failed to rollback broken transaction on logout: %s", rollback_error)
        auth_logout(request)
    return redirect(next_url)


@login_required
def patient_list(request):
    """List all patients - OPTIMIZED for mobile performance"""
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    from datetime import datetime
    from django.core.cache import cache
    
    # Try to import LegacyPatient, but handle gracefully if it doesn't exist
    LegacyPatient = None
    legacy_table_exists = False
    try:
        from .models_legacy_patients import LegacyPatient
        # Check if the table actually exists in the database
        try:
            LegacyPatient.objects.count()  # This will fail if table doesn't exist
            legacy_table_exists = True
        except Exception:
            legacy_table_exists = False
            LegacyPatient = None
    except ImportError:
        LegacyPatient = None
    
    query = request.GET.get('q', '').strip()
    source_filter = request.GET.get('source', 'all')  # 'all', 'new', 'legacy'
    page_number = request.GET.get('page', 1)
    per_page = 25  # Reduced from 50 for faster mobile loading
    
    # Cache counts for 10 minutes to reduce database load on network devices
    cache_key = f'patient_counts_{source_filter}_{query}'
    counts = cache.get(cache_key)
    if not counts:
        new_count = Patient.objects.filter(is_deleted=False).exclude(id__isnull=True).count()
        # Safely get legacy count - handle if table doesn't exist
        legacy_count = 0
        if LegacyPatient and legacy_table_exists:
            try:
                legacy_count = LegacyPatient.objects.count()
            except Exception as e:
                logger.warning(f"Could not count legacy patients: {e}")
                legacy_count = 0
        counts = {'new': new_count, 'legacy': legacy_count, 'total': new_count + legacy_count}
        cache.set(cache_key, counts, 600)  # Cache for 10 minutes (longer for network devices)
    
    new_count = counts['new']
    legacy_count = counts['legacy']
    total_count = counts['total']
    
    # OPTIMIZED: Use database-level pagination instead of loading all into memory
    # This is MUCH faster, especially on network devices
    all_patients = []
    
    try:
        # Filter out patients with invalid IDs at database level
        django_patients_qs = Patient.objects.filter(
            is_deleted=False
        ).exclude(
            id__isnull=True
        ).select_related('primary_insurance').only(
            'id', 'first_name', 'last_name', 'middle_name', 'mrn', 'date_of_birth',
            'gender', 'phone_number', 'created', 'primary_insurance_id',
            'insurance_company',
        ).order_by('-created')
    except Exception as e:
        logger.error(f"Error loading patient list: {e}")
        django_patients_qs = Patient.objects.none()
    
    # Filter based on source selection
    if source_filter == 'new':
        # Only new Django patients - use database pagination
        if query:
            # Enhanced search: also search by full name combination
            query_parts = query.strip().split()
            search_query = Q(
                Q(first_name__icontains=query) |
                Q(last_name__icontains=query) |
                Q(middle_name__icontains=query) |
                Q(mrn__icontains=query) |
                Q(phone_number__icontains=query) |
                Q(national_id__icontains=query)
            )
            
            # If query has multiple words, search for full name combinations
            # This allows searching "John Doe" and finding patients with first_name="John" and last_name="Doe"
            if len(query_parts) >= 2:
                # Try first word(s) in first_name and remaining words in last_name
                first_part = query_parts[0]
                last_parts = ' '.join(query_parts[1:])
                search_query |= Q(
                    Q(first_name__icontains=first_part) &
                    Q(last_name__icontains=last_parts)
                )
                # Also try: first part could be in last_name, last part in first_name
                search_query |= Q(
                    Q(first_name__icontains=last_parts) &
                    Q(last_name__icontains=first_part)
                )
                # Try each word individually as well for better matching
                for part in query_parts:
                    search_query |= Q(first_name__icontains=part) | Q(last_name__icontains=part)
            
            filtered_qs = django_patients_qs.filter(search_query).distinct()
        else:
            filtered_qs = django_patients_qs
        
        # Use database pagination - don't load all into memory
        paginator = Paginator(filtered_qs, per_page)
        try:
            patients_page = paginator.page(page_number)
        except (PageNotAnInteger, EmptyPage):
            patients_page = paginator.page(1)
        
        # Only process the current page (25 patients)
        # Import validation utility
        from .utils_patient_validation import is_valid_patient_id
        
        for p in patients_page:
            # Validate patient ID - skip if invalid
            if not is_valid_patient_id(p.id):
                logger.warning(f"Skipping patient with invalid ID: MRN={p.mrn}, Name={p.full_name}, ID={p.id}")
                continue
            
            all_patients.append({
                'id': str(p.id),
                'name': p.full_name,
                'mrn': p.mrn,
                'dob': p.date_of_birth,
                'age': p.age,
                'gender': p.get_gender_display(),
                'phone': p.phone_number,
                'source': 'new',
                'initials': f"{p.first_name[0] if p.first_name else ''}{p.last_name[0] if p.last_name else ''}",
                'view_url': f"/hms/patients/{p.id}/",
                'edit_url': f"/hms/patients/{p.id}/edit/",
                'quick_visit_url': f"/hms/patients/{p.id}/quick-visit/",
                'django_patient': p,
            })
        
        # Use the paginator from database query
        pagination_html = get_pagination_html(request, patients_page, 25)
        context = {
            'page_obj': patients_page,
            'patients': all_patients,
            'query': query,
            'source': source_filter,
            'total_count': paginator.count,
            'visible_count': len(all_patients),
            'new_count': paginator.count,
            'legacy_count': 0,
            'is_paginated': patients_page.has_other_pages(),
            'page_range': paginator.get_elided_page_range(page_number, on_each_side=2, on_ends=1),
            'pagination_html': pagination_html,
        }
        return render(request, 'hospital/patient_list.html', context)
    
    elif source_filter == 'legacy':
        # Only legacy patients - use database pagination
        if not LegacyPatient or not legacy_table_exists:
            all_patients = []
            paginator = Paginator([], per_page)
            patients_page = paginator.page(1)
        else:
            try:
                legacy_qs = LegacyPatient.objects.all().only(
                    'id', 'pid', 'fname', 'lname', 'mname', 'DOB', 'sex', 'phone_cell', 'pmc_mrn'
                ).order_by('lname', 'fname')
                
                if query:
                    # Enhanced search for legacy patients: also search by full name
                    query_parts = query.strip().split()
                    search_query = Q(
                        Q(fname__icontains=query) |
                        Q(lname__icontains=query) |
                        Q(mname__icontains=query) |
                        Q(pid__icontains=query) |
                        Q(phone_cell__icontains=query) |
                        Q(pmc_mrn__icontains=query)
                    )
                    
                    # If query has multiple words, search for full name combinations
                    if len(query_parts) >= 2:
                        first_part = query_parts[0]
                        last_parts = ' '.join(query_parts[1:])
                        search_query |= Q(
                            Q(fname__icontains=first_part) &
                            Q(lname__icontains=last_parts)
                        )
                        search_query |= Q(
                            Q(fname__icontains=last_parts) &
                            Q(lname__icontains=first_part)
                        )
                    
                    legacy_qs = legacy_qs.filter(search_query).distinct()
                
                # Use database pagination
                paginator = Paginator(legacy_qs, per_page)
                try:
                    patients_page = paginator.page(page_number)
                except (PageNotAnInteger, EmptyPage):
                    patients_page = paginator.page(1)
                
                # Only process current page
                for lp in patients_page:
                    # Calculate age from DOB string
                    age_str = ''
                    try:
                        if lp.DOB:
                            dob = datetime.strptime(lp.DOB, '%Y-%m-%d').date()
                            today = datetime.today().date()
                            age = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
                            age_str = f"{age} years"
                    except:
                        age_str = 'Unknown'
                    
                    all_patients.append({
                        'id': f"legacy-{lp.id}",
                        'name': lp.full_name,
                        'mrn': lp.mrn_display,
                        'dob': lp.DOB or 'Unknown',
                        'age': age_str,
                        'gender': lp.sex or 'Unknown',
                        'phone': lp.display_phone,
                        'source': 'legacy',
                        'initials': f"{lp.fname[0] if lp.fname else ''}{lp.lname[0] if lp.lname else ''}",
                        'view_url': f"/hms/patients/legacy/{lp.id}/",
                        'edit_url': None,  # Legacy patients are read-only
                    })
            except Exception as e:
                logger.warning(f"Could not fetch legacy patients: {e}")
                all_patients = []
                paginator = Paginator([], per_page)
                patients_page = paginator.page(1)
        
        pagination_html = get_pagination_html(request, patients_page, 25)
        context = {
            'page_obj': patients_page,
            'patients': all_patients,
            'query': query,
            'source': source_filter,
            'total_count': paginator.count,
            'visible_count': len(all_patients),
            'new_count': 0,
            'legacy_count': paginator.count,
            'is_paginated': patients_page.has_other_pages(),
            'page_range': paginator.get_elided_page_range(page_number, on_each_side=2, on_ends=1),
            'pagination_html': pagination_html,
        }
        return render(request, 'hospital/patient_list.html', context)
    
    else:  # 'all' - show both (OPTIMIZED: database-level pagination)
        # NETWORK OPTIMIZATION: Use database pagination - only load current page
        # This is MUCH faster on network devices and shows ALL patients
        
        # For 'all' view, prioritize new patients (they're more relevant)
        # If no new patients, fall back to showing legacy patients
        
        # Enhanced search: also search by full name combination
        if query:
            query_parts = query.strip().split()
            search_query = Q(
                Q(first_name__icontains=query) |
                Q(last_name__icontains=query) |
                Q(middle_name__icontains=query) |
                Q(mrn__icontains=query) |
                Q(phone_number__icontains=query) |
                Q(national_id__icontains=query)
            )
            
            # If query has multiple words, search for full name combinations
            if len(query_parts) >= 2:
                first_part = query_parts[0]
                last_parts = ' '.join(query_parts[1:])
                search_query |= Q(
                    Q(first_name__icontains=first_part) &
                    Q(last_name__icontains=last_parts)
                )
                search_query |= Q(
                    Q(first_name__icontains=last_parts) &
                    Q(last_name__icontains=first_part)
                )
            
            new_filtered = django_patients_qs.filter(search_query).distinct()
        else:
            new_filtered = django_patients_qs
        
        # Use database-level pagination (Django handles this efficiently)
        paginator = Paginator(new_filtered, per_page)
        try:
            patients_page = paginator.page(page_number)
        except (PageNotAnInteger, EmptyPage):
            patients_page = paginator.page(1)
        
        # Only process the current page (25 patients) - MUCH faster!
        # Import validation utility
        from .utils_patient_validation import is_valid_patient_id
        
        for p in patients_page:
            # Validate patient ID - skip if invalid
            if not is_valid_patient_id(p.id):
                logger.warning(f"Skipping patient with invalid ID: MRN={p.mrn}, Name={p.full_name}, ID={p.id}")
                continue
            
            all_patients.append({
                'id': str(p.id),
                'name': p.full_name,
                'mrn': p.mrn,
                'dob': p.date_of_birth,
                'age': p.age,
                'gender': p.get_gender_display(),
                'phone': p.phone_number,
                'source': 'new',
                'initials': f"{p.first_name[0] if p.first_name else ''}{p.last_name[0] if p.last_name else ''}",
                'view_url': f"/hms/patients/{p.id}/",
                'edit_url': f"/hms/patients/{p.id}/edit/",
                'quick_visit_url': f"/hms/patients/{p.id}/quick-visit/",
                'django_patient': p,
            })
        
        # If no new patients found and we have legacy patients, show legacy instead
        if len(all_patients) == 0 and legacy_count > 0 and LegacyPatient and legacy_table_exists:
            # Fall back to showing legacy patients
            try:
                legacy_qs = LegacyPatient.objects.all().only(
                    'id', 'pid', 'fname', 'lname', 'mname', 'DOB', 'sex', 'phone_cell', 'pmc_mrn'
                ).order_by('lname', 'fname')
                
                if query:
                    legacy_qs = legacy_qs.filter(
                        Q(fname__icontains=query) |
                        Q(lname__icontains=query) |
                        Q(mname__icontains=query) |
                        Q(pid__icontains=query) |
                        Q(phone_cell__icontains=query) |
                        Q(pmc_mrn__icontains=query)
                    )
                
                # Use database pagination
                legacy_paginator = Paginator(legacy_qs, per_page)
                try:
                    legacy_page = legacy_paginator.page(page_number)
                except (PageNotAnInteger, EmptyPage):
                    legacy_page = legacy_paginator.page(1)
                
                # Process legacy patients
                for lp in legacy_page:
                    age_str = ''
                    try:
                        if lp.DOB:
                            dob = datetime.strptime(lp.DOB, '%Y-%m-%d').date()
                            today = datetime.today().date()
                            age = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
                            age_str = f"{age} years"
                    except:
                        age_str = 'Unknown'
                    
                    all_patients.append({
                        'id': f"legacy-{lp.id}",
                        'name': lp.full_name,
                        'mrn': lp.mrn_display,
                        'dob': lp.DOB or 'Unknown',
                        'age': age_str,
                        'gender': lp.sex or 'Unknown',
                        'phone': lp.display_phone,
                        'source': 'legacy',
                        'initials': f"{lp.fname[0] if lp.fname else ''}{lp.lname[0] if lp.lname else ''}",
                        'view_url': f"/hms/patients/legacy/{lp.id}/",
                        'edit_url': None,
                    })
                
                # Update paginator for legacy patients
                patients_page = legacy_page
                paginator = legacy_paginator
            except Exception as e:
                logger.warning(f"Could not fetch legacy patients as fallback: {e}")
        
        # Get legacy count for display (cached)
        legacy_display_count = legacy_count if legacy_table_exists else 0
        
        pagination_html = get_pagination_html(request, patients_page, 25)
        context = {
            'page_obj': patients_page,
            'patients': all_patients,
            'query': query,
            'source': source_filter,
            'total_count': paginator.count + legacy_display_count,
            'visible_count': len(all_patients),
            'new_count': new_count,
            'legacy_count': legacy_display_count,
            'is_paginated': patients_page.has_other_pages() if hasattr(patients_page, 'has_other_pages') else False,
            'page_range': paginator.get_elided_page_range(page_number, on_each_side=2, on_ends=1) if hasattr(paginator, 'get_elided_page_range') else [],
            'pagination_html': pagination_html,
        }
        
        return render(request, 'hospital/patient_list.html', context)


@login_required
def patient_create(request):
    """Create a new patient with insurance enrollment
    Uses transaction to prevent duplicate creation in concurrent environments (Docker).
    Supports ?next=/path for redirect after success (e.g. cashier patient list).
    """
    from django.db import transaction, IntegrityError

    next_url = (request.GET.get('next') or '').strip()
    use_next_redirect = bool(next_url and next_url.startswith('/') and not next_url.startswith('//'))
    
    # CRITICAL FIX: Access Patient through module import to avoid local variable shadowing
    # Do NOT import Patient locally - any local import makes Python treat it as local
    # Do NOT use globals()['Patient'] - Python still sees this as a reference to Patient
    # Instead, import the models module and access Patient through it
    import hospital.models as models_module
    PatientModel = models_module.Patient  # Access Patient through module, not direct reference
    
    # Initialize variables at function level to avoid scope issues
    encounter = None
    default_dept = None
    patient = None
    
    if request.method == 'POST':
        # CRITICAL: Check if this is an auto-save request - IGNORE IT to prevent duplicates
        is_auto_save = request.POST.get('auto_save') == 'true' or \
                      request.META.get('HTTP_X_AUTO_SAVE') == 'true'
        
        if is_auto_save:
            # Auto-save should NOT create patients - return success but don't save
            logger.warning(
                f"🚨 AUTO-SAVE BLOCKED on patient registration - "
                f"User: {request.user.username}, "
                f"IP: {request.META.get('REMOTE_ADDR', 'Unknown')}, "
                f"Timestamp: {timezone.now()}"
            )
            return JsonResponse({'status': 'ignored', 'message': 'Patient registration cannot be auto-saved'})
        
        # CRITICAL: Check for duplicate submission using session token
        submission_token = request.POST.get('submission_token')
        session_key = f'patient_submission_{submission_token}'
        
        if submission_token and request.session.get(session_key):
            # This submission was already processed - prevent duplicate
            logger.error(f"DUPLICATE SUBMISSION DETECTED! Token: {submission_token}")
            messages.error(request, 'This form was already submitted. Please do not refresh the page.')
            return redirect('hospital:patient_list')
        
        # Mark this submission as processed
        if submission_token:
            request.session[session_key] = True
            # Expire after 5 minutes
            request.session.set_expiry(300)
        
        # Check if user wants to proceed with duplicate (for family members, etc.)
        proceed_with_duplicate = request.POST.get('proceed_with_duplicate') == 'true'
        
        # CRITICAL: If user wants to proceed, create mutable POST copy with the flag
        if proceed_with_duplicate:
            from django.http import QueryDict
            # Create mutable copy of POST data
            mutable_post = request.POST.copy()
            mutable_post['proceed_with_duplicate'] = 'true'
            form = PatientForm(mutable_post)
        else:
            form = PatientForm(request.POST)
        
        # CRITICAL: Form validation MUST run first (includes duplicate checks in clean())
        # This is the FIRST line of defense against duplicates
        # If user wants to proceed with duplicate, the form's clean() will skip duplicate checks
        form_valid = form.is_valid()
        bypass_duplicate_check = False
        
        # If form is invalid, check if it's a duplicate error and user wants to proceed
        if not form_valid:
            if proceed_with_duplicate:
                # Check if the only errors are duplicate-related
                has_duplicate_error = False
                has_other_errors = False
                
                if form.non_field_errors():
                    for error in form.non_field_errors():
                        if 'duplicate' in str(error).lower():
                            has_duplicate_error = True
                        else:
                            has_other_errors = True
                
                # If only duplicate errors, clear them and manually populate cleaned_data
                if has_duplicate_error and not has_other_errors:
                    logger.info("Clearing duplicate errors - user confirmed to proceed")
                    # Clear only non-field errors (duplicate errors)
                    if '__all__' in form._errors:
                        del form._errors['__all__']
                    
                    # Manually populate cleaned_data from POST
                    if not hasattr(form, 'cleaned_data') or not form.cleaned_data:
                        form.cleaned_data = {}
                    
                    # Get all field values from POST (use mutable_post if available)
                    post_data = mutable_post if mutable_post else request.POST
                    for field_name in form.fields:
                        if field_name in post_data:
                            value = post_data.get(field_name)
                            # Handle date fields
                            if field_name == 'date_of_birth' and value:
                                try:
                                    from django.utils.dateparse import parse_date
                                    form.cleaned_data[field_name] = parse_date(value)
                                except:
                                    form.cleaned_data[field_name] = value
                            else:
                                form.cleaned_data[field_name] = value
                        elif field_name in request.FILES:
                            form.cleaned_data[field_name] = request.FILES.get(field_name)
                    
                    logger.info(f"Form data manually populated with {len(form.cleaned_data)} fields, marking form as valid")
                    # Mark form as valid by clearing all errors
                    form._errors = {}
                    form_valid = True  # Override validation result
                    bypass_duplicate_check = True  # Skip duplicate checks in transaction
                else:
                    # Other validation errors - show them
                    if form.errors:
                        logger.warning(f"Form validation failed - errors: {form.errors}")
                    return render(request, 'hospital/patient_form.html', {'form': form, 'title': 'Register New Patient'})
            else:
                # Form has validation errors (including duplicate checks from clean())
                # Log the errors for debugging
                if form.errors:
                    logger.warning(f"Form validation failed - errors: {form.errors}")
                    # Check if it's a duplicate error
                    if form.non_field_errors():
                        for error in form.non_field_errors():
                            if 'duplicate' in str(error).lower():
                                logger.error(f"DUPLICATE DETECTED IN FORM VALIDATION: {error}")
                                # Try to find existing patient for display
                                try:
                                    # Extract values from POST data or cleaned_data for error handling
                                    first_name_val = (request.POST.get('first_name') or '').strip()
                                    last_name_val = (request.POST.get('last_name') or '').strip()
                                    phone_number_val = (request.POST.get('phone_number') or '').strip()
                                    
                                    # Normalize phone for comparison
                                    def normalize_phone(phone):
                                        if not phone:
                                            return ''
                                        phone = phone.replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
                                        if phone.startswith('0') and len(phone) == 10:
                                            phone = '233' + phone[1:]
                                        elif phone.startswith('+'):
                                            phone = phone[1:]
                                        return phone
                                    
                                    normalized_phone_val = normalize_phone(phone_number_val)
                                    
                                    existing_patient = None
                                    if first_name_val and last_name_val and normalized_phone_val:
                                        existing_patient = PatientModel.objects.filter(
                                            first_name__iexact=first_name_val,
                                            last_name__iexact=last_name_val,
                                            is_deleted=False
                                        ).first()
                                except Exception as e:
                                    logger.warning(f"Error extracting patient info for duplicate display: {e}")
                                    existing_patient = None
                                
                                context = {
                                    'form': form,
                                    'title': 'Register New Patient',
                                    'duplicate_warning': True,
                                    'duplicate_reason': str(error),
                                    'existing_patient': existing_patient
                                }
                                return render(request, 'hospital/patient_form.html', context)
                # Return form with errors - DO NOT PROCEED TO SAVE
                return render(request, 'hospital/patient_form.html', {'form': form, 'title': 'Register New Patient'})
        
        # Form is valid (or was manually validated) - proceed with additional duplicate checks inside transaction
        # This is the SECOND line of defense (transaction-based with row locking)
        # Wrap in transaction to ensure atomicity and handle duplicates
        # IMPORTANT: All duplicate checks must be INSIDE the transaction to prevent race conditions
        try:
            with transaction.atomic():
                # Normalize phone number for comparison
                def normalize_phone(phone):
                    if not phone:
                        return ''
                    phone = phone.replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
                    if phone.startswith('0') and len(phone) == 10:
                        phone = '233' + phone[1:]
                    elif phone.startswith('+'):
                        phone = phone[1:]
                    return phone
                
                # Get cleaned data
                cleaned_data = form.cleaned_data
                # Safely handle None values - use or '' to ensure we always have a string before calling strip()
                first_name = (cleaned_data.get('first_name') or '').strip()
                last_name = (cleaned_data.get('last_name') or '').strip()
                date_of_birth = cleaned_data.get('date_of_birth')
                phone_number = (cleaned_data.get('phone_number') or '').strip()
                email = (cleaned_data.get('email') or '').strip()
                national_id = (cleaned_data.get('national_id') or '').strip()
                
                normalized_phone = normalize_phone(phone_number)
                
                # CRITICAL: Duplicate checks INSIDE transaction with SELECT FOR UPDATE to prevent race conditions
                # Use select_for_update() to lock rows and prevent concurrent duplicate creation
                from django.db.models import Q
                
                duplicate_found = False
                existing_patient = None
                error_message = None
                
                # Check by name + DOB + phone (most reliable match) - WITH ROW LOCKING
                # RELAXED: Only flag as duplicate if name + DOB + phone match (strong match)
                # Phone-only matches are allowed (family members can share phones)
                if first_name and last_name and normalized_phone and date_of_birth and date_of_birth != '2000-01-01':
                    # Strong match: name + DOB + phone - this is likely a duplicate
                    candidates = PatientModel.objects.select_for_update().filter(
                        first_name__iexact=first_name,
                        last_name__iexact=last_name,
                        date_of_birth=date_of_birth,
                        is_deleted=False
                    )
                    
                    for candidate in candidates:
                        if normalize_phone(candidate.phone_number) == normalized_phone:
                            existing_patient = candidate
                            duplicate_found = True
                            error_message = (
                                f'⚠️ Potential duplicate detected! A patient with the same name ({first_name} {last_name}), '
                                f'date of birth ({date_of_birth}), and phone number ({phone_number}) already exists. '
                                f'MRN: {existing_patient.mrn}. If this is a different person (e.g., family member), you can proceed.'
                            )
                            logger.warning(f"Potential duplicate patient: {first_name} {last_name} - {phone_number} - Existing MRN: {existing_patient.mrn}")
                            break
                
                # Check by email (if not already found) - WITH ROW LOCKING
                if not duplicate_found and email:
                    existing_patient = PatientModel.objects.select_for_update().filter(
                        email__iexact=email,
                        is_deleted=False
                    ).first()
                    if existing_patient:
                        duplicate_found = True
                        error_message = (
                            f'⚠️ Duplicate patient detected! A patient with email {email} already exists. '
                            f'Name: {existing_patient.full_name}, MRN: {existing_patient.mrn}'
                        )
                
                # Check by national_id (if not already found) - WITH ROW LOCKING
                if not duplicate_found and national_id:
                    existing_patient = PatientModel.objects.select_for_update().filter(
                        national_id=national_id,
                        is_deleted=False
                    ).first()
                    if existing_patient:
                        duplicate_found = True
                        error_message = (
                            f'⚠️ Duplicate patient detected! A patient with National ID {national_id} already exists. '
                            f'Name: {existing_patient.full_name}, MRN: {existing_patient.mrn}'
                        )
                
                # If duplicate found, check if user wants to proceed anyway
                if duplicate_found and existing_patient:
                    # Check if user confirmed they want to proceed (for family members, etc.)
                    # Use bypass_duplicate_check flag if set, otherwise check POST
                    proceed_anyway = bypass_duplicate_check or request.POST.get('proceed_with_duplicate') == 'true'
                    
                    if not proceed_anyway:
                        # Show warning with option to proceed
                        # Return form with duplicate warning and option to proceed
                        context = {
                            'form': form,
                            'title': 'Register New Patient',
                            'duplicate_warning': True,
                            'existing_patient': existing_patient,
                            'duplicate_reason': error_message,
                            'show_modal': True  # Show modal popup
                        }
                        return render(request, 'hospital/patient_form.html', context)
                    else:
                        # User confirmed - log the bypass and proceed
                        logger.warning(
                            f"User bypassed duplicate check for {first_name} {last_name} - "
                            f"Existing: {existing_patient.mrn}, Proceeding anyway"
                        )
                        messages.info(
                            request,
                            f'⚠️ Proceeding with registration despite potential duplicate. '
                            f'Existing patient: {existing_patient.mrn}. '
                            f'Please verify this is a different person or family member.'
                        )
                
                # FINAL SAFETY CHECK: One more check right before save (catches any edge cases)
                # This ensures no duplicate was created in the microseconds between our check and save
                # Skip this check if user confirmed they want to proceed (bypass_duplicate_check)
                # RELAXED: Only check for strong duplicates (name + DOB + phone) - not just name + phone
                if not bypass_duplicate_check and first_name and last_name and normalized_phone and date_of_birth and date_of_birth != '2000-01-01':
                    # Only check for strong match: name + DOB + phone
                    last_second_check = PatientModel.objects.select_for_update().filter(
                        first_name__iexact=first_name,
                        last_name__iexact=last_name,
                        date_of_birth=date_of_birth,
                        is_deleted=False
                    ).first()
                    
                    if last_second_check and normalize_phone(last_second_check.phone_number) == normalized_phone:
                        messages.error(
                            request,
                            f'⚠️ Duplicate patient detected! A patient with the same information was just created. '
                            f'MRN: {last_second_check.mrn}. Please check the patient list.'
                        )
                        logger.warning(f"Final check caught duplicate: {first_name} {last_name} - {phone_number} - Existing MRN: {last_second_check.mrn}")
                        return render(request, 'hospital/patient_form.html', {'form': form, 'title': 'Register New Patient'})
                
                # Create patient - this is the critical operation that must be atomic
                # The database unique constraints on mrn and national_id provide final protection
                # RELAXED: Removed redundant last-second check - we already checked above
                # The model.save() method will also perform duplicate checks as a final safety net
                
                try:
                    # CRITICAL: Log before save to track if this is called multiple times
                    logger.info(f"About to save patient: {first_name} {last_name} - {phone_number} - Token: {submission_token}")
                    # patient is already initialized at function level as None
                    patient = form.save()
                    logger.info(f"Patient saved successfully: {patient.mrn} - ID: {patient.id}")
                except IntegrityError as save_error:
                    # Catch database-level duplicate errors (unique constraint violations)
                    # Ensure patient is set to None if save failed
                    patient = None
                    error_str = str(save_error).lower()
                    if 'unique' in error_str or 'duplicate' in error_str or 'mrn' in error_str or 'national_id' in error_str:
                        # Database-level duplicate detected - this is the final safety net
                        messages.error(
                            request,
                            f'⚠️ Duplicate patient detected at database level! '
                            f'This patient may have been created by another user. Please check the patient list.'
                        )
                        logger.error(f"Database duplicate error during patient creation: {save_error}", exc_info=True)
                        return render(request, 'hospital/patient_form.html', {'form': form, 'title': 'Register New Patient'})
                    else:
                        # Some other IntegrityError occurred - ensure patient is None before re-raising
                        patient = None
                        raise
                except ValidationError as validation_error:
                    # Catch ValidationError from model.save() duplicate checks
                    # Ensure patient is set to None if save failed
                    patient = None
                    error_message = str(validation_error)
                    if 'duplicate' in error_message.lower():
                        messages.error(request, f'⚠️ {error_message}')
                        logger.error(f"ValidationError during patient creation: {validation_error}", exc_info=True)
                        return render(request, 'hospital/patient_form.html', {'form': form, 'title': 'Register New Patient'})
                    else:
                        # Re-raise if not a duplicate error - patient already set to None above
                        raise
                except Exception as save_error:
                    # Catch any other save errors
                    # Ensure patient is set to None if save failed
                    patient = None
                    logger.error(f"Error creating patient: {save_error}", exc_info=True)
                    messages.error(request, f'Error creating patient: {str(save_error)}. Please try again.')
                    return render(request, 'hospital/patient_form.html', {'form': form, 'title': 'Register New Patient'})
                # MRN is already generated in model.save() - no need to save again
                # REMOVED: This extra save() was causing duplicate patient creation
                # The model.save() method already handles MRN generation
                # Safety check: ensure patient was assigned
                if patient is None:
                    logger.error("Patient is None after form.save() - this should not happen")
                    messages.error(request, 'Error: Patient was not created properly. Please try again.')
                    return render(request, 'hospital/patient_form.html', {'form': form, 'title': 'Register New Patient'})
                
                if not patient.mrn:
                    # This should never happen, but if it does, log it
                    logger.error(f"Patient created without MRN! ID: {patient.id}")
                    # Use PatientModel to avoid variable shadowing issues
                    patient.mrn = PatientModel.generate_mrn()
                    # Use refresh_from_db to avoid triggering another save
                    patient.refresh_from_db()
                
                # QR code is created automatically by signal - no need to call ensure_qr_profile()
                # REMOVED: This was redundant and could cause issues
                # The post_save signal already creates the QR profile
                # Only call if signal didn't work (shouldn't happen)
                try:
                    if not hasattr(patient, 'qr_profile') or not patient.qr_profile:
                        logger.warning(f"QR profile not created by signal for {patient.mrn}, creating manually")
                        patient.ensure_qr_profile()
                except Exception as qr_error:
                    logger.warning(f"Failed to provision patient QR card: {qr_error}", exc_info=True)
                
                from .patient_payer import apply_patient_payer_from_form
                apply_patient_payer_from_form(request, patient, form)

                # Auto-create encounter (wrap in try-except to not break patient registration)
                # timezone is already imported at module level - do not re-import (would shadow and break insurance block above)
                from django.db import transaction
                from .models import Encounter, Department, Staff
                from .models_workflow import PatientFlowStage
                
                # encounter and default_dept already initialized at function level
                try:
                    # Get or create default department for registration
                    default_dept = Department.objects.filter(name__icontains='outpatient').first()
                    if not default_dept:
                        default_dept = Department.objects.first()
                    
                    # Do NOT set provider to current user (front desk) - registration is not a clinical encounter.
                    # Provider remains None until a doctor is assigned; prevents front desk staff appearing as "Attending Physician".
                    
                    # Check for existing encounter today to prevent duplicates
                    today = timezone.now().date()
                    existing_encounter = Encounter.objects.filter(
                        patient=patient,
                        status='active',
                        is_deleted=False,
                        started_at__date=today
                    ).order_by('-started_at').first()
                    
                    if existing_encounter:
                        encounter = existing_encounter
                        logger.info(f"Patient registration: Reusing existing encounter {encounter.id} for patient {patient.mrn}")
                    else:
                        # Create encounter only if none exists today (provider=None: no front desk as "physician")
                        encounter = Encounter.objects.create(
                            patient=patient,
                            encounter_type='outpatient',
                            status='active',
                            started_at=timezone.now(),
                            location=None,
                            provider=None,
                            chief_complaint='New patient registration',
                            notes='Auto-created during registration'
                        )
                        logger.info(f"Patient registration: Created new encounter {encounter.id} for patient {patient.mrn}")
                    
                    # Create vital signs stage in patient flow (prevent duplicates)
                    try:
                        existing_stage = PatientFlowStage.objects.filter(
                            encounter=encounter,
                            stage_type='vitals',
                            is_deleted=False
                        ).first()
                        
                        if not existing_stage:
                            PatientFlowStage.objects.create(
                                encounter=encounter,
                                stage_type='vitals',
                                status='pending'
                            )
                    except Exception as flow_error:
                        # Log but don't fail - flow stage is optional
                        logger.warning(f"Could not create patient flow stage: {flow_error}")
                        
                except Exception as encounter_error:
                    # Log error but don't fail patient registration - encounter can be created later
                    logger.error(f"Could not create encounter during patient registration: {encounter_error}", exc_info=True)
                    messages.warning(request, 'Patient registered successfully, but encounter creation failed. You can create an encounter manually.')
                
                # Transaction completed successfully - patient is saved
                
        except Exception as transaction_error:
            # If transaction fails, log and show error
            # Ensure patient is None if transaction failed (it may not have been assigned)
            patient = None
            logger.error(f"Transaction error during patient creation: {transaction_error}", exc_info=True)
            messages.error(request, f'Error creating patient: {str(transaction_error)}. Please try again.')
            return render(request, 'hospital/patient_form.html', {'form': form, 'title': 'Register New Patient'})
        
        # Do NOT clear submission token here - clear it only when we actually redirect (below).
        # That way if redirect fails, next submit will be blocked and user goes to patient list.
        
        # Operations outside transaction (non-critical, can fail without rolling back patient creation)
        # CRITICAL SAFETY CHECK: Verify patient was actually created (not a duplicate that slipped through)
        # Refresh patient from database to ensure we have latest data
        # Patient is already imported at module level - don't re-import to avoid shadowing
        from .models_advanced import SMSLog
        
        # Safety check: ensure patient was created
        if patient is None:
            logger.error("Patient is None after transaction - patient creation may have failed")
            messages.error(request, 'Error: Patient was not created properly. Please try again.')
            return redirect('hospital:patient_list')
        
        # Wrap all post-save steps in try so we ALWAYS redirect on success (never re-show form)
        try:
            try:
                patient.refresh_from_db()
            except PatientModel.DoesNotExist:
                logger.error(f"Patient {patient.id if patient else 'None'} does not exist after transaction")
                messages.error(request, 'Error: Patient was not created properly. Please try again.')
                return redirect('hospital:patient_list')
            except Exception as refresh_error:
                logger.warning(f"Error refreshing patient from database: {refresh_error}")
                # Continue anyway - patient might still exist
            
            # Send welcome SMS to new patient - CHECK FOR DUPLICATE FIRST
            logger.info(f"🔔 SMS CHECK: Patient {patient.mrn} - Phone: {patient.phone_number}")
            phone_number = (patient.phone_number or '').strip() if patient.phone_number else None
            logger.info(f"🔔 SMS CHECK: phone_number after strip: {phone_number}")
            if phone_number:
                logger.info(f"🔔 SMS CHECK: Phone number exists, attempting to send SMS to {phone_number}")
                try:
                    from .services.sms_service import sms_service
                    # timezone already imported at module level
                    from datetime import timedelta
                    
                    # ADDITIONAL SAFETY: Check if duplicate patient with same MRN exists (should never happen)
                    duplicate_check = PatientModel.objects.filter(
                        mrn=patient.mrn,
                        is_deleted=False
                    ).exclude(id=patient.id).first()
                    if duplicate_check:
                        logger.error(f"CRITICAL: Duplicate patient found with same MRN {patient.mrn}! Existing ID: {duplicate_check.id}, New ID: {patient.id}")
                        messages.warning(request, f'Warning: Another patient with MRN {patient.mrn} exists. Please verify patient list.')
                    
                    # CRITICAL: Check if SMS was already sent to this patient within the last hour
                    recent_cutoff = timezone.now() - timedelta(hours=1)
                    
                    def normalize_phone_for_sms(phone):
                        if not phone:
                            return ''
                        phone = phone.replace(' ', '').replace('-', '').replace('(', '').replace(')', '').replace('+', '')
                        if phone.startswith('0') and len(phone) == 10:
                            phone = '233' + phone[1:]
                        return phone
                    
                    normalized_sms_phone = normalize_phone_for_sms(phone_number)
                    
                    existing_sms_by_patient = SMSLog.objects.filter(
                        related_object_id=patient.id,
                        related_object_type='Patient',
                        message_type='patient_registration',
                        created__gte=recent_cutoff
                    ).exclude(status='failed').first()
                    
                    existing_sms_by_phone = None
                    if normalized_sms_phone:
                        existing_sms_by_phone = SMSLog.objects.filter(
                            recipient_phone__icontains=normalized_sms_phone[-9:],
                            message_type='patient_registration',
                            created__gte=recent_cutoff
                        ).exclude(status='failed').first()
                    
                    if existing_sms_by_patient or existing_sms_by_phone:
                        existing_sms = existing_sms_by_patient or existing_sms_by_phone
                        logger.info(f"Registration SMS already sent to {phone_number} for patient {patient.mrn} at {existing_sms.created}. Skipping duplicate SMS.")
                        messages.success(request, f'Patient registered successfully! (SMS was already sent previously)')
                    else:
                        logger.info(f"🔔 Attempting to send registration SMS to {phone_number} for patient {patient.mrn}")
                        message = (
                            f"Welcome to PrimeCare Hospital, {patient.first_name}!\n\n"
                            f"Your Medical Record Number (MRN): {patient.mrn}\n"
                            f"Please keep this number for future visits.\n\n"
                            f"Thank you for choosing us for your healthcare needs.\n\n"
                            f"PrimeCare Hospital\n"
                            f"Call us: [Hospital Contact]"
                        )
                        sms_log = sms_service.send_sms(
                            phone_number=phone_number,
                            message=message,
                            message_type='patient_registration',
                            recipient_name=patient.full_name,
                            related_object_id=patient.id,
                            related_object_type='Patient'
                        )
                        if sms_log:
                            logger.info(f"SMS log created: status={sms_log.status}, phone={sms_log.recipient_phone}")
                            if sms_log.status == 'sent':
                                messages.success(request, f'Patient registered successfully! Welcome SMS sent to {phone_number}.')
                            elif sms_log.status == 'failed':
                                error_msg = sms_log.error_message or 'Unknown error'
                                logger.warning(f"Registration SMS failed for {phone_number}: {error_msg}")
                                if sms_log.provider_response:
                                    logger.warning(f"SMS provider response: {sms_log.provider_response}")
                                messages.warning(request, f'Patient registered successfully, but SMS failed: {error_msg}')
                            else:
                                logger.warning(f"SMS status is {sms_log.status} for {phone_number}")
                                messages.warning(request, f'Patient registered successfully, but SMS status is {sms_log.status}. Check SMS logs.')
                        else:
                            logger.error(f"SMS service returned None for {phone_number}")
                            messages.warning(request, 'Patient registered successfully, but SMS service returned no result.')
                except Exception as e:
                    # Use patient.phone_number as fallback if phone_number variable not accessible
                    phone_display = phone_number if 'phone_number' in locals() else (patient.phone_number if patient and patient.phone_number else 'N/A')
                    logger.error(f"Exception sending registration SMS to {phone_display}: {str(e)}", exc_info=True)
                    messages.warning(request, f'Patient registered successfully, but SMS could not be sent: {str(e)}')
            else:
                logger.warning(f"🔔 NO PHONE NUMBER: Patient {patient.mrn} has no phone number - skipping SMS")
                messages.success(request, 'Patient registered successfully! No phone number provided for SMS.')
            
            # 🎫 QUEUE SYSTEM: Assign queue number and send SMS (runs regardless of SMS status)
            try:
                from .services.queue_service import queue_service
                from .services.queue_notification_service import queue_notification_service
                
                if encounter and default_dept:
                    queue_entry = queue_service.create_queue_entry(
                        patient=patient,
                        encounter=encounter,
                        department=default_dept,
                        assigned_doctor=None,
                        priority=3,  # 3 = Normal priority
                        notes='New patient registration'
                    )
                    queue_notification_service.send_check_in_notification(queue_entry)
                    logger.info(f"✅ Queue entry created: {queue_entry.queue_number} for {patient.full_name}")
                else:
                    logger.warning(f"Queue entry not created - encounter or department missing for {patient.full_name}")
            except Exception as e:
                logger.error(f"Error creating queue entry: {str(e)}", exc_info=True)
                # Don't fail patient creation if queue fails
            
            # Registration fee is NOT auto-added. Cashier adds it via Add Services when the new patient pays.
            # Success message - clear and prominent
            if encounter:
                messages.success(
                    request,
                    f'✅ <strong>Patient Registration Successful!</strong><br>'
                    f'Patient: <strong>{patient.full_name}</strong><br>'
                    f'Patient ID (MRN): <strong>{patient.mrn}</strong><br>'
                    f'Please record vital signs.',
                    extra_tags='html'
                )
            else:
                messages.success(
                    request,
                    f'✅ <strong>Patient Registration Successful!</strong><br>'
                    f'Patient: <strong>{patient.full_name}</strong><br>'
                    f'Patient ID (MRN): <strong>{patient.mrn}</strong>',
                    extra_tags='html'
                )

            # CRITICAL: Clear submission token to prevent reuse (run for both encounter and no-encounter)
            if submission_token:
                try:
                    session_key_clear = f'patient_submission_{submission_token}'
                    if session_key_clear in request.session:
                        del request.session[session_key_clear]
                        request.session.save()
                except Exception:
                    pass

            # CRITICAL: Always redirect after success to prevent POST resubmission and show success
            from django.http import HttpResponseRedirect
            from django.urls import reverse
            try:
                if use_next_redirect:
                    response = HttpResponseRedirect(next_url)
                    response.status_code = 303
                    return response
                if encounter:
                    vitals_url = reverse('hospital:record_vitals', args=[encounter.pk])
                    response = HttpResponseRedirect(vitals_url)
                    response.status_code = 303
                    return response
                patient_url = reverse('hospital:patient_detail', args=[patient.pk])
                response = HttpResponseRedirect(patient_url)
                response.status_code = 303
                return response
            except Exception as redirect_error:
                logger.warning(f"Redirect error: {redirect_error}", exc_info=True)
                try:
                    if use_next_redirect:
                        return redirect(next_url)
                    patient_url = reverse('hospital:patient_detail', args=[patient.pk])
                    return HttpResponseRedirect(patient_url)
                except Exception:
                    return redirect('hospital:patient_list')
        except Exception as post_save_error:
            # Any unexpected error after save: still redirect so user never sees form again
            logger.exception("Post-registration step failed; redirecting to patient anyway")
            messages.success(
                request,
                f'Patient <strong>{patient.full_name}</strong> registered (MRN: <strong>{patient.mrn}</strong>).',
                extra_tags='html'
            )
            try:
                if use_next_redirect:
                    return redirect(next_url)
                return redirect('hospital:patient_detail', pk=patient.pk)
            except Exception:
                return redirect('hospital:patient_list')
    else:
        form = PatientForm()
    
    context = {'form': form, 'title': 'Register New Patient'}
    return render(request, 'hospital/patient_form.html', context)


@login_required
def patient_outstanding_api(request, patient_id):
    """
    Single URL for patient outstanding across the app. Returns JSON.
    Use this URL everywhere you need outstanding / amount_due_after_deposit so values are always the same.
    """
    from .models import Patient
    from .services.patient_outstanding_service import get_patient_outstanding
    patient = get_object_or_404(Patient, pk=patient_id, is_deleted=False)
    from decimal import Decimal
    data = get_patient_outstanding(patient, include_deposit_breakdown=True)
    # Serialize Decimals as numbers for JSON
    return JsonResponse({
        k: float(v) if isinstance(v, Decimal) else v
        for k, v in data.items()
    })


@login_required
def patient_detail(request, pk):
    """OPTIMIZED patient detail view for fast mobile loading"""
    from .models import Encounter, VitalSign, Order, Invoice, LabResult
    from django.core.cache import cache
    from django.http import Http404
    
    # Import validation utility
    from .utils_patient_validation import is_valid_patient_id
    
    # Check if pk is invalid
    if not is_valid_patient_id(pk):
        messages.error(request, 'Invalid patient ID. Please select a valid patient from the patient list.')
        return redirect('hospital:patient_list')
    
    # Get patient with optimized query
    try:
        patient = get_object_or_404(
            Patient.objects.select_related('primary_insurance'),
            pk=pk, 
            is_deleted=False
        )
    except (ValueError, ValidationError) as e:
        # Invalid UUID format
        messages.error(request, f'Invalid patient ID format. Please select a valid patient from the patient list.')
        return redirect('hospital:patient_list')
    
    # MOBILE OPTIMIZATION: Limit initial data load
    # Use smaller limits for faster page load - users can click "View More" if needed
    ENCOUNTER_LIMIT = 10  # Reduced from 20
    VITALS_LIMIT = 10  # Reduced from 50
    RESULTS_LIMIT = 10  # Reduced from 30
    ORDERS_LIMIT = 10  # Reduced from 30
    INVOICES_LIMIT = 10  # Reduced from 20
    
    # Get encounters with optimized query (exclude front-desk registration-only from record list)
    encounters_qs = Encounter.objects.filter(
        patient=patient,
        is_deleted=False
    ).exclude(chief_complaint__iexact='New patient registration').select_related(
        'provider__user', 'location'
    ).order_by('-started_at', '-id')
    # Deduplicate: one encounter per day (same day same time = show only one; keep most recent per day)
    seen_dates = {}
    encounter_ids_dedup = []
    for enc in encounters_qs:
        d = (enc.started_at or enc.created).date() if (enc.started_at or enc.created) else None
        if d is None:
            encounter_ids_dedup.append(enc.id)
            continue
        if d not in seen_dates:
            seen_dates[d] = enc.id
            encounter_ids_dedup.append(enc.id)
    all_encounters = Encounter.objects.filter(
        pk__in=encounter_ids_dedup
    ).select_related('provider__user', 'location').order_by('-started_at', '-id')
    
    encounters = all_encounters[:ENCOUNTER_LIMIT]
    active_encounters = all_encounters.filter(status='active')[:5]
    completed_encounters = all_encounters.filter(status='completed')[:5]
    last_encounter = all_encounters.first()
    last_visit_date = last_encounter.started_at if last_encounter else None
    
    # Get vital signs (limited)
    all_vitals = VitalSign.objects.filter(
        encounter__patient=patient,
        is_deleted=False
    ).select_related('encounter').order_by('-recorded_at')[:VITALS_LIMIT]
    
    # Get orders (limited)
    all_orders_qs = Order.objects.filter(
        encounter__patient=patient,
        is_deleted=False
    ).select_related('encounter', 'requested_by__user').order_by('-requested_at')
    
    # DEDUPLICATION: Remove duplicates - keep only the most recent order per encounter+order_type
    seen_order_keys = {}
    all_orders = []
    status_priority = {'completed': 10, 'in_progress': 8, 'pending': 5, 'cancelled': 1}
    
    for order in all_orders_qs:
        key = (order.encounter_id, order.order_type)
        if key not in seen_order_keys:
            seen_order_keys[key] = order
            all_orders.append(order)
        else:
            # Keep the one with higher status or more recent
            existing = seen_order_keys[key]
            existing_time = existing.requested_at or existing.created
            current_time = order.requested_at or order.created
            existing_priority = status_priority.get(existing.status, 0)
            current_priority = status_priority.get(order.status, 0)
            
            if current_priority > existing_priority or (current_priority == existing_priority and current_time > existing_time):
                all_orders.remove(existing)
                seen_order_keys[key] = order
                all_orders.append(order)
    
    # Limit to ORDERS_LIMIT
    all_orders = sorted(all_orders, key=lambda x: x.requested_at or x.created, reverse=True)[:ORDERS_LIMIT]
    
    # Get lab results (limited)
    all_lab_results = LabResult.objects.filter(
        order__encounter__patient=patient,
        is_deleted=False
    ).select_related('test', 'order__encounter').order_by('-verified_at')[:RESULTS_LIMIT]
    
    # Get medications (limited)
    try:
        from .models import Prescription
        all_prescriptions = Prescription.objects.filter(
            order__encounter__patient=patient,
            is_deleted=False
        ).select_related('drug', 'order__encounter').order_by('-created')[:ORDERS_LIMIT]
    except:
        all_prescriptions = []
    
    # Get invoices (limited for list display)
    all_invoices = Invoice.objects.filter(
        patient=patient,
        is_deleted=False
    ).order_by('-issued_at')[:INVOICES_LIMIT]

    # Refresh unpaid invoice totals so balance reflects deposit and payments (not stale "initial before deposit")
    for inv in Invoice.objects.filter(
        patient=patient,
        is_deleted=False,
        status__in=['issued', 'partially_paid', 'overdue'],
    ):
        try:
            inv.update_totals()
        except Exception:
            pass

    # Cache expensive count queries for 2 minutes
    stats_cache_key = f'patient_stats_{pk}'
    stats = cache.get(stats_cache_key)
    if not stats:
        total_encounters = all_encounters.count()
        total_vitals = VitalSign.objects.filter(encounter__patient=patient, is_deleted=False).count()
        total_orders = Order.objects.filter(encounter__patient=patient, is_deleted=False).count()
        stats = {'encounters': total_encounters, 'vitals': total_vitals, 'orders': total_orders}
        cache.set(stats_cache_key, stats, 120)  # Cache for 2 minutes
    
    total_encounters = stats['encounters']
    total_vitals = stats['vitals']
    total_orders = stats['orders']
    total_lab_tests = LabResult.objects.filter(order__encounter__patient=patient, is_deleted=False).count()
    
    # Get latest vital signs
    latest_vitals = all_vitals.first()
    
    # Single source of truth for outstanding (same as patient_outstanding_api URL)
    from decimal import Decimal
    from .services.patient_outstanding_service import get_patient_outstanding
    outstanding_data = get_patient_outstanding(patient, include_deposit_breakdown=True)
    total_billed = outstanding_data['total_billed']
    total_paid = outstanding_data['total_paid']
    deposit_balance = outstanding_data['deposit_balance']
    total_outstanding = outstanding_data['total_outstanding']
    amount_due_after_deposit = outstanding_data['amount_due_after_deposit']
    total_deposited = outstanding_data.get('total_deposited', Decimal('0.00'))
    deposit_used = outstanding_data.get('deposit_used', Decimal('0.00'))
    
    hospital_settings = HospitalSettings.get_settings()
    prepared_by = request.user.get_full_name() or request.user.username
    
    # Ensure QR profile exists - create and generate if missing
    qr_profile = getattr(patient, 'qr_profile', None)
    if not qr_profile:
        try:
            logger.info(f"[PATIENT DETAIL] Creating QR profile for patient {patient.mrn}")
            qr_profile = patient.ensure_qr_profile()
            logger.info(f"[PATIENT DETAIL] QR profile created: {qr_profile}")
        except Exception as qr_exc:
            logger.error(f"[PATIENT DETAIL] Failed to create QR profile: {qr_exc}", exc_info=True)
            qr_profile = None
    
    # CRITICAL: If QR profile exists but image is missing, regenerate it
    if qr_profile and (not qr_profile.qr_code_image or not qr_profile.qr_code_data or not qr_profile.qr_token):
        try:
            logger.info(f"[PATIENT DETAIL] QR profile exists but image/data missing - regenerating for patient {patient.mrn}")
            qr_profile.refresh_qr(force_token=True)
            logger.info(f"[PATIENT DETAIL] QR code regenerated successfully")
        except Exception as regen_exc:
            logger.error(f"[PATIENT DETAIL] Failed to regenerate QR code: {regen_exc}", exc_info=True)
            # Try to get fresh instance
            try:
                qr_profile.refresh_from_db()
            except:
                pass
    
    user_role = get_user_role(request.user)
    # Front desk/receptionist cannot view vitals - only medical staff
    can_view_vitals = user_role not in {'receptionist', 'frontdesk'}
    
    # Payer/company for billing (insurance or corporate) – same as pharmacy "Bill to" so profile shows company name
    try:
        from .utils_billing import get_patient_payer_info
        payer_info = get_patient_payer_info(patient, None)
    except Exception:
        payer_info = {'name': 'Cash', 'is_insurance_or_corporate': False}
    
    context = {
        'patient': patient,
        'payer_info': payer_info,
        'encounters': encounters,
        'active_encounters': active_encounters,
        'completed_encounters': completed_encounters,
        'all_vitals': all_vitals,
        'latest_vitals': latest_vitals,
        'all_orders': all_orders,
        'all_lab_results': all_lab_results,
        'all_prescriptions': all_prescriptions,
        'invoices': all_invoices,
        # Statistics
        'total_encounters': total_encounters,
        'total_vitals': total_vitals,
        'total_orders': total_orders,
        'total_lab_tests': total_lab_tests,
        'total_billed': total_billed,
        'total_paid': total_paid,
        'total_outstanding': total_outstanding,
        'deposit_balance': deposit_balance,
        'total_deposited': total_deposited,
        'deposit_used': deposit_used,
        'amount_due_after_deposit': amount_due_after_deposit,
        'now': timezone.now(),
        'last_visit_date': last_visit_date,
        'hospital_settings': hospital_settings,
        'prepared_by': prepared_by,
        'qr_profile': qr_profile,
        'qr_card_url': reverse('hospital:patient_qr_card', args=[str(patient.pk)]) if patient.pk else None,
        'qr_checkin_url': reverse('hospital:patient_qr_checkin'),
        'user_role': user_role,
        'can_view_vitals': can_view_vitals,
        'latest_encounter': encounters.first() if encounters.exists() else None,
    }
    return render(request, 'hospital/patient_medical_record_sheet.html', context)


@login_required
def patient_qr_card(request, patient_pk):
    """Printable patient ID card with enhanced QR code and authentication"""
    try:
        patient = get_object_or_404(Patient, pk=patient_pk, is_deleted=False)
        
        logger.info(f"[QR CARD] Rendering card for patient: {patient.full_name} ({patient.mrn})")
        
        # Ensure QR profile exists with enhanced authentication
        qr_profile = None
        try:
            qr_profile = patient.ensure_qr_profile()
            logger.info(f"[QR CARD] QR profile found: {qr_profile}")
            
            # Force refresh if QR code is missing or token is missing (for enhanced security)
            if not qr_profile.qr_code_image or not qr_profile.qr_token or not qr_profile.qr_code_data:
                logger.info(f"[QR CARD] Refreshing QR profile for patient {patient.mrn}")
                qr_profile.refresh_qr(force_token=True)
                logger.info(f"[QR CARD] QR refreshed successfully")
        except Exception as exc:
            logger.error(f"[QR CARD] Failed to refresh QR card for patient {patient_pk}: {exc}", exc_info=True)
            qr_profile = getattr(patient, 'qr_profile', None)
            # Try to create if it doesn't exist
            if not qr_profile:
                try:
                    logger.info(f"[QR CARD] Creating new QR profile for patient {patient.mrn}")
                    qr_profile, _ = PatientQRCode.objects.get_or_create(patient=patient)
                    qr_profile.refresh_qr(force_token=True)
                    logger.info(f"[QR CARD] QR profile created successfully")
                except Exception as create_exc:
                    logger.error(f"[QR CARD] Failed to create QR profile for patient {patient_pk}: {create_exc}", exc_info=True)
                    qr_profile = None
        
        # Get hospital settings
        try:
            hospital_settings = HospitalSettings.get_settings()
        except Exception as settings_exc:
            logger.error(f"[QR CARD] Failed to get hospital settings: {settings_exc}", exc_info=True)
            hospital_settings = None
        
        context = {
            'patient': patient,
            'qr_profile': qr_profile,
            'hospital_settings': hospital_settings,
            'generated_at': timezone.now(),
        }
        
        logger.info(f"[QR CARD] Rendering template with context keys: {list(context.keys())}")
        
        return render(request, 'hospital/patient_qr_card.html', context)
        
    except Exception as e:
        logger.error(f"[QR CARD] Critical error in patient_qr_card view: {e}", exc_info=True)
        from django.http import HttpResponse
        return HttpResponse(f"Error loading patient card: {str(e)}", status=500)


@login_required
def patient_qr_verify(request):
    """Enhanced QR code verification endpoint with authentication hash validation"""
    if request.method != 'POST' and request.method != 'GET':
        return JsonResponse({'success': False, 'error': 'POST or GET method required'}, status=405)
    
    qr_data = request.POST.get('qr_data', '').strip() or request.GET.get('qr_data', '').strip()
    if not qr_data:
        return JsonResponse({'success': False, 'error': 'QR code data is required'}, status=400)
    
    logger.info(f"QR Verification attempt - Raw data: {qr_data[:100]}...")
    
    # Use the robust lookup system with enhanced authentication
    qr_profile = PatientQRCode.find_by_qr_data(qr_data)
    
    if not qr_profile:
        # Try to extract patient UUID and create QR profile
        patient_uuid = PatientQRCode.extract_patient_uuid(qr_data)
        if patient_uuid:
            patient = Patient.objects.filter(pk=patient_uuid, is_deleted=False).first()
            if patient:
                qr_profile = getattr(patient, 'qr_profile', None)
                if not qr_profile:
                    qr_profile, _ = PatientQRCode.objects.get_or_create(patient=patient)
                    qr_profile.refresh_qr(save=True)
                    logger.info(f"Created QR profile for patient {patient.mrn} during verification")
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Patient not found',
                    'verified': False,
                    'debug': {'patient_uuid': str(patient_uuid)}
                }, status=404)
        else:
            # Try parsing payload to get more info
            parsed = PatientQRCode.parse_qr_payload(qr_data)
            return JsonResponse({
                'success': False,
                'verified': False,
                'error': 'Could not identify patient from QR code',
                'debug': {
                    'scanned_data': qr_data[:100],
                    'parsed_data': parsed,
                    'hint': 'QR code should contain a patient UUID, MRN, or valid QR payload'
                }
            }, status=400)
    
    if not qr_profile:
        return JsonResponse({
            'success': False,
            'verified': False,
            'error': 'QR profile not found'
        }, status=404)
    
    # Enhanced verification with authentication hash
    is_verified, verification_message = qr_profile.verify_qr_data(qr_data)
    
    # Parse payload for additional debug info
    parsed = PatientQRCode.parse_qr_payload(qr_data)
    
    # Record scan if verified
    if is_verified:
        qr_profile.mark_scan(request.user if request.user.is_authenticated else None)
        logger.info(f"QR verified successfully for patient {qr_profile.patient.mrn}")
    
    return JsonResponse({
        'success': is_verified,
        'verified': is_verified,
        'verification_message': verification_message,
        'patient': {
            'mrn': qr_profile.patient.mrn,
            'name': qr_profile.patient.full_name,
            'uuid': str(qr_profile.patient.id),
            'age': qr_profile.patient.age,
            'gender': qr_profile.patient.get_gender_display(),
            'blood_type': qr_profile.patient.blood_type or 'Not specified'
        },
        'qr_profile': {
            'has_qr_data': bool(qr_profile.qr_code_data),
            'is_active': qr_profile.is_active,
            'scan_count': qr_profile.scan_count,
            'last_scanned_at': qr_profile.last_scanned_at.isoformat() if qr_profile.last_scanned_at else None,
            'last_generated_at': qr_profile.last_generated_at.isoformat() if qr_profile.last_generated_at else None
        },
        'parsed': parsed,
        'error': None if is_verified else verification_message,
        'timestamp': timezone.now().isoformat()
    })


@login_required
def patient_qr_checkin(request):
    """Receptionist QR scanning console for instant visit creation"""
    departments = Department.objects.filter(is_deleted=False).order_by('name')
    default_department = departments.filter(name__icontains='outpatient').first() or departments.first()
    
    # Get available doctors for assignment - more inclusive query
    available_doctors = Staff.objects.filter(
        profession='doctor',
        is_deleted=False
    ).select_related('user', 'department').filter(
        user__isnull=False,
        user__is_active=True
    ).order_by('is_active', 'user__first_name', 'user__last_name')
    
    context = {
        'departments': departments,
        'default_department': default_department,
        'hospital_settings': HospitalSettings.get_settings(),
        'available_doctors': available_doctors,
    }
    return render(request, 'hospital/patient_qr_checkin.html', context)


@login_required
def patient_qr_checkin_api(request):
    """AJAX endpoint triggered by QR scanner to auto-create visits"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=405)
    
    qr_data = request.POST.get('qr_data', '').strip()
    if not qr_data:
        return JsonResponse({'success': False, 'error': 'QR code data is required.'}, status=400)
    
    # Log the raw QR data for debugging
    logger.info(f"QR Check-in attempt - Raw data: {qr_data[:100]}...")
    
    # SIMPLIFIED APPROACH: Direct database lookup with multiple methods
    patient = None
    qr_profile = None
    
    # Method 1: Find QR profile by scanning database directly
    qr_profile = PatientQRCode.find_by_qr_data(qr_data)
    
    if qr_profile:
        patient = qr_profile.patient
        logger.info(f"QR profile found for patient {patient.mrn} via database lookup")
    else:
        # Method 2: Extract patient UUID and find patient directly
        patient_uuid = PatientQRCode.extract_patient_uuid(qr_data)
        if patient_uuid:
            patient = Patient.objects.filter(pk=patient_uuid, is_deleted=False).select_related('primary_insurance').first()
            if patient:
                logger.info(f"Patient found by UUID extraction: {patient.mrn}")
                # Create QR profile if missing
                qr_profile, _ = PatientQRCode.objects.get_or_create(patient=patient)
                if not qr_profile.qr_code_data:
                    qr_profile.refresh_qr(save=True)
                    logger.info(f"Created/refreshed QR profile for patient {patient.mrn}")
        
        # Method 3: Try MRN lookup (for manual entry like "PMC2026000031")
        if not patient:
            parsed = PatientQRCode.parse_qr_payload(qr_data)
            mrn = parsed.get('mrn')
            
            # Also try direct MRN match if qr_data looks like an MRN
            if not mrn:
                # Check if qr_data itself is an MRN pattern
                import re
                mrn_pattern = re.compile(r'PMC\d{10}', re.IGNORECASE)
                mrn_match = mrn_pattern.search(qr_data)
                if mrn_match:
                    mrn = mrn_match.group().upper()
            
            if mrn:
                patient = Patient.objects.filter(mrn__iexact=mrn, is_deleted=False).select_related('primary_insurance').first()
                if patient:
                    logger.info(f"Patient found by MRN lookup: {patient.mrn}")
                    # Create QR profile if missing
                    qr_profile, _ = PatientQRCode.objects.get_or_create(patient=patient)
                    if not qr_profile.qr_code_data:
                        qr_profile.refresh_qr(save=True)
                    logger.info(f"Created/refreshed QR profile for patient {patient.mrn}")
        
        # Method 4: Try finding by partial MRN (last resort for manual entry)
        if not patient:
            # If qr_data is short and looks like it might be a partial identifier
            qr_data_clean = qr_data.strip()
            if len(qr_data_clean) > 3 and len(qr_data_clean) < 20:
                # Try as partial MRN
                patient = Patient.objects.filter(mrn__icontains=qr_data_clean, is_deleted=False).select_related('primary_insurance').first()
                if patient:
                    logger.info(f"Patient found by partial MRN match: {patient.mrn}")
                    qr_profile, _ = PatientQRCode.objects.get_or_create(patient=patient)
                    if not qr_profile.qr_code_data:
                        qr_profile.refresh_qr(save=True)
        
        # If still no patient found, return error
        if not patient:
            logger.warning(f"QR check-in - Could not identify patient from: {qr_data[:100]}")
            parsed = PatientQRCode.parse_qr_payload(qr_data)
            return JsonResponse({
                'success': False, 
                'error': 'Patient not found. Please check the QR code or MRN and try again.',
                'debug': {
                    'received_data': qr_data[:100],
                    'parsed': parsed,
                    'hint': 'QR code should contain a patient UUID or MRN (e.g., PMC2026000031)'
                }
            }, status=404)
    
    # Ensure we have both patient and qr_profile
    if patient and not qr_profile:
        qr_profile, _ = PatientQRCode.objects.get_or_create(patient=patient)
        if not qr_profile.qr_code_data:
            qr_profile.refresh_qr(save=True)
    
    # If we found the patient, verification always passes (we found them by UUID/MRN)
    # This ensures QR codes always work even if format is slightly different
    if qr_profile and patient:
        is_verified = True
        verification_message = f"Patient found: {patient.mrn}"
        # Update QR data if it doesn't match (for future scans)
        if not qr_profile.qr_code_data or qr_profile.qr_code_data != str(patient.id):
            logger.info(f"Updating QR data for patient {patient.mrn}")
            qr_profile.refresh_qr(save=True)
        # Mark the scan
        qr_profile.mark_scan(request.user)
        logger.info(f"QR check-in successful: {verification_message}")
    else:
        # This should not happen due to checks above, but just in case
        logger.error(f"QR check-in - Patient or QR profile missing after lookup")
        return JsonResponse({
            'success': False,
            'error': 'Patient or QR profile not found after lookup.',
            'debug': {'qr_data': qr_data[:100]}
        }, status=500)
    
    encounter_type = request.POST.get('encounter_type', 'outpatient')
    chief_complaint = request.POST.get('chief_complaint', '').strip() or 'QR check-in at reception'
    assigned_doctor_id = request.POST.get('assigned_doctor', '').strip()
    department_id = request.POST.get('department_id')
    
    department = None
    if department_id:
        try:
            department = Department.objects.get(pk=department_id, is_deleted=False)
        except (Department.DoesNotExist, ValueError):
            department = None
    
    if not department:
        department = Department.objects.filter(name__icontains='outpatient', is_deleted=False).first()
    if not department:
        department = Department.objects.filter(is_deleted=False).first()
    if not department:
        return JsonResponse({'success': False, 'error': 'No active departments configured.'}, status=400)
    
    # Get assigned doctor if provided
    assigned_doctor_staff = None
    assigned_doctor_user = None
    if assigned_doctor_id:
        try:
            assigned_doctor_staff = Staff.objects.get(
                pk=assigned_doctor_id,
                profession='doctor',
                is_active=True,
                is_deleted=False
            )
            assigned_doctor_user = assigned_doctor_staff.user if assigned_doctor_staff else None
        except (Staff.DoesNotExist, ValueError):
            logger.warning(f"QR check-in: Assigned doctor {assigned_doctor_id} not found")
    
    from .models_queue import QueueEntry
    today = timezone.now().date()
    active_queue = QueueEntry.objects.filter(
        patient=patient,
        queue_date=today,
        status__in=['checked_in', 'called', 'vitals_completed', 'in_progress'],
        is_deleted=False
    ).select_related('department', 'encounter').order_by('-created').first()
    
    if active_queue:
        queue_position = None
        try:
            from .services.queue_service import queue_service
            queue_position = queue_service.get_position_in_queue(active_queue)
        except Exception:
            queue_position = None
        qr_profile.mark_scan(request.user)
        return JsonResponse({
            'success': True,
            'already_checked_in': True,
            'message': f'{patient.full_name} is already in queue (ticket {active_queue.display_ticket_number}).',
            'patient': {
                'name': patient.full_name,
                'mrn': patient.mrn,
            },
            'queue': {
                'number': active_queue.display_ticket_number,
                'status': active_queue.get_status_display(),
                'position': queue_position,
                'department': active_queue.department.name if active_queue.department else '',
            },
            'encounter': {
                'id': str(active_queue.encounter_id) if active_queue.encounter_id else None,
                'type': active_queue.encounter.get_encounter_type_display() if active_queue.encounter else encounter_type,
            }
        })
    
    current_staff = assigned_doctor_staff or getattr(request.user, 'staff', None)
    encounter = None
    encounter_created = False
    
    try:
        with transaction.atomic():
            # Check for existing encounter today - use select_for_update to prevent race conditions
            existing_encounter = Encounter.objects.filter(
                patient=patient,
                status='active',
                is_deleted=False,
                started_at__date=today
            ).select_for_update().order_by('-started_at').first()
            
            if existing_encounter:
                encounter = existing_encounter
                logger.info(f"QR check-in: Reusing existing encounter {encounter.id} for patient {patient.mrn}")
            else:
                # Double-check with a broader filter to prevent duplicates
                recent_encounter = Encounter.objects.filter(
                    patient=patient,
                    status='active',
                    is_deleted=False,
                    started_at__gte=timezone.now() - timedelta(hours=1)
                ).select_for_update().order_by('-started_at').first()
                
                if recent_encounter:
                    encounter = recent_encounter
                    logger.info(f"QR check-in: Reusing recent encounter {encounter.id} for patient {patient.mrn}")
                else:
                    encounter = Encounter.objects.create(
                        patient=patient,
                        encounter_type=encounter_type,
                        status='active',
                        started_at=timezone.now(),
                        provider=assigned_doctor_staff or current_staff,
                        chief_complaint=chief_complaint,
                        notes=f'QR check-in by {request.user.get_full_name() or request.user.username}'
                    )
                    encounter_created = True
                    logger.info(f"QR check-in: Created new encounter {encounter.id} for patient {patient.mrn}")
                    
                    # Send notification to assigned doctor if doctor was assigned
                    if assigned_doctor_staff and assigned_doctor_staff.user:
                        try:
                            from .models import Notification
                            Notification.objects.create(
                                recipient=assigned_doctor_staff.user,
                                notification_type='order_urgent',
                                title='New Patient Assigned',
                                message=f'Patient {patient.full_name} (MRN: {patient.mrn}) has been assigned to you via QR check-in. Chief Complaint: {chief_complaint[:100]}',
                                related_object_id=encounter.id,
                                related_object_type='Encounter'
                            )
                            logger.info(f"QR check-in: Notification sent to doctor {assigned_doctor_staff.user.username}")
                        except Exception as notif_error:
                            logger.warning(f"QR check-in: Failed to send notification to doctor: {notif_error}")
                # Initialize patient flow stage (prevent duplicates)
                try:
                    existing_stage = PatientFlowStage.objects.filter(
                        encounter=encounter,
                        stage_type='vitals',
                        is_deleted=False
                    ).first()
                    
                    if not existing_stage:
                        PatientFlowStage.objects.create(
                            encounter=encounter,
                            stage_type='vitals',
                            status='pending'
                        )
                except Exception:
                    pass
            
            from .services.queue_service import queue_service
            from .services.queue_notification_service import queue_notification_service
            
            queue_entry = queue_service.create_queue_entry(
                patient=patient,
                encounter=encounter,
                department=department,
                assigned_doctor=assigned_doctor_user or (current_staff.user if current_staff else None),
                priority=1 if encounter_type == 'emergency' else 3,
                notes=chief_complaint
            )
            queue_position = queue_service.get_position_in_queue(queue_entry)
            try:
                queue_notification_service.send_check_in_notification(queue_entry)
            except Exception as notify_error:
                logger.warning(f"QR check-in notification failed: {notify_error}")
    except Exception as exc:
        logger.error(f"QR check-in failed for patient {patient.pk}: {exc}", exc_info=True)
        return JsonResponse({'success': False, 'error': 'Unable to complete QR check-in. Please process manually.'}, status=500)
    
    qr_profile.mark_scan(request.user)
    
    return JsonResponse({
        'success': True,
        'message': f'{patient.full_name} checked in successfully.',
        'already_checked_in': False,
        'patient': {
            'name': patient.full_name,
            'mrn': patient.mrn,
        },
        'encounter': {
            'id': str(encounter.id),
            'type': encounter.get_encounter_type_display(),
            'created': encounter_created,
            'started_at': timezone.localtime(encounter.started_at).strftime('%Y-%m-%d %H:%M'),
        },
        'queue': {
            'number': queue_entry.display_ticket_number,
            'position': queue_position,
            'status': queue_entry.get_status_display(),
            'estimated_wait': queue_entry.estimated_wait_minutes,
            'department': department.name if department else '',
        }
    })


@login_required
def patient_edit(request, pk):
    """Edit patient"""
    patient = get_object_or_404(
        Patient.objects.select_related('primary_insurance'),
        pk=pk,
        is_deleted=False,
    )
    
    if request.method == 'POST':
        # Check if user wants to proceed with duplicate (for family members, etc.)
        proceed_with_duplicate = request.POST.get('proceed_with_duplicate') == 'true'
        
        # Create mutable POST data if needed
        if proceed_with_duplicate:
            from django.http import QueryDict
            mutable_post = request.POST.copy()
            mutable_post['proceed_with_duplicate'] = 'true'
            form = PatientForm(mutable_post, instance=patient)
        else:
            form = PatientForm(request.POST, instance=patient)
        
        if form.is_valid():
            try:
                updated_patient = form.save()
                from .patient_payer import apply_patient_payer_from_form
                apply_patient_payer_from_form(request, updated_patient, form)
                updated_patient.refresh_from_db()
                messages.success(
                    request,
                    f'Patient {updated_patient.full_name} updated successfully.',
                )
                return redirect('hospital:patient_detail', pk=updated_patient.pk)
            except Exception as e:
                logger.error(f"Error saving patient {patient.mrn} during edit: {e}", exc_info=True)
                messages.error(request, f'Error updating patient: {str(e)}. Please try again.')
        else:
            # Log form errors for debugging
            logger.warning(f"Patient edit form validation failed for {patient.mrn}: {form.errors}")
            if form.non_field_errors():
                messages.warning(request, 'Please review the form errors below.')
            # Show field errors if any
            for field, errors in form.errors.items():
                for error in errors:
                    if field == '__all__':
                        messages.error(request, f'Error: {error}')
                    else:
                        messages.error(request, f'{field}: {error}')
    else:
        form = PatientForm(instance=patient)
    
    context = {
        'form': form,
        'patient': patient,
        'title': 'Edit Patient',
    }
    return render(request, 'hospital/patient_form.html', context)


@login_required
def encounter_list(request):
    """List all encounters with deduplication"""
    status_filter = request.GET.get('status', '')
    type_filter = request.GET.get('type', '')
    query = request.GET.get('q', '')
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    
    # Parse date range (YYYY-MM-DD)
    from datetime import datetime
    date_from_obj = None
    date_to_obj = None
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        except ValueError:
            date_from = ''
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
        except ValueError:
            date_to = ''
    
    # CRITICAL: Deduplicate encounters - keep only most recent per patient per day
    # Use DISTINCT ON approach via raw SQL since UUID fields don't support MAX()
    from django.db import connection
    from django.db.models import Q
    
    # Build WHERE clause based on filters
    where_clauses = ["e.is_deleted = false"]
    params = []
    
    if status_filter:
        where_clauses.append("e.status = %s")
        params.append(status_filter)
    
    if type_filter:
        where_clauses.append("e.encounter_type = %s")
        params.append(type_filter)
    
    if date_from_obj:
        where_clauses.append("DATE(COALESCE(e.started_at, e.created)) >= %s")
        params.append(date_from_obj)
    
    if date_to_obj:
        where_clauses.append("DATE(COALESCE(e.started_at, e.created)) <= %s")
        params.append(date_to_obj)
    
    if query:
        where_clauses.append("""
            (p.first_name ILIKE %s OR p.last_name ILIKE %s OR p.mrn ILIKE %s)
        """)
        query_param = f"%{query}%"
        params.extend([query_param, query_param, query_param])
    
    where_sql = " AND ".join(where_clauses)
    
    # Get most recent encounter ID per patient per day using DISTINCT ON (PostgreSQL)
    # This ensures only ONE encounter per patient per day (highest ID wins ties)
    # Use COALESCE to handle NULL started_at by falling back to created timestamp
    try:
        with connection.cursor() as cursor:
            # Use parameterized query to prevent SQL injection
            sql_query = f"""
                SELECT DISTINCT ON (e.patient_id, DATE(COALESCE(e.started_at, e.created))) e.id
                FROM hospital_encounter e
                INNER JOIN hospital_patient p ON p.id = e.patient_id
                WHERE {where_sql}
                ORDER BY e.patient_id, DATE(COALESCE(e.started_at, e.created)), e.id DESC
            """
            # Get full encounter data directly from SQL to avoid Django ORM MAX() optimization issues
            full_sql_query = f"""
                SELECT DISTINCT ON (e.patient_id, DATE(COALESCE(e.started_at, e.created))) 
                    e.id, e.patient_id, e.provider_id, e.location_id, e.encounter_type, 
                    e.status, e.started_at, e.ended_at, e.created, e.modified
                FROM hospital_encounter e
                INNER JOIN hospital_patient p ON p.id = e.patient_id
                WHERE {where_sql}
                ORDER BY e.patient_id, DATE(COALESCE(e.started_at, e.created)), e.id DESC
            """
            cursor.execute(full_sql_query, params)
            latest_ids = [row[0] for row in cursor.fetchall()]  # Get UUID IDs
        
        # If no IDs found, return empty queryset
        if not latest_ids:
            encounters = Encounter.objects.none()
        else:
            # Convert UUIDs to strings for SQL IN clause
            id_strings = [str(uuid_id) for uuid_id in latest_ids]
            id_placeholders = ','.join(['%s'] * len(id_strings))
            
            # Use raw SQL query to completely bypass Django ORM optimizations
            # This avoids any MAX() function calls on UUID fields
            raw_sql = f"""
                SELECT e.*
                FROM hospital_encounter e
                WHERE e.id::text IN ({id_placeholders})
                  AND e.is_deleted = false
                ORDER BY e.started_at DESC NULLS LAST, e.id DESC
            """
            encounters = Encounter.objects.raw(raw_sql, id_strings)
            
            # Convert RawQuerySet to a list, then create a new queryset from IDs
            # This is necessary because raw() returns RawQuerySet which doesn't support pagination
            encounter_ids = [enc.id for enc in encounters]
            if encounter_ids:
                # Create a fresh queryset from the IDs
                encounters = Encounter.objects.filter(
                    pk__in=encounter_ids
                ).defer('current_activity').select_related(
                    'patient', 'provider__user', 'location'
                ).prefetch_related('provider__department').order_by('-started_at', '-id')
            else:
                encounters = Encounter.objects.none()
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error in encounter deduplication: {str(e)}", exc_info=True)
        # Fallback: return queryset without deduplication
        encounters = Encounter.objects.filter(is_deleted=False).defer('current_activity').select_related(
            'patient', 'provider__user', 'location'
        ).prefetch_related('provider__department')
        
        if query:
            encounters = encounters.filter(
                Q(patient__first_name__icontains=query) |
                Q(patient__last_name__icontains=query) |
                Q(patient__mrn__icontains=query)
            )
        if status_filter:
            encounters = encounters.filter(status=status_filter)
        if type_filter:
            encounters = encounters.filter(encounter_type=type_filter)
        if date_from_obj or date_to_obj:
            from django.db.models.functions import Coalesce, TruncDate
            from django.db.models import F
            encounters = encounters.annotate(
                effective_date=TruncDate(Coalesce(F('started_at'), F('created')))
            )
            if date_from_obj:
                encounters = encounters.filter(effective_date__gte=date_from_obj)
            if date_to_obj:
                encounters = encounters.filter(effective_date__lte=date_to_obj)
        
        encounters = encounters.order_by('-started_at', '-id')
    
    # Pagination
    paginator = Paginator(encounters, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'encounters': page_obj,
        'status_filter': status_filter,
        'type_filter': type_filter,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'hospital/encounter_list.html', context)


@login_required
def encounter_create(request):
    """Create a new encounter with duplicate prevention"""
    from django.db import transaction
    
    if request.method == 'POST':
        form = EncounterForm(request.POST)
        if form.is_valid():
            # PREVENT DUPLICATES: Check for very recent duplicate encounter
            patient = form.cleaned_data.get('patient')
            encounter_type = form.cleaned_data.get('encounter_type')
            chief_complaint = form.cleaned_data.get('chief_complaint', '')
            
            with transaction.atomic():
                five_minutes_ago = timezone.now() - timedelta(minutes=5)
                existing_encounter = Encounter.objects.select_for_update().filter(
                    patient=patient,
                    encounter_type=encounter_type,
                    chief_complaint=chief_complaint,
                    status='active',
                    started_at__gte=five_minutes_ago,
                    is_deleted=False
                ).order_by('-created').first()
                
                if existing_encounter:
                    messages.info(
                        request,
                        f'An active encounter already exists for {patient.full_name} with the same complaint. '
                        f'Redirecting to existing encounter.'
                    )
                    return redirect('hospital:encounter_detail', pk=existing_encounter.pk)
                
                encounter = form.save()
                messages.success(request, 'Encounter created successfully.')
                return redirect('hospital:encounter_detail', pk=encounter.pk)
    else:
        form = EncounterForm()
    
    context = {'form': form}
    return render(request, 'hospital/encounter_form.html', context)


def _expand_antenatal_doctor_cards_for_quick_visit(doctors_with_pricing):
    """
    Doctors with anc_visit (e.g. Dr. Ali Samba) get a second card for ANC pricing.
    Antenatal duplicate rows are placed FIRST so the Antenatal group appears at the top
    of the page (easy to see without scrolling the specialty list).
    """
    if not doctors_with_pricing:
        return doctors_with_pricing
    from decimal import Decimal, InvalidOperation
    antenatal_front = []
    rest = []
    for entry in doctors_with_pricing:
        pi = entry.get('pricing_info') or {}
        amt = pi.get('anc_visit')
        if amt is not None:
            try:
                if not isinstance(amt, Decimal):
                    amt = Decimal(str(amt))
            except (ValueError, TypeError, InvalidOperation):
                amt = None
        if amt is not None:
            anc_pi = dict(pi)
            for k in (
                'first_visit', 'subsequent_visit', 'cash_first_visit', 'cash_subsequent_visit',
                'insurance_first_visit', 'insurance_subsequent_visit',
                'corporate_first_visit', 'corporate_subsequent_visit',
            ):
                if k in anc_pi:
                    anc_pi[k] = amt
            anc_display = dict(entry.get('display_info') or {})
            anc_display['specialty'] = 'Antenatal / ANC — 235 GHS'
            antenatal_front.append({
                'doctor': entry['doctor'],
                'pricing_info': anc_pi,
                'display_info': anc_display,
                'is_first_visit': True,
                'radio_id_suffix': '_antenatal',
            })
        rest.append(entry)
    return antenatal_front + rest


@login_required
def patient_quick_visit_create(request, patient_pk):
    """Quick visit/encounter creation for returning patients"""
    patient = get_object_or_404(Patient, pk=patient_pk, is_deleted=False)
    
    # Ensure patient has a valid payer (fix invalid payers)
    from .models import Payer
    cash_payer, _ = Payer.objects.get_or_create(
        name='Cash',
        defaults={'payer_type': 'cash', 'is_active': True}
    )
    
    # Fix invalid payer - refresh from DB to get latest data
    try:
        patient.refresh_from_db()
        if patient.primary_insurance_id:
            try:
                payer = patient.primary_insurance
                payer_type = getattr(payer, 'payer_type', None) if payer else None
                payer_name = getattr(payer, 'name', None) if payer else None
                
                # Check if payer is invalid (empty name, empty type, or invalid type)
                valid_types = ['cash', 'insurance', 'private', 'nhis', 'corporate']
                if (not payer or payer.is_deleted or
                    not payer_name or payer_name.strip() == '' or 
                    not payer_type or payer_type.strip() == '' or 
                    payer_type not in valid_types):
                    # Invalid payer - reset to cash
                    patient.primary_insurance = cash_payer
                    patient.save(update_fields=['primary_insurance'])
            except (AttributeError, Payer.DoesNotExist):
                # Payer doesn't exist or is deleted - set to cash
                patient.primary_insurance = cash_payer
                patient.save(update_fields=['primary_insurance'])
        elif not patient.primary_insurance_id:
            # No payer - set to cash
            patient.primary_insurance = cash_payer
            patient.save(update_fields=['primary_insurance'])
    except Exception as e:
        # If anything fails, just ensure cash payer is set
        logger.warning(f"Error fixing payer for patient {patient.mrn}: {e}")
        patient.primary_insurance = cash_payer
        patient.save(update_fields=['primary_insurance'])
    
    # Get available doctors for assignment - more inclusive query
    # Include all doctors with active users (regardless of staff.is_active status)
    # Use distinct() to prevent duplicates
    available_doctors = Staff.objects.filter(
        profession='doctor',
        is_deleted=False,
        user__isnull=False,
        user__is_active=True
    ).select_related('user', 'department').distinct().order_by('is_active', 'user__first_name', 'user__last_name')
    
    # Get doctor pricing information for display with insurance/corporate pricing
    from .utils_doctor_pricing import DoctorPricingService
    from .services.pricing_engine_service import pricing_engine
    from .models import ServiceCode
    from decimal import Decimal, InvalidOperation
    
    doctors_with_pricing = []
    seen_doctor_ids = set()  # Track to prevent duplicates
    
    # Get patient's payer for pricing
    payer = patient.primary_insurance if hasattr(patient, 'primary_insurance') else None
    payer_type = payer.payer_type if payer else 'cash'
    
    # Determine consultation service code (general vs specialist)
    # We'll determine this per doctor based on their specialty
    general_consultation_code = ServiceCode.objects.filter(code='CON001').first()
    specialist_consultation_code = ServiceCode.objects.filter(code='CON002').first()
    
    # Ensure we iterate over the queryset
    doctors_list = list(available_doctors) if hasattr(available_doctors, '__iter__') else []
    logger.info(f"Processing {len(doctors_list)} doctors for pricing info")
    
    for doctor in doctors_list:
        # Skip if we've already processed this doctor
        if doctor.id in seen_doctor_ids:
            continue
        seen_doctor_ids.add(doctor.id)
        
        try:
            pricing_info = DoctorPricingService.get_doctor_pricing_info(doctor)
            display_info = DoctorPricingService.get_doctor_display_info(doctor)
            is_first_visit = DoctorPricingService.is_first_visit_to_doctor(patient, doctor)
            
            # Get insurance/corporate pricing if patient has insurance/corporate
            # Ensure all values are valid Decimals with proper fallbacks
            default_price = Decimal('150.00')
            
            # Get base cash prices with validation
            cash_first_visit = pricing_info.get('first_visit')
            if not cash_first_visit or not isinstance(cash_first_visit, Decimal):
                try:
                    cash_first_visit = Decimal(str(cash_first_visit)) if cash_first_visit else default_price
                except (ValueError, TypeError, InvalidOperation):
                    cash_first_visit = default_price
            
            cash_subsequent_visit = pricing_info.get('subsequent_visit')
            if not cash_subsequent_visit or not isinstance(cash_subsequent_visit, Decimal):
                try:
                    cash_subsequent_visit = Decimal(str(cash_subsequent_visit)) if cash_subsequent_visit else default_price
                except (ValueError, TypeError, InvalidOperation):
                    cash_subsequent_visit = default_price
            
            # Ensure values are positive
            if cash_first_visit <= 0:
                cash_first_visit = default_price
            if cash_subsequent_visit <= 0:
                cash_subsequent_visit = default_price
            
            # Determine which service code to use
            service_code = specialist_consultation_code if pricing_info.get('is_specialist', False) else general_consultation_code
            
            # Initialize payer-specific pricing with cash prices as defaults
            insurance_first_visit = cash_first_visit
            insurance_subsequent_visit = cash_subsequent_visit
            corporate_first_visit = cash_first_visit
            corporate_subsequent_visit = cash_subsequent_visit
            
            if service_code:
                try:
                    # Get insurance price
                    insurance_price = pricing_engine.get_service_price(service_code, patient, payer)
                    if insurance_price and isinstance(insurance_price, Decimal) and insurance_price > 0:
                        insurance_first_visit = insurance_price
                        insurance_subsequent_visit = insurance_price
                    else:
                        # Fallback to cash pricing if insurance pricing fails
                        insurance_first_visit = cash_first_visit
                        insurance_subsequent_visit = cash_subsequent_visit
                    
                    # Get corporate price
                    corporate_price = pricing_engine.get_service_price(service_code, patient, payer)
                    if corporate_price and isinstance(corporate_price, Decimal) and corporate_price > 0:
                        corporate_first_visit = corporate_price
                        corporate_subsequent_visit = corporate_price
                    else:
                        # Fallback to cash pricing if corporate pricing fails
                        corporate_first_visit = cash_first_visit
                        corporate_subsequent_visit = cash_subsequent_visit
                except Exception as pricing_error:
                    logger.warning(f"Error getting payer pricing for doctor {doctor.id}: {pricing_error}")
                    # Use cash prices as fallback
                    insurance_first_visit = cash_first_visit
                    insurance_subsequent_visit = cash_subsequent_visit
                    corporate_first_visit = cash_first_visit
                    corporate_subsequent_visit = cash_subsequent_visit
            
            # Final validation - ensure all values are valid Decimals
            def ensure_decimal(value, default=default_price):
                """Ensure value is a valid Decimal"""
                if value is None:
                    return default
                if isinstance(value, Decimal):
                    return value if value > 0 else default
                try:
                    decimal_value = Decimal(str(value))
                    return decimal_value if decimal_value > 0 else default
                except (ValueError, TypeError, InvalidOperation):
                    return default
            
            # Add payer-specific pricing to pricing_info with validation
            pricing_info['cash_first_visit'] = ensure_decimal(cash_first_visit)
            pricing_info['cash_subsequent_visit'] = ensure_decimal(cash_subsequent_visit)
            pricing_info['insurance_first_visit'] = ensure_decimal(insurance_first_visit)
            pricing_info['insurance_subsequent_visit'] = ensure_decimal(insurance_subsequent_visit)
            pricing_info['corporate_first_visit'] = ensure_decimal(corporate_first_visit)
            pricing_info['corporate_subsequent_visit'] = ensure_decimal(corporate_subsequent_visit)
            
            doctors_with_pricing.append({
                'doctor': doctor,
                'pricing_info': pricing_info,
                'display_info': display_info,
                'is_first_visit': is_first_visit,
            })
        except Exception as e:
            # If pricing fails, still include doctor with default pricing
            logger.warning(f"Error getting pricing for doctor {doctor.id}: {e}", exc_info=True)
            try:
                default_fee = DoctorPricingService.DEFAULT_CONSULTATION_FEE
                # Ensure default_fee is a valid Decimal
                if not isinstance(default_fee, Decimal):
                    try:
                        default_fee = Decimal(str(default_fee))
                    except (ValueError, TypeError, InvalidOperation):
                        default_fee = Decimal('150.00')
                
                default_pricing = {
                    'first_visit': default_fee,
                    'subsequent_visit': default_fee,
                    'cash_first_visit': default_fee,
                    'cash_subsequent_visit': default_fee,
                    'insurance_first_visit': default_fee,
                    'insurance_subsequent_visit': default_fee,
                    'corporate_first_visit': default_fee,
                    'corporate_subsequent_visit': default_fee,
                    'specialty': doctor.specialization or 'General Consultation',
                    'show_price': False,
                    'is_specialist': False,
                }
                doctors_with_pricing.append({
                    'doctor': doctor,
                    'pricing_info': default_pricing,
                    'display_info': {
                        'specialty': doctor.specialization or 'General Consultation',
                        'is_specialist': False,
                        'show_price': False,
                        'price_text': 'Consultation Fee',
                    },
                    'is_first_visit': True,
                })
            except Exception as e2:
                logger.error(f"Critical error adding doctor {doctor.id} to list: {e2}", exc_info=True)
    
    if request.method == 'POST':
        encounter_type = request.POST.get('encounter_type', 'outpatient')
        chief_complaint = request.POST.get('chief_complaint', 'Follow-up visit')
        visit_reason = request.POST.get('visit_reason', 'new').strip().lower()  # 'new' or 'review'
        assigned_doctor_id = request.POST.get('assigned_doctor', '').strip()
        
        # Get assigned doctor if provided
        assigned_doctor_staff = None
        if assigned_doctor_id:
            try:
                assigned_doctor_staff = Staff.objects.get(
                    pk=assigned_doctor_id,
                    profession='doctor',
                    is_active=True,
                    is_deleted=False
                )
            except (Staff.DoesNotExist, ValueError):
                messages.warning(request, 'Selected doctor not found. Visit will be created without doctor assignment.')
        
        # Get current staff if no doctor assigned (for provider field)
        current_staff = assigned_doctor_staff
        if not current_staff and hasattr(request.user, 'staff'):
            current_staff = request.user.staff
        
        # PREVENT DUPLICATES: Check for very recent duplicate encounter (within 5 minutes)
        from django.db import transaction
        with transaction.atomic():
            five_minutes_ago = timezone.now() - timedelta(minutes=5)
            existing_encounter = Encounter.objects.select_for_update().filter(
                patient=patient,
                encounter_type=encounter_type,
                chief_complaint=chief_complaint,
                status='active',
                started_at__gte=five_minutes_ago,
                is_deleted=False
            ).order_by('-created').first()
            
            if existing_encounter:
                messages.info(
                    request,
                    f'An active encounter already exists for {patient.full_name} with the same complaint. '
                    f'Redirecting to existing encounter.'
                )
                return redirect('hospital:encounter_detail', pk=existing_encounter.pk)
            
            # Prepare notes - add review marker if review visit
            visit_notes = f'Visit created by {request.user.get_full_name() or request.user.username}'
            if visit_reason == 'review':
                visit_notes = f'[REVIEW_VISIT] {visit_notes}'
                # Also prepend to chief_complaint if not already there
                if 'review' not in chief_complaint.lower() and 'follow-up' not in chief_complaint.lower():
                    chief_complaint = f'Review: {chief_complaint}'
            
            # Create encounter/visit
            encounter = Encounter.objects.create(
                patient=patient,
                encounter_type=encounter_type,
                status='active',
                started_at=timezone.now(),
                provider=assigned_doctor_staff or current_staff,
                chief_complaint=chief_complaint,
                notes=visit_notes
            )
        
        # Send notification to assigned doctor if doctor was assigned
        if assigned_doctor_staff and assigned_doctor_staff.user:
            try:
                from .models import Notification
                Notification.objects.create(
                    recipient=assigned_doctor_staff.user,
                    notification_type='order_urgent',
                    title='New Patient Assigned',
                    message=f'Patient {patient.full_name} (MRN: {patient.mrn}) has been assigned to you. Chief Complaint: {chief_complaint[:100]}',
                    related_object_id=encounter.id,
                    related_object_type='Encounter'
                )
                # Also send SMS notification if doctor has phone number
                if assigned_doctor_staff.user.email or (hasattr(assigned_doctor_staff.user, 'phone_number') and assigned_doctor_staff.user.phone_number):
                    try:
                        from .services.sms_service import sms_service
                        doctor_name = assigned_doctor_staff.user.get_full_name() or assigned_doctor_staff.user.username
                        message = (
                            f"New Patient Assignment\n\n"
                            f"Patient: {patient.full_name}\n"
                            f"MRN: {patient.mrn}\n"
                            f"Complaint: {chief_complaint[:100]}\n\n"
                            f"Please check your consultation dashboard."
                        )
                        phone = getattr(assigned_doctor_staff.user, 'phone_number', None) or assigned_doctor_staff.phone_number if hasattr(assigned_doctor_staff, 'phone_number') else None
                        if phone:
                            sms_service.send_sms(
                                phone_number=phone,
                                message=message,
                                message_type='doctor_assignment',
                                recipient_name=doctor_name,
                                related_object_id=encounter.id,
                                related_object_type='Encounter'
                            )
                    except Exception as sms_error:
                        logger.warning(f"Failed to send SMS to doctor: {sms_error}")
            except Exception as notif_error:
                logger.warning(f"Failed to send notification to doctor: {notif_error}")
        
        # Sync payer type and create encounter invoice so cashier sees cash payment immediately
        consultation_amount = None
        try:
            from .services.visit_payer_sync_service import visit_payer_sync_service
            payer_type = request.POST.get('payer_type', '').strip()
            if not payer_type and patient.primary_insurance:
                payer_type = getattr(patient.primary_insurance, 'payer_type', None) or 'cash'
            if not payer_type:
                payer_type = 'cash'
            sync_result = visit_payer_sync_service.verify_and_set_payer_type(
                encounter=encounter,
                payer_type=payer_type
            )
            if sync_result['success']:
                if request.POST.get('payer_type', '').strip():
                    messages.info(
                        request,
                        f"Payer type set to {sync_result['payer_type']} ({sync_result['payer'].name}). "
                        f"Pricing will use {sync_result['payer_type']} rates."
                    )
        except Exception as sync_error:
            # Don't fail visit creation if sync fails
            logger.warning(f"Payer type sync failed for encounter {encounter.id}: {sync_error}")
        
        # 💰 AUTO-CREATE BILL FOR NEW VISITS (not review visits)
        # Use doctor-specific pricing with insurance/corporate pricing if applicable
        if visit_reason == 'new':
            try:
                from .utils_billing import CONSULTATION_LINE_SERVICE_CODES, add_consultation_charge
                from .utils_doctor_pricing import DoctorPricingService
                from .services.pricing_engine_service import pricing_engine
                from .models import ServiceCode, InvoiceLine
                from decimal import Decimal
                
                # Check for manual price override
                manual_price = request.POST.get('manual_price', '').strip()
                if manual_price:
                    try:
                        from decimal import InvalidOperation
                        manual_price = Decimal(manual_price)
                        if manual_price > 0:
                            logger.info(f"Using manual price override: GHS {manual_price}")
                    except (ValueError, InvalidOperation):
                        manual_price = None
                
                # Determine consultation type: Antenatal (235), Gynae/Special (260) so cashier sees price immediately
                consultation_type = 'general'
                if encounter_type == 'specialist' or encounter_type == 'gynae':
                    consultation_type = 'specialist'
                elif encounter_type == 'antenatal':
                    pass  # add_consultation_charge uses encounter.encounter_type and bills MAT-ANC 235 GHC
                elif assigned_doctor_staff:
                    pricing_info = DoctorPricingService.get_doctor_pricing_info(assigned_doctor_staff)
                    if pricing_info['is_specialist']:
                        consultation_type = 'specialist'
                
                # Get payer type for pricing
                payer_type = request.POST.get('payer_type', '').strip()
                if not payer_type and patient.primary_insurance:
                    payer_type = patient.primary_insurance.payer_type or 'cash'
                payer_type = payer_type or 'cash'
                
                # Determine if this is first visit to doctor
                is_first_visit = True
                if assigned_doctor_staff:
                    is_first_visit = DoctorPricingService.is_first_visit_to_doctor(patient, assigned_doctor_staff)
                
                # Get service code
                service_code = ServiceCode.objects.filter(code='CON002').first() if consultation_type == 'specialist' else ServiceCode.objects.filter(code='CON001').first()
                
                # Get the correct price based on payer type and visit type
                consultation_price = None
                if manual_price:
                    consultation_price = manual_price
                elif service_code:
                    # Use pricing engine to get payer-specific price
                    consultation_price = pricing_engine.get_service_price(service_code, patient, payer=patient.primary_insurance)
                    
                    # If pricing engine returns 0 or None, fall back to doctor pricing
                    if not consultation_price or consultation_price == 0:
                        if assigned_doctor_staff:
                            pricing_info = DoctorPricingService.get_doctor_pricing_info(assigned_doctor_staff)
                            if payer_type in ['insurance', 'private', 'nhis']:
                                consultation_price = pricing_info.get('insurance_first_visit' if is_first_visit else 'insurance_subsequent_visit', pricing_info.get('first_visit' if is_first_visit else 'subsequent_visit', Decimal('150.00')))
                            elif payer_type == 'corporate':
                                consultation_price = pricing_info.get('corporate_first_visit' if is_first_visit else 'corporate_subsequent_visit', pricing_info.get('first_visit' if is_first_visit else 'subsequent_visit', Decimal('150.00')))
                            else:
                                consultation_price = pricing_info.get('first_visit' if is_first_visit else 'subsequent_visit', Decimal('150.00'))
                        else:
                            consultation_price = Decimal('150.00')
                
                # Auto-create consultation bill with correct pricing
                invoice = add_consultation_charge(
                    encounter, 
                    consultation_type=consultation_type,
                    doctor_staff=assigned_doctor_staff
                )
                
                # Override price if manual price or payer-specific price was determined (include MAT-ANC for antenatal)
                if invoice and consultation_price:
                    consultation_line = InvoiceLine.objects.filter(
                        invoice_id=invoice.pk,
                        service_code__code__in=CONSULTATION_LINE_SERVICE_CODES,
                        is_deleted=False
                    ).order_by('-modified', '-created').first()
                    
                    if consultation_line:
                        # Update the price if it's different
                        if consultation_line.unit_price != consultation_price:
                            consultation_line.unit_price = consultation_price
                            consultation_line.line_total = consultation_price * Decimal(consultation_line.quantity)
                            consultation_line.save()
                            invoice.refresh_from_db()
                            invoice.update_totals()
                        
                        consultation_amount = consultation_line.unit_price
                        logger.info(
                            f"✅ Auto-created consultation bill for {patient.full_name}: "
                            f"GHS {consultation_amount} (Payer: {payer_type}, "
                            f"{'First' if is_first_visit else 'Review'} Visit)"
                        )
                        messages.info(
                            request,
                            f"💰 Bill created: {consultation_type.title()} Consultation - GHS {consultation_amount}. "
                            f"Patient should proceed to cashier for payment."
                        )
                    else:
                        consultation_amount = invoice.total_amount
                        logger.info(f"✅ Consultation bill created: GHS {consultation_amount}")
                else:
                    logger.warning(f"⚠️ Failed to create consultation bill for encounter {encounter.id}")
            except Exception as bill_error:
                logger.error(f"❌ Error creating consultation bill: {str(bill_error)}", exc_info=True)
                messages.warning(request, "Visit created, but bill creation failed. Please create bill manually.")
        
        # Create vital signs stage in patient flow
        try:
            from .models_workflow import PatientFlowStage
            # Check for existing stage before creating
            existing_stage = PatientFlowStage.objects.filter(
                encounter=encounter,
                stage_type='vitals',
                is_deleted=False
            ).first()
            
            if not existing_stage:
                PatientFlowStage.objects.create(
                    encounter=encounter,
                    stage_type='vitals',
                    status='pending'
                )
        except:
            pass
        
        # 🎫 QUEUE SYSTEM: Assign queue number and send SMS notification
        queue_number = None
        queue_position = None
        try:
            from .services.queue_service import queue_service
            from .services.queue_notification_service import queue_notification_service
            
            # Get department (try to get from encounter type or use default OPD)
            department = Department.objects.filter(name__icontains='outpatient').first()
            if not department:
                department = Department.objects.first()
            
            # Determine priority based on encounter type
            priority = 1 if encounter_type == 'emergency' else 3  # 1=Emergency, 3=Normal
            
            # Create queue entry
            assigned_doctor_user = assigned_doctor_staff.user if assigned_doctor_staff and assigned_doctor_staff.user else (current_staff.user if current_staff else None)
            queue_entry = queue_service.create_queue_entry(
                patient=patient,
                encounter=encounter,
                department=department,
                assigned_doctor=assigned_doctor_user,
                priority=priority,
                notes=f'Visit: {chief_complaint}'
            )
            
            patient_ticket = queue_entry.display_ticket_number
            queue_position = queue_service.get_position_in_queue(queue_entry)
            
            # Send queue SMS (Cashier reminder when applicable; no amounts in message)
            sms_sent = queue_notification_service.send_check_in_notification(queue_entry, consultation_amount=consultation_amount)
            
            logger.info(
                f"✅ Queue entry created: {patient_ticket} ({queue_entry.queue_number}) for {patient.full_name} "
                f"(Position: {queue_position}, Priority: {priority}, SMS sent: {sms_sent})"
            )
            
            # If queue SMS failed, try fallback SMS
            if not sms_sent and patient.phone_number:
                logger.warning(f"Queue SMS failed, trying fallback SMS for {patient.full_name}")
                try:
                    from .services.sms_service import sms_service
                    visit_date = encounter.started_at.strftime('%d/%m/%Y at %I:%M %p')
                    
                    payment_info = ""
                    if consultation_amount:
                        payment_info = (
                            "\nPlease proceed to the Cashier to make payment before consultation.\n"
                            "Your correct bill is available at the desk.\n"
                        )
                    
                    message = (
                        f"Dear {patient.first_name},\n\n"
                        f"Your visit has been registered at PrimeCare Hospital.\n\n"
                        f"Visit Type: {encounter.get_encounter_type_display()}\n"
                        f"Date/Time: {visit_date}\n"
                        f"MRN: {patient.mrn}\n"
                        f"Queue Number: {patient_ticket}\n"
                        f"{payment_info}"
                        f"Please proceed to the waiting area.\n\n"
                        f"Thank you,\nPrimeCare Hospital"
                    )
                    sms_result = sms_service.send_sms(
                        phone_number=patient.phone_number,
                        message=message,
                        message_type='visit_created',
                        recipient_name=patient.full_name,
                        related_object_id=encounter.id,
                        related_object_type='Encounter'
                    )
                    if sms_result.status == 'sent':
                        sms_sent = True
                        logger.info(f"✅ Fallback SMS sent successfully to {patient.full_name}")
                    else:
                        logger.warning(f"⚠️ Fallback SMS failed: {sms_result.error_message}")
                except Exception as sms_error:
                    logger.error(f"❌ Fallback SMS exception: {str(sms_error)}", exc_info=True)
            
            # Add success message to display
            if sms_sent:
                messages.success(
                    request,
                    f"✅ Visit created! Ticket: {patient_ticket}, Position: {queue_position}. SMS sent. Please record vital signs."
                )
            else:
                messages.warning(
                    request,
                    f"✅ Visit created! Ticket: {patient_ticket}, Position: {queue_position}. SMS could not be sent - check patient phone number. Please record vital signs."
                )
            
        except Exception as e:
            logger.error(f"❌ Error creating queue entry: {str(e)}", exc_info=True)
            # Fallback to old SMS if queue fails
            if patient.phone_number:
                try:
                    from .services.sms_service import sms_service
                    visit_date = encounter.started_at.strftime('%d/%m/%Y at %I:%M %p')
                    
                    payment_info = ""
                    if consultation_amount:
                        payment_info = (
                            "\nPlease proceed to the Cashier to make payment before consultation.\n"
                            "Your correct bill is available at the desk.\n"
                        )
                    
                    message = (
                        f"Dear {patient.first_name},\n\n"
                        f"Your visit has been registered at PrimeCare Hospital.\n\n"
                        f"Visit Type: {encounter.get_encounter_type_display()}\n"
                        f"Date/Time: {visit_date}\n"
                        f"MRN: {patient.mrn}\n"
                        f"{payment_info}"
                        f"Please proceed to the waiting area.\n\n"
                        f"Thank you,\nPrimeCare Hospital"
                    )
                    sms_result = sms_service.send_sms(
                        phone_number=patient.phone_number,
                        message=message,
                        message_type='visit_created',
                        recipient_name=patient.full_name,
                        related_object_id=encounter.id,
                        related_object_type='Encounter'
                    )
                    if sms_result.status == 'sent':
                        messages.success(request, f'New visit created for {patient.full_name}. SMS confirmation sent. Please record vital signs.')
                    else:
                        messages.warning(request, f'New visit created for {patient.full_name}, but SMS could not be sent: {sms_result.error_message}. Please record vital signs.')
                except Exception as sms_error:
                    logger.error(f"❌ Fallback SMS exception: {str(sms_error)}", exc_info=True)
                    messages.warning(request, f'New visit created for {patient.full_name}, but SMS could not be sent: {str(sms_error)}. Please record vital signs.')
            else:
                messages.warning(request, f'New visit created for {patient.full_name}, but patient has no phone number. Please record vital signs.')
        
        return redirect('hospital:record_vitals', encounter_id=encounter.pk)
    
    # For GET request, show the quick form
    # Prepare safe context for template
    
    # Fallback: If doctors_with_pricing is empty but available_doctors exists, create basic entries
    if not doctors_with_pricing:
        doctors_count = available_doctors.count() if hasattr(available_doctors, 'count') else len(list(available_doctors)) if hasattr(available_doctors, '__iter__') else 0
        logger.warning(f"doctors_with_pricing is empty but {doctors_count} doctors exist - creating fallback entries")
        
        if doctors_count > 0:
            from .utils_doctor_pricing import DoctorPricingService
            doctors_list = list(available_doctors) if hasattr(available_doctors, '__iter__') else []
            for doctor in doctors_list:
                # Skip if already in list
                if any(d['doctor'].id == doctor.id for d in doctors_with_pricing):
                    continue
                    
                try:
                    pricing_info = DoctorPricingService.get_doctor_pricing_info(doctor)
                    display_info = DoctorPricingService.get_doctor_display_info(doctor)
                    is_first_visit = DoctorPricingService.is_first_visit_to_doctor(patient, doctor)
                    
                    # Ensure all pricing fields exist with valid values
                    from decimal import Decimal, InvalidOperation
                    default_fee = Decimal('150.00')
                    
                    def ensure_decimal_field(value, default=default_fee):
                        if value is None:
                            return default
                        if isinstance(value, Decimal):
                            return value if value > 0 else default
                        try:
                            decimal_value = Decimal(str(value))
                            return decimal_value if decimal_value > 0 else default
                        except (ValueError, TypeError, InvalidOperation):
                            return default
                    
                    # Add payer-specific pricing if not present
                    if 'cash_first_visit' not in pricing_info:
                        cash_first = ensure_decimal_field(pricing_info.get('first_visit', default_fee))
                        cash_subsequent = ensure_decimal_field(pricing_info.get('subsequent_visit', default_fee))
                        pricing_info['cash_first_visit'] = cash_first
                        pricing_info['cash_subsequent_visit'] = cash_subsequent
                        pricing_info['insurance_first_visit'] = cash_first
                        pricing_info['insurance_subsequent_visit'] = cash_subsequent
                        pricing_info['corporate_first_visit'] = cash_first
                        pricing_info['corporate_subsequent_visit'] = cash_subsequent
                    else:
                        # Validate existing fields
                        pricing_info['cash_first_visit'] = ensure_decimal_field(pricing_info.get('cash_first_visit'))
                        pricing_info['cash_subsequent_visit'] = ensure_decimal_field(pricing_info.get('cash_subsequent_visit'))
                        pricing_info['insurance_first_visit'] = ensure_decimal_field(pricing_info.get('insurance_first_visit'))
                        pricing_info['insurance_subsequent_visit'] = ensure_decimal_field(pricing_info.get('insurance_subsequent_visit'))
                        pricing_info['corporate_first_visit'] = ensure_decimal_field(pricing_info.get('corporate_first_visit'))
                        pricing_info['corporate_subsequent_visit'] = ensure_decimal_field(pricing_info.get('corporate_subsequent_visit'))
                    
                    doctors_with_pricing.append({
                        'doctor': doctor,
                        'pricing_info': pricing_info,
                        'display_info': display_info,
                        'is_first_visit': is_first_visit,
                    })
                except Exception as e:
                    logger.error(f"Error creating fallback pricing for doctor {doctor.id}: {e}", exc_info=True)
                    # Add doctor with default pricing - ensure all fields are valid Decimals
                    from decimal import Decimal, InvalidOperation
                    default_fee = DoctorPricingService.DEFAULT_CONSULTATION_FEE
                    if not isinstance(default_fee, Decimal):
                        try:
                            default_fee = Decimal(str(default_fee))
                        except (ValueError, TypeError, InvalidOperation):
                            default_fee = Decimal('150.00')
                    
                    default_pricing = {
                        'first_visit': default_fee,
                        'subsequent_visit': default_fee,
                        'cash_first_visit': default_fee,
                        'cash_subsequent_visit': default_fee,
                        'insurance_first_visit': default_fee,
                        'insurance_subsequent_visit': default_fee,
                        'corporate_first_visit': default_fee,
                        'corporate_subsequent_visit': default_fee,
                        'specialty': doctor.specialization or 'General Consultation',
                        'show_price': False,
                        'is_specialist': False,
                    }
                    doctors_with_pricing.append({
                        'doctor': doctor,
                        'pricing_info': default_pricing,
                        'display_info': {
                            'specialty': doctor.specialization or 'General Consultation',
                            'is_specialist': False,
                            'show_price': False,
                            'price_text': 'Consultation Fee',
                        },
                        'is_first_visit': True,
                    })
    
    # Check if patient has previous visits to suggest review visit
    has_previous_visits = Encounter.objects.filter(
        patient=patient,
        is_deleted=False
    ).exclude(status='cancelled').exists()

    doctors_with_pricing = _expand_antenatal_doctor_cards_for_quick_visit(doctors_with_pricing)
    
    context = {
        'patient': patient,
        'available_doctors': available_doctors,  # Add available doctors to context
        'doctors_with_pricing': doctors_with_pricing,  # Doctors with pricing information
        'has_previous_visits': has_previous_visits,
    }
    
    # Safely get payer info for display
    try:
        if patient.primary_insurance_id:
            payer = patient.primary_insurance
            if payer and not payer.is_deleted:
                context['current_payer_name'] = payer.name or 'Cash'
                context['current_payer_type'] = payer.payer_type or 'cash'
                context['current_payer'] = payer
            else:
                context['current_payer_name'] = 'Cash'
                context['current_payer_type'] = 'cash'
                context['current_payer'] = None
        else:
            context['current_payer_name'] = 'Cash'
            context['current_payer_type'] = 'cash'
            context['current_payer'] = None
    except Exception:
        context['current_payer_name'] = 'Cash'
        context['current_payer_type'] = 'cash'
        context['current_payer'] = None
    
    # Final safety check - ensure doctors_with_pricing is not None and is a list
    if doctors_with_pricing is None:
        doctors_with_pricing = []
        logger.error("doctors_with_pricing was None - reset to empty list")
    
    # Ensure it's always a list
    if not isinstance(doctors_with_pricing, list):
        logger.error(f"doctors_with_pricing is not a list, it's {type(doctors_with_pricing)} - converting")
        doctors_with_pricing = list(doctors_with_pricing) if hasattr(doctors_with_pricing, '__iter__') else []
    
    # Final fallback - if still empty but doctors exist, create minimal entries
    if not doctors_with_pricing:
        from decimal import Decimal
        doctors_list = list(available_doctors) if hasattr(available_doctors, '__iter__') else []
        if doctors_list:
            logger.warning(f"Final fallback: Creating minimal doctor entries for {len(doctors_list)} doctors")
            default_fee = Decimal('150.00')
            for doctor in doctors_list:
                # Skip if already in list
                if any(d['doctor'].id == doctor.id for d in doctors_with_pricing):
                    continue
                    
                doctors_with_pricing.append({
                    'doctor': doctor,
                    'pricing_info': {
                        'first_visit': default_fee,
                        'subsequent_visit': default_fee,
                        'cash_first_visit': default_fee,
                        'cash_subsequent_visit': default_fee,
                        'insurance_first_visit': default_fee,
                        'insurance_subsequent_visit': default_fee,
                        'corporate_first_visit': default_fee,
                        'corporate_subsequent_visit': default_fee,
                        'specialty': doctor.specialization or 'General Consultation',
                        'show_price': False,
                        'is_specialist': False,
                    },
                    'display_info': {
                        'specialty': doctor.specialization or 'General Consultation',
                        'is_specialist': False,
                        'show_price': False,
                        'price_text': 'Consultation Fee',
                    },
                    'is_first_visit': True,
                })
    
    # Debug: Log context keys
    logger.info(f"Visit creation context keys: {list(context.keys())}")
    doctors_count = available_doctors.count() if hasattr(available_doctors, 'count') else len(list(available_doctors)) if hasattr(available_doctors, '__iter__') else 0
    logger.info(f"Available doctors count: {doctors_count}")
    logger.info(f"Doctors with pricing count: {len(doctors_with_pricing)}")
    
    # Ensure context is set
    context['doctors_with_pricing'] = doctors_with_pricing
    context['available_doctors'] = available_doctors
    
    return render(request, 'hospital/quick_visit_form.html', context)


@login_required
def encounter_detail(request, pk):
    """Enhanced Encounter detail view with super intelligent patient flow tracking"""
    encounter = get_object_or_404(Encounter, pk=pk, is_deleted=False)
    
    # Handle POST requests (e.g., adding nurse notes)
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add_nurse_note':
            # Add nurse note
            try:
                staff = Staff.objects.get(user=request.user, is_active=True)
                note_type = request.POST.get('note_type', 'progress')
                objective = request.POST.get('objective', '')
                notes = request.POST.get('notes', '').strip()
                
                if not notes:
                    messages.error(request, 'Note cannot be empty.')
                else:
                    from .models_advanced import ClinicalNote
                    ClinicalNote.objects.create(
                        encounter=encounter,
                        note_type=note_type,
                        objective=objective,
                        notes=notes,
                        created_by=staff
                    )
                    messages.success(request, 'Nurse note added successfully.')
            except Staff.DoesNotExist:
                messages.error(request, 'Staff profile not found.')
            except Exception as e:
                logger.error(f"Error adding nurse note: {e}", exc_info=True)
                messages.error(request, f'Error adding note: {str(e)}')
            
            return redirect('hospital:encounter_detail', pk=pk)
    
    # ========== SUPER INTELLIGENT PATIENT FLOW TRACKING ==========
    # Get all flow stages with comprehensive data
    flow_stages = PatientFlowStage.objects.filter(
        encounter=encounter,
        is_deleted=False
    ).select_related('completed_by__user').order_by('created')
    
    # Calculate comprehensive flow analytics
    flow_analytics = {}
    if flow_stages.exists():
        total_stages = flow_stages.count()
        completed_stages = flow_stages.filter(status='completed')
        in_progress_stages = flow_stages.filter(status='in_progress')
        pending_stages = flow_stages.filter(status='pending')
        
        # Calculate progress percentage
        flow_analytics['progress_percentage'] = round((completed_stages.count() / total_stages * 100) if total_stages > 0 else 0)
        
        # Calculate total time spent
        total_duration = timedelta()
        for stage in completed_stages:
            if stage.started_at and stage.completed_at:
                total_duration += (stage.completed_at - stage.started_at)
        flow_analytics['total_duration'] = total_duration
        flow_analytics['total_time_formatted'] = f"{int(total_duration.total_seconds() // 60)} min" if total_duration else "0 min"
        
        # Calculate average time per stage
        completed_count = completed_stages.count()
        if completed_count > 0:
            avg_seconds = total_duration.total_seconds() / completed_count
            flow_analytics['avg_time_per_stage'] = f"{int(avg_seconds // 60)} min"
        else:
            flow_analytics['avg_time_per_stage'] = "N/A"
        
        # Get current stage
        current_stage = in_progress_stages.first() or pending_stages.first()
        flow_analytics['current_stage'] = current_stage
        
        # Calculate current wait time (if in progress)
        if current_stage and current_stage.status == 'in_progress' and current_stage.started_at:
            wait_duration = timezone.now() - current_stage.started_at
            wait_minutes = int(wait_duration.total_seconds() // 60)
            flow_analytics['current_wait_time'] = wait_minutes
            flow_analytics['current_wait_formatted'] = f"{wait_minutes} min"
            # Flag if waiting too long (>60 min blinks)
            flow_analytics['is_long_wait'] = wait_minutes > 60
        else:
            flow_analytics['current_wait_time'] = 0
            flow_analytics['current_wait_formatted'] = "0 min"
            flow_analytics['is_long_wait'] = False
        
        # Determine current location based on stages
        current_location = "Unknown"
        location_stage_map = {
            'registration': 'Registration Desk',
            'triage': 'Triage Station',
            'vitals': 'Vitals Station',
            'consultation': 'Consultation Room',
            'laboratory': 'Laboratory',
            'imaging': 'Imaging/Radiolgy',
            'pharmacy': 'Pharmacy',
            'treatment': 'Treatment Room',
            'admission': encounter.location.name if encounter.location else 'Ward',
            'billing': 'Billing Desk',
            'payment': 'Cashier Desk',
            'discharge': 'Discharge Area',
        }
        
        if current_stage:
            current_location = location_stage_map.get(current_stage.stage_type, current_stage.get_stage_type_display())
        
        # Also check encounter.current_activity
        if encounter.current_activity:
            activities = encounter.current_activity.split(',')
            if activities:
                activity_map = {
                    'Consulting': 'Consultation Room',
                    'Lab': 'Laboratory',
                    'Pharmacy': 'Pharmacy',
                    'Imaging': 'Imaging/Radiolgy',
                }
                latest_activity = activities[-1].strip()
                current_location = activity_map.get(latest_activity, latest_activity) or current_location
        
        flow_analytics['current_location'] = current_location
        
        # Calculate time for each stage with intelligent insights
        enhanced_stages = []
        for stage in flow_stages:
            stage_data = {
                'stage': stage,
                'duration': "N/A",
                'elapsed_time': "N/A",
                'time_minutes': 0,
                'elapsed_minutes': 0,
                'is_slow': False,
                'is_current': stage == current_stage,
            }
            
            # For completed stages
            if stage.status == 'completed' and stage.started_at and stage.completed_at:
                duration = stage.completed_at - stage.started_at
                minutes = int(duration.total_seconds() // 60)
                seconds = int(duration.total_seconds() % 60)
                stage_data['duration'] = f"{minutes}m {seconds}s" if minutes > 0 else f"{seconds}s"
                stage_data['time_minutes'] = minutes
                # Flag if stage took too long (>90 min is slow)
                stage_data['is_slow'] = minutes > 90
            
            # For in-progress stages
            elif stage.status == 'in_progress' and stage.started_at:
                elapsed = timezone.now() - stage.started_at
                minutes = int(elapsed.total_seconds() // 60)
                stage_data['elapsed_time'] = f"{minutes} min" if minutes > 0 else "< 1 min"
                stage_data['elapsed_minutes'] = minutes
                stage_data['is_slow'] = minutes > 60  # Flag if waiting too long
            
            enhanced_stages.append(stage_data)
        
        flow_analytics['enhanced_stages'] = enhanced_stages
        flow_analytics['total_stages'] = total_stages
        flow_analytics['completed_count'] = completed_stages.count()
        flow_analytics['in_progress_count'] = in_progress_stages.count()
        flow_analytics['pending_count'] = pending_stages.count()
        
        # Get staff involved
        staff_involved = flow_stages.filter(completed_by__isnull=False).values(
            'completed_by__user__first_name',
            'completed_by__user__last_name',
            'completed_by__user__username'
        ).distinct()
        flow_analytics['staff_count'] = staff_involved.count()
        flow_analytics['staff_involved'] = list(staff_involved)
        
        # Calculate overall encounter duration
        if encounter.started_at:
            if encounter.ended_at:
                encounter_duration = encounter.ended_at - encounter.started_at
            else:
                encounter_duration = timezone.now() - encounter.started_at
            flow_analytics['encounter_duration'] = encounter_duration
            flow_analytics['encounter_duration_formatted'] = f"{int(encounter_duration.total_seconds() // 60)} min"
        else:
            flow_analytics['encounter_duration'] = timedelta()
            flow_analytics['encounter_duration_formatted'] = "N/A"
    else:
        # No flow stages yet
        flow_analytics = {
            'progress_percentage': 0,
            'total_stages': 0,
            'completed_count': 0,
            'in_progress_count': 0,
            'pending_count': 0,
            'current_stage': None,
            'current_location': encounter.location.name if encounter.location else "Unknown",
            'current_wait_time': 0,
            'current_wait_formatted': "0 min",
            'is_long_wait': False,
            'enhanced_stages': [],
            'total_time_formatted': "0 min",
            'avg_time_per_stage': "N/A",
            'staff_count': 0,
            'staff_involved': [],
            'encounter_duration_formatted': "0 min" if encounter.started_at else "N/A",
        }
    
    # Get vital signs
    all_vitals = encounter.vitals.filter(is_deleted=False).order_by('-recorded_at')
    vitals = all_vitals[:10]
    latest_vitals = all_vitals.first()
    
    # Get orders
    orders = encounter.orders.filter(is_deleted=False).order_by('-created')[:20]
    
    # Get referrals
    referrals = []
    try:
        from .models_specialists import Referral
        referrals = Referral.objects.filter(
            encounter=encounter,
            is_deleted=False
        ).select_related('specialist__staff__user', 'specialty', 'referring_doctor__user').order_by('-referred_date')
    except:
        pass
    
    # Get care plans (doctor's treatment plans) - FOR NURSES
    care_plans = []
    try:
        from .models_advanced import CarePlan
        care_plans = CarePlan.objects.filter(
            encounter=encounter,
            is_deleted=False
        ).select_related('created_by__user').order_by('-created')
    except:
        pass
    
    # Get clinical notes with plan field (doctor's plans) – no limit so all notes visible
    clinical_notes_with_plan = []
    nurse_notes = []
    doctor_notes = []
    other_notes = []
    try:
        from .models_advanced import ClinicalNote
        clinical_notes_with_plan = ClinicalNote.objects.filter(
            encounter=encounter,
            is_deleted=False
        ).exclude(plan='').exclude(plan__isnull=True).select_related('created_by__user').order_by('-created')
        # Progress notes split by author for dedicated Nurse / Doctor sections
        _progress = ClinicalNote.objects.filter(
            encounter=encounter,
            is_deleted=False,
            note_type='progress'
        ).select_related('created_by__user').order_by('-created')
        nurse_notes = [n for n in _progress if getattr(getattr(n, 'created_by', None), 'profession', None) == 'nurse']
        doctor_notes = [n for n in _progress if getattr(getattr(n, 'created_by', None), 'profession', None) == 'doctor']
        # All other notes (SOAP, consultation, discharge, progress by other roles) so every saved note shows
        all_encounter_notes = ClinicalNote.objects.filter(
            encounter=encounter,
            is_deleted=False
        ).select_related('created_by__user').order_by('-created')
        nurse_ids = {n.id for n in nurse_notes}
        doctor_ids = {n.id for n in doctor_notes}
        other_notes = [n for n in all_encounter_notes if n.id not in nurse_ids and n.id not in doctor_ids]
    except Exception:
        other_notes = []
    
    user_role = get_user_role(request.user)
    # Front desk/receptionist cannot view vitals - only medical staff
    can_view_vitals = user_role not in {'receptionist', 'frontdesk'}
    is_nurse = user_role == 'nurse'

    # Pre-employment / Pre-admission screening: report and flags for UI
    screening_report = None
    is_screening_encounter = encounter.encounter_type in ('pre_employment', 'pre_admission')
    try:
        from .models_screening import ScreeningReport
        screening_report = ScreeningReport.objects.filter(encounter=encounter).first()
    except Exception:
        pass

    context = {
        'encounter': encounter,
        'vitals': vitals,
        'latest_vitals': latest_vitals,
        'orders': orders,
        'referrals': referrals,
        'care_plans': care_plans,
        'clinical_notes_with_plan': clinical_notes_with_plan,
        'nurse_notes': nurse_notes,
        'doctor_notes': doctor_notes,
        'other_notes': other_notes,
        'user_role': user_role,
        'can_view_vitals': can_view_vitals,
        'is_nurse': is_nurse,
        # Enhanced patient flow data
        'flow_analytics': flow_analytics,
        'flow_stages': flow_stages,
        # Pre-employment / Pre-admission screening
        'screening_report': screening_report,
        'is_screening_encounter': is_screening_encounter,
    }
    return render(request, 'hospital/encounter_detail.html', context)


@login_required
@csrf_exempt
def encounter_flow_ajax(request, pk):
    """AJAX endpoint for real-time patient flow updates"""
    if request.method != 'GET':
        return JsonResponse({'error': 'GET method required'}, status=405)
    
    try:
        encounter = get_object_or_404(Encounter, pk=pk, is_deleted=False)
        
        # Get flow stages
        flow_stages = PatientFlowStage.objects.filter(
            encounter=encounter,
            is_deleted=False
        ).select_related('completed_by__user').order_by('created')
        
        # Build response data
        response_data = {
            'success': True,
            'encounter_id': str(encounter.pk),
            'current_location': "Unknown",
            'current_stage': None,
            'current_wait_time': 0,
            'current_wait_formatted': "0 min",
            'is_long_wait': False,
            'progress_percentage': 0,
            'stages': [],
        }
        
        if flow_stages.exists():
            completed_stages = flow_stages.filter(status='completed')
            in_progress_stages = flow_stages.filter(status='in_progress')
            pending_stages = flow_stages.filter(status='pending')
            
            total_stages = flow_stages.count()
            response_data['progress_percentage'] = round((completed_stages.count() / total_stages * 100) if total_stages > 0 else 0)
            
            # Current stage
            current_stage = in_progress_stages.first() or pending_stages.first()
            if current_stage:
                response_data['current_stage'] = {
                    'id': str(current_stage.pk),
                    'type': current_stage.stage_type,
                    'type_display': current_stage.get_stage_type_display(),
                    'status': current_stage.status,
                }
                
                # Calculate wait time
                if current_stage.status == 'in_progress' and current_stage.started_at:
                    wait_duration = timezone.now() - current_stage.started_at
                    wait_minutes = int(wait_duration.total_seconds() // 60)
                    response_data['current_wait_time'] = wait_minutes
                    response_data['current_wait_formatted'] = f"{wait_minutes} min"
                    response_data['is_long_wait'] = wait_minutes > 60
                
                # Location mapping
                location_stage_map = {
                    'registration': 'Registration Desk',
                    'triage': 'Triage Station',
                    'vitals': 'Vitals Station',
                    'consultation': 'Consultation Room',
                    'laboratory': 'Laboratory',
                    'imaging': 'Imaging/Radiolgy',
                    'pharmacy': 'Pharmacy',
                    'treatment': 'Treatment Room',
                    'admission': encounter.location.name if encounter.location else 'Ward',
                    'billing': 'Billing Desk',
                    'payment': 'Cashier Desk',
                    'discharge': 'Discharge Area',
                }
                response_data['current_location'] = location_stage_map.get(current_stage.stage_type, current_stage.get_stage_type_display())
            
            # Also check encounter.current_activity
            if encounter.current_activity:
                activities = encounter.current_activity.split(',')
                if activities:
                    activity_map = {
                        'Consulting': 'Consultation Room',
                        'Lab': 'Laboratory',
                        'Pharmacy': 'Pharmacy',
                        'Imaging': 'Imaging/Radiolgy',
                    }
                    latest_activity = activities[-1].strip()
                    response_data['current_location'] = activity_map.get(latest_activity, latest_activity) or response_data['current_location']
            
            # Build stages data
            stages_data = []
            for stage in flow_stages:
                stage_info = {
                    'id': str(stage.pk),
                    'type': stage.stage_type,
                    'type_display': stage.get_stage_type_display(),
                    'status': stage.status,
                    'duration': "N/A",
                    'elapsed_time': "N/A",
                    'time_minutes': 0,
                    'elapsed_minutes': 0,
                    'is_slow': False,
                    'is_current': stage == current_stage,
                }
                
                # Completed stages
                if stage.status == 'completed' and stage.started_at and stage.completed_at:
                    duration = stage.completed_at - stage.started_at
                    minutes = int(duration.total_seconds() // 60)
                    seconds = int(duration.total_seconds() % 60)
                    stage_info['duration'] = f"{minutes}m {seconds}s" if minutes > 0 else f"{seconds}s"
                    stage_info['time_minutes'] = minutes
                    stage_info['is_slow'] = minutes > 90
                
                # In-progress stages
                elif stage.status == 'in_progress' and stage.started_at:
                    elapsed = timezone.now() - stage.started_at
                    minutes = int(elapsed.total_seconds() // 60)
                    stage_info['elapsed_time'] = f"{minutes} min" if minutes > 0 else "< 1 min"
                    stage_info['elapsed_minutes'] = minutes
                    stage_info['is_slow'] = minutes > 60
                
                stages_data.append(stage_info)
            
            response_data['stages'] = stages_data
        
        return JsonResponse(response_data)
        
    except Exception as e:
        logger.error(f"Error in encounter_flow_ajax: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@csrf_exempt
def surgery_control(request, encounter_id):
    """Control surgery start/complete/notes"""
    try:
        if request.method != 'POST':
            return JsonResponse({'error': 'Method not allowed'}, status=405)
        
        logger.info(f"Surgery control request for encounter {encounter_id}, action: {request.POST.get('action')}")
        
        encounter = get_object_or_404(Encounter, pk=encounter_id, is_deleted=False)
        
        # Verify it's a surgery encounter
        if encounter.encounter_type != 'surgery':
            return JsonResponse({'error': 'This is not a surgery encounter'}, status=400)
        
        action = request.POST.get('action')
    except Exception as e:
        logger.error(f"Error in surgery_control: {str(e)}", exc_info=True)
        return JsonResponse({'error': f'Server error: {str(e)}'}, status=500)
    
    if action == 'start':
        # Start surgery
        if encounter.status != 'active':
            return JsonResponse({'error': 'Encounter is not active'}, status=400)
        
        # Update encounter start time and add initial note
        encounter.started_at = timezone.now()
        current_notes = encounter.notes or ''
        encounter.notes = f"[SURGERY STARTED: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}]\n\n{current_notes}"
        encounter.save()
        
        messages.success(request, 'Surgery started successfully!')
        return JsonResponse({
            'success': True,
            'message': 'Surgery started',
            'started_at': encounter.started_at.isoformat()
        })
    
    elif action == 'complete':
        # Complete surgery
        if encounter.status != 'active':
            return JsonResponse({'error': 'Encounter is not active'}, status=400)
        
        # Mark as completed
        encounter.status = 'completed'
        encounter.ended_at = timezone.now()
        current_notes = encounter.notes or ''
        duration_minutes = encounter.get_duration_minutes() or 0
        encounter.notes = f"{current_notes}\n\n[SURGERY COMPLETED: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}]\nDuration: {duration_minutes} minutes"
        encounter.save()
        
        messages.success(request, 'Surgery completed successfully!')
        return JsonResponse({
            'success': True,
            'message': 'Surgery completed',
            'ended_at': encounter.ended_at.isoformat(),
            'duration_minutes': duration_minutes
        })
    
    elif action == 'add_note':
        # Add surgical note
        note = request.POST.get('note', '').strip()
        if not note:
            return JsonResponse({'error': 'Note cannot be empty'}, status=400)
        
        current_notes = encounter.notes or ''
        timestamp = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        user_name = request.user.get_full_name() or request.user.username
        encounter.notes = f"{current_notes}\n\n[{timestamp}] {user_name}: {note}"
        encounter.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Note added successfully'
        })
    
    elif action == 'report_issue':
        # Report complication/issue
        issue = request.POST.get('issue', '').strip()
        if not issue:
            return JsonResponse({'error': 'Issue description cannot be empty'}, status=400)
        
        current_notes = encounter.notes or ''
        timestamp = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        user_name = request.user.get_full_name() or request.user.username
        encounter.notes = f"{current_notes}\n\n[⚠️ COMPLICATION - {timestamp}] Reported by {user_name}:\n{issue}"
        encounter.save()
        
        # You could also create an alert or notification here
        messages.warning(request, f'Complication reported: {issue}')
        
        return JsonResponse({
            'success': True,
            'message': 'Issue reported and logged'
        })
    
    else:
        return JsonResponse({'error': 'Invalid action'}, status=400)


@login_required
def admission_list(request):
    """List all admissions"""
    status_filter = request.GET.get('status', 'admitted')
    
    admissions = Admission.objects.filter(is_deleted=False).select_related(
        'encounter__patient', 'ward', 'bed', 'admitting_doctor__user'
    ).order_by('-admit_date')
    
    if status_filter:
        admissions = admissions.filter(status=status_filter)
    
    # Calculate statistics
    total_admissions = Admission.objects.filter(is_deleted=False).count()
    current_admissions = Admission.objects.filter(status='admitted', is_deleted=False).count()
    discharged_today = Admission.objects.filter(
        is_deleted=False,
        discharge_date__date=timezone.now().date()
    ).count()
    
    # Bed statistics
    total_beds = Bed.objects.filter(is_deleted=False, is_active=True).count()
    available_beds = Bed.objects.filter(is_deleted=False, is_active=True, status='available').count()
    occupied_beds = Bed.objects.filter(is_deleted=False, is_active=True, status='occupied').count()
    bed_occupancy_rate = round((occupied_beds / total_beds * 100) if total_beds > 0 else 0, 1)
    
    stats = {
        'current_admissions': current_admissions,
        'available_beds': available_beds,
        'total_beds': total_beds,
        'bed_occupancy_rate': bed_occupancy_rate,
        'discharged_today': discharged_today,
    }
    
    paginator = Paginator(admissions, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'admissions': page_obj,
        'status_filter': status_filter,
        'total_admissions': total_admissions,
        'current_admissions': current_admissions,
        'stats': stats,
    }
    return render(request, 'hospital/admission_list.html', context)


@login_required
@block_pharmacy_from_invoice_and_payment
def invoice_list(request):
    """List invoices - supports combined by patient (accountant) or individual invoices"""
    from decimal import Decimal
    
    status_filter = request.GET.get('status', '')
    payer_type_filter = request.GET.get('payer_type', '')
    query = request.GET.get('q', '')
    view_mode = request.GET.get('view', 'combined')  # 'combined' = by patient, 'invoices' = individual
    
    # Ensure prescribe (pharmacy) sales have invoices so they all show on this page
    try:
        from hospital.models_pharmacy_walkin import WalkInPharmacySale
        from hospital.services.pharmacy_walkin_service import WalkInPharmacyService
        # Find prescribe sales that have items but no invoice yet (limit per request to avoid slow load)
        sales_without_invoice = []
        for sale in WalkInPharmacySale.objects.filter(
            is_deleted=False,
            items__is_deleted=False,
        ).distinct().order_by('-sale_date')[:50]:
            if not sale.items.filter(is_deleted=False).exists():
                continue
            if Invoice.all_objects.filter(
                patient__isnull=False,
                is_deleted=False,
                lines__description__icontains=sale.sale_number,
            ).exclude(status__in=['cancelled', 'paid']).exists():
                continue
            sales_without_invoice.append(sale)
        for sale in sales_without_invoice:
            try:
                patient = WalkInPharmacyService.ensure_sale_patient(sale)
                WalkInPharmacyService.ensure_sale_invoice(sale, patient)
            except Exception:
                pass
    except Exception:
        pass

    base_invoices = Invoice.objects.filter(
        is_deleted=False,
        patient__isnull=False,
        patient__is_deleted=False
    ).select_related('patient', 'payer', 'encounter')
    
    if query:
        query_clean = query.strip()
        base_invoices = base_invoices.filter(
            Q(patient__first_name__icontains=query_clean) |
            Q(patient__last_name__icontains=query_clean) |
            Q(patient__middle_name__icontains=query_clean) |
            Q(patient__mrn__icontains=query_clean) |
            Q(invoice_number__icontains=query_clean) |
            Q(payer__name__icontains=query_clean)  # corporate: search by company (payer) name
        )
        # Ensure prescribe (pharmacy) sales have an invoice so they appear when searching (e.g. "joe" / "Joe Anokye")
        if query_clean:
            try:
                from hospital.models import Patient
                from hospital.models_pharmacy_walkin import WalkInPharmacySale
                from hospital.services.pharmacy_walkin_service import WalkInPharmacyService
                matching_patients = Patient.objects.filter(
                    is_deleted=False
                ).filter(
                    Q(first_name__icontains=query_clean) |
                    Q(last_name__icontains=query_clean) |
                    Q(middle_name__icontains=query_clean) |
                    Q(mrn__icontains=query_clean)
                )[:20]
                for patient in matching_patients:
                    for sale in WalkInPharmacySale.objects.filter(
                        patient=patient, is_deleted=False,
                        items__is_deleted=False,
                    ).distinct():
                        if not sale.items.filter(is_deleted=False).exists():
                            continue
                        if Invoice.all_objects.filter(
                            patient=patient,
                            is_deleted=False,
                            lines__description__icontains=sale.sale_number,
                        ).exclude(status__in=['cancelled', 'paid']).exists():
                            continue
                        try:
                            p = WalkInPharmacyService.ensure_sale_patient(sale)
                            WalkInPharmacyService.ensure_sale_invoice(sale, p)
                        except Exception:
                            pass
            except Exception:
                pass
    if status_filter:
        base_invoices = base_invoices.filter(status=status_filter)
    if payer_type_filter == 'insurance':
        base_invoices = base_invoices.filter(payer__payer_type__in=['private', 'nhis'])
    elif payer_type_filter and payer_type_filter != 'all':
        base_invoices = base_invoices.filter(payer__payer_type=payer_type_filter)
    
    # Stats: use filtered base_invoices so cards match the list (e.g. corporate + company search)
    stats_invoices = base_invoices
    stats = {
        'total_invoices': stats_invoices.count(),
        'paid_invoices': stats_invoices.filter(status='paid').count(),
        'total_revenue': stats_invoices.filter(status='paid').aggregate(
            Sum('total_amount'))['total_amount__sum'] or 0,
        'outstanding_balance': stats_invoices.filter(
            status__in=['issued', 'partially_paid', 'overdue'],
            balance__gt=0
        ).aggregate(Sum('balance'))['balance__sum'] or 0,
    }
    
    if view_mode == 'combined':
        # Combined by patient: include ALL payers (insurance + corporate + cash) so we combine
        # invoices of individuals who have both insurance and corporate into one row per patient.
        outstanding_all_payers = Invoice.objects.filter(
            is_deleted=False,
            patient__isnull=False,
            patient__is_deleted=False,
            status__in=['issued', 'partially_paid', 'overdue'],
            balance__gt=0,
        ).select_related('patient', 'payer')
        if query:
            query_clean = query.strip()
            outstanding_all_payers = outstanding_all_payers.filter(
                Q(patient__first_name__icontains=query_clean) |
                Q(patient__last_name__icontains=query_clean) |
                Q(patient__middle_name__icontains=query_clean) |
                Q(patient__mrn__icontains=query_clean) |
                Q(invoice_number__icontains=query_clean) |
                Q(payer__name__icontains=query_clean)
            )
        if status_filter:
            outstanding_all_payers = outstanding_all_payers.filter(status=status_filter)
        # Optional: restrict which patients appear by payer type (balance shown is still full combined)
        if payer_type_filter == 'insurance':
            patient_ids_with_payer = set(
                base_invoices.filter(
                    status__in=['issued', 'partially_paid', 'overdue'],
                    balance__gt=0,
                ).values_list('patient_id', flat=True).distinct()
            )
        elif payer_type_filter and payer_type_filter != 'all':
            patient_ids_with_payer = set(
                base_invoices.filter(
                    status__in=['issued', 'partially_paid', 'overdue'],
                    balance__gt=0,
                ).values_list('patient_id', flat=True).distinct()
            )
        else:
            patient_ids_with_payer = None  # show all patients with any outstanding
        from django.db.models import Count
        patient_aggs = outstanding_all_payers.values(
            'patient', 'patient__first_name', 'patient__last_name', 'patient__mrn'
        ).annotate(
            invoice_count=Count('id'),
            total_balance=Sum('balance'),
            total_amount=Sum('total_amount'),
        ).order_by('-total_balance')
        # Build simple list for template
        patient_rows = []
        for agg in patient_aggs:
            if patient_ids_with_payer is not None and agg['patient'] not in patient_ids_with_payer:
                continue
            from hospital.models import Patient
            try:
                p = Patient.objects.get(pk=agg['patient'], is_deleted=False)
            except Patient.DoesNotExist:
                continue
            patient_rows.append({
                'patient': p,
                'invoice_count': agg['invoice_count'],
                'total_balance': agg['total_balance'] or Decimal('0'),
                'total_amount': agg['total_amount'] or Decimal('0'),
            })
        paginator = Paginator(patient_rows, 20)
        page_obj = paginator.get_page(request.GET.get('page'))
        context = {
            'view_mode': 'combined',
            'patient_rows': page_obj,
            'page_obj': page_obj,
            'status_filter': status_filter,
            'payer_type_filter': payer_type_filter,
            'stats': stats,
        }
    else:
        invoices = base_invoices.order_by('-issued_at')
        paginator = Paginator(invoices, 20)
        page_obj = paginator.get_page(request.GET.get('page'))
        context = {
            'view_mode': 'invoices',
            'invoices': page_obj,
            'page_obj': page_obj,
            'status_filter': status_filter,
            'payer_type_filter': payer_type_filter,
            'stats': stats,
        }
    return render(request, 'hospital/invoice_list.html', context)


def _group_invoice_lines_by_category(lines):
    """Group invoice lines by service category for categorical display (Consultation, Lab, Pharmacy, etc.)"""
    from itertools import groupby

    from .utils_billing import invoice_line_display_category

    sorted_lines = sorted(lines, key=invoice_line_display_category)
    return [(cat, list(grp)) for cat, grp in groupby(sorted_lines, key=invoice_line_display_category)]


def _line_visit_datetime(line, invoice):
    """Best-effort datetime for when the charge was posted (for visit grouping)."""
    return getattr(line, 'created', None) or getattr(invoice, 'issued_at', None)


def _group_invoice_lines_by_visit_then_category(lines, invoice):
    """
    Group lines by calendar day of line.created (fallback: invoice.issued_at), then by category.
    Shows date + time span when multiple times occur the same day.
    """
    from collections import defaultdict

    buckets = defaultdict(list)
    for line in lines:
        dt = _line_visit_datetime(line, invoice)
        d = dt.date() if dt else None
        buckets[d].append(line)

    visits_out = []
    for d in sorted(buckets.keys(), key=lambda x: x or date.min, reverse=True):
        vlines = buckets[d]
        vlines.sort(key=lambda x: _line_visit_datetime(x, invoice) or timezone.now())
        times = [
            _line_visit_datetime(x, invoice)
            for x in vlines
            if _line_visit_datetime(x, invoice)
        ]
        time_part = ''
        if times:
            t_min = min(times).time()
            t_max = max(times).time()
            if t_min != t_max:
                time_part = f' · {t_min.strftime("%H:%M")}–{t_max.strftime("%H:%M")}'
            else:
                time_part = f' · {t_min.strftime("%H:%M")}'
        if d:
            label = f'{d.strftime("%a %d %b %Y")}{time_part}'
        else:
            label = 'Charges'
        visits_out.append(
            {
                'visit_date': d,
                'visit_label': label,
                'lines_by_category': _group_invoice_lines_by_category(vlines),
            }
        )
    return visits_out


@login_required
@block_pharmacy_from_invoice_and_payment
def invoice_detail(request, pk):
    """Invoice detail view. Uses all_objects so write-off period, zero-amount, and soft-deleted invoices can still be viewed by direct link."""
    from .utils_roles import user_can_waive
    # Use all_objects + no is_deleted filter so any invoice is viewable by direct URL (avoid 404 for write-off/deleted)
    invoice = get_object_or_404(
        Invoice.all_objects.select_related('patient', 'payer', 'encounter'),
        pk=pk
    )
    try:
        from .utils_invoice_line import heal_invoice_zero_line_prices
        heal_invoice_zero_line_prices(invoice)
    except Exception:
        import logging
        logging.getLogger(__name__).exception('heal_invoice_zero_line_prices failed')
    invoice.update_totals()
    # Exclude waived lines - they contribute 0 to total and should not appear on invoice
    invoice_lines = list(invoice.lines.filter(is_deleted=False).filter(waived_at__isnull=True).select_related('service_code'))
    lines_by_category = _group_invoice_lines_by_category(invoice_lines)
    invoice_visits = _group_invoice_lines_by_visit_then_category(invoice_lines, invoice)
    days_overdue = invoice.get_days_overdue() if hasattr(invoice, 'get_days_overdue') else 0
    can_waive = user_can_waive(request.user)
    context = {
        'invoice': invoice,
        'lines': invoice_lines,
        'lines_by_category': lines_by_category,
        'invoice_visits': invoice_visits,
        'days_overdue': days_overdue,
        'can_waive': can_waive,
        'invoice_removed': getattr(invoice, 'is_deleted', False),
    }
    return render(request, 'hospital/invoice_detail.html', context)


@login_required
@block_pharmacy_from_invoice_and_payment
def invoice_combined_patient(request, patient_id):
    """Combined bill for one patient - all unpaid invoices + pending pharmacy/prescription (especially pharmacy). Waive items (administrators only)."""
    from decimal import Decimal
    from itertools import groupby
    from .models import Patient, InvoiceLine
    from .utils_roles import user_can_waive

    patient = get_object_or_404(Patient, pk=patient_id, is_deleted=False)
    # Include draft + unpaid so all billable items show (drugs, ECG, lab, etc.).
    # Also include invoices where DB balance > 0 even if status is "paid" (stale/corrupt);
    # update_totals() then decides the real balance so lines do not disappear incorrectly.
    unpaid_qs = Invoice.objects.filter(
        patient=patient,
        is_deleted=False,
    ).filter(
        Q(status__in=['draft', 'issued', 'partially_paid', 'overdue']) | Q(balance__gt=0)
    ).exclude(status='cancelled').select_related('payer', 'encounter').order_by('-issued_at')

    unpaid = []
    for inv in unpaid_qs:
        inv.update_totals()
        # Paid in full or removed: no balance — do not list even if line rows still exist
        if (inv.balance or Decimal('0.00')) <= 0:
            continue
        unpaid.append(inv)

    from .utils_invoice_line import heal_invoice_zero_line_prices
    all_lines = []
    for inv in unpaid:
        try:
            heal_invoice_zero_line_prices(inv)
        except Exception:
            pass
        for line in inv.lines.filter(is_deleted=False).filter(waived_at__isnull=True).select_related('service_code'):
            all_lines.append({'invoice': inv, 'line': line})

    get_cat = lambda x: (getattr(x['line'].service_code, 'category', None) or 'Other') if x['line'].service_code else 'Other'
    sorted_items = sorted(all_lines, key=get_cat)
    lines_by_category = [(cat, list(grp)) for cat, grp in groupby(sorted_items, key=get_cat)]
    # Numbered items for invoice print (# column)
    running = 0
    numbered_by_category = []
    for cat, grp in lines_by_category:
        items = list(grp)
        numbered = [(running + i + 1, items[i]) for i in range(len(items))]
        running += len(items)
        numbered_by_category.append((cat, numbered))

    total_balance = sum((inv.balance or 0) for inv in unpaid)
    # Initial (total from items): sum of displayed line items so it matches the itemized list exactly
    if all_lines:
        total_amount = sum(
            (item['line'].display_line_total or Decimal('0.00')) for item in all_lines
        )
    else:
        total_amount = sum((inv.total_amount or Decimal('0.00')) for inv in unpaid)

    # Amount paid to date: receipts for this patient; fallback to invoice (total - balance) when no receipts
    from .models_accounting import PaymentReceipt
    from django.db.models import Sum
    total_paid = (
        PaymentReceipt.objects.filter(patient=patient, is_deleted=False).aggregate(s=Sum('amount_paid'))['s'] or Decimal('0.00')
    )
    if total_paid == 0:
        inv_total = Invoice.objects.filter(patient=patient, is_deleted=False).exclude(status='cancelled').aggregate(s=Sum('total_amount'))['s'] or Decimal('0.00')
        inv_balance = Invoice.objects.filter(patient=patient, is_deleted=False).exclude(status='cancelled').aggregate(s=Sum('balance'))['s'] or Decimal('0.00')
        if (inv_total - inv_balance) > 0:
            total_paid = inv_total - inv_balance

    # Deposit applied (deduped: matches Invoice.calculate_totals — not receipt sum + DA sum)
    deposit_applied_to_bill = Decimal('0.00')
    if unpaid:
        from .services.deposit_payment_service import deposit_amount_applied_to_invoices_for_display
        deposit_applied_to_bill = deposit_amount_applied_to_invoices_for_display(unpaid)

    # Pending pharmacy prescriptions (not yet on invoice or not paid) – show on combined bill
    pending_pharmacy_prescriptions = []
    try:
        from .models_payment_verification import PharmacyDispensing
        from .models_accounting import PaymentReceipt
        from .utils_billing import get_drug_price_for_prescription
        from .services.auto_billing_service import AutoBillingService
        rx_query = Prescription.objects.filter(
            order__encounter__patient=patient,
            is_deleted=False
        ).select_related(
            'drug',
            'order__encounter',
            'order__encounter__patient',
            'dispensing_record',
            'dispensing_record__substitute_drug',
        ).distinct()
        for rx in rx_query:
            rx_paid = False
            try:
                if hasattr(rx, 'dispensing_record'):
                    disp = rx.dispensing_record
                    if disp and getattr(disp, 'payment_receipt_id', None) is not None:
                        rx_paid = True
            except (PharmacyDispensing.DoesNotExist, AttributeError):
                pass
            if not rx_paid and InvoiceLine.objects.filter(
                prescription_id=rx.id, invoice__patient=patient, invoice__status='paid',
                invoice__is_deleted=False, is_deleted=False
            ).exists():
                rx_paid = True
            if not rx_paid:
                for rec in PaymentReceipt.objects.filter(
                    patient=patient, is_deleted=False,
                    service_type__in=('pharmacy', 'pharmacy_prescription')
                )[:20]:
                    d = getattr(rec, 'service_details', None) or {}
                    if isinstance(d, dict) and d.get('prescription_id') == str(rx.id):
                        rx_paid = True
                        break
            rx_on_unpaid = InvoiceLine.objects.filter(
                prescription_id=rx.id, invoice__patient=patient, invoice__is_deleted=False,
                invoice__status__in=('issued', 'draft', 'partially_paid'),
                invoice__balance__gt=0, is_deleted=False
            ).exists()
            if rx_paid or rx_on_unpaid:
                continue
            encounter = getattr(getattr(rx, 'order', None), 'encounter', None)
            payer = AutoBillingService._ensure_payer(patient, encounter)
            price_drug = rx.drug
            eff_qty = rx.quantity
            try:
                disp = getattr(rx, 'dispensing_record', None)
                if disp and getattr(disp, 'drug_to_dispense', None):
                    price_drug = disp.drug_to_dispense
                if disp and (disp.quantity_ordered or 0) > 0:
                    eff_qty = disp.quantity_ordered
            except Exception:
                pass
            unit_price = get_drug_price_for_prescription(price_drug, payer=payer)
            if not unit_price or unit_price == 0:
                li = InvoiceLine.objects.filter(
                    prescription_id=rx.id, invoice__patient=patient,
                    invoice__is_deleted=False, is_deleted=False
                ).first()
                if li:
                    li_up = getattr(li, 'unit_price', Decimal('0.00')) or Decimal('0.00')
                    if li_up > 0:
                        unit_price = li_up
            qty = Decimal(str(eff_qty))
            line_total = unit_price * qty
            drug_display = getattr(rx.drug, 'name', '') or 'Drug'
            if getattr(rx.drug, 'strength', None):
                drug_display = f"{rx.drug.name} {rx.drug.strength}"
            pending_pharmacy_prescriptions.append({
                'description': drug_display,
                'quantity': qty,
                'unit_price': unit_price,
                'amount': line_total,
                'date': getattr(rx.order, 'created', None) if getattr(rx, 'order', None) else None,
            })
    except Exception:
        pass

    # Pending Prescribe Sales (pharmacy walk-in) with line-item breakdown
    pending_prescribe_sales = []
    try:
        from .models_pharmacy_walkin import WalkInPharmacySale, WalkInPharmacySaleItem
        walkin_sales = WalkInPharmacySale.objects.filter(
            patient=patient,
            is_deleted=False,
            payment_status__in=['pending', 'partial'],
            waived_at__isnull=True
        ).filter(amount_due__gt=0)
        for sale in walkin_sales:
            if getattr(sale, 'amount_due', 0) <= 0:
                continue
            items = WalkInPharmacySaleItem.objects.filter(
                sale=sale, is_deleted=False
            ).select_related('drug').order_by('created')
            breakdown = []
            for item in items:
                drug_display = item.drug.name
                if getattr(item.drug, 'strength', None):
                    drug_display = f"{item.drug.name} {item.drug.strength}"
                breakdown.append({
                    'description': drug_display,
                    'quantity': item.quantity,
                    'unit_price': item.unit_price,
                    'amount': item.line_total,
                })
            if not breakdown:
                breakdown = [{'description': f"Prescribe Sale {getattr(sale, 'sale_number', '')}", 'quantity': 1, 'unit_price': sale.amount_due, 'amount': sale.amount_due}]
            pending_prescribe_sales.append({
                'sale_id': sale.pk,
                'sale_number': getattr(sale, 'sale_number', ''),
                'total': getattr(sale, 'amount_due', sale.total_amount),
                'date': getattr(sale, 'sale_date', None),
                'breakdown': breakdown,
            })
    except Exception:
        pass

    # All deposits on this patient's account (for statement / combined bill)
    patient_deposits = []
    try:
        from .models_patient_deposits import PatientDeposit
        patient_deposits = list(
            PatientDeposit.objects.filter(patient=patient, is_deleted=False).order_by('-deposit_date')
        )
    except Exception:
        pass

    can_waive = user_can_waive(request.user)
    try:
        from .models_settings import HospitalSettings
        hospital_settings = HospitalSettings.get_settings()
    except Exception:
        hospital_settings = None
    context = {
        'patient': patient,
        'unpaid_invoices': unpaid,
        'all_lines': all_lines,
        'lines_by_category': lines_by_category,
        'numbered_by_category': numbered_by_category,
        'total_balance': total_balance,
        'total_amount': total_amount,
        'total_paid': total_paid,
        'deposit_applied_to_bill': deposit_applied_to_bill,
        'patient_deposits': patient_deposits,
        'can_waive': can_waive,
        'pending_pharmacy_prescriptions': pending_pharmacy_prescriptions,
        'pending_prescribe_sales': pending_prescribe_sales,
        'hospital_settings': hospital_settings,
    }
    return render(request, 'hospital/invoice_combined_patient.html', context)


@login_required
@block_pharmacy_from_invoice_and_payment
def update_invoice_line_quantity(request):
    """Update quantity for an invoice line. Redirects back to referrer (invoice or cashier detail)."""
    from decimal import Decimal
    from .models import InvoiceLine

    if request.method != 'POST':
        return redirect('hospital:invoice_list')
    
    line_id = request.POST.get('line_id')
    quantity = request.POST.get('quantity')
    
    if not line_id or quantity is None or quantity == '':
        messages.error(request, 'Line and quantity are required.')
        redirect_url = request.META.get('HTTP_REFERER') or reverse('hospital:invoice_list')
        return redirect(redirect_url)
    
    try:
        qty = Decimal(str(quantity))
        if qty <= 0:
            messages.error(request, 'Quantity must be greater than zero.')
            redirect_url = request.META.get('HTTP_REFERER') or reverse('hospital:invoice_list')
            return redirect(redirect_url)
    except (ValueError, TypeError):
        messages.error(request, 'Invalid quantity.')
        redirect_url = request.META.get('HTTP_REFERER') or reverse('hospital:invoice_list')
        return redirect(redirect_url)
    
    try:
        line = InvoiceLine.objects.select_related('invoice', 'service_code').get(pk=line_id, is_deleted=False)
    except (InvoiceLine.DoesNotExist, ValueError):
        messages.error(request, 'Invoice line not found.')
        redirect_url = request.META.get('HTTP_REFERER') or reverse('hospital:invoice_list')
        return redirect(redirect_url)
    
    # Lab lines: quantity is fixed at 1, cannot be changed once billed (sent to cashier is final)
    if line.service_code and getattr(line.service_code, 'code', '').startswith('LAB-'):
        messages.error(
            request,
            'Lab test quantity cannot be changed. Each lab result is billed once (quantity 1). '
            'Once sent to cashier from the lab, the invoice line is final.'
        )
        redirect_url = request.META.get('HTTP_REFERER') or reverse('hospital:invoice_detail', args=[line.invoice_id])
        return redirect(redirect_url)
    
    if line.waived_at:
        messages.error(request, 'Cannot edit quantity for a waived line.')
        redirect_url = request.META.get('HTTP_REFERER') or reverse('hospital:invoice_list')
        return redirect(redirect_url)
    
    line.quantity = qty
    line.save()  # save() recalculates line_total; signal updates invoice totals
    
    messages.success(request, f'Quantity updated to {qty}')
    redirect_url = request.META.get('HTTP_REFERER') or reverse('hospital:invoice_detail', args=[line.invoice_id])
    return redirect(redirect_url)


@login_required
@block_pharmacy_from_invoice_and_payment
def invoice_print(request, pk):
    """Printable invoice view with detailed services"""
    invoice = get_object_or_404(
        Invoice.all_objects.select_related('patient', 'payer', 'encounter'),
        pk=pk,
        is_deleted=False,
    )
    try:
        from .utils_invoice_line import heal_invoice_zero_line_prices
        heal_invoice_zero_line_prices(invoice)
    except Exception:
        import logging
        logging.getLogger(__name__).exception('heal_invoice_zero_line_prices failed')
    invoice.update_totals()

    # Get hospital settings for logo and name
    from .models_settings import HospitalSettings
    from django.conf import settings as django_settings
    hospital_settings = None
    try:
        hospital_settings = HospitalSettings.get_settings()
        # Ensure we have a valid hospital_settings object
        if not hospital_settings:
            hospital_settings = HospitalSettings.objects.create(
                hospital_name=getattr(django_settings, 'HOSPITAL_NAME', 'PrimeCare Hospital'),
                hospital_tagline='Professional Healthcare Services'
            )
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error loading hospital settings: {e}", exc_info=True)
        # Fallback to settings file if model fails
        class FallbackSettings:
            def __init__(self):
                self.hospital_name = getattr(django_settings, 'HOSPITAL_NAME', 'PrimeCare Hospital')
                self.hospital_tagline = 'Professional Healthcare Services'
                self.logo = None
                self.logo_width = 150
                self.logo_height = 150
                self.address = ''
                self.city = ''
                self.country = 'Ghana'
                self.phone = ''
                self.email = ''
        hospital_settings = FallbackSettings()

    cover_page = (request.GET.get('cover_page') or '').lower() in ('1', 'true', 'yes')
    if cover_page:
        from .models_workflow import Bill as WorkflowBill
        group_bills = list(
            WorkflowBill.objects.filter(invoice=invoice, is_deleted=False)
            .select_related('patient', 'encounter', 'issued_by')
            .order_by('issued_at')
        )
        lines = list(
            invoice.lines.filter(is_deleted=False, waived_at__isnull=True).select_related('service_code')
        )
        lines_by_category = _group_invoice_lines_by_category(lines)
        context = {
            'invoice': invoice,
            'group_bills': group_bills,
            'lines': lines,
            'lines_by_category': lines_by_category,
            'now': timezone.now(),
            'hospital_settings': hospital_settings,
        }
        return render(request, 'hospital/billing/invoice_group_cover_print.html', context)

    # Exclude waived lines - they contribute 0 to total and should not appear on printed invoice
    lines = list(invoice.lines.filter(is_deleted=False).filter(waived_at__isnull=True).select_related('service_code'))
    lines_by_category = _group_invoice_lines_by_category(lines)
    as_statement = request.GET.get('as') == 'statement'

    context = {
        'invoice': invoice,
        'lines': lines,
        'lines_by_category': lines_by_category,
        'as_statement': as_statement,
        'now': timezone.now(),
        'hospital_settings': hospital_settings,
    }
    return render(request, 'hospital/invoice_print.html', context)


@login_required
def bed_availability(request):
    """Bed availability view"""
    ward_filter = request.GET.get('ward', '')
    
    beds = Bed.objects.filter(is_deleted=False).select_related('ward').order_by('ward', 'bed_number')
    
    if ward_filter:
        beds = beds.filter(ward_id=ward_filter)
    
    # Group by ward
    from .models import Ward
    wards = Ward.objects.filter(is_active=True, is_deleted=False)
    
    bed_stats = {}
    for ward in wards:
        ward_beds = beds.filter(ward=ward)
        bed_stats[ward] = {
            'total': ward_beds.count(),
            'available': ward_beds.filter(status='available').count(),
            'occupied': ward_beds.filter(status='occupied').count(),
        }
    
    context = {
        'beds': beds,
        'wards': wards,
        'bed_stats': bed_stats,
        'ward_filter': ward_filter,
    }
    return render(request, 'hospital/bed_availability.html', context)


@login_required
def daily_report(request):
    """Daily activity report with real-time data"""
    from datetime import date
    from django.utils import timezone
    
    report_date = request.GET.get('date')
    selected_date = None
    
    if report_date:
        try:
            selected_date = date.fromisoformat(report_date)
        except ValueError:
            selected_date = None
    
    # Default to today if no date provided
    if not selected_date:
        selected_date = timezone.now().date()
    
    from .utils import generate_daily_report
    report = generate_daily_report(selected_date)
    
    context = {
        'report': report,
        'selected_date': selected_date,
        'report_date': selected_date,  # For backward compatibility
    }
    return render(request, 'hospital/daily_report.html', context)


@login_required
def api_stats(request):
    """API endpoint for comprehensive dashboard statistics with real-time updates"""
    from decimal import Decimal
    from .models_accounting import PaymentReceipt
    from .models_advanced import ImagingStudy
    
    stats = get_dashboard_stats()
    today_date = timezone.now().date()
    
    # Financial stats
    today_payments = PaymentReceipt.objects.filter(
        receipt_date__date=today_date,
        is_deleted=False
    )
    today_revenue = today_payments.aggregate(Sum('amount_paid'))['amount_paid__sum'] or Decimal('0')
    
    # Department stats
    lab_pending = LabResult.objects.filter(
        status__in=['pending', 'in_progress'],
        is_deleted=False
    ).count()
    
    pharmacy_pending = Prescription.objects.filter(
        is_deleted=False
    ).count()
    
    imaging_pending = 0
    try:
        imaging_pending = ImagingStudy.objects.filter(
            status__in=['pending', 'in_progress'],
            is_deleted=False
        ).count()
    except:
        pass
    
    # Add real-time stats
    stats['today_revenue'] = str(today_revenue)
    stats['today_payment_count'] = today_payments.count()
    stats['lab_pending'] = lab_pending
    stats['pharmacy_pending'] = pharmacy_pending
    stats['imaging_pending'] = imaging_pending
    
    return JsonResponse(stats)


@login_required
def export_patients_csv(request):
    """Export patients to CSV"""
    patients = Patient.objects.filter(is_deleted=False).exclude(id__isnull=True)
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="patients.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['MRN', 'First Name', 'Last Name', 'Date of Birth', 'Gender', 'Phone', 'Email'])
    
    for patient in patients:
        writer.writerow([
            patient.mrn,
            patient.first_name,
            patient.last_name,
            patient.date_of_birth,
            patient.get_gender_display(),
            patient.phone_number,
            patient.email,
        ])
    
    return response


@login_required
def export_invoices_csv(request):
    """Export invoices to CSV"""
    invoices = Invoice.objects.filter(is_deleted=False).select_related('patient')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="invoices.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Invoice Number', 'Patient', 'Issue Date', 'Status', 'Total', 'Balance'])
    
    for invoice in invoices:
        writer.writerow([
            invoice.invoice_number,
            invoice.patient.full_name,
            invoice.issued_at.date(),
            invoice.get_status_display(),
            invoice.total_amount,
            invoice.balance,
        ])
    
    return response


@login_required
def export_encounters_csv(request):
    """Export encounters to CSV"""
    encounters = Encounter.objects.filter(is_deleted=False).defer('current_activity').select_related('patient')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="encounters.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Patient', 'Type', 'Status', 'Started At', 'Ended At', 'Location'])
    
    for encounter in encounters:
        writer.writerow([
            encounter.patient.full_name,
            encounter.get_encounter_type_display(),
            encounter.get_status_display(),
            encounter.started_at,
            encounter.ended_at or '',
            encounter.location.name if encounter.location else '',
        ])
    
    return response


@login_required
def financial_report_view(request):
    """Interactive financial report with export options"""
    period = request.GET.get('period', 'month') or 'month'
    report = generate_financial_report(period)
    context = {
        'report': report,
        'period': period,
    }
    return render(request, 'hospital/financial_report.html', context)


@login_required
def financial_report_export_excel(request):
    """Export financial report to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse(
            'openpyxl is required for Excel export. Please install it and try again.',
            status=500
        )
    
    period = request.GET.get('period', 'month') or 'month'
    report = generate_financial_report(period)
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Financial Report'
    
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = f'Financial Report ({period.title()})'
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    current_row = 2
    ws.cell(row=current_row, column=1, value='Metric').font = Font(bold=True)
    ws.cell(row=current_row, column=2, value='Value').font = Font(bold=True)
    current_row += 1
    
    metrics = [
        ('Date Range', f"{report['date_from']} to {report['date_to']}"),
        ('Total Invoiced', float(report['total_invoiced'])),
        ('Total Collected', float(report['total_collected'])),
        ('Outstanding', float(report['outstanding'])),
        ('Collection Rate (%)', float(report['collection_rate'])),
        ('Invoice Count', report['invoice_count']),
    ]
    for label, value in metrics:
        ws.cell(row=current_row, column=1, value=label)
        ws.cell(row=current_row, column=2, value=value)
        current_row += 1
    
    current_row += 1
    ws.cell(row=current_row, column=1, value='Breakdown by Status').font = Font(bold=True)
    current_row += 1
    ws.cell(row=current_row, column=1, value='Status').font = Font(bold=True)
    ws.cell(row=current_row, column=2, value='Count').font = Font(bold=True)
    ws.cell(row=current_row, column=3, value='Total Amount').font = Font(bold=True)
    current_row += 1
    for item in report['by_status']:
        ws.cell(row=current_row, column=1, value=item['status'])
        ws.cell(row=current_row, column=2, value=item['count'])
        ws.cell(row=current_row, column=3, value=float(item['total'] or 0))
        current_row += 1
    
    current_row += 1
    ws.cell(row=current_row, column=1, value='Breakdown by Payer').font = Font(bold=True)
    current_row += 1
    ws.cell(row=current_row, column=1, value='Payer').font = Font(bold=True)
    ws.cell(row=current_row, column=2, value='Count').font = Font(bold=True)
    ws.cell(row=current_row, column=3, value='Total Amount').font = Font(bold=True)
    current_row += 1
    for item in report['by_payer']:
        ws.cell(row=current_row, column=1, value=item['payer__name'] or 'Unknown')
        ws.cell(row=current_row, column=2, value=item['count'])
        ws.cell(row=current_row, column=3, value=float(item['total'] or 0))
        current_row += 1
    
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        # Filter out MergedCell objects - they don't have .value
        actual_cells = [cell for cell in column_cells if hasattr(cell, 'value')]
        if actual_cells:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in actual_cells)
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = max(15, length + 2)
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="financial_report_{period}.xlsx"'
    return response


@login_required
def financial_report_export_pdf(request):
    """Export financial report to PDF"""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import inch
    except ImportError:
        return HttpResponse(
            'ReportLab is required for PDF export. Please install it and try again.',
            status=500
        )
    
    period = request.GET.get('period', 'month') or 'month'
    report = generate_financial_report(period)
    
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    pdf.setFont('Helvetica-Bold', 16)
    pdf.drawString(1 * inch, height - 1 * inch, f'Financial Report ({period.title()})')
    
    pdf.setFont('Helvetica', 11)
    y = height - 1.5 * inch
    lines = [
        f"Date Range: {report['date_from']} to {report['date_to']}",
        f"Total Invoiced: GHS {report['total_invoiced']}",
        f"Total Collected: GHS {report['total_collected']}",
        f"Outstanding: GHS {report['outstanding']}",
        f"Collection Rate: {report['collection_rate']:.2f}%",
        f"Invoice Count: {report['invoice_count']}",
    ]
    for line in lines:
        pdf.drawString(1 * inch, y, line)
        y -= 0.25 * inch
    
    y -= 0.25 * inch
    pdf.setFont('Helvetica-Bold', 12)
    pdf.drawString(1 * inch, y, 'Breakdown by Status')
    y -= 0.3 * inch
    pdf.setFont('Helvetica', 10)
    pdf.drawString(1 * inch, y, 'Status')
    pdf.drawString(3 * inch, y, 'Count')
    pdf.drawString(4 * inch, y, 'Total (GHS)')
    y -= 0.2 * inch
    for item in report['by_status']:
        pdf.drawString(1 * inch, y, str(item['status']))
        pdf.drawString(3 * inch, y, str(item['count']))
        pdf.drawString(4 * inch, y, f"{item['total'] or 0:.2f}")
        y -= 0.2 * inch
        if y < 1 * inch:
            pdf.showPage()
            y = height - 1 * inch
    
    y -= 0.2 * inch
    pdf.setFont('Helvetica-Bold', 12)
    pdf.drawString(1 * inch, y, 'Breakdown by Payer')
    y -= 0.3 * inch
    pdf.setFont('Helvetica', 10)
    pdf.drawString(1 * inch, y, 'Payer')
    pdf.drawString(3 * inch, y, 'Count')
    pdf.drawString(4 * inch, y, 'Total (GHS)')
    y -= 0.2 * inch
    for item in report['by_payer']:
        payer = item['payer__name'] or 'Unknown'
        pdf.drawString(1 * inch, y, payer[:30])
        pdf.drawString(3 * inch, y, str(item['count']))
        pdf.drawString(4 * inch, y, f"{item['total'] or 0:.2f}")
        y -= 0.2 * inch
        if y < 1 * inch:
            pdf.showPage()
            y = height - 1 * inch
    
    pdf.showPage()
    pdf.save()
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="financial_report_{period}.pdf"'
    return response


@login_required
def patient_statistics_report_view(request):
    """Patient statistics report"""
    from .utils import get_patient_demographics
    report = get_patient_demographics()
    context = {'report': report}
    return render(request, 'hospital/patient_statistics_report.html', context)


@login_required
def encounter_report_view(request):
    """Encounter report view"""
    from .utils import get_encounter_statistics
    report = get_encounter_statistics()
    context = {'report': report}
    return render(request, 'hospital/encounter_report.html', context)


@login_required
def admission_report_view(request):
    """Admission report view"""
    from .reports import generate_admission_report
    
    # Get period from request, default to 'month'
    period = request.GET.get('period', 'month')
    
    # Generate the report with proper data structure
    report = generate_admission_report(period=period)
    
    context = {
        'report': report,
        'period': period,
    }
    return render(request, 'hospital/admission_report.html', context)


@login_required
def department_performance_report_view(request):
    """Department performance report"""
    from .models import Department, Encounter, Staff, Ward
    from django.db.models import Count, Q
    from django.utils import timezone
    from datetime import timedelta
    
    # Get period filter
    period = request.GET.get('period', 'month')
    today = timezone.now().date()
    
    # Calculate date range based on period
    if period == 'today':
        date_from = today
        date_to = today
    elif period == 'week':
        date_from = today - timedelta(days=7)
        date_to = today
    elif period == 'month':
        date_from = today.replace(day=1)
        date_to = today
    elif period == 'quarter':
        quarter = (today.month - 1) // 3
        date_from = today.replace(month=quarter * 3 + 1, day=1)
        date_to = today
    elif period == 'year':
        date_from = today.replace(month=1, day=1)
        date_to = today
    else:
        date_from = today.replace(day=1)
        date_to = today
    
    # Get all active departments
    departments = Department.objects.filter(is_active=True, is_deleted=False).order_by('name')
    
    # Calculate performance metrics for each department
    performance_data = []
    total_encounters = 0
    total_staff = 0
    
    for dept in departments:
        # Count encounters in period
        encounters = Encounter.objects.filter(
            provider__department=dept,
            is_deleted=False,
            started_at__date__gte=date_from,
            started_at__date__lte=date_to
        ).count()
        
        # Count active staff
        staff_count = dept.staff.filter(is_active=True, is_deleted=False).count()
        
        # Count wards
        try:
            wards_count = dept.wards.filter(is_active=True, is_deleted=False).count()
        except:
            wards_count = 0
        
        # Calculate encounters per staff (productivity metric)
        encounters_per_staff = round(encounters / staff_count, 2) if staff_count > 0 else 0
        
        performance_data.append({
            'department': dept,
            'name': dept.name,
            'code': dept.code or '-',
            'head': dept.head_of_department.user.get_full_name() if dept.head_of_department else '-',
            'encounters': encounters,
            'staff_count': staff_count,
            'wards_count': wards_count,
            'encounters_per_staff': encounters_per_staff,
        })
        
        total_encounters += encounters
        total_staff += staff_count
    
    # Calculate averages
    avg_encounters_per_dept = round(total_encounters / len(performance_data), 2) if performance_data else 0
    avg_staff_per_dept = round(total_staff / len(performance_data), 2) if performance_data else 0
    
    report = {
        'period': period,
        'date_from': date_from,
        'date_to': date_to,
        'departments': performance_data,
        'total_departments': len(performance_data),
        'total_encounters': total_encounters,
        'total_staff': total_staff,
        'avg_encounters_per_dept': avg_encounters_per_dept,
        'avg_staff_per_dept': avg_staff_per_dept,
    }
    
    context = {'report': report}
    return render(request, 'hospital/department_performance_report.html', context)


@login_required
@login_required
def bed_utilization_report_view(request):
    """Bed utilization report"""
    from .reports import generate_bed_utilization_report
    
    # Generate the report with proper data structure
    report = generate_bed_utilization_report()
    
    context = {'report': report}
    return render(request, 'hospital/bed_utilization_report.html', context)


@login_required
def global_search(request):
    """Enhanced global search across all models with filters"""
    query = request.GET.get('q', '').strip()
    
    # Filter parameters
    category_filter = request.GET.get('category', '').strip()
    status_filter = request.GET.get('status', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    
    # Results per category limit
    limit_per_category = 20
    
    results = {
        'patients': [],
        'encounters': [],
        'invoices': [],
        'appointments': [],
        'staff': [],
        'drugs': [],
        'lab_results': [],
        'prescriptions': [],
        'orders': [],
        'medical_records': [],
        'referrals': [],
        'insurance': [],
        'insurance_companies': [],
        'departments': [],
        'specialties': [],
        'wards': [],
        'beds': [],
    }
    
    result_counts = {
        'patients': 0,
        'encounters': 0,
        'invoices': 0,
        'appointments': 0,
        'staff': 0,
        'drugs': 0,
        'lab_results': 0,
        'prescriptions': 0,
        'orders': 0,
        'medical_records': 0,
        'referrals': 0,
        'insurance': 0,
        'insurance_companies': 0,
        'departments': 0,
        'specialties': 0,
        'wards': 0,
        'beds': 0,
    }
    
    if query and len(query) >= 2:
        # Build date filter if provided
        date_filter = {}
        if date_from:
            try:
                from datetime import datetime
                date_filter['start'] = datetime.strptime(date_from, '%Y-%m-%d')
            except ValueError:
                pass
        if date_to:
            try:
                from datetime import datetime
                date_filter['end'] = datetime.strptime(date_to, '%Y-%m-%d')
            except ValueError:
                pass
        
        # Search patients
        if not category_filter or category_filter == 'patients':
            qs = Patient.objects.filter(
                Q(first_name__icontains=query) |
                Q(last_name__icontains=query) |
                Q(middle_name__icontains=query) |
                Q(mrn__icontains=query) |
                Q(phone_number__icontains=query) |
                Q(national_id__icontains=query) |
                Q(email__icontains=query),
                is_deleted=False
            ).order_by('-created')[:limit_per_category]
            results['patients'] = list(qs)
            result_counts['patients'] = Patient.objects.filter(
                Q(first_name__icontains=query) |
                Q(last_name__icontains=query) |
                Q(middle_name__icontains=query) |
                Q(mrn__icontains=query) |
                Q(phone_number__icontains=query) |
                Q(national_id__icontains=query) |
                Q(email__icontains=query),
                is_deleted=False
            ).count()
        
        # Search encounters
        if not category_filter or category_filter == 'encounters':
            qs = Encounter.objects.filter(
                Q(patient__first_name__icontains=query) |
                Q(patient__last_name__icontains=query) |
                Q(patient__mrn__icontains=query) |
                Q(chief_complaint__icontains=query) |
                Q(notes__icontains=query),
                is_deleted=False
            )
            if status_filter:
                qs = qs.filter(status=status_filter)
            if date_from:
                qs = qs.filter(started_at__gte=date_filter.get('start'))
            if date_to:
                qs = qs.filter(started_at__lte=date_filter.get('end'))
            qs = qs.select_related('patient', 'provider__user', 'provider__department', 'location').order_by('-started_at')[:limit_per_category]
            results['encounters'] = list(qs)
            result_counts['encounters'] = Encounter.objects.filter(
                Q(patient__first_name__icontains=query) |
                Q(patient__last_name__icontains=query) |
                Q(patient__mrn__icontains=query) |
                Q(chief_complaint__icontains=query) |
                Q(notes__icontains=query),
                is_deleted=False
            ).count()
        
        # Search invoices
        if not category_filter or category_filter == 'invoices':
            qs = Invoice.objects.filter(
                Q(patient__first_name__icontains=query) |
                Q(patient__last_name__icontains=query) |
                Q(invoice_number__icontains=query),
                is_deleted=False
            )
            if status_filter:
                qs = qs.filter(status=status_filter)
            if date_from:
                qs = qs.filter(issued_at__gte=date_filter.get('start'))
            if date_to:
                qs = qs.filter(issued_at__lte=date_filter.get('end'))
            qs = qs.select_related('patient', 'payer').order_by('-issued_at')[:limit_per_category]
            results['invoices'] = list(qs)
            result_counts['invoices'] = Invoice.objects.filter(
                Q(patient__first_name__icontains=query) |
                Q(patient__last_name__icontains=query) |
                Q(invoice_number__icontains=query),
                is_deleted=False
            ).count()
        
        # Search appointments
        if not category_filter or category_filter == 'appointments':
            from .models import Appointment
            qs = Appointment.objects.filter(
                Q(patient__first_name__icontains=query) |
                Q(patient__last_name__icontains=query) |
                Q(provider__user__first_name__icontains=query) |
                Q(provider__user__last_name__icontains=query) |
                Q(reason__icontains=query),
                is_deleted=False
            )
            if status_filter:
                qs = qs.filter(status=status_filter)
            if date_from:
                qs = qs.filter(appointment_date__gte=date_filter.get('start'))
            if date_to:
                qs = qs.filter(appointment_date__lte=date_filter.get('end'))
            qs = qs.select_related('patient', 'provider__user', 'department').order_by('-appointment_date')[:limit_per_category]
            results['appointments'] = list(qs)
            result_counts['appointments'] = Appointment.objects.filter(
                Q(patient__first_name__icontains=query) |
                Q(patient__last_name__icontains=query) |
                Q(provider__user__first_name__icontains=query) |
                Q(provider__user__last_name__icontains=query) |
                Q(reason__icontains=query),
                is_deleted=False
            ).count()
        
        # Search staff
        if not category_filter or category_filter == 'staff':
            from .models import Staff
            from .models_advanced import LeaveRequest
            from django.utils import timezone
            from django.db.models import OuterRef, Subquery
            
            # Get the most recent staff record ID for each user to avoid duplicates
            from django.db.models import OuterRef, Subquery
            base_qs = Staff.objects.filter(is_deleted=False)
            if status_filter == 'active':
                base_qs = base_qs.filter(is_active=True)
            elif status_filter == 'inactive':
                base_qs = base_qs.filter(is_active=False)
            
            latest_staff = Staff.objects.filter(
                is_deleted=False,
                user=OuterRef('user')
            )
            if status_filter == 'active':
                latest_staff = latest_staff.filter(is_active=True)
            elif status_filter == 'inactive':
                latest_staff = latest_staff.filter(is_active=False)
            latest_staff = latest_staff.order_by('-created')[:1]
            
            latest_staff_ids = base_qs.annotate(
                latest_id=Subquery(latest_staff.values('id'))
            ).values_list('latest_id', flat=True).distinct()
            
            # latest_staff_ids already includes status filter, so we don't need to apply it again
            qs = Staff.objects.filter(
                id__in=latest_staff_ids,
                is_deleted=False
            ).filter(
                Q(user__first_name__icontains=query) |
                Q(user__last_name__icontains=query) |
                Q(user__username__icontains=query) |
                Q(user__email__icontains=query) |
                Q(employee_id__icontains=query) |
                Q(registration_number__icontains=query) |
                Q(phone_number__icontains=query)
            )
            qs = qs.select_related('user', 'department', 'leave_balance').order_by('user__last_name', 'user__first_name')[:limit_per_category]
            
            # Add current leave information to each staff member
            staff_list = list(qs)
            today = timezone.now().date()
            for staff in staff_list:
                # Get current approved leave (if any)
                current_leave = LeaveRequest.objects.filter(
                    staff=staff,
                    status='approved',
                    start_date__lte=today,
                    end_date__gte=today,
                    is_deleted=False
                ).first()
                staff.current_leave = current_leave
            
            results['staff'] = staff_list
            result_counts['staff'] = Staff.objects.filter(
                Q(user__first_name__icontains=query) |
                Q(user__last_name__icontains=query) |
                Q(user__username__icontains=query) |
                Q(user__email__icontains=query) |
                Q(employee_id__icontains=query) |
                Q(registration_number__icontains=query) |
                Q(phone_number__icontains=query),
                is_deleted=False
            ).count()
        
        # Search drugs
        if not category_filter or category_filter == 'drugs':
            qs = Drug.objects.filter(
                Q(name__icontains=query) |
                Q(generic_name__icontains=query) |
                Q(atc_code__icontains=query) |
                Q(strength__icontains=query) |
                Q(form__icontains=query),
                is_deleted=False
            )
            if status_filter == 'active':
                qs = qs.filter(is_active=True)
            elif status_filter == 'inactive':
                qs = qs.filter(is_active=False)
            qs = qs.order_by('name')[:limit_per_category]
            results['drugs'] = list(qs)
            result_counts['drugs'] = Drug.objects.filter(
                Q(name__icontains=query) |
                Q(generic_name__icontains=query) |
                Q(atc_code__icontains=query) |
                Q(strength__icontains=query) |
                Q(form__icontains=query),
                is_deleted=False
            ).count()
        
        # Search lab results
        if not category_filter or category_filter == 'lab_results':
            from .models import LabResult, Order
            qs = LabResult.objects.filter(
                Q(order__encounter__patient__first_name__icontains=query) |
                Q(order__encounter__patient__last_name__icontains=query) |
                Q(order__encounter__patient__mrn__icontains=query) |
                Q(test__name__icontains=query) |
                Q(test__code__icontains=query) |
                Q(value__icontains=query),
                is_deleted=False
            )
            if status_filter:
                qs = qs.filter(status=status_filter)
            if date_from:
                qs = qs.filter(created__gte=date_filter.get('start'))
            if date_to:
                qs = qs.filter(created__lte=date_filter.get('end'))
            qs = qs.select_related('order__encounter__patient', 'test', 'order__requested_by__user').order_by('-created')[:limit_per_category]
            results['lab_results'] = list(qs)
            result_counts['lab_results'] = LabResult.objects.filter(
                Q(order__encounter__patient__first_name__icontains=query) |
                Q(order__encounter__patient__last_name__icontains=query) |
                Q(order__encounter__patient__mrn__icontains=query) |
                Q(test__name__icontains=query) |
                Q(test__code__icontains=query) |
                Q(value__icontains=query),
                is_deleted=False
            ).count()
        
        # Search prescriptions
        if not category_filter or category_filter == 'prescriptions':
            from .models import Prescription
            qs = Prescription.objects.filter(
                Q(order__encounter__patient__first_name__icontains=query) |
                Q(order__encounter__patient__last_name__icontains=query) |
                Q(order__encounter__patient__mrn__icontains=query) |
                Q(drug__name__icontains=query) |
                Q(drug__generic_name__icontains=query) |
                Q(instructions__icontains=query),
                is_deleted=False
            )
            # Note: Prescription model doesn't have status field
            if date_from:
                qs = qs.filter(created__gte=date_filter.get('start'))
            if date_to:
                qs = qs.filter(created__lte=date_filter.get('end'))
            qs = qs.select_related('order__encounter__patient', 'drug', 'prescribed_by__user').order_by('-created')[:limit_per_category]
            results['prescriptions'] = list(qs)
            result_counts['prescriptions'] = Prescription.objects.filter(
                Q(order__encounter__patient__first_name__icontains=query) |
                Q(order__encounter__patient__last_name__icontains=query) |
                Q(order__encounter__patient__mrn__icontains=query) |
                Q(drug__name__icontains=query) |
                Q(drug__generic_name__icontains=query) |
                Q(instructions__icontains=query),
                is_deleted=False
            ).count()
        
        # Search orders
        if not category_filter or category_filter == 'orders':
            from .models import Order
            qs = Order.objects.filter(
                Q(encounter__patient__first_name__icontains=query) |
                Q(encounter__patient__last_name__icontains=query) |
                Q(encounter__patient__mrn__icontains=query) |
                Q(notes__icontains=query),
                is_deleted=False
            )
            if status_filter:
                qs = qs.filter(status=status_filter)
            if date_from:
                qs = qs.filter(requested_at__gte=date_filter.get('start'))
            if date_to:
                qs = qs.filter(requested_at__lte=date_filter.get('end'))
            qs = qs.select_related('encounter__patient', 'requested_by__user', 'requested_by__department').order_by('-requested_at')[:limit_per_category]
            results['orders'] = list(qs)
            result_counts['orders'] = Order.objects.filter(
                Q(encounter__patient__first_name__icontains=query) |
                Q(encounter__patient__last_name__icontains=query) |
                Q(encounter__patient__mrn__icontains=query) |
                Q(notes__icontains=query),
                is_deleted=False
            ).count()
        
        # Search medical records
        if not category_filter or category_filter == 'medical_records':
            from .models import MedicalRecord
            qs = MedicalRecord.objects.filter(
                Q(patient__first_name__icontains=query) |
                Q(patient__last_name__icontains=query) |
                Q(patient__mrn__icontains=query) |
                Q(title__icontains=query) |
                Q(content__icontains=query),
                is_deleted=False
            )
            if date_from:
                qs = qs.filter(created__gte=date_filter.get('start'))
            if date_to:
                qs = qs.filter(created__lte=date_filter.get('end'))
            qs = qs.select_related('patient', 'created_by__user').order_by('-created')[:limit_per_category]
            results['medical_records'] = list(qs)
            result_counts['medical_records'] = MedicalRecord.objects.filter(
                Q(patient__first_name__icontains=query) |
                Q(patient__last_name__icontains=query) |
                Q(patient__mrn__icontains=query) |
                Q(title__icontains=query) |
                Q(content__icontains=query),
                is_deleted=False
            ).count()
        
        # Search insurance
        if not category_filter or category_filter == 'insurance':
            try:
                from .models_insurance import PatientInsurance
                qs = PatientInsurance.objects.filter(
                    Q(patient__first_name__icontains=query) |
                    Q(patient__last_name__icontains=query) |
                    Q(patient__mrn__icontains=query) |
                    Q(insurance_company__name__icontains=query) |
                    Q(policy_number__icontains=query) |
                    Q(member_id__icontains=query),
                    is_deleted=False
                )
                if status_filter == 'active':
                    qs = qs.filter(is_active=True)
                elif status_filter == 'inactive':
                    qs = qs.filter(is_active=False)
                if date_from:
                    qs = qs.filter(effective_date__gte=date_filter.get('start'))
                if date_to:
                    qs = qs.filter(effective_date__lte=date_filter.get('end'))
                qs = qs.select_related('patient', 'insurance_company').order_by('-effective_date')[:limit_per_category]
                results['insurance'] = list(qs)
                result_counts['insurance'] = PatientInsurance.objects.filter(
                    Q(patient__first_name__icontains=query) |
                    Q(patient__last_name__icontains=query) |
                    Q(patient__mrn__icontains=query) |
                    Q(insurance_company__name__icontains=query) |
                    Q(policy_number__icontains=query) |
                    Q(member_id__icontains=query),
                    is_deleted=False
                ).count()
            except ImportError:
                results['insurance'] = []
                result_counts['insurance'] = 0
        
        # Search insurance companies
        if not category_filter or category_filter == 'insurance_companies':
            try:
                from .models_insurance import InsuranceCompany
                qs = InsuranceCompany.objects.filter(
                    Q(name__icontains=query) |
                    Q(code__icontains=query) |
                    Q(contact_phone__icontains=query) |
                    Q(contact_email__icontains=query) |
                    Q(address__icontains=query),
                    is_deleted=False
                )
                if status_filter == 'active':
                    qs = qs.filter(is_active=True)
                elif status_filter == 'inactive':
                    qs = qs.filter(is_active=False)
                qs = qs.order_by('name')[:limit_per_category]
                results['insurance_companies'] = list(qs)
                result_counts['insurance_companies'] = InsuranceCompany.objects.filter(
                    Q(name__icontains=query) |
                    Q(code__icontains=query) |
                    Q(contact_phone__icontains=query) |
                    Q(contact_email__icontains=query) |
                    Q(address__icontains=query),
                    is_deleted=False
                ).count()
            except ImportError:
                results['insurance_companies'] = []
                result_counts['insurance_companies'] = 0
        
        # Search referrals
        if not category_filter or category_filter == 'referrals':
            try:
                from .models_specialists import Referral
                qs = Referral.objects.filter(
                    Q(encounter__patient__first_name__icontains=query) |
                    Q(encounter__patient__last_name__icontains=query) |
                    Q(encounter__patient__mrn__icontains=query) |
                    Q(specialist__staff__user__first_name__icontains=query) |
                    Q(specialist__staff__user__last_name__icontains=query) |
                    Q(specialty__name__icontains=query) |
                    Q(reason__icontains=query) |
                    Q(specialist_notes__icontains=query) |
                    Q(clinical_summary__icontains=query),
                    is_deleted=False
                )
                if status_filter:
                    qs = qs.filter(status=status_filter)
                if date_from:
                    qs = qs.filter(referred_date__gte=date_filter.get('start'))
                if date_to:
                    qs = qs.filter(referred_date__lte=date_filter.get('end'))
                qs = qs.select_related('encounter__patient', 'specialist__staff__user', 'specialty', 'referring_doctor__user').order_by('-referred_date')[:limit_per_category]
                results['referrals'] = list(qs)
                result_counts['referrals'] = Referral.objects.filter(
                    Q(encounter__patient__first_name__icontains=query) |
                    Q(encounter__patient__last_name__icontains=query) |
                    Q(encounter__patient__mrn__icontains=query) |
                    Q(specialist__staff__user__first_name__icontains=query) |
                    Q(specialist__staff__user__last_name__icontains=query) |
                    Q(specialty__name__icontains=query) |
                    Q(reason__icontains=query) |
                    Q(specialist_notes__icontains=query) |
                    Q(clinical_summary__icontains=query),
                    is_deleted=False
                ).count()
            except ImportError:
                results['referrals'] = []
                result_counts['referrals'] = 0
        
        # Search departments
        if not category_filter or category_filter == 'departments':
            from .models import Department
            qs = Department.objects.filter(
                Q(name__icontains=query) |
                Q(code__icontains=query) |
                Q(description__icontains=query),
                is_deleted=False
            )
            if status_filter == 'active':
                qs = qs.filter(is_active=True)
            elif status_filter == 'inactive':
                qs = qs.filter(is_active=False)
            qs = qs.order_by('name')[:limit_per_category]
            results['departments'] = list(qs)
            result_counts['departments'] = Department.objects.filter(
                Q(name__icontains=query) |
                Q(code__icontains=query) |
                Q(description__icontains=query),
                is_deleted=False
            ).count()
        
        # Search specialties
        if not category_filter or category_filter == 'specialties':
            try:
                from .models_specialists import Specialty
                qs = Specialty.objects.filter(
                    Q(name__icontains=query) |
                    Q(code__icontains=query) |
                    Q(description__icontains=query),
                    is_deleted=False
                )
                if status_filter == 'active':
                    qs = qs.filter(is_active=True)
                elif status_filter == 'inactive':
                    qs = qs.filter(is_active=False)
                qs = qs.order_by('name')[:limit_per_category]
                results['specialties'] = list(qs)
                result_counts['specialties'] = Specialty.objects.filter(
                    Q(name__icontains=query) |
                    Q(code__icontains=query) |
                    Q(description__icontains=query),
                    is_deleted=False
                ).count()
            except ImportError:
                results['specialties'] = []
                result_counts['specialties'] = 0
        
        # Search wards
        if not category_filter or category_filter == 'wards':
            from .models import Ward
            qs = Ward.objects.filter(
                Q(name__icontains=query) |
                Q(code__icontains=query) |
                Q(ward_type__icontains=query),
                is_deleted=False
            )
            if status_filter == 'active':
                qs = qs.filter(is_active=True)
            elif status_filter == 'inactive':
                qs = qs.filter(is_active=False)
            qs = qs.select_related('department').order_by('name')[:limit_per_category]
            results['wards'] = list(qs)
            result_counts['wards'] = Ward.objects.filter(
                Q(name__icontains=query) |
                Q(code__icontains=query) |
                Q(ward_type__icontains=query),
                is_deleted=False
            ).count()
        
        # Search beds
        if not category_filter or category_filter == 'beds':
            from .models import Bed
            qs = Bed.objects.filter(
                Q(bed_number__icontains=query) |
                Q(ward__name__icontains=query) |
                Q(ward__code__icontains=query),
                is_deleted=False
            )
            if status_filter == 'available':
                qs = qs.filter(is_occupied=False)
            elif status_filter == 'occupied':
                qs = qs.filter(is_occupied=True)
            qs = qs.select_related('ward', 'ward__department').order_by('ward__name', 'bed_number')[:limit_per_category]
            results['beds'] = list(qs)
            result_counts['beds'] = Bed.objects.filter(
                Q(bed_number__icontains=query) |
                Q(ward__name__icontains=query) |
                Q(ward__code__icontains=query),
                is_deleted=False
            ).count()
    
    # Calculate totals
    total_results = sum(len(v) if isinstance(v, list) else 0 for v in results.values())
    
    context = {
        'query': query,
        'results': results,
        'result_counts': result_counts,
        'has_results': total_results > 0,
        'total_results': total_results,
        'category_filter': category_filter,
        'status_filter': status_filter,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'hospital/global_search.html', context)


@login_required
def tabular_lab_report(request, result_id):
    """Tabular lab report for structured tests (FBC, LFT, RFT). Uses result_id from URL only."""
    lab_result = LabResult.objects.filter(pk=result_id).select_related(
        'order', 'order__encounter', 'order__encounter__patient', 'test'
    ).first()
    if not lab_result:
        return render(
            request,
            'hospital/errors/404.html',
            {
                'message': 'The requested lab result was not found. It may have been removed or the link may be outdated.',
                'dashboard_url': reverse('hospital:laboratory_dashboard'),
                'lab_results_url': reverse('hospital:lab_results_list'),
            },
            status=404,
        )
    if lab_result.is_deleted:
        return render(
            request,
            'hospital/errors/404.html',
            {
                'message': 'This lab result has been removed and is no longer available.',
                'dashboard_url': reverse('hospital:laboratory_dashboard'),
                'lab_results_url': reverse('hospital:lab_results_list'),
            },
            status=404,
        )

    # Get patient and test details (ensure we have a valid order/encounter/patient)
    try:
        patient = lab_result.order.encounter.patient
        test = lab_result.test
    except Exception:
        return render(
            request,
            'hospital/errors/404.html',
            {
                'message': 'This lab result is not linked to a valid patient or encounter.',
                'dashboard_url': reverse('hospital:laboratory_dashboard'),
                'lab_results_url': reverse('hospital:lab_results_list'),
            },
            status=404,
        )

    # Use central mapping so the correct template is shown for this test (FBC, LFT, RFT, etc.)
    test_type = get_lab_result_template_type(test)
    # Ensure Fasting Blood Sugar / FBS always uses glucose form, never LFT (even if test is mis-coded)
    _name = (getattr(test, 'name', None) or '').lower()
    _code = (getattr(test, 'code', None) or '').upper()
    if _code == 'FBS' or 'fasting' in _name or 'blood sugar' in _name or 'fbs' in _name or 'glucose' in _name:
        test_type = 'glucose'

    # Single-value tests: send GET to simple form. POST must stay here so saves are not dropped by redirect.
    if request.method == 'GET' and is_single_value_template(test):
        return redirect('hospital:edit_lab_result', result_id=result_id)

    if request.method == 'POST':
        post_data = request.POST
        # Prevent mixing patients: only accept save for THIS result (URL result_id). Reject if form was from another tab/patient.
        post_result_id = post_data.get('lab_result_id', '').strip()
        if str(result_id) != post_result_id:
            messages.error(
                request,
                'This form does not match the lab result you opened. To avoid mixing patients, please go back to the Lab list, '
                'find the correct patient (check name and MRN), then click Enter on that result.',
            )
            return redirect('hospital:tabular_lab_report', result_id=result_id)
        # Re-fetch by URL result_id only (never trust POST for which record to update)
        lab_result = LabResult.objects.filter(pk=result_id, is_deleted=False).select_related(
            'order', 'order__encounter', 'order__encounter__patient', 'test'
        ).first()
        if not lab_result:
            messages.error(request, 'This lab result is no longer available. Please use the Lab list to open the correct result.')
            return redirect('hospital:laboratory_dashboard')
        patient = lab_result.order.encounter.patient
        test = lab_result.test
        test_type = get_lab_result_template_type(test)
        _name = (getattr(test, 'name', None) or '').lower()
        _code = (getattr(test, 'code', None) or '').upper()
        if _code == 'FBS' or 'fasting' in _name or 'blood sugar' in _name or 'fbs' in _name or 'glucose' in _name:
            test_type = 'glucose'
        decimal_field_names = [
            'wbc', 'lymph_count', 'mid_count', 'gran_count', 'lymph_perc', 'mid_perc', 'gran_perc',
            'plt', 'mpv', 'pdw', 'pct', 'rbc', 'hgb', 'hct', 'mcv', 'mch', 'mchc', 'rdw_cv', 'rdw_sd',
            'neut_perc', 'mono_perc', 'eos_perc', 'baso_perc',
            'total_bili', 'direct_bili', 'indirect_bili', 'alt', 'ast', 'alp', 'ggt',
            'total_protein', 'albumin', 'globulin', 'ag_ratio', 'ast_alt_ratio',
            'urea', 'bun', 'creatinine', 'egfr', 'uric_acid',
            'sodium', 'potassium', 'chloride', 'bicarbonate', 'calcium', 'magnesium', 'phosphorus',
            'total_chol', 'triglycerides', 'hdl', 'ldl', 'vldl', 'chol_hdl_ratio', 'ldl_hdl_ratio', 'non_hdl',
            'tsh', 'free_t4', 'total_t4', 'free_t3', 'total_t3',
            'fbs', 'rbs', 'hba1c', 'ppbs',
            'urine_ph', 'urine_sgravity',
            'coag_pt', 'coag_inr', 'coag_aptt', 'coag_fibrinogen',
            'semen_volume', 'semen_ph', 'semen_count', 'semen_motility', 'semen_morphology',
        ]
        # Bypass form validation - extract data directly from POST to avoid save failures
        def _sanitize_decimal(val):
            if val is None or val == '':
                return None
            s = str(val).strip()
            if s in ('', '-', '.', 'N/A', 'n/a', '--', 'na', 'NA'):
                return None
            s = s.replace(',', '')
            if s.endswith('.'):
                try:
                    part = s[:-1]
                    if part and part not in ('-', '+'):
                        float(part)
                        return part if '.' in part else part + '.0'
                except (ValueError, TypeError):
                    return None
            try:
                float(s)
                return s
            except (ValueError, TypeError):
                return None

        excluded = {'test_type', 'status', 'qualitative_result', 'notes', 'csrfmiddlewaretoken',
                    'lab_result_id', 'verified_by'}
        details = {}
        for key in post_data:
            if key in excluded:
                continue
            val = post_data.get(key)
            if key in decimal_field_names:
                sanitized = _sanitize_decimal(val)
                if sanitized is not None:
                    details[key] = sanitized
            else:
                if val is not None and str(val).strip():
                    details[key] = str(val).strip()

        # Only keep detail keys that belong to THIS test type (stops FBC from storing uric_acid, etc.)
        allowed_keys = get_allowed_detail_keys_for_save(test_type)
        if allowed_keys:
            details = {k: v for k, v in details.items() if k in allowed_keys}

        uploaded_file = request.FILES.get('attachment')
        if uploaded_file:
            lab_result.attachment = uploaded_file

        lab_result.details = details
        lab_result.status = post_data.get('status', 'completed') or 'completed'
        lab_result.qualitative_result = (post_data.get('qualitative_result') or '').strip()
        lab_result.notes = (post_data.get('notes') or '').strip()

        try:
            staff = Staff.objects.get(user=request.user, is_deleted=False)
            lab_result.verified_by = staff
            lab_result.verified_at = timezone.now()
        except Staff.DoesNotExist:
            pass

        lab_result.save()

        try:
            order = lab_result.order
            if order and order.status != 'completed':
                order.status = 'completed'
                order.save(update_fields=['status', 'modified'])
        except Exception:
            pass

        messages.success(request, f'Lab result for {test.name} saved successfully.')
        return redirect('hospital:laboratory_dashboard')
    else:
        # Pre-populate form with existing data
        initial_data = {
            'test_type': test_type,
            'status': lab_result.status,
            'qualitative_result': lab_result.qualitative_result or '',
            'notes': lab_result.notes or '',
        }
        
        # Add existing details if any
        if lab_result.details:
            initial_data.update(lab_result.details)
        
        form = TabularLabReportForm(initial=initial_data)

    # For glucose form: show only the matching row (RBS->rbs only, FBS->fbs only, etc.)
    glucose_show_param = None
    if test_type == 'glucose' and test:
        code = (test.code or '').upper()
        name = (test.name or '').upper()
        if code == 'RBS' or 'RANDOM' in name or '(RBS)' in name:
            glucose_show_param = 'rbs'
        elif code == 'FBS' or 'FASTING' in name or '(FBS)' in name:
            glucose_show_param = 'fbs'
        elif code == 'HBA1C' or 'HBA1C' in name or 'GLYCATED' in name:
            glucose_show_param = 'hba1c'
        elif code == 'OGTT' or code == 'PPBS' or 'POST-PRANDIAL' in name or 'OGTT' in name:
            glucose_show_param = 'ogtt'  # show both fbs and ppbs

    context = {
        'form': form,
        'lab_result': lab_result,
        'lab_result_id': result_id,
        'patient': patient,
        'patient_name': patient.full_name,
        'test': test,
        'test_name': test.name,
        'test_type': test_type,
        'glucose_show_param': glucose_show_param,
        'details': lab_result.details or {},
        'notes': lab_result.notes or '',
        'qualitative_result': lab_result.qualitative_result or '',
        'qualitative_options': ['Seen', 'Not seen', 'Negative', 'Positive', 'Reactive', 'Non-Reactive', 'Normal', 'Abnormal'],
        'current_user': request.user.get_full_name() or request.user.username,
    }
    
    return render(request, 'hospital/lab_report_tabular.html', context)


@login_required
def print_lab_report(request, result_id):
    """Print-friendly lab report with logo and department info"""
    result = get_object_or_404(LabResult, pk=result_id, is_deleted=False)
    settings = HospitalSettings.get_settings()
    patient_gender = None
    if result.order and result.order.encounter and result.order.encounter.patient:
        patient_gender = getattr(result.order.encounter.patient, 'gender', None)

    # Build display rows for TEST RESULTS so single-value tests (Typhidot, etc.) show test name, not WBC
    result_rows = []
    details = result.details or {}
    if isinstance(details, dict):
        result_value = details.get('result_value') or details.get('RESULT_VALUE')
        if result_value:
            # Single-value test: one row with test name and result
            ref = '-'
            if result.range_low and result.range_high:
                ref = f'{result.range_low} - {result.range_high}'
            elif result.range_low:
                ref = result.range_low
            elif result.range_high:
                ref = result.range_high
            units_val = (details.get('result_unit') or details.get('RESULT_UNIT') or result.units or
                        infer_units_from_result(result_value) or 'N/A')
            value_flag = compute_value_flag(result_value, ref, patient_gender)
            flag_val = value_flag if value_flag else get_qualitative_flag(result_value, result.is_abnormal)
            result_rows.append({
                'parameter': result.test.name if result.test else 'Result',
                'result': result_value,
                'units': units_val,
                'ref_range': ref,
                'flag': flag_val,
                'value_flag': value_flag,
            })
        else:
            # Panel test (FBC, LFT, urine, stool): one row per detail, skip result_value/result_unit
            test_type = get_lab_result_template_type(result.test) if result.test else None
            ordered_keys = None
            if test_type == TEMPLATE_FBC:
                ordered_keys = FBC_ORDERED_KEYS
            elif test_type == TEMPLATE_RFT:
                ordered_keys = RFT_ORDERED_KEYS
            elif test_type == TEMPLATE_LFT:
                ordered_keys = LFT_ORDERED_KEYS
            elif test_type == TEMPLATE_URINE:
                ordered_keys = URINE_ORDERED_KEYS
            elif test_type == TEMPLATE_STOOL:
                ordered_keys = STOOL_ORDERED_KEYS
            elif test_type == TEMPLATE_MALARIA:
                ordered_keys = MALARIA_ORDERED_KEYS
            elif test_type == TEMPLATE_BLOOD_GROUP:
                ordered_keys = BLOOD_GROUP_ORDERED_KEYS
            elif test_type == TEMPLATE_SICKLE:
                ordered_keys = SICKLE_ORDERED_KEYS
            elif test_type == TEMPLATE_COAGULATION:
                ordered_keys = COAGULATION_ORDERED_KEYS
            elif test_type == TEMPLATE_SEROLOGY:
                ordered_keys = SEROLOGY_ORDERED_KEYS
            elif test_type == TEMPLATE_SEMEN:
                ordered_keys = SEMEN_ORDERED_KEYS
            elif test_type == TEMPLATE_AFB:
                ordered_keys = AFB_ORDERED_KEYS
            keys_to_iterate = ordered_keys if ordered_keys else list(details.keys())
            for key in keys_to_iterate:
                if key.lower() in ('result_value', 'result_unit'):
                    continue
                if key.endswith('_unit'):
                    continue
                value = details.get(key)
                if value is None or str(value).strip() == '':
                    continue
                param_label = get_param_display_name(key)
                ref_range = get_param_ref_range(key)
                units_val = get_param_units(key, details)
                value_flag = compute_value_flag(str(value), ref_range, patient_gender)
                flag_val = value_flag if value_flag else get_qualitative_flag(str(value), result.is_abnormal)
                result_rows.append({
                    'parameter': param_label,
                    'result': value,
                    'units': units_val,
                    'ref_range': ref_range,
                    'flag': flag_val,
                    'value_flag': value_flag,
                })
    if not result_rows and result.value:
        ref = '-'
        if result.range_low and result.range_high:
            ref = f'{result.range_low} - {result.range_high}'
        elif result.range_low or result.range_high:
            ref = (result.range_low or '') + (' - ' if result.range_low and result.range_high else '') + (result.range_high or '')
        units_val = result.units or 'N/A'
        value_flag = compute_value_flag(result.value, ref, patient_gender)
        flag_val = value_flag if value_flag else ('ABNORMAL' if result.is_abnormal else 'NORMAL')
        result_rows.append({
            'parameter': result.test.name if result.test else 'Result',
            'result': result.value,
            'units': units_val,
            'ref_range': ref,
            'flag': flag_val,
            'value_flag': value_flag,
        })
    if not result_rows and result.qualitative_result:
        flag_val = get_qualitative_flag(result.qualitative_result, result.is_abnormal)
        units_val = infer_units_from_result(result.qualitative_result) or 'Qualitative'
        result_rows.append({
            'parameter': result.test.name if result.test else 'Result',
            'result': result.qualitative_result,
            'units': units_val,
            'ref_range': 'Reactive / Non-Reactive / Equivocal',
            'flag': flag_val,
            'value_flag': None,
        })

    context = {
        'result': result,
        'result_rows': result_rows,
        'settings': settings,
        'now': timezone.now(),
    }

    return render(request, 'hospital/lab_report_print.html', context)


@login_required
def hospital_settings_view(request):
    """Hospital settings configuration page - Accessible by Admin and Medical Director"""
    # Check if user is admin or Medical Director
    is_admin = user_has_role_access(request.user, 'admin')
    is_medical_director = False
    
    # Check if user is Medical Director
    try:
        from .models import Staff
        staff = Staff.objects.filter(user=request.user, is_deleted=False).first()
        if staff:
            specialization = (staff.specialization or '').lower()
            is_medical_director = (
                'medical director' in specialization or
                (request.user.is_staff and staff.profession == 'doctor' and 'director' in specialization)
            )
    except Exception:
        pass
    
    if not (is_admin or is_medical_director or request.user.is_superuser):
        messages.error(request, 'You do not have permission to manage hospital settings. Only Administrators and Medical Directors have access.')
        return redirect('hospital:dashboard')
    
    settings = HospitalSettings.get_settings()
    
    if request.method == 'POST':
        # Update settings
        settings.hospital_name = request.POST.get('hospital_name', settings.hospital_name)
        settings.hospital_tagline = request.POST.get('hospital_tagline', settings.hospital_tagline)
        settings.address = request.POST.get('address', settings.address)
        settings.city = request.POST.get('city', settings.city)
        settings.state = request.POST.get('state', settings.state)
        settings.postal_code = request.POST.get('postal_code', settings.postal_code)
        settings.country = request.POST.get('country', settings.country)
        settings.phone = request.POST.get('phone', settings.phone)
        settings.email = request.POST.get('email', settings.email)
        settings.website = request.POST.get('website', settings.website)
        
        # Lab settings
        settings.lab_department_name = request.POST.get('lab_department_name', settings.lab_department_name)
        settings.lab_phone = request.POST.get('lab_phone', settings.lab_phone)
        settings.lab_email = request.POST.get('lab_email', settings.lab_email)
        settings.lab_accreditation = request.POST.get('lab_accreditation', settings.lab_accreditation)
        settings.lab_license_number = request.POST.get('lab_license_number', settings.lab_license_number)
        
        # Radiology
        settings.radiology_department_name = request.POST.get('radiology_department_name', settings.radiology_department_name)
        settings.radiology_phone = request.POST.get('radiology_phone', settings.radiology_phone)
        settings.radiology_email = request.POST.get('radiology_email', settings.radiology_email)
        
        # Pharmacy
        settings.pharmacy_department_name = request.POST.get('pharmacy_department_name', settings.pharmacy_department_name)
        settings.pharmacy_phone = request.POST.get('pharmacy_phone', settings.pharmacy_phone)
        settings.pharmacy_license_number = request.POST.get('pharmacy_license_number', settings.pharmacy_license_number)
        
        # System (currency & date)
        settings.currency = request.POST.get('currency', settings.currency) or 'GHS'
        settings.currency_symbol = request.POST.get('currency_symbol', settings.currency_symbol) or '₵'
        settings.date_format = request.POST.get('date_format', settings.date_format) or '%d/%m/%Y'
        settings.time_format = request.POST.get('time_format', settings.time_format) or '%H:%M'
        
        # Logo upload
        if 'logo' in request.FILES:
            settings.logo = request.FILES['logo']
        
        settings.logo_width = int(request.POST.get('logo_width', settings.logo_width))
        settings.logo_height = int(request.POST.get('logo_height', settings.logo_height))
        settings.report_header_color = request.POST.get('report_header_color', settings.report_header_color)
        settings.report_footer_text = request.POST.get('report_footer_text', settings.report_footer_text)
        
        # Printer & POS receipt
        try:
            settings.pos_receipt_width_mm = int(request.POST.get('pos_receipt_width_mm', getattr(settings, 'pos_receipt_width_mm', 80)))
        except (TypeError, ValueError):
            pass
        try:
            settings.pos_receipt_printable_width_mm = float(request.POST.get('pos_receipt_printable_width_mm', getattr(settings, 'pos_receipt_printable_width_mm', 78)))
        except (TypeError, ValueError):
            pass
        try:
            settings.pos_receipt_length_mm = int(request.POST.get('pos_receipt_length_mm', getattr(settings, 'pos_receipt_length_mm', 810)))
        except (TypeError, ValueError):
            pass
        try:
            settings.pos_receipt_font_size_body = int(request.POST.get('pos_receipt_font_size_body', getattr(settings, 'pos_receipt_font_size_body', 10)))
        except (TypeError, ValueError):
            pass
        try:
            settings.pos_receipt_font_size_header = int(request.POST.get('pos_receipt_font_size_header', getattr(settings, 'pos_receipt_font_size_header', 12)))
        except (TypeError, ValueError):
            pass
        try:
            settings.pos_receipt_font_size_footer = int(request.POST.get('pos_receipt_font_size_footer', getattr(settings, 'pos_receipt_font_size_footer', 8)))
        except (TypeError, ValueError):
            pass
        try:
            settings.pos_receipt_margin_mm = int(request.POST.get('pos_receipt_margin_mm', getattr(settings, 'pos_receipt_margin_mm', 4)))
        except (TypeError, ValueError):
            pass
        try:
            settings.pos_receipt_padding_mm = int(request.POST.get('pos_receipt_padding_mm', getattr(settings, 'pos_receipt_padding_mm', 3)))
        except (TypeError, ValueError):
            pass
        settings.pos_receipt_show_qr = request.POST.get('pos_receipt_show_qr') in ('on', '1', 'true')
        try:
            settings.pos_receipt_qr_size_px = int(request.POST.get('pos_receipt_qr_size_px', getattr(settings, 'pos_receipt_qr_size_px', 56)))
        except (TypeError, ValueError):
            pass
        try:
            settings.default_print_copies = max(1, int(request.POST.get('default_print_copies', getattr(settings, 'default_print_copies', 1))))
        except (TypeError, ValueError):
            pass
        settings.invoice_paper_size = request.POST.get('invoice_paper_size', getattr(settings, 'invoice_paper_size', 'A4'))
        if settings.invoice_paper_size not in ('A4', 'Letter'):
            settings.invoice_paper_size = 'A4'
        try:
            settings.label_printer_width_mm = int(request.POST.get('label_printer_width_mm', getattr(settings, 'label_printer_width_mm', 60)))
        except (TypeError, ValueError):
            pass
        settings.report_paper_orientation = request.POST.get('report_paper_orientation', getattr(settings, 'report_paper_orientation', 'portrait'))
        if settings.report_paper_orientation not in ('portrait', 'landscape'):
            settings.report_paper_orientation = 'portrait'
        settings.default_printer_name = request.POST.get('default_printer_name', getattr(settings, 'default_printer_name', '')) or ''
        
        # Session & security
        try:
            settings.session_timeout_minutes = int(request.POST.get('session_timeout_minutes', getattr(settings, 'session_timeout_minutes', 30)))
        except (TypeError, ValueError):
            pass
        settings.require_login_for_receipt_verify = request.POST.get('require_login_for_receipt_verify') in ('on', '1', 'true')
        try:
            settings.max_login_attempts = int(request.POST.get('max_login_attempts', getattr(settings, 'max_login_attempts', 5)))
        except (TypeError, ValueError):
            pass
        try:
            settings.lockout_duration_minutes = int(request.POST.get('lockout_duration_minutes', getattr(settings, 'lockout_duration_minutes', 15)))
        except (TypeError, ValueError):
            pass
        
        # Business & display
        settings.timezone = request.POST.get('timezone', getattr(settings, 'timezone', 'Africa/Accra')) or 'Africa/Accra'
        settings.business_hours_start = request.POST.get('business_hours_start', getattr(settings, 'business_hours_start', '08:00')) or '08:00'
        settings.business_hours_end = request.POST.get('business_hours_end', getattr(settings, 'business_hours_end', '17:00')) or '17:00'
        
        # Notifications & integrations
        settings.sms_enabled = request.POST.get('sms_enabled') in ('on', '1', 'true')
        settings.email_notifications_enabled = request.POST.get('email_notifications_enabled') in ('on', '1', 'true')
        try:
            settings.backup_retention_days = int(request.POST.get('backup_retention_days', getattr(settings, 'backup_retention_days', 30)))
        except (TypeError, ValueError):
            pass
        settings.patient_portal_enabled = request.POST.get('patient_portal_enabled') in ('on', '1', 'true')
        settings.show_prices_to_patient = request.POST.get('show_prices_to_patient') in ('on', '1', 'true')
        
        settings.updated_by = request.user
        settings.save()
        
        messages.success(request, 'Hospital settings updated successfully.')
        return redirect('hospital:hospital_settings')
    
    context = {
        'settings': settings,
    }
    
    return render(request, 'hospital/hospital_settings.html', context)


# ==================== DRUG FORMULARY MANAGEMENT ====================

def drug_formulary_filtered_queryset(request):
    """Apply the same GET filters as the drug formulary list (for list, print, PDF)."""
    drugs = Drug.objects.filter(is_deleted=False).order_by('category', 'name')
    search_query = request.GET.get('q', '').strip()
    if search_query:
        drugs = drugs.filter(
            Q(name__icontains=search_query) |
            Q(generic_name__icontains=search_query) |
            Q(atc_code__icontains=search_query)
        )
    category_filter = request.GET.get('category', '')
    if category_filter:
        drugs = drugs.filter(category=category_filter)
    form_filter = request.GET.get('form', '')
    if form_filter:
        drugs = drugs.filter(form__icontains=form_filter)
    status_filter = request.GET.get('status', '')
    if status_filter == 'active':
        drugs = drugs.filter(is_active=True)
    elif status_filter == 'inactive':
        drugs = drugs.filter(is_active=False)
    controlled_filter = request.GET.get('controlled', '')
    if controlled_filter == 'yes':
        drugs = drugs.filter(is_controlled=True)
    elif controlled_filter == 'no':
        drugs = drugs.filter(is_controlled=False)
    return drugs


@login_required
def drug_formulary_list(request):
    """List all drugs in formulary with search and filtering, organized by category"""
    drugs = drug_formulary_filtered_queryset(request)
    search_query = request.GET.get('q', '').strip()
    category_filter = request.GET.get('category', '')
    form_filter = request.GET.get('form', '')
    status_filter = request.GET.get('status', '')
    controlled_filter = request.GET.get('controlled', '')

    # Organize drugs by category
    drugs_by_category = {}
    for drug in drugs:
        category_code = drug.category or 'other'
        if category_code not in drugs_by_category:
            drugs_by_category[category_code] = []
        drugs_by_category[category_code].append(drug)
    
    # Pagination
    paginator = Paginator(drugs, 20)
    page = request.GET.get('page')
    drugs_page = paginator.get_page(page)
    
    # Get unique forms for filter dropdown
    forms = Drug.objects.filter(is_deleted=False).values_list('form', flat=True).distinct().order_by('form')
    
    context = {
        'drugs': drugs_page,
        'drugs_by_category': drugs_by_category,
        'forms': forms,
        'drug_categories': Drug.CATEGORIES,
        'search_query': search_query,
        'category_filter': category_filter,
        'form_filter': form_filter,
        'status_filter': status_filter,
        'controlled_filter': controlled_filter,
        'total_drugs': drugs.count(),
        'can_edit_stock': is_admin_user(request.user),
    }
    return render(request, 'hospital/drug_formulary_list.html', context)


@login_required
def drug_formulary_print(request):
    """Print-friendly full drug list (same filters as formulary list); use browser Print / Save as PDF."""
    from .models_settings import HospitalSettings

    drugs_qs = drug_formulary_filtered_queryset(request)
    drugs_flat = list(drugs_qs.order_by('category', 'name'))

    settings = HospitalSettings.get_settings()
    category_filter_raw = request.GET.get('category', '')
    category_filter_label = (
        dict(Drug.CATEGORIES).get(category_filter_raw, category_filter_raw) if category_filter_raw else ''
    )
    context = {
        'formulary_rows': drugs_flat,
        'search_query': request.GET.get('q', '').strip(),
        'category_filter': category_filter_raw,
        'category_filter_label': category_filter_label,
        'form_filter': request.GET.get('form', ''),
        'status_filter': request.GET.get('status', ''),
        'controlled_filter': request.GET.get('controlled', ''),
        'total_drugs': len(drugs_flat),
        'settings': settings,
        'generated_at': timezone.now(),
    }
    return render(request, 'hospital/drug_formulary_print.html', context)


@login_required
def drug_formulary_pdf(request):
    """Download full drug formulary as PDF (same filters as list). Uses wrapped text for long names/categories."""
    try:
        from xml.sax.saxutils import escape
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        return HttpResponse(
            'ReportLab is required for PDF export. Please install it and try again.',
            status=500,
        )

    def _cell_plain(val, max_len=48):
        s = '' if val is None else str(val).replace('\r', ' ').replace('\n', ' ')
        return s[:max_len] if len(s) > max_len else s

    def _pcell(text, style):
        safe = escape(str(text or '')).replace('\n', '<br/>')
        return Paragraph(safe, style)

    drugs_qs = drug_formulary_filtered_queryset(request)
    drugs = list(drugs_qs.order_by('category', 'name'))
    from django.db.models import Sum, Count

    drug_ids_pdf = [d.pk for d in drugs]
    stock_by_drug_pdf = {}
    if drug_ids_pdf:
        for row in (
            PharmacyStock.objects.filter(is_deleted=False, drug_id__in=drug_ids_pdf)
            .values('drug_id')
            .annotate(total_qty=Sum('quantity_on_hand'), batch_lines=Count('id'))
        ):
            stock_by_drug_pdf[row['drug_id']] = (
                int(row['total_qty'] or 0),
                int(row['batch_lines'] or 0),
            )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        topMargin=0.4 * inch,
        bottomMargin=0.4 * inch,
        leftMargin=0.35 * inch,
        rightMargin=0.35 * inch,
    )
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=14,
        textColor=colors.HexColor('#7c3aed'),
        spaceAfter=8,
        alignment=TA_CENTER,
    )
    meta_style = ParagraphStyle('Meta', parent=styles['Normal'], fontSize=8, textColor=colors.grey)
    p_name = ParagraphStyle(
        'CellName', parent=styles['Normal'], fontSize=7, leading=8.5, alignment=TA_LEFT, spaceAfter=0, spaceBefore=0,
    )
    p_cat = ParagraphStyle(
        'CellCat', parent=styles['Normal'], fontSize=6.5, leading=7.8, alignment=TA_LEFT, spaceAfter=0, spaceBefore=0,
    )
    p_small = ParagraphStyle(
        'CellSm', parent=styles['Normal'], fontSize=7, leading=8, alignment=TA_LEFT, spaceAfter=0, spaceBefore=0,
    )

    elements.append(Paragraph('<b>Drug Formulary</b>', title_style))
    elements.append(Paragraph(f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M")}', meta_style))
    elements.append(Paragraph(f'Total drugs: {len(drugs)}', meta_style))
    fq = request.GET.get('q', '').strip()
    if fq:
        elements.append(Paragraph(f'Filter — search: {_cell_plain(fq, 120)}', meta_style))
    elements.append(Spacer(1, 0.12 * inch))

    header_row = [
        'Name', 'Generic', 'Strength', 'Form', 'Pack', 'Therapeutic category',
        'Active', 'Ctrl', 'Unit (GHS)', 'Cost (GHS)', 'Stock qty', 'Batches', 'ATC',
    ]
    table_data = [header_row]
    for drug in drugs:
        sq, nb = stock_by_drug_pdf.get(drug.pk, (0, 0))
        table_data.append([
            _pcell(drug.name, p_name),
            _pcell(drug.generic_name, p_small),
            _cell_plain(drug.strength, 20),
            _cell_plain(drug.form, 22),
            _cell_plain(drug.pack_size, 16),
            _pcell(drug.get_category_display(), p_cat),
            'Yes' if drug.is_active else 'No',
            'Yes' if drug.is_controlled else 'No',
            f'{drug.unit_price:.2f}',
            f'{drug.cost_price:.2f}',
            str(sq),
            str(nb),
            _cell_plain(drug.atc_code, 24),
        ])

    # ~10.0" usable width: prioritize category + name; keep narrow flags and money columns
    col_widths = [
        1.65 * inch, 0.88 * inch, 0.5 * inch, 0.58 * inch, 0.44 * inch,
        2.85 * inch, 0.36 * inch, 0.32 * inch, 0.48 * inch, 0.48 * inch,
        0.42 * inch, 0.4 * inch, 0.48 * inch,
    ]
    table = Table(table_data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7c3aed')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 5),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('ALIGN', (6, 1), (7, -1), 'CENTER'),
        ('ALIGN', (8, 1), (9, -1), 'RIGHT'),
        ('ALIGN', (10, 1), (11, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.read(), content_type='application/pdf')
    fname = f'drug_formulary_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


@login_required
def drug_formulary_excel(request):
    """Multi-sheet Excel: summary + full formulary with wrapped columns and GHS number formats."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from .models_settings import HospitalSettings

    HEADER = '5B21B6'
    HEADER_LIGHT = '7C3AED'
    LABEL_BG = 'EDE9FE'
    BORDER_HEX = 'D1D5DB'
    ZEBRA = 'F9FAFB'
    thin = Side(style='thin', color=BORDER_HEX)
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_top = Alignment(wrap_text=True, vertical='top')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    drugs = list(drug_formulary_filtered_queryset(request).order_by('category', 'name'))
    from django.db.models import Sum, Count

    drug_ids = [d.pk for d in drugs]
    stock_by_drug = {}
    if drug_ids:
        for row in (
            PharmacyStock.objects.filter(is_deleted=False, drug_id__in=drug_ids)
            .values('drug_id')
            .annotate(total_qty=Sum('quantity_on_hand'), batch_lines=Count('id'))
        ):
            stock_by_drug[row['drug_id']] = (
                int(row['total_qty'] or 0),
                int(row['batch_lines'] or 0),
            )

    settings = HospitalSettings.get_settings()
    search_query = request.GET.get('q', '').strip()
    category_filter = request.GET.get('category', '')
    form_filter = request.GET.get('form', '')
    status_filter = request.GET.get('status', '')
    controlled_filter = request.GET.get('controlled', '')
    category_filter_label = (
        dict(Drug.CATEGORIES).get(category_filter, category_filter) if category_filter else 'All categories'
    )
    now = timezone.now()
    gen_str = now.strftime('%Y-%m-%d %H:%M:%S')

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = 'Summary'
    ws_sum.sheet_properties.tabColor = HEADER_LIGHT
    ws_sum.merge_cells('A1:F1')
    t1 = ws_sum['A1']
    t1.value = 'Drug formulary export'
    t1.font = Font(bold=True, size=22, color=HEADER)
    t1.alignment = Alignment(horizontal='center', vertical='center')
    ws_sum.row_dimensions[1].height = 36
    ws_sum.merge_cells('A2:F2')
    ws_sum['A2'] = settings.hospital_name or 'Hospital'
    ws_sum['A2'].font = Font(size=12, color='374151')
    ws_sum['A2'].alignment = Alignment(horizontal='center')
    start_meta = 4
    if settings.address:
        ws_sum.merge_cells('A3:F3')
        ws_sum['A3'] = settings.address
        ws_sum['A3'].alignment = Alignment(horizontal='center', wrap_text=True)
        ws_sum['A3'].font = Font(size=10, color='6B7280')
        start_meta = 5

    meta = [
        ('Generated', gen_str),
        ('Total drugs', len(drugs)),
        ('Search', search_query or '—'),
        ('Category filter', category_filter_label),
        ('Form filter', form_filter or '—'),
        ('Status filter', status_filter or '—'),
        ('Controlled filter', controlled_filter or '—'),
    ]
    for i, (label, value) in enumerate(meta):
        r = start_meta + i
        ws_sum.cell(r, 1, label).font = Font(bold=True, size=11)
        ws_sum.cell(r, 1).fill = PatternFill(start_color=LABEL_BG, end_color=LABEL_BG, fill_type='solid')
        ws_sum.cell(r, 1).border = grid
        ws_sum.cell(r, 2, value).border = grid
        ws_sum.cell(r, 2).alignment = Alignment(vertical='center', wrap_text=True)
    ws_sum.column_dimensions['A'].width = 22
    ws_sum.column_dimensions['B'].width = 70

    tip_r = start_meta + len(meta) + 2
    ws_sum.merge_cells(start_row=tip_r, start_column=1, end_row=tip_r, end_column=5)
    ws_sum.cell(tip_r, 1).value = (
        'Open the "Formulary" sheet: use Data → Filter on row 1, freeze panes, and widen columns as needed. '
        'Category and Name cells wrap text. "Total stock qty" is the sum of pharmacy batch on-hand quantities; '
        '"No. of batches" is how many PharmacyStock rows exist for that drug.'
    )
    ws_sum.cell(tip_r, 1).font = Font(italic=True, size=10, color='6B7280')

    ws_d = wb.create_sheet('Formulary', 1)
    ws_d.sheet_properties.tabColor = HEADER
    headers = [
        'Category code',
        'Therapeutic category (full)',
        'Drug name',
        'Generic name',
        'Strength',
        'Form',
        'Pack size',
        'Active',
        'Controlled substance',
        'Unit price (GHS)',
        'Cost price (GHS)',
        'Total stock qty',
        'No. of batches',
        'ATC code',
    ]
    ncols = len(headers)
    for col, h in enumerate(headers, 1):
        c = ws_d.cell(1, col, h)
        c.fill = PatternFill(start_color=HEADER, end_color=HEADER, fill_type='solid')
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.alignment = center
        c.border = grid

    for idx, drug in enumerate(drugs):
        row = 2 + idx
        total_qty, n_batches = stock_by_drug.get(drug.pk, (0, 0))
        vals = [
            drug.category or '',
            drug.get_category_display() or '',
            drug.name,
            drug.generic_name or '',
            drug.strength,
            drug.form,
            drug.pack_size or '',
            'Yes' if drug.is_active else 'No',
            'Yes' if drug.is_controlled else 'No',
            float(drug.unit_price or 0),
            float(drug.cost_price or 0),
            total_qty,
            n_batches,
            drug.atc_code or '',
        ]
        zebra = idx % 2 == 1
        for col, val in enumerate(vals, 1):
            cell = ws_d.cell(row, col, val)
            cell.border = grid
            if zebra:
                cell.fill = PatternFill(start_color=ZEBRA, end_color=ZEBRA, fill_type='solid')
            if col in (2, 3, 4):
                cell.alignment = wrap_top
            elif col in (8, 9):
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col in (10, 11):
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '#,##0.00'
            elif col in (12, 13):
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '#,##0'
            else:
                cell.alignment = Alignment(vertical='top')

    last = 1 + len(drugs)
    ws_d.auto_filter.ref = f'A1:{get_column_letter(ncols)}{max(last, 1)}'
    ws_d.freeze_panes = 'A2'
    widths = [14, 52, 38, 24, 12, 14, 12, 10, 12, 16, 16, 14, 12, 14]
    for i, w in enumerate(widths, 1):
        if i <= ncols:
            ws_d.column_dimensions[get_column_letter(i)].width = w

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    fname = f'drug_formulary_{now.strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


@login_required
def drug_detail(request, pk):
    """View drug details"""
    from datetime import timedelta
    from .models_payment_verification import PharmacyDispenseHistory
    from .models_pharmacy_walkin import WalkInPharmacySaleItem
    drug = get_object_or_404(Drug, pk=pk, is_deleted=False)
    
    # Get stock information
    stock_entries_qs = PharmacyStock.objects.filter(drug=drug, is_deleted=False).order_by('expiry_date')
    total_stock_on_hand = stock_entries_qs.aggregate(total=Sum('quantity_on_hand'))['total'] or 0
    total_initial_stock = stock_entries_qs.aggregate(total=Sum('initial_quantity'))['total'] or 0

    # Get recent prescriptions
    recent_prescriptions = Prescription.objects.filter(
        drug=drug,
        is_deleted=False
    ).select_related('order__encounter__patient', 'prescribed_by__user').order_by('-created')[:10]
    
    prescription_count = Prescription.objects.filter(drug=drug, is_deleted=False).count()
    dispensed_txn_count = PharmacyDispenseHistory.objects.filter(
        drug=drug,
        is_deleted=False,
    ).count()
    dispensed_qty_prescription = PharmacyDispenseHistory.objects.filter(
        drug=drug,
        is_deleted=False,
    ).aggregate(total=Sum('quantity_dispensed'))['total'] or 0
    dispensed_qty_walkin = WalkInPharmacySaleItem.objects.filter(
        drug=drug,
        is_deleted=False,
        sale__is_deleted=False,
        sale__is_dispensed=True,
    ).aggregate(total=Sum('quantity'))['total'] or 0
    total_dispensed_qty = dispensed_qty_prescription + dispensed_qty_walkin
    total_used_stock = total_dispensed_qty
    total_stock = total_initial_stock - total_dispensed_qty
    stock_entries = list(stock_entries_qs)
    for stock in stock_entries:
        stock.used_quantity = (stock.initial_quantity or 0) - (stock.quantity_on_hand or 0)
    
    # Calculate stock value
    stock_value = total_stock * drug.unit_price
    
    # Dates for expiry checking
    today = timezone.now().date()
    expiring_soon = today + timedelta(days=30)
    
    context = {
        'drug': drug,
        'stock_entries': stock_entries,
        'total_stock': total_stock,
        'total_stock_on_hand': total_stock_on_hand,
        'total_initial_stock': total_initial_stock,
        'total_used_stock': total_used_stock,
        'stock_value': stock_value,
        'recent_prescriptions': recent_prescriptions,
        'prescription_count': prescription_count,
        'dispensed_txn_count': dispensed_txn_count,
        'total_dispensed_qty': total_dispensed_qty,
        'today': today,
        'expiring_soon': expiring_soon,
        'can_edit_stock': is_admin_user(request.user),
    }
    return render(request, 'hospital/drug_detail.html', context)


@login_required
def drug_audit_trail(request, pk):
    """
    Comprehensive Drug Audit Trail (Ledger)
    Shows complete history of drug movements: transfers, dispensing, adjustments, etc.
    World-Class Pharmacy System Standard
    """
    from django.db.models import Sum, Q, F
    from django.core.paginator import Paginator
    from decimal import Decimal
    from .models_procurement import StoreTransfer, StoreTransferLine, InventoryItem
    from .models_payment_verification import PharmacyDispensing
    from .models_pharmacy_walkin import WalkInPharmacySale, WalkInPharmacySaleItem
    from .models_inventory_advanced import InventoryTransaction
    from datetime import timedelta
    
    drug = get_object_or_404(Drug, pk=pk, is_deleted=False)
    
    # Get date range filter
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Build comprehensive audit trail
    audit_entries = []

    # Pharmacy batches (main stock path for many sites) — query once for ledger + current_stock summary
    pharmacy_stocks_qs = PharmacyStock.objects.filter(
        drug=drug,
        is_deleted=False,
    ).select_related('created_by', 'drug').order_by('-created')

    # 0. Formulary / catalog — when the drug record was created (new drugs had no ledger rows before)
    created_ts = getattr(drug, 'created', None)
    if created_ts:
        audit_entries.append({
            'type': 'catalog_created',
            'date': created_ts.date(),
            'datetime': created_ts,
            'reference': str(drug.pk)[:8].upper(),
            'description': 'Drug added to formulary (catalog entry)',
            'quantity': 0,
            'quantity_type': 'catalog',
            'notes': None,
        })
    
    # 1. STORE TRANSFERS (where drug was transferred to/from)
    # First get inventory items for this drug
    inventory_items = InventoryItem.objects.filter(
        drug=drug,
        is_deleted=False
    ).select_related('store')
    
    # Get all item codes for this drug
    item_codes = list(inventory_items.values_list('item_code', flat=True))
    item_names = list(inventory_items.values_list('item_name', flat=True))
    
    # Build transfer query
    transfer_q = Q()
    if drug.name:
        transfer_q |= Q(item_name__icontains=drug.name)
    if item_codes:
        transfer_q |= Q(item_code__in=item_codes)
    if item_names:
        for name in item_names:
            if name:
                transfer_q |= Q(item_name__icontains=name)
    
    transfer_lines = StoreTransferLine.objects.filter(
        transfer__is_deleted=False,
        is_deleted=False
    ).filter(transfer_q).select_related(
        'transfer__from_store',
        'transfer__to_store',
        'transfer__requested_by__user',
        'transfer__approved_by__user',
        'transfer__received_by__user'
    ).distinct()
    
    for line in transfer_lines:
        transfer = line.transfer
        audit_entries.append({
            'type': 'transfer',
            'date': transfer.transfer_date or transfer.created.date(),
            'datetime': transfer.created,
            'reference': transfer.transfer_number,
            'description': f'Transfer: {transfer.from_store.name} → {transfer.to_store.name}',
            'quantity': line.quantity,
            'quantity_type': 'transferred',
            'store_from': transfer.from_store.name,
            'store_to': transfer.to_store.name,
            'requested_by': transfer.requested_by.user.get_full_name() if transfer.requested_by and transfer.requested_by.user else 'System',
            'approved_by': transfer.approved_by.user.get_full_name() if transfer.approved_by and transfer.approved_by.user else None,
            'received_by': transfer.received_by.user.get_full_name() if transfer.received_by and transfer.received_by.user else None,
            'status': transfer.get_status_display(),
            'notes': line.notes or transfer.notes,
            'unit_cost': line.unit_cost,
        })
    
    # 2. PHARMACY DISPENSING (prescription-based)
    dispensings = PharmacyDispensing.objects.filter(
        prescription__drug=drug,
        is_deleted=False
    ).select_related(
        'prescription__order__encounter__patient',
        'prescription__prescribed_by__user',
        'dispensed_by__user',
        'payment_receipt'
    ).order_by('-dispensed_at', '-created')
    
    for disp in dispensings:
        if disp.dispensed_at:
            audit_entries.append({
                'type': 'dispensing',
                'date': disp.dispensed_at.date(),
                'datetime': disp.dispensed_at,
                'reference': f"RX-{disp.prescription.id}",
                'description': f'Dispensed to Patient: {disp.patient.full_name}',
                'quantity': disp.quantity_dispensed or 0,
                'quantity_type': 'dispensed',
                'patient_name': disp.patient.full_name,
                'patient_mrn': disp.patient.mrn,
                'prescribed_by': disp.prescription.prescribed_by.user.get_full_name() if disp.prescription.prescribed_by and disp.prescription.prescribed_by.user else 'Unknown',
                'dispensed_by': disp.dispensed_by.user.get_full_name() if disp.dispensed_by and disp.dispensed_by.user else None,
                'status': disp.get_dispensing_status_display(),
                'payment_receipt': disp.payment_receipt.receipt_number if disp.payment_receipt else None,
                'notes': disp.dispensing_notes,
            })
    
    # 3. WALK-IN SALES
    walkin_items = WalkInPharmacySaleItem.objects.filter(
        drug=drug,
        is_deleted=False
    ).select_related(
        'sale__served_by__user',
        'sale__dispensed_by__user',
        'sale__patient'
    ).order_by('-sale__sale_date')
    
    for item in walkin_items:
        sale = item.sale
        if sale.is_dispensed and sale.dispensed_at:
            audit_entries.append({
                'type': 'walkin_sale',
                'date': sale.dispensed_at.date(),
                'datetime': sale.dispensed_at,
                'reference': sale.sale_number,
                'description': f'Prescribe Sale to: {sale.customer_name}',
                'quantity': item.quantity,
                'quantity_type': 'dispensed',
                'customer_name': sale.customer_name,
                'customer_phone': sale.customer_phone,
                'patient_name': sale.patient.full_name if sale.patient else None,
                'patient_mrn': sale.patient.mrn if sale.patient else None,
                'served_by': sale.served_by.user.get_full_name() if sale.served_by and sale.served_by.user else None,
                'dispensed_by': sale.dispensed_by.user.get_full_name() if sale.dispensed_by and sale.dispensed_by.user else None,
                'unit_price': item.unit_price,
                'total_amount': item.line_total,
                'payment_status': sale.get_payment_status_display(),
                'notes': sale.notes,
            })

    # 3b. PHARMACY STOCK BATCHES (receipts at Main Pharmacy / locations — was missing from ledger)
    for stock in pharmacy_stocks_qs:
        created_ts = stock.created
        qty_received = int(stock.initial_quantity or 0)
        if qty_received <= 0:
            qty_received = int(stock.quantity_on_hand or 0)
        added_by = 'System'
        if stock.created_by_id:
            added_by = (
                stock.created_by.get_full_name()
                or getattr(stock.created_by, 'username', None)
                or 'System'
            )
        audit_entries.append({
            'type': 'pharmacy_stock',
            'date': created_ts.date(),
            'datetime': created_ts,
            'reference': f'BATCH-{stock.batch_number}',
            'description': (
                f'Pharmacy stock received · {stock.location or "Pharmacy"} · '
                f'Batch {stock.batch_number} · Expires {stock.expiry_date}'
            ),
            'quantity': max(qty_received, 0),
            'quantity_type': 'added',
            'batch_number': stock.batch_number,
            'expiry_date': stock.expiry_date,
            'location': stock.location or 'Pharmacy',
            'quantity_on_hand': stock.quantity_on_hand,
            'initial_quantity': stock.initial_quantity,
            'created_by': added_by,
            'unit_cost': stock.unit_cost,
            'notes': None,
        })
    
    # 4. INVENTORY TRANSACTIONS (adjustments, receipts, etc.)
    # InventoryTransaction has performed_by (not created_by); use only valid FK names in select_related
    for inv_item in inventory_items:
        transactions = InventoryTransaction.objects.filter(
            inventory_item=inv_item,
            is_deleted=False
        ).select_related('store', 'performed_by__user').order_by('-transaction_date', '-created')
        
        for trans in transactions:
            performer_name = 'System'
            if getattr(trans, 'performed_by', None) and getattr(trans.performed_by, 'user', None):
                performer_name = trans.performed_by.user.get_full_name() or 'System'
            audit_entries.append({
                'type': 'inventory_transaction',
                'date': trans.transaction_date.date() if trans.transaction_date else trans.created.date(),
                'datetime': trans.transaction_date or trans.created,
                'reference': trans.transaction_number,
                'description': f'{trans.get_transaction_type_display()}: {trans.store.name}',
                'quantity': abs(trans.quantity),
                'quantity_type': 'added' if trans.quantity > 0 else 'removed',
                'store': trans.store.name,
                'quantity_before': trans.quantity_before,
                'quantity_after': trans.quantity_after,
                'created_by': performer_name,
                'unit_cost': trans.unit_cost,
                'notes': trans.notes,
            })
    
    # 5. CURRENT STOCK LEVELS across all stores
    current_stock = {}
    for inv_item in inventory_items:
        store_name = inv_item.store.name
        if store_name not in current_stock:
            current_stock[store_name] = {
                'quantity': 0,
                'value': Decimal('0.00'),
                'unit_cost': inv_item.unit_cost or Decimal('0.00'),
            }
        current_stock[store_name]['quantity'] += inv_item.quantity_on_hand
        current_stock[store_name]['value'] += (inv_item.quantity_on_hand * (inv_item.unit_cost or Decimal('0.00')))
    
    # Also check PharmacyStock (same queryset as ledger above)
    for stock in pharmacy_stocks_qs:
        location = stock.location or 'Pharmacy'
        if location not in current_stock:
            current_stock[location] = {
                'quantity': 0,
                'value': Decimal('0.00'),
                'unit_cost': stock.unit_cost or Decimal('0.00'),
            }
        current_stock[location]['quantity'] += stock.quantity_on_hand
        current_stock[location]['value'] += (stock.quantity_on_hand * (stock.unit_cost or Decimal('0.00')))
    
    # Sort audit entries by datetime (most recent first)
    audit_entries.sort(key=lambda x: x['datetime'], reverse=True)
    
    # Apply date filters
    if date_from:
        try:
            from datetime import datetime
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            audit_entries = [e for e in audit_entries if e['date'] >= date_from_obj]
        except:
            pass
    
    if date_to:
        try:
            from datetime import datetime
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            audit_entries = [e for e in audit_entries if e['date'] <= date_to_obj]
        except:
            pass
    
    # Calculate statistics
    total_transferred = sum(e['quantity'] for e in audit_entries if e['type'] == 'transfer')
    total_dispensed = sum(e['quantity'] for e in audit_entries if e['type'] in ['dispensing', 'walkin_sale'])
    total_adjusted = (
        sum(e['quantity'] for e in audit_entries if e['type'] == 'inventory_transaction' and e['quantity_type'] == 'added')
        + sum(e['quantity'] for e in audit_entries if e['type'] == 'pharmacy_stock')
    )
    total_removed = sum(e['quantity'] for e in audit_entries if e['type'] == 'inventory_transaction' and e['quantity_type'] == 'removed')
    
    # Pagination
    paginator = Paginator(audit_entries, 50)
    page = request.GET.get('page', 1)
    try:
        audit_page = paginator.get_page(page)
    except:
        audit_page = paginator.get_page(1)
    
    context = {
        'drug': drug,
        'audit_entries': audit_page,
        'current_stock': current_stock,
        'total_transferred': total_transferred,
        'total_dispensed': total_dispensed,
        'total_adjusted': total_adjusted,
        'total_removed': total_removed,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'hospital/drug_audit_trail.html', context)


def _can_access_drug_catalog_edit(user):
    """Admin, account/finance, and procurement/store manager can access drug catalog edit."""
    return is_admin_user(user) or is_account_or_finance(user) or is_procurement_staff(user)


@login_required
@user_passes_test(_can_access_drug_catalog_edit, login_url='/admin/login/')
def drug_create(request):
    """Create a new drug. Anyone who can access this page can set selling price; cost price is admin/procurement only."""
    # Access already restricted by decorator: allow selling price edit for all who reach the form
    can_edit_selling_price = True
    can_edit_cost_price = is_admin_user(request.user) or is_procurement_staff(request.user)
    if request.method == 'POST':
        try:
            unit_price = (request.POST.get('unit_price', 0) or 0) if can_edit_selling_price else 0
            cost_price = (request.POST.get('cost_price', 0) or 0) if can_edit_cost_price else 0
            drug = Drug.objects.create(
                atc_code=request.POST.get('atc_code', ''),
                name=request.POST['name'],
                generic_name=request.POST.get('generic_name', ''),
                strength=request.POST['strength'],
                form=request.POST['form'],
                pack_size=request.POST.get('pack_size', ''),
                category=request.POST.get('category', 'other'),
                is_controlled=request.POST.get('is_controlled') == 'on',
                is_active=request.POST.get('is_active', 'on') == 'on',
                unit_price=unit_price,
                cost_price=cost_price,
            )
            messages.success(request, f'Drug "{drug.name}" created successfully.')
            return redirect('hospital:drug_detail', pk=drug.pk)
        except Exception as e:
            messages.error(request, f'Error creating drug: {str(e)}')
    
    # Minimal drug-like object for create form (template expects drug.unit_price, etc.)
    from types import SimpleNamespace
    create_drug = SimpleNamespace(
        name='', generic_name='', atc_code='', strength='', form='', pack_size='',
        category='other', is_controlled=False, is_active=True, unit_price=0, cost_price=0,
    )
    context = {
        'action': 'Create',
        'drug': create_drug,
        'drug_categories': Drug.CATEGORIES,
        'can_edit_selling_price': can_edit_selling_price,
        'can_edit_cost_price': can_edit_cost_price,
    }
    return render(request, 'hospital/drug_form.html', context)


@login_required
@user_passes_test(_can_access_drug_catalog_edit, login_url='/admin/login/')
def drug_edit(request, pk):
    """Edit drug. Anyone who can access this page can edit selling/unit price; cost price and stock expiry are admin/procurement only."""
    drug = get_object_or_404(Drug, pk=pk, is_deleted=False)
    # Access already restricted by decorator: allow selling price edit for all who reach the form
    can_edit_selling_price = True
    can_edit_cost_price = is_admin_user(request.user) or is_procurement_staff(request.user)
    can_edit_stock = is_admin_user(request.user) or is_procurement_staff(request.user)
    
    if request.method == 'POST':
        try:
            drug.atc_code = request.POST.get('atc_code', '')
            drug.name = request.POST['name']
            drug.generic_name = request.POST.get('generic_name', '')
            drug.strength = request.POST['strength']
            drug.form = request.POST['form']
            drug.pack_size = request.POST.get('pack_size', '')
            drug.category = request.POST.get('category', 'other')
            drug.is_controlled = request.POST.get('is_controlled') == 'on'
            drug.is_active = request.POST.get('is_active', 'on') == 'on'
            if can_edit_selling_price:
                drug.unit_price = request.POST.get('unit_price', 0) or 0
            if can_edit_cost_price:
                drug.cost_price = request.POST.get('cost_price', 0) or 0
            drug.save()
            
            # Update expiry dates for PharmacyStock batches (admin only)
            if can_edit_stock:
                from datetime import datetime
                from .models import PharmacyStock
                for key, value in request.POST.items():
                    if key.startswith('stock_expiry_') and value:
                        try:
                            stock_pk = key.replace('stock_expiry_', '')
                            stock = PharmacyStock.objects.get(
                                pk=stock_pk,
                                drug=drug,
                                is_deleted=False
                            )
                            stock.expiry_date = datetime.strptime(value, '%Y-%m-%d').date()
                            stock.save()
                        except (ValueError, PharmacyStock.DoesNotExist):
                            pass
            
            messages.success(request, f'Drug "{drug.name}" updated successfully.')
            return redirect('hospital:drug_detail', pk=drug.pk)
        except Exception as e:
            messages.error(request, f'Error updating drug: {str(e)}')
    
    # Get stock batches for expiry editing
    stock_entries = []
    try:
        from .models import PharmacyStock
        stock_entries = list(
            PharmacyStock.objects.filter(
                drug=drug,
                is_deleted=False
            ).order_by('expiry_date')
        )
    except Exception:
        pass
    
    context = {
        'drug': drug,
        'action': 'Edit',
        'drug_categories': Drug.CATEGORIES,
        'stock_entries': stock_entries,
        'can_edit_selling_price': can_edit_selling_price,
        'can_edit_cost_price': can_edit_cost_price,
        'can_edit_stock': can_edit_stock,
    }
    return render(request, 'hospital/drug_form.html', context)


# ==================== LAB TESTS CATALOG MANAGEMENT ====================

@login_required
def lab_tests_catalog(request):
    """List all lab tests with search and filtering"""
    tests = LabTest.objects.filter(is_deleted=False).order_by('name')
    
    # Search
    search_query = request.GET.get('q', '')
    if search_query:
        tests = tests.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query) |
            Q(specimen_type__icontains=search_query)
        )
    
    # Filter by specimen type
    specimen_filter = request.GET.get('specimen', '')
    if specimen_filter:
        tests = tests.filter(specimen_type__icontains=specimen_filter)
    
    # Filter by active status
    status_filter = request.GET.get('status', '')
    if status_filter == 'active':
        tests = tests.filter(is_active=True)
    elif status_filter == 'inactive':
        tests = tests.filter(is_active=False)
    
    # Pagination
    paginator = Paginator(tests, 20)
    page = request.GET.get('page')
    tests_page = paginator.get_page(page)
    
    # Get unique specimen types for filter dropdown
    specimen_types = LabTest.objects.filter(is_deleted=False).values_list('specimen_type', flat=True).distinct().order_by('specimen_type')
    
    can_manage_catalog = _can_access_lab_catalog_edit(request.user)
    context = {
        'tests': tests_page,
        'specimen_types': specimen_types,
        'search_query': search_query,
        'specimen_filter': specimen_filter,
        'status_filter': status_filter,
        'total_tests': tests.count(),
        'can_manage_catalog': can_manage_catalog,
    }
    return render(request, 'hospital/lab_tests_catalog.html', context)


@login_required
def lab_test_detail(request, pk):
    """View lab test details"""
    test = get_object_or_404(LabTest, pk=pk, is_deleted=False)
    
    # Get recent results for this test
    recent_results = LabResult.objects.filter(
        test=test,
        is_deleted=False
    ).select_related('order__encounter__patient').order_by('-created')[:10]
    
    # Get statistics
    total_results = LabResult.objects.filter(test=test, is_deleted=False).count()
    completed_results = LabResult.objects.filter(test=test, status='completed', is_deleted=False).count()
    abnormal_results = LabResult.objects.filter(test=test, is_abnormal=True, is_deleted=False).count()
    
    try:
        lab_test_delete_url = reverse('hospital:lab_test_delete', kwargs={'pk': test.pk})
    except Exception:
        lab_test_delete_url = None
    context = {
        'test': test,
        'recent_results': recent_results,
        'total_results': total_results,
        'completed_results': completed_results,
        'abnormal_results': abnormal_results,
        'lab_test_delete_url': lab_test_delete_url,
        'can_manage_catalog': _can_access_lab_catalog_edit(request.user),
    }
    return render(request, 'hospital/lab_test_detail.html', context)


def _parse_tat_minutes(val, default=60):
    """Safely parse tat_minutes from form; avoids int('') error."""
    if val is None or str(val).strip() == '':
        return default
    try:
        n = int(val)
        return max(1, n) if n > 0 else default
    except (ValueError, TypeError):
        return default


def _can_access_lab_catalog_edit(user):
    """Who may create/edit/delete LabTest catalog rows: lab staff, admins, finance, procurement."""
    if not user or not getattr(user, 'is_authenticated', False):
        return False
    if getattr(user, 'is_superuser', False):
        return True
    try:
        from .utils_roles import get_user_role
        if get_user_role(user) in ('lab_technician', 'admin'):
            return True
    except Exception:
        pass
    if is_admin_user(user) or is_account_or_finance(user) or is_procurement_staff(user):
        return True
    try:
        if user.has_perm('hospital.change_labtest'):
            return True
    except Exception:
        pass
    try:
        from .models import Staff
        staff = Staff.objects.filter(user=user, is_deleted=False).order_by('-created').first()
        if staff and staff.profession:
            prof = (staff.profession or '').strip().lower().replace(' ', '_').replace('-', '_')
            lab_professions = {
                'lab_technician', 'lab_tech', 'laboratory_technician', 'laboratory_scientist',
                'medical_laboratory_scientist', 'lab', 'laboratory', 'mls',
            }
            if prof in lab_professions or ('lab' in prof and 'assistant' not in prof):
                return True
    except Exception:
        pass
    if user.groups.filter(
        name__in=['Laboratory', 'Lab', 'Lab Technician', 'Lab Technicians', 'Laboratory Technician']
    ).exists():
        return True
    return False


@login_required
@user_passes_test(_can_access_lab_catalog_edit, login_url='/admin/login/')
def lab_test_create(request):
    """Create a new lab test. Anyone who can access this page can set price."""
    can_edit_price = True
    if request.method == 'POST':
        try:
            price = (request.POST.get('price', 0) or 0) if can_edit_price else 0
            test = LabTest.objects.create(
                code=request.POST['code'],
                name=request.POST['name'],
                specimen_type=request.POST['specimen_type'],
                tat_minutes=_parse_tat_minutes(request.POST.get('tat_minutes'), 60),
                price=price,
                is_active=request.POST.get('is_active', 'on') == 'on',
            )
            messages.success(request, f'Lab test "{test.name}" created successfully.')
            return redirect('hospital:lab_test_detail', pk=test.pk)
        except Exception as e:
            messages.error(request, f'Error creating lab test: {str(e)}')
    
    return render(request, 'hospital/lab_test_form.html', {
        'action': 'Create',
        'can_edit_price': can_edit_price,
    })


@login_required
@user_passes_test(_can_access_lab_catalog_edit, login_url='/admin/login/')
def lab_test_edit(request, pk):
    """Edit lab test. Anyone who can access this page can change price."""
    test = get_object_or_404(LabTest, pk=pk, is_deleted=False)
    can_edit_price = True
    
    if request.method == 'POST':
        try:
            test.code = request.POST['code']
            test.name = request.POST['name']
            test.specimen_type = request.POST['specimen_type']
            test.tat_minutes = _parse_tat_minutes(request.POST.get('tat_minutes'), test.tat_minutes or 60)
            if can_edit_price:
                test.price = request.POST.get('price', 0) or 0
            test.is_active = request.POST.get('is_active', 'on') == 'on'
            test.save()
            
            messages.success(request, f'Lab test "{test.name}" updated successfully.')
            return redirect('hospital:lab_test_detail', pk=test.pk)
        except Exception as e:
            messages.error(request, f'Error updating lab test: {str(e)}')
    
    try:
        delete_url = reverse('hospital:lab_test_delete', kwargs={'pk': test.pk})
    except Exception:
        delete_url = None
    context = {
        'test': test,
        'action': 'Edit',
        'lab_test_delete_url': delete_url,
        'can_edit_price': can_edit_price,
    }
    return render(request, 'hospital/lab_test_form.html', context)


@login_required
@user_passes_test(_can_access_lab_catalog_edit, login_url='/admin/login/')
def lab_test_delete(request, pk):
    """Soft-delete a lab test (sets is_deleted=True)"""
    test = get_object_or_404(LabTest, pk=pk, is_deleted=False)
    if request.method == 'POST':
        try:
            test.is_deleted = True
            test.save()
            messages.success(request, f'Lab test "{test.name}" has been deleted.')
            return redirect('hospital:lab_tests_catalog')
        except Exception as e:
            messages.error(request, f'Error deleting lab test: {str(e)}')
            return redirect('hospital:lab_test_detail', pk=pk)
    # GET: show confirmation
    context = {'test': test}
    return render(request, 'hospital/lab_test_confirm_delete.html', context)


# ==================== DEPARTMENTS & WARDS MANAGEMENT ====================

from .models import Department, Ward

@login_required
def departments_list(request):
    """List all departments"""
    departments = Department.objects.filter(is_deleted=False).select_related('head_of_department__user').order_by('name')
    
    # Search
    search_query = request.GET.get('q', '')
    if search_query:
        departments = departments.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Filter by active status
    status_filter = request.GET.get('status', '')
    if status_filter == 'active':
        departments = departments.filter(is_active=True)
    elif status_filter == 'inactive':
        departments = departments.filter(is_active=False)
    
    # Add counts
    for dept in departments:
        dept.staff_count = Staff.objects.filter(department=dept, is_active=True, is_deleted=False).count()
        dept.wards_count = Ward.objects.filter(department=dept, is_active=True, is_deleted=False).count()
    
    context = {
        'departments': departments,
        'search_query': search_query,
        'status_filter': status_filter,
        'total_departments': departments.count(),
    }
    return render(request, 'hospital/departments_list.html', context)


@login_required
def department_detail(request, pk):
    """View department details"""
    department = get_object_or_404(Department, pk=pk, is_deleted=False)
    
    # Get staff in this department
    staff = Staff.objects.filter(department=department, is_deleted=False).select_related('user').order_by('user__last_name')
    
    # Get wards in this department
    wards = Ward.objects.filter(department=department, is_deleted=False).order_by('name')
    
    context = {
        'department': department,
        'staff': staff,
        'wards': wards,
    }
    return render(request, 'hospital/department_detail.html', context)


@login_required
def department_create(request):
    """Create a new department"""
    if request.method == 'POST':
        try:
            head_id = request.POST.get('head_of_department')
            head = Staff.objects.get(pk=head_id) if head_id else None
            
            department = Department.objects.create(
                name=request.POST['name'],
                code=request.POST.get('code', ''),
                description=request.POST.get('description', ''),
                head_of_department=head,
                is_active=request.POST.get('is_active', 'on') == 'on',
            )
            messages.success(request, f'Department "{department.name}" created successfully.')
            return redirect('hospital:department_detail', pk=department.pk)
        except Exception as e:
            messages.error(request, f'Error creating department: {str(e)}')
    
    staff_list = Staff.objects.filter(is_active=True, is_deleted=False).select_related('user').order_by('user__last_name')
    return render(request, 'hospital/department_form.html', {'action': 'Create', 'staff_list': staff_list})


@login_required
def department_edit(request, pk):
    """Edit an existing department"""
    department = get_object_or_404(Department, pk=pk, is_deleted=False)
    
    if request.method == 'POST':
        try:
            head_id = request.POST.get('head_of_department')
            head = Staff.objects.get(pk=head_id) if head_id else None
            
            department.name = request.POST['name']
            department.code = request.POST.get('code', '')
            department.description = request.POST.get('description', '')
            department.head_of_department = head
            department.is_active = request.POST.get('is_active', 'on') == 'on'
            department.save()
            
            messages.success(request, f'Department "{department.name}" updated successfully.')
            return redirect('hospital:department_detail', pk=department.pk)
        except Exception as e:
            messages.error(request, f'Error updating department: {str(e)}')
    
    staff_list = Staff.objects.filter(is_active=True, is_deleted=False).select_related('user').order_by('user__last_name')
    context = {
        'department': department,
        'action': 'Edit',
        'staff_list': staff_list
    }
    return render(request, 'hospital/department_form.html', context)


@login_required
def wards_list(request):
    """List all wards"""
    wards = Ward.objects.filter(is_deleted=False).select_related('department').order_by('name')
    
    # Search
    search_query = request.GET.get('q', '')
    if search_query:
        wards = wards.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query)
        )
    
    # Filter by department
    dept_filter = request.GET.get('department', '')
    if dept_filter:
        wards = wards.filter(department_id=dept_filter)
    
    # Filter by ward type
    type_filter = request.GET.get('type', '')
    if type_filter:
        wards = wards.filter(ward_type=type_filter)
    
    # Filter by active status
    status_filter = request.GET.get('status', '')
    if status_filter == 'active':
        wards = wards.filter(is_active=True)
    elif status_filter == 'inactive':
        wards = wards.filter(is_active=False)
    
    # Add bed counts
    for ward in wards:
        beds = Bed.objects.filter(ward=ward, is_deleted=False)
        ward.total_beds = beds.count()
        ward.available_beds = beds.filter(status='available').count()
        ward.occupied_beds = beds.filter(status='occupied').count()
    
    departments = Department.objects.filter(is_active=True, is_deleted=False).order_by('name')
    
    context = {
        'wards': wards,
        'departments': departments,
        'search_query': search_query,
        'dept_filter': dept_filter,
        'type_filter': type_filter,
        'status_filter': status_filter,
        'total_wards': wards.count(),
    }
    return render(request, 'hospital/wards_list.html', context)


@login_required
def ward_detail(request, pk):
    """View ward details"""
    ward = get_object_or_404(Ward, pk=pk, is_deleted=False)
    
    # Get beds in this ward
    beds = Bed.objects.filter(ward=ward, is_deleted=False).order_by('bed_number')
    
    # Get current admissions
    admissions = Admission.objects.filter(
        ward=ward,
        status='admitted',
        is_deleted=False
    ).select_related('encounter__patient').order_by('-admit_date')
    
    context = {
        'ward': ward,
        'beds': beds,
        'admissions': admissions,
    }
    return render(request, 'hospital/ward_detail.html', context)


@login_required
def ward_create(request):
    """Create a new ward"""
    if request.method == 'POST':
        try:
            ward = Ward.objects.create(
                name=request.POST['name'],
                code=request.POST.get('code', ''),
                ward_type=request.POST['ward_type'],
                department_id=request.POST['department'],
                capacity=int(request.POST.get('capacity', 1)),
                is_active=request.POST.get('is_active', 'on') == 'on',
            )
            messages.success(request, f'Ward "{ward.name}" created successfully.')
            return redirect('hospital:ward_detail', pk=ward.pk)
        except Exception as e:
            messages.error(request, f'Error creating ward: {str(e)}')
    
    departments = Department.objects.filter(is_active=True, is_deleted=False).order_by('name')
    return render(request, 'hospital/ward_form.html', {'action': 'Create', 'departments': departments})


@login_required
def ward_edit(request, pk):
    """Edit an existing ward"""
    ward = get_object_or_404(Ward, pk=pk, is_deleted=False)
    
    if request.method == 'POST':
        try:
            ward.name = request.POST['name']
            ward.code = request.POST.get('code', '')
            ward.ward_type = request.POST['ward_type']
            ward.department_id = request.POST['department']
            ward.capacity = int(request.POST.get('capacity', 1))
            ward.is_active = request.POST.get('is_active', 'on') == 'on'
            ward.save()
            
            messages.success(request, f'Ward "{ward.name}" updated successfully.')
            return redirect('hospital:ward_detail', pk=ward.pk)
        except Exception as e:
            messages.error(request, f'Error updating ward: {str(e)}')
    
    departments = Department.objects.filter(is_active=True, is_deleted=False).order_by('name')
    context = {
        'ward': ward,
        'action': 'Edit',
        'departments': departments
    }
    return render(request, 'hospital/ward_form.html', context)


# ==================== MEDICAL RECORDS MANAGEMENT ====================

from .models import MedicalRecord

@login_required
def medical_records_list(request):
    """List all medical records"""
    # Auto-generate records for encounters that don't have documentation yet
    # OPTIMIZED: Better duplicate prevention - check for ANY record of this type for this encounter
    # Exclude front-desk registration-only encounters (not clinical; do not create notes for them)
    missing_encounters = Encounter.objects.filter(
        is_deleted=False
    ).exclude(chief_complaint__iexact='New patient registration').select_related(
        'patient', 'provider'
    ).annotate(
        record_count=models.Count('medical_records', filter=Q(medical_records__is_deleted=False))
    ).filter(
        record_count=0
    )[:50]  # limit to avoid heavy load
    
    for encounter in missing_encounters:
        if not encounter.patient:
            continue
        
        record_type_map = {
            'lab': 'lab_result',
            'imaging': 'imaging',
            'admission': 'discharge_summary',
            'surgery': 'surgical_report',
            'consultation': 'consultation_note',
            'outpatient': 'consultation_note',
            'inpatient': 'discharge_summary',
            'emergency': 'consultation_note',
        }
        record_type = record_type_map.get(encounter.encounter_type, 'consultation_note')
        title = f"{encounter.patient.full_name} - {encounter.encounter_type.title()} Note" if encounter.encounter_type else f"Encounter Note - {encounter.patient.full_name}"
        content = encounter.diagnosis or encounter.chief_complaint or getattr(encounter, 'summary', '') or ''
        
        # STRICT duplicate check: same patient + encounter + record_type + title
        existing_record = MedicalRecord.objects.filter(
            patient=encounter.patient,
            encounter=encounter,
            record_type=record_type,
            title=title,
            is_deleted=False
        ).first()
        
        if existing_record:
            # Update existing record if content is missing
            if not existing_record.content and content:
                existing_record.content = content
                existing_record.save()
            continue
        
        # Additional check: same patient + encounter + record_type (prevent same type duplicates)
        same_type_exists = MedicalRecord.objects.filter(
            patient=encounter.patient,
            encounter=encounter,
            record_type=record_type,
            is_deleted=False
        ).exists()
        
        if same_type_exists:
            continue  # Already have a record of this type for this encounter
        
        try:
            MedicalRecord.objects.create(
                patient=encounter.patient,
                encounter=encounter,
                record_type=record_type,
                title=title,
                content=content,
                created_by=encounter.provider,
            )
        except Exception:
            continue
    
    records = MedicalRecord.objects.filter(is_deleted=False).exclude(
        encounter__chief_complaint__iexact='New patient registration'
    ).select_related(
        'patient', 'encounter', 'created_by__user'
    ).order_by('-created')
    
    # Search
    search_query = request.GET.get('q', '')
    if search_query:
        records = records.filter(
            Q(title__icontains=search_query) |
            Q(content__icontains=search_query) |
            Q(patient__first_name__icontains=search_query) |
            Q(patient__last_name__icontains=search_query) |
            Q(patient__mrn__icontains=search_query)
        )
    
    # Date filter (record created date)
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    from datetime import datetime
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            records = records.filter(created__date__gte=date_from_obj)
        except ValueError:
            date_from = ''
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            records = records.filter(created__date__lte=date_to_obj)
        except ValueError:
            date_to = ''
    
    # Filter by record type
    type_filter = request.GET.get('type', '')
    if type_filter:
        records = records.filter(record_type=type_filter)
    
    # Filter by patient
    patient_filter = request.GET.get('patient', '')
    if patient_filter:
        records = records.filter(patient_id=patient_filter)
    
    # Group records by patient (folder view); dedupe so same encounter+type+title+datetime = one row
    patient_map = OrderedDict()
    for record in records:
        if not record.patient:
            continue
        pid = record.patient_id
        if pid not in patient_map:
            patient_map[pid] = {
                'patient': record.patient,
                'records': [],
                'record_count': 0,
                'last_created': record.created,
                '_seen': set(),
            }
        folder = patient_map[pid]
        # Same date+time = duplicate: use date+time to the second so duplicates collapse
        enc_key = record.encounter_id if record.encounter_id is not None else ''
        created_sec = record.created.strftime('%Y-%m-%d %H:%M:%S') if record.created else ''
        dedupe_key = (enc_key, (record.record_type or '').strip(), (record.title or '').strip(), created_sec)
        if dedupe_key in folder['_seen']:
            continue
        folder['_seen'].add(dedupe_key)
        folder['record_count'] += 1
        if record.created > folder['last_created']:
            folder['last_created'] = record.created
        folder['records'].append(record)
    for folder in patient_map.values():
        folder.pop('_seen', None)
    patient_records = sorted(
        patient_map.values(),
        key=lambda item: item['last_created'],
        reverse=True
    )
    for folder in patient_records:
        folder['service_counts'] = Counter(r.record_type for r in folder['records']).most_common(3)
    
    context = {
        'patient_records': patient_records,
        'search_query': search_query,
        'type_filter': type_filter,
        'patient_filter': patient_filter,
        'date_from': date_from,
        'date_to': date_to,
        'total_records': records.count(),
    }
    return render(request, 'hospital/medical_records_list.html', context)


@login_required
def medical_record_detail(request, pk):
    """View medical record details; nurses and doctors can add clinical notes for the encounter."""
    record = get_object_or_404(MedicalRecord, pk=pk, is_deleted=False)
    
    # Handle adding a clinical note (nurse or doctor) for this record's encounter
    if request.method == 'POST' and request.POST.get('action') == 'add_clinical_note':
        if not record.encounter_id:
            messages.error(request, 'This record is not linked to an encounter. Cannot add note here.')
        else:
            try:
                staff = Staff.objects.get(user=request.user, is_active=True, is_deleted=False)
                note_type = request.POST.get('note_type', 'progress').strip() or 'progress'
                objective = request.POST.get('objective', '').strip()
                notes = request.POST.get('notes', '').strip()
                if not notes:
                    messages.error(request, 'Note text is required.')
                else:
                    from .models_advanced import ClinicalNote
                    ClinicalNote.objects.create(
                        encounter_id=record.encounter_id,
                        note_type=note_type,
                        objective=objective,
                        notes=notes,
                        created_by=staff,
                    )
                    messages.success(request, 'Clinical note added successfully.')
            except Staff.DoesNotExist:
                messages.error(request, 'You must have an active staff profile to add clinical notes.')
            except Exception as e:
                logger.error(f"Error adding clinical note on medical record detail: {e}", exc_info=True)
                messages.error(request, f'Error saving note: {str(e)}')
        return redirect('hospital:medical_record_detail', pk=pk)
    
    # All clinical notes for this encounter so every saved note is visible (no filter by type or profession)
    clinical_notes = []
    nurse_notes = []
    doctor_notes = []
    try:
        if record.encounter_id:
            from .models_advanced import ClinicalNote
            clinical_notes = ClinicalNote.objects.filter(
                encounter_id=record.encounter_id,
                is_deleted=False
            ).select_related('created_by__user').order_by('-created')
            _progress = [n for n in clinical_notes if getattr(n, 'note_type', None) == 'progress']
            nurse_notes = [n for n in _progress if getattr(getattr(n, 'created_by', None), 'profession', None) == 'nurse']
            doctor_notes = [n for n in _progress if getattr(getattr(n, 'created_by', None), 'profession', None) == 'doctor']
    except Exception:
        clinical_notes = []
        nurse_notes = []
        doctor_notes = []
    
    # Can current user add a note? (staff only, and record must have an encounter)
    can_add_note = False
    try:
        if record.encounter_id and Staff.objects.filter(user=request.user, is_active=True, is_deleted=False).exists():
            can_add_note = True
    except Exception:
        pass
    
    context = {
        'record': record,
        'clinical_notes': clinical_notes,
        'nurse_notes': nurse_notes,
        'doctor_notes': doctor_notes,
        'can_add_note': can_add_note,
    }
    return render(request, 'hospital/medical_record_detail.html', context)


@login_required
def medical_record_create(request):
    """Create a new medical record"""
    if request.method == 'POST':
        try:
            staff_profile = getattr(request.user, 'staff_profile', None)
            encounter_id = request.POST.get('encounter')
            encounter = Encounter.objects.get(pk=encounter_id) if encounter_id else None
            
            record = MedicalRecord.objects.create(
                patient_id=request.POST['patient'],
                encounter=encounter,
                record_type=request.POST['record_type'],
                title=request.POST['title'],
                content=request.POST.get('content', ''),
                created_by=staff_profile,
            )
            
            # Handle document upload
            if 'document' in request.FILES:
                record.document = request.FILES['document']
                record.save()
            
            messages.success(request, f'Medical record "{record.title}" created successfully.')
            return redirect('hospital:medical_record_detail', pk=record.pk)
        except Exception as e:
            messages.error(request, f'Error creating medical record: {str(e)}')
    
    patients = Patient.objects.filter(is_deleted=False).exclude(id__isnull=True).order_by('last_name')
    # Get distinct encounters - prefer most recent per patient per day
    # Handle exact timestamp matches by using ID as tie-breaker
    from django.db.models import Max
    
    # Get most recent encounter ID per patient per day using DISTINCT ON (UUID-compatible)
    from django.db import connection
    
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT DISTINCT ON (patient_id, started_at::date) id
            FROM hospital_encounter
            WHERE is_deleted = false
            ORDER BY patient_id, started_at::date, id DESC
        """)
        latest_ids = [row[0] for row in cursor.fetchall()]
    
    encounters = Encounter.objects.filter(
        id__in=latest_ids
    ).select_related('patient').order_by('-started_at', '-id')[:50]
    
    return render(request, 'hospital/medical_record_form.html', {
        'action': 'Create',
        'patients': patients,
        'encounters': encounters
    })


# ==================== ORDERS MANAGEMENT ====================

@login_required
def orders_list(request):
    """List all orders"""
    all_orders = Order.objects.filter(is_deleted=False).select_related(
        'encounter__patient', 'requested_by__user'
    ).order_by('-requested_at')

    # Search
    search_query = request.GET.get('q', '')
    if search_query:
        all_orders = all_orders.filter(
            Q(encounter__patient__first_name__icontains=search_query) |
            Q(encounter__patient__last_name__icontains=search_query) |
            Q(encounter__patient__mrn__icontains=search_query) |
            Q(notes__icontains=search_query)
        )

    # Filter by order type
    type_filter = request.GET.get('type', '')
    if type_filter:
        all_orders = all_orders.filter(order_type=type_filter)

    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        all_orders = all_orders.filter(status=status_filter)

    # Filter by priority
    priority_filter = request.GET.get('priority', '')
    if priority_filter:
        all_orders = all_orders.filter(priority=priority_filter)
    
    # DEDUPLICATION: Remove duplicates before displaying
    seen_order_keys = {}
    unique_orders = []
    status_priority = {'completed': 10, 'in_progress': 8, 'pending': 5, 'cancelled': 1}
    
    for order in all_orders:
        # Create unique key: encounter + order_type
        key = (order.encounter_id, order.order_type)
        if key not in seen_order_keys:
            seen_order_keys[key] = order
            unique_orders.append(order)
        else:
            # Keep the one with higher status or more recent
            existing = seen_order_keys[key]
            existing_time = existing.requested_at or existing.created
            current_time = order.requested_at or order.created
            existing_priority = status_priority.get(existing.status, 0)
            current_priority = status_priority.get(order.status, 0)
            
            if current_priority > existing_priority or (current_priority == existing_priority and current_time > existing_time):
                # Replace with better one
                unique_orders.remove(existing)
                seen_order_keys[key] = order
                unique_orders.append(order)
    
    # Sort unique orders by requested_at
    orders = sorted(unique_orders, key=lambda x: x.requested_at or x.created, reverse=True)
    
    # Pagination
    paginator = Paginator(orders, 20)
    page = request.GET.get('page')
    orders_page = paginator.get_page(page)
    
    context = {
        'orders': orders_page,
        'search_query': search_query,
        'type_filter': type_filter,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
        'total_orders': len(orders),
    }
    return render(request, 'hospital/orders_list.html', context)


@login_required
def order_detail(request, pk):
    """View order details"""
    order = get_object_or_404(Order, pk=pk, is_deleted=False)
    
    # Get related items based on order type
    lab_results = None
    prescriptions = None
    imaging_studies = []
    
    if order.order_type == 'lab':
        lab_results = LabResult.objects.filter(order=order, is_deleted=False)
    elif order.order_type == 'medication':
        prescriptions = Prescription.objects.filter(order=order, is_deleted=False)
    
    try:
        from .models_advanced import ImagingStudy
    except ImportError:
        ImagingStudy = None
    
    if ImagingStudy:
        imaging_queryset = ImagingStudy.objects.filter(
            is_deleted=False
        ).prefetch_related('images')
        
        imaging_studies = imaging_queryset.filter(
            order=order
        ).order_by('-performed_at', '-created')
        
        if not imaging_studies.exists():
            imaging_studies = imaging_queryset.filter(
                encounter=order.encounter
            ).order_by('-performed_at', '-created')
    else:
        imaging_studies = []
    
    context = {
        'order': order,
        'lab_results': lab_results,
        'prescriptions': prescriptions,
        'imaging_studies': imaging_studies,
    }
    return render(request, 'hospital/order_detail.html', context)


@login_required
def order_create(request):
    """Create a new order"""
    if request.method == 'POST':
        try:
            staff_profile = getattr(request.user, 'staff_profile', None)
            
            order = Order.objects.create(
                encounter_id=request.POST['encounter'],
                order_type=request.POST['order_type'],
                status='pending',
                priority=request.POST.get('priority', 'routine'),
                requested_by=staff_profile,
                notes=request.POST.get('notes', ''),
            )
            
            messages.success(request, 'Order created successfully.')
            return redirect('hospital:order_detail', pk=order.pk)
        except Exception as e:
            messages.error(request, f'Error creating order: {str(e)}')
    
    # Get distinct encounters - prefer most recent per patient per day
    # Handle exact timestamp matches by using ID as tie-breaker
    from django.db.models import Max
    
    # Get most recent encounter ID per patient per day using DISTINCT ON (UUID-compatible)
    from django.db import connection
    
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT DISTINCT ON (patient_id, started_at::date) id
            FROM hospital_encounter
            WHERE is_deleted = false AND status = 'active'
            ORDER BY patient_id, started_at::date, id DESC
        """)
        latest_ids = [row[0] for row in cursor.fetchall()]
    
    encounters = Encounter.objects.filter(
        id__in=latest_ids
    ).select_related('patient').order_by('-started_at', '-id')
    
    return render(request, 'hospital/order_form.html', {
        'action': 'Create',
        'encounters': encounters
    })
