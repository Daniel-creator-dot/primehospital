"""
Advanced Accounting System - Admin Interface
Complete admin for journals, ledgers, vouchers, and financial management
"""

from django.contrib import admin
from django.utils.html import format_html
from django.db.models import Sum, Q
from django.urls import reverse
from django.utils.safestring import mark_safe
from django.utils import timezone
from .models_accounting_advanced import (
    AccountCategory, FiscalYear, AccountingPeriod,
    Journal, AdvancedJournalEntry, AdvancedJournalEntryLine, AdvancedGeneralLedger,
    PaymentVoucher, ReceiptVoucher, Cheque,
    RevenueCategory, Revenue,
    ExpenseCategory, Expense,
    AdvancedAccountsReceivable, AccountsPayable,
    BankAccount, BankTransaction,
    Budget, BudgetLine, TaxRate,
    AccountingAuditLog,
    Cashbook, BankReconciliation, BankReconciliationItem,
    InsuranceReceivable, ProcurementPurchase,
    AccountingPayroll, AccountingPayrollEntry, DoctorCommission,
    IncomeGroup, ProfitLossReport,
    RegistrationFee, CashSale, AccountingCorporateAccount,
    WithholdingReceivable, WithholdingTaxPayable, Deposit, InitialRevaluation,
    PettyCashTransaction
)
from .models_primecare_accounting import InsuranceReceivableEntry, InsurancePaymentReceived, CorporateReceivableEntry


# ==================== INLINE ADMINS ====================

class AdvancedJournalEntryLineInline(admin.TabularInline):
    model = AdvancedJournalEntryLine
    extra = 2
    fields = ['line_number', 'account', 'cost_center', 'description', 'debit_amount', 'credit_amount']
    
    def get_readonly_fields(self, request, obj=None):
        if obj and obj.status == 'posted':
            return ['line_number', 'account', 'cost_center', 'description', 'debit_amount', 'credit_amount']
        return []


class BudgetLineInline(admin.TabularInline):
    model = BudgetLine
    extra = 1
    fields = ['account', 'cost_center', 'budgeted_amount', 'actual_amount', 'variance', 'variance_percent']
    readonly_fields = ['actual_amount', 'variance', 'variance_percent']


# ==================== ACCOUNT MANAGEMENT ====================

@admin.register(AccountCategory)
class AccountCategoryAdmin(admin.ModelAdmin):
    list_display = ['code', 'name', 'category_type', 'is_active']
    list_filter = ['category_type', 'is_active']
    search_fields = ['code', 'name']
    
    fieldsets = (
        (None, {
            'fields': ('code', 'name', 'category_type', 'description')
        }),
        ('Status', {
            'fields': ('is_active',)
        }),
    )


@admin.register(FiscalYear)
class FiscalYearAdmin(admin.ModelAdmin):
    list_display = ['name', 'start_date', 'end_date', 'is_current_display', 'is_closed', 'closed_date']
    list_filter = ['is_closed', 'start_date']
    search_fields = ['name']
    readonly_fields = ['is_current_display']
    
    fieldsets = (
        ('Fiscal Year Information', {
            'fields': ('name', 'start_date', 'end_date')
        }),
        ('Status', {
            'fields': ('is_closed', 'closed_date', 'closed_by', 'is_current_display')
        }),
    )
    
    def is_current_display(self, obj):
        if obj.is_current:
            return format_html('<span style="color: green; font-weight: bold;">✓ Current</span>')
        return format_html('<span style="color: gray;">Not Current</span>')
    is_current_display.short_description = 'Status'


@admin.register(AccountingPeriod)
class AccountingPeriodAdmin(admin.ModelAdmin):
    list_display = ['name', 'fiscal_year', 'period_number', 'start_date', 'end_date', 'is_closed']
    list_filter = ['fiscal_year', 'is_closed', 'start_date']
    search_fields = ['name']
    
    fieldsets = (
        ('Period Information', {
            'fields': ('fiscal_year', 'period_number', 'name', 'start_date', 'end_date')
        }),
        ('Status', {
            'fields': ('is_closed', 'closed_date')
        }),
    )


# ==================== JOURNALS & LEDGERS ====================

@admin.register(Journal)
class JournalAdmin(admin.ModelAdmin):
    list_display = ['code', 'name', 'journal_type', 'is_active']
    list_filter = ['journal_type', 'is_active']
    search_fields = ['code', 'name']
    
    fieldsets = (
        ('Journal Information', {
            'fields': ('code', 'name', 'journal_type', 'description')
        }),
        ('Default Accounts', {
            'fields': ('default_debit_account', 'default_credit_account'),
            'classes': ('collapse',)
        }),
        ('Status', {
            'fields': ('is_active',)
        }),
    )


@admin.register(AdvancedJournalEntry)
class AdvancedJournalEntryAdmin(admin.ModelAdmin):
    list_display = ['entry_number', 'entry_date', 'journal', 'description_short', 'total_debit', 'total_credit', 'status_badge', 'is_balanced_display']
    list_filter = ['status', 'journal', 'entry_date', 'fiscal_year']
    search_fields = ['entry_number', 'description', 'reference']
    readonly_fields = ['entry_number', 'total_debit', 'total_credit', 'is_balanced_display', 'posted_by', 'posting_date']
    date_hierarchy = 'entry_date'
    
    inlines = [AdvancedJournalEntryLineInline]
    
    fieldsets = (
        ('Journal Entry Information', {
            'fields': ('entry_number', 'journal', 'entry_date', 'posting_date', 'reference')
        }),
        ('Description', {
            'fields': ('description', 'notes')
        }),
        ('Accounting Period', {
            'fields': ('fiscal_year', 'accounting_period')
        }),
        ('Totals', {
            'fields': ('total_debit', 'total_credit', 'is_balanced_display')
        }),
        ('Status & Tracking', {
            'fields': ('status', 'created_by', 'posted_by')
        }),
        ('Links', {
            'fields': ('invoice', 'reversed_entry'),
            'classes': ('collapse',)
        }),
    )
    
    def description_short(self, obj):
        return obj.description[:50] + '...' if len(obj.description) > 50 else obj.description
    description_short.short_description = 'Description'
    
    def status_badge(self, obj):
        colors = {
            'draft': 'secondary',
            'posted': 'success',
            'void': 'danger',
            'reversed': 'warning'
        }
        color = colors.get(obj.status, 'secondary')
        return format_html(
            '<span class="badge bg-{}">{}</span>',
            color,
            obj.get_status_display()
        )
    status_badge.short_description = 'Status'
    
    def is_balanced_display(self, obj):
        if obj.is_balanced:
            return format_html('<span style="color: green;">✓ Balanced</span>')
        return format_html('<span style="color: red;">✗ Not Balanced</span>')
    is_balanced_display.short_description = 'Balanced?'
    
    def get_readonly_fields(self, request, obj=None):
        readonly = list(self.readonly_fields)
        if obj and obj.status == 'posted':
            readonly.extend(['journal', 'entry_date', 'fiscal_year', 'accounting_period'])
        return readonly
    
    actions = ['post_entries', 'void_entries']
    
    def post_entries(self, request, queryset):
        count = 0
        for entry in queryset.filter(status='draft'):
            try:
                entry.post(request.user)
                count += 1
            except Exception as e:
                self.message_user(request, f"Error posting {entry.entry_number}: {e}", level='error')
        
        self.message_user(request, f"Successfully posted {count} journal entries", level='success')
    post_entries.short_description = "Post selected journal entries"
    
    def void_entries(self, request, queryset):
        count = queryset.filter(status='posted').update(status='void')
        self.message_user(request, f"Voided {count} journal entries", level='warning')
    void_entries.short_description = "Void selected journal entries"


@admin.register(AdvancedGeneralLedger)
class AdvancedGeneralLedgerAdmin(admin.ModelAdmin):
    list_display = ['transaction_date', 'account', 'description_short', 'debit_amount', 'credit_amount', 'balance', 'journal_entry_link']
    list_filter = ['account', 'transaction_date', 'fiscal_year', 'accounting_period', 'is_voided']
    search_fields = ['description', 'journal_entry__entry_number', 'account__account_code', 'account__account_name']
    readonly_fields = ['journal_entry', 'journal_entry_line', 'account', 'transaction_date', 'posting_date', 'debit_amount', 'credit_amount', 'balance']
    date_hierarchy = 'transaction_date'
    
    def has_add_permission(self, request):
        return False  # Ledger entries created automatically from journal entries
    
    def has_delete_permission(self, request, obj=None):
        return False  # Cannot delete ledger entries
    
    def description_short(self, obj):
        return obj.description[:40] + '...' if len(obj.description) > 40 else obj.description
    description_short.short_description = 'Description'
    
    def journal_entry_link(self, obj):
        url = reverse('admin:hospital_advancedjournalentry_change', args=[obj.journal_entry.pk])
        return format_html('<a href="{}">{}</a>', url, obj.journal_entry.entry_number)
    journal_entry_link.short_description = 'Journal Entry'


# ==================== PAYMENT & RECEIPT VOUCHERS ====================

@admin.register(PaymentVoucher)
class PaymentVoucherAdmin(admin.ModelAdmin):
    list_display = ['voucher_number', 'voucher_date', 'payee_name', 'payment_type', 'amount', 'status_badge', 'payment_method']
    list_filter = ['status', 'payment_type', 'payment_method', 'voucher_date']
    search_fields = ['voucher_number', 'payee_name', 'description', 'payment_reference']
    readonly_fields = ['voucher_number', 'journal_entry', 'approved_date', 'created', 'modified']
    date_hierarchy = 'voucher_date'
    autocomplete_fields = ['expense_account', 'payment_account']
    
    fieldsets = (
        ('Voucher Information', {
            'fields': ('voucher_number', 'voucher_date', 'payment_type', 'status'),
            'description': 'Basic voucher identification and type'
        }),
        ('Payee Details', {
            'fields': ('payee_name', 'payee_type', 'amount'),
            'description': 'Who is being paid and how much'
        }),
        ('Description', {
            'fields': ('description',),
        }),
        ('Payment Details', {
            'fields': ('payment_method', 'payment_date', 'payment_reference', 'bank_name', 'account_number', 'cheque_number'),
            'description': 'Payment processing information',
            'classes': ('collapse',)
        }),
        ('Accounting Links', {
            'fields': ('expense_account', 'payment_account', 'journal_entry'),
            'description': 'Linked accounting accounts and journal entries'
        }),
        ('Supporting Documents', {
            'fields': ('invoice_number', 'po_number'),
            'classes': ('collapse',)
        }),
        ('Approval Workflow', {
            'fields': ('requested_by', 'approved_by', 'approved_date', 'paid_by'),
            'description': 'Approval and payment audit trail'
        }),
        ('Additional Information', {
            'fields': ('notes', 'created', 'modified'),
            'classes': ('collapse',)
        }),
    )
    
    def get_readonly_fields(self, request, obj=None):
        """Make certain fields readonly for existing vouchers"""
        readonly = ['voucher_number', 'journal_entry', 'approved_date', 'created', 'modified']
        if obj and obj.status in ['paid', 'void']:
            # Can't edit paid or voided vouchers
            readonly.extend(['amount', 'payee_name', 'expense_account', 'payment_account'])
        return readonly
    
    def status_badge(self, obj):
        colors = {
            'draft': 'secondary',
            'pending_approval': 'warning',
            'approved': 'info',
            'paid': 'success',
            'rejected': 'danger',
            'void': 'dark'
        }
        color = colors.get(obj.status, 'secondary')
        return format_html(
            '<span class="badge bg-{}">{}</span>',
            color,
            obj.get_status_display()
        )
    status_badge.short_description = 'Status'
    
    actions = ['approve_vouchers', 'mark_as_paid', 'export_to_excel']
    
    list_per_page = 50
    
    def approve_vouchers(self, request, queryset):
        count = 0
        for voucher in queryset.filter(status='pending_approval'):
            try:
                voucher.approve(request.user)
                count += 1
            except Exception as e:
                self.message_user(request, f"Error approving {voucher.voucher_number}: {e}", level='error')
        
        self.message_user(request, f"Approved {count} payment vouchers", level='success')
    approve_vouchers.short_description = "Approve selected vouchers"
    
    def mark_as_paid(self, request, queryset):
        """Mark selected approved vouchers as paid"""
        count = 0
        for voucher in queryset.filter(status='approved'):
            try:
                # Simple mark as paid without journal entry (for now)
                voucher.status = 'paid'
                voucher.payment_date = timezone.now().date()
                voucher.paid_by = request.user
                voucher.save()
                count += 1
            except Exception as e:
                self.message_user(request, f"Error marking {voucher.voucher_number} as paid: {e}", level='error')
        
        if count > 0:
            self.message_user(request, f"✅ Marked {count} vouchers as paid successfully!", level='success')
    mark_as_paid.short_description = "✅ Mark selected as paid"
    
    def export_to_excel(self, request, queryset):
        """Export selected vouchers to Excel"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        from datetime import datetime
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Payment Vouchers"
        
        # Headers
        headers = ['Voucher #', 'Date', 'Payee', 'Type', 'Amount (GHS)', 'Status', 'Payment Date', 'Reference']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Data
        for voucher in queryset:
            ws.append([
                voucher.voucher_number,
                voucher.voucher_date.strftime('%Y-%m-%d'),
                voucher.payee_name,
                voucher.get_payment_type_display(),
                float(voucher.amount),
                voucher.get_status_display(),
                voucher.payment_date.strftime('%Y-%m-%d') if voucher.payment_date else '',
                voucher.payment_reference or '',
            ])
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="payment_vouchers_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response
    export_to_excel.short_description = "📊 Export to Excel"


@admin.register(PettyCashTransaction)
class PettyCashTransactionAdmin(admin.ModelAdmin):
    list_display = ['transaction_number', 'transaction_date', 'payee_name', 'amount', 'status_badge', 'expense_account', 'created_by']
    list_filter = ['status', 'transaction_date', 'expense_account']
    search_fields = ['transaction_number', 'payee_name', 'description', 'invoice_number']
    readonly_fields = ['transaction_number', 'journal_entry', 'approved_date', 'rejected_date', 'created', 'modified']
    date_hierarchy = 'transaction_date'
    autocomplete_fields = ['expense_account', 'cost_center']
    
    fieldsets = (
        ('Transaction Information', {
            'fields': ('transaction_number', 'transaction_date', 'status'),
            'description': 'Basic transaction identification'
        }),
        ('Payee Details', {
            'fields': ('payee_name', 'payee_type', 'amount'),
            'description': 'Who is being paid and how much'
        }),
        ('Description', {
            'fields': ('description',),
        }),
        ('Accounting', {
            'fields': ('expense_account', 'cost_center', 'journal_entry'),
            'description': 'Expense account and cost center'
        }),
        ('Payment Details', {
            'fields': ('payment_date', 'receipt_number', 'invoice_number'),
            'description': 'Payment processing information',
            'classes': ('collapse',)
        }),
        ('Approval Workflow', {
            'fields': ('created_by', 'approved_by', 'approved_date', 'rejected_by', 'rejected_date', 'rejection_reason'),
            'description': 'Approval and rejection audit trail'
        }),
        ('Additional Information', {
            'fields': ('notes', 'created', 'modified'),
            'classes': ('collapse',)
        }),
    )
    
    def get_readonly_fields(self, request, obj=None):
        """Make certain fields readonly for existing transactions"""
        readonly = ['transaction_number', 'journal_entry', 'approved_date', 'rejected_date', 'created', 'modified']
        if obj and obj.status in ['paid', 'void']:
            readonly.extend(['amount', 'payee_name', 'expense_account'])
        return readonly
    
    def status_badge(self, obj):
        """Display status as colored badge"""
        colors = {
            'draft': 'secondary',
            'pending_approval': 'warning',
            'approved': 'info',
            'paid': 'success',
            'rejected': 'danger',
            'void': 'dark',
        }
        color = colors.get(obj.status, 'secondary')
        badge_text = obj.get_status_display()
        if obj.amount > 500 and obj.status == 'pending_approval':
            badge_text += ' (High Amount)'
        return format_html('<span class="badge bg-{}">{}</span>', color, badge_text)
    status_badge.short_description = 'Status'


@admin.register(ReceiptVoucher)
class ReceiptVoucherAdmin(admin.ModelAdmin):
    list_display = ['receipt_number', 'receipt_date', 'received_from', 'amount', 'payment_method', 'status_badge']
    list_filter = ['status', 'payment_method', 'receipt_date']
    search_fields = ['receipt_number', 'received_from', 'description', 'reference']
    readonly_fields = ['receipt_number', 'journal_entry', 'received_by']
    date_hierarchy = 'receipt_date'
    
    fieldsets = (
        ('Receipt Information', {
            'fields': ('receipt_number', 'receipt_date')
        }),
        ('From', {
            'fields': ('received_from', 'patient', 'description')
        }),
        ('Payment', {
            'fields': ('amount', 'payment_method', 'reference')
        }),
        ('Accounting', {
            'fields': ('revenue_account', 'cash_account', 'journal_entry', 'invoice')
        }),
        ('Status', {
            'fields': ('status', 'received_by')
        }),
    )
    
    def status_badge(self, obj):
        colors = {'draft': 'secondary', 'issued': 'success', 'void': 'danger'}
        color = colors.get(obj.status, 'secondary')
        return format_html('<span class="badge bg-{}">{}</span>', color, obj.get_status_display())
    status_badge.short_description = 'Status'


# ==================== REVENUE & EXPENSE ====================

@admin.register(RevenueCategory)
class RevenueCategoryAdmin(admin.ModelAdmin):
    list_display = ['code', 'name', 'account', 'is_active']
    list_filter = ['is_active']
    search_fields = ['code', 'name']


@admin.register(Revenue)
class RevenueAdmin(admin.ModelAdmin):
    list_display = ['revenue_number', 'revenue_date', 'category', 'amount', 'patient', 'payment_method']
    list_filter = ['category', 'payment_method', 'revenue_date']
    search_fields = ['revenue_number', 'description', 'patient__first_name', 'patient__last_name']
    readonly_fields = ['revenue_number', 'journal_entry', 'recorded_by']
    date_hierarchy = 'revenue_date'
    
    fieldsets = (
        ('Revenue Information', {
            'fields': ('revenue_number', 'revenue_date', 'category', 'description')
        }),
        ('Amount', {
            'fields': ('amount',)
        }),
        ('Source', {
            'fields': ('patient', 'invoice')
        }),
        ('Payment', {
            'fields': ('payment_method', 'reference')
        }),
        ('Accounting', {
            'fields': ('journal_entry', 'receipt_voucher', 'recorded_by')
        }),
    )


@admin.register(ExpenseCategory)
class ExpenseCategoryAdmin(admin.ModelAdmin):
    list_display = ['code', 'name', 'account', 'requires_approval', 'approval_limit', 'is_active']
    list_filter = ['requires_approval', 'is_active']
    search_fields = ['code', 'name']


@admin.register(Expense)
class ExpenseAdmin(admin.ModelAdmin):
    list_display = ['expense_number', 'expense_date', 'category', 'vendor_name', 'amount', 'status_badge']
    list_filter = ['status', 'category', 'expense_date']
    search_fields = ['expense_number', 'vendor_name', 'description', 'vendor_invoice_number']
    readonly_fields = ['expense_number', 'payment_voucher', 'journal_entry', 'approved_by']
    date_hierarchy = 'expense_date'
    
    fieldsets = (
        ('Expense Information', {
            'fields': ('expense_number', 'expense_date', 'category', 'description')
        }),
        ('Vendor', {
            'fields': ('vendor_name', 'vendor_invoice_number')
        }),
        ('Amount', {
            'fields': ('amount',)
        }),
        ('Status & Approval', {
            'fields': ('status', 'recorded_by', 'approved_by')
        }),
        ('Accounting', {
            'fields': ('payment_voucher', 'journal_entry')
        }),
    )
    
    def status_badge(self, obj):
        colors = {
            'draft': 'secondary',
            'pending': 'warning',
            'approved': 'info',
            'paid': 'success',
            'rejected': 'danger'
        }
        color = colors.get(obj.status, 'secondary')
        return format_html('<span class="badge bg-{}">{}</span>', color, obj.get_status_display())
    status_badge.short_description = 'Status'


# ==================== RECEIVABLES & PAYABLES ====================

@admin.register(AdvancedAccountsReceivable)
class AdvancedAccountsReceivableAdmin(admin.ModelAdmin):
    list_display = ['invoice', 'patient', 'invoice_amount', 'amount_paid', 'balance_due', 'due_date', 'aging_badge', 'overdue_days']
    list_filter = ['is_overdue', 'aging_bucket', 'due_date']
    search_fields = ['invoice__invoice_number', 'patient__first_name', 'patient__last_name']
    readonly_fields = ['balance_due', 'is_overdue', 'days_overdue', 'aging_bucket']
    date_hierarchy = 'due_date'
    
    def aging_badge(self, obj):
        colors = {
            'current': 'success',
            '0-30': 'info',
            '31-60': 'warning',
            '61-90': 'danger',
            '90+': 'dark'
        }
        color = colors.get(obj.aging_bucket, 'secondary')
        return format_html('<span class="badge bg-{}">{}</span>', color, obj.aging_bucket)
    aging_badge.short_description = 'Aging'
    
    def overdue_days(self, obj):
        if obj.is_overdue:
            return format_html('<span style="color: red;">{} days</span>', obj.days_overdue)
        return '-'
    overdue_days.short_description = 'Overdue'


@admin.register(AccountsPayable)
class AccountsPayableAdmin(admin.ModelAdmin):
    list_display = ['bill_number', 'vendor_name', 'bill_date', 'due_date', 'amount', 'balance_due', 'is_overdue', 'payment_voucher']
    list_filter = ['is_overdue', 'bill_date', 'due_date']
    search_fields = ['bill_number', 'vendor_name', 'vendor_invoice', 'description']
    readonly_fields = ['balance_due', 'is_overdue', 'days_overdue']
    date_hierarchy = 'due_date'


# ==================== BANK & CASH ====================

@admin.register(BankAccount)
class BankAccountAdmin(admin.ModelAdmin):
    list_display = ['account_name', 'bank_name', 'account_number', 'account_type', 'current_balance', 'currency', 'is_active']
    list_filter = ['bank_name', 'account_type', 'currency', 'is_active']
    search_fields = ['account_name', 'account_number', 'bank_name']
    
    fieldsets = (
        ('Account Information', {
            'fields': ('account_name', 'account_number', 'account_type', 'currency')
        }),
        ('Bank Details', {
            'fields': ('bank_name', 'branch')
        }),
        ('Balances', {
            'fields': ('opening_balance', 'current_balance')
        }),
        ('Accounting', {
            'fields': ('gl_account',)
        }),
        ('Status', {
            'fields': ('is_active',)
        }),
    )


@admin.register(BankTransaction)
class BankTransactionAdmin(admin.ModelAdmin):
    list_display = ['bank_account', 'transaction_date', 'transaction_type', 'description_short', 'amount', 'reconciled_badge']
    list_filter = ['bank_account', 'transaction_type', 'is_reconciled', 'transaction_date']
    search_fields = ['description', 'reference']
    date_hierarchy = 'transaction_date'
    
    def description_short(self, obj):
        return obj.description[:40] + '...' if len(obj.description) > 40 else obj.description
    description_short.short_description = 'Description'
    
    def reconciled_badge(self, obj):
        if obj.is_reconciled:
            return format_html('<span class="badge bg-success">Reconciled</span>')
        return format_html('<span class="badge bg-warning">Pending</span>')
    reconciled_badge.short_description = 'Reconciliation'
    
    actions = ['mark_reconciled']
    
    def mark_reconciled(self, request, queryset):
        count = queryset.update(is_reconciled=True, reconciled_date=timezone.now().date())
        self.message_user(request, f"Marked {count} transactions as reconciled", level='success')
    mark_reconciled.short_description = "Mark as reconciled"


# ==================== BUDGETING ====================

@admin.register(Budget)
class BudgetAdmin(admin.ModelAdmin):
    list_display = ['name', 'fiscal_year', 'start_date', 'end_date', 'total_revenue_budget', 'total_expense_budget', 'is_approved']
    list_filter = ['is_approved', 'fiscal_year', 'start_date']
    search_fields = ['name', 'description']
    
    inlines = [BudgetLineInline]
    
    fieldsets = (
        ('Budget Information', {
            'fields': ('name', 'fiscal_year', 'accounting_period', 'start_date', 'end_date', 'description')
        }),
        ('Totals', {
            'fields': ('total_revenue_budget', 'total_expense_budget')
        }),
        ('Approval', {
            'fields': ('is_approved', 'approved_by')
        }),
    )


@admin.register(TaxRate)
class TaxRateAdmin(admin.ModelAdmin):
    list_display = ['code', 'name', 'rate', 'account', 'effective_date', 'is_active']
    list_filter = ['is_active', 'effective_date']
    search_fields = ['code', 'name']


# ==================== AUDIT LOG ====================

@admin.register(AccountingAuditLog)
class AccountingAuditLogAdmin(admin.ModelAdmin):
    list_display = ['timestamp', 'user', 'action', 'model_name', 'object_repr']
    list_filter = ['action', 'model_name', 'timestamp']
    search_fields = ['object_repr', 'user__username']
    readonly_fields = ['user', 'action', 'timestamp', 'model_name', 'object_id', 'object_repr', 'changes', 'ip_address']
    date_hierarchy = 'timestamp'
    
    def has_add_permission(self, request):
        return False
    
    def has_delete_permission(self, request, obj=None):
        return False


# ==================== CASHBOOK ====================

@admin.register(Cashbook)
class CashbookAdmin(admin.ModelAdmin):
    list_display = ['entry_number', 'entry_type', 'entry_date', 'amount', 'payee_or_payer', 'status', 'held_until', 'can_classify_now']
    list_filter = ['entry_type', 'status', 'entry_date', 'payment_method']
    search_fields = ['entry_number', 'payee_or_payer', 'description', 'reference']
    readonly_fields = ['entry_number', 'held_until', 'classified_at', 'classified_by']
    date_hierarchy = 'entry_date'
    actions = ['classify_ready_entries']
    
    fieldsets = (
        ('Entry Information', {
            'fields': ('entry_number', 'entry_type', 'entry_date', 'amount', 'payee_or_payer', 'description', 'reference')
        }),
        ('Payment Details', {
            'fields': ('payment_method', 'cheque')
        }),
        ('Links', {
            'fields': ('patient', 'invoice')
        }),
        ('Account Classification', {
            'fields': ('cash_account', 'revenue_account', 'expense_account', 'status', 'held_until', 'classified_at', 'classified_by')
        }),
        ('Accounting', {
            'fields': ('journal_entry',)
        }),
    )
    
    def can_classify_now(self, obj):
        """Display if entry can be classified now"""
        if obj and obj.pk:  # Only check for saved objects
            try:
                return obj.can_classify()
            except (TypeError, AttributeError):
                return False
        return False
    can_classify_now.boolean = True
    can_classify_now.short_description = 'Ready to Classify'
    
    def classify_ready_entries(self, request, queryset):
        """Classify all ready cashbook entries to revenue/expense"""
        from django.utils import timezone
        from django.contrib import messages
        
        count = 0
        errors = []
        
        for entry in queryset.filter(status='pending'):
            if entry.can_classify():
                try:
                    # Use existing accounts if set, otherwise skip
                    if entry.entry_type == 'receipt' and entry.revenue_account:
                        entry.classify_to_revenue(request.user, entry.revenue_account)
                        count += 1
                    elif entry.entry_type == 'payment' and entry.expense_account:
                        entry.classify_to_revenue(request.user, expense_account=entry.expense_account)
                        count += 1
                    else:
                        errors.append(f"{entry.entry_number}: Missing {'revenue' if entry.entry_type == 'receipt' else 'expense'} account")
                except Exception as e:
                    errors.append(f"{entry.entry_number}: {str(e)}")
        
        if count > 0:
            self.message_user(request, f"Successfully classified {count} cashbook entries", level='success')
        if errors:
            self.message_user(request, f"Errors: {'; '.join(errors)}", level='error')
    classify_ready_entries.short_description = "Classify ready entries to revenue/expense"
    
    def get_readonly_fields(self, request, obj=None):
        readonly = list(self.readonly_fields)
        if obj and obj.status == 'classified':
            readonly.extend(['entry_type', 'amount', 'revenue_account', 'expense_account'])
        return readonly


# ==================== BANK RECONCILIATION ====================

class BankReconciliationItemInline(admin.TabularInline):
    model = BankReconciliationItem
    extra = 1
    fields = ['transaction', 'journal_entry', 'cheque', 'description', 'amount', 'is_matched']


@admin.register(BankReconciliation)
class BankReconciliationAdmin(admin.ModelAdmin):
    list_display = ['reconciliation_number', 'bank_account', 'statement_date', 'statement_balance', 'adjusted_balance', 'difference', 'status']
    list_filter = ['status', 'statement_date', 'bank_account']
    search_fields = ['reconciliation_number', 'bank_account__account_name']
    readonly_fields = ['reconciliation_number', 'book_balance', 'adjusted_balance', 'difference']
    inlines = [BankReconciliationItemInline]
    
    fieldsets = (
        ('Reconciliation Information', {
            'fields': ('reconciliation_number', 'bank_account', 'statement_date', 'statement_balance', 'book_balance')
        }),
        ('Adjustments', {
            'fields': ('deposits_in_transit', 'outstanding_cheques', 'bank_charges', 'interest_earned', 'other_adjustments')
        }),
        ('Results', {
            'fields': ('adjusted_balance', 'difference', 'status')
        }),
        ('Completion', {
            'fields': ('reconciled_by', 'reconciled_at', 'notes')
        }),
    )


# ==================== INSURANCE RECEIVABLE ====================

@admin.register(InsuranceReceivable)
class InsuranceReceivableAdmin(admin.ModelAdmin):
    list_display = ['receivable_number', 'insurance_company', 'patient', 'total_amount', 'amount_paid', 'balance_due', 'status', 'due_date']
    list_filter = ['status', 'claim_date', 'insurance_company']
    search_fields = ['receivable_number', 'claim_number', 'patient__first_name', 'patient__last_name', 'insurance_company__name']
    readonly_fields = ['receivable_number', 'balance_due']
    date_hierarchy = 'claim_date'
    
    fieldsets = (
        ('Receivable Information', {
            'fields': ('receivable_number', 'insurance_company', 'patient', 'invoice', 'claim_number', 'claim_date')
        }),
        ('Amounts', {
            'fields': ('total_amount', 'amount_paid', 'balance_due', 'due_date', 'payment_date')
        }),
        ('Status Tracking', {
            'fields': ('status', 'submitted_date', 'approved_date', 'rejection_reason')
        }),
        ('Accounting', {
            'fields': ('receivable_account', 'journal_entry')
        }),
        ('Notes', {
            'fields': ('notes',)
        }),
    )


@admin.register(InsuranceReceivableEntry)
class InsuranceReceivableEntryAdmin(admin.ModelAdmin):
    """Admin for Insurance Receivable Entry (PrimeCare Accounting) - INSURANCE ONLY, EXCLUDES CORPORATE"""
    list_display = ['entry_number', 'payer', 'entry_date', 'total_amount', 'outstanding_amount', 'status', 'entry_date']
    list_filter = ['status', 'entry_date', 'payer']
    search_fields = ['entry_number', 'payer__name', 'notes']
    readonly_fields = ['entry_number', 'outstanding_amount']
    date_hierarchy = 'entry_date'
    
    def get_queryset(self, request):
        """Filter out corporate payers - only show insurance receivables (private/nhis)"""
        qs = super().get_queryset(request)
        return qs.exclude(payer__payer_type='corporate')
    
    def changelist_view(self, request, extra_context=None):
        """Add a note about corporate receivables being excluded"""
        extra_context = extra_context or {}
        extra_context['title'] = 'Insurance Receivable Entries (Insurance Companies Only)'
        return super().changelist_view(request, extra_context)
    
    fieldsets = (
        ('Entry Information', {
            'fields': ('entry_number', 'payer', 'entry_date', 'status')
        }),
        ('Amounts', {
            'fields': ('total_amount', 'amount_received', 'amount_rejected', 'withholding_tax', 'outstanding_amount')
        }),
        ('Revenue Breakdown', {
            'fields': (
                'registration_amount', 'consultation_amount', 'laboratory_amount',
                'pharmacy_amount', 'surgeries_amount', 'admissions_amount',
                'radiology_amount', 'dental_amount', 'physiotherapy_amount'
            )
        }),
        ('Accounting', {
            'fields': ('journal_entry', 'matched_at', 'matched_by')
        }),
        ('Notes', {
            'fields': ('notes',)
        }),
    )


@admin.register(CorporateReceivableEntry)
class CorporateReceivableEntryAdmin(admin.ModelAdmin):
    """Admin for Corporate Receivable Entry - CORPORATE CLIENTS ONLY"""
    list_display = ['entry_number', 'payer', 'entry_date', 'total_amount', 'outstanding_amount', 'status', 'entry_date']
    list_filter = ['status', 'entry_date', 'payer']
    search_fields = ['entry_number', 'payer__name', 'notes']
    readonly_fields = ['entry_number', 'outstanding_amount']
    date_hierarchy = 'entry_date'
    
    def get_queryset(self, request):
        """Only show corporate payers"""
        qs = super().get_queryset(request)
        return qs.filter(payer__payer_type='corporate')
    
    def changelist_view(self, request, extra_context=None):
        """Add a note about this being corporate receivables only"""
        extra_context = extra_context or {}
        extra_context['title'] = 'Corporate Receivable Entries (Corporate Clients Only)'
        return super().changelist_view(request, extra_context)
    
    fieldsets = (
        ('Entry Information', {
            'fields': ('entry_number', 'payer', 'entry_date', 'status')
        }),
        ('Amounts', {
            'fields': ('total_amount', 'amount_received', 'amount_rejected', 'withholding_tax', 'outstanding_amount')
        }),
        ('Revenue Breakdown', {
            'fields': (
                'registration_amount', 'consultation_amount', 'laboratory_amount',
                'pharmacy_amount', 'surgeries_amount', 'admissions_amount',
                'radiology_amount', 'dental_amount', 'physiotherapy_amount'
            )
        }),
        ('Accounting', {
            'fields': ('journal_entry', 'matched_at', 'matched_by')
        }),
        ('Notes', {
            'fields': ('notes',)
        }),
    )


# ==================== PROCUREMENT PURCHASE ====================

@admin.register(ProcurementPurchase)
class ProcurementPurchaseAdmin(admin.ModelAdmin):
    list_display = ['purchase_number', 'purchase_type', 'supplier_name', 'total_amount', 'status', 'purchase_date']
    list_filter = ['purchase_type', 'status', 'purchase_date']
    search_fields = ['purchase_number', 'supplier_name', 'supplier_invoice', 'description']
    readonly_fields = ['purchase_number', 'net_amount']
    date_hierarchy = 'purchase_date'
    
    fieldsets = (
        ('Purchase Information', {
            'fields': ('purchase_number', 'purchase_type', 'purchase_date', 'supplier_name', 'supplier_invoice', 'description', 'classification')
        }),
        ('Amounts', {
            'fields': ('total_amount', 'tax_amount', 'net_amount')
        }),
        ('Status', {
            'fields': ('status', 'created_by', 'approved_by')
        }),
        ('Accounting', {
            'fields': ('expense_account', 'liability_account', 'payment_account', 'accounts_payable', 'journal_entry')
        }),
    )


# ==================== PAYROLL ====================

class AccountingPayrollEntryInline(admin.TabularInline):
    model = AccountingPayrollEntry
    extra = 0
    fields = [
        'staff', 'sheet_serial', 'department_snapshot', 'basic_salary', 'other_allowances',
        'medical_allowance', 'risk_emergency_allowance', 'personal_relief', 'taxable_income',
        'loan_deduction', 'gross_pay',
        'ssnit_employee', 'pension_employee', 'paye_tax', 'deductions', 'net_pay',
    ]
    readonly_fields = ['gross_pay', 'deductions', 'net_pay']


@admin.register(AccountingPayroll)
class AccountingPayrollAdmin(admin.ModelAdmin):
    list_display = [
        'payroll_number', 'payroll_period_start', 'payroll_period_end', 'pay_date',
        'total_net_pay', 'deduction_apply_percentages', 'status',
    ]
    list_filter = ['status', 'deduction_apply_percentages', 'payroll_period_start', 'payroll_period_end']
    search_fields = ['payroll_number']
    readonly_fields = ['payroll_number', 'total_net_pay']
    inlines = [AccountingPayrollEntryInline]
    date_hierarchy = 'payroll_period_end'
    
    fieldsets = (
        ('Payroll Information', {
            'fields': ('payroll_number', 'payroll_period_start', 'payroll_period_end', 'pay_date', 'period_label', 'import_source_filename', 'status')
        }),
        ('Automatic deduction percentages', {
            'description': (
                'Turn on the switch, set rates, then save. All staff lines are recalculated from earnings '
                '(SSF/PF use basic salary, or gross earnings if basic is 0). Leave a rate at 0 to type that amount manually on each line. '
                'PAYE uses Taxable income when set; otherwise (gross earnings - personal relief) when the PAYE % is above zero.'
            ),
            'fields': (
                'deduction_apply_percentages',
                'deduction_ssnit_employee_pct',
                'deduction_pension_employee_pct',
                'deduction_paye_pct',
                'deduction_other_deduction_pct',
            ),
        }),
        ('Totals', {
            'fields': ('total_gross_pay', 'total_deductions', 'total_net_pay')
        }),
        ('Approval', {
            'fields': ('created_by', 'approved_by')
        }),
    )

    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)
        obj = form.instance
        if getattr(obj, 'deduction_apply_percentages', False):
            obj.apply_percentage_deductions_to_all_entries()


@admin.register(DoctorCommission)
class DoctorCommissionAdmin(admin.ModelAdmin):
    list_display = ['commission_number', 'doctor', 'service_type', 'total_fee', 'doctor_share', 'operational_share', 'hospital_share', 'is_paid', 'service_date']
    list_filter = ['service_type', 'is_paid', 'service_date']
    search_fields = ['commission_number', 'doctor__user__first_name', 'doctor__user__last_name']
    readonly_fields = ['commission_number', 'doctor_share', 'operational_share', 'hospital_share']
    date_hierarchy = 'service_date'
    
    fieldsets = (
        ('Commission Information', {
            'fields': ('commission_number', 'doctor', 'service_type', 'service_date', 'invoice', 'total_fee')
        }),
        ('Commission Splits', {
            'fields': ('doctor_share', 'operational_share', 'hospital_share')
        }),
        ('Payment', {
            'fields': ('is_paid', 'paid_date')
        }),
        ('Accounting', {
            'fields': ('doctor_receivable_account', 'hospital_revenue_account', 'operational_account', 'journal_entry')
        }),
    )


# ==================== INCOME GROUPING & PROFIT/LOSS ====================

@admin.register(IncomeGroup)
class IncomeGroupAdmin(admin.ModelAdmin):
    list_display = ['code', 'name', 'account', 'is_active']
    list_filter = ['is_active']
    search_fields = ['code', 'name']


@admin.register(ProfitLossReport)
class ProfitLossReportAdmin(admin.ModelAdmin):
    list_display = ['report_number', 'report_period', 'period_start', 'period_end', 'total_revenue', 'total_expenses', 'net_profit', 'profit_percentage']
    list_filter = ['report_period', 'fiscal_year', 'period_start']
    search_fields = ['report_number']
    readonly_fields = ['report_number', 'gross_profit', 'net_profit', 'profit_percentage', 'generated_at', 'generated_by']
    date_hierarchy = 'period_end'
    
    fieldsets = (
        ('Report Information', {
            'fields': ('report_number', 'report_period', 'period_start', 'period_end', 'fiscal_year')
        }),
        ('Revenue', {
            'fields': ('total_revenue', 'revenue_by_category'),
            'description': 'Enter total revenue and optionally break down by category as JSON'
        }),
        ('Expenses', {
            'fields': ('total_expenses', 'expenses_by_category'),
            'description': 'Enter total expenses and optionally break down by category as JSON'
        }),
        ('Profit/Loss', {
            'fields': ('gross_profit', 'net_profit', 'profit_percentage'),
            'description': 'Calculated automatically from revenue and expenses'
        }),
        ('Generation', {
            'fields': ('generated_at', 'generated_by'),
            'classes': ('collapse',)
        }),
    )
    
    def save_model(self, request, obj, form, change):
        """Set generated_by when creating new report"""
        if not change:  # New object
            obj.generated_by = request.user
        super().save_model(request, obj, form, change)


# ==================== REGISTRATION FEE ====================

@admin.register(RegistrationFee)
class RegistrationFeeAdmin(admin.ModelAdmin):
    list_display = ['fee_number', 'patient', 'registration_date', 'fee_amount', 'payment_method']
    list_filter = ['payment_method', 'registration_date']
    search_fields = ['fee_number', 'patient__first_name', 'patient__last_name', 'patient__mrn']
    readonly_fields = ['fee_number']
    date_hierarchy = 'registration_date'
    
    fieldsets = (
        ('Fee Information', {
            'fields': ('fee_number', 'patient', 'registration_date', 'fee_amount', 'payment_method')
        }),
        ('Accounting', {
            'fields': ('revenue_account', 'journal_entry')
        }),
        ('Notes', {
            'fields': ('notes',)
        }),
    )


# ==================== CASH SALES & CORPORATE ====================

@admin.register(CashSale)
class CashSaleAdmin(admin.ModelAdmin):
    list_display = ['sale_number', 'customer_name', 'sale_date', 'total_amount', 'payment_method']
    list_filter = ['payment_method', 'sale_date']
    search_fields = ['sale_number', 'customer_name', 'description']
    readonly_fields = ['sale_number']
    date_hierarchy = 'sale_date'
    
    fieldsets = (
        ('Sale Information', {
            'fields': ('sale_number', 'sale_date', 'customer_name', 'description', 'total_amount', 'payment_method')
        }),
        ('Accounting', {
            'fields': ('revenue_account', 'cash_account', 'journal_entry', 'created_by')
        }),
    )


@admin.register(AccountingCorporateAccount)
class AccountingCorporateAccountAdmin(admin.ModelAdmin):
    list_display = ['account_number', 'company_name', 'contact_person', 'credit_limit_display', 'current_balance_display', 'is_active']
    list_filter = ['is_active']
    search_fields = ['account_number', 'company_name', 'contact_person', 'contact_email']
    
    fieldsets = (
        ('Company Information', {
            'fields': ('account_number', 'company_name', 'contact_person', 'contact_email', 'contact_phone')
        }),
        ('Credit Management', {
            'fields': ('credit_limit', 'current_balance')
        }),
        ('Accounting', {
            'fields': ('receivable_account',)
        }),
        ('Status', {
            'fields': ('is_active',)
        }),
    )
    
    def credit_limit_display(self, obj):
        """Display credit limit formatted"""
        if obj.credit_limit is None:
            return '-'
        try:
            limit = float(obj.credit_limit)
            # Return plain string, not SafeString
            return "GHS {:,.2f}".format(limit)
        except (ValueError, TypeError, AttributeError):
            return str(obj.credit_limit) if obj.credit_limit else '-'
    credit_limit_display.short_description = 'Credit Limit'
    credit_limit_display.admin_order_field = 'credit_limit'
    
    def current_balance_display(self, obj):
        """Display current balance formatted"""
        if obj.current_balance is None:
            return '-'
        try:
            balance = float(obj.current_balance)
            # Format number as plain string first
            balance_formatted = "{:,.2f}".format(balance)
            if obj.credit_limit and balance >= float(obj.credit_limit):
                # Pass the formatted string (not SafeString) to format_html
                return format_html('<span style="color: red; font-weight: bold;">GHS {}</span>', balance_formatted)
            # Return plain string for normal cases
            return "GHS {}".format(balance_formatted)
        except (ValueError, TypeError, AttributeError):
            return str(obj.current_balance) if obj.current_balance else '-'
    current_balance_display.short_description = 'Current Balance'
    current_balance_display.admin_order_field = 'current_balance'


# ==================== WITHHOLDING RECEIVABLE ====================

@admin.register(WithholdingReceivable)
class WithholdingReceivableAdmin(admin.ModelAdmin):
    list_display = ['withholding_number', 'payer', 'withholding_date', 'amount_withheld', 'amount_recovered', 'balance', 'expected_recovery_date']
    list_filter = ['withholding_date', 'expected_recovery_date']
    search_fields = ['withholding_number', 'payer', 'description']
    readonly_fields = ['withholding_number', 'balance']
    date_hierarchy = 'withholding_date'
    
    fieldsets = (
        ('Withholding Information', {
            'fields': ('withholding_number', 'withholding_date', 'payer', 'description', 'amount_withheld')
        }),
        ('Recovery', {
            'fields': ('amount_recovered', 'balance', 'expected_recovery_date', 'recovered_date')
        }),
        ('Accounting', {
            'fields': ('receivable_account',)
        }),
        ('Notes', {
            'fields': ('notes',)
        }),
    )


# ==================== WITHHOLDING TAX PAYABLE ====================

@admin.register(WithholdingTaxPayable)
class WithholdingTaxPayableAdmin(admin.ModelAdmin):
    list_display = ['withholding_number', 'supplier_name', 'withholding_date', 'gross_amount', 'withholding_rate', 'withholding_amount', 'balance_due', 'supply_type']
    list_filter = ['withholding_date', 'supply_type', 'is_exempted', 'due_date']
    search_fields = ['withholding_number', 'supplier_name', 'supplier_tin', 'description']
    readonly_fields = ['withholding_number', 'withholding_amount', 'net_amount_paid', 'balance_due']
    date_hierarchy = 'withholding_date'
    
    fieldsets = (
        ('Withholding Tax Information', {
            'fields': ('withholding_number', 'withholding_date', 'supplier_name', 'supplier_tin', 'is_exempted', 'supply_type')
        }),
        ('Amounts', {
            'fields': ('gross_amount', 'withholding_rate', 'withholding_amount', 'net_amount_paid')
        }),
        ('Payment Status', {
            'fields': ('amount_paid', 'balance_due', 'due_date', 'paid_date')
        }),
        ('Source Transaction', {
            'fields': ('accounts_payable', 'payment_voucher')
        }),
        ('Accounting', {
            'fields': ('payable_account', 'journal_entry')
        }),
        ('Details', {
            'fields': ('description', 'notes')
        }),
    )


# ==================== DEPOSITS ====================

@admin.register(Deposit)
class DepositAdmin(admin.ModelAdmin):
    list_display = ['deposit_number', 'deposit_type', 'deposit_date', 'amount', 'from_account', 'to_account']
    list_filter = ['deposit_type', 'deposit_date']
    search_fields = ['deposit_number', 'description', 'reference']
    readonly_fields = ['deposit_number']
    date_hierarchy = 'deposit_date'
    
    fieldsets = (
        ('Deposit Information', {
            'fields': ('deposit_number', 'deposit_type', 'deposit_date', 'amount', 'description', 'reference')
        }),
        ('From/To Accounts', {
            'fields': ('from_account', 'to_account', 'from_bank_account', 'to_bank_account')
        }),
        ('Accounting', {
            'fields': ('journal_entry', 'created_by')
        }),
    )


# ==================== INITIAL REVALUATIONS ====================

@admin.register(InitialRevaluation)
class InitialRevaluationAdmin(admin.ModelAdmin):
    list_display = ['revaluation_number', 'account', 'revaluation_date', 'previous_value', 'new_value', 'revaluation_amount', 'revaluation_type']
    list_filter = ['revaluation_type', 'revaluation_date', 'effective_date']
    search_fields = ['revaluation_number', 'account__account_name', 'asset_description']
    readonly_fields = ['revaluation_number', 'revaluation_amount', 'revaluation_type']
    date_hierarchy = 'revaluation_date'
    
    fieldsets = (
        ('Revaluation Information', {
            'fields': ('revaluation_number', 'revaluation_date', 'effective_date', 'account', 'asset_description')
        }),
        ('Values', {
            'fields': ('previous_value', 'new_value', 'revaluation_amount', 'revaluation_type')
        }),
        ('Details', {
            'fields': ('reason',)
        }),
        ('Approval', {
            'fields': ('approved_by', 'approved_at')
        }),
        ('Accounting', {
            'fields': ('journal_entry',)
        }),
    )

