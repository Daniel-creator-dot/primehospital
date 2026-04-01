"""
Comprehensive Billing & Claims Management Views
All billing features for accountants including claims, corporate bills, insurance bills, and reports
"""
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count, F, Case, When, DecimalField, Value, Max, Exists, OuterRef, CharField, Subquery, Prefetch, ExpressionWrapper
from django.db.models.functions import TruncDate, TruncWeek, TruncMonth, Coalesce
from django.utils import timezone
from django.utils.text import slugify
from django.core.paginator import Paginator, Page
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from decimal import Decimal
from datetime import date, datetime, timedelta
from types import SimpleNamespace
from urllib.parse import urlencode
from collections import defaultdict
import csv
import io
import json
import logging

logger = logging.getLogger(__name__)

from .models import Patient, Invoice, InvoiceLine, Payer, Staff
from .models_workflow import Bill
from .models_insurance import InsuranceClaimItem, MonthlyInsuranceClaim
from .models_enterprise_billing import (
    CorporateAccount,
    MonthlyStatement,
    CorporateEmployee,
    StatementLine,
    STATEMENT_STATUS_CHOICES,
)
from .utils_roles import get_user_role
from .decorators import role_required
from .models_primecare_accounting import InsuranceReceivableEntry, InsurancePaymentReceived
from .insurance_claim_query import insurance_claim_item_deduped_q


def is_accountant(user):
    """Check if user is accountant or senior account officer"""
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    role = get_user_role(user)
    return role in ('accountant', 'senior_account_officer')


# ==================== CLAIMS BILLS HUB ====================

@login_required
@role_required('accountant', 'senior_account_officer')
def claims_bills_hub(request):
    """Main hub for all billing and claims features"""
    today = timezone.now().date()
    start_of_month = today.replace(day=1)
    
    # Statistics (pending = unpaid: draft, issued, partially_paid; exclude cancelled/paid)
    bill_base = Bill.objects.filter(is_deleted=False)
    stats = {
        'total_bills': bill_base.count(),
        'pending_bills': bill_base.filter(
            status__in=['draft', 'issued', 'partially_paid']
        ).count(),
        'paid_bills': bill_base.filter(status='paid').count(),
        'total_bill_amount': bill_base.aggregate(
            total=Sum('total_amount')
        )['total'] or Decimal('0.00'),
        'outstanding_bills': bill_base.filter(
            status__in=['issued', 'partially_paid']
        ).aggregate(total=Sum('patient_portion'))['total'] or Decimal('0.00'),
        'insurance_claims_pending': InsuranceClaimItem.objects.filter(
            claim_status='pending',
            is_deleted=False,
        )
        .filter(insurance_claim_item_deduped_q())
        .count(),
        'corporate_accounts': CorporateAccount.objects.filter(is_deleted=False).count(),
        'monthly_statements': MonthlyStatement.objects.filter(
            statement_date__gte=start_of_month,
            is_deleted=False
        ).count(),
    }
    
    # Recent bills
    recent_bills = Bill.objects.filter(is_deleted=False).select_related(
        'patient', 'issued_by'
    ).order_by('-issued_at')[:10]
    
    # Recent claims
    recent_claims = (
        InsuranceClaimItem.objects.filter(is_deleted=False)
        .filter(insurance_claim_item_deduped_q())
        .select_related('patient', 'payer')
        .order_by('-created')[:10]
    )
    
    context = {
        'stats': stats,
        'recent_bills': recent_bills,
        'recent_claims': recent_claims,
        'today': today,
    }
    
    return render(request, 'hospital/billing/claims_bills_hub.html', context)


# ==================== BILLS MANAGEMENT ====================

def _normalize_filter_value(val):
    """Treat INVALID/empty as no filter so form and results behave correctly."""
    if val is None or (isinstance(val, str) and val.strip().lower() == 'invalid'):
        return ''
    return val.strip() if isinstance(val, str) else val


def _patient_search_q(search):
    """
    Build Q object for patient search. Do NOT use patient__full_name - it's a
    Python @property, not a DB field; using it in filter causes FieldError.
    """
    return (
        Q(patient__first_name__icontains=search) |
        Q(patient__last_name__icontains=search) |
        Q(patient__middle_name__icontains=search) |
        Q(patient__mrn__icontains=search)
    )


def _invoice_search_q(search):
    """Build Q for invoice/patient search (Invoice model)."""
    return (
        Q(invoice_number__icontains=search) |
        Q(patient__first_name__icontains=search) |
        Q(patient__last_name__icontains=search) |
        Q(patient__middle_name__icontains=search) |
        Q(patient__mrn__icontains=search)
    )


def _invoice_bill_type_q(bill_type_filter):
    """Map bill_type filter to Invoice queryset Q (by payer type). Used when showing invoices as bills."""
    if not bill_type_filter:
        return Q()
    if bill_type_filter == 'insurance':
        return Q(payer__payer_type__in=['private', 'nhis', 'insurance'])
    if bill_type_filter == 'cash':
        return Q(payer__isnull=True) | Q(payer__payer_type='cash')
    if bill_type_filter == 'corporate':
        return Q(payer__payer_type='corporate')
    # 'mixed' = no filter (show all)
    return Q()


def _invoice_fallback_base_queryset():
    """
    Invoices shown when the site has no workflow Bill records — same inclusion rules as bills_list.
    Uses default Invoice manager (excludes write-off window, zero-total unless unwaived lines).
    """
    has_unwaived_line = Exists(
        InvoiceLine.objects.filter(
            invoice_id=OuterRef('pk'),
            is_deleted=False,
            waived_at__isnull=True
        )
    )
    return Invoice.objects.filter(
        is_deleted=False,
        status__in=['draft', 'issued', 'partially_paid', 'overdue', 'paid'],
    ).filter(Q(total_amount__gt=0) | has_unwaived_line)


def _invoice_accountant_summary_queryset():
    """
    Invoices for Bill Summary aggregates. Uses all_objects so rows in the write-off archive
    window (Invoice.WRITE_OFF_START–WRITE_OFF_END) still count — default Invoice.manager hides them.
    """
    has_unwaived_line = Exists(
        InvoiceLine.objects.filter(
            invoice_id=OuterRef('pk'),
            is_deleted=False,
            waived_at__isnull=True
        )
    )
    return Invoice.all_objects.filter(
        is_deleted=False,
        status__in=['draft', 'issued', 'partially_paid', 'overdue', 'paid'],
    ).filter(Q(total_amount__gt=0) | has_unwaived_line)


def _use_invoice_for_billing_aggregates():
    """True when invoice rows should drive accountant bill summary/total reports (see bill_summary)."""
    n_bills = Bill.objects.filter(is_deleted=False).count()
    n_invoices = _invoice_accountant_summary_queryset().count()
    return n_invoices > 0 and (n_bills == 0 or n_invoices > n_bills)


def _invoice_summary_category_case():
    """Annotate bucket for summary-by-type (cash / insurance / corporate)."""
    return Case(
        When(payer__payer_type__in=['private', 'nhis', 'insurance'], then=Value('insurance')),
        When(payer__payer_type='corporate', then=Value('corporate')),
        default=Value('cash'),
        output_field=CharField(),
    )


def _invoice_bill_type_display(invoice_or_payer_type):
    """Return (bill_type, display_label) for an invoice or payer_type string. For adapter/grouped row."""
    if hasattr(invoice_or_payer_type, 'payer'):
        payer = invoice_or_payer_type.payer
        pt = payer.payer_type if payer else 'cash'
    else:
        pt = invoice_or_payer_type or 'cash'
    if pt in ('private', 'nhis', 'insurance'):
        return 'insurance', 'Insurance Bill'
    if pt == 'corporate':
        return 'cash', 'Corporate'
    return 'cash', 'Cash Bill'


_STATUS_DISPLAY = {'draft': 'Draft', 'issued': 'Issued', 'partially_paid': 'Partially Paid', 'paid': 'Paid', 'overdue': 'Overdue', 'cancelled': 'Cancelled'}


def _batch_invoice_payer_names(invoice_qs, patient_ids):
    """Distinct payer names per patient (for combined / All Bills invoice fallback rows)."""
    if not patient_ids:
        return {}
    labels = defaultdict(set)
    for pid, pname in invoice_qs.filter(patient_id__in=patient_ids).values_list('patient_id', 'payer__name'):
        if pname and str(pname).strip():
            labels[pid].add(str(pname).strip())
    return labels


def _batch_invoice_amount_paid(invoice_qs, patient_ids):
    """Sum of (total_amount - balance) per patient — payments applied across invoices."""
    if not patient_ids:
        return {}
    rows = invoice_qs.filter(patient_id__in=patient_ids).values('patient_id').annotate(
        paid=Sum(F('total_amount') - F('balance'), output_field=DecimalField(max_digits=18, decimal_places=2))
    )
    return {r['patient_id']: (r['paid'] or Decimal('0')) for r in rows}


def _batch_patient_deposit_available(patient_ids):
    """Sum of active deposit available_balance per patient."""
    if not patient_ids:
        return {}
    try:
        from .models_patient_deposits import PatientDeposit
        rows = PatientDeposit.objects.filter(
            patient_id__in=patient_ids,
            is_deleted=False,
            status='active',
        ).values('patient_id').annotate(s=Sum('available_balance'))
        return {r['patient_id']: (r['s'] or Decimal('0')) for r in rows}
    except Exception:
        return {}


# Short labels for Payer column on cash bills (from Transaction.payment_method)
_PAYMENT_CHANNEL_SHORT = {
    'cash': 'Cash',
    'mobile_money': 'MoMo',
    'card': 'Card',
    'bank_transfer': 'Bank transfer',
    'cheque': 'Cheque',
    'insurance': 'Insurance',
    'deposit': 'Deposit',
}
_PAYMENT_CHANNEL_ORDER = (
    'cash', 'mobile_money', 'card', 'bank_transfer', 'cheque', 'deposit', 'insurance',
)


def _format_payment_channel_labels(methods):
    """Turn distinct payment_method codes into a stable display string."""
    if not methods:
        return ''
    s = set(methods)
    parts = []
    for m in _PAYMENT_CHANNEL_ORDER:
        if m in s:
            parts.append(_PAYMENT_CHANNEL_SHORT.get(m, m.replace('_', ' ').title()))
    for m in sorted(s):
        if m not in _PAYMENT_CHANNEL_ORDER:
            parts.append(_PAYMENT_CHANNEL_SHORT.get(m, str(m)))
    return ', '.join(parts)


def _txn_payment_methods_by_invoice(invoice_ids):
    """
    invoice_id -> set of payment_method codes.

    Uses Transaction rows, PaymentReceipt (Transaction.invoice is sometimes null until
    after receipt creation), and PaymentAllocation (combined payments attach extra
    invoices only via allocations).
    """
    if not invoice_ids:
        return {}
    from .models_accounting import Transaction, PaymentReceipt, PaymentAllocation

    out = defaultdict(set)

    rows = Transaction.objects.filter(
        invoice_id__in=invoice_ids,
        transaction_type='payment_received',
        is_deleted=False,
    ).values_list('invoice_id', 'payment_method')
    for inv_id, pm in rows:
        if inv_id and pm:
            out[inv_id].add(pm)

    for inv_id, pm in PaymentReceipt.objects.filter(
        invoice_id__in=invoice_ids,
        is_deleted=False,
    ).values_list('invoice_id', 'payment_method'):
        if inv_id and pm:
            out[inv_id].add(pm)

    for inv_id, pm in PaymentAllocation.objects.filter(
        invoice_id__in=invoice_ids,
        is_deleted=False,
        payment_transaction__transaction_type='payment_received',
        payment_transaction__is_deleted=False,
    ).values_list('invoice_id', 'payment_transaction__payment_method'):
        if inv_id and pm:
            out[inv_id].add(pm)

    return out


def _batch_payment_channel_labels_by_patient(invoice_qs, patient_ids):
    """One formatted channel string per patient (all invoices in invoice_qs)."""
    if not patient_ids:
        return {}
    pairs = list(invoice_qs.filter(patient_id__in=patient_ids).values_list('id', 'patient_id'))
    if not pairs:
        return {}
    inv_ids = [p[0] for p in pairs]
    inv_to_patient = dict(pairs)
    by_inv = _txn_payment_methods_by_invoice(inv_ids)
    by_patient = defaultdict(set)
    for inv_id, methods in by_inv.items():
        pid = inv_to_patient.get(inv_id)
        if pid:
            by_patient[pid].update(methods)
    return {pid: _format_payment_channel_labels(ms) for pid, ms in by_patient.items()}


def _batch_payment_channel_labels_by_invoice(invoice_ids):
    """invoice_id -> formatted label string."""
    by_inv = _txn_payment_methods_by_invoice(invoice_ids)
    return {iid: _format_payment_channel_labels(ms) for iid, ms in by_inv.items()}


def _invoice_is_cash_bill(inv):
    """True when invoice is patient/cash responsibility (not insurance/corporate payer)."""
    if not inv:
        return False
    pay = getattr(inv, 'payer', None)
    if not pay:
        return True
    return getattr(pay, 'payer_type', None) == 'cash'


def _compose_payer_column_for_patient(payer_names, channel_label_str):
    """
    Merge insurance/corporate payer names with cash payment channels for grouped rows.
    payer_names: set of payer__name from invoices; channel_label_str from transactions.
    """
    payer_names = set(payer_names or [])
    channel_label_str = (channel_label_str or '').strip()
    ordered_non_cash = sorted(
        n for n in payer_names if n and str(n).strip().lower() != 'cash'
    )
    parts = list(ordered_non_cash)
    if channel_label_str:
        parts.append(channel_label_str)
    elif any(not n or str(n).strip().lower() == 'cash' for n in payer_names):
        parts.append('Cash')
    if not parts:
        return '—'
    return ', '.join(parts)


def _empty_bills_analytics(mode='invoice'):
    z = 0.0
    return {
        'mode': mode,
        'summary': {
            'patient_count': 0,
            'invoice_or_bill_count': 0,
            'total_billed': z,
            'total_balance': z,
            'total_paid': z,
            'deposits_available': z,
            'negative_balance_invoices': 0,
            'insurance_covered': z,
            'patient_portion_sum': z,
        },
        'charts': {
            'status': {'labels': [], 'counts': [], 'amounts': []},
            'payer_type': {'labels': [], 'counts': [], 'amounts': []},
            'bill_type': {'labels': [], 'counts': [], 'amounts': []},
            'daily': {'labels': [], 'amounts': [], 'counts': []},
        },
    }


def _fdec(v):
    if v is None:
        return 0.0
    return float(v)


def analytics_for_invoice_queryset(qs):
    """Aggregates for All Bills when listing Invoices (respects same filters as the table)."""
    out = _empty_bills_analytics('invoice')
    cnt = qs.count()
    if cnt == 0:
        return out

    agg = qs.aggregate(
        st=Sum('total_amount'),
        sb=Sum('balance'),
        sp=Sum(F('total_amount') - F('balance'), output_field=DecimalField(max_digits=20, decimal_places=2)),
    )
    neg = qs.filter(balance__lt=0).count()
    patients = qs.aggregate(n=Count('patient_id', distinct=True))['n'] or 0

    deposits = Decimal('0')
    try:
        from .models_patient_deposits import PatientDeposit
        deposits = (
            PatientDeposit.objects.filter(
                patient_id__in=Subquery(qs.values('patient_id').distinct()),
                is_deleted=False,
                status='active',
            ).aggregate(s=Sum('available_balance'))['s']
            or Decimal('0')
        )
    except Exception:
        pass

    payer_case = Case(
        When(Q(payer__isnull=True) | Q(payer__payer_type='cash'), then=Value('cash')),
        When(payer__payer_type='corporate', then=Value('corporate')),
        When(payer__payer_type__in=['private', 'nhis', 'insurance'], then=Value('insurance')),
        default=Value('other'),
        output_field=CharField(),
    )
    payer_labels = {
        'cash': 'Cash',
        'insurance': 'Insurance / NHIS',
        'corporate': 'Corporate',
        'other': 'Other',
    }
    prow = list(
        qs.annotate(_pc=payer_case)
        .values('_pc')
        .annotate(c=Count('id'), amt=Sum('total_amount'))
        .order_by('-amt')
    )

    srows = list(
        qs.values('status').annotate(c=Count('id'), amt=Sum('total_amount')).order_by('-c')
    )

    daily_rows = list(
        qs.annotate(d=TruncDate('issued_at'))
        .values('d')
        .annotate(amt=Sum('total_amount'), c=Count('id'))
        .order_by('-d')[:14]
    )
    daily_rows.reverse()

    out['summary'].update(
        {
            'patient_count': patients,
            'invoice_or_bill_count': cnt,
            'total_billed': _fdec(agg['st']),
            'total_balance': _fdec(agg['sb']),
            'total_paid': _fdec(agg['sp']),
            'deposits_available': _fdec(deposits),
            'negative_balance_invoices': neg,
            'insurance_covered': 0.0,
            'patient_portion_sum': _fdec(agg['sb']),
        }
    )
    out['charts']['status'] = {
        'labels': [_STATUS_DISPLAY.get(r['status'], r['status'] or '') for r in srows],
        'counts': [r['c'] for r in srows],
        'amounts': [_fdec(r['amt']) for r in srows],
    }
    out['charts']['payer_type'] = {
        'labels': [payer_labels.get(r['_pc'], r['_pc'] or 'Other') for r in prow],
        'counts': [r['c'] for r in prow],
        'amounts': [_fdec(r['amt']) for r in prow],
    }
    out['charts']['daily'] = {
        'labels': [
            r['d'].strftime('%b %d') if hasattr(r['d'], 'strftime') else str(r['d'] or '')
            for r in daily_rows
        ],
        'amounts': [_fdec(r['amt']) for r in daily_rows],
        'counts': [r['c'] for r in daily_rows],
    }
    return out


def analytics_for_bill_queryset(qs):
    """Aggregates when workflow Bill records are used."""
    out = _empty_bills_analytics('bill')
    cnt = qs.count()
    if cnt == 0:
        return out

    agg = qs.aggregate(
        st=Sum('total_amount'),
        si=Sum('insurance_covered'),
        sp=Sum('patient_portion'),
    )
    patients = qs.aggregate(n=Count('patient_id', distinct=True))['n'] or 0

    srows = list(
        qs.values('status').annotate(c=Count('id'), amt=Sum('total_amount')).order_by('-c')
    )
    trows = list(
        qs.values('bill_type').annotate(c=Count('id'), amt=Sum('total_amount')).order_by('-c')
    )
    type_labels = dict(Bill.BILL_TYPES)

    daily_rows = list(
        qs.annotate(d=TruncDate('issued_at'))
        .values('d')
        .annotate(amt=Sum('total_amount'), c=Count('id'))
        .order_by('-d')[:14]
    )
    daily_rows.reverse()

    out['summary'].update(
        {
            'patient_count': patients,
            'invoice_or_bill_count': cnt,
            'total_billed': _fdec(agg['st']),
            'total_balance': _fdec(agg['sp']),
            'total_paid': None,
            'deposits_available': 0.0,
            'negative_balance_invoices': 0,
            'insurance_covered': _fdec(agg['si']),
            'patient_portion_sum': _fdec(agg['sp']),
        }
    )
    out['charts']['status'] = {
        'labels': [_STATUS_DISPLAY.get(r['status'], r['status'] or '') for r in srows],
        'counts': [r['c'] for r in srows],
        'amounts': [_fdec(r['amt']) for r in srows],
    }
    out['charts']['bill_type'] = {
        'labels': [type_labels.get(r['bill_type'], r['bill_type'] or '') for r in trows],
        'counts': [r['c'] for r in trows],
        'amounts': [_fdec(r['amt']) for r in trows],
    }
    out['charts']['daily'] = {
        'labels': [
            r['d'].strftime('%b %d') if hasattr(r['d'], 'strftime') else str(r['d'] or '')
            for r in daily_rows
        ],
        'amounts': [_fdec(r['amt']) for r in daily_rows],
        'counts': [r['c'] for r in daily_rows],
    }
    return out


def finalize_bills_list_analytics(qs, *, mode, grouped_by_patient, view_mode, showing_invoices_fallback):
    """Build analytics payload for bills_list.html charts (template: json_script filter)."""
    if mode == 'invoice':
        payload = analytics_for_invoice_queryset(qs)
    else:
        payload = analytics_for_bill_queryset(qs)
    payload['meta'] = {
        'grouped_by_patient': grouped_by_patient,
        'view_mode': view_mode,
        'showing_invoices_fallback': showing_invoices_fallback,
    }
    return payload


def _invoice_as_bill_adapter(inv, deposit_available=None, channels_by_invoice=None):
    """Adapt Invoice to bill-like object for template display."""
    inv.update_totals()
    bill_type, type_label = _invoice_bill_type_display(inv)
    if inv.balance is not None:
        bal = inv.balance
    else:
        bal = inv.total_amount if inv.total_amount is not None else Decimal('0')
    tot = inv.total_amount or Decimal('0')
    credit_overpay = abs(bal) if bal < 0 else Decimal('0')
    dep = deposit_available if deposit_available is not None else Decimal('0')
    payer_display = None
    if _invoice_is_cash_bill(inv):
        ch = (channels_by_invoice or {}).get(inv.id, '') if channels_by_invoice is not None else ''
        payer_display = ch.strip() if ch else 'Cash'
    obj = SimpleNamespace(
        bill_number=inv.invoice_number,
        patient=inv.patient,
        bill_type=bill_type,
        total_amount=tot,
        insurance_covered=Decimal('0'),
        patient_portion=bal,
        status=inv.status,
        issued_at=inv.issued_at,
        invoice=inv,
        is_invoice_fallback=True,
        amount_paid=inv.amount_paid,
        deposit_available=dep,
        credit_overpay=credit_overpay,
        payer_display=payer_display,
    )
    obj.get_status_display = lambda val=_STATUS_DISPLAY.get(inv.status, inv.status): val
    obj.get_bill_type_display = lambda: type_label
    return obj


@login_required
@role_required('accountant', 'senior_account_officer')
def bills_list(request):
    """List all bills with filtering. Falls back to Invoices when no Bills exist."""
    # Redirect to clean URL when any param is INVALID so form does not show it
    invalid_keys = [k for k, v in request.GET.items() if (v or '').strip().lower() == 'invalid']
    if invalid_keys:
        cleaned = {k: v for k, v in request.GET.items() if k not in invalid_keys}
        target = request.path + ('?' + urlencode(cleaned) if cleaned else '')
        return redirect(target)
    
    bills = Bill.objects.filter(is_deleted=False).select_related(
        'patient', 'issued_by', 'invoice', 'invoice__payer'
    ).order_by('-issued_at')
    
    # All insurance payers for accountant filter (private, nhis, insurance)
    insurance_payers = Payer.objects.filter(
        payer_type__in=['private', 'nhis', 'insurance'],
        is_active=True,
        is_deleted=False
    ).order_by('name')
    # Corporate payers for when bill type = Corporate
    corporate_payers = Payer.objects.filter(
        payer_type='corporate',
        is_active=True,
        is_deleted=False
    ).order_by('name')

    # Filters (treat INVALID as empty so nothing is filtered by it)
    status_filter = _normalize_filter_value(request.GET.get('status', ''))
    bill_type_filter = _normalize_filter_value(request.GET.get('bill_type', ''))
    payer_filter = _normalize_filter_value(request.GET.get('payer', ''))
    date_from = _normalize_filter_value(request.GET.get('date_from', ''))
    date_to = _normalize_filter_value(request.GET.get('date_to', ''))
    search = _normalize_filter_value(request.GET.get('search', ''))
    
    if status_filter:
        bills = bills.filter(status=status_filter)
    if bill_type_filter:
        bills = bills.filter(bill_type=bill_type_filter)
    if payer_filter:
        bills = bills.filter(invoice__payer_id=payer_filter)
    if date_from:
        bills = bills.filter(issued_at__date__gte=date_from)
    if date_to:
        bills = bills.filter(issued_at__date__lte=date_to)
    if search:
        bills = bills.filter(Q(bill_number__icontains=search) | _patient_search_q(search))
    
    bills_count = bills.count()
    bills_exist_in_system = Bill.objects.filter(is_deleted=False).exists()
    view_mode = _normalize_filter_value(request.GET.get('view', ''))
    if view_mode not in ('patient', 'invoice'):
        view_mode = 'patient'  # Default: group by patient
    
    # When no Bills exist in system at all, show Invoices as fallback
    showing_invoices_fallback = False
    grouped_by_patient = False
    grouped_invoice_bills = False
    if bills_count == 0 and not bills_exist_in_system:
        # Include invoices with total>0 OR at least one unwaived line (so drugs, ECG, etc. all show)
        has_unwaived_line = Exists(
            InvoiceLine.objects.filter(
                invoice_id=OuterRef('pk'),
                is_deleted=False,
                waived_at__isnull=True
            )
        )
        invoices = Invoice.objects.filter(
            is_deleted=False,
            status__in=['draft', 'issued', 'partially_paid', 'overdue', 'paid'],
        ).filter(Q(total_amount__gt=0) | has_unwaived_line).select_related('patient', 'payer').order_by('-issued_at')

        if status_filter:
            invoices = invoices.filter(status=status_filter)
        if bill_type_filter:
            invoices = invoices.filter(_invoice_bill_type_q(bill_type_filter))
        if payer_filter:
            invoices = invoices.filter(payer_id=payer_filter)
        if date_from:
            invoices = invoices.filter(issued_at__date__gte=date_from)
        if date_to:
            invoices = invoices.filter(issued_at__date__lte=date_to)
        if search:
            invoices = invoices.filter(_invoice_search_q(search))
        
        if view_mode == 'patient':
            # Group by patient: combine payers per patient (filtered by bill_type when set).
            # Include invoices with total>0 OR at least one unwaived line so all items (drugs, ECG, etc.) show.
            has_unwaived_line_patient = Exists(
                InvoiceLine.objects.filter(
                    invoice_id=OuterRef('pk'),
                    is_deleted=False,
                    waived_at__isnull=True
                )
            )
            invoices_all_payers = Invoice.objects.filter(
                is_deleted=False,
                patient__isnull=False,
                status__in=['draft', 'issued', 'partially_paid', 'overdue', 'paid'],
            ).filter(Q(total_amount__gt=0) | has_unwaived_line_patient).select_related('patient', 'payer')
            if status_filter:
                invoices_all_payers = invoices_all_payers.filter(status=status_filter)
            if bill_type_filter:
                invoices_all_payers = invoices_all_payers.filter(_invoice_bill_type_q(bill_type_filter))
            if date_from:
                invoices_all_payers = invoices_all_payers.filter(issued_at__date__gte=date_from)
            if date_to:
                invoices_all_payers = invoices_all_payers.filter(issued_at__date__lte=date_to)
            if search:
                invoices_all_payers = invoices_all_payers.filter(_invoice_search_q(search))
            patient_groups = (
                invoices_all_payers.values('patient_id', 'patient__first_name', 'patient__last_name', 'patient__mrn')
                .annotate(
                    invoice_count=Count('id'),
                    total_amount=Sum('total_amount'),
                    total_balance=Sum('balance'),
                    latest_issued=Max('issued_at'),
                )
                .order_by('-latest_issued')
            )
            # If payer filter set, only show patients that have at least one invoice of that payer
            if payer_filter:
                patient_ids_with_payer = set(invoices.values_list('patient_id', flat=True).distinct())
            else:
                patient_ids_with_payer = None

            patient_group_rows = list(patient_groups)
            pending_groups = []
            for row in patient_group_rows:
                patient_id = row['patient_id']
                if not patient_id:
                    continue
                if patient_ids_with_payer is not None and patient_id not in patient_ids_with_payer:
                    continue
                pending_groups.append((patient_id, row))

            batch_pids = [p[0] for p in pending_groups]
            payer_by_patient = _batch_invoice_payer_names(invoices_all_payers, batch_pids)
            channels_by_patient = _batch_payment_channel_labels_by_patient(
                invoices_all_payers, batch_pids
            )
            paid_by_patient = _batch_invoice_amount_paid(invoices_all_payers, batch_pids)
            deposit_by_patient = _batch_patient_deposit_available(batch_pids)

            grouped_rows = []
            for patient_id, row in pending_groups:
                total_amt = row['total_amount'] or Decimal('0')
                total_bal = row['total_balance']
                if total_bal is None:
                    total_bal = total_amt
                status = 'paid' if total_bal <= 0 else (
                    'partially_paid' if total_bal < total_amt else 'issued'
                )
                patient = Patient.objects.get(pk=patient_id)
                names = payer_by_patient.get(patient_id) or set()
                ch = channels_by_patient.get(patient_id, '')
                payer_display = _compose_payer_column_for_patient(names, ch)
                amount_paid_row = paid_by_patient.get(patient_id, Decimal('0'))
                deposit_avail = deposit_by_patient.get(patient_id, Decimal('0'))
                credit_overpay = abs(total_bal) if total_bal < 0 else Decimal('0')
                if bill_type_filter == 'insurance':
                    row_bill_type, row_type_label = 'insurance', 'Insurance Bill'
                elif bill_type_filter == 'cash':
                    row_bill_type, row_type_label = 'cash', 'Cash Bill'
                elif bill_type_filter == 'corporate':
                    row_bill_type, row_type_label = 'cash', 'Corporate'
                elif bill_type_filter == 'mixed':
                    row_bill_type, row_type_label = 'mixed', 'Mixed (Cash + Insurance)'
                else:
                    row_bill_type, row_type_label = 'mixed', 'Combined'
                obj = SimpleNamespace(
                    bill_number=f"{row['invoice_count']} invoice(s)",
                    patient=patient,
                    patient_id=patient_id,
                    bill_type=row_bill_type,
                    total_amount=total_amt,
                    insurance_covered=Decimal('0'),
                    patient_portion=total_bal,
                    status=status,
                    issued_at=row['latest_issued'],
                    invoice_count=row['invoice_count'],
                    invoice=None,
                    payer_display=payer_display,
                    amount_paid=amount_paid_row,
                    deposit_available=deposit_avail,
                    credit_overpay=credit_overpay,
                    is_invoice_fallback=True,
                    is_grouped=True,
                )
                obj.get_status_display = lambda val=_STATUS_DISPLAY.get(status, status): val
                obj.get_bill_type_display = lambda: row_type_label
                grouped_rows.append(obj)
            
            grouped_by_patient = True
            grp_paginator = Paginator(grouped_rows, 50)
            bills_page = grp_paginator.get_page(request.GET.get('page'))
        elif view_mode == 'invoice' and bill_type_filter in ('cash', 'insurance', 'corporate'):
            inv_list = list(invoices.select_related('patient', 'payer'))
            groups_map = defaultdict(list)
            labels = {}
            if bill_type_filter == 'cash':
                for inv in inv_list:
                    k = inv.patient_id or 'nopatient'
                    if k not in labels:
                        if inv.patient:
                            labels[k] = (getattr(inv.patient, 'full_name', None) or '—')
                        else:
                            labels[k] = '—'
                    groups_map[k].append(inv)
            else:
                for inv in inv_list:
                    lk = inv.payer_id if inv.payer_id is not None else 'nopayer'
                    if lk not in labels:
                        labels[lk] = (inv.payer.name or '—') if inv.payer else '—'
                    groups_map[lk].append(inv)

            group_rows = []
            for k, invs in groups_map.items():
                latest = max((i.issued_at for i in invs if i.issued_at), default=None)
                tot = sum((i.total_amount or Decimal('0')) for i in invs)
                total_bal = sum(
                    (
                        i.balance
                        if i.balance is not None
                        else (i.total_amount or Decimal('0'))
                    )
                    for i in invs
                )
                total_paid = sum(
                    (i.total_amount or Decimal('0'))
                    - (i.balance if i.balance is not None else (i.total_amount or Decimal('0')))
                    for i in invs
                )
                group_rows.append(
                    {
                        'label': labels[k],
                        'group_key': k,
                        'invoices': invs,
                        'latest': latest,
                        'total_amount': tot,
                        'total_balance': total_bal,
                        'total_paid': total_paid,
                        'invoice_count': len(invs),
                    }
                )
            group_rows.sort(key=lambda g: (g['latest'] is None, g['latest']), reverse=True)
            grp_paginator = Paginator(group_rows, 50)
            gpage = grp_paginator.get_page(request.GET.get('page'))
            page_groups = []
            for g in gpage.object_list:
                pids = list({inv.patient_id for inv in g['invoices'] if inv.patient_id})
                ch_map = _batch_payment_channel_labels_by_invoice([inv.id for inv in g['invoices']])
                dep_map = _batch_patient_deposit_available(pids)
                invs_sorted = sorted(
                    g['invoices'],
                    key=lambda x: x.issued_at or timezone.now(),
                    reverse=True,
                )
                adapted_invoices = [
                    _invoice_as_bill_adapter(
                        inv,
                        dep_map.get(inv.patient_id, Decimal('0')) if inv.patient_id else Decimal('0'),
                        ch_map,
                    )
                    for inv in invs_sorted
                ]
                page_groups.append({**g, 'adapted_invoices': adapted_invoices})
            bills_page = Page(page_groups, gpage.number, grp_paginator)
            grouped_invoice_bills = True
        else:
            # By Invoice: one row per invoice
            inv_paginator = Paginator(invoices, 50)
            inv_page = inv_paginator.get_page(request.GET.get('page'))
            dep_map = _batch_patient_deposit_available(list({inv.patient_id for inv in inv_page}))
            ch_map = _batch_payment_channel_labels_by_invoice([inv.id for inv in inv_page])
            adapted_list = [
                _invoice_as_bill_adapter(
                    inv, dep_map.get(inv.patient_id, Decimal('0')), ch_map
                )
                for inv in inv_page
            ]
            bills_page = Page(adapted_list, inv_page.number, inv_paginator)
        
        showing_invoices_fallback = True
    else:
        paginator = Paginator(bills, 50)
        page = request.GET.get('page')
        bills_page = paginator.get_page(page)
        inv_ids_page = [b.invoice_id for b in bills_page if getattr(b, 'invoice_id', None)]
        ch_by_inv_page = _batch_payment_channel_labels_by_invoice(inv_ids_page)
        for b in bills_page:
            setattr(b, 'amount_paid', None)
            setattr(b, 'deposit_available', None)
            setattr(b, 'credit_overpay', None)
            inv = getattr(b, 'invoice', None)
            if inv and _invoice_is_cash_bill(inv):
                ch = ch_by_inv_page.get(inv.id, '')
                setattr(b, 'payer_display', ch.strip() if ch else 'Cash')
            else:
                setattr(b, 'payer_display', None)

    ba = finalize_bills_list_analytics(
        invoices if showing_invoices_fallback else bills,
        mode='invoice' if showing_invoices_fallback else 'bill',
        grouped_by_patient=grouped_by_patient,
        view_mode=view_mode if showing_invoices_fallback else None,
        showing_invoices_fallback=showing_invoices_fallback,
    )

    context = {
        'bills': bills_page,
        'showing_invoices_fallback': showing_invoices_fallback,
        'grouped_by_patient': grouped_by_patient,
        'grouped_invoice_bills': grouped_invoice_bills,
        'view_mode': view_mode,
        'status_filter': status_filter,
        'bill_type_filter': bill_type_filter,
        'payer_filter': payer_filter,
        'insurance_payers': insurance_payers,
        'corporate_payers': corporate_payers,
        'date_from': date_from,
        'date_to': date_to,
        'search': search,
        'bills_analytics': ba,
    }

    return render(request, 'hospital/billing/bills_list.html', context)


@login_required
@role_required('accountant', 'senior_account_officer')
def bills_by_normal_invoice(request):
    """Bills list (same as All Bills) - title indicates invoice-linked view; show all bills or invoices so page is never empty."""
    # Redirect to clean URL when any param is INVALID so form does not show it
    invalid_keys = [k for k, v in request.GET.items() if (v or '').strip().lower() == 'invalid']
    if invalid_keys:
        cleaned = {k: v for k, v in request.GET.items() if k not in invalid_keys}
        target = request.path + ('?' + urlencode(cleaned) if cleaned else '')
        return redirect(target)
    # Use same queryset as bills_list so bills always show when they exist
    bills = Bill.objects.filter(is_deleted=False).select_related(
        'patient', 'invoice', 'invoice__payer', 'issued_by'
    ).order_by('-issued_at')

    # All insurance and corporate payers for filter dropdowns
    insurance_payers = Payer.objects.filter(
        payer_type__in=['private', 'nhis', 'insurance'],
        is_active=True,
        is_deleted=False
    ).order_by('name')
    corporate_payers = Payer.objects.filter(
        payer_type='corporate',
        is_active=True,
        is_deleted=False
    ).order_by('name')

    status_filter = _normalize_filter_value(request.GET.get('status', ''))
    bill_type_filter = _normalize_filter_value(request.GET.get('bill_type', ''))
    payer_filter = _normalize_filter_value(request.GET.get('payer', ''))
    date_from = _normalize_filter_value(request.GET.get('date_from', ''))
    date_to = _normalize_filter_value(request.GET.get('date_to', ''))
    search = _normalize_filter_value(request.GET.get('search', ''))

    if status_filter:
        bills = bills.filter(status=status_filter)
    if bill_type_filter:
        bills = bills.filter(bill_type=bill_type_filter)
    if payer_filter:
        bills = bills.filter(invoice__payer_id=payer_filter)
    if date_from:
        bills = bills.filter(issued_at__date__gte=date_from)
    if date_to:
        bills = bills.filter(issued_at__date__lte=date_to)
    if search:
        bills = bills.filter(Q(bill_number__icontains=search) | _patient_search_q(search))

    bills_count = bills.count()
    bills_exist_in_system = Bill.objects.filter(is_deleted=False).exists()

    # When no Bill records exist, show Invoices as fallback; support view=patient (combined) or view=invoice
    if bills_count == 0 and not bills_exist_in_system:
        view_mode = _normalize_filter_value(request.GET.get('view', ''))
        if view_mode not in ('patient', 'invoice'):
            view_mode = 'invoice'  # Default on this page: by invoice

        has_unwaived_line = Exists(
            InvoiceLine.objects.filter(
                invoice_id=OuterRef('pk'),
                is_deleted=False,
                waived_at__isnull=True
            )
        )
        invoices = Invoice.objects.filter(
            is_deleted=False,
            status__in=['draft', 'issued', 'partially_paid', 'overdue', 'paid'],
        ).filter(Q(total_amount__gt=0) | has_unwaived_line).select_related('patient', 'payer').order_by('-issued_at')
        if status_filter:
            invoices = invoices.filter(status=status_filter)
        if bill_type_filter:
            invoices = invoices.filter(_invoice_bill_type_q(bill_type_filter))
        if payer_filter:
            invoices = invoices.filter(payer_id=payer_filter)
        if date_from:
            invoices = invoices.filter(issued_at__date__gte=date_from)
        if date_to:
            invoices = invoices.filter(issued_at__date__lte=date_to)
        if search:
            invoices = invoices.filter(_invoice_search_q(search))

        if view_mode == 'patient':
            # By Patient (Combined): one row per patient, all payers combined
            has_unwaived_line_patient = Exists(
                InvoiceLine.objects.filter(
                    invoice_id=OuterRef('pk'),
                    is_deleted=False,
                    waived_at__isnull=True
                )
            )
            invoices_all_payers = Invoice.objects.filter(
                is_deleted=False,
                patient__isnull=False,
                status__in=['draft', 'issued', 'partially_paid', 'overdue', 'paid'],
            ).filter(Q(total_amount__gt=0) | has_unwaived_line_patient).select_related('patient', 'payer')
            if status_filter:
                invoices_all_payers = invoices_all_payers.filter(status=status_filter)
            if bill_type_filter:
                invoices_all_payers = invoices_all_payers.filter(_invoice_bill_type_q(bill_type_filter))
            if date_from:
                invoices_all_payers = invoices_all_payers.filter(issued_at__date__gte=date_from)
            if date_to:
                invoices_all_payers = invoices_all_payers.filter(issued_at__date__lte=date_to)
            if search:
                invoices_all_payers = invoices_all_payers.filter(_invoice_search_q(search))
            patient_groups = (
                invoices_all_payers.values('patient_id', 'patient__first_name', 'patient__last_name', 'patient__mrn')
                .annotate(
                    invoice_count=Count('id'),
                    total_amount=Sum('total_amount'),
                    total_balance=Sum('balance'),
                    latest_issued=Max('issued_at'),
                )
                .order_by('-latest_issued')
            )
            if payer_filter:
                patient_ids_with_payer = set(invoices.values_list('patient_id', flat=True).distinct())
            else:
                patient_ids_with_payer = None

            patient_group_rows = list(patient_groups)
            pending_groups = []
            for row in patient_group_rows:
                patient_id = row['patient_id']
                if not patient_id:
                    continue
                if patient_ids_with_payer is not None and patient_id not in patient_ids_with_payer:
                    continue
                pending_groups.append((patient_id, row))

            batch_pids = [p[0] for p in pending_groups]
            payer_by_patient = _batch_invoice_payer_names(invoices_all_payers, batch_pids)
            channels_by_patient = _batch_payment_channel_labels_by_patient(
                invoices_all_payers, batch_pids
            )
            paid_by_patient = _batch_invoice_amount_paid(invoices_all_payers, batch_pids)
            deposit_by_patient = _batch_patient_deposit_available(batch_pids)

            grouped_rows = []
            for patient_id, row in pending_groups:
                total_amt = row['total_amount'] or Decimal('0')
                total_bal = row['total_balance']
                if total_bal is None:
                    total_bal = total_amt
                status = 'paid' if total_bal <= 0 else (
                    'partially_paid' if total_bal < total_amt else 'issued'
                )
                patient = Patient.objects.get(pk=patient_id)
                names = payer_by_patient.get(patient_id) or set()
                ch = channels_by_patient.get(patient_id, '')
                payer_display = _compose_payer_column_for_patient(names, ch)
                amount_paid_row = paid_by_patient.get(patient_id, Decimal('0'))
                deposit_avail = deposit_by_patient.get(patient_id, Decimal('0'))
                credit_overpay = abs(total_bal) if total_bal < 0 else Decimal('0')
                if bill_type_filter == 'insurance':
                    row_bill_type, row_type_label = 'insurance', 'Insurance Bill'
                elif bill_type_filter == 'cash':
                    row_bill_type, row_type_label = 'cash', 'Cash Bill'
                elif bill_type_filter == 'corporate':
                    row_bill_type, row_type_label = 'cash', 'Corporate'
                elif bill_type_filter == 'mixed':
                    row_bill_type, row_type_label = 'mixed', 'Mixed (Cash + Insurance)'
                else:
                    row_bill_type, row_type_label = 'mixed', 'Combined'
                obj = SimpleNamespace(
                    bill_number=f"{row['invoice_count']} invoice(s)",
                    patient=patient,
                    patient_id=patient_id,
                    bill_type=row_bill_type,
                    total_amount=total_amt,
                    insurance_covered=Decimal('0'),
                    patient_portion=total_bal,
                    status=status,
                    issued_at=row['latest_issued'],
                    invoice_count=row['invoice_count'],
                    invoice=None,
                    payer_display=payer_display,
                    amount_paid=amount_paid_row,
                    deposit_available=deposit_avail,
                    credit_overpay=credit_overpay,
                    is_invoice_fallback=True,
                    is_grouped=True,
                )
                obj.get_status_display = lambda val=_STATUS_DISPLAY.get(status, status): val
                obj.get_bill_type_display = lambda: row_type_label
                grouped_rows.append(obj)
            grp_paginator = Paginator(grouped_rows, 50)
            bills_page = grp_paginator.get_page(request.GET.get('page'))
            ba = finalize_bills_list_analytics(
                invoices,
                mode='invoice',
                grouped_by_patient=True,
                view_mode='patient',
                showing_invoices_fallback=True,
            )
            context = {
                'bills': bills_page,
                'title': 'Bills by Normal Invoice',
                'search': search,
                'status_filter': status_filter,
                'bill_type_filter': bill_type_filter,
                'payer_filter': payer_filter,
                'insurance_payers': insurance_payers,
                'corporate_payers': corporate_payers,
                'date_from': date_from,
                'date_to': date_to,
                'showing_invoices_fallback': True,
                'grouped_by_patient': True,
                'grouped_invoice_bills': False,
                'view_mode': 'patient',
                'bills_analytics': ba,
            }
            return render(request, 'hospital/billing/bills_list.html', context)

        # view_mode == 'invoice': one row per invoice, or grouped when type is cash/insurance/corporate
        grouped_invoice_bills = False
        if bill_type_filter in ('cash', 'insurance', 'corporate'):
            inv_list = list(invoices.select_related('patient', 'payer'))
            groups_map = defaultdict(list)
            labels = {}
            if bill_type_filter == 'cash':
                for inv in inv_list:
                    k = inv.patient_id or 'nopatient'
                    if k not in labels:
                        if inv.patient:
                            labels[k] = (getattr(inv.patient, 'full_name', None) or '—')
                        else:
                            labels[k] = '—'
                    groups_map[k].append(inv)
            else:
                for inv in inv_list:
                    lk = inv.payer_id if inv.payer_id is not None else 'nopayer'
                    if lk not in labels:
                        labels[lk] = (inv.payer.name or '—') if inv.payer else '—'
                    groups_map[lk].append(inv)
            group_rows = []
            for k, invs in groups_map.items():
                latest = max((i.issued_at for i in invs if i.issued_at), default=None)
                tot = sum((i.total_amount or Decimal('0')) for i in invs)
                total_bal = sum(
                    (
                        i.balance
                        if i.balance is not None
                        else (i.total_amount or Decimal('0'))
                    )
                    for i in invs
                )
                total_paid = sum(
                    (i.total_amount or Decimal('0'))
                    - (i.balance if i.balance is not None else (i.total_amount or Decimal('0')))
                    for i in invs
                )
                group_rows.append(
                    {
                        'label': labels[k],
                        'group_key': k,
                        'invoices': invs,
                        'latest': latest,
                        'total_amount': tot,
                        'total_balance': total_bal,
                        'total_paid': total_paid,
                        'invoice_count': len(invs),
                    }
                )
            group_rows.sort(key=lambda g: (g['latest'] is None, g['latest']), reverse=True)
            grp_paginator = Paginator(group_rows, 50)
            gpage = grp_paginator.get_page(request.GET.get('page'))
            page_groups = []
            for g in gpage.object_list:
                pids = list({inv.patient_id for inv in g['invoices'] if inv.patient_id})
                ch_map = _batch_payment_channel_labels_by_invoice([inv.id for inv in g['invoices']])
                dep_map = _batch_patient_deposit_available(pids)
                invs_sorted = sorted(
                    g['invoices'],
                    key=lambda x: x.issued_at or timezone.now(),
                    reverse=True,
                )
                adapted_invoices = [
                    _invoice_as_bill_adapter(
                        inv,
                        dep_map.get(inv.patient_id, Decimal('0')) if inv.patient_id else Decimal('0'),
                        ch_map,
                    )
                    for inv in invs_sorted
                ]
                page_groups.append({**g, 'adapted_invoices': adapted_invoices})
            bills_page = Page(page_groups, gpage.number, grp_paginator)
            grouped_invoice_bills = True
        else:
            inv_paginator = Paginator(invoices, 50)
            inv_page = inv_paginator.get_page(request.GET.get('page'))
            dep_map = _batch_patient_deposit_available(list({inv.patient_id for inv in inv_page}))
            ch_map = _batch_payment_channel_labels_by_invoice([inv.id for inv in inv_page])
            adapted_list = [
                _invoice_as_bill_adapter(
                    inv, dep_map.get(inv.patient_id, Decimal('0')), ch_map
                )
                for inv in inv_page
            ]
            bills_page = Page(adapted_list, inv_page.number, inv_paginator)
        ba = finalize_bills_list_analytics(
            invoices,
            mode='invoice',
            grouped_by_patient=False,
            view_mode='invoice',
            showing_invoices_fallback=True,
        )
        context = {
            'bills': bills_page,
            'title': 'Bills by Normal Invoice',
            'search': search,
            'status_filter': status_filter,
            'bill_type_filter': bill_type_filter,
            'payer_filter': payer_filter,
            'insurance_payers': insurance_payers,
            'corporate_payers': corporate_payers,
            'date_from': date_from,
            'date_to': date_to,
            'showing_invoices_fallback': True,
            'grouped_by_patient': False,
            'grouped_invoice_bills': grouped_invoice_bills,
            'view_mode': 'invoice',
            'bills_analytics': ba,
        }
        return render(request, 'hospital/billing/bills_list.html', context)

    paginator = Paginator(bills, 50)
    page = request.GET.get('page')
    bills_page = paginator.get_page(page)
    inv_ids_page = [b.invoice_id for b in bills_page if getattr(b, 'invoice_id', None)]
    ch_by_inv_page = _batch_payment_channel_labels_by_invoice(inv_ids_page)
    for b in bills_page:
        setattr(b, 'amount_paid', None)
        setattr(b, 'deposit_available', None)
        setattr(b, 'credit_overpay', None)
        inv = getattr(b, 'invoice', None)
        if inv and _invoice_is_cash_bill(inv):
            ch = ch_by_inv_page.get(inv.id, '')
            setattr(b, 'payer_display', ch.strip() if ch else 'Cash')
        else:
            setattr(b, 'payer_display', None)

    ba = finalize_bills_list_analytics(
        bills,
        mode='bill',
        grouped_by_patient=False,
        view_mode=None,
        showing_invoices_fallback=False,
    )
    context = {
        'bills': bills_page,
        'title': 'Bills by Normal Invoice',
        'search': search,
        'status_filter': status_filter,
        'bill_type_filter': bill_type_filter,
        'payer_filter': payer_filter,
        'insurance_payers': insurance_payers,
        'corporate_payers': corporate_payers,
        'date_from': date_from,
        'date_to': date_to,
        'showing_invoices_fallback': False,
        'grouped_by_patient': False,
        'grouped_invoice_bills': False,
        'view_mode': None,
        'bills_analytics': ba,
    }
    return render(request, 'hospital/billing/bills_list.html', context)


@login_required
@role_required('accountant', 'senior_account_officer')
def bills_by_invoice_group2_cover_page(request):
    """Bills grouped by invoice with cover page format. Uses all_objects so write-off period invoices show."""
    invalid_keys = [k for k, v in request.GET.items() if (v or '').strip().lower() == 'invalid']
    if invalid_keys:
        cleaned = {k: v for k, v in request.GET.items() if k not in invalid_keys}
        target = request.path + ('?' + urlencode(cleaned) if cleaned else '')
        return redirect(target)

    insurance_payers = Payer.objects.filter(
        payer_type__in=['private', 'nhis', 'insurance'],
        is_active=True,
        is_deleted=False,
    ).order_by('name')
    corporate_payers = Payer.objects.filter(
        payer_type='corporate',
        is_active=True,
        is_deleted=False,
    ).order_by('name')

    status_filter = _normalize_filter_value(request.GET.get('status', ''))
    bill_type_filter = _normalize_filter_value(request.GET.get('bill_type', ''))
    payer_filter = _normalize_filter_value(request.GET.get('payer', ''))
    date_from = _normalize_filter_value(request.GET.get('date_from', ''))
    date_to = _normalize_filter_value(request.GET.get('date_to', ''))
    search = _normalize_filter_value(request.GET.get('search', ''))

    # List (almost) all invoices: do not require linked Bills or Lines — some DBs/environments
    # had zero rows with the Count filter; draft/empty invoices are still valid rows to show.
    bills_prefetch = Prefetch(
        'bills',
        queryset=Bill.objects.filter(is_deleted=False).select_related('issued_by', 'encounter').order_by('-issued_at'),
    )

    invoices_with_bills = (
        Invoice.all_objects.filter(is_deleted=False)
        .exclude(status='cancelled')
        .annotate(
            active_bill_count=Count('bills', filter=Q(bills__is_deleted=False)),
            active_line_count=Count('lines', filter=Q(lines__is_deleted=False)),
        )
        .select_related('patient', 'payer', 'encounter')
        .prefetch_related(bills_prefetch)
        .order_by('-issued_at')
    )

    if status_filter:
        invoices_with_bills = invoices_with_bills.filter(status=status_filter)
    if search:
        invoices_with_bills = invoices_with_bills.filter(_invoice_search_q(search))
    if date_from:
        invoices_with_bills = invoices_with_bills.filter(issued_at__date__gte=date_from)
    if date_to:
        invoices_with_bills = invoices_with_bills.filter(issued_at__date__lte=date_to)
    if payer_filter:
        invoices_with_bills = invoices_with_bills.filter(payer_id=payer_filter)
    if bill_type_filter == 'mixed':
        invoices_with_bills = invoices_with_bills.filter(
            Exists(
                Bill.objects.filter(
                    invoice_id=OuterRef('pk'),
                    is_deleted=False,
                    bill_type='mixed',
                )
            )
        )
    elif bill_type_filter:
        invoices_with_bills = invoices_with_bills.filter(_invoice_bill_type_q(bill_type_filter))

    paginator = Paginator(invoices_with_bills, 50)
    invoices_page = paginator.get_page(request.GET.get('page'))

    for inv in invoices_page:
        try:
            inv.update_totals()
        except Exception:
            logger.exception('bills_by_invoice_group: update_totals failed for invoice %s', inv.pk)

    filter_qs = request.GET.copy()
    filter_qs.pop('page', None)
    filter_querystring = filter_qs.urlencode()

    bills_without_invoice = Bill.objects.filter(
        invoice__isnull=True,
        is_deleted=False,
    ).select_related('patient', 'encounter', 'issued_by')
    if status_filter:
        bills_without_invoice = bills_without_invoice.filter(status=status_filter)
    if bill_type_filter and bill_type_filter != 'corporate':
        bills_without_invoice = bills_without_invoice.filter(bill_type=bill_type_filter)
    if date_from:
        bills_without_invoice = bills_without_invoice.filter(issued_at__date__gte=date_from)
    if date_to:
        bills_without_invoice = bills_without_invoice.filter(issued_at__date__lte=date_to)
    if search:
        bills_without_invoice = bills_without_invoice.filter(
            Q(bill_number__icontains=search) | _patient_search_q(search)
        )
    bills_without_invoice = bills_without_invoice.order_by('-issued_at')[:100]

    context = {
        'invoices': invoices_page,
        'bills_without_invoice': bills_without_invoice,
        'title': 'Bills by Invoice Group (Cover Page Format)',
        'status_filter': status_filter,
        'bill_type_filter': bill_type_filter,
        'payer_filter': payer_filter,
        'insurance_payers': insurance_payers,
        'corporate_payers': corporate_payers,
        'date_from': date_from,
        'date_to': date_to,
        'search': search,
        'filter_querystring': filter_querystring,
    }

    return render(request, 'hospital/billing/bills_by_invoice_group.html', context)


@login_required
@role_required('accountant', 'senior_account_officer')
def bill_summary(request):
    """
    Bill summary report. Prefers invoice-backed aggregates when that is where volume lives:
    no workflow bills, or fewer Bill rows than qualifying invoices (avoids empty page when a
    few stray Bill records exist). Invoice path uses all_objects so write-off-archive invoices
    are included.
    """
    today = timezone.now().date()
    start_of_month = today.replace(day=1)
    if _use_invoice_for_billing_aggregates():
        base = _invoice_accountant_summary_queryset()
        monthly_qs = base.filter(issued_at__date__gte=start_of_month)
        monthly_summary = {
            'total_bills': monthly_qs.count(),
            'total_amount': monthly_qs.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00'),
            'paid_amount': monthly_qs.filter(status='paid').aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00'),
            'outstanding_amount': monthly_qs.filter(
                status__in=['issued', 'partially_paid', 'overdue']
            ).aggregate(total=Sum('balance'))['total'] or Decimal('0.00'),
        }
        summary_by_status = list(
            base.values('status')
            .annotate(count=Count('id'), total_amount=Sum('total_amount'))
            .order_by('status')
        )
        summary_by_type = list(
            base.annotate(bill_type=_invoice_summary_category_case())
            .values('bill_type')
            .annotate(count=Count('id'), total_amount=Sum('total_amount'))
            .order_by('bill_type')
        )
        context = {
            'summary_by_status': summary_by_status,
            'summary_by_type': summary_by_type,
            'monthly_summary': monthly_summary,
            'current_month': start_of_month,
            'showing_invoices_fallback': True,
            'summary_uses_workflow_bills': False,
        }
        return render(request, 'hospital/billing/bill_summary.html', context)

    monthly_bills = Bill.objects.filter(
        issued_at__date__gte=start_of_month,
        is_deleted=False,
    )
    monthly_summary = {
        'total_bills': monthly_bills.count(),
        'total_amount': monthly_bills.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00'),
        'paid_amount': monthly_bills.filter(status='paid').aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00'),
        'outstanding_amount': monthly_bills.filter(
            status__in=['issued', 'partially_paid']
        ).aggregate(total=Sum('patient_portion'))['total'] or Decimal('0.00'),
    }
    summary_by_status = list(
        Bill.objects.filter(is_deleted=False)
        .values('status')
        .annotate(count=Count('id'), total_amount=Sum('total_amount'))
        .order_by('status')
    )
    summary_by_type = list(
        Bill.objects.filter(is_deleted=False)
        .values('bill_type')
        .annotate(count=Count('id'), total_amount=Sum('total_amount'))
        .order_by('bill_type')
    )
    context = {
        'summary_by_status': summary_by_status,
        'summary_by_type': summary_by_type,
        'monthly_summary': monthly_summary,
        'current_month': start_of_month,
        'showing_invoices_fallback': False,
        'summary_uses_workflow_bills': True,
    }
    return render(request, 'hospital/billing/bill_summary.html', context)


@login_required
@role_required('accountant', 'senior_account_officer')
def bill_total(request):
    """Total bills report for a date range; mirrors bill_summary invoice vs workflow Bill choice."""
    today = timezone.now().date()
    date_from = _normalize_filter_value(request.GET.get('date_from', '')) or today.replace(day=1).isoformat()
    date_to = _normalize_filter_value(request.GET.get('date_to', '')) or today.isoformat()

    if _use_invoice_for_billing_aggregates():
        inv = _invoice_accountant_summary_queryset().filter(
            issued_at__date__gte=date_from,
            issued_at__date__lte=date_to,
        )
        ins_payer = Q(payer__payer_type__in=['private', 'nhis', 'insurance'])
        totals = {
            'total_bills': inv.count(),
            'total_amount': inv.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00'),
            'insurance_covered': inv.filter(ins_payer).aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00'),
            'patient_portion': inv.exclude(ins_payer).aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00'),
            'paid_amount': inv.aggregate(
                total=Sum(F('total_amount') - F('balance'), output_field=DecimalField(max_digits=18, decimal_places=2))
            )['total'] or Decimal('0.00'),
            'outstanding_amount': inv.filter(status__in=['issued', 'partially_paid', 'overdue']).aggregate(
                total=Sum('balance')
            )['total'] or Decimal('0.00'),
        }
        by_type = list(
            inv.annotate(bill_type=_invoice_summary_category_case())
            .values('bill_type')
            .annotate(count=Count('id'), amount=Sum('total_amount'))
            .order_by('bill_type')
        )
        context = {
            'totals': totals,
            'by_type': by_type,
            'date_from': date_from,
            'date_to': date_to,
            'showing_invoices_fallback': True,
        }
        return render(request, 'hospital/billing/bill_total.html', context)

    bills = Bill.objects.filter(
        issued_at__date__gte=date_from,
        issued_at__date__lte=date_to,
        is_deleted=False,
    )
    totals = {
        'total_bills': bills.count(),
        'total_amount': bills.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00'),
        'insurance_covered': bills.aggregate(total=Sum('insurance_covered'))['total'] or Decimal('0.00'),
        'patient_portion': bills.aggregate(total=Sum('patient_portion'))['total'] or Decimal('0.00'),
        'paid_amount': bills.filter(status='paid').aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00'),
        'outstanding_amount': bills.filter(
            status__in=['issued', 'partially_paid']
        ).aggregate(total=Sum('patient_portion'))['total'] or Decimal('0.00'),
    }
    by_type = list(
        Bill.objects.filter(
            issued_at__date__gte=date_from,
            issued_at__date__lte=date_to,
            is_deleted=False,
        )
        .values('bill_type')
        .annotate(count=Count('id'), amount=Sum('total_amount'))
        .order_by('bill_type')
    )
    context = {
        'totals': totals,
        'by_type': by_type,
        'date_from': date_from,
        'date_to': date_to,
        'showing_invoices_fallback': False,
    }
    return render(request, 'hospital/billing/bill_total.html', context)


# ==================== COMPANY/CORPORATE BILLS ====================

def _company_bill_statement_sort_date(stmt):
    """Return a datetime.date for sorting — avoids TypeError mixing date vs datetime in max()/sort."""
    raw = None
    if hasattr(stmt, 'statement_date') and stmt.statement_date:
        raw = stmt.statement_date
    elif hasattr(stmt, 'period_start') and stmt.period_start:
        raw = stmt.period_start
    else:
        return date(2000, 1, 1)
    if isinstance(raw, datetime):
        return timezone.localdate(raw)
    if isinstance(raw, date):
        return raw
    return date(2000, 1, 1)


def _company_display_name_for_statement(stmt):
    ca = getattr(stmt, 'corporate_account', None)
    if ca and getattr(ca, 'company_name', None):
        n = (ca.company_name or '').strip()
        if n:
            return n
    p = getattr(stmt, 'payer', None)
    if p and getattr(p, 'name', None):
        n = (p.name or '').strip()
        if n:
            return n
    return 'N/A'


def _company_bill_decimal(v):
    if v is None:
        return Decimal('0.00')
    if isinstance(v, Decimal):
        return v
    return Decimal(str(v))


def _corporate_invoice_nonzero_q():
    """
    Invoices worth showing on corporate print/download: omit shells with no charges and no balance movement.
    Keeps fully paid invoices (total > 0, balance 0) so the company still sees billed activity.
    """
    return Q(total_amount__gt=0) | Q(balance__gt=0) | Q(balance__lt=0)


def _format_company_bill_patient_names(name_set, max_show=4):
    if not name_set:
        return '—'
    names = sorted((n for n in name_set if n and str(n).strip()), key=lambda x: str(x).lower())
    if not names:
        return '—'
    if len(names) <= max_show:
        return ', '.join(names)
    rest = len(names) - max_show
    return ', '.join(names[:max_show]) + f' (+{rest} more)'


def _tag_company_bill_row_kinds(rows):
    """Mark each row for UI badges (monthly statement vs invoice roll-up vs receivable)."""
    for s in rows:
        if getattr(s, 'is_invoice_summary', False):
            setattr(s, 'row_kind', 'invoices')
            setattr(s, 'row_kind_label', 'Open invoices')
        elif getattr(s, 'is_receivable', False):
            setattr(s, 'row_kind', 'receivable')
            setattr(s, 'row_kind_label', 'Receivable entry')
        elif isinstance(s, MonthlyStatement):
            setattr(s, 'row_kind', 'statement')
            setattr(s, 'row_kind_label', 'Monthly statement')
        else:
            setattr(s, 'row_kind', 'other')
            setattr(s, 'row_kind_label', 'Other')


def _corporate_bills_analytics_payload(group_rows, all_statements):
    """
    JSON-serializable analytics for corporate bills dashboard (Chart.js + KPI cards).
    """
    from collections import Counter

    def _fdec(x):
        try:
            return float(x or 0)
        except (TypeError, ValueError):
            return 0.0

    tc_all = sum(_company_bill_decimal(getattr(s, 'total_charges', None)) for s in all_statements)
    tp_all = sum(_company_bill_decimal(getattr(s, 'total_payments', None)) for s in all_statements)
    cl_all = sum(_company_bill_decimal(getattr(s, 'closing_balance', None)) for s in all_statements)
    companies_outstanding = sum(1 for g in group_rows if _company_bill_decimal(g.get('closing_balance')) > 0)

    by_kind = Counter()
    for s in all_statements:
        if getattr(s, 'is_invoice_summary', False):
            by_kind['invoices'] += 1
        elif getattr(s, 'is_receivable', False):
            by_kind['receivable'] += 1
        elif isinstance(s, MonthlyStatement):
            by_kind['statement'] += 1
        else:
            by_kind['other'] += 1

    status_labels = Counter()
    for s in all_statements:
        disp = None
        try:
            fn = getattr(s, 'get_status_display', None)
            if callable(fn):
                disp = fn()
        except Exception:
            disp = None
        if not disp:
            disp = str(getattr(s, 'status', '') or 'unknown').replace('_', ' ').title()
        status_labels[disp] += 1

    top_sorted = sorted(
        group_rows,
        key=lambda g: _company_bill_decimal(g.get('closing_balance')),
        reverse=True,
    )[:12]
    top_labels = []
    top_amounts = []
    for g in top_sorted:
        name = (g.get('company_name') or '')[:42]
        if len(g.get('company_name') or '') > 42:
            name += '…'
        top_labels.append(name or '—')
        top_amounts.append(_fdec(g.get('closing_balance')))

    kind_labels = []
    kind_counts = []
    kind_key_title = {
        'statement': 'Monthly statements',
        'invoices': 'Open invoice roll-ups',
        'receivable': 'Receivable entries',
        'other': 'Other',
    }
    for key in ('statement', 'invoices', 'receivable', 'other'):
        c = by_kind.get(key, 0)
        if c:
            kind_labels.append(kind_key_title.get(key, key))
            kind_counts.append(c)

    st_labels = list(status_labels.keys())[:14]
    st_counts = [status_labels[l] for l in st_labels]

    top_table = []
    for g in top_sorted:
        pid = g.get('primary_payer_id')
        top_table.append(
            {
                'name': g.get('company_name') or '—',
                'closing': _fdec(g.get('closing_balance')),
                'charges': _fdec(g.get('total_charges')),
                'payer_id': str(pid) if pid is not None else None,
            }
        )

    return {
        'summary': {
            'companies': len(group_rows),
            'lines': len(all_statements),
            'total_charges': _fdec(tc_all),
            'total_payments': _fdec(tp_all),
            'total_closing': _fdec(cl_all),
            'companies_with_outstanding': companies_outstanding,
        },
        'by_kind': {'labels': kind_labels, 'counts': kind_counts},
        'status': {'labels': st_labels, 'counts': st_counts},
        'top_outstanding': {'labels': top_labels, 'amounts': top_amounts},
        'top_table': top_table,
    }


def _enrich_company_bill_rows_with_patient_display(rows, corporate_invoice_qs):
    """Set patient_display on MonthlyStatement, invoice-summary, and receivable rows for company bill list."""
    from .models_enterprise_billing import StatementLine

    monthly_ids = [s.id for s in rows if isinstance(s, MonthlyStatement)]
    by_stmt = defaultdict(set)
    if monthly_ids:
        for line in StatementLine.objects.filter(
            statement_id__in=monthly_ids, is_deleted=False
        ).select_related('patient'):
            p = line.patient
            if not p:
                continue
            nm = (getattr(p, 'full_name', None) or f'{p.first_name or ""} {p.last_name or ""}').strip() or '—'
            by_stmt[line.statement_id].add(nm)

    payer_to_names = defaultdict(set)
    payer_ids = {getattr(s, 'payer_id', None) for s in rows if getattr(s, 'is_invoice_summary', False)}
    payer_ids.discard(None)
    if payer_ids:
        for inv in corporate_invoice_qs.filter(payer_id__in=payer_ids).select_related('patient'):
            if not inv.patient_id or not inv.patient:
                continue
            p = inv.patient
            nm = (getattr(p, 'full_name', None) or f'{p.first_name or ""} {p.last_name or ""}').strip() or '—'
            payer_to_names[inv.payer_id].add(nm)

    for s in rows:
        if isinstance(s, MonthlyStatement):
            setattr(s, 'patient_display', _format_company_bill_patient_names(by_stmt.get(s.id, set())))
        elif getattr(s, 'is_invoice_summary', False):
            pid = getattr(s, 'payer_id', None)
            setattr(s, 'patient_display', _format_company_bill_patient_names(payer_to_names.get(pid, set())))
        elif getattr(s, 'is_receivable', False):
            setattr(s, 'patient_display', '—')
        else:
            setattr(s, 'patient_display', '—')


@login_required
@role_required('accountant', 'senior_account_officer')
def company_bill_list(request):
    """List all company/corporate bills - includes MonthlyStatements, Corporate Receivables, and unbilled corporate invoices"""
    from django.db.models import Min, Max
    from types import SimpleNamespace
    from decimal import Decimal
    
    # Get MonthlyStatements: linked to corporate account OR payer is corporate
    statements = MonthlyStatement.objects.filter(
        Q(corporate_account__isnull=False) | Q(payer__payer_type='corporate'),
        is_deleted=False
    ).select_related('corporate_account', 'payer').order_by('-statement_date')
    # Omit empty statements (nothing to send on a company pack)
    statements = statements.exclude(
        (Q(total_charges__isnull=True) | Q(total_charges=0))
        & (Q(closing_balance__isnull=True) | Q(closing_balance=0))
    )
    
    # Get Corporate Receivables (from InsuranceReceivableEntry)
    # Handle case where table doesn't exist (migration not run)
    try:
        corporate_receivables = InsuranceReceivableEntry.objects.filter(
            payer__payer_type='corporate',
            outstanding_amount__gt=0,
            is_deleted=False
        ).select_related('payer').order_by('-entry_date')
    except Exception:
        # Table doesn't exist or other database error - use empty queryset
        corporate_receivables = InsuranceReceivableEntry.objects.none()
    
    # Unbilled corporate invoices: use all_objects so write-off period doesn't hide them; group by payer
    corporate_invoices = (
        Invoice.all_objects.filter(
            payer__payer_type='corporate',
            is_deleted=False,
        )
        .exclude(status='cancelled')
        .filter(_corporate_invoice_nonzero_q())
        .select_related('payer', 'patient')
    )
    
    # Filters
    company_filter = request.GET.get('company', '')
    status_filter = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if company_filter:
        statements = statements.filter(
            Q(corporate_account__company_name__icontains=company_filter) |
            Q(payer__name__icontains=company_filter)
        )
        try:
            corporate_receivables = corporate_receivables.filter(payer__name__icontains=company_filter)
        except Exception:
            pass  # Table doesn't exist, skip filtering
        corporate_invoices = corporate_invoices.filter(payer__name__icontains=company_filter)
    
    if status_filter:
        statements = statements.filter(status=status_filter)
        # Map status filter to receivable status
        try:
            if status_filter == 'paid':
                corporate_receivables = corporate_receivables.filter(status='paid')
            elif status_filter == 'overdue':
                corporate_receivables = corporate_receivables.filter(status__in=['pending', 'partially_paid'])
        except Exception:
            pass  # Table doesn't exist, skip filtering
    
    if date_from:
        statements = statements.filter(statement_date__gte=date_from)
        try:
            corporate_receivables = corporate_receivables.filter(entry_date__gte=date_from)
        except Exception:
            pass  # Table doesn't exist, skip filtering
        corporate_invoices = corporate_invoices.filter(issued_at__date__gte=date_from)
    if date_to:
        statements = statements.filter(statement_date__lte=date_to)
        try:
            corporate_receivables = corporate_receivables.filter(entry_date__lte=date_to)
        except Exception:
            pass  # Table doesn't exist, skip filtering
        corporate_invoices = corporate_invoices.filter(issued_at__date__lte=date_to)
    
    # Aggregate corporate invoices by payer (unbilled / activity view)
    invoice_summary_statements = []
    payer_ids_with_invoices = set()
    try:
        from django.db.models import Sum, Count
        inv_agg = corporate_invoices.values('payer_id', 'payer__name').annotate(
            total_charges=Sum('total_amount'),
            closing_balance=Sum('balance'),
            invoice_count=Count('id'),
            period_start=Min('issued_at'),
            period_end=Max('issued_at'),
        ).order_by('-period_end')
        for row in inv_agg:
            if not row.get('payer_id'):
                continue
            _tc = row.get('total_charges') or Decimal('0')
            _cb = row.get('closing_balance') or Decimal('0')
            if _tc == 0 and _cb == 0:
                continue
            payer_ids_with_invoices.add(row['payer_id'])
            stmt = SimpleNamespace()
            payer_name = row.get('payer__name') or 'N/A'
            stmt.id = f"inv-{row['payer_id']}"
            stmt.statement_number = f"Invoices ({row['invoice_count']})"
            stmt.is_invoice_summary = True
            stmt.payer_id = row['payer_id']
            stmt.payer = SimpleNamespace(name=payer_name)
            stmt.corporate_account = SimpleNamespace(company_name=payer_name)
            stmt.period_start = row['period_start'].date() if row.get('period_start') else None
            stmt.period_end = row['period_end'].date() if row.get('period_end') else None
            stmt.statement_date = row['period_end'] or row['period_start']
            if hasattr(stmt.statement_date, 'date'):
                stmt.statement_date = stmt.statement_date.date()
            stmt.total_charges = row['total_charges'] or Decimal('0.00')
            stmt.total_payments = Decimal('0.00')
            stmt.closing_balance = row['closing_balance'] or Decimal('0.00')
            stmt.status = 'issued' if (row['closing_balance'] or 0) > 0 else 'paid'
            _status = stmt.status
            stmt.get_status_display = lambda: 'Unpaid' if _status == 'issued' else 'Paid'
            invoice_summary_statements.append(stmt)
    except Exception:
        pass

    # Convert corporate receivables to statement-like objects for display
    receivable_statements = []
    try:
        status_choices = dict(InsuranceReceivableEntry.STATUS_CHOICES)
    except:
        status_choices = {}
    
    # Only process if table exists and has data
    try:
        if corporate_receivables.exists():
            for rec in corporate_receivables:
                # Create a statement-like object
                stmt = SimpleNamespace()
                stmt.id = rec.id
                stmt.statement_number = rec.entry_number
                stmt.is_receivable = True  # Flag to identify as receivable
                stmt.corporate_account = SimpleNamespace(company_name=rec.payer.name)
                stmt.period_start = rec.entry_date
                stmt.period_end = rec.entry_date
                stmt.statement_date = rec.entry_date
                stmt.total_charges = rec.total_amount
                stmt.total_payments = rec.amount_received
                stmt.closing_balance = rec.outstanding_amount
                stmt.status = rec.status
                # Create a method that returns the status display (using closure to capture rec.status)
                rec_status = rec.status  # Capture in local variable
                stmt.get_status_display = lambda: status_choices.get(rec_status, rec_status.title())
                stmt.payer_id = rec.payer_id
                receivable_statements.append(stmt)
    except Exception:
        # Table doesn't exist or error accessing it - skip processing
        pass
    
    # Combine all three: statements, receivables, and unbilled invoice summary by corporate payer
    all_statements = list(statements) + receivable_statements + invoice_summary_statements
    all_statements.sort(key=_company_bill_statement_sort_date, reverse=True)
    _enrich_company_bill_rows_with_patient_display(all_statements, corporate_invoices)
    _tag_company_bill_row_kinds(all_statements)

    # Group by company (case-insensitive); parent row shows rolled-up totals, children expandable
    by_key = defaultdict(list)
    key_display = {}
    for stmt in all_statements:
        label = _company_display_name_for_statement(stmt)
        gk = label.strip().lower() or 'n/a'
        if gk not in key_display:
            key_display[gk] = label
        by_key[gk].append(stmt)

    group_rows = []
    for gk in sorted(by_key.keys(), key=lambda k: max(_company_bill_statement_sort_date(s) for s in by_key[k]), reverse=True):
        children = sorted(by_key[gk], key=_company_bill_statement_sort_date, reverse=True)
        g_tc = sum(_company_bill_decimal(getattr(s, 'total_charges', None)) for s in children)
        g_tp = sum(_company_bill_decimal(getattr(s, 'total_payments', None)) for s in children)
        g_cl = sum(_company_bill_decimal(getattr(s, 'closing_balance', None)) for s in children)
        _ppid = None
        for ch in children:
            pid = getattr(ch, 'payer_id', None)
            if pid is not None:
                _ppid = pid
                break
        group_rows.append(
            {
                'company_name': key_display[gk],
                'total_charges': g_tc,
                'total_payments': g_tp,
                'closing_balance': g_cl,
                'row_count': len(children),
                'children': children,
                'primary_payer_id': _ppid,
            }
        )

    # Drop company groups with no monetary activity (cleaner list & print)
    group_rows = [
        g
        for g in group_rows
        if (
            _company_bill_decimal(g.get('total_charges')) != 0
            or _company_bill_decimal(g.get('total_payments')) != 0
            or _company_bill_decimal(g.get('closing_balance')) != 0
        )
    ]

    visible_statements = [s for g in group_rows for s in g['children']]
    corporate_bills_analytics = _corporate_bills_analytics_payload(group_rows, visible_statements)
    corporate_bills_charts = {
        k: corporate_bills_analytics[k]
        for k in ('summary', 'by_kind', 'status', 'top_outstanding')
        if k in corporate_bills_analytics
    }

    total_line_count = len(visible_statements)
    total_group_count = len(group_rows)
    paginator = Paginator(group_rows, 50)
    page = request.GET.get('page')
    statements_page = paginator.get_page(page)

    context = {
        'statement_groups': statements_page,
        'total_corporate_bills_count': total_group_count,
        'total_corporate_line_count': total_line_count,
        'corporate_bills_analytics': corporate_bills_analytics,
        'corporate_bills_charts': corporate_bills_charts,
        'title': 'Company Bills',
        'company_filter': company_filter,
        'status_filter': status_filter,
        'date_from': date_from,
        'date_to': date_to,
    }

    return render(request, 'hospital/billing/company_bill_list.html', context)


@login_required
@role_required('accountant', 'senior_account_officer')
def company_bill_with_cover_page(request, statement_id):
    """Company bill with cover page format"""
    statement = get_object_or_404(MonthlyStatement, pk=statement_id, is_deleted=False)
    
    # Get statement lines
    try:
        from .models_enterprise_billing import StatementLine
        lines = StatementLine.objects.filter(
            statement=statement,
            is_deleted=False
        ).select_related('invoice', 'patient').order_by('line_number')
    except:
        lines = []
    
    context = {
        'statement': statement,
        'lines': lines,
        'cover_page': True,
    }
    
    return render(request, 'hospital/billing/company_bill_cover_page.html', context)


@login_required
@role_required('accountant', 'senior_account_officer')
def company_bill_with_offer(request, statement_id):
    """Company bill with offer/promotion details"""
    statement = get_object_or_404(MonthlyStatement, pk=statement_id, is_deleted=False)
    
    try:
        from .models_enterprise_billing import StatementLine
        lines = StatementLine.objects.filter(
            statement=statement,
            is_deleted=False
        ).select_related('invoice', 'patient').order_by('line_number')
    except:
        lines = []
    
    context = {
        'statement': statement,
        'lines': lines,
        'show_offer': True,
    }
    
    return render(request, 'hospital/billing/company_bill_offer.html', context)


def _group_invoice_lines_by_category(lines):
    """Group invoice lines by service category (Consultation, Lab, Pharmacy, Imaging, etc.) for display."""
    from itertools import groupby

    from .utils_billing import invoice_line_display_category

    sorted_lines = sorted(lines, key=invoice_line_display_category)
    return [(cat, list(grp)) for cat, grp in groupby(sorted_lines, key=invoice_line_display_category)]


def _corporate_bill_pack_invoices_with_lines(payer):
    """
    Invoices and line items for a corporate payer (detail page and Excel/CSV exports).
    Returns (invoices_with_lines, total_charges, total_balance).
    """
    invoices = (
        Invoice.all_objects.filter(payer=payer, is_deleted=False)
        .exclude(status='cancelled')
        .filter(_corporate_invoice_nonzero_q())
        .select_related('patient', 'encounter')
        .order_by('-issued_at')
    )
    invoices_with_lines = []
    for inv in invoices:
        inv.update_totals()
        try:
            from hospital.utils_invoice_line import heal_invoice_zero_line_prices
            heal_invoice_zero_line_prices(inv)
        except Exception:
            pass
        _tot = inv.total_amount or Decimal('0')
        _bal = inv.balance if inv.balance is not None else _tot
        if _tot == 0 and _bal == 0:
            continue
        lines = list(
            inv.lines.filter(is_deleted=False).filter(waived_at__isnull=True).select_related('service_code')
        )
        lines_by_category = _group_invoice_lines_by_category(lines)
        invoices_with_lines.append({'invoice': inv, 'lines_by_category': lines_by_category, 'lines': lines})
    total_charges = sum((x['invoice'].total_amount or Decimal('0')) for x in invoices_with_lines)
    total_balance = sum((x['invoice'].balance or Decimal('0')) for x in invoices_with_lines)
    return invoices_with_lines, total_charges, total_balance


def _invoice_line_category_sums_by_payer(date_from, date_to, payer_types):
    """
    Sum line totals by payer for labs / pharmacy / imaging-style categories (matches ServiceCode usage in billing).
    Returns dict[payer_id] -> {'labs', 'drugs', 'scans'} as Decimals.
    """
    dec = DecimalField(max_digits=12, decimal_places=2)
    lab_q = (
        Q(service_code__category__icontains='laboratory')
        | Q(service_code__category__icontains='lab')
    )
    drug_q = Q(service_code__category__icontains='pharmacy') | Q(service_code__category__icontains='drug')
    scan_q = (
        Q(service_code__category__icontains='imaging')
        | Q(service_code__category__icontains='radiology')
        | Q(service_code__category__icontains='scan')
    )
    base = InvoiceLine.objects.filter(
        is_deleted=False,
        waived_at__isnull=True,
        invoice__is_deleted=False,
        invoice__payer__is_deleted=False,
        invoice__payer__payer_type__in=list(payer_types),
        invoice__issued_at__date__gte=date_from,
        invoice__issued_at__date__lte=date_to,
    ).exclude(invoice__status='cancelled')
    rows = base.values('invoice__payer_id').annotate(
        labs=Coalesce(
            Sum(Case(When(lab_q, then=F('line_total')), default=Value(Decimal('0.00')), output_field=dec)),
            Value(Decimal('0.00')),
            output_field=dec,
        ),
        drugs=Coalesce(
            Sum(Case(When(drug_q, then=F('line_total')), default=Value(Decimal('0.00')), output_field=dec)),
            Value(Decimal('0.00')),
            output_field=dec,
        ),
        scans=Coalesce(
            Sum(Case(When(scan_q, then=F('line_total')), default=Value(Decimal('0.00')), output_field=dec)),
            Value(Decimal('0.00')),
            output_field=dec,
        ),
    )
    return {r['invoice__payer_id']: r for r in rows}


def _claim_item_category_sums_by_payer(date_from, date_to):
    """Insurance claim billed amounts by payer and high-level category (service_code.category)."""
    dec = DecimalField(max_digits=12, decimal_places=2)
    lab_q = (
        Q(service_code__category__icontains='laboratory')
        | Q(service_code__category__icontains='lab')
    )
    drug_q = Q(service_code__category__icontains='pharmacy') | Q(service_code__category__icontains='drug')
    scan_q = (
        Q(service_code__category__icontains='imaging')
        | Q(service_code__category__icontains='radiology')
        | Q(service_code__category__icontains='scan')
    )
    base = InsuranceClaimItem.objects.filter(
        is_deleted=False,
        service_date__gte=date_from,
        service_date__lte=date_to,
        payer__is_deleted=False,
        payer__payer_type__in=['nhis', 'private', 'insurance'],
    ).filter(insurance_claim_item_deduped_q())
    rows = base.values('payer_id').annotate(
        labs=Coalesce(
            Sum(Case(When(lab_q, then=F('billed_amount')), default=Value(Decimal('0.00')), output_field=dec)),
            Value(Decimal('0.00')),
            output_field=dec,
        ),
        drugs=Coalesce(
            Sum(Case(When(drug_q, then=F('billed_amount')), default=Value(Decimal('0.00')), output_field=dec)),
            Value(Decimal('0.00')),
            output_field=dec,
        ),
        scans=Coalesce(
            Sum(Case(When(scan_q, then=F('billed_amount')), default=Value(Decimal('0.00')), output_field=dec)),
            Value(Decimal('0.00')),
            output_field=dec,
        ),
    )
    return {r['payer_id']: r for r in rows}


def _statement_line_category_sums_by_corporate_account(date_from, date_to):
    """StatementLine totals by corporate account (category field and/or ServiceCode) for monthly statements."""
    dec = DecimalField(max_digits=12, decimal_places=2)
    lab_q = (
        Q(category__icontains='laboratory')
        | Q(category__icontains='lab')
        | Q(service_code__category__icontains='laboratory')
        | Q(service_code__category__icontains='lab')
    )
    drug_q = (
        Q(category__icontains='pharmacy')
        | Q(category__icontains='drug')
        | Q(service_code__category__icontains='pharmacy')
        | Q(service_code__category__icontains='drug')
    )
    scan_q = (
        Q(category__icontains='imaging')
        | Q(category__icontains='radiology')
        | Q(category__icontains='scan')
        | Q(service_code__category__icontains='imaging')
        | Q(service_code__category__icontains='radiology')
        | Q(service_code__category__icontains='scan')
    )
    base = StatementLine.objects.filter(
        is_deleted=False,
        statement__is_deleted=False,
        statement__corporate_account__isnull=False,
        statement__statement_date__gte=date_from,
        statement__statement_date__lte=date_to,
    )
    rows = base.values('statement__corporate_account_id').annotate(
        labs=Coalesce(
            Sum(Case(When(lab_q, then=F('line_total')), default=Value(Decimal('0.00')), output_field=dec)),
            Value(Decimal('0.00')),
            output_field=dec,
        ),
        drugs=Coalesce(
            Sum(Case(When(drug_q, then=F('line_total')), default=Value(Decimal('0.00')), output_field=dec)),
            Value(Decimal('0.00')),
            output_field=dec,
        ),
        scans=Coalesce(
            Sum(Case(When(scan_q, then=F('line_total')), default=Value(Decimal('0.00')), output_field=dec)),
            Value(Decimal('0.00')),
            output_field=dec,
        ),
    )
    return {r['statement__corporate_account_id']: r for r in rows}


@login_required
@role_required('accountant', 'senior_account_officer')
def corporate_bill_detail_by_payer(request, payer_id):
    """Show individual corporate bill: all invoices for this payer with full line items (labs, scans, drugs, etc.)."""
    payer = get_object_or_404(Payer, pk=payer_id, payer_type='corporate', is_deleted=False)
    invoices_with_lines, total_charges, total_balance = _corporate_bill_pack_invoices_with_lines(payer)
    patient_invoice_groups = _corporate_invoices_grouped_by_patient(invoices_with_lines)
    context = {
        'payer': payer,
        'invoices_with_lines': invoices_with_lines,
        'patient_invoice_groups': patient_invoice_groups,
        'total_charges': total_charges,
        'total_balance': total_balance,
        'invoice_count': len(invoices_with_lines),
    }
    return render(request, 'hospital/billing/corporate_bill_detail_by_payer.html', context)


@login_required
@role_required('accountant', 'senior_account_officer')
def corporate_bill_pack_export_excel(request, payer_id):
    """Download corporate bill pack as .xlsx (one section per invoice, lines by category)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return HttpResponse(
            'Excel export requires openpyxl. Install with: pip install openpyxl',
            content_type='text/plain',
            status=500,
        )

    payer = get_object_or_404(Payer, pk=payer_id, payer_type='corporate', is_deleted=False)
    invoices_with_lines, total_charges, total_balance = _corporate_bill_pack_invoices_with_lines(payer)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Corporate bill pack'

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0D9488', end_color='0D9488', fill_type='solid')
    inv_header_fill = PatternFill(start_color='64748B', end_color='64748B', fill_type='solid')

    row = 1
    ws.cell(row=row, column=1, value='Corporate bill pack').font = title_font
    row += 1
    ws.cell(row=row, column=1, value=f'Payer: {payer.name}')
    row += 1
    ws.cell(row=row, column=1, value=f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M")}')
    row += 1
    ws.cell(row=row, column=1, value=f'Invoices: {len(invoices_with_lines)}  |  Total charges (GHS): {total_charges}  |  Balance due (GHS): {total_balance}')
    row += 2

    if not invoices_with_lines:
        ws.cell(row=row, column=1, value='No invoices in this pack.')
    else:
        for item in invoices_with_lines:
            inv = item['invoice']
            patient = inv.patient
            pname = patient.full_name if patient else '—'
            mrn = getattr(patient, 'mrn', None) or '—'
            issued = inv.issued_at.strftime('%Y-%m-%d') if inv.issued_at else '—'

            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            c = ws.cell(row=row, column=1, value=f"Invoice {inv.invoice_number}  |  {pname}  |  {issued}  |  MRN {mrn}")
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = inv_header_fill
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            row += 1

            for col, h in enumerate(['Category', 'Service', 'Qty', 'Unit (GHS)', 'Line total (GHS)'], start=1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center' if col >= 3 else 'left', vertical='center')
            row += 1

            for category, cat_lines in item['lines_by_category']:
                for line in cat_lines:
                    ws.cell(row=row, column=1, value=str(category))
                    ws.cell(row=row, column=2, value=(line.description or '')[:500])
                    ws.cell(row=row, column=3, value=float(line.quantity or 0))
                    ws.cell(row=row, column=4, value=float(line.display_unit_price or 0))
                    ws.cell(row=row, column=5, value=float(line.display_line_total or 0))
                    row += 1

            ws.cell(row=row, column=4, value='Invoice total (GHS):').font = Font(bold=True)
            ws.cell(row=row, column=5, value=float(inv.total_amount or 0)).font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=4, value='Balance due (GHS):').font = Font(bold=True)
            ws.cell(row=row, column=5, value=float(inv.balance or 0)).font = Font(bold=True)
            row += 2

        ws.cell(row=row, column=4, value='Pack total charges (GHS):').font = Font(bold=True)
        ws.cell(row=row, column=5, value=float(total_charges)).font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=4, value='Pack balance due (GHS):').font = Font(bold=True)
        ws.cell(row=row, column=5, value=float(total_balance)).font = Font(bold=True)

    for col_letter, width in [('A', 22), ('B', 48), ('C', 10), ('D', 14), ('E', 16)]:
        ws.column_dimensions[col_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    base = slugify(payer.name)[:80] or 'corporate_bill_pack'
    fname = f'{base}_{timezone.now().strftime("%Y%m%d")}.xlsx'
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


@login_required
@role_required('accountant', 'senior_account_officer')
def corporate_bill_pack_export_csv(request, payer_id):
    """Download corporate bill pack as CSV (UTF-8, Excel-friendly)."""
    payer = get_object_or_404(Payer, pk=payer_id, payer_type='corporate', is_deleted=False)
    invoices_with_lines, total_charges, total_balance = _corporate_bill_pack_invoices_with_lines(payer)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Corporate bill pack'])
    writer.writerow(['Payer', payer.name])
    writer.writerow(['Generated', timezone.now().strftime('%Y-%m-%d %H:%M')])
    writer.writerow(['Invoice count', len(invoices_with_lines)])
    writer.writerow(['Total charges (GHS)', str(total_charges)])
    writer.writerow(['Balance due (GHS)', str(total_balance)])
    writer.writerow([])

    if not invoices_with_lines:
        writer.writerow(['No invoices in this pack.'])
    else:
        writer.writerow(
            ['Invoice #', 'Patient', 'Invoice date', 'MRN', 'Category', 'Service', 'Qty', 'Unit (GHS)', 'Line total (GHS)']
        )
        for item in invoices_with_lines:
            inv = item['invoice']
            patient = inv.patient
            pname = patient.full_name if patient else '—'
            mrn = getattr(patient, 'mrn', None) or '—'
            issued = inv.issued_at.strftime('%Y-%m-%d') if inv.issued_at else '—'
            inv_no = inv.invoice_number
            for category, cat_lines in item['lines_by_category']:
                for line in cat_lines:
                    writer.writerow(
                        [
                            inv_no,
                            pname,
                            issued,
                            mrn,
                            category,
                            (line.description or '').replace('\r\n', ' ').replace('\n', ' '),
                            str(line.quantity or 0),
                            str(line.display_unit_price or 0),
                            str(line.display_line_total or 0),
                        ]
                    )
            writer.writerow([inv_no, pname, issued, mrn, '', 'Invoice total (GHS)', '', '', str(inv.total_amount or 0)])
            writer.writerow([inv_no, pname, issued, mrn, '', 'Balance due (GHS)', '', '', str(inv.balance or 0)])
            writer.writerow([])
        writer.writerow(['', '', '', '', '', 'Pack total charges (GHS)', '', '', str(total_charges)])
        writer.writerow(['', '', '', '', '', 'Pack balance due (GHS)', '', '', str(total_balance)])

    base = slugify(payer.name)[:80] or 'corporate_bill_pack'
    fname = f'{base}_{timezone.now().strftime("%Y%m%d")}.csv'
    response = HttpResponse('\ufeff' + output.getvalue(), content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


@login_required
@role_required('accountant', 'senior_account_officer')
def receivable_entry_detail(request, entry_id):
    """Read-only detail for an insurance/corporate receivable entry (replaces admin change view for View links)."""
    entry = get_object_or_404(
        InsuranceReceivableEntry.objects.select_related('payer', 'matched_by', 'journal_entry'),
        pk=entry_id,
        is_deleted=False,
    )
    payments = (
        InsurancePaymentReceived.objects.filter(receivable_entry=entry, is_deleted=False)
        .select_related('bank_account', 'payer', 'processed_by')
        .order_by('-entry_date', '-created')
    )
    context = {
        'entry': entry,
        'payments': payments,
        'is_corporate': entry.payer.payer_type == 'corporate',
    }
    return render(request, 'hospital/billing/receivable_entry_detail.html', context)


@login_required
@role_required('accountant', 'senior_account_officer')
def corporate_bill_list(request):
    """List all corporate bills"""
    return company_bill_list(request)  # Same as company bills


@login_required
@role_required('accountant', 'senior_account_officer')
def corporate_bill_with_subscriber(request, statement_id):
    """Corporate bill with subscriber/employee details"""
    statement = get_object_or_404(MonthlyStatement, pk=statement_id, is_deleted=False)
    
    # Get employees/subscribers
    employees = CorporateEmployee.objects.filter(
        corporate_account=statement.corporate_account,
        is_active=True,
        is_deleted=False
    ).select_related('patient').order_by('patient__last_name', 'patient__first_name')
    
    try:
        from .models_enterprise_billing import StatementLine
        lines = StatementLine.objects.filter(
            statement=statement,
            is_deleted=False
        ).select_related('invoice', 'patient').order_by('patient__last_name', 'patient__first_name', 'line_number')
    except:
        lines = []
    
    context = {
        'statement': statement,
        'employees': employees,
        'lines': lines,
        'show_subscribers': True,
    }
    
    return render(request, 'hospital/billing/corporate_bill_subscriber.html', context)


# ==================== INSURANCE BILLS ====================


def _normalize_claim_description_for_dedupe(s):
    if not s:
        return ''
    return (str(s).strip().lower())[:500]


def _dedupe_insurance_claim_items_by_invoice_line(claims):
    """
    Keep one claim row per invoice line (newest by created/pk). Collapse orphan rows
    (no invoice_line) that repeat the same billing snapshot — fixes thousands of
    duplicate claim rows on one invoice.
    """
    by_line = {}
    no_line = []
    for c in claims:
        lid = getattr(c, 'invoice_line_id', None)
        if lid:
            prev = by_line.get(lid)
            if prev is None:
                by_line[lid] = c
            else:
                pc, cc = getattr(prev, 'created', None), getattr(c, 'created', None)
                if cc and pc and cc > pc:
                    by_line[lid] = c
                elif cc and not pc:
                    by_line[lid] = c
                elif not cc and not pc and c.pk > prev.pk:
                    by_line[lid] = c
        else:
            no_line.append(c)

    by_snap = {}
    no_line_deduped = []
    for c in sorted(
        no_line,
        key=lambda x: (getattr(x, 'created', None) or timezone.now(), x.pk),
        reverse=True,
    ):
        nd = _normalize_claim_description_for_dedupe(c.service_description)
        if c.invoice_id:
            snap_key = (
                'inv',
                c.invoice_id,
                c.patient_id,
                c.payer_id,
                c.service_date,
                str(c.billed_amount or Decimal('0')),
                nd,
                c.service_code_id,
            )
        else:
            snap_key = (
                'solo',
                c.patient_id,
                c.payer_id,
                c.encounter_id,
                c.service_date,
                str(c.billed_amount or Decimal('0')),
                nd,
                c.service_code_id,
            )
        if snap_key in by_snap:
            continue
        by_snap[snap_key] = True
        no_line_deduped.append(c)

    merged = list(by_line.values()) + no_line_deduped
    merged.sort(
        key=lambda x: (
            x.service_date or date.min,
            getattr(x, 'created', None) or timezone.now(),
            x.pk,
        ),
        reverse=True,
    )
    return merged


def _invoice_billable_line_stats(invoice):
    """
    Returns (raw_line_count, distinct_service_code_count, duplicate_row_count).
    duplicate_row_count = how many extra lines would be removed by merging same service_code.
    """
    from django.db.models import Count

    from .models import InvoiceLine

    rows = list(
        InvoiceLine.objects.filter(
            invoice=invoice, is_deleted=False, waived_at__isnull=True
        )
        .values('service_code_id')
        .annotate(c=Count('id'))
    )
    raw = sum(r['c'] for r in rows) if rows else 0
    distinct = len(rows)
    duplicate = max(0, raw - distinct)
    return raw, distinct, duplicate


def _visit_bucket_key_for_claim(c):
    """Group claim lines into one visit (one combined bill) per invoice when possible."""
    if c.invoice_id:
        return ('inv', c.invoice_id)
    if c.encounter_id:
        return ('enc', c.encounter_id, c.service_date)
    return ('solo', c.id)


def _rollup_insurance_claim_status(items):
    """Single status for a visit group when lines differ."""
    statuses = [i.claim_status for i in items]
    if len(set(statuses)) == 1:
        return statuses[0]
    if 'pending' in statuses:
        return 'pending'
    if 'rejected' in statuses:
        return 'rejected'
    if 'reversed' in statuses:
        return 'reversed'
    order = ['submitted', 'processing', 'approved', 'partially_paid', 'paid']
    for s in order:
        if s in statuses:
            return s
    return statuses[0]


def _insurance_visit_groups_for_payer(claim_items):
    """
    Under one payer: group claim items by visit (invoice / encounter), combine amounts.
    """
    deduped = _dedupe_insurance_claim_items_by_invoice_line(list(claim_items))
    buckets = defaultdict(list)
    for c in deduped:
        buckets[_visit_bucket_key_for_claim(c)].append(c)
    groups = []
    for _key, items in buckets.items():
        items.sort(
            key=lambda x: (getattr(x, 'created', None) or timezone.now(), x.pk),
            reverse=True,
        )
        first = items[0]
        inv = getattr(first, 'invoice', None)
        visit_dt = None
        if inv and inv.issued_at:
            visit_dt = inv.issued_at
        else:
            visit_dt = max(
                (i.created for i in items if getattr(i, 'created', None)),
                default=first.created,
            )
        if inv is not None:
            raw_lines, distinct_svcs, dup_rows = _invoice_billable_line_stats(inv)
            line_count = raw_lines
            billed = inv.total_amount or Decimal('0')
            paid = (inv.total_amount or Decimal('0')) - (inv.balance or Decimal('0'))
        else:
            raw_lines, distinct_svcs, dup_rows = (len(items), len(items), 0)
            line_count = len(items)
            billed = sum((i.billed_amount or Decimal('0') for i in items), Decimal('0'))
            paid = sum((i.paid_amount or Decimal('0') for i in items), Decimal('0'))
        rolled_status = _rollup_insurance_claim_status(items)
        status_label = dict(InsuranceClaimItem.CLAIM_STATUS_CHOICES).get(
            rolled_status, rolled_status.replace('_', ' ').title()
        )
        groups.append(
            {
                'patient': first.patient,
                'payer': first.payer,
                'invoice': inv,
                'visit_dt': visit_dt,
                'service_date': first.service_date,
                'line_count': line_count,
                'invoice_line_rows': raw_lines,
                'invoice_distinct_services': distinct_svcs,
                'invoice_duplicate_rows': dup_rows,
                'claim_row_count': len(items),
                'billed_total': billed,
                'paid_total': paid,
                'claim_status': rolled_status,
                'claim_status_label': status_label,
                'claims': items,
            }
        )
    groups.sort(
        key=lambda g: (g['visit_dt'] or timezone.now(), g['service_date'] or date.min),
        reverse=True,
    )
    return groups


def _corporate_invoices_grouped_by_patient(invoices_with_lines):
    """Group invoice packs by patient for corporate payer detail (visits = invoices, sorted by time)."""
    buckets = defaultdict(list)
    for item in invoices_with_lines:
        inv = item['invoice']
        pid = inv.patient_id if getattr(inv, 'patient_id', None) else None
        key = pid or f'nopatient-{inv.pk}'
        buckets[key].append(item)
    groups = []
    for _key, items in buckets.items():
        items.sort(
            key=lambda x: x['invoice'].issued_at or timezone.now(),
            reverse=True,
        )
        inv0 = items[0]['invoice']
        patient = getattr(inv0, 'patient', None)
        tot_ch = sum((x['invoice'].total_amount or Decimal('0') for x in items), Decimal('0'))
        tot_bal = sum((x['invoice'].balance or Decimal('0') for x in items), Decimal('0'))
        latest = max(
            (x['invoice'].issued_at for x in items if x['invoice'].issued_at),
            default=None,
        )
        groups.append(
            {
                'patient': patient,
                'invoices': items,
                'invoice_count': len(items),
                'total_charges': tot_ch,
                'total_balance': tot_bal,
                'latest_issued': latest,
            }
        )
    groups.sort(key=lambda g: g['latest_issued'] or timezone.now(), reverse=True)
    return groups


@login_required
@role_required('accountant', 'senior_account_officer')
def insurance_bill_list(request):
    """List all insurance bills/claims"""
    # Redirect to clean URL when any param is INVALID so form does not show it
    invalid_keys = [k for k, v in request.GET.items() if (v or '').strip().lower() == 'invalid']
    if invalid_keys:
        cleaned = {k: v for k, v in request.GET.items() if k not in invalid_keys}
        target = request.path + ('?' + urlencode(cleaned) if cleaned else '')
        return redirect(target)
    claims = (
        InsuranceClaimItem.objects.filter(is_deleted=False)
        .filter(insurance_claim_item_deduped_q())
        .select_related('patient', 'payer', 'invoice')
        .order_by('-created')
    )
    
    # Filters (treat INVALID as empty)
    payer_filter = _normalize_filter_value(request.GET.get('payer', ''))
    status_filter = _normalize_filter_value(request.GET.get('status', ''))
    date_from = _normalize_filter_value(request.GET.get('date_from', ''))
    date_to = _normalize_filter_value(request.GET.get('date_to', ''))
    
    if payer_filter:
        claims = claims.filter(payer__name__icontains=payer_filter)
    if status_filter:
        claims = claims.filter(claim_status=status_filter)
    if date_from:
        claims = claims.filter(service_date__gte=date_from)
    if date_to:
        claims = claims.filter(service_date__lte=date_to)

    total_claims_count = claims.count()
    payer_rows = list(
        claims.values('payer_id', 'payer__name')
        .annotate(
            billed_total=Sum('billed_amount'),
            paid_total=Sum('paid_amount'),
            claim_count=Count('id'),
            latest_service=Max('service_date'),
        )
        .order_by('-latest_service')
    )
    total_payer_groups = len(payer_rows)
    paginator = Paginator(payer_rows, 30)
    page = request.GET.get('page')
    payer_page = paginator.get_page(page)

    page_payer_ids = []
    page_has_null_payer = False
    for row in payer_page.object_list:
        if row['payer_id'] is None:
            page_has_null_payer = True
        else:
            page_payer_ids.append(row['payer_id'])

    by_payer = defaultdict(list)
    if payer_page.object_list:
        child_filter = Q()
        if page_payer_ids:
            child_filter |= Q(payer_id__in=page_payer_ids)
        if page_has_null_payer:
            child_filter |= Q(payer_id__isnull=True)
        children_qs = claims.filter(child_filter).select_related('patient', 'payer', 'invoice').order_by(
            '-service_date', '-created'
        )
        for c in children_qs:
            by_payer[c.payer_id].append(c)

    claim_groups = []
    for row in payer_page.object_list:
        pid = row['payer_id']
        visit_groups = _insurance_visit_groups_for_payer(by_payer.get(pid, []))
        billed_total = sum((vg['billed_total'] for vg in visit_groups), Decimal('0'))
        paid_total = sum((vg['paid_total'] for vg in visit_groups), Decimal('0'))
        line_count = sum((vg['line_count'] for vg in visit_groups), 0)
        claim_groups.append(
            {
                'payer_id': pid,
                'payer_name': (row.get('payer__name') or '').strip() or '—',
                'billed_total': billed_total,
                'paid_total': paid_total,
                'claim_count': line_count,
                'visit_count': len(visit_groups),
                'visit_groups': visit_groups,
            }
        )

    context = {
        'claim_groups': Page(claim_groups, payer_page.number, paginator),
        'total_claims_count': total_claims_count,
        'total_payer_groups': total_payer_groups,
        'title': 'Insurance Bills',
        'payer_filter': payer_filter,
        'status_filter': status_filter,
        'date_from': date_from,
        'date_to': date_to,
    }

    return render(request, 'hospital/billing/insurance_bill_list.html', context)


# ==================== MEDICAL BILLS STATEMENT ====================

def _parse_iso_date(val, fallback):
    try:
        if val:
            return date.fromisoformat(str(val).strip()[:10])
    except ValueError:
        pass
    return fallback


@login_required
@role_required('accountant', 'senior_account_officer')
def medical_bills_statement_processing(request):
    """Process and generate medical bills statements with period analytics, export, and print."""
    today = timezone.now().date()
    start_of_month = today.replace(day=1)

    analytics_from = _parse_iso_date(request.GET.get('analytics_from'), start_of_month)
    analytics_to = _parse_iso_date(request.GET.get('analytics_to'), today)
    if analytics_from > analytics_to:
        analytics_from, analytics_to = analytics_to, analytics_from

    statements_in_period = MonthlyStatement.objects.filter(
        statement_date__gte=analytics_from,
        statement_date__lte=analytics_to,
        is_deleted=False,
    ).select_related('corporate_account', 'payer')

    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        fname = f'medical_bills_processing_{analytics_from}_{analytics_to}.csv'
        response['Content-Disposition'] = f'attachment; filename="{fname}"'
        response.write('\ufeff')
        w = csv.writer(response)
        w.writerow(['Medical bills statement processing — analytics export'])
        w.writerow(['Period from', analytics_from.isoformat(), 'to', analytics_to.isoformat()])
        w.writerow([])
        agg = statements_in_period.aggregate(
            cnt=Count('id'),
            charges=Sum('total_charges'),
            payments=Sum('total_payments'),
            closing=Sum('closing_balance'),
            lines=Sum('total_line_items'),
            patients=Sum('total_patients_served'),
        )
        w.writerow(['Statements (count)', agg['cnt'] or 0])
        w.writerow(['Total charges (GHS)', str(agg['charges'] or Decimal('0'))])
        w.writerow(['Total payments (GHS)', str(agg['payments'] or Decimal('0'))])
        w.writerow(['Closing balance sum (GHS)', str(agg['closing'] or Decimal('0'))])
        w.writerow(['Line items (sum)', agg['lines'] or 0])
        w.writerow(['Patients served (sum)', agg['patients'] or 0])
        w.writerow([])
        w.writerow(['Status breakdown'])
        for row in statements_in_period.values('status').annotate(c=Count('id')).order_by('status'):
            w.writerow([row['status'], row['c']])
        w.writerow([])
        w.writerow(
            [
                'Statement #',
                'Statement date',
                'Account',
                'Period start',
                'Period end',
                'Due date',
                'Status',
                'Opening (GHS)',
                'Charges (GHS)',
                'Payments (GHS)',
                'Adjustments (GHS)',
                'Closing (GHS)',
                'Line items',
                'Patients served',
            ]
        )
        for s in statements_in_period.order_by('-statement_date', '-created'):
            acct = (
                s.corporate_account.company_name
                if s.corporate_account_id
                else (s.payer.name if s.payer_id else '')
            )
            w.writerow(
                [
                    s.statement_number,
                    s.statement_date.isoformat(),
                    acct,
                    s.period_start.isoformat(),
                    s.period_end.isoformat(),
                    s.due_date.isoformat() if s.due_date else '',
                    s.status,
                    str(s.opening_balance),
                    str(s.total_charges),
                    str(s.total_payments),
                    str(s.total_adjustments),
                    str(s.closing_balance),
                    s.total_line_items,
                    s.total_patients_served,
                ]
            )
        return response

    if request.method == 'POST':
        period_start = request.POST.get('period_start', start_of_month.isoformat())
        period_end = request.POST.get('period_end', today.isoformat())
        account_type = request.POST.get('account_type', 'all')
        period_start_d = _parse_iso_date(period_start, start_of_month)
        period_end_d = _parse_iso_date(period_end, today)
        redirect_from = _parse_iso_date(request.POST.get('analytics_from'), analytics_from)
        redirect_to = _parse_iso_date(request.POST.get('analytics_to'), analytics_to)
        if redirect_from > redirect_to:
            redirect_from, redirect_to = redirect_to, redirect_from

        try:
            from .services.monthly_billing_service import MonthlyBillingService

            billing_service = MonthlyBillingService()

            if account_type == 'corporate':
                corporate_accounts = CorporateAccount.objects.filter(is_deleted=False)
                generated = 0
                for account in corporate_accounts:
                    statement = billing_service.generate_corporate_statement(
                        account, period_start_d, period_end_d
                    )
                    if statement:
                        generated += 1
                messages.success(request, f'Generated {generated} corporate statements')
            elif account_type == 'insurance':
                messages.success(request, 'Processed insurance claims')
            else:
                result = billing_service.generate_all_monthly_statements(billing_month=period_end_d)
                total = result.get('total_accounts', 0)
                ok = result.get('successful', 0)
                failed = result.get('failed', 0)
                skipped = result.get('skipped', 0)
                errs = result.get('errors', [])
                if total == 0:
                    messages.info(
                        request,
                        'No accounts are set for monthly billing. '
                        'Add corporate accounts with billing cycle "monthly" to generate statements.',
                    )
                elif ok or failed or skipped:
                    msg = f'Statements: {ok} generated, {skipped} skipped (no charges), {failed} failed.'
                    if errs:
                        msg += ' ' + errs[0] if len(errs) == 1 else f' {len(errs)} errors occurred.'
                    messages.success(request, msg)
                else:
                    messages.success(request, 'Processing complete. No statements generated for the period.')

            q = urlencode(
                {
                    'analytics_from': redirect_from.isoformat(),
                    'analytics_to': redirect_to.isoformat(),
                }
            )
            return redirect(f"{request.path}?{q}")
        except Exception as e:
            messages.error(request, f'Error processing statements: {str(e)}')

    pending_corporate = CorporateAccount.objects.filter(
        is_deleted=False,
        current_balance__gt=0,
    ).count()

    pending_insurance = (
        InsuranceClaimItem.objects.filter(claim_status='pending', is_deleted=False)
        .filter(insurance_claim_item_deduped_q())
        .count()
    )

    eligible_monthly_accounts = CorporateAccount.objects.filter(
        is_active=True,
        billing_cycle='monthly',
        is_deleted=False,
    ).count()

    statement_summary = statements_in_period.aggregate(
        total_statements=Count('id'),
        total_charges=Sum('total_charges'),
        total_payments=Sum('total_payments'),
        total_closing_balance=Sum('closing_balance'),
        total_line_items=Sum('total_line_items'),
        total_patients_served=Sum('total_patients_served'),
    )
    n_stmt = statement_summary['total_statements'] or 0
    sum_close = statement_summary['total_closing_balance']
    if n_stmt and sum_close is not None:
        statement_summary['avg_closing'] = sum_close / n_stmt
    elif n_stmt:
        statement_summary['avg_closing'] = Decimal('0.00')
    else:
        statement_summary['avg_closing'] = Decimal('0.00')

    outstanding_due = statements_in_period.exclude(
        status__in=['paid', 'written_off'],
    ).aggregate(t=Sum('closing_balance'))['t'] or Decimal('0.00')

    statements_by_status = []
    status_labels = dict(MonthlyStatement._meta.get_field('status').choices)
    for row in statements_in_period.values('status').annotate(c=Count('id')).order_by('status'):
        statements_by_status.append(
            {
                'status': row['status'],
                'label': status_labels.get(row['status'], row['status']),
                'count': row['c'],
            }
        )

    recent_statements = list(statements_in_period.order_by('-statement_date', '-created')[:40])

    insurance_pending_by_payer = list(
        InsuranceClaimItem.objects.filter(claim_status='pending', is_deleted=False)
        .filter(insurance_claim_item_deduped_q())
        .values('payer__name')
        .annotate(n=Count('id'))
        .order_by('-n')[:8]
    )

    context = {
        'pending_corporate': pending_corporate,
        'pending_insurance': pending_insurance,
        'default_start': start_of_month,
        'default_end': today,
        'analytics_from': analytics_from,
        'analytics_to': analytics_to,
        'statement_summary': statement_summary,
        'statements_by_status': statements_by_status,
        'recent_statements': recent_statements,
        'eligible_monthly_accounts': eligible_monthly_accounts,
        'outstanding_due': outstanding_due,
        'insurance_pending_by_payer': insurance_pending_by_payer,
        'export_csv_url': request.path
        + '?'
        + urlencode(
            {
                'analytics_from': analytics_from.isoformat(),
                'analytics_to': analytics_to.isoformat(),
                'export': 'csv',
            }
        ),
    }

    return render(request, 'hospital/billing/medical_bills_statement_processing.html', context)


def _statement_report_decimal_to_float(v):
    if v is None:
        return 0.0
    return float(v)


@login_required
@role_required('accountant', 'senior_account_officer')
def medical_bills_statement_report(request):
    """Medical bills statement report with filters and chart analytics."""
    today = timezone.now().date()
    start_of_month = today.replace(day=1)

    date_from_str = request.GET.get('date_from', start_of_month.isoformat())
    date_to_str = request.GET.get('date_to', today.isoformat())
    try:
        date_from = date.fromisoformat(date_from_str)
    except ValueError:
        date_from = start_of_month
    try:
        date_to = date.fromisoformat(date_to_str)
    except ValueError:
        date_to = today
    if date_from > date_to:
        date_from, date_to = date_to, date_from
    date_from_str = date_from.isoformat()
    date_to_str = date_to.isoformat()

    account_scope = (request.GET.get('account_scope') or 'all').strip().lower()
    if account_scope not in ('all', 'corporate', 'individual'):
        account_scope = 'all'
    status_filter = (request.GET.get('status') or '').strip()
    valid_statuses = {c[0] for c in STATEMENT_STATUS_CHOICES}
    if status_filter and status_filter not in valid_statuses:
        status_filter = ''
    search = (request.GET.get('search') or '').strip()
    sort = (request.GET.get('sort') or 'statement_date_desc').strip()
    sort_keys = {
        'statement_date_desc': '-statement_date',
        'statement_date_asc': 'statement_date',
        'closing_balance_desc': '-closing_balance',
        'total_charges_desc': '-total_charges',
    }
    order_by = sort_keys.get(sort, '-statement_date')

    base = MonthlyStatement.objects.filter(
        statement_date__gte=date_from,
        statement_date__lte=date_to,
        is_deleted=False,
    ).select_related('corporate_account', 'payer')

    if account_scope == 'corporate':
        base = base.filter(corporate_account__isnull=False)
    elif account_scope == 'individual':
        base = base.filter(corporate_account__isnull=True)
    if status_filter:
        base = base.filter(status=status_filter)
    if search:
        base = base.filter(
            Q(statement_number__icontains=search)
            | Q(corporate_account__company_name__icontains=search)
            | Q(payer__name__icontains=search)
        )

    summary = base.aggregate(
        total_statements=Count('id'),
        total_charges=Sum('total_charges'),
        total_payments=Sum('total_payments'),
        closing_balance=Sum('closing_balance'),
    )

    corporate_qs = base.filter(corporate_account__isnull=False)
    non_corporate_qs = base.filter(corporate_account__isnull=True)
    by_type = [
        {
            'label': 'Corporate',
            'count': corporate_qs.count(),
            'total_charges': corporate_qs.aggregate(t=Sum('total_charges'))['t'] or Decimal('0.00'),
            'total_payments': corporate_qs.aggregate(t=Sum('total_payments'))['t'] or Decimal('0.00'),
            'closing_balance': corporate_qs.aggregate(t=Sum('closing_balance'))['t'] or Decimal('0.00'),
        },
        {
            'label': 'Individual (payer)',
            'count': non_corporate_qs.count(),
            'total_charges': non_corporate_qs.aggregate(t=Sum('total_charges'))['t'] or Decimal('0.00'),
            'total_payments': non_corporate_qs.aggregate(t=Sum('total_payments'))['t'] or Decimal('0.00'),
            'closing_balance': non_corporate_qs.aggregate(t=Sum('closing_balance'))['t'] or Decimal('0.00'),
        },
    ]

    span_days = (date_to - date_from).days + 1
    status_labels_map = dict(STATEMENT_STATUS_CHOICES)
    if span_days <= 62:
        granularity = 'daily'
        trend_rows = (
            base.annotate(bucket=TruncDate('statement_date'))
            .values('bucket')
            .annotate(
                charges=Sum('total_charges'),
                payments=Sum('total_payments'),
                cnt=Count('id'),
            )
            .order_by('bucket')
        )
    else:
        granularity = 'monthly'
        trend_rows = (
            base.annotate(bucket=TruncMonth('statement_date'))
            .values('bucket')
            .annotate(
                charges=Sum('total_charges'),
                payments=Sum('total_payments'),
                cnt=Count('id'),
            )
            .order_by('bucket')
        )

    trend_labels = []
    trend_charges = []
    trend_payments = []
    trend_counts = []
    for row in trend_rows:
        b = row['bucket']
        if b is None:
            continue
        if hasattr(b, 'date'):
            d = b.date()
        else:
            d = b
        if granularity == 'daily':
            trend_labels.append(d.isoformat())
        else:
            trend_labels.append(f'{d.year:04d}-{d.month:02d}')
        trend_charges.append(_statement_report_decimal_to_float(row['charges']))
        trend_payments.append(_statement_report_decimal_to_float(row['payments']))
        trend_counts.append(row['cnt'])

    status_rows = (
        base.values('status')
        .annotate(cnt=Count('id'), bal=Sum('closing_balance'))
        .order_by('-cnt')
    )
    st_labels = []
    st_counts = []
    st_amounts = []
    for row in status_rows:
        st_labels.append(status_labels_map.get(row['status'], row['status'] or '—'))
        st_counts.append(row['cnt'])
        st_amounts.append(_statement_report_decimal_to_float(row['bal']))

    analytics_data = {
        'meta': {'granularity': granularity},
        'charts': {
            'trend': {
                'labels': trend_labels,
                'charges': trend_charges,
                'payments': trend_payments,
                'counts': trend_counts,
            },
            'status': {
                'labels': st_labels,
                'counts': st_counts,
                'amounts': st_amounts,
            },
            'account_type': {
                'labels': [b['label'] for b in by_type],
                'counts': [b['count'] for b in by_type],
                'balances': [
                    _statement_report_decimal_to_float(b['closing_balance']) for b in by_type
                ],
            },
        },
    }

    ordered = base.order_by(order_by)
    paginator = Paginator(ordered, 40)
    statements_page = paginator.get_page(request.GET.get('page'))

    q = request.GET.copy()
    q.pop('page', None)
    page_querystring = q.urlencode()

    context = {
        'statements': statements_page,
        'summary': summary,
        'by_type': by_type,
        'date_from': date_from_str,
        'date_to': date_to_str,
        'account_scope': account_scope,
        'status_filter': status_filter,
        'search': search,
        'sort': sort,
        'statement_status_choices': STATEMENT_STATUS_CHOICES,
        'page_querystring': page_querystring,
        'analytics_data': analytics_data,
    }

    return render(request, 'hospital/billing/medical_bills_statement_report.html', context)


# ==================== STAFF DEP BILL ====================

@login_required
@role_required('accountant', 'senior_account_officer')
def staff_dep_bill_list(request):
    """Staff dependency bills (bills for staff dependents)"""
    # Redirect to clean URL when any param is INVALID so form does not show it
    invalid_keys = [k for k, v in request.GET.items() if (v or '').strip().lower() == 'invalid']
    if invalid_keys:
        cleaned = {k: v for k, v in request.GET.items() if k not in invalid_keys}
        target = request.path + ('?' + urlencode(cleaned) if cleaned else '')
        return redirect(target)
    bills = Bill.objects.filter(
        is_deleted=False
    ).select_related('patient', 'issued_by', 'invoice').order_by('-issued_at')
    # Filters (treat INVALID as empty)
    status_filter = _normalize_filter_value(request.GET.get('status', ''))
    bill_type_filter = _normalize_filter_value(request.GET.get('bill_type', ''))
    date_from = _normalize_filter_value(request.GET.get('date_from', ''))
    date_to = _normalize_filter_value(request.GET.get('date_to', ''))
    search = _normalize_filter_value(request.GET.get('search', ''))
    if status_filter:
        bills = bills.filter(status=status_filter)
    if bill_type_filter:
        bills = bills.filter(bill_type=bill_type_filter)
    if date_from:
        bills = bills.filter(issued_at__date__gte=date_from)
    if date_to:
        bills = bills.filter(issued_at__date__lte=date_to)
    if search:
        bills = bills.filter(Q(bill_number__icontains=search) | _patient_search_q(search))
    paginator = Paginator(bills, 50)
    page = request.GET.get('page')
    bills_page = paginator.get_page(page)
    summary = bills.aggregate(
        total_bills=Count('id'),
        total_amount=Sum('total_amount'),
        outstanding=Sum('patient_portion', filter=Q(status__in=['issued', 'partially_paid']))
    )
    context = {
        'bills': bills_page,
        'summary': summary,
        'title': 'Staff Dependency Bills',
        'search': search,
        'status_filter': status_filter,
        'bill_type_filter': bill_type_filter,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'hospital/billing/staff_dep_bill_list.html', context)


# ==================== REVENUE/PAYMENT ANALYSIS ====================

@login_required
@role_required('accountant', 'senior_account_officer')
def revenue_payment_analysis(request):
    """Revenue and payment analysis: KPIs, bill-type breakdown, and Chart.js payloads."""
    today = timezone.now().date()
    start_of_month = today.replace(day=1)

    def _parse_date(key, default):
        raw = (request.GET.get(key) or '').strip()
        if not raw:
            return default
        try:
            return date.fromisoformat(raw)
        except (ValueError, TypeError):
            return default

    date_from = _parse_date('date_from', start_of_month)
    date_to = _parse_date('date_to', today)
    if date_from > date_to:
        date_from, date_to = date_to, date_from

    date_from_s = date_from.isoformat()
    date_to_s = date_to.isoformat()

    base = Bill.objects.filter(
        issued_at__date__gte=date_from,
        issued_at__date__lte=date_to,
        is_deleted=False,
    )

    bill_revenue = base.aggregate(
        total_revenue=Sum('total_amount'),
        cash_revenue=Sum('total_amount', filter=Q(bill_type='cash')),
        insurance_revenue=Sum('insurance_covered'),
        patient_payments=Sum('patient_portion'),
    )

    payments = base.filter(status='paid').aggregate(
        total_paid=Sum('total_amount'),
        count=Count('id'),
    )

    outstanding = base.filter(status__in=['issued', 'partially_paid']).aggregate(
        total_outstanding=Sum('patient_portion'),
        count=Count('id'),
    )

    total_rev = bill_revenue['total_revenue'] or Decimal('0')
    total_paid_amt = payments['total_paid'] or Decimal('0')
    out_amt = outstanding['total_outstanding'] or Decimal('0')
    collection_rate_pct = float((total_paid_amt / total_rev) * 100) if total_rev > 0 else 0.0
    outstanding_pct = float((out_amt / total_rev) * 100) if total_rev > 0 else 0.0

    type_labels = dict(Bill.BILL_TYPES)
    by_type = list(
        base.values('bill_type')
        .annotate(
            count=Count('id'),
            revenue=Sum('total_amount'),
            paid_amount=Sum('total_amount', filter=Q(status='paid')),
            outstanding=Sum('patient_portion', filter=Q(status__in=['issued', 'partially_paid'])),
        )
        .order_by('bill_type')
    )
    for row in by_type:
        row['type_label'] = type_labels.get(row['bill_type'], row['bill_type'] or '—')

    total_bills = base.count()
    paid_bills = payments['count'] or 0
    bill_close_rate_pct = (paid_bills / total_bills * 100) if total_bills else 0.0
    avg_bill = (total_rev / total_bills) if total_bills else Decimal('0')
    avg_paid = (total_paid_amt / paid_bills) if paid_bills else Decimal('0')

    days_span = (date_to - date_from).days + 1
    if days_span <= 62:
        daily_rows = list(
            base.annotate(period=TruncDate('issued_at'))
            .values('period')
            .annotate(
                revenue=Sum('total_amount'),
                collected=Sum('total_amount', filter=Q(status='paid')),
                bill_count=Count('id'),
            )
            .order_by('period')
        )
        granularity = 'day'
    else:
        daily_rows = list(
            base.annotate(period=TruncWeek('issued_at'))
            .values('period')
            .annotate(
                revenue=Sum('total_amount'),
                collected=Sum('total_amount', filter=Q(status='paid')),
                bill_count=Count('id'),
            )
            .order_by('period')
        )
        granularity = 'week'

    daily_labels = [
        r['period'].strftime('%b %d') if r.get('period') and hasattr(r['period'], 'strftime') else ''
        for r in daily_rows
    ]
    trend_chart = {
        'labels': daily_labels,
        'revenue': [_fdec(r['revenue']) for r in daily_rows],
        'collected': [_fdec(r['collected']) for r in daily_rows],
        'bill_counts': [r['bill_count'] for r in daily_rows],
    }

    srows = list(
        base.values('status').annotate(c=Count('id'), amt=Sum('total_amount')).order_by('-c')
    )
    status_chart = {
        'labels': [_STATUS_DISPLAY.get(r['status'], r['status'] or '') for r in srows],
        'counts': [r['c'] for r in srows],
        'amounts': [_fdec(r['amt']) for r in srows],
    }

    bill_type_rows = list(
        base.values('bill_type').annotate(c=Count('id'), amt=Sum('total_amount')).order_by('-amt')
    )
    bill_type_chart = {
        'labels': [type_labels.get(r['bill_type'], r['bill_type'] or '') for r in bill_type_rows],
        'counts': [r['c'] for r in bill_type_rows],
        'amounts': [_fdec(r['amt']) for r in bill_type_rows],
    }

    inv_qs = Invoice.objects.filter(
        issued_at__date__gte=date_from,
        issued_at__date__lte=date_to,
        is_deleted=False,
    )
    invoice_stats = inv_qs.aggregate(n=Count('id'), total=Sum('total_amount'))

    analytics_data = {
        'charts': {
            'status': status_chart,
            'bill_type': bill_type_chart,
            'trend': trend_chart,
        },
        'meta': {
            'date_from': date_from_s,
            'date_to': date_to_s,
            'granularity': granularity,
        },
    }

    context = {
        'bill_revenue': bill_revenue,
        'payments': payments,
        'outstanding': outstanding,
        'by_type': by_type,
        'date_from': date_from_s,
        'date_to': date_to_s,
        'collection_rate_pct': round(collection_rate_pct, 1),
        'outstanding_pct': round(outstanding_pct, 1),
        'bill_close_rate_pct': round(bill_close_rate_pct, 1),
        'avg_bill': avg_bill,
        'avg_paid': avg_paid,
        'total_bills': total_bills,
        'invoice_count': invoice_stats['n'] or 0,
        'invoice_total': invoice_stats['total'] or Decimal('0'),
        'analytics_data': analytics_data,
    }

    return render(request, 'hospital/billing/revenue_payment_analysis.html', context)


# ==================== SALES DETAILS CLAIMS ====================

@login_required
@role_required('accountant', 'senior_account_officer')
def sales_details_claims(request):
    """Sales details with claims breakdown"""
    today = timezone.now().date()
    start_of_month = today.replace(day=1)
    
    # Date range filter
    date_from = request.GET.get('date_from', start_of_month.isoformat())
    date_to = request.GET.get('date_to', today.isoformat())
    
    # Get all invoices with claims - use correct related name 'claim_items'
    invoices = Invoice.objects.filter(
        issued_at__date__gte=date_from,
        issued_at__date__lte=date_to,
        is_deleted=False
    ).prefetch_related('lines', 'claim_items').order_by('-issued_at')
    
    # Summary
    summary = {
        'total_invoices': invoices.count(),
        'total_amount': invoices.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00'),
        'with_claims': invoices.filter(claim_items__isnull=False).distinct().count(),
        'cash_only': invoices.filter(
            claim_items__isnull=True,
            payer__isnull=True
        ).count(),
    }
    
    paginator = Paginator(invoices, 50)
    page = request.GET.get('page')
    invoices_page = paginator.get_page(page)
    
    context = {
        'invoices': invoices_page,
        'summary': summary,
        'date_from': date_from,
        'date_to': date_to,
    }
    
    return render(request, 'hospital/billing/sales_details_claims.html', context)


# ==================== SALES SUMMARY BY COMPANY ====================

@login_required
@role_required('accountant', 'senior_account_officer')
def sales_summary_by_company(request):
    """Sales summary grouped by corporate payer / insurance company, with category splits and correct totals."""
    today = timezone.now().date()
    start_of_month = today.replace(day=1)

    date_from = request.GET.get('date_from', start_of_month.isoformat())
    date_to = request.GET.get('date_to', today.isoformat())

    dec = DecimalField(max_digits=12, decimal_places=2)
    inv_f = Q(
        invoices__issued_at__date__gte=date_from,
        invoices__issued_at__date__lte=date_to,
        invoices__is_deleted=False,
    ) & ~Q(invoices__status='cancelled')
    paid_expr = ExpressionWrapper(
        F('invoices__total_amount') - F('invoices__balance'),
        output_field=dec,
    )

    corp_line_cat = _invoice_line_category_sums_by_payer(date_from, date_to, ['corporate'])

    corporate_payer_qs = Payer.objects.filter(payer_type='corporate', is_deleted=False).annotate(
        invoice_count=Count('invoices', filter=inv_f, distinct=True),
        total_billed=Coalesce(Sum('invoices__total_amount', filter=inv_f), Value(Decimal('0.00')), output_field=dec),
        total_paid=Coalesce(Sum(paid_expr, filter=inv_f), Value(Decimal('0.00')), output_field=dec),
        outstanding=Coalesce(Sum('invoices__balance', filter=inv_f), Value(Decimal('0.00')), output_field=dec),
    ).filter(
        Q(invoice_count__gt=0) | Q(total_billed__gt=0) | Q(outstanding__gt=0)
    ).order_by('-total_billed')

    corporate_summary = []
    names_seen = set()
    for p in corporate_payer_qs:
        c = corp_line_cat.get(p.id, {})
        tb = p.total_billed or Decimal('0.00')
        tp = p.total_paid or Decimal('0.00')
        os_bal = p.outstanding if p.outstanding is not None else (tb - tp)
        corporate_summary.append({
            'name': p.name,
            'source_label': 'Invoices',
            'invoice_count': p.invoice_count,
            'total_billed': tb,
            'total_paid': tp,
            'outstanding': os_bal,
            'labs': c.get('labs') or Decimal('0.00'),
            'drugs': c.get('drugs') or Decimal('0.00'),
            'scans': c.get('scans') or Decimal('0.00'),
            'credit_status': None,
            'credit_status_display': None,
        })
        names_seen.add(p.name.strip().lower())

    stmt_f = Q(
        statements__statement_date__gte=date_from,
        statements__statement_date__lte=date_to,
        statements__is_deleted=False,
    )
    stmt_line_cat = _statement_line_category_sums_by_corporate_account(date_from, date_to)
    account_qs = CorporateAccount.objects.filter(is_deleted=False).annotate(
        total_charges=Coalesce(
            Sum('statements__total_charges', filter=stmt_f),
            Value(Decimal('0.00')),
            output_field=dec,
        ),
        total_payments=Coalesce(
            Sum('statements__total_payments', filter=stmt_f),
            Value(Decimal('0.00')),
            output_field=dec,
        ),
        outstanding_stmt=Coalesce(
            Sum('statements__closing_balance', filter=stmt_f),
            Value(Decimal('0.00')),
            output_field=dec,
        ),
        stmt_count=Count('statements', filter=stmt_f, distinct=True),
    ).filter(
        Q(stmt_count__gt=0) | Q(total_charges__gt=0) | Q(outstanding_stmt__gt=0)
    )

    for acc in account_qs:
        key = acc.company_name.strip().lower()
        if key in names_seen:
            continue
        sc = stmt_line_cat.get(acc.id, {})
        corporate_summary.append({
            'name': acc.company_name,
            'source_label': 'Monthly statement',
            'invoice_count': acc.stmt_count,
            'total_billed': acc.total_charges or Decimal('0.00'),
            'total_paid': acc.total_payments or Decimal('0.00'),
            'outstanding': acc.outstanding_stmt or Decimal('0.00'),
            'labs': sc.get('labs') or Decimal('0.00'),
            'drugs': sc.get('drugs') or Decimal('0.00'),
            'scans': sc.get('scans') or Decimal('0.00'),
            'credit_status': acc.credit_status,
            'credit_status_display': acc.get_credit_status_display(),
        })

    corporate_summary.sort(key=lambda r: r['total_billed'], reverse=True)

    corporate_totals = {
        'invoice_count': sum(r['invoice_count'] for r in corporate_summary),
        'total_billed': sum(r['total_billed'] for r in corporate_summary),
        'total_paid': sum(r['total_paid'] for r in corporate_summary),
        'outstanding': sum(r['outstanding'] for r in corporate_summary),
        'labs': sum(r['labs'] for r in corporate_summary),
        'drugs': sum(r['drugs'] for r in corporate_summary),
        'scans': sum(r['scans'] for r in corporate_summary),
    }

    claim_f = Q(
        claim_items__service_date__gte=date_from,
        claim_items__service_date__lte=date_to,
        claim_items__is_deleted=False,
    )
    claim_cat = _claim_item_category_sums_by_payer(date_from, date_to)

    insurance_qs = Payer.objects.filter(
        payer_type__in=['nhis', 'private', 'insurance'],
        is_deleted=False,
    ).annotate(
        total_claims=Count('claim_items', filter=claim_f, distinct=True),
        total_billed=Coalesce(
            Sum('claim_items__billed_amount', filter=claim_f),
            Value(Decimal('0.00')),
            output_field=dec,
        ),
        total_paid=Coalesce(
            Sum('claim_items__paid_amount', filter=claim_f),
            Value(Decimal('0.00')),
            output_field=dec,
        ),
    ).filter(
        Q(total_claims__gt=0) | Q(total_billed__gt=0)
    ).order_by('-total_billed')

    insurance_summary = []
    for payer in insurance_qs:
        cc = claim_cat.get(payer.id, {})
        tb = payer.total_billed or Decimal('0.00')
        tp = payer.total_paid or Decimal('0.00')
        insurance_summary.append({
            'name': payer.name,
            'total_claims': payer.total_claims,
            'total_billed': tb,
            'total_paid': tp,
            'outstanding': tb - tp,
            'labs': cc.get('labs') or Decimal('0.00'),
            'drugs': cc.get('drugs') or Decimal('0.00'),
            'scans': cc.get('scans') or Decimal('0.00'),
        })

    insurance_totals = {
        'total_claims': sum(r['total_claims'] for r in insurance_summary),
        'total_billed': sum(r['total_billed'] for r in insurance_summary),
        'total_paid': sum(r['total_paid'] for r in insurance_summary),
        'outstanding': sum(r['outstanding'] for r in insurance_summary),
        'labs': sum(r['labs'] for r in insurance_summary),
        'drugs': sum(r['drugs'] for r in insurance_summary),
        'scans': sum(r['scans'] for r in insurance_summary),
    }

    context = {
        'corporate_summary': corporate_summary,
        'insurance_summary': insurance_summary,
        'corporate_totals': corporate_totals,
        'insurance_totals': insurance_totals,
        'date_from': date_from,
        'date_to': date_to,
    }

    return render(request, 'hospital/billing/sales_summary_by_company.html', context)

