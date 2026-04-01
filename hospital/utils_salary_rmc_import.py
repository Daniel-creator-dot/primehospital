"""
Import staff payroll from Raphal Medical Centre — Sample Salary-RMC.xlsx layout.

Header row is detected automatically (row containing "Serial" and "Emp. Code").
Period text is read from a row like "PERIOD :JANUARY 2026".
"""
from __future__ import annotations

import calendar
import re
from dataclasses import dataclass, field
from decimal import Decimal
from typing import Any, BinaryIO, Optional

MONTH_NAMES = {
    'january': 1, 'february': 2, 'march': 3, 'april': 4, 'may': 5, 'june': 6,
    'july': 7, 'august': 8, 'september': 9, 'october': 10, 'november': 11, 'december': 12,
}


def normalize_header(value: Any) -> str:
    if value is None:
        return ''
    s = str(value).strip().lower()
    s = re.sub(r'\s+', ' ', s)
    return s


def normalize_emp_code(value: Any) -> str:
    if value is None or value == '':
        return ''
    if isinstance(value, float):
        if value == int(value):
            value = int(value)
    if isinstance(value, int):
        return str(value)
    s = str(value).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return s


def to_decimal(value: Any) -> Decimal:
    if value is None or value == '':
        return Decimal('0')
    if isinstance(value, (int, Decimal)):
        return Decimal(str(value))
    s = str(value).strip().replace(',', '')
    if not s:
        return Decimal('0')
    return Decimal(s)


def find_header_row(ws, max_row: int = 25) -> Optional[int]:
    for r in range(1, max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if a is None:
            continue
        na = normalize_header(a)
        nb = normalize_header(b) if b else ''
        if 'serial' in na and ('emp' in nb and 'code' in nb):
            return r
    return None


def parse_period_from_sheet(ws) -> tuple[Optional[str], Optional[tuple[int, int, int]]]:
    """
    Returns (label, (year, month, day)) where day is last day of month; or Nones.
    """
    for r in range(1, 16):
        v = ws.cell(r, 1).value
        if not v:
            continue
        s = str(v).strip()
        if 'period' not in s.lower():
            continue
        m = re.search(
            r'(january|february|march|april|may|june|july|august|september|october|november|december)\s+(\d{4})',
            s,
            re.I,
        )
        if not m:
            continue
        month_name = m.group(1).lower()
        year = int(m.group(2))
        month = MONTH_NAMES.get(month_name)
        if not month:
            continue
        last_day = calendar.monthrange(year, month)[1]
        label = f'{month_name.title()} {year}'
        return label, (year, month, last_day)
    return None, None


def classify_column(normalized_header: str) -> Optional[str]:
    nh = normalized_header
    if nh.startswith('serial') or (nh.startswith('serial') and '#' in nh):
        return 'sheet_serial'
    if 'emp' in nh and 'code' in nh:
        return 'emp_code'
    if 'employee' in nh and 'name' in nh:
        return 'employee_name'
    if nh == 'department' or nh.startswith('department'):
        return 'department'
    if nh == 'rank' or nh.startswith('rank'):
        return 'rank'
    if 'service' in nh or 'lenth' in nh or 'length' in nh:
        return 'service_length'
    if 'basic' in nh and 'salary' in nh:
        return 'basic_salary'
    if 'other' in nh and 'allowance' in nh:
        return 'other_allowances'
    if 'medical' in nh and 'allowance' in nh:
        return 'medical_allowance'
    if 'risk' in nh or ('emergency' in nh and 'allowance' in nh):
        return 'risk_emergency_allowance'
    if 'total' in nh and 'gross' in nh:
        return 'total_gross_salary'
    if '5.5' in nh and 'ssf' in nh:
        return 'ssnit_employee'
    if '5.0' in nh and 'pf' in nh and 'employer' in nh:
        return 'pf_employer_contribution'
    if '5.0' in nh and 'pf' in nh and 'employee' in nh:
        return 'pension_employee'
    if 'personal' in nh and 'relief' in nh:
        return 'personal_relief'
    if '13%' in nh and 'ssf' in nh:
        return 'ssf_employer_contribution'
    if 'total' in nh and 'relief' in nh:
        return 'total_relief'
    if 'taxable' in nh and 'income' in nh:
        return 'taxable_income'
    if 'paye' in nh:
        return 'paye_tax'
    if 'net' in nh and 'pay' in nh:
        return 'net_pay_sheet'
    return None


def build_column_map(ws, header_row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        raw = ws.cell(header_row, c).value
        key = classify_column(normalize_header(raw))
        if key and key not in mapping:
            mapping[key] = c
    return mapping


@dataclass
class RMCImportRow:
    sheet_serial: Optional[int] = None
    emp_code: str = ''
    employee_name: str = ''
    department: str = ''
    rank: str = ''
    service_length: str = ''
    basic_salary: Decimal = field(default_factory=lambda: Decimal('0'))
    other_allowances: Decimal = field(default_factory=lambda: Decimal('0'))
    medical_allowance: Decimal = field(default_factory=lambda: Decimal('0'))
    risk_emergency_allowance: Decimal = field(default_factory=lambda: Decimal('0'))
    total_gross_salary: Decimal = field(default_factory=lambda: Decimal('0'))
    ssnit_employee: Decimal = field(default_factory=lambda: Decimal('0'))
    pf_employer_contribution: Decimal = field(default_factory=lambda: Decimal('0'))
    pension_employee: Decimal = field(default_factory=lambda: Decimal('0'))
    personal_relief: Decimal = field(default_factory=lambda: Decimal('0'))
    ssf_employer_contribution: Decimal = field(default_factory=lambda: Decimal('0'))
    total_relief: Decimal = field(default_factory=lambda: Decimal('0'))
    taxable_income: Decimal = field(default_factory=lambda: Decimal('0'))
    paye_tax: Decimal = field(default_factory=lambda: Decimal('0'))
    net_pay_sheet: Decimal = field(default_factory=lambda: Decimal('0'))


def row_to_rmc_dataclass(ws, r: int, colmap: dict[str, int]) -> RMCImportRow:
    def cell(field: str) -> Any:
        ci = colmap.get(field)
        if not ci:
            return None
        return ws.cell(r, ci).value

    serial_raw = cell('sheet_serial')
    serial = None
    if serial_raw is not None and str(serial_raw).strip() != '':
        try:
            serial = int(float(serial_raw))
        except (TypeError, ValueError):
            serial = None

    return RMCImportRow(
        sheet_serial=serial,
        emp_code=normalize_emp_code(cell('emp_code')),
        employee_name=(str(cell('employee_name')).strip() if cell('employee_name') else ''),
        department=(str(cell('department')).strip() if cell('department') else ''),
        rank=(str(cell('rank')).strip() if cell('rank') else ''),
        service_length=(str(cell('service_length')).strip() if cell('service_length') else ''),
        basic_salary=to_decimal(cell('basic_salary')),
        other_allowances=to_decimal(cell('other_allowances')),
        medical_allowance=to_decimal(cell('medical_allowance')),
        risk_emergency_allowance=to_decimal(cell('risk_emergency_allowance')),
        total_gross_salary=to_decimal(cell('total_gross_salary')),
        ssnit_employee=to_decimal(cell('ssnit_employee')),
        pf_employer_contribution=to_decimal(cell('pf_employer_contribution')),
        pension_employee=to_decimal(cell('pension_employee')),
        personal_relief=to_decimal(cell('personal_relief')),
        ssf_employer_contribution=to_decimal(cell('ssf_employer_contribution')),
        total_relief=to_decimal(cell('total_relief')),
        taxable_income=to_decimal(cell('taxable_income')),
        paye_tax=to_decimal(cell('paye_tax')),
        net_pay_sheet=to_decimal(cell('net_pay_sheet')),
    )


def iter_data_rows(ws, header_row: int, colmap: dict[str, int]):
    for r in range(header_row + 1, ws.max_row + 1):
        row = row_to_rmc_dataclass(ws, r, colmap)
        if not row.emp_code and not row.employee_name:
            continue
        if row.basic_salary == 0 and row.total_gross_salary == 0 and row.net_pay_sheet == 0:
            continue
        yield row


def parse_rmc_workbook(fileobj: BinaryIO) -> tuple[Optional[str], Optional[tuple[int, int, int]], list[RMCImportRow], list[str]]:
    """
    Returns (period_label, (y,m,last_day), rows, warnings).
    """
    import openpyxl

    wb = openpyxl.load_workbook(fileobj, data_only=True)
    ws = wb[wb.sheetnames[0]]

    warnings: list[str] = []
    period_label, ymd = parse_period_from_sheet(ws)

    hr = find_header_row(ws)
    if not hr:
        return period_label, ymd, [], ['Could not find header row (Serial # / Emp. Code).']

    colmap = build_column_map(ws, hr)
    required = ('emp_code', 'employee_name', 'basic_salary')
    for req in required:
        if req not in colmap:
            warnings.append(f'Missing expected column for: {req}')

    rows = list(iter_data_rows(ws, hr, colmap))
    return period_label, ymd, rows, warnings


def resolve_staff_for_rmc(emp_code: str, employee_name: str):
    """Match Staff by employee_id (exact / zero-padded) then by first+last name."""
    from django.db.models import Q
    from hospital.models import Staff

    code = (emp_code or '').strip()
    if not code and not (employee_name or '').strip():
        return None

    qs = Staff.objects.filter(is_deleted=False).select_related('user')
    if code:
        st = qs.filter(employee_id__iexact=code).first()
        if st:
            return st
        if code.isdigit():
            st = qs.filter(employee_id__iexact=code.zfill(4)).first()
            if st:
                return st
            st = qs.filter(employee_id__iexact=code.zfill(3)).first()
            if st:
                return st

    name = (employee_name or '').strip()
    if name:
        parts = name.split()
        if len(parts) >= 2:
            st = qs.filter(
                user__first_name__icontains=parts[0],
                user__last_name__icontains=parts[-1],
            ).first()
            if st:
                return st
        st = qs.filter(
            Q(user__first_name__icontains=name) | Q(user__last_name__icontains=name)
        ).first()
        if st:
            return st
    return None


def apply_rmc_rows_to_payroll(payroll, rows: list[RMCImportRow], replace: bool = True) -> tuple[list[str], int]:
    """Create AccountingPayrollEntry rows. Returns (error_messages, created_count)."""
    from django.db import transaction
    from hospital.models_accounting_advanced import AccountingPayrollEntry

    errors: list[str] = []
    created = 0

    with transaction.atomic():
        if replace:
            AccountingPayrollEntry.objects.filter(payroll=payroll).delete()

        seen_staff: set = set()
        for row in rows:
            staff = resolve_staff_for_rmc(row.emp_code, row.employee_name)
            if not staff:
                label = row.sheet_serial if row.sheet_serial is not None else '?'
                errors.append(
                    f'Row {label}: no HMS staff for Emp. Code {row.emp_code!r} — {row.employee_name}'
                )
                continue
            if staff.id in seen_staff:
                errors.append(
                    f'Row {row.sheet_serial or "?"}: duplicate Emp. Code {row.emp_code!r} in file'
                )
                continue
            seen_staff.add(staff.id)

            entry = AccountingPayrollEntry(
                payroll=payroll,
                staff=staff,
                sheet_serial=row.sheet_serial,
                department_snapshot=row.department[:200] if row.department else '',
                rank_snapshot=row.rank[:120] if row.rank else '',
                service_length_snapshot=row.service_length[:80] if row.service_length else '',
                basic_salary=row.basic_salary,
                other_allowances=row.other_allowances,
                medical_allowance=row.medical_allowance,
                risk_emergency_allowance=row.risk_emergency_allowance,
                ssnit_employee=row.ssnit_employee,
                pf_employer_contribution=row.pf_employer_contribution,
                pension_employee=row.pension_employee,
                personal_relief=row.personal_relief,
                ssf_employer_contribution=row.ssf_employer_contribution,
                total_relief=row.total_relief,
                taxable_income=row.taxable_income,
                paye_tax=row.paye_tax,
                gross_pay=Decimal('0'),
                deductions=Decimal('0'),
                net_pay=Decimal('0'),
            )
            entry.save()
            created += 1

        payroll.recalculate_totals_from_entries()

    return errors, created


def build_rmc_template_workbook_bytes() -> bytes:
    """Empty workbook matching Sample Salary-RMC layout (row 5–11 titles + header)."""
    import openpyxl
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    ws.merge_cells('A5:T5')
    ws['A5'] = 'STAFF  PAYROLL'
    ws.merge_cells('A6:T6')
    ws['A6'] = 'RAPHAL  MEDICAL  CENTRE'
    ws['A8'] = 'PERIOD :JANUARY 2026'

    ws['H10'] = 'Taxable Allowances'
    ws['L10'] = 'Pension'

    headers = [
        'Serial #', 'Emp. Code', 'Employee Name', 'Department', 'Rank', 'Lenth of service',
        ' Basic Salary', 'Other Allowance/Overtime', 'Medical Allowance', 'Risk/Emergency Allowance',
        'Total Gross Salary', ' 5.5 %SSF Employee', ' 5.0 %PF Employer', ' 5.0 %PF Employee',
        'Personal Relief', '13% SSF Employer', 'Total Relief', 'Taxable Income', 'PAYE Tax Deduct.',
        'Net Pay After Tax',
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(11, i, value=h)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
