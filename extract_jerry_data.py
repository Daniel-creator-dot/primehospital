import openpyxl
from decimal import Decimal
import json

def extract_excel_data():
    """Extract debtor and creditor balances from JERRY.xlsx"""
    wb = openpyxl.load_workbook('insurance excel/JERRY.xlsx', data_only=True)  # data_only=True gets calculated values
    
    results = {
        'debtors': [],  # Insurance companies (receivables)
        'creditors': []  # Suppliers (payables)
    }
    
    # Extract DEBTORS (Insurance Companies)
    if 'DEBTOR BALANCES' in wb.sheetnames:
        ws = wb['DEBTOR BALANCES']
        print("Processing DEBTOR BALANCES...")
        
        for row in ws.iter_rows(min_row=5, values_only=True):  # Start from row 5 (skip headers)
            company_name = row[0] if row[0] else None
            balance = row[1] if len(row) > 1 else None
            
            if company_name and company_name.strip():
                # Try to convert balance to Decimal
                balance_value = None
                if balance is not None:
                    try:
                        if isinstance(balance, (int, float)):
                            balance_value = Decimal(str(balance))
                        elif isinstance(balance, str):
                            # Remove currency symbols and commas
                            balance_str = balance.replace('GHC', '').replace('GHS', '').replace(',', '').strip()
                            if balance_str:
                                balance_value = Decimal(balance_str)
                    except:
                        balance_value = Decimal('0.00')
                else:
                    balance_value = Decimal('0.00')
                
                if balance_value and balance_value != Decimal('0.00'):
                    results['debtors'].append({
                        'name': company_name.strip(),
                        'balance': float(balance_value)
                    })
                    print(f"  Debtor: {company_name.strip()} = {balance_value}")
    
    # Extract CREDITORS (Suppliers)
    if 'CREDITOR BALANCES' in wb.sheetnames:
        ws = wb['CREDITOR BALANCES']
        print("\nProcessing CREDITOR BALANCES...")
        
        for row in ws.iter_rows(min_row=3, values_only=True):  # Start from row 3
            supplier_name = row[1] if len(row) > 1 and row[1] else None
            balance = row[2] if len(row) > 2 and row[2] else None
            
            if supplier_name and supplier_name.strip():
                # Try to convert balance to Decimal
                balance_value = None
                if balance is not None:
                    try:
                        if isinstance(balance, (int, float)):
                            balance_value = Decimal(str(balance))
                        elif isinstance(balance, str):
                            # Remove currency symbols and commas
                            balance_str = balance.replace('GHC', '').replace('GHS', '').replace(',', '').strip()
                            if balance_str:
                                balance_value = Decimal(balance_str)
                    except:
                        balance_value = Decimal('0.00')
                else:
                    balance_value = Decimal('0.00')
                
                if balance_value and balance_value != Decimal('0.00'):
                    results['creditors'].append({
                        'name': supplier_name.strip(),
                        'balance': float(balance_value)
                    })
                    print(f"  Creditor: {supplier_name.strip()} = {balance_value}")
    
    # Save to JSON for review
    with open('jerry_extracted_data.json', 'w') as f:
        json.dump(results, f, indent=2)
    
    print(f"\n\nSummary:")
    print(f"Total Debtors (Insurance): {len(results['debtors'])}")
    print(f"Total Creditors (Suppliers): {len(results['creditors'])}")
    print(f"\nData saved to jerry_extracted_data.json")
    
    return results

if __name__ == '__main__':
    extract_excel_data()


