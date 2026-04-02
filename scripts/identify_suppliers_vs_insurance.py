"""
Identify Suppliers vs Insurance Companies from Excel Data
This script will help identify which entries are suppliers vs insurance companies
"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'hms.settings')
import django
django.setup()

from openpyxl import load_workbook

# Keywords that indicate suppliers (not insurance)
SUPPLIER_KEYWORDS = [
    'pharma', 'pharmaceutical', 'enterprise', 'logistics', 'supply', 
    'electricals', 'motors', 'company limited', 'ltd', 'limited',
    'bank', 'plc', 'church', 'university', 'college', 'school',
    'total', 'oil', 'petroleum', 'gas', 'energy', 'power'
]

# Keywords that indicate insurance
INSURANCE_KEYWORDS = [
    'insurance', 'health insurance', 'medical insurance', 'nhis',
    'mutual health', 'care', 'health', 'medical', 'assurance'
]

def classify_entity(name):
    """Classify an entity as supplier, insurance, or unknown"""
    name_lower = name.lower()
    
    # Check for insurance keywords
    for keyword in INSURANCE_KEYWORDS:
        if keyword in name_lower:
            return 'insurance'
    
    # Check for supplier keywords
    for keyword in SUPPLIER_KEYWORDS:
        if keyword in name_lower:
            return 'supplier'
    
    # Check specific patterns
    if 'health insurance' in name_lower or 'medical insurance' in name_lower:
        return 'insurance'
    
    if any(word in name_lower for word in ['pharma', 'pharmaceutical', 'enterprise', 'logistics']):
        return 'supplier'
    
    return 'unknown'

if __name__ == "__main__":
    excel_file = "insurance excel/JERRY.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"File not found: {excel_file}")
        sys.exit(1)
    
    wb = load_workbook(excel_file, read_only=True, data_only=True)
    
    print("="*80)
    print("CLASSIFYING ENTITIES FROM DEBTOR BALANCES")
    print("="*80)
    
    # Process DEBTOR BALANCES sheet
    if 'DEBTOR BALANCES' in wb.sheetnames:
        ws = wb['DEBTOR BALANCES']
        print("\nDEBTOR BALANCES (Entities we're owed money from):\n")
        
        insurance_entities = []
        supplier_entities = []
        unknown_entities = []
        
        # Find data start
        data_start = None
        for row_idx in range(1, min(35, ws.max_row + 1)):
            row = [cell.value for cell in ws[row_idx]]
            if row and len(row) >= 2 and isinstance(row[0], str) and isinstance(row[1], (int, float)):
                data_start = row_idx
                break
        
        if data_start:
            for row_idx in range(data_start, ws.max_row + 1):
                row = [cell.value for cell in ws[row_idx]]
                if not row or len(row) < 2:
                    continue
                
                entity_name = str(row[0]).strip() if row[0] else None
                amount = row[1] if isinstance(row[1], (int, float)) else None
                
                if entity_name and amount and amount != 0:
                    classification = classify_entity(entity_name)
                    
                    if classification == 'insurance':
                        insurance_entities.append((entity_name, amount))
                        print(f"  [INSURANCE] {entity_name}: GHS {amount:,.2f}")
                    elif classification == 'supplier':
                        supplier_entities.append((entity_name, amount))
                        print(f"  [SUPPLIER]  {entity_name}: GHS {amount:,.2f}")
                    else:
                        unknown_entities.append((entity_name, amount))
                        print(f"  [UNKNOWN]   {entity_name}: GHS {amount:,.2f}")
        
        print(f"\n\nSummary:")
        print(f"  Insurance Companies: {len(insurance_entities)}")
        print(f"  Suppliers: {len(supplier_entities)}")
        print(f"  Unknown: {len(unknown_entities)}")
        
        if supplier_entities:
            print(f"\n\n[WARNING] SUPPLIERS FOUND IN DEBTOR BALANCES (Should be in CREDITOR BALANCES):")
            for name, amount in supplier_entities:
                print(f"    - {name}: GHS {amount:,.2f}")
    
    print("\n\n" + "="*80)
    print("CREDITOR BALANCES (Suppliers we owe money to)")
    print("="*80)
    
    # Process CREDITOR BALANCES sheet
    if 'CREDITOR BALANCES' in wb.sheetnames:
        ws = wb['CREDITOR BALANCES']
        print("\nFirst 20 suppliers:\n")
        
        data_start = None
        for row_idx in range(1, min(25, ws.max_row + 1)):
            row = [cell.value for cell in ws[row_idx]]
            if row and len(row) >= 2 and isinstance(row[1], str) and row[1].strip():
                if isinstance(row[2], (int, float)):
                    data_start = row_idx
                    break
        
        if data_start:
            count = 0
            for row_idx in range(data_start, min(data_start + 20, ws.max_row + 1)):
                row = [cell.value for cell in ws[row_idx]]
                if not row or len(row) < 3:
                    continue
                
                supplier_name = str(row[1]).strip() if row[1] else None
                amount = row[2] if isinstance(row[2], (int, float)) else None
                
                if supplier_name and amount is not None:
                    count += 1
                    print(f"  {count}. {supplier_name}: GHS {amount:,.2f}")

