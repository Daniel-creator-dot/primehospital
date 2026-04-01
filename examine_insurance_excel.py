"""
Examine Insurance Excel Files Structure
This script will help us understand the format of the adjudication reports
"""
import os
import sys

# Add the project root to the path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# Set up Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'hms.settings')
import django
django.setup()

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False
    print("Pandas not available, trying openpyxl...")
    try:
        from openpyxl import load_workbook
        HAS_OPENPYXL = True
    except ImportError:
        HAS_OPENPYXL = False
        print("Neither pandas nor openpyxl available")

def examine_excel_file(file_path):
    """Examine an Excel file and print its structure"""
    print(f"\n{'='*80}")
    print(f"Examining: {file_path}")
    print(f"{'='*80}\n")
    
    if HAS_PANDAS:
        try:
            # Read first few rows
            df = pd.read_excel(file_path, nrows=20)
            print(f"Shape: {df.shape[0]} rows x {df.shape[1]} columns")
            print(f"\nColumns ({len(df.columns)}):")
            for i, col in enumerate(df.columns, 1):
                print(f"  {i}. {col}")
            
            print(f"\nFirst 5 rows:")
            print(df.head().to_string())
            
            print(f"\nData types:")
            print(df.dtypes)
            
            print(f"\nSample values from each column:")
            for col in df.columns:
                sample = df[col].dropna().head(3).tolist()
                if sample:
                    print(f"  {col}: {sample}")
            
            return df
        except Exception as e:
            print(f"Error reading with pandas: {e}")
            return None
    
    elif HAS_OPENPYXL:
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            print(f"Sheet names: {wb.sheetnames}")
            
            for sheet_name in wb.sheetnames[:3]:  # First 3 sheets
                ws = wb[sheet_name]
                print(f"\nSheet: {sheet_name}")
                print(f"  Dimensions: {ws.max_row} rows x {ws.max_column} columns")
                
                # Get header row
                headers = []
                for cell in ws[1]:
                    headers.append(cell.value)
                print(f"  Headers ({len(headers)}):")
                for i, header in enumerate(headers[:20], 1):  # First 20 columns
                    print(f"    {i}. {header}")
                
                # Find actual data start (skip header rows)
                data_start_row = 1
                for row_idx in range(1, min(20, ws.max_row + 1)):
                    row_values = [cell.value for cell in ws[row_idx] if cell.value is not None]
                    if len(row_values) > 2:  # Likely a data row
                        data_start_row = row_idx
                        break
                
                print(f"  Data appears to start at row: {data_start_row}")
                
                # Get header row (row before data starts)
                if data_start_row > 1:
                    header_row = [cell.value for cell in ws[data_start_row - 1]]
                    print(f"  Header row ({len([h for h in header_row if h])} columns):")
                    for i, header in enumerate(header_row[:15], 1):
                        if header:
                            print(f"    {i}. {header}")
                
                # Get first 10 data rows
                print(f"\n  First 10 data rows (starting from row {data_start_row}):")
                for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, max_row=min(data_start_row + 9, ws.max_row), values_only=True), data_start_row):
                    row_data = [val for val in row if val is not None][:15]  # First 15 non-null values
                    if row_data:
                        print(f"    Row {row_idx}: {row_data}")
                
        except Exception as e:
            print(f"Error reading with openpyxl: {e}")
            return None
    
    else:
        print("No Excel library available. Please install pandas or openpyxl")
        return None

if __name__ == "__main__":
    excel_folder = "insurance excel"
    
    if not os.path.exists(excel_folder):
        print(f"Folder '{excel_folder}' not found!")
        sys.exit(1)
    
    # Get all Excel files (case-insensitive)
    excel_files = []
    for file in os.listdir(excel_folder):
        if file.lower().endswith(('.xls', '.xlsx')):
            excel_files.append(os.path.join(excel_folder, file))
    
    if not excel_files:
        print(f"No Excel files found in '{excel_folder}'")
        sys.exit(1)
    
    print(f"Found {len(excel_files)} Excel file(s):")
    for f in excel_files:
        print(f"  - {os.path.basename(f)}")
    
    # Prioritize adjudication reports (XLS files)
    adjudication_files = [f for f in excel_files if f.endswith('.XLS') or 'ADJUDICATION' in os.path.basename(f).upper()]
    other_files = [f for f in excel_files if f not in adjudication_files]
    
    # Examine adjudication reports first
    if adjudication_files:
        print(f"\n\n{'='*80}")
        print("ADJUDICATION REPORTS (Priority):")
        print(f"{'='*80}\n")
        for f in adjudication_files:
            examine_excel_file(f)
    
    # Then examine other files
    if other_files:
        print(f"\n\n{'='*80}")
        print("OTHER FILES:")
        print(f"{'='*80}\n")
        for f in other_files:
            examine_excel_file(f)

